from __future__ import annotations

import importlib
import os
import sys
import tempfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))


SUPPLY_HEADERS = [
    "순서",
    "매입거래처",
    "매출거래처",
    "거래구분",
    "장부입력확인",
    "주문일",
    "출고일",
    "주문자",
    "발신자연락처",
    "수령자",
    "수령자연락처",
    "제 품 명",
    "수량",
    "상 세 주 소",
    "택배사**",
    "배송번호",
    "특이(요청)사항",
    "주문상품고유번호",
    "상품코드",
    "주문번호",
    "고객선택옵션",
]


def create_supply_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Supply"
    worksheet.append(SUPPLY_HEADERS)
    worksheet.append(
        [
            1,
            "(주)소일브릿지(본사)",
            "토스",
            "",
            "",
            "2026-06-15 09:30:46",
            "",
            "김주문",
            "010-1111-2222",
            "이수령",
            "010-3333-4444",
            "[테스트] 감식초 1000ml",
            2,
            "서울시 테스트구 테스트로 1",
            "롯데택배",
            "1234567890",
            "문 앞",
            "111385084",
            "SOSO27788588",
            "WSO260615-000000021",
            "기본",
        ]
    )
    workbook.save(path)


def load_app(tmp_path: Path):
    os.environ["WORKHUB_DATA_DIR"] = str(tmp_path)
    sys.modules.pop("workhub_delivery_app", None)
    return importlib.import_module("workhub_delivery_app")


def test_delivery_summary_accepts_supply_format(tmp_path: Path) -> None:
    from delivery_text_summary import summarize_workbook

    workbook_path = tmp_path / "supply.xlsx"
    create_supply_workbook(workbook_path)

    text, sheet_names = summarize_workbook(workbook_path, sort_mode="first")

    assert sheet_names == ["Supply"]
    assert "[테스트] 감식초 1000ml - 2개 (1건)" in text


def test_management_import_accepts_supply_header_on_first_row(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    workbook_path = tmp_path / "supply.xlsx"
    create_supply_workbook(workbook_path)

    result = app.import_management_workbook(workbook_path)
    rows = app.list_management_records(limit=None)

    assert result["inserted"] == 1
    assert result["skipped"] == 0
    assert rows[0]["receiver_name"] == "이수령"
    assert rows[0]["invoice_number"] == "1234567890"
    assert rows[0]["order_item_id"] == "111385084"
    assert rows[0]["product_code"] == "SOSO27788588"
    assert rows[0]["order_number"] == "WSO260615-000000021"
    assert rows[0]["ship_date"] == "2026-06-15"
    assert rows[0]["ledger_checked"] == "\uc785\ub825 \uc644\ub8cc"
    assert rows[0]["customer_option"] == "기본"


def test_management_import_defaults_blank_ship_date_and_ledger_check_only(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    records = [
        {
            "purchase_vendor": "(주)소일브릿지(본사)",
            "order_date": "2026-06-15 09:30:46",
            "ship_date": "",
            "transaction_type": "",
            "ledger_checked": "",
        },
        {
            "purchase_vendor": "키친쿡",
            "order_date": "2026-06-16",
            "ship_date": "",
            "transaction_type": "",
            "ledger_checked": "\ud655\uc778 \ubcf4\ub958",
        },
    ]

    app.normalize_management_import_records(records)

    assert records[0]["order_date"] == "2026-06-15"
    assert records[0]["ship_date"] == "2026-06-15"
    assert records[0]["transaction_type"] == "매출"
    assert records[0]["ledger_checked"] == "\uc785\ub825 \uc644\ub8cc"
    assert records[1]["ship_date"] == "2026-06-16"
    assert records[1]["transaction_type"] == "매입/매출"
    assert records[1]["ledger_checked"] == "\ud655\uc778 \ubcf4\ub958"


def test_management_import_records_are_sorted_before_insert(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    records = [
        {"quantity": "2", "product_name": "나상품", "purchase_vendor": "키친쿡", "sales_vendor": "B"},
        {"quantity": "1", "product_name": "다상품", "purchase_vendor": "키친쿡", "sales_vendor": "B"},
        {"quantity": "1", "product_name": "가상품", "purchase_vendor": "모드니", "sales_vendor": "B"},
        {"quantity": "1", "product_name": "가상품", "purchase_vendor": "가매입", "sales_vendor": "A"},
    ]

    app.sort_management_import_records(records)

    assert [
        (row["quantity"], row["product_name"], row["purchase_vendor"], row["sales_vendor"])
        for row in records
    ] == [
        ("1", "가상품", "가매입", "A"),
        ("1", "가상품", "모드니", "B"),
        ("1", "다상품", "키친쿡", "B"),
        ("2", "나상품", "키친쿡", "B"),
    ]


def test_manual_management_record_creation_normalizes_defaults(tmp_path: Path) -> None:
    app = load_app(tmp_path)

    record_id = app.create_management_manual_record({
        "purchase_vendor": "키친쿡",
        "sales_vendor": "모드니",
        "order_date": "2026-06-24",
        "ship_date": "",
        "product_name": "수기 상품",
        "quantity": "3",
        "receiver_name": "홍길동",
    })
    record = app.get_management_record(record_id)

    assert record["source_file"] == "수기입력"
    assert record["source_sheet"] == "수기추가"
    assert record["ship_date"] == "2026-06-24"
    assert record["transaction_type"] == "매입/매출"
    assert record["ledger_checked"] == "입력 완료"
    assert record["product_name"] == "수기 상품"


def test_management_template_export_uses_supply_columns_and_filename(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    rows = [
        {
            "purchase_vendor": "(주)소일브릿지(본사)",
            "sales_vendor": "토스",
            "order_date": "2026-06-15",
            "receiver_name": "이수령",
            "receiver_phone": "010-3333-4444",
            "product_name": "[테스트] 감식초 1000ml",
            "quantity": "2",
            "receiver_address": "서울시 테스트구 테스트로 1",
            "courier": "롯데택배",
            "invoice_number": "1234567890",
            "memo": "문 앞",
            "order_item_id": "111385084",
            "product_code": "SOSO27788588",
            "order_number": "WSO260615-000000021",
            "customer_option": "기본",
        }
    ]

    exported = load_workbook(BytesIO(app.management_workbook_bytes_from_template(rows)), data_only=True)
    worksheet = exported.worksheets[0]

    assert [worksheet.cell(1, column).value for column in range(1, 22)] == SUPPLY_HEADERS
    assert worksheet.cell(2, 1).value == "1"
    assert worksheet.cell(2, 18).value == "111385084"
    assert worksheet.cell(2, 20).value == "WSO260615-000000021"
    assert (
        app.management_template_filename_stem(rows, {}, datetime(2026, 6, 15))
        == "통합관리대장 양식 2026년 20260615"
    )


def main() -> None:
    with tempfile.TemporaryDirectory() as directory:
        test_delivery_summary_accepts_supply_format(Path(directory) / "summary")
    with tempfile.TemporaryDirectory() as directory:
        test_management_import_accepts_supply_header_on_first_row(Path(directory) / "import")
    with tempfile.TemporaryDirectory() as directory:
        test_management_import_defaults_blank_ship_date_and_ledger_check_only(Path(directory) / "normalize")
    with tempfile.TemporaryDirectory() as directory:
        test_management_import_records_are_sorted_before_insert(Path(directory) / "sort")
    with tempfile.TemporaryDirectory() as directory:
        test_manual_management_record_creation_normalizes_defaults(Path(directory) / "manual")
    with tempfile.TemporaryDirectory() as directory:
        test_management_template_export_uses_supply_columns_and_filename(Path(directory) / "export")


if __name__ == "__main__":
    main()
