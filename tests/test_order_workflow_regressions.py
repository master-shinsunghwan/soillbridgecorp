from __future__ import annotations

import importlib.util
import tempfile
import unittest
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side


ROOT = Path(__file__).resolve().parents[1]
MODULE_SETS = [
    ROOT / "scripts",
    ROOT / "_workhub_zip_inspect" / "scripts",
]
TEMPLATE_PATHS = [
    ROOT / "templates" / "vehicle_receipt_template.xlsx",
    ROOT / "_workhub_zip_inspect" / "templates" / "vehicle_receipt_template.xlsx",
]


def load_module(module_path: Path):
    spec = importlib.util.spec_from_file_location(module_path.stem, module_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def create_delivery_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Supply"
    worksheet.append(
        [
            "\uc21c\uc11c",
            "\uc8fc\ubb38\uc77c",
            "\uc8fc\ubb38\ubc88\ud638",
            "\uc218\ub839\uc790",
            "\uc218\ub839\uc790\uc5f0\ub77d\ucc98",
            "\uc81c \ud488 \uba85",
            "\uc218\ub7c9",
            "\uc0c1 \uc138 \uc8fc \uc18c",
        ]
    )
    worksheet.append(
        [
            1,
            "2026-06-15 09:30:46",
            "WSO260615-000000001",
            "\ud64d\uae38\ub3d9",
            "010-1111-2222",
            "\ud14c\uc2a4\ud2b8 \uc0c1\ud488",
            2,
            "\uc11c\uc6b8\uc2dc \ud14c\uc2a4\ud2b8\ub85c 1",
        ]
    )
    workbook.save(path)


def create_safe_number_delivery_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SafeNumber"
    worksheet.append(
        [
            "순서",
            "주문일",
            "주문번호",
            "수령자",
            "수령자연락처",
            "제 품 명",
            "수량",
            "상 세 주 소",
        ]
    )
    worksheet.append(
        [
            1,
            "2026-06-18 09:30:46",
            "WSO260618-000000001",
            "홍길동",
            "0504-1111-2222",
            "아이제나흐 에센 스텐 찜솥 28cm",
            6,
            "서울시 테스트로 1",
        ]
    )
    worksheet.append(
        [
            2,
            "2026-06-18 09:31:10",
            "WSO260618-000000002",
            "홍길동",
            "0504-3333-4444",
            "아이제나흐 에쎈 레인지 쿡 12종/24P",
            1,
            "서울시 테스트로 1",
        ]
    )
    workbook.save(path)


def create_invoice_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoices"
    worksheet.append(
        [
            "\uc218\ud558\uc778\uba85",
            "\uc1a1\uc7a5\ubc88\ud638",
            "\uc8fc\ubb38\ubc88\ud638",
            "\uc218\ud558\uc778\uae30\ubcf8\uc8fc\uc18c",
        ]
    )
    worksheet.append(["\ud64d\uae38\ub3d9", "1234567890", "ORDER-1", "\uc11c\uc6b8"])
    worksheet.append(["\ud64d\uae38\ub3d9", "1234567890", "ORDER-1", "\uc11c\uc6b8"])
    workbook.save(path)


def force_bad_sheet_dimension(path: Path) -> None:
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
        temp_dir = Path(tmp)
        extracted = temp_dir / "xlsx"
        with zipfile.ZipFile(path) as archive:
            archive.extractall(extracted)

        sheet_path = extracted / "xl" / "worksheets" / "sheet1.xml"
        text = sheet_path.read_text(encoding="utf-8")
        text = text.replace('<dimension ref="A1:D3"/>', '<dimension ref="A1"/>')
        sheet_path.write_text(text, encoding="utf-8")

        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for item in extracted.rglob("*"):
                if item.is_file():
                    archive.write(item, item.relative_to(extracted).as_posix())


def create_lotte_source_workbook(path: Path, source_headers: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Source"
    worksheet.append(source_headers)
    worksheet.append([f"테스트{idx}" for idx, _ in enumerate(source_headers, start=1)])
    workbook.save(path)


def create_lotte_template_workbook(path: Path, template_headers: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Template"
    worksheet.append(template_headers)
    worksheet.append(["" for _ in template_headers])

    left_border = Side(style="thin", color="000000")
    for column in range(2, 6):
        cell = worksheet.cell(2, column)
        cell.border = Border(left=left_border)

    workbook.save(path)


class OrderWorkflowRegressionTests(unittest.TestCase):
    def test_delivery_summary_accepts_receiver_headers_in_all_script_copies(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            workbook_path = Path(tmp) / "delivery.xlsx"
            create_delivery_workbook(workbook_path)

            for scripts_dir in MODULE_SETS:
                with self.subTest(scripts_dir=scripts_dir):
                    module = load_module(scripts_dir / "delivery_text_summary.py")
                    summary, sheet_names = module.summarize_workbook(workbook_path)

                    self.assertEqual(sheet_names, ["Supply"])
                    self.assertIn("\ud14c\uc2a4\ud2b8 \uc0c1\ud488 - 2\uac1c (1\uac74)", summary)

    def test_delivery_summary_builds_safe_number_approval_payload_in_all_script_copies(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            workbook_path = Path(tmp) / "delivery-safe-number.xlsx"
            create_safe_number_delivery_workbook(workbook_path)

            for scripts_dir in MODULE_SETS:
                with self.subTest(scripts_dir=scripts_dir):
                    module = load_module(scripts_dir / "delivery_text_summary.py")
                    payload = module.build_summary_payload(workbook_path)

                    self.assertEqual(payload["sheet_names"], ["SafeNumber"])
                    self.assertEqual(len(payload["safe_number_candidates"]), 1)
                    self.assertIn("아이제나흐 에센 스텐 찜솥 28cm - 6개 (1건)", payload["text"])
                    self.assertIn("아이제나흐 에쎈 레인지 쿡 12종/24P - 1개 (1건)", payload["text"])
                    self.assertIn(
                        "★합포장 확인 요청★아이제나흐 에센 스텐 찜솥 28cm - 6개 + 아이제나흐 에쎈 레인지 쿡 12종/24P - 1개",
                        payload["approved_text"],
                    )

    def test_invoice_export_preserves_duplicate_invoice_numbers_in_all_script_copies(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            workbook_path = Path(tmp) / "invoice.xlsx"
            create_invoice_workbook(workbook_path)

            for scripts_dir in MODULE_SETS:
                with self.subTest(scripts_dir=scripts_dir):
                    module = load_module(scripts_dir / "invoice_number_exporter.py")
                    rows = module.extract_invoice_rows(workbook_path)
                    output_path = module.export_invoice_numbers(
                        workbook_path,
                        Path(tmp) / f"invoice-output-{scripts_dir.name}",
                    )
                    workbook = load_workbook(output_path)
                    worksheet = workbook.active

                    self.assertEqual(rows, [("\ud64d\uae38\ub3d9", "1234567890 / 1234567890")])
                    self.assertTrue(output_path.exists())
                    self.assertEqual(worksheet.max_row, 2)
                    self.assertEqual(worksheet["A2"].value, "\ud64d\uae38\ub3d9")
                    self.assertEqual(worksheet["B2"].value, "1234567890 / 1234567890")

    def test_invoice_export_reads_lotte_files_with_bad_sheet_dimension(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            workbook_path = Path(tmp) / "lotte-invoice.xlsx"
            create_invoice_workbook(workbook_path)
            force_bad_sheet_dimension(workbook_path)

            for scripts_dir in MODULE_SETS:
                with self.subTest(scripts_dir=scripts_dir):
                    module = load_module(scripts_dir / "invoice_number_exporter.py")
                    rows = module.extract_invoice_rows(workbook_path)

                    self.assertEqual(rows, [("\ud64d\uae38\ub3d9", "1234567890 / 1234567890")])

    def test_lotte_order_export_removes_left_borders_from_columns_b_to_e(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            temp_dir = Path(tmp)

            for scripts_dir in MODULE_SETS:
                with self.subTest(scripts_dir=scripts_dir):
                    module = load_module(scripts_dir / "lotte_order_form_converter.py")
                    source_path = temp_dir / f"source-{scripts_dir.name}.xlsx"
                    template_path = temp_dir / f"template-{scripts_dir.name}.xlsx"
                    output_dir = temp_dir / f"output-{scripts_dir.name}"

                    create_lotte_source_workbook(source_path, list(module.SOURCE_TO_TEMPLATE.keys()))
                    create_lotte_template_workbook(
                        template_path,
                        list(module.SOURCE_TO_TEMPLATE.values()),
                    )

                    output_path = module.convert_lotte_order_form(
                        source_path,
                        template_path,
                        output_dir,
                        output_date=date(2026, 6, 16),
                    )
                    workbook = load_workbook(output_path)
                    worksheet = workbook.active

                    for column in range(2, 6):
                        self.assertIsNone(
                            worksheet.cell(2, column).border.left.style,
                            f"{worksheet.cell(2, column).coordinate} should not have a left border",
                        )

    def test_vehicle_receipt_template_formatting_matches_print_requirements(self) -> None:
        red_values = {"FFFF0000", "00FF0000", "FF0000"}

        for template_path in TEMPLATE_PATHS:
            with self.subTest(template_path=template_path):
                workbook = load_workbook(template_path)
                worksheet = workbook.active

                for row in range(8, 39):
                    color = worksheet[f"J{row}"].font.color
                    rgb = str(color.rgb).upper() if color and color.type == "rgb" else None
                    self.assertNotIn(rgb, red_values, f"J{row} should not be red")

                b40_color = worksheet["B40"].font.color
                b40_rgb = (
                    str(b40_color.rgb).upper()
                    if b40_color and b40_color.type == "rgb"
                    else None
                )
                self.assertIn(b40_rgb, red_values)

                for row in range(1, 45):
                    self.assertIsNone(
                        worksheet[f"A{row}"].border.left.style,
                        f"A{row} should not have a left border",
                    )

    def test_vehicle_receipt_output_starts_at_main_border_without_extra_left_line(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            temp_dir = Path(tmp)

            for scripts_dir, template_path in zip(MODULE_SETS, TEMPLATE_PATHS):
                with self.subTest(scripts_dir=scripts_dir):
                    module = load_module(scripts_dir / "vehicle_receipt_generator.py")
                    output_path = module.generate_vehicle_receipt(
                        supplier="테스트 거래처",
                        items=[{"product_name": "테스트 제품", "quantity": "1", "pack_quantity": ""}],
                        delivery_place="테스트 납품장소",
                        manager="테스트 담당자",
                        output_dir=temp_dir / scripts_dir.name,
                        template_path=template_path,
                        output_date=date(2026, 6, 16),
                    )
                    workbook = load_workbook(output_path)
                    worksheet = workbook.active

                    self.assertTrue(
                        worksheet.print_area.startswith("'모드니인수증'!$B$1:$M$"),
                        worksheet.print_area,
                    )
                    self.assertTrue(worksheet.column_dimensions["A"].hidden)
                    for row in range(1, 44):
                        self.assertIsNone(
                            worksheet[f"A{row}"].border.left.style,
                            f"A{row} should not have a left border",
                        )
                        self.assertIsNone(
                            worksheet[f"A{row}"].border.right.style,
                            f"A{row} should not have a right border",
                        )


if __name__ == "__main__":
    unittest.main()
