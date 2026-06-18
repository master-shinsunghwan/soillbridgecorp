from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
MODULE_SETS = [
    ROOT / "scripts",
    ROOT / "_workhub_zip_inspect" / "scripts",
]

SUPPLY_HEADERS = [
    "순서",
    "매입거래처",
    "매출거래처",
    "거래구분",
    "장부입력확인",
    "주문일",
    "출고일",
    "주문자",
    "발신자연락처",
    "수령자",
    "수령자연락처",
    "제 품 명",
    "수량",
    "상 세 주 소",
    "택배사**",
    "배송번호",
    "특이(요청)사항",
    "주문상품고유번호",
    "상품코드",
    "주문번호",
    "고객선택옵션",
]


def load_module(module_path: Path):
    spec = importlib.util.spec_from_file_location(module_path.stem, module_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def create_source_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Supply"
    worksheet.append(SUPPLY_HEADERS)
    worksheet.append(
        [
            1,
            "(주)소일브릿지(본사)",
            "토스",
            "",
            "",
            "2026-06-17 09:11:00",
            "",
            "홍길동",
            "010-0000-0000",
            "김서울",
            "010-1111-1111",
            "소일 테스트 상품",
            1,
            "서울시 테스트로 1",
            "CJ대한통운",
            "123",
            "",
            "ITEM-1",
            "P-1",
            "ORDER-1",
            "",
        ]
    )
    worksheet.append(
        [
            2,
            "(주)소일브릿지(본사)",
            "토스",
            "",
            "",
            "2026-06-17 09:12:00",
            "",
            "홍길동",
            "010-0000-0000",
            "이부산",
            "010-2222-2222",
            "소일 테스트 상품",
            1,
            "부산시 테스트로 2",
            "CJ대한통운",
            "124",
            "",
            "ITEM-2",
            "P-1",
            "ORDER-2",
            "",
        ]
    )
    worksheet.append(
        [
            3,
            "탑스미넬(주) ",
            "후후커머스",
            "",
            "",
            "2026-06-17 09:13:00",
            "",
            "홍길동",
            "010-0000-0000",
            "박대구",
            "010-3333-3333",
            "햄튼 컴포트 샤프너",
            1,
            "대구시 테스트로 3",
            "CJ대한통운",
            "125",
            "",
            "ITEM-3",
            "P-2",
            "ORDER-3",
            "",
        ]
    )
    worksheet.append(
        [
            4,
            "(주)소일브릿지(본사)",
            "후후커머스",
            "",
            "",
            "2026-06-18 09:14:00",
            "",
            "홍길동",
            "010-0000-0000",
            "최광주",
            "010-4444-4444",
            "크레마우디 계란말이팬",
            2,
            "광주시 테스트로 4",
            "CJ대한통운",
            "126",
            "",
            "ITEM-4",
            "P-3",
            "ORDER-4",
            "",
        ]
    )
    workbook.save(path)


class SalesVendorSummaryConverterTests(unittest.TestCase):
    def test_sales_vendor_workbook_is_split_into_vendor_sheets_with_summaries(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            source_path = Path(tmp) / "source.xlsx"
            create_source_workbook(source_path)

            for scripts_dir in MODULE_SETS:
                with self.subTest(scripts_dir=scripts_dir):
                    module = load_module(scripts_dir / "sales_vendor_summary_converter.py")
                    output_path = module.convert_sales_vendor_workbook(source_path, Path(tmp) / scripts_dir.name)

                    self.assertEqual(output_path.name, "주소일브릿지_매출처별_정리 2026년 6월 17일.xlsx")
                    workbook = load_workbook(output_path, data_only=True)
                    self.assertEqual(workbook.sheetnames, ["토스", "후후커머스"])

                    toss_sheet = workbook["토스"]
                    self.assertEqual(toss_sheet["A1"].value, "토스 요약")
                    self.assertEqual(toss_sheet["A3"].value, "소일 테스트 상품 - 1개 (2건)")
                    self.assertIn("A1:H1", [str(range_ref) for range_ref in toss_sheet.merged_cells.ranges])
                    self.assertIn("A3:H3", [str(range_ref) for range_ref in toss_sheet.merged_cells.ranges])
                    self.assertEqual([toss_sheet.cell(5, col).value for col in range(1, 22)], SUPPLY_HEADERS)
                    self.assertEqual(toss_sheet["A6"].value, 1)
                    self.assertEqual(toss_sheet["L7"].value, "소일 테스트 상품")

                    huhu_sheet = workbook["후후커머스"]
                    self.assertEqual(huhu_sheet["A1"].value, "후후커머스 요약")
                    self.assertEqual(
                        huhu_sheet["A3"].value,
                        "★탑스미넬(주) 매입건★\n햄튼 컴포트 샤프너 - 1개 (1건)\n\n크레마우디 계란말이팬 - 2개 (1건)",
                    )
                    self.assertIn("A3:H6", [str(range_ref) for range_ref in huhu_sheet.merged_cells.ranges])
                    self.assertEqual([huhu_sheet.cell(8, col).value for col in range(1, 22)], SUPPLY_HEADERS)
                    self.assertEqual(huhu_sheet["A9"].value, 3)
                    self.assertEqual(huhu_sheet["A10"].value, 4)


if __name__ == "__main__":
    unittest.main()
