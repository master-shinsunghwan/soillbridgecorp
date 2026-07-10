from __future__ import annotations

import importlib
import os
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))


class WorkhubAppFeatureParityTests(unittest.TestCase):
    def load_app(self):
        for module_name in ("workhub_delivery_app", "workhub_crm"):
            sys.modules.pop(module_name, None)
        tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(tempdir.cleanup)
        os.environ["WORKHUB_DATA_DIR"] = tempdir.name
        return importlib.import_module("workhub_delivery_app")

    def test_root_app_exposes_crm_feature_set_from_packaged_app(self) -> None:
        crm = importlib.import_module("workhub_crm")
        app = self.load_app()

        permission_keys = {key for key, _, _ in app.PERMISSION_DEFINITIONS}

        self.assertTrue(hasattr(crm, "crm_dashboard_payload"))
        self.assertIn("crm_view", permission_keys)
        self.assertIn("crm_manage", permission_keys)
        self.assertIn("crm_message_manage", permission_keys)
        self.assertIn("crm_view", app.DEFAULT_ROLE_PERMISSIONS["user"])
        self.assertIn("sub_admin", app.DEFAULT_ROLE_PERMISSIONS)
        self.assertIn('data-open="crm"', app.render_app_html({"display_name": "???", "role": "admin", "permissions": app.ALL_PERMISSIONS}))

    def test_order_workspace_cards_open_existing_modals(self) -> None:
        html_source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")

        for mode in ("delivery", "invoice", "lotte", "salesVendor", "vehicle"):
            self.assertIn(f"{mode}: {{", html_source)

        self.assertIn("ORDER_MODAL_MODES", html_source)
        self.assertIn("ORDER_MODAL_TITLES", html_source)
        self.assertIn("function openOrderModal(mode)", html_source)
        self.assertIn('data-open="order"', html_source)
        self.assertIn('data-order-card="delivery"', html_source)
        self.assertIn('data-order-execute="delivery"', html_source)
        self.assertIn('data-order-card="salesVendor"', html_source)
        self.assertIn('data-order-execute="salesVendor"', html_source)
        self.assertIn('/api/sales-vendor-summary', html_source)
        self.assertIn('data-icon="📦"', html_source)
        self.assertIn('data-icon="🔎"', html_source)
        self.assertIn('data-icon="▦"', html_source)
        self.assertIn('data-icon="🚚"', html_source)
        for mode in ("delivery", "invoice", "lotte", "salesVendor", "vehicle"):
            self.assertNotIn(f'data-order-card="{mode}" role="button" tabindex="0"', html_source)
        self.assertNotIn("onclick=\"openOrderModal('delivery')\"", html_source)
        self.assertNotIn("onclick=\"event.stopPropagation();openOrderModal('delivery');\"", html_source)
        self.assertIn("setActiveNav(\"order\")", html_source)
        self.assertIn("setPageTitle(ORDER_MODAL_TITLES[currentOrderMode]", html_source)
        self.assertIn('sidebar.addEventListener("click"', html_source)
        self.assertIn('event.target.closest("[data-open]")', html_source)
        self.assertIn('document.addEventListener("keydown"', html_source)
        self.assertIn("event.stopImmediatePropagation()", html_source)

    def test_dashboard_entry_shows_notice_import_schedule_and_calendar(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn('id="sidebarNoticePreview"', html_source)
            self.assertIn('id="dashboardImportScheduleCard"', html_source)
            self.assertIn('data-company-panel="calendar"', html_source)
            self.assertIn('class="company-panel active dashboard-calendar-panel" data-company-panel="calendar"', html_source)
            notice_index = html_source.index('id="sidebarNoticePreview"')
            import_index = html_source.index('id="dashboardImportScheduleCard"')
            calendar_index = html_source.index('data-company-panel="calendar"')
            self.assertLess(notice_index, import_index)
            self.assertLess(import_index, calendar_index)
            self.assertIn("수입제품 입고 일정", html_source)
            self.assertIn('id="dashboardImportScheduleSummary"', html_source)
            self.assertIn('id="dashboardImportScheduleBody"', html_source)
            self.assertIn('class="company-card dashboard-calendar-card"', html_source)
            self.assertIn('id="dashboardSalesPanel"', html_source)
            self.assertIn("매출 현황", html_source)
            sales_panel_start = html_source.index('id="dashboardSalesPanel"')
            sales_panel_end = html_source.index('</aside>', sales_panel_start)
            sales_panel_html = html_source[sales_panel_start:sales_panel_end]
            self.assertIn("매출현황 및 관리 데이터 연결 대기 중", sales_panel_html)
            self.assertNotIn("선택한 날짜", sales_panel_html)
            self.assertNotIn("이번 달 요약", sales_panel_html)
            self.assertIn(".dashboard-calendar-panel", html_source)
            self.assertIn("grid-template-columns: minmax(0, 1fr) minmax(0, 1fr)", html_source)
            self.assertIn("function renderDashboardImportSchedule()", html_source)
            self.assertIn("renderDashboardImportSchedule();", html_source)
            self.assertIn('companyActiveTab === "notice" && panel.dataset.companyPanel === "calendar"', html_source)
            self.assertIn("loadDashboardEntryData().catch", html_source)
        app_html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")
        self.assertIn('id="automationOverviewCard"', app_html_source)
        self.assertIn('id="automationOverviewStatus"', app_html_source)
        self.assertIn('id="automationOverviewBody"', app_html_source)
        self.assertIn('/api/automation-overview', app_html_source)
        self.assertIn('function renderAutomationOverview(data = {})', app_html_source)
        self.assertIn("업무 자동화 점검", app_html_source)

    def test_portal_notice_records_are_stored_in_server_db(self) -> None:
        app = self.load_app()
        app.init_db()

        saved = app.save_portal_notice(
            {
                "date": "2026-06-26",
                "title": "업무 포털 업데이트 안내",
                "owner": "관리자",
                "body": "공지사항 서버 저장 테스트",
            },
            {"username": "admin", "display_name": "Admin"},
        )
        rows = app.list_portal_notices()

        self.assertEqual(rows[0]["id"], saved["id"])
        self.assertEqual(rows[0]["date"], "2026-06-26")
        self.assertEqual(rows[0]["title"], "업무 포털 업데이트 안내")
        self.assertEqual(rows[0]["body"], "공지사항 서버 저장 테스트")
        self.assertEqual(app.delete_portal_notice(int(saved["id"])), 1)
        self.assertEqual(app.list_portal_notices(), [])

    def test_sidebar_navigation_uses_hierarchical_text_colors(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        for color_token in (
            "--sidebar-text-primary",
            "--sidebar-text-secondary",
            "--sidebar-text-muted",
            "--sidebar-text-subtle",
            "--sidebar-accent",
        ):
            self.assertIn(color_token, html_source)
        self.assertIn(".nav-item .nav-label span", html_source)
        self.assertIn(".nav-item svg", html_source)
        self.assertIn('.nav-item[data-nav-tone="home"]', html_source)
        self.assertIn('.nav-item[data-nav-tone="management"]', html_source)
        self.assertIn('.nav-item[data-nav-tone="cs"]', html_source)
        self.assertIn('.nav-item[data-nav-tone="admin"]', html_source)
        self.assertIn('data-nav-tone="home"', html_source)
        self.assertIn('data-nav-tone="management"', html_source)
        self.assertIn('data-nav-tone="cs"', html_source)
        self.assertIn(".nav-subitem.active", html_source)
        self.assertIn("color: var(--sidebar-text-primary)", html_source)
        self.assertIn("color: var(--sidebar-text-secondary)", html_source)
        self.assertIn("color: var(--sidebar-text-muted)", html_source)

    def test_company_portal_click_always_opens_submenu_after_workspace_switch(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("const companyGroup = document.querySelector(\"#companyNavGroup\")", html_source)
        self.assertIn("companyGroup?.classList.add(\"open\")", html_source)
        self.assertNotIn("wasCompanyGroupOpen", html_source)
        self.assertNotIn('companyGroup?.classList.toggle("open"', html_source)
        self.assertNotIn('document.querySelector("#companyNavGroup").classList.toggle("open");', html_source)

    def test_ledger_sidebar_groups_open_directly_without_upload_subtrees(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        management_group_start = html_source.index('id="managementNavGroup"')
        ledger_group_start = html_source.index('id="ledgerNavGroup"')
        crm_group_start = html_source.index('id="crmNavGroup"')
        admin_group_start = html_source.index('id="adminNavGroup"')
        admin_group_end = html_source.index('id="leaveWorkspace"', admin_group_start)
        management_group = html_source[management_group_start:ledger_group_start]
        ledger_group = html_source[ledger_group_start:crm_group_start]
        admin_group = html_source[admin_group_start:admin_group_end]

        self.assertIn('data-open="management"', management_group)
        self.assertNotIn('id="managementImportOpen"', management_group)
        self.assertNotIn('data-management-import-mode="daily"', management_group)
        self.assertNotIn("통합관리대장 업로드", management_group)
        self.assertNotIn('class="nav-submenu"', management_group)

        self.assertIn('data-open="ledger"', ledger_group)
        self.assertNotIn('id="ledgerImportOpen"', ledger_group)
        self.assertNotIn('data-ledger-import-mode="daily"', ledger_group)
        self.assertNotIn("CS처리대장 업로드", ledger_group)
        self.assertNotIn('data-mail-popup="cs"', ledger_group)
        self.assertNotIn("CS처리 요청", ledger_group)
        self.assertNotIn('class="nav-submenu"', ledger_group)

        self.assertIn('data-management-import-mode="replace"', admin_group)
        self.assertIn('data-ledger-import-mode="replace"', admin_group)
        self.assertNotIn('data-management-import-mode="daily"', admin_group)
        self.assertNotIn('data-ledger-import-mode="daily"', admin_group)
        self.assertIn('#managementNavToggle', html_source)
        self.assertIn('#ledgerNavToggle', html_source)

    def test_management_and_cs_ledgers_default_to_500_rows(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        ledger_size_start = html_source.index('id="ledgerPageSize"')
        ledger_size_end = html_source.index("</select>", ledger_size_start)
        ledger_size_slice = html_source[ledger_size_start:ledger_size_end]
        management_size_start = html_source.index('id="managementPageSize"')
        management_size_end = html_source.index("</select>", management_size_start)
        management_size_slice = html_source[management_size_start:management_size_end]

        self.assertIn('<option value="500" selected>500개씩 보기</option>', ledger_size_slice)
        self.assertNotIn('<option value="1000" selected>', ledger_size_slice)
        self.assertIn('<option value="500" selected>500개씩 보기</option>', management_size_slice)
        self.assertNotIn('<option value="1000" selected>', management_size_slice)
        self.assertIn('new URLSearchParams({ limit: ledgerPageSize.value || "500" })', html_source)
        self.assertIn("const hasColorFilters = managementHasActiveColorFilters();", html_source)
        self.assertIn('new URLSearchParams({ limit: (hasColumnFilters || hasColorFilters) ? "50000" : (managementPageSize.value || "500") })', html_source)
        self.assertIn('ledgerPageSize.value = "500";', html_source)
        self.assertEqual(html_source.count('managementPageSize.value = "500";'), 2)
        self.assertNotIn('managementPageSize.value = "1000";', html_source)
        self.assertIn('int(params.get("limit", ["500"])[0])', html_source)
        self.assertEqual(html_source.count('limit = 500'), 2)

    def test_management_workspace_navigation_preserves_selected_period(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        workspace_start = html_source.index('if (showManagement) {\n        setPageTitle("통합관리대장 관리");')
        workspace_end = html_source.index('} else if (showLedger) {', workspace_start)
        workspace_block = html_source[workspace_start:workspace_end]

        self.assertNotIn('managementYearFilter.value = "";', workspace_block)
        self.assertNotIn('managementMonthFilter.value = "";', workspace_block)
        self.assertIn("loadManagementWorkspaceData();", workspace_block)

        modal_start = html_source.index('modalTitle.textContent = "통합관리대장 관리";')
        modal_end = html_source.index('const fileDrop = document.querySelector("label[for=\'fileInput\']");', modal_start)
        modal_block = html_source[modal_start:modal_end]

        self.assertIn('managementYearFilter.value = "";', modal_block)
        self.assertIn('managementMonthFilter.value = "";', modal_block)

    def test_cs_request_from_ledger_menu_supports_mail_attachments(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        ledger_group_start = html_source.index('id="ledgerNavGroup"')
        crm_group_start = html_source.index('id="crmNavGroup"')
        ledger_group = html_source[ledger_group_start:crm_group_start]

        self.assertNotIn('data-mail-popup="cs"', ledger_group)
        self.assertNotIn("CS처리 요청", ledger_group)
        self.assertIn('id="csAttachmentInput"', html_source)
        self.assertIn('id="csAttachmentDropzone"', html_source)
        self.assertIn('id="csAttachmentChoose"', html_source)
        self.assertIn('id="csAttachmentDropMain"', html_source)
        self.assertIn('id="csAttachmentSummary"', html_source)
        self.assertIn('accept="image/*,video/*,.pdf,.xlsx,.xls,.doc,.docx,.zip"', html_source)
        self.assertIn("setupDropzone(csAttachmentDropzone", html_source)
        self.assertIn('csAttachmentChoose?.addEventListener("click"', html_source)
        self.assertIn("파일 선택", html_source)
        self.assertIn("드래그", html_source)
        self.assertIn('id="sendCsMailButton"', html_source)
        self.assertIn("async function sendCurrentCsMail()", html_source)
        self.assertIn('sendCsMailButton?.addEventListener("click"', html_source)
        self.assertIn("appendCsMailPayload", html_source)
        self.assertIn("collect_mail_attachments", html_source)
        self.assertIn("attachments=attachments", html_source)

    def test_ledger_cs_popup_and_cell_edit_activation_are_workspace_safe(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("function mountCsFieldsInLedgerWorkspace()", html_source)
        self.assertIn("ledgerWorkspaceMount.appendChild(csFields)", html_source)
        self.assertIn("#ledgerWorkspace .cs-fields.ledger-cs-popup", html_source)
        self.assertIn("position: fixed;", html_source)
        self.assertIn("transform: translate(-50%, -50%);", html_source)
        self.assertIn("grid-template-columns: repeat(2, minmax(0, 1fr));", html_source)
        self.assertIn("text-field cs-wide", html_source)
        self.assertIn("restoreCsFieldsToModal();", html_source)

        self.assertIn('tabindex="0"', html_source)
        self.assertIn("function selectEditableCell(scope, cell)", html_source)
        self.assertIn('managementBody.addEventListener("dblclick"', html_source)
        self.assertIn('ledgerBody.addEventListener("dblclick"', html_source)
        self.assertIn("function handleEditableCellNavigation(scope, event)", html_source)
        self.assertIn('data-ledger-edit-row', html_source)
        management_click_start = html_source.index('managementBody.addEventListener("click"')
        management_click_end = html_source.index('managementBody.addEventListener("dblclick"', management_click_start)
        ledger_click_start = html_source.index('ledgerBody.addEventListener("click"')
        ledger_click_end = html_source.index('ledgerBody.addEventListener("dblclick"', ledger_click_start)
        self.assertNotIn('openCellEditor("management"', html_source[management_click_start:management_click_end])
        self.assertIn('const editButton = event.target.closest("[data-ledger-edit-row]");', html_source[ledger_click_start:ledger_click_end])
        self.assertIn('if (editableCell) openCellEditor("ledger", editableCell);', html_source[ledger_click_start:ledger_click_end])
        ledger_cell_click = html_source[
            html_source.index('const editableCell = event.target.closest(".editable-cell[data-field]");', ledger_click_start):ledger_click_end
        ]
        self.assertNotIn('openCellEditor("ledger"', ledger_cell_click)

    def test_sheet_like_tables_support_drag_selection_and_numeric_summary(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("sheetRangeSelection", html_source)
        self.assertIn("sheet-selection-summary", html_source)
        self.assertIn("function applySheetRangeSelection(scope, anchor, current)", html_source)
        self.assertIn("function updateSheetSelectionSummary(cells)", html_source)
        self.assertIn("function numericValueFromSheetCell(cell)", html_source)
        self.assertIn("합계 ${formatSheetNumber(sum)}", html_source)
        self.assertIn("평균 ${formatSheetNumber(average)}", html_source)
        self.assertIn('managementBody.addEventListener("mousedown"', html_source)
        self.assertIn('ledgerBody.addEventListener("mousedown"', html_source)
        self.assertIn('managementBody.addEventListener("mousemove"', html_source)
        self.assertIn('ledgerBody.addEventListener("mousemove"', html_source)
        self.assertIn('document.addEventListener("mouseup", finishSheetRangeSelection)', html_source)

    def test_crm_daily_work_logs_are_available_from_work_management(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")
        crm_source = (ROOT / "scripts" / "workhub_crm.py").read_text(encoding="utf-8")

        self.assertIn('data-crm-nav-tab="daily"', html_source)
        self.assertIn('id="crmTabDaily"', html_source)
        self.assertIn('id="crmPanelDaily"', html_source)
        self.assertIn('id="crmDailyLogForm"', html_source)
        self.assertIn('id="crmDailyLogBody"', html_source)
        self.assertIn("crm-daily-widget", html_source)
        self.assertIn("crm-daily-log-list", html_source)
        self.assertIn("crm-daily-log-card", html_source)
        self.assertIn("grid-template-columns: repeat(auto-fit, minmax(min(100%, 380px), 1fr));", html_source)
        self.assertIn("function crmDailyReadableText", html_source)
        self.assertIn("function crmDailyNeedsToggle", html_source)
        self.assertIn("function crmDailyDisplayLines", html_source)
        self.assertIn("function crmDailyLineListHtml", html_source)
        self.assertIn("function crmDailyDomToken", html_source)
        self.assertIn("data-crm-daily-toggle", html_source)
        self.assertIn('aria-controls="${safeTextId}"', html_source)
        self.assertIn("crm-daily-log-line-list", html_source)
        self.assertIn("crm-daily-log-line-index", html_source)
        self.assertIn(".crm-daily-log-card.has-entry", html_source)
        self.assertIn("전체보기", html_source)
        self.assertIn("white-space: pre-line;", html_source)
        self.assertIn("overflow-wrap: anywhere;", html_source)
        self.assertIn("word-break: normal;", html_source)
        self.assertIn("오늘 한 일", html_source)
        self.assertIn("내일 할 일", html_source)
        self.assertNotIn('id="crmDailyFormSummary"', html_source)
        self.assertNotIn('id="crmDailyDraftButton"', html_source)
        self.assertIn("/api/crm-daily-logs", html_source)
        self.assertIn("/api/crm-daily-log-save", html_source)
        self.assertIn("function renderCrmDailyLogs", html_source)
        self.assertIn("function saveCrmDailyLogForm", html_source)
        self.assertIn("crm_daily_logs", crm_source)
        self.assertIn("completed_today_tasks", crm_source)
        self.assertIn("open_task_samples", crm_source)
        self.assertIn("def list_crm_daily_logs", crm_source)
        self.assertIn("def save_crm_daily_log", crm_source)

    def test_ledger_completed_rows_ignore_whitespace_variants_for_yellow_highlight(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("function isOverallCompletedStatus(statusValue)", html_source)
        self.assertIn('status.includes("전체") && status.includes("처리") && status.includes("완료")', html_source)
        self.assertIn("if (isOverallCompletedStatus(statusValue)) return true;", html_source)
        self.assertIn("const type = normalizedLedgerText(typeValue);", html_source)
        self.assertIn("const status = normalizedLedgerText(statusValue);", html_source)
        self.assertIn("function isLedgerCompletedCase(csCase)", html_source)
        self.assertIn('if (status.includes("완료")) return true;', html_source)
        self.assertIn("if (String(csCase.completed_at || '').trim()) return true;", html_source)
        self.assertIn(".ledger-table tbody tr.completed-cs td", html_source)
        self.assertIn("background: #fff8d8 !important;", html_source)
        self.assertIn('field: "completed_at", label: "완료일", value: csCase.completed_at', html_source)
        self.assertIn("const completedAt = fieldValue(row.querySelector('[data-field=\"completed_at\"]'));", html_source)
        self.assertIn('row.classList.add("completed-cs")', html_source)
        self.assertIn('row.classList.toggle("completed-cs", isCompletedByValues(csType, status) || Boolean(completedAt));', html_source)

    def test_cs_followup_alert_uses_spacious_cards_and_direct_row_shortcuts(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn(".app-confirm.wide", html_source)
        self.assertIn(".cs-followup-alert-card", html_source)
        self.assertIn('data-cs-followup-open="${escapeHtml(item.id)}"', html_source)
        self.assertIn("function openCsCaseFromAlert(caseId)", html_source)
        self.assertIn('if (ledgerPageSize) ledgerPageSize.value = "5000";', html_source)
        self.assertIn('tr[data-case-id="${CSS.escape(id)}"]', html_source)
        self.assertIn("openCsCaseFromAlert(followupButton.dataset.csFollowupOpen)", html_source)
        self.assertIn("highlightHtml: `<div class=\"cs-followup-alert-list\">${cards.join(\"\")}</div>`", html_source)
        self.assertIn("wide: true", html_source)

    def test_desktop_launchers_register_user_startup_on_first_run(self) -> None:
        vps_launcher = (ROOT / "scripts" / "workhub_vps_desktop_app.py").read_text(encoding="utf-8")
        local_launcher = (ROOT / "scripts" / "workhub_desktop_launcher.py").read_text(encoding="utf-8")
        desktop_installer = (ROOT / "install_workhub_desktop_app.ps1").read_text(encoding="utf-8")
        local_installer = (ROOT / "install_workhub_app.ps1").read_text(encoding="utf-8")
        desktop_uninstaller = (ROOT / "uninstall_workhub_desktop_app.ps1").read_text(encoding="utf-8")
        local_uninstaller = (ROOT / "uninstall_workhub_app.ps1").read_text(encoding="utf-8")

        self.assertIn("def register_startup_launch()", vps_launcher)
        self.assertIn("SoilbridgeWorkhubDesktop_AutoStart.vbs", vps_launcher)
        self.assertIn("register_startup_launch()", vps_launcher)
        self.assertIn("def register_startup_launch()", local_launcher)
        self.assertIn("Workhub_AutoStart.vbs", local_launcher)
        self.assertIn("register_startup_launch()", local_launcher)
        self.assertIn("WORKHUB_DESKTOP_DISABLE_AUTOSTART", vps_launcher)
        self.assertIn("WORKHUB_DESKTOP_DISABLE_AUTOSTART", local_launcher)
        self.assertIn("SoilbridgeWorkhubDesktop_AutoStart.vbs", desktop_installer)
        self.assertIn("Workhub_AutoStart.vbs", local_installer)
        self.assertIn("SoilbridgeWorkhubDesktop_AutoStart.vbs", desktop_uninstaller)
        self.assertIn("Workhub_AutoStart.vbs", local_uninstaller)

    def test_management_duplicate_rows_override_even_row_striping(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn(".ledger-table tbody tr:nth-child(even) td", html_source)
        self.assertIn(".ledger-table tbody tr.management-duplicate td", html_source)
        self.assertIn("background-color: var(--duplicate-row-color, #eef6ff);", html_source)
        self.assertIn('row.classList.add("management-duplicate")', html_source)

    def test_hermes_workspace_menu_and_configurable_agent_bridge_exist(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="hermesNavGroup"', html_source)
        self.assertIn('data-open="hermes"', html_source)
        self.assertIn("AI 업무채팅", html_source)
        self.assertIn("자동화 요청", html_source)
        self.assertIn("작업내역", html_source)
        self.assertIn("관리자 설정", html_source)
        self.assertIn('id="hermesWorkspace"', html_source)
        self.assertIn('id="hermesBaseUrl"', html_source)
        self.assertIn('id="hermesConnectionTest"', html_source)
        self.assertIn("/api/hermes-status", html_source)
        self.assertIn("/api/hermes-chat", html_source)
        self.assertIn("/api/hermes-automation", html_source)
        self.assertIn('data-hermes-chat-mode="auto"', html_source)
        self.assertIn('data-hermes-chat-mode="automation"', html_source)
        self.assertIn('data-hermes-chat-mode="general"', html_source)
        self.assertIn('data-hermes-chat-mode="search"', html_source)
        self.assertIn('data-hermes-chat-mode="image"', html_source)
        self.assertIn("Workhub 업무 자동화와 분리해서 일반 Codex/GPT 영역으로 답합니다.", html_source)
        self.assertIn("생성된 파일은 먼저 다운로드 링크를 제공하고, 업무파일 저장은 승인 후에만 처리합니다.", html_source)
        self.assertIn("function setHermesChatMode(mode)", html_source)
        self.assertIn('effective_mode not in {"general", "search", "image"}', html_source)
        self.assertIn('"workhub_context": workhub_context', html_source)
        self.assertIn("body: JSON.stringify({ message, mode: requestedMode })", html_source)
        self.assertIn("save_hermes_text_result", html_source)
        self.assertIn("generated_text_file", html_source)
        self.assertIn("hermes-result-link", html_source)
        self.assertIn("HERMES_SETTINGS_PATH", html_source)
        self.assertIn("hermes_request", html_source)
        self.assertIn("hermes_use", html_source)
        self.assertIn("hermes_automation", html_source)
        self.assertIn("hermes_admin", html_source)
        self.assertIn("hermes-mark", html_source)
        self.assertIn("/static/hermes-icon.png", html_source)
        self.assertIn('data-hermes-preset="vps"', html_source)
        self.assertIn('data-hermes-preset="local"', html_source)
        self.assertIn("HERMES_PROFILE", html_source)
        self.assertIn("describe_hermes_connection_error", html_source)
        self.assertIn('api_key.lower().startswith(("basic ", "bearer "))', html_source)
        self.assertIn("승인형 Workhub 실행 API", html_source)
        self.assertIn('id="hermesActionTaskTitle"', html_source)
        self.assertIn("/api/hermes-actions", html_source)
        self.assertIn("/api/hermes-action-preview", html_source)
        self.assertIn("/api/hermes-action-execute", html_source)
        self.assertIn("execute_hermes_workhub_action", html_source)

    def test_hermes_chat_history_can_be_summarized_and_filtered(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="hermesSummaryCreate"', html_source)
        self.assertIn('id="hermesHistoryFilter"', html_source)
        self.assertIn('id="hermesSummaryList"', html_source)
        self.assertIn("/api/hermes-summary", html_source)
        self.assertIn("summarize_hermes_chat_history", html_source)
        self.assertIn("renderHermesSummaries", html_source)
        self.assertIn("createHermesSummary", html_source)
        self.assertIn('kind: "summary"', html_source)
        self.assertEqual(html_source.count("function renderHermesHistory(items = [])"), 1)

    def test_hermes_chat_uses_half_width_answer_and_quick_actions(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("hermes-chat-layout", html_source)
        self.assertIn("hermes-side-grid", html_source)
        self.assertIn("hermes-quick-actions", html_source)
        self.assertIn('data-hermes-quick="summary"', html_source)
        self.assertIn('data-hermes-quick="history"', html_source)
        self.assertIn('data-hermes-quick="automation"', html_source)
        self.assertIn('data-hermes-quick="settings"', html_source)
        self.assertIn('setHermesTab("history")', html_source)
        self.assertIn('setHermesTab("automation")', html_source)
        self.assertIn('setHermesTab("settings")', html_source)

    def test_dashboard_sales_amounts_show_full_numbers_not_million_suffix(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        start = html_source.index("function formatSalesMillion")
        end = html_source.index("function formatSalesPercent", start)
        formatter_slice = html_source[start:end]
        self.assertIn('Math.abs(number).toLocaleString("ko-KR")', formatter_slice)
        self.assertIn('}원`', formatter_slice)
        self.assertNotIn("백만", formatter_slice)

    def test_admin_navigation_and_admin_pages_can_scroll(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("height: 100vh;", html_source)
        self.assertIn("max-height: 100vh;", html_source)
        self.assertIn("overscroll-behavior: contain;", html_source)
        self.assertIn("#userAdminWorkspace.active", html_source)
        self.assertIn("#backupWorkspace.active", html_source)
        self.assertIn("#systemUpdateWorkspace.active", html_source)
        self.assertIn("main.workspace-scroll-mode", html_source)
        self.assertIn("main:has(#userAdminWorkspace.active)", html_source)
        self.assertIn("main:has(#backupWorkspace.active)", html_source)
        self.assertIn("main:has(#systemUpdateWorkspace.active)", html_source)
        self.assertIn('"workspace-scroll-mode"', html_source)
        self.assertIn("showUserAdmin || showBackup || showSystemUpdate", html_source)
        self.assertIn("#userAdminPermissions.permission-grid", html_source)
        self.assertIn("max-height: min(40vh, 380px);", html_source)
        self.assertIn("#userAdminWorkspace .admin-table-wrap", html_source)
        self.assertIn("max-height: min(44vh, 460px);", html_source)
        self.assertIn("position: sticky;", html_source)

    def test_user_admin_can_delete_edit_and_review_deleted_history(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("/api/users-delete", html_source)
        self.assertIn("delete_user_account", html_source)
        self.assertIn("deleted_user_accounts", html_source)
        self.assertIn("list_deleted_user_accounts", html_source)
        self.assertIn('id="userAdminDeletedBody"', html_source)
        self.assertIn('data-user-edit="${user.id}"', html_source)
        self.assertIn('data-user-delete="${user.id}"', html_source)
        self.assertIn("비밀번호는 보안상 확인할 수 없으며", html_source)
        self.assertNotIn("password_hash TEXT", html_source[html_source.index("CREATE TABLE IF NOT EXISTS deleted_user_accounts"):html_source.index("CREATE TABLE IF NOT EXISTS shared_files")])

    def test_ledger_arrow_keys_move_between_cells_before_editing(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("function handleEditableCellNavigation(scope, event)", html_source)
        self.assertIn("function moveEditableCell(scope, cell, rowDelta, colDelta)", html_source)
        self.assertIn("ArrowUp", html_source)
        self.assertIn("ArrowDown", html_source)
        self.assertIn("ArrowLeft", html_source)
        self.assertIn("ArrowRight", html_source)
        self.assertIn("event.preventDefault();", html_source)
        self.assertIn('handleEditableCellNavigation("management", event);', html_source)
        self.assertIn('handleEditableCellNavigation("ledger", event);', html_source)
        self.assertIn('event.key === "F2"', html_source)

    def test_cell_editor_apply_saves_and_refreshes_visible_ledgers(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("async function saveEditedCellRow(scope, row)", html_source)
        self.assertIn("function scheduleCellEditorAutoApply(scope)", html_source)
        self.assertIn("function commitCellEditorOnChange(scope)", html_source)
        self.assertIn("control.addEventListener(\"input\", () => scheduleCellEditorAutoApply(scope));", html_source)
        self.assertIn("control.addEventListener(\"change\", () => commitCellEditorOnChange(scope));", html_source)
        self.assertIn("control.addEventListener(\"blur\", () => commitCellEditorOnChange(scope));", html_source)
        apply_start = html_source.index("async function applyCellEditor(scope")
        apply_end = html_source.index("function dirtyRows", apply_start)
        apply_block = html_source[apply_start:apply_end]
        self.assertIn("await saveEditedCellRow(scope, row);", apply_block)
        self.assertIn('notice.textContent = scope === "management"', apply_block)

        save_start = html_source.index("async function saveEditedCellRow(scope, row)")
        save_end = html_source.index("function dirtyRows", save_start)
        save_block = html_source[save_start:save_end]
        self.assertIn("await saveManagementPayload(payload);", save_block)
        self.assertIn("updateManagementRecordCache(payload);", save_block)
        self.assertIn("applyManagementFilters();", save_block)
        self.assertIn("await saveLedgerPayload(payload);", save_block)
        self.assertIn("updateLedgerCaseCache(payload);", save_block)
        self.assertIn("applyLedgerFilters();", save_block)

    def test_ledger_rows_expose_visible_edit_button_and_order_field_updates(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('data-ledger-edit-row', html_source)
        self.assertIn('openCellEditor("ledger", editableCell);', html_source)
        self.assertIn('sales_vendor: fieldValue(row.querySelector(\'[data-field="sales_vendor"]\'))', html_source)
        self.assertIn('purchase_vendor: fieldValue(row.querySelector(\'[data-field="purchase_vendor"]\'))', html_source)
        self.assertIn('order_date: fieldValue(row.querySelector(\'[data-field="order_date"]\'))', html_source)
        self.assertIn('receiver_address: fieldValue(row.querySelector(\'[data-field="receiver_address"]\'))', html_source)
        self.assertIn('original_invoice: fieldValue(row.querySelector(\'[data-field="original_invoice"]\'))', html_source)
        self.assertIn("CS_CASE_UPDATE_FIELDS", html_source)
        self.assertIn('"receiver_address"', html_source)
        self.assertIn('"original_invoice"', html_source)

    def test_checked_rows_can_receive_selected_cell_value_in_bulk(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="managementBulkApply"', html_source)
        self.assertIn('id="ledgerBulkApply"', html_source)
        self.assertIn("async function applySelectedCellToCheckedRows(scope)", html_source)
        self.assertIn("selectedRows(body, rowSelector)", html_source)
        self.assertIn("const fieldSelector = scope === \"management\"", html_source)
        self.assertIn("setEditableCellValue(targetCell, value);", html_source)
        self.assertIn("await saveCurrentWorkspaceRows({ mode: scope, selectedOnly: true });", html_source)
        self.assertIn('managementBulkApply.addEventListener("click"', html_source)
        self.assertIn('ledgerBulkApply.addEventListener("click"', html_source)

    def test_management_manual_add_uses_popup_workflow(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="managementAddManual"', html_source)
        self.assertIn('id="managementManualFields"', html_source)
        self.assertIn('id="managementManualClose"', html_source)
        self.assertIn('id="saveManagementManual"', html_source)
        self.assertIn("function openManagementManualPopup()", html_source)
        self.assertIn("function closeManagementManualPopup()", html_source)
        self.assertIn("async function saveManagementManualRecord()", html_source)
        self.assertIn('id="manualManagementReceiverAddress" data-management-manual-field="receiver_address" rows="2"', html_source)
        self.assertIn('id="manualManagementMemo" data-management-manual-field="memo" rows="2"', html_source)
        self.assertIn(".management-manual-fields #manualManagementReceiverAddress", html_source)
        self.assertIn(".management-manual-fields #manualManagementMemo", html_source)
        self.assertIn("height: 52px;", html_source)
        self.assertIn('fetch("/api/management-record-create"', html_source)
        self.assertIn('managementAddManual.addEventListener("click"', html_source)

    def test_workhub_uses_in_app_dialogs_instead_of_native_browser_dialogs(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("function requestAppConfirm", html_source)
        self.assertNotIn("window.alert(", html_source)
        self.assertNotIn("window.confirm(", html_source)
        self.assertNotIn("if (!confirm(", html_source)
        self.assertNotIn("alert(error.message)", html_source)

    def test_automation_center_popup_and_apis_exist(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="automationCenterOpen"', html_source)
        self.assertIn('id="automationCenterPopup"', html_source)
        self.assertIn("automation-center-popup", html_source)
        self.assertIn("automation-center-body", html_source)
        self.assertIn('id="automationLogBody"', html_source)
        self.assertIn("/api/automation-operation-logs", html_source)
        self.assertIn("function canViewAutomationCenter()", html_source)
        self.assertIn("can_view_automation_center", html_source)
        self.assertIn("/api/automation-center", html_source)
        self.assertIn("/api/automation-center-preview", html_source)
        self.assertIn("/api/automation-center-execute", html_source)
        self.assertIn("automation_operation_logs", html_source)
        self.assertIn("preview_management_rules", html_source)
        self.assertIn("execute_bulk_db_change", html_source)
        self.assertIn("execute_cs_bulk_mail", html_source)
        self.assertIn("execute_notice_auto", html_source)
        self.assertIn("list_automation_operation_logs", html_source)

    def test_excel_style_filter_reset_controls_exist_for_ledgers(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="ledgerFilterResetAll"', html_source)
        self.assertIn("function clearActiveLedgerFilter()", html_source)
        self.assertIn("function clearAllLedgerFilters()", html_source)
        self.assertIn("function clearAllManagementFilters()", html_source)
        self.assertIn("delete ledgerFilters[activeLedgerFilterField];", html_source)
        self.assertIn("delete managementFilters[activeManagementFilterField];", html_source)
        self.assertIn("Object.keys(ledgerFilters).forEach", html_source)
        self.assertIn("Object.keys(managementFilters).forEach", html_source)
        self.assertIn('id="ledgerFilterScope"', html_source)
        self.assertIn("월 전체 데이터에서 필터 후보를 보여주고", html_source)
        self.assertIn('ledgerFilterResetAll.addEventListener("click"', html_source)

    def test_ledger_filter_options_follow_other_active_filters(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("function matchesLedgerFiltersExcept(csCase, excludedField = \"\")", html_source)
        self.assertIn("function matchesManagementFiltersExcept(record, excludedField = \"\")", html_source)
        self.assertIn(".filter((csCase) => matchesLedgerFiltersExcept(csCase, field))", html_source)
        self.assertIn("/api/management-filter-options?", html_source)
        self.assertIn("appendManagementFilterParams(params, { excludeField: field });", html_source)
        self.assertIn("filter_${field}", html_source)
        self.assertIn("function defaultManagementPeriod()", html_source)
        self.assertIn("managementMonthFilter.innerHTML = periodOptions", html_source)
        self.assertNotIn('managementMonthFilter.innerHTML = `<option value="">전체 선택</option>', html_source)
        self.assertIn("if (field === excludedField) return true;", html_source)

    def test_management_filters_can_show_only_colored_cells(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="managementFilterColoredOnly"', html_source)
        self.assertIn("const managementColorFilters = {};", html_source)
        self.assertIn("function managementCellHasColor(record, field, duplicateCounts = managementColorFilterDuplicateCounts)", html_source)
        self.assertIn("function matchesManagementColorFilters(record)", html_source)
        self.assertIn("managementFilterColoredOnly.addEventListener", html_source)
        self.assertIn("managementColorFilters[activeManagementFilterField] = \"colored\";", html_source)
        self.assertIn("Object.keys(managementColorFilters).forEach", html_source)
        self.assertIn("colorActive: Boolean(managementColorFilters[field])", html_source)

    def test_ledger_filter_headers_show_active_filter_state(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("function syncColumnFilterIndicator(button", html_source)
        self.assertIn('button.closest("th")', html_source)
        self.assertIn('header.classList.toggle("filter-active", active);', html_source)
        self.assertIn('header.classList.toggle("text-filter-active", Boolean(textActive));', html_source)
        self.assertIn('header.classList.toggle("color-filter-active", Boolean(colorActive));', html_source)
        self.assertIn('.ledger-table th.filter-active', html_source)
        self.assertIn('.ledger-table th.color-filter-active', html_source)
        self.assertIn('.ledger-table th.text-filter-active.color-filter-active', html_source)
        self.assertIn('syncColumnFilterIndicator(button, { textActive: Boolean(value), value });', html_source)
        self.assertIn('syncColumnFilterIndicator(button, {', html_source)
        self.assertIn('colorActive: Boolean(managementColorFilters[field])', html_source)

    def test_crm_task_board_uses_collapsible_advanced_filters_from_branch(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="crmTaskAdvancedToggle"', html_source)
        self.assertIn('id="crmAdvancedFilters" hidden', html_source)
        self.assertIn('aria-controls="crmAdvancedFilters"', html_source)
        self.assertIn(".crm-advanced-filters", html_source)
        self.assertIn("crmTaskAdvancedToggle?.addEventListener", html_source)
        self.assertIn('crmTaskAdvancedToggle.textContent = open ? "필터 닫기" : "고급 필터"', html_source)
        self.assertIn('data-crm-nav-tab="dashboard">업무 현황</button>', html_source)
        self.assertIn('data-crm-tab="accounts">직원 현황</button>', html_source)
        self.assertIn('data-crm-tab="messages">연동 로그</button>', html_source)

    def test_leave_workflow_has_multi_step_approval_cancel_and_accrual_ui(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            for permission in (
                "leave_approve_team",
                "leave_approve_director",
                "leave_approve_ceo",
                "leave_director_override",
            ):
                self.assertIn(permission, html_source)
            for table_name in ("company_holidays", "leave_notifications"):
                self.assertIn(table_name, html_source)
            for endpoint in ("/api/leave-cancel", "/api/leave-accrual", "/api/company-holiday"):
                self.assertIn(endpoint, html_source)
            for function_name in (
                "save_company_holiday",
                "apply_annual_leave_accrual",
                "cancel_leave_request",
                "list_leave_notifications",
                "actor_can_override_leave",
            ):
                self.assertIn(f"def {function_name}", html_source)
            self.assertIn('id="leaveReservedDays"', html_source)
            self.assertIn('id="leaveNotificationList"', html_source)
            self.assertIn('id="leaveAccrualApply"', html_source)
            self.assertIn('id="leaveAdminUserSearch"', html_source)
            self.assertIn('id="leaveAdminUserList"', html_source)
            self.assertIn('data-leave-tab="mine">내 연차</button>', html_source)
            self.assertIn('data-leave-tab="request">신청하기</button>', html_source)
            self.assertIn('data-leave-tab="approvals">승인 관리</button>', html_source)
            self.assertIn('data-leave-tab="admin">직원 기준</button>', html_source)
            self.assertIn('data-leave-tab="holidays">공휴일 설정</button>', html_source)
            self.assertIn('data-leave-tab="history">이력</button>', html_source)
            self.assertIn('id="leaveTabHolidays"', html_source)
            self.assertIn('id="leaveTabHistory"', html_source)
            self.assertIn("main:has(#leaveWorkspace.active)", html_source)
            self.assertIn("#leaveWorkspace.active", html_source)
            self.assertIn("can_view_staff", html_source)
            self.assertIn("function renderLeaveAdminUserList", html_source)
            self.assertIn("function selectLeaveAdminUser", html_source)
            self.assertIn("data-leave-admin-user", html_source)
            self.assertIn('id="leaveHolidayDateInput"', html_source)
            self.assertIn('id="leaveHolidaySave"', html_source)
            self.assertIn('data-leave-comment="${row.id}"', html_source)
            self.assertIn('data-leave-cancel="${row.id}"', html_source)
            self.assertIn('data-leave-decision="override"', html_source)

    def test_leave_notifications_are_available_from_topbar_alerts(self) -> None:
        html_source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("/api/leave-notifications", html_source)
        self.assertIn("/api/leave-notifications-read", html_source)
        self.assertIn("function fetchLeaveNotifications", html_source)
        self.assertIn("function showLeaveNotificationWidget", html_source)
        self.assertIn("function startLeaveNotificationWatcher", html_source)
        self.assertIn("function markLeaveNotificationsRead", html_source)
        self.assertIn("function canUseLeaveNotifications", html_source)
        self.assertIn("leaveNotificationText", html_source)
        self.assertIn("data-leave-notification-read", html_source)
        self.assertIn("확인 완료", html_source)
        self.assertIn("연차 알림", html_source)

    def test_delivery_modal_title_matches_menu_label(self) -> None:
        html_source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('data-order-execute="delivery">실행</button>', html_source)
        self.assertIn('modalTitle.textContent = "개별 택배건 정리";', html_source)

    def test_delivery_summary_supports_safe_number_confirmation(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn("build_summary_payload", html_source)
            self.assertIn("safeNumberConfirmMessage", html_source)
            self.assertIn("safe_number_candidates", html_source)
            self.assertIn("approved_text", html_source)
            self.assertIn("안심번호 합포 후보", html_source)
            self.assertIn('id="safeNumberPackageDialog"', html_source)
            self.assertIn("function requestSafeNumberPackageApproval", html_source)
            self.assertIn('id="safeNumberPackageApprove"', html_source)
            self.assertIn('id="safeNumberPackageReject"', html_source)
            self.assertNotIn("window.confirm(safeNumberConfirmMessage", html_source)

    def test_ledger_imports_are_split_into_daily_and_replace_modes(self) -> None:
        html_source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn("통합관리대장 업로드", html_source)
        self.assertIn("CS처리대장 업로드", html_source)
        self.assertIn("전체 데이터 교체 업로드", html_source)
        self.assertIn("/api/management-import-preview", html_source)
        self.assertIn("/api/cs-cases-import-preview", html_source)
        self.assertIn("function requestImportWarningApproval", html_source)
        self.assertIn('id="importWarningDialog"', html_source)
        self.assertIn('id="importCorrectionDialog"', html_source)
        self.assertIn("function requestImportCorrectionApproval", html_source)
        self.assertIn("data-correction-field", html_source)
        self.assertIn('formData.append("corrections"', html_source)
        self.assertIn('mode", mode', html_source)
        self.assertIn('mode === "replace"', html_source)
        self.assertIn("preview_management_import", html_source)
        self.assertIn("preview_cs_cases_import", html_source)

    def test_ledger_imports_show_progress_dialog(self) -> None:
        html_source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="importProgressDialog"', html_source)
        self.assertIn('id="importProgressStage"', html_source)
        self.assertIn('id="importProgressSteps"', html_source)
        self.assertIn("function showImportProgress", html_source)
        self.assertIn("function updateImportProgress", html_source)
        self.assertIn("function finishImportProgress", html_source)
        self.assertIn('updateImportProgress("preview"', html_source)
        self.assertIn('updateImportProgress("saving"', html_source)
        self.assertIn('updateImportProgress("done"', html_source)

    def test_backup_workspace_supports_auto_and_selected_backup_settings(self) -> None:
        html_source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="backupAutoEnabled"', html_source)
        self.assertIn('id="backupAutoHour"', html_source)
        self.assertIn('id="backupRetentionDays"', html_source)
        self.assertIn('id="backupDirInput"', html_source)
        self.assertIn('id="backupSettingsSave"', html_source)
        self.assertIn('id="backupCreateSelected"', html_source)
        self.assertIn('id="backupOfflineDownload"', html_source)
        self.assertIn('id="backupExternalEnabled"', html_source)
        self.assertIn('id="backupRcloneRemote"', html_source)
        self.assertIn('id="backupRclonePath"', html_source)
        self.assertIn('id="backupRcloneExecutable"', html_source)
        self.assertIn('id="backupExternalStatus"', html_source)
        self.assertIn("/api/backup-settings", html_source)
        self.assertIn("/api/backup-offline-download", html_source)
        self.assertIn("function saveBackupSettings", html_source)
        self.assertIn("function createBackupAtSelectedPath", html_source)
        self.assertIn("function downloadOfflineBackup", html_source)
        self.assertIn("Google Drive 업로드 상태", html_source)
        self.assertIn("rclone_remote", html_source)
        self.assertIn("load_backup_settings", html_source)
        self.assertIn("save_backup_settings", html_source)
        self.assertIn("BACKUP_DATA_DIRECTORIES", html_source)
        self.assertIn("write_backup_directory", html_source)
        self.assertIn("restore_backup_directory", html_source)
        self.assertIn('"backup_scope"', html_source)

    def test_order_workspace_has_right_side_execution_cards(self) -> None:
        html_source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="orderWorkspace"', html_source)
        self.assertIn('id="orderWorkspaceTitle"', html_source)
        self.assertIn('id="orderWorkspaceCards"', html_source)
        self.assertIn('id="orderRecentDownloads"', html_source)
        self.assertIn('id="orderDownloadList"', html_source)
        self.assertIn('/api/order-downloads', html_source)
        self.assertIn('/api/order-download?id=', html_source)
        self.assertIn("order-exec-card", html_source)
        self.assertIn("ORDER_WORKFLOWS", html_source)
        self.assertIn("function showOrderWorkspace(mode)", html_source)
        self.assertIn("function loadOrderDownloads()", html_source)
        self.assertIn("function renderOrderDownloads(downloads)", html_source)
        self.assertIn('orderWorkspace.classList.toggle("active"', html_source)

    def test_upload_dialog_does_not_use_daisyui_modal_class(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn('class="workhub-modal" role="dialog"', html_source)
            self.assertIn('modal.querySelector(".workhub-modal")', html_source)
            self.assertNotIn('class="modal" role="dialog"', html_source)
            self.assertNotIn(".modal-backdrop.open .modal *", html_source)
            self.assertNotIn(".workhub-modal-backdrop.open .workhub-modal *", html_source)

    def test_naver_mail_defaults_are_managed_from_admin_workspace(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn('id="adminNaverEmailInput"', html_source)
            self.assertIn('id="adminNaverPasswordInput"', html_source)
            self.assertIn('id="adminSaveMailCredentials"', html_source)
            self.assertIn('id="adminMailSettingsSave"', html_source)
            self.assertIn("function loadAdminMailSettings()", html_source)
            self.assertIn("function saveAdminMailSettings()", html_source)
            self.assertIn('"/api/mail-settings"', html_source)
            self.assertIn("메일 기본정보 저장", html_source)

            cs_fields_start = html_source.index('class="cs-fields" id="csFields"')
            vendor_select = html_source.index('id="vendorContactSelect"', cs_fields_start)
            cs_mail_account_slice = html_source[cs_fields_start:vendor_select]
            self.assertNotIn('id="naverEmailInput"', cs_mail_account_slice)
            self.assertNotIn('id="naverPasswordInput"', cs_mail_account_slice)
            self.assertNotIn('id="saveMailCredentials"', cs_mail_account_slice)

    def test_bulk_mail_technical_settings_are_managed_from_admin_workspace(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn('id="adminSmtpPort"', html_source)
            self.assertIn('id="adminSmtpSecurity"', html_source)
            self.assertIn('id="adminBulkBatchSize"', html_source)
            self.assertIn('id="adminBulkSendInterval"', html_source)
            self.assertIn('id="adminBulkBatchPause"', html_source)
            self.assertIn('id="adminBulkTestRecipient"', html_source)
            self.assertIn('id="adminMailTechnicalSave"', html_source)
            self.assertIn('id="adminMailTestSend"', html_source)
            self.assertIn("function saveAdminMailTechnicalSettings()", html_source)
            self.assertIn("function sendAdminMailTestMessage()", html_source)
            self.assertIn('"/api/mail-test"', html_source)

    def test_naver_mail_integration_admin_ui_uses_clear_operating_text(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn("네이버 메일 연동", html_source)
            self.assertIn("네이버 메일 아이디", html_source)
            self.assertIn("네이버 메일 비밀번호", html_source)
            self.assertIn("연동 테스트 메일 발송", html_source)
            self.assertIn("저장된 네이버 메일 계정으로 1건만 발송합니다.", html_source)
            self.assertIn("테스트 메일을 발송했습니다.", html_source)

    def test_mail_settings_redirects_cleanly_when_login_session_expires(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn("function handleLoginRequiredResponse", html_source)
            self.assertIn("response.status === 401", html_source)
            self.assertIn("로그인이 만료되었습니다. 다시 로그인해주세요.", html_source)
            self.assertIn('window.location.href = "/login"', html_source)
            self.assertIn('credentials: "same-origin"', html_source)

    def test_distribution_mail_tree_only_opens_writable_mail_flows(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertNotIn('data-mail-popup="supplier"', html_source)
            self.assertNotIn('data-mail-popup="seller"', html_source)
            self.assertIn('data-mail-popup="cs"', html_source)
            self.assertIn('data-mail-popup="stock"', html_source)
            self.assertIn("입고 및 품절 공지", html_source)
            self.assertIn('openModal(type === "stock" ? "mail-stock" : "cs")', html_source)
            self.assertIn('class="stock-notice-fields" id="stockNoticeFields"', html_source)
            self.assertIn('id="stockNoticeDateInput"', html_source)
            self.assertNotIn('id="stockManagerNameInput"', html_source)
            self.assertNotIn('id="stockManagerPhoneInput"', html_source)
            self.assertNotIn('id="stockSenderEmailInput"', html_source)
            self.assertIn("function defaultStockContactInfo()", html_source)
            self.assertIn("cachedMailSettings.naver_email ||", html_source)
            self.assertIn('id="stockVendorTypeFilter"', html_source)
            self.assertIn('id="stockVendorSearchInput"', html_source)
            self.assertIn('id="stockVendorPickerButton"', html_source)
            self.assertIn('id="stockVendorTree"', html_source)
            self.assertIn('id="stockSelectedVendorLabel"', html_source)
            self.assertIn('id="stockRecipientEmailInput" type="hidden"', html_source)
            self.assertIn('id="stockVendorNameInput" type="hidden"', html_source)
            self.assertNotIn('for="stockRecipientEmailInput">받는 업체 메일', html_source)
            self.assertNotIn('for="stockVendorNameInput">업체명', html_source)
            self.assertIn("function renderStockVendorContacts()", html_source)
            self.assertIn("let selectedStockVendors = [];", html_source)
            self.assertIn("data-stock-vendor-select-all", html_source)
            self.assertIn("data-stock-vendor-checkbox", html_source)
            self.assertIn("function stockVendorMatchesSearch(contact", html_source)
            self.assertIn("function visibleStockVendorContactsForPicker()", html_source)
            self.assertIn("function toggleAllStockVendorsForType(checked)", html_source)
            self.assertIn("async function sendCurrentStockNoticeMail()", html_source)
            self.assertIn("renderStockVendorContacts();", html_source)
            self.assertIn("async function loadVendorContacts()", html_source)
            self.assertIn("function applySelectedVendor()", html_source)
            self.assertIn('vendorContactSelect.addEventListener("change", applySelectedVendor)', html_source)
            self.assertIn("function applySelectedStockVendor()", html_source)
            self.assertIn('stockRecipientEmailInput.value = first?.email || "";', html_source)
            self.assertIn('stockVendorPickerButton?.addEventListener("click"', html_source)
            self.assertIn('stockVendorSearchInput?.addEventListener("input"', html_source)
            self.assertIn('stockVendorTree?.addEventListener("change"', html_source)
            self.assertIn("function defaultStockNoticeBody()", html_source)
            self.assertIn("function collectStockNoticePayload", html_source)
            self.assertIn('id="stockMailHistoryList"', html_source)
            self.assertIn("function renderStockMailHistory(logs)", html_source)
            self.assertIn("async function loadStockMailHistory()", html_source)
            self.assertIn('stockNoticeFields.style.display = "block"', html_source)
            self.assertIn('"/api/mail-send"', html_source)
            self.assertIn('send_general_mail(payload)', html_source)
            self.assertIn("/api/vendor-mail-send-logs", html_source)
            self.assertIn("vendor_mail_send_logs", html_source)
            self.assertIn("save_vendor_mail_send_log", html_source)
            self.assertIn("list_vendor_mail_send_logs", html_source)
            self.assertIn("chunk_mail_recipients", html_source)
            self.assertIn("vendor_contact_delete_approvals", html_source)
            self.assertIn("approve_failed_vendor_contact_delete", html_source)
            self.assertIn("/api/vendor-contact-failure-delete", html_source)
            self.assertIn("승인 삭제", html_source)
            self.assertEqual(1, html_source.count("function defaultStockNoticeBody()"))
            self.assertIn('} else if (mode === "cs") {', html_source)
            self.assertIn('} else if (mode === "mail-stock") {', html_source)
            self.assertNotIn('mode === "cs" || mode === "mail-stock"', html_source)
            self.assertNotIn('currentMode === "cs" || currentMode === "mail-stock"', html_source)
            self.assertNotIn("제품 입고 및 품절 안내드립니다", html_source)
            self.assertNotIn("안내 내용", html_source)
            self.assertNotIn("확인 후 관련 일정", html_source)
            self.assertNotIn("csBodyInput.value = defaultStockNoticeBody();", html_source)
        app_html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")
        self.assertIn("bcc_emails: bccEmails", app_html_source)
        self.assertIn("숨은참조 방식으로 1회 발송", app_html_source)
        self.assertIn("숨은참조 방식으로 {batch_count}회 나눠 발송", app_html_source)
        self.assertNotIn("for (const recipient of recipients)", app_html_source)

    def test_vendor_contact_upload_is_managed_from_admin_workspace(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn("업체 메일 주소록", html_source)
            self.assertIn('id="vendorContactsFileInput"', html_source)
            self.assertIn('id="vendorContactsDropMain"', html_source)
            self.assertIn("function uploadVendorContactsWorkbook()", html_source)

            admin_start = html_source.index('id="userAdminWorkspace"')
            admin_end = html_source.index('id="userAdminBody"', admin_start)
            admin_slice = html_source[admin_start:admin_end]
            self.assertIn('id="vendorContactsFileInput"', admin_slice)
            self.assertIn('id="vendorContactsDropMain"', admin_slice)

            cs_fields_start = html_source.index('class="cs-fields" id="csFields"')
            recipient_field = html_source.index('id="recipientEmailInput"', cs_fields_start)
            cs_contact_slice = html_source[cs_fields_start:recipient_field]
            self.assertNotIn('id="vendorContactsFileInput"', cs_contact_slice)
            self.assertNotIn('id="vendorContactsDropMain"', cs_contact_slice)

    def test_sales_report_has_dedicated_sales_workspace_mode_with_upload_form(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('id="salesReportUploadCard"', html_source)
        self.assertIn('id="salesReportNavGroup"', html_source)
        self.assertIn('id="salesReportNavToggle"', html_source)
        self.assertIn('"sales_report_manage"', html_source)
        self.assertIn("__SALES_REPORT_NAV__", html_source)
        self.assertNotIn('data-open="salesReport">매출현황</button>', html_source)
        self.assertIn('mode === "salesReport"', html_source)
        self.assertIn("sales-report-only", html_source)
        self.assertIn("#userAdminWorkspace.sales-report-only > .workspace-head", html_source)
        self.assertIn('!can("sales_report_manage")', html_source)
        self.assertIn('"/api/sales-report-upload"', html_source)
        self.assertIn('"/api/sales-report-uploads"', html_source)
        self.assertIn('"/api/sales-report-products"', html_source)
        self.assertIn('"/api/sales-report-manual-options"', html_source)
        self.assertIn('"/api/sales-report-manual-entry"', html_source)
        self.assertIn("function uploadSalesReportWorkbook()", html_source)
        self.assertIn("function openSalesReportUploadPicker()", html_source)
        self.assertIn("function saveSalesManualEntry()", html_source)
        self.assertIn("function loadSalesReportUploads()", html_source)
        self.assertIn('document.querySelector("#salesReportNavToggle")?.addEventListener("click"', html_source)
        self.assertIn('showWorkspace("salesReport")', html_source)
        self.assertNotIn('button.closest("#salesReportNavGroup")', html_source)
        self.assertIn('id="salesReportFileInput"', html_source)
        self.assertIn('id="salesReportManualUpload"', html_source)
        self.assertIn('id="salesReportManualEntryToggle"', html_source)
        self.assertIn('id="salesManualSellerName"', html_source)
        self.assertIn('id="salesManualSupplierName"', html_source)
        self.assertIn('id="salesManualProductName"', html_source)
        self.assertIn('id="salesManualSellerOptions"', html_source)
        self.assertIn('id="salesManualSupplierOptions"', html_source)
        self.assertIn('id="salesManualProductOptions"', html_source)
        self.assertIn('id="salesManualEntrySave"', html_source)
        self.assertIn('id="salesReportUploadMessage"', html_source)
        self.assertIn('id="salesReportRecentList"', html_source)
        self.assertIn("매출표 파일을 직접 업로드해서 현황을 갱신합니다.", html_source)
        self.assertNotIn('id="salesReportDropMain"', html_source)
        self.assertNotIn('id="salesReportChooseFile"', html_source)
        self.assertNotIn('label[for=\'salesReportFileInput\']', html_source)

        admin_nav_start = html_source.index("ADMIN_TOOLS_NAV_HTML")
        admin_nav_end = html_source.index("SALES_REPORT_NAV_HTML", admin_nav_start)
        admin_nav_slice = html_source[admin_nav_start:admin_nav_end]
        self.assertNotIn('data-admin-focus="salesReport"', admin_nav_slice)
        self.assertNotIn('data-open="salesReport"', admin_nav_slice)

        admin_start = html_source.index('id="userAdminWorkspace"')
        admin_end = html_source.index('id="userAdminBody"', admin_start)
        admin_slice = html_source[admin_start:admin_end]
        self.assertIn('>매출현황</div>', admin_slice)
        self.assertIn('id="salesReportFileInput"', admin_slice)
        self.assertIn('id="salesReportManualUpload"', admin_slice)
        self.assertNotIn('id="salesNasImportDir"', admin_slice)
        self.assertNotIn('id="salesNasScanNow"', admin_slice)
        self.assertNotIn("NAS 자동 업로드", admin_slice)

    def test_sales_report_permission_shows_menu_for_non_admin_staff(self) -> None:
        app = self.load_app()

        html_source = app.render_app_html({
            "username": "sales-staff",
            "display_name": "매출담당",
            "role": "user",
            "permissions": ["sales_report_manage"],
        })

        self.assertIn('id="salesReportNavGroup"', html_source)
        self.assertNotIn('data-open="salesReport">매출현황</button>', html_source)
        self.assertIn('id="userAdminWorkspace"', html_source)
        self.assertIn('id="salesReportUploadCard"', html_source)
        self.assertNotIn('id="adminToolsNavGroup"', html_source)
        self.assertIn('mode === "userAdmin" && (!userAdminWorkspace || currentUser.role !== "admin")', html_source)

    def test_import_cost_program_is_controlled_by_dedicated_permission(self) -> None:
        app = self.load_app()

        admin_html = app.render_app_html({
            "username": "admin",
            "display_name": "관리자",
            "role": "admin",
            "permissions": app.ALL_PERMISSIONS,
        })
        director_html = app.render_app_html({
            "username": "ssh19",
            "display_name": "신성환 실장",
            "role": "user",
            "permissions": ["ledger_edit"],
        })
        ceo_html = app.render_app_html({
            "username": "ceo",
            "display_name": "신성민 대표",
            "role": "user",
            "permissions": ["ledger_edit"],
        })
        staff_html = app.render_app_html({
            "username": "staff",
            "display_name": "직원",
            "role": "user",
            "permissions": ["import_cost_manage"],
        })

        self.assertIn("import_cost_manage", app.ALL_PERMISSIONS)
        self.assertIn('id="importCostNavGroup"', admin_html)
        self.assertIn('id="importCostWorkspace"', admin_html)
        self.assertNotIn('id="importCostNavGroup"', director_html)
        self.assertNotIn('id="importCostWorkspace"', director_html)
        self.assertNotIn('id="importCostNavGroup"', ceo_html)
        self.assertNotIn('id="importCostWorkspace"', ceo_html)
        self.assertIn('id="importCostNavGroup"', staff_html)
        self.assertIn('id="importCostWorkspace"', staff_html)
        self.assertNotIn('id="salesReportNavGroup"', ceo_html)
        self.assertNotIn('id="salesReportDashboard"', ceo_html)
        self.assertIn("function calculateImportCost()", admin_html)
        self.assertIn('"/api/import-cost-calculate"', admin_html)
        self.assertIn('id="importCostFileInput"', admin_html)
        self.assertIn('"/api/import-cost-upload"', admin_html)
        self.assertIn("selectedImportCostFiles", admin_html)
        self.assertIn("appendImportCostSelectedFiles(files)", admin_html)
        self.assertIn('id="importCostIncludeImportVat" type="checkbox" checked', admin_html)
        self.assertIn('id="importCostIncludeServiceVat" type="checkbox" checked', admin_html)
        self.assertIn("includeImportVat.checked = true", admin_html)
        self.assertIn("includeServiceVat.checked = true", admin_html)
        self.assertIn("const importCostWonInputIds = new Set", admin_html)
        self.assertIn("function normalizeImportCostMoneyValue", admin_html)
        self.assertIn("function formatImportCostMoneyInput", admin_html)
        self.assertIn("function formatImportCostRate", admin_html)
        self.assertIn('id="importCostDocFee" type="text" inputmode="numeric"', admin_html)
        self.assertIn('id="importCostImportVat" type="text" inputmode="numeric"', admin_html)
        self.assertIn('placeholder="예) 2,191,192원"', admin_html)
        self.assertIn("importCostWonInputIds.has(target.id)", admin_html)
        self.assertIn('id="importCostRunStatus"', admin_html)
        self.assertIn('id="importCostChargeSummary"', admin_html)
        self.assertIn("function renderImportCostChargeSummary", admin_html)
        self.assertIn("function importCostMissingSettlementWarnings", admin_html)
        self.assertIn("import-cost-unit-cost", admin_html)
        self.assertIn("import-cost-unit-price-card", admin_html)
        self.assertIn("핵심 원가", admin_html)
        self.assertIn("D/O/운임비", admin_html)
        self.assertIn('["수입부가세", formatImportCostWon(vat), ""]', admin_html)
        self.assertIn("const freight = values.docFee + values.otherCost + values.serviceVat;", admin_html)
        self.assertIn("function setImportCostRunStatus", admin_html)
        self.assertIn('id="importCostExportReport"', admin_html)
        self.assertIn('id="importCostSaveReport"', admin_html)
        self.assertIn('id="importCostFinalizeReport"', admin_html)
        self.assertIn('id="importCostSaveFinalReport"', admin_html)
        self.assertIn('id="importCostTabs"', admin_html)
        self.assertIn('data-import-cost-tab="calculate"', admin_html)
        self.assertIn('data-import-cost-tab="saved"', admin_html)
        self.assertIn('data-import-cost-tab="files"', admin_html)
        self.assertIn('data-import-cost-tab="history"', admin_html)
        self.assertIn('id="importCostSavedListNav"', admin_html)
        self.assertIn('id="importCostHistoryNav"', admin_html)
        self.assertIn('id="importCostSavedCards"', admin_html)
        self.assertIn("function saveImportCostManagedProductName", admin_html)
        self.assertIn('"/api/import-cost-report-managed-product"', admin_html)
        self.assertIn('id="importCostSavedSearch"', admin_html)
        self.assertIn('id="importCostSavedStatusFilter"', admin_html)
        self.assertIn('id="importCostSavedMonthFilter"', admin_html)
        self.assertIn('id="importCostSavedResetFilters"', admin_html)
        self.assertIn('id="importCostOriginalFilesPanel"', admin_html)
        self.assertIn('id="importCostHistoryPanel"', admin_html)
        self.assertIn("data-import-cost-file-analyze", admin_html)
        self.assertIn('id="importCostSavedBody"', admin_html)
        self.assertIn("import-cost-report-table-wrap", admin_html)
        self.assertIn("import-cost-report-table", admin_html)
        self.assertIn("importCostReportBasisDate", admin_html)
        self.assertIn("import-cost-managed-product-inline", admin_html)
        self.assertIn("index + 1", admin_html)
        self.assertIn('class="import-cost-card import-cost-result-card"', admin_html)
        self.assertLess(admin_html.index('id="importCostResultBody"'), admin_html.index('id="importCostSavedBody"'))
        self.assertIn("formatImportCostRate(report.remittance_rate)", admin_html)
        self.assertIn("function setImportCostTab", admin_html)
        self.assertIn("function loadImportCostSavedReports", admin_html)
        self.assertIn("function saveCurrentImportCostReport", admin_html)
        self.assertIn("function saveAndFinalizeCurrentImportCostReport", admin_html)
        self.assertIn("import-cost-report-warning", admin_html)
        self.assertIn('id="importCostReviewBackdrop"', admin_html)
        self.assertIn("function openImportCostChargeReview", admin_html)
        self.assertIn("function applyImportCostChargeReview", admin_html)
        self.assertIn('formData.append("review_only", "1")', admin_html)
        source = (SCRIPTS / "workhub_delivery_app.py").read_text(encoding="utf-8")
        self.assertIn("review_only = str(fields.get(\"review_only\")", source)
        self.assertIn("if result and payload and not review_only:", source)
        self.assertIn("function formatImportCostProductNumber", admin_html)
        self.assertIn("formatImportCostProductNumber(product.quantity, true)", admin_html)
        self.assertIn("formatImportCostProductNumber(product.unit_usd, false, 4)", admin_html)
        self.assertIn("formatImportCostProductNumber(product.amount_usd, false, 2)", admin_html)
        self.assertIn("formatImportCostProductNumber(product.gross_weight, false, 4)", admin_html)
        self.assertIn("formatImportCostProductNumber(product.cbm, false, 4)", admin_html)
        self.assertIn("function finalizeCurrentImportCostReport", admin_html)
        self.assertIn("function analyzeSavedImportCostFile", admin_html)
        self.assertIn('"/api/import-cost-report-save"', admin_html)
        self.assertIn('"/api/import-cost-report-status"', admin_html)
        self.assertIn('"/api/import-cost-original-analyze"', admin_html)
        self.assertIn('"/api/import-cost-reports"', admin_html)
        self.assertIn("import-cost-rate-field", admin_html)
        self.assertIn("제품 원가 계산의 기준 환율입니다.", admin_html)
        self.assertIn("function exportImportCostReport", admin_html)
        self.assertIn('"/api/import-cost-report-export"', admin_html)
        self.assertIn('await downloadWorkbookResponse(response, "수입원가_계산보고서.xlsx")', admin_html)
        self.assertIn('setImportCostRunStatus("running"', admin_html)
        self.assertIn('setImportCostRunStatus("done"', admin_html)
        self.assertIn('setImportCostRunStatus("error"', admin_html)

    def test_import_cost_report_tracks_managed_product_name(self) -> None:
        app = self.load_app()
        payload = {
            "hbl_no": "XLTNGB26040216",
            "invoice_no": "SXT20260420",
            "remittance_rate": "1482.04",
            "allocation_basis": "amount",
            "managed_product_name": "28CM POT INTERNAL",
            "products": [{
                "name": "28CM POT",
                "quantity": "3985",
                "unit_usd": "3.86",
                "amount_usd": "15382.10",
                "gross_weight": "6463.7",
                "cbm": "68",
            }],
        }
        result = app.calculate_import_cost(payload)
        report = app.save_import_cost_report(payload, result, user={"display_name": "Admin", "role": "admin"})

        self.assertEqual(report["managed_product_name"], "28CM POT INTERNAL")
        self.assertEqual(app.list_import_cost_reports()[0]["managed_product_name"], "28CM POT INTERNAL")

        updated = app.update_import_cost_report_managed_product(
            report["id"],
            "노르디쿡 IH 무쇠팬 28cm",
            user={"display_name": "Admin", "role": "admin"},
        )

        self.assertEqual(updated["managed_product_name"], "노르디쿡 IH 무쇠팬 28cm")
        detailed = app.get_import_cost_report(report["id"])
        self.assertEqual(detailed["managed_product_name"], "노르디쿡 IH 무쇠팬 28cm")
        self.assertEqual(detailed["history"][-1]["action"], "managed_product")

    def test_import_cost_saved_reports_do_not_read_import_shipment_dates(self) -> None:
        app = self.load_app()
        app.init_db()

        def save_report(hbl_no: str, warehouse_due_date: str) -> dict:
            app.save_import_shipment(
                {
                    "warehouse_due_date": warehouse_due_date,
                    "hbl_no": hbl_no,
                    "item": f"Product {hbl_no}",
                    "quantity": "1",
                }
            )
            payload = {
                "hbl_no": hbl_no,
                "invoice_no": f"INV-{hbl_no}",
                "managed_product_name": f"Managed {hbl_no}",
                "remittance_rate": "1512",
                "allocation_basis": "amount",
                "doc_fee": "1000",
                "duty": "0",
                "broker_fee": "0",
                "other_cost": "",
                "import_vat": "0",
                "service_vat": "0",
                "include_import_vat": True,
                "include_service_vat": True,
                "products": [
                    {
                        "name": f"Product {hbl_no}",
                        "quantity": "10",
                        "unit_usd": "1",
                        "amount_usd": "10",
                        "gross_weight": "1",
                        "cbm": "1",
                    }
                ],
            }
            return app.save_import_cost_report(
                payload,
                app.calculate_import_cost(payload),
                user={"display_name": "Admin", "role": "admin"},
            )

        save_report("NEW-DATE-HBL", "2026-07-09")
        save_report("OLD-DATE-HBL", "2026-06-10")

        reports = app.list_import_cost_reports()

        self.assertEqual({report["hbl_no"] for report in reports}, {"NEW-DATE-HBL", "OLD-DATE-HBL"})
        self.assertTrue(all("import_ledger_date" not in report for report in reports))
        self.assertTrue(all("import_ledger_date_label" not in report for report in reports))

    def test_import_cost_calculation_allocates_to_product_unit_cost(self) -> None:
        app = self.load_app()

        result = app.calculate_import_cost({
            "remittance_rate": "1482.04",
            "allocation_basis": "amount",
            "doc_fee": "2240795",
            "duty": "0",
            "broker_fee": "48400",
            "import_vat": "2418290",
            "service_vat": "4840",
            "include_import_vat": False,
            "include_service_vat": False,
            "products": [{
                "name": "28CM POT",
                "quantity": "3985",
                "unit_usd": "3.86",
                "amount_usd": "15382.10",
                "gross_weight": "6463.7",
                "cbm": "68",
            }],
        })

        self.assertEqual(result["summary"]["invoice_total_usd"], 15382.10)
        self.assertEqual(result["summary"]["purchase_total_krw"], 22796887)
        self.assertEqual(result["summary"]["allocated_cost_total"], 2289195)
        self.assertEqual(result["products"][0]["landed_total"], 25086082)
        self.assertEqual(result["products"][0]["landed_unit"], 6295.13)

    def test_import_cost_report_workbook_contains_summary_charges_and_product_rows(self) -> None:
        app = self.load_app()
        payload = {
            "hbl_no": "XLTNGB26040216",
            "invoice_no": "SXT20260420",
            "remittance_rate": "1482.04",
            "allocation_basis": "amount",
            "doc_fee": "2191192",
            "duty": "0",
            "broker_fee": "53240",
            "other_cost": "",
            "import_vat": "2418290",
            "service_vat": "49603",
            "include_import_vat": True,
            "include_service_vat": True,
            "products": [{
                "name": "28CM POT",
                "quantity": "3985",
                "unit_usd": "3.86",
                "amount_usd": "15382.10",
                "gross_weight": "6463.7",
                "cbm": "68",
            }],
        }

        result = app.calculate_import_cost(payload)
        data = app.import_cost_report_workbook_bytes(payload, result)
        workbook = load_workbook(BytesIO(data), data_only=True)
        try:
            sheet = workbook["수입원가 계산보고서"]
            values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
            self.assertIn("수입 원가 계산 보고서", values)
            self.assertIn("HBL/컨테이너 번호", values)
            self.assertIn("XLTNGB26040216", values)
            self.assertIn("정산 비용", values)
            self.assertIn("D/O/운임비", values)
            self.assertIn(2240795, values)
            self.assertIn("D/O 비용에 포함", values)
            self.assertIn("D/O/운임비 행에 합산 표시", values)
            self.assertIn("제품별 원가 계산 결과", values)
            self.assertIn("28CM POT", values)
            self.assertIn(result["summary"]["landed_total"], values)
        finally:
            workbook.close()

    def test_import_cost_report_is_saved_without_linking_to_import_shipments(self) -> None:
        app = self.load_app()
        payload = {
            "hbl_no": "XLTNGB26040216",
            "invoice_no": "SXT20260420",
            "remittance_rate": "1482.04",
            "allocation_basis": "amount",
            "doc_fee": "2191192",
            "duty": "0",
            "broker_fee": "53240",
            "other_cost": "",
            "import_vat": "2418290",
            "service_vat": "49603",
            "include_import_vat": True,
            "include_service_vat": True,
            "products": [{
                "name": "28CM POT",
                "quantity": "3985",
                "unit_usd": "3.86",
                "amount_usd": "15382.10",
                "gross_weight": "6463.7",
                "cbm": "68",
            }],
        }
        result = app.calculate_import_cost(payload)
        upload_path = Path(os.environ["WORKHUB_DATA_DIR"]) / "XLTNGB26040216.pdf"
        upload_path.write_bytes(b"original pdf bytes")

        report = app.save_import_cost_report(
            payload,
            result,
            user={"display_name": "신성환 실장"},
            upload_paths=[upload_path],
        )

        self.assertEqual(report["hbl_no"], "XLTNGB26040216")
        self.assertEqual(report["remittance_rate"], "1482.04")
        self.assertEqual(report["landed_total"], result["summary"]["landed_total"])
        self.assertEqual(len(report["files"]), 1)
        stored_path, metadata = app.import_cost_file_download_info(report["files"][0]["id"])
        self.assertTrue(stored_path.exists())
        self.assertEqual(metadata["original_name"], "XLTNGB26040216.pdf")
        reports = app.list_import_cost_reports()
        self.assertEqual(reports[0]["id"], report["id"])
        self.assertEqual(app.list_import_shipments(), [])

        shipment_id = app.save_import_shipment({
            "hbl_no": "XLTNGB26040216",
            "item": "사용자 입력 제품명",
            "quantity": "77",
            "progress_status": "출항예정",
        })
        before = app.list_import_shipments()
        payload["products"][0]["name"] = "원가 계산 제품명"
        payload["products"][0]["quantity"] = "999"
        app.save_import_cost_report(payload, app.calculate_import_cost(payload), user={"display_name": "신성환 실장"})
        after = app.list_import_shipments()

        self.assertEqual(len(after), 1)
        self.assertEqual(after, before)
        self.assertEqual(after[0]["id"], shipment_id)
        self.assertEqual(after[0]["item"], "사용자 입력 제품명")
        self.assertEqual(after[0]["quantity"], "77")
        self.assertNotIn("import_cost_report_id", after[0])
        self.assertNotIn("import_cost_landed_total", after[0])

    def test_import_cost_reports_track_status_version_and_history(self) -> None:
        app = self.load_app()
        payload = {
            "hbl_no": "XLTNGB26040216",
            "invoice_no": "SXT20260420",
            "remittance_rate": "1482.04",
            "allocation_basis": "amount",
            "doc_fee": "2191192",
            "duty": "0",
            "broker_fee": "53240",
            "other_cost": "",
            "import_vat": "2418290",
            "service_vat": "49603",
            "include_import_vat": True,
            "include_service_vat": True,
            "products": [{
                "name": "28CM POT",
                "quantity": "3985",
                "unit_usd": "3.86",
                "amount_usd": "15382.10",
                "gross_weight": "6463.7",
                "cbm": "68",
            }],
        }
        result = app.calculate_import_cost(payload)

        first = app.save_import_cost_report(payload, result, user={"display_name": "Admin", "role": "admin"})
        self.assertEqual(first["status"], "saved")
        self.assertEqual(first["version"], 1)
        self.assertIn("product_summary", first)

        payload["remittance_rate"] = "1550"
        second_result = app.calculate_import_cost(payload)
        second = app.save_import_cost_report(payload, second_result, user={"display_name": "Admin", "role": "admin"})
        self.assertEqual(second["id"], first["id"])
        self.assertEqual(second["version"], 2)

        finalized = app.set_import_cost_report_status(first["id"], "final", user={"display_name": "Admin", "role": "admin"})
        self.assertEqual(finalized["status"], "final")

        detailed = app.get_import_cost_report(first["id"])
        self.assertGreaterEqual(len(detailed["history"]), 3)
        self.assertEqual(detailed["history"][-1]["action"], "status")

    def test_import_cost_saved_original_file_can_be_reanalyzed(self) -> None:
        app = self.load_app()
        payload = {
            "hbl_no": "XLTNGB26040216",
            "invoice_no": "SXT20260420",
            "remittance_rate": "1482.04",
            "allocation_basis": "amount",
            "products": [{
                "name": "28CM POT",
                "quantity": "3985",
                "unit_usd": "3.86",
                "amount_usd": "15382.10",
                "gross_weight": "6463.7",
                "cbm": "68",
            }],
        }
        result = app.calculate_import_cost(payload)
        upload_path = Path(os.environ["WORKHUB_DATA_DIR"]) / "XLTNGB26040216.xlsx"
        upload_path.write_bytes(b"stored workbook bytes")
        report = app.save_import_cost_report(
            payload,
            result,
            user={"display_name": "Admin", "role": "admin"},
            upload_paths=[upload_path],
        )

        original_analyzer = app.analyze_import_cost_files

        def fake_analyzer(paths, options=None):
            self.assertEqual(len(paths), 1)
            self.assertTrue(paths[0].exists())
            self.assertEqual((options or {}).get("remittance_rate"), "1550")
            return {
                "payload": {**payload, "remittance_rate": "1550"},
                "result": {"summary": {"landed_total": 1234}},
            }

        app.analyze_import_cost_files = fake_analyzer
        try:
            analysis = app.analyze_import_cost_original_file(
                report["files"][0]["id"],
                {"remittance_rate": "1550"},
            )
        finally:
            app.analyze_import_cost_files = original_analyzer

        self.assertEqual(analysis["source_file"]["original_name"], "XLTNGB26040216.xlsx")
        self.assertEqual(analysis["payload"]["remittance_rate"], "1550")
        self.assertEqual(analysis["result"]["summary"]["landed_total"], 1234)

    def test_import_cost_saved_reports_can_be_recalculated_from_original_charges(self) -> None:
        app = self.load_app()
        payload = {
            "hbl_no": "XLTNGB26040216",
            "invoice_no": "SXT20260420",
            "remittance_rate": "1482.04",
            "allocation_basis": "amount",
            "doc_fee": "2191192",
            "include_import_vat": True,
            "include_service_vat": True,
            "products": [{
                "name": "28CM POT",
                "quantity": "3985",
                "unit_usd": "3.86",
                "amount_usd": "15382.10",
                "gross_weight": "6463.7",
                "cbm": "68",
            }],
        }
        old_result = app.calculate_import_cost(payload)
        upload_path = Path(os.environ["WORKHUB_DATA_DIR"]) / "XLTNGB26040216.pdf"
        upload_path.write_bytes(b"stored settlement pdf bytes")
        report = app.save_import_cost_report(
            payload,
            old_result,
            user={"display_name": "Admin", "role": "admin"},
            upload_paths=[upload_path],
        )

        original_analyzer = app.analyze_import_cost_files

        def fake_analyzer(paths, options=None):
            self.assertEqual(len(paths), 1)
            self.assertTrue(paths[0].exists())
            self.assertEqual((options or {}).get("remittance_rate"), "1482.04")
            return {
                "payload": {
                    "broker_fee": "53240",
                    "import_vat": "2418290",
                    "duty": "0",
                    "service_vat": "49603",
                },
                "details": ["re-read saved customs charges"],
            }

        app.analyze_import_cost_files = fake_analyzer
        try:
            fixed = app.recalculate_import_cost_reports_from_originals(
                report_id=report["id"],
                user={"display_name": "Admin", "role": "admin"},
            )
            second = app.recalculate_import_cost_reports_from_originals(
                report_id=report["id"],
                user={"display_name": "Admin", "role": "admin"},
            )
        finally:
            app.analyze_import_cost_files = original_analyzer

        updated = app.get_import_cost_report(report["id"])
        self.assertEqual(fixed["updated"], 1)
        self.assertEqual(updated["payload"]["broker_fee"], "53240")
        self.assertEqual(updated["payload"]["import_vat"], "2418290")
        self.assertGreater(updated["result"]["summary"]["allocated_cost_total"], old_result["summary"]["allocated_cost_total"])
        self.assertEqual(updated["history"][-1]["action"], "recalculate")
        self.assertEqual(second["updated"], 0)
        self.assertEqual(second["errors"][0]["reason"], "already current")

    def test_import_cost_missing_settlement_values_are_flagged_and_not_finalized(self) -> None:
        app = self.load_app()
        payload = {
            "hbl_no": "XLTSWA26030027",
            "invoice_no": "",
            "remittance_rate": "1512",
            "allocation_basis": "amount",
            "doc_fee": "2092337",
            "service_vat": "49627",
            "include_import_vat": True,
            "include_service_vat": True,
            "products": [{
                "name": "아이제나흐 베틴 IH 스텐냄비",
                "quantity": "1800",
                "unit_usd": "3.18",
                "amount_usd": "5724",
                "gross_weight": "1728",
                "cbm": "69",
            }],
        }
        result = app.calculate_import_cost(payload)
        report = app.save_import_cost_report(
            payload,
            result,
            user={"display_name": "Admin", "role": "admin"},
        )

        self.assertIn("통관수수료", report["warnings"])
        self.assertIn("수입부가세", report["warnings"])
        with self.assertRaisesRegex(ValueError, "통관수수료.*수입부가세"):
            app.set_import_cost_report_status(
                report["id"],
                "final",
                user={"display_name": "Admin", "role": "admin"},
            )

    def test_import_cost_upload_analysis_reads_invoice_and_packing_files(self) -> None:
        app = self.load_app()
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            invoice_path = root / "INV.xlsx"
            proforma_path = root / "PROFORMA.xlsx"
            packing_path = root / "PACKING LIST.xlsx"

            invoice = Workbook()
            ws = invoice.active
            ws.append(["COMMERCIAL INVOICE"])
            ws.append(["", "", "INVOICE NO.", "", "DATE:"])
            ws.append(["", "", "SXT20260420", "", "2026-04-20"])
            ws.append(["SHIPPING MARK", "DESCRIPTIONS", "UNIT PRICE (FOB)", "QUANTITES (PCS)", "AMOUNT"])
            ws.append(["N/M", "28CM POT", 3.86, 3985, 15382.10])
            ws.append(["TOTAL", "", "", 3985, 15382.10])
            invoice.save(invoice_path)

            proforma = Workbook()
            ws = proforma.active
            ws.append(["PROFORMA INVOICE"])
            ws.append(["NO", "", "Description of Goods", "", "Qty.", "", "UNITPRICE", "AMOUNT"])
            ws.append([1, "", "", "28cm무쇄냄비", 3985, "PCS", 3.86, 15382.10])
            ws.append(["TOTEL", "", "", "", 3985, "", "", 15382.10])
            proforma.save(proforma_path)

            packing = Workbook()
            ws = packing.active
            ws.append(["PACKING LIST"])
            ws.append(["SHIPPING MARK", "DESCRIPTION", "QUANTITIES (CTNS)", "MEASUREMENT (CBM)", "G.W/N.W (KGS)"])
            ws.append(["N/M", "28CM POT", "797CTNS /3985SETS", 68, "6463.7 KGS / 5738.4KGS"])
            ws.append(["TOTAL", "", "797CTNS /3985SETS", "68 CBM", "6463.7 KGS / 5738.4KGS"])
            packing.save(packing_path)

            analysis = app.analyze_import_cost_files(
                [invoice_path, packing_path, proforma_path],
                {
                    "remittance_rate": "1482.04",
                    "allocation_basis": "amount",
                    "doc_fee": "2240795",
                    "broker_fee": "48400",
                },
            )

            self.assertEqual(analysis["payload"]["invoice_no"], "SXT20260420")
            self.assertTrue(analysis["payload"]["include_import_vat"])
            self.assertTrue(analysis["payload"]["include_service_vat"])
            self.assertEqual(len(analysis["payload"]["products"]), 1)
            product = analysis["payload"]["products"][0]
            self.assertEqual(product["name"], "28CM POT")
            self.assertEqual(product["quantity"], "3985")
            self.assertEqual(product["gross_weight"], "6463.7")
            self.assertEqual(product["cbm"], "68")
            self.assertEqual(analysis["result"]["products"][0]["landed_unit"], 6295.13)

    def test_import_cost_settlement_text_extracts_cost_items(self) -> None:
        app = self.load_app()

        charges = app.parse_import_cost_settlement_text("""
        DOC / FEE
        2,240,795
        관 세 0
        부 가 세 2,418,290
        통관수수료 48,400
        부가세 4,840
        """)

        self.assertEqual(charges["doc_fee"], "2240795")
        self.assertEqual(charges["duty"], "0")
        self.assertEqual(charges["import_vat"], "2418290")
        self.assertEqual(charges["broker_fee"], "48400")
        self.assertEqual(charges["service_vat"], "4840")

    def test_import_cost_settlement_text_treats_negative_duty_ocr_noise_as_zero(self) -> None:
        app = self.load_app()

        charges = app.parse_import_cost_settlement_text("""
        DOC / FEE
        2,141,964
        包 技 -61
        何 啊 技 3,923,910
        烹包荐荐丰 86,350
        """)

        self.assertNotEqual(charges.get("duty"), "-61")
        self.assertEqual(charges["doc_fee"], "2141964")

    def test_import_cost_scanned_pdf_ocr_does_not_auto_apply_costs(self) -> None:
        app = self.load_app()
        with tempfile.TemporaryDirectory() as directory:
            pdf_path = Path(directory) / "scanned.pdf"
            pdf_path.write_bytes(b"not a readable pdf")

            app.ocr_import_cost_pdf_text = lambda path: ("""
            DOC / FEE 000
            관 세 -00261
            부 가 세 24102909
            통관수수료 00000
            """, ["스캔 PDF 1페이지를 OCR로 읽었습니다."])

            parsed = app.parse_import_cost_pdf_text(pdf_path)

        self.assertEqual(parsed["charges"], {})
        self.assertTrue(any("자동 반영하지 않았습니다" in detail for detail in parsed["details"]))

    def test_import_cost_domestic_settlement_pdf_extracts_verified_charges(self) -> None:
        app = self.load_app()

        parsed = app.parse_import_cost_domestic_settlement_text("""
        통관자금(청구)정산서
        B / L XLTSWA26050065
        관 세 0
        부가가치세 2,798,180
        통관수수료 61,600
        D/O 비용 2,101,843
        청구 금액 4,961,623
        JTS SHIPPING CO., LTD.
        OCEAN FREIGHT
        TERMINAL HANDLING CHARGE
        TRUCKING CHARGE/김포신항
        X-RAY 검사료
        INVOICE
        합 계 750.00 2,034,740 67,103
        TOTAL AMOUNT: KRW 2,101,843
        """)

        self.assertTrue(parsed["trusted"])
        self.assertEqual(parsed["hbl_no"], "XLTSWA26050065")
        self.assertEqual(parsed["charges"]["duty"], "0")
        self.assertEqual(parsed["charges"]["import_vat"], "2798180")
        self.assertEqual(parsed["charges"]["broker_fee"], "61600")
        self.assertEqual(parsed["charges"]["doc_fee"], "2034740")
        self.assertEqual(parsed["charges"]["service_vat"], "67103")
        detail_text = "\n".join(parsed["details"])
        self.assertIn("OCEAN FREIGHT: 해상운임", detail_text)
        self.assertIn("TERMINAL HANDLING CHARGE: 터미널 작업료", detail_text)
        self.assertIn("TRUCKING CHARGE: 내륙 운송료", detail_text)
        self.assertIn("X-RAY 검사료", detail_text)

    def test_import_cost_domestic_settlement_keeps_labeled_customs_without_claim_total(self) -> None:
        app = self.load_app()

        parsed = app.parse_import_cost_domestic_settlement_text("""
        통관자금(청구)정산서
        B / L XLTSWA26030027
        관세 0
        부가가치세 2,418,290
        통관수수료 53,240
        JTS SHIPPING CO., LTD.
        TOTAL AMOUNT: KRW 2,141,964
        """)

        self.assertTrue(parsed["trusted"])
        self.assertEqual(parsed["charges"]["duty"], "0")
        self.assertEqual(parsed["charges"]["import_vat"], "2418290")
        self.assertEqual(parsed["charges"]["broker_fee"], "53240")

    def test_import_cost_domestic_settlement_treats_negative_duty_ocr_noise_as_zero(self) -> None:
        app = self.load_app()

        parsed = app.parse_import_cost_domestic_settlement_text("""
        통관자금(청구)정산서
        B / L XLTSWA26030027
        관세 -61
        부가가치세 3,923,910
        통관수수료 86,350
        JTS SHIPPING CO., LTD.
        TOTAL AMOUNT: KRW 2,141,964
        """)

        self.assertTrue(parsed["trusted"])
        self.assertEqual(parsed["charges"]["duty"], "0")
        self.assertEqual(parsed["charges"]["import_vat"], "3923910")
        self.assertEqual(parsed["charges"]["broker_fee"], "86350")

    def test_import_cost_verified_customs_charges_keeps_do_total_separate_when_jts_exists(self) -> None:
        app = self.load_app()

        charges, trusted = app.import_cost_verified_customs_charges({
            "duty": "0",
            "import_vat": "2418290",
            "broker_fee": "53240",
            "do_total": "2240795",
            "claim_amount": "4712325",
        })

        self.assertTrue(trusted)
        self.assertEqual(charges["duty"], "0")
        self.assertEqual(charges["import_vat"], "2418290")
        self.assertEqual(charges["broker_fee"], "53240")
        self.assertNotIn("doc_fee", charges)

    def test_import_cost_jts_invoice_line_items_are_summed_when_total_row_is_ocr_broken(self) -> None:
        app = self.load_app()

        amount, vat = app.import_cost_jts_line_totals([
            "1 | OCEAN FREIGHT USD 1,506.60 700.00 700.00 1,054,620",
            "2 | TERMINAL HANDLING CHARGE KRW 1.00 210,000 210,000",
            "3| CONTAINER CLEANING FEE KRW 1.00 50,000 50,000",
            "4| DOCUMENT FEE KRW 1.00 40,000 40,000",
            "5 | WHARFAGE KRW 1.00 8,400 8,400",
            "6| PORT FACILITY SECURITY KRW 1.00 172 172",
            "7| PORT SAFETY MANAGEMENT CHARGE KRW 1.00 518 518",
            "8 | HANDLING CHARGE(VAT) USD 1,506.60 50.00 50.00 75,330 7,533",
            "9| 검역수수료 KRW 1.00 50,000 100,000 10,000",
            "10| TRUCKING CHARGE/김포신항 KRW 1.00 370,700 370,700 37,070",
            "11] X-RAY 검사료 KRW 1.00 125,000 125,000 12,500",
            "1| OCEAN FREIGHT USD. 1,506.60 700.00 700.00 1,054,620",
            "8 | HANDLING CHARGE(VAT) USD 1,506.60 50.00 50.00 75,330 7,533",
        ])

        self.assertEqual(amount, "2034740")
        self.assertEqual(vat, "67103")
        self.assertEqual(app.import_cost_amounts_from_text("TOTAL AMOUNT: KAW 2,101.843")[-1], "2101843")
        self.assertEqual(app.import_cost_amounts_from_text("USD 1,506.60")[-1], "1506.60")

    def test_import_cost_jts_invoice_includes_freight_surcharge_items(self) -> None:
        app = self.load_app()

        amount, vat = app.import_cost_jts_line_totals([
            "1| OCEAN FREIGHT USD 1,482.04 406.62 406.62 602,640",
            "2| B.A.F USD 1,482.04 386.30 386.30 572,508",
            "3| C.A.F USD 1,482.04 61.00 61.00 90,396",
            "4| CONTAINER IMBALANCE CHARGE USD 1,482.04 81.33 81.33 120,528",
            "5| TERMINAL HANDLING CHARGE KRW 1.00 210,000 210,000",
            "6| CONTAINER CLEANING FEE KRW 1.00 50,000 50,000",
            "7| DOCUMENT FEE KRW 1.00 40,000 40,000",
            "8| WHARFAGE KRW 1.00 8,400 8,400",
            "9| PORT FACILITY SECURITY KRW 1.00 172 172",
            "10| PORT SAFETY MANAGEMENT CHARGE KRW 1.00 518 518",
            "11| HANDLING CHARGE(VAT) USD 1,506.60 50.83 50.83 75,330 7,533",
            "12| 검역수수료 KRW 1.00 50,000 50,000 5,000",
            "13| TRUCKING CHARGE/김포신항 KRW 1.00 370,700 370,700 37,070",
        ])

        self.assertEqual(amount, "2191192")
        self.assertEqual(vat, "49603")

    def test_import_cost_domestic_settlement_uses_jts_total_when_one_line_is_missed_by_ocr(self) -> None:
        app = self.load_app()

        parsed = app.parse_import_cost_domestic_settlement_text("""
        통관자금(청구)정산서
        B / L XLTNGB26040216
        청구 금액 4,712,325
        JTS SHIPPING CO., LTD.
        OCEAN FREIGHT USD 1,506.60 400.00 400.00 602,640
        B.A.F USD 1,506.60 380.00 380.00 572,508
        C.A.F USD 1,506.60 60.00 60.00 90,396
        CONTAINER IMBALANCE CHARGE USD 1,506.60 80.00 80.00 120,528
        TERMINAL HANDLING CHARGE KRW 1.00 210,000 210,000
        CONTAINER CLEANING FEE KRW 1.00 50,000 50,000
        WHARFAGE KRW 1.00 8,400 8,400
        PORT FACILITY SECURITY KRW 1.00 172 172
        PORT SAFETY MANAGEMENT CHARGE KRW 1.00 518 518
        HANDLING CHARGE(VAT) USD 1,506.60 50.00 50.00 75,330 7,533
        검역수수료 KRW 1.00 50,000 50,000 5,000
        TRUCKING CHARGE/김포신항 KRW 1.00 370,700 370,700 37,070
        TOTAL AMOUNT: KRW 2,240,795
        """)

        self.assertTrue(parsed["trusted"])
        self.assertEqual(parsed["charges"]["doc_fee"], "2191192")
        self.assertEqual(parsed["charges"]["service_vat"], "49603")
        self.assertTrue(any("TOTAL AMOUNT" in detail for detail in parsed["details"]))

    def test_sales_report_dashboard_layout_uses_three_report_types(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn('"/api/sales-report-dashboard"', html_source)
        self.assertIn("sales_report_dashboard_payload", html_source)
        self.assertIn("function loadSalesReportDashboard()", html_source)
        self.assertIn("function renderSalesReportDashboard(data)", html_source)
        self.assertIn("function renderSalesReportPeriodOptions", html_source)
        self.assertIn("function formatSalesNumber(value)", html_source)
        self.assertIn("function formatSalesPercent(value)", html_source)
        self.assertIn('id="salesReportPeriodLabel"', html_source)
        self.assertIn('id="salesReportPeriodSelect"', html_source)
        self.assertIn('params.set("period", activeSalesReportPeriod)', html_source)
        self.assertIn('salesReportPeriodSelect?.addEventListener("change"', html_source)
        self.assertIn('id="salesReportKpiGrid"', html_source)
        self.assertIn("매출 마감 자동 점검", html_source)
        self.assertIn("sales_report_closing_check", html_source)
        self.assertIn('"closing_check": closing_check', html_source)
        self.assertIn('id="salesReportDailyBody"', html_source)
        self.assertIn('id="salesReportSellerBody"', html_source)
        self.assertIn('id="salesReportProductBody"', html_source)
        self.assertIn('id="salesReportReviewBody"', html_source)
        self.assertIn('function dashboardRecentSalesRows(dailyRows = [], baseDateText = "")', html_source)
        self.assertIn("data.recent_daily_rows || data.daily_rows || []", html_source)
        self.assertIn('grid-template-columns: repeat(${recentRows.length}, minmax(0, 1fr));', html_source)
        self.assertIn("날짜별", html_source)
        self.assertIn("상품별", html_source)
        self.assertIn("매출처별", html_source)
        self.assertIn("매입처별 총합계 금액", html_source)
        self.assertNotIn('["파일 검증"', html_source)

    def test_daily_sales_detail_popup_uses_full_height_with_compact_density(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertIn(".sales-detail-popup.compact", html_source)
        self.assertIn(".sales-detail-popup.expanded", html_source)
        self.assertIn(".sales-detail-popup.compact .sales-detail-table-wrap", html_source)
        self.assertIn(".sales-detail-popup.expanded .sales-detail-table-wrap", html_source)
        self.assertIn(".sales-detail-sections", html_source)
        self.assertIn(".sales-detail-popup.compact .sales-detail-sections", html_source)
        self.assertIn(".sales-detail-popup.expanded .sales-detail-sections", html_source)
        self.assertIn('salesDetailBody.classList.toggle("has-note", Boolean(data.note));', html_source)
        self.assertIn('salesDetailBody.innerHTML = `${metricHtml}${noteHtml}<div class="sales-detail-sections">${sectionsHtml}</div>`;', html_source)
        self.assertIn("height: calc(100vh - 40px);", html_source)
        self.assertIn("grid-template-rows: auto auto minmax(0, 1fr);", html_source)
        self.assertIn("min-height: 0;", html_source)
        self.assertIn("height: 100% !important;", html_source)
        self.assertIn('const salesDetailPanel = document.querySelector("#salesDetailPopup .sales-detail-popup");', html_source)
        self.assertIn("function setSalesDetailPopupMode(kind)", html_source)
        self.assertIn('salesDetailPanel.classList.toggle("compact", compact);', html_source)
        self.assertIn('salesDetailPanel.classList.toggle("expanded", !compact);', html_source)
        self.assertIn("setSalesDetailPopupMode(data.kind);", html_source)
        self.assertIn("setSalesDetailPopupMode(kind);", html_source)

    def test_lucide_imports_do_not_request_missing_arrow_right_export(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        self.assertNotIn("Search, ArrowRight, LogOut", html_source)
        self.assertNotIn('"arrow-right": ArrowRight', html_source)
        self.assertIn('"arrow-right": ChevronRight', html_source)

    def test_tools_sidebar_is_shared_file_library_only(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn('data-open="fileLibrary"', html_source)
            self.assertIn('id="fileLibraryWorkspace"', html_source)
            self.assertIn('id="sharedFileInput"', html_source)
            self.assertIn('id="sharedFileBody"', html_source)
            self.assertIn('"/api/shared-files"', html_source)
            self.assertIn('"/api/shared-file-upload"', html_source)
            self.assertIn('"/api/shared-file-delete"', html_source)
            self.assertIn('/api/shared-file-download?id=', html_source)
            self.assertIn("function loadSharedFiles()", html_source)
            self.assertIn("function uploadSharedFile()", html_source)
            self.assertIn("function downloadSharedFile", html_source)
            self.assertNotIn('data-open="invoice"><i data-lucide="file-spreadsheet"></i>', html_source)
            self.assertNotIn('data-open="vehicle"><i data-lucide="truck"></i>', html_source)

    def test_work_management_tree_links_to_shared_file_library(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")
        crm_group_start = html_source.index('id="crmNavGroup"')
        crm_group_end = html_source.index("__HERMES_NAV__", crm_group_start)
        crm_group = html_source[crm_group_start:crm_group_end]

        self.assertIn('data-open="fileLibrary"', crm_group)
        self.assertIn("업무 파일", crm_group)
        self.assertIn('id="fileLibraryWorkspace"', html_source)
        self.assertIn('id="sharedFileInput"', html_source)
        self.assertIn('id="sharedFileBody"', html_source)

    def test_excel_downloads_keep_object_url_until_browser_starts_download(self) -> None:
        for app_file in (
            ROOT / "scripts" / "workhub_delivery_app.py",
            ROOT / "_workhub_zip_inspect" / "scripts" / "workhub_delivery_app.py",
        ):
            html_source = app_file.read_text(encoding="utf-8")

            self.assertIn("function downloadWorkbookResponse(response, fallbackName)", html_source)
            self.assertIn("window.setTimeout(() => URL.revokeObjectURL(url), 1000)", html_source)
            self.assertIn('await downloadWorkbookResponse(response, "차량인수증.xlsx")', html_source)
            self.assertIn(
                'await downloadWorkbookResponse(\n            response,\n            currentMode === "invoice" ? "송장번호_추출.xlsx" : "롯데택배_발주서.xlsx"\n          )',
                html_source,
            )

    def test_import_and_cargo_schedule_delete_buttons_are_wired(self) -> None:
        html_source = (ROOT / "scripts" / "workhub_delivery_app.py").read_text(encoding="utf-8")

        for token in (
            'id="importShipmentDelete"',
            'id="cargoShipmentDelete"',
            "function deleteImportShipment()",
            "function deleteCargoShipment()",
            '"/api/import-shipment-delete"',
            '"/api/cargo-shipment-delete"',
            "delete_import_shipment(",
            "delete_cargo_shipment(",
        ):
            self.assertIn(token, html_source)

    def test_import_and_cargo_schedules_can_be_deleted(self) -> None:
        app = self.load_app()
        app.init_db()

        import_id = app.save_import_shipment(
            {
                "warehouse_due_date": "2026-07-10",
                "item": "Test import item",
                "quantity": "1",
            }
        )
        cargo_id = app.save_cargo_shipment(
            {
                "cargo_type": "outbound",
                "ship_date": "2026-07-10",
                "customer": "Test customer",
                "item": "Test cargo item",
                "quantity": "1",
            }
        )

        self.assertEqual([row["id"] for row in app.list_import_shipments()], [import_id])
        self.assertEqual([row["id"] for row in app.list_cargo_shipments()], [cargo_id])
        self.assertEqual(app.delete_import_shipment(import_id), 1)
        self.assertEqual(app.delete_cargo_shipment(cargo_id), 1)
        self.assertEqual(app.list_import_shipments(), [])
        self.assertEqual(app.list_cargo_shipments(), [])


if __name__ == "__main__":
    unittest.main()
