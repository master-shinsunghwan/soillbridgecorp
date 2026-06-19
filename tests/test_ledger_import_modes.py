from __future__ import annotations

import importlib
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))


MANAGEMENT_HEADERS = [
    "매입거래처",
    "매출거래처",
    "거래구분",
    "장부입력확인",
    "주문일",
    "출고일",
    "주문자",
    "발신자연락처",
    "수령자",
    "수령자연락처",
    "제 품 명",
    "수량",
    "상 세 주 소",
    "택배사**",
    "배송번호",
    "특이(요청)사항",
    "주문상품고유번호",
    "상품코드",
    "주문번호",
    "고객선택옵션",
]


CS_HEADERS = [
    "발생일",
    "발생거래처",
    "처리거래처",
    "처리진행상태",
    "처리완료일",
    "주문일자",
    "출고일",
    "주문자",
    "주문자연락처",
    "수령자",
    "수령자연락처",
    "제품명",
    "수량",
    "상세주소",
    "택배사",
    "기존운송장번호",
    "처리내용",
    "CS내용",
]


def load_app(tmp_path: Path):
    os.environ["WORKHUB_DATA_DIR"] = str(tmp_path)
    sys.modules.pop("workhub_delivery_app", None)
    return importlib.import_module("workhub_delivery_app")


def create_management_workbook(path: Path, *, order_item_id: str, receiver_name: str = "이수령") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "통합"
    worksheet.append(MANAGEMENT_HEADERS)
    worksheet.append(
        [
            "거래처A",
            "판매처A",
            "매출",
            "",
            "2026-06-19",
            "2026-06-20",
            "김주문",
            "010-1111-2222",
            receiver_name,
            "010-3333-4444",
            "테스트 제품",
            "1",
            "서울시 테스트구 테스트로 1",
            "롯데택배",
            "1234567890",
            "",
            order_item_id,
            "P-001",
            "ORDER-001",
            "기본",
        ]
    )
    workbook.save(path)


def create_invalid_management_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "통합"
    worksheet.append(MANAGEMENT_HEADERS)
    worksheet.append(
        [
            "거래처A",
            "판매처A",
            "매출",
            "",
            "2026-06-19",
            "2026-06-20",
            "김주문",
            "010-1111-2222",
            "",
            "010-3333-4444",
            "테스트 제품",
            "한개",
            "서울시 테스트구 테스트로 1",
            "롯데택배",
            "1234567890",
            "",
            "ITEM-INVALID",
            "P-001",
            "ORDER-INVALID",
            "기본",
        ]
    )
    workbook.save(path)


def create_cs_workbook(path: Path, *, cs_content: str = "파손 접수") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CS"
    worksheet.append(["CS 처리대장"])
    worksheet.append(CS_HEADERS)
    worksheet.append(
        [
            "2026-06-19",
            "판매처A",
            "거래처A",
            "접수",
            "",
            "2026-06-18",
            "2026-06-19",
            "김주문",
            "010-1111-2222",
            "이수령",
            "010-3333-4444",
            "테스트 제품",
            "1",
            "서울시 테스트구 테스트로 1",
            "롯데택배",
            "1234567890",
            "재발송",
            cs_content,
        ]
    )
    workbook.save(path)


def create_invalid_cs_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CS"
    worksheet.append(["CS 처리대장"])
    worksheet.append(CS_HEADERS)
    worksheet.append(
        [
            "2026-06-19",
            "판매처A",
            "거래처A",
            "접수",
            "",
            "2026-06-18",
            "2026-06-19",
            "김주문",
            "010-1111-2222",
            "이수령",
            "연락처확인",
            "테스트 제품",
            "한개",
            "서울시 테스트구 테스트로 1",
            "롯데택배",
            "",
            "재발송",
            "파손 접수",
        ]
    )
    workbook.save(path)


def test_management_daily_upload_detects_business_duplicates_across_files(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    create_management_workbook(first, order_item_id="ITEM-001")
    create_management_workbook(second, order_item_id="ITEM-001")

    assert app.import_management_workbook(first) == (1, 0)

    preview = app.preview_management_import(second)
    assert preview["total"] == 1
    assert preview["insertable"] == 0
    assert preview["duplicate_existing"] == 1
    assert "이수령" in preview["duplicates"][0]["summary"]

    assert app.import_management_workbook(second) == (0, 1)


def test_management_upload_requires_fixing_invalid_text_and_number_fields(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    workbook_path = tmp_path / "invalid-management.xlsx"
    create_invalid_management_workbook(workbook_path)

    preview = app.preview_management_import(workbook_path)
    assert preview["total"] == 1
    assert preview["insertable"] == 0
    assert preview["invalid_count"] == 1
    invalid_row = preview["invalid_rows"][0]
    issue_fields = {issue["field"] for issue in invalid_row["issues"]}
    assert {"receiver_name", "quantity"}.issubset(issue_fields)

    assert app.import_management_workbook(workbook_path) == (0, 1)
    assert app.import_management_workbook(
        workbook_path,
        corrections=[
            {
                "source_sheet": "통합",
                "source_row": 2,
                "receiver_name": "이수령",
                "quantity": "1",
            }
        ],
    ) == (1, 0)


def test_management_replace_upload_clears_existing_rows(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    first = tmp_path / "first.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    create_management_workbook(first, order_item_id="ITEM-001", receiver_name="이수령")
    create_management_workbook(replacement, order_item_id="ITEM-002", receiver_name="박수령")

    assert app.import_management_workbook(first) == (1, 0)
    assert app.import_management_workbook(replacement, mode="replace") == (1, 0)

    rows = app.list_management_records(limit=None)
    assert len(rows) == 1
    assert rows[0]["receiver_name"] == "박수령"


def test_cs_daily_upload_detects_business_duplicates_across_files(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    first = tmp_path / "cs-first.xlsx"
    second = tmp_path / "cs-second.xlsx"
    create_cs_workbook(first)
    create_cs_workbook(second)

    assert app.import_cs_cases_from_workbook(first) == (1, 0)

    preview = app.preview_cs_cases_import(second)
    assert preview["total"] == 1
    assert preview["insertable"] == 0
    assert preview["duplicate_existing"] == 1
    assert "파손 접수" in preview["duplicates"][0]["summary"]

    assert app.import_cs_cases_from_workbook(second) == (0, 1)


def test_cs_upload_requires_fixing_invalid_text_and_number_fields(tmp_path: Path) -> None:
    app = load_app(tmp_path)
    workbook_path = tmp_path / "invalid-cs.xlsx"
    create_invalid_cs_workbook(workbook_path)

    preview = app.preview_cs_cases_import(workbook_path)
    assert preview["total"] == 1
    assert preview["insertable"] == 0
    assert preview["invalid_count"] == 1
    invalid_row = preview["invalid_rows"][0]
    issue_fields = {issue["field"] for issue in invalid_row["issues"]}
    assert {"original_invoice", "receiver_phone", "quantity"}.issubset(issue_fields)

    assert app.import_cs_cases_from_workbook(workbook_path) == (0, 1)
    assert app.import_cs_cases_from_workbook(
        workbook_path,
        corrections=[
            {
                "source_sheet": "CS",
                "source_row": 3,
                "original_invoice": "1234567890",
                "receiver_phone": "010-3333-4444",
                "quantity": "1",
            }
        ],
    ) == (1, 0)


def main() -> None:
    with tempfile.TemporaryDirectory() as directory:
        test_management_daily_upload_detects_business_duplicates_across_files(Path(directory) / "management-daily")
    with tempfile.TemporaryDirectory() as directory:
        test_management_upload_requires_fixing_invalid_text_and_number_fields(Path(directory) / "management-invalid")
    with tempfile.TemporaryDirectory() as directory:
        test_management_replace_upload_clears_existing_rows(Path(directory) / "management-replace")
    with tempfile.TemporaryDirectory() as directory:
        test_cs_daily_upload_detects_business_duplicates_across_files(Path(directory) / "cs-daily")
    with tempfile.TemporaryDirectory() as directory:
        test_cs_upload_requires_fixing_invalid_text_and_number_fields(Path(directory) / "cs-invalid")


if __name__ == "__main__":
    main()
