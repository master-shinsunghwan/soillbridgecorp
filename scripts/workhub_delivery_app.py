from __future__ import annotations

import json
import mimetypes
import os
import re
import smtplib
import ssl
import sqlite3
import subprocess
import sys
import time
import threading
import base64
import ctypes
import hashlib
import hmac
import secrets
import tempfile
import zipfile
from copy import copy
from io import BytesIO
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from email.utils import formataddr
from html import escape as html_escape
from html.parser import HTMLParser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from delivery_text_summary import build_summary_payload
from invoice_number_exporter import export_invoice_numbers, extract_invoice_rows
from lotte_order_form_converter import convert_lotte_order_form
from sales_vendor_summary_converter import convert_sales_vendor_workbook
from vehicle_receipt_generator import generate_vehicle_receipt
from workhub_crm import (
    add_crm_task_comment,
    change_crm_task_status,
    crm_dashboard_payload,
    delete_crm_saved_view,
    ensure_webhook_token,
    handle_crm_messenger_webhook,
    init_crm_db,
    list_crm_accounts,
    list_crm_message_events,
    list_crm_messenger_users,
    list_crm_saved_views,
    list_crm_task_comments,
    list_crm_tasks,
    rotate_webhook_token,
    save_crm_account,
    save_crm_messenger_user,
    save_crm_saved_view,
    save_crm_task,
)


if getattr(sys, "frozen", False):
    ROOT = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    RUNTIME_ROOT = Path(os.environ.get("LOCALAPPDATA", str(Path.home()))) / "Workhub"
else:
    ROOT = Path(__file__).resolve().parents[1]
    RUNTIME_ROOT = ROOT

RUNTIME_ROOT = Path(os.environ.get("WORKHUB_DATA_DIR", str(RUNTIME_ROOT)))

OUTPUT_DIR = RUNTIME_ROOT / "output" / "workhub_app"
UPLOAD_DIR = OUTPUT_DIR / "uploads"
DOWNLOAD_DIR = OUTPUT_DIR / "downloads"
ORDER_DOWNLOAD_DIR = DOWNLOAD_DIR / "order_outputs"
ORDER_DOWNLOAD_HISTORY_PATH = ORDER_DOWNLOAD_DIR / "history.json"
ORDER_DOWNLOAD_LIMIT = 10
SHARED_FILE_DIR = RUNTIME_ROOT / "shared_files"
SALES_REPORT_DIR = RUNTIME_ROOT / "sales_reports"
CONFIG_DIR = RUNTIME_ROOT / "config"
DB_PATH = CONFIG_DIR / "workhub.db"
MAIL_SETTINGS_PATH = CONFIG_DIR / "mail_settings.json"
VENDOR_CONTACTS_PATH = CONFIG_DIR / "vendor_contacts.json"
CRM_WEBHOOK_TOKEN_PATH = CONFIG_DIR / "crm_webhook_token.txt"
BACKUP_SETTINGS_PATH = CONFIG_DIR / "backup_settings.json"
BACKUP_DIR = Path(os.environ.get("WORKHUB_BACKUP_DIR", str(RUNTIME_ROOT / "backups")))
LUCIDE_DIR = ROOT / "node_modules" / "lucide"
STATIC_DIR = ROOT / "static"
LOTTE_TEMPLATE = ROOT / "templates" / "lotte_order_form_template.xlsx"
MANAGEMENT_EXPORT_TEMPLATE = ROOT / "templates" / "management_ledger_export_template.xlsx"
NAVER_SMTP_HOST = "smtp.naver.com"
NAVER_SMTP_PORT = 465
DEFAULT_MAIL_TECHNICAL_SETTINGS = {
    "smtp_host": NAVER_SMTP_HOST,
    "smtp_port": NAVER_SMTP_PORT,
    "smtp_security": "ssl",
    "bulk_batch_size": 20,
    "bulk_send_interval_seconds": 15,
    "bulk_batch_pause_minutes": 5,
    "bulk_test_recipient": "",
}
SECRET_KEY_PATH = CONFIG_DIR / "secret.key"
TOKEN_PREFIX_DPAPI = "dpapi:"
TOKEN_PREFIX_KEY = "key1:"
SESSION_COOKIE_NAME = "workhub_session"
SESSION_SECONDS = 60 * 60 * 12
SESSION_IDLE_SECONDS = 60 * 60 * 2
LOGIN_MAX_FAILURES = 5
LOGIN_FAILURE_WINDOW_SECONDS = 15 * 60
LOGIN_LOCK_SECONDS = 15 * 60
PASSWORD_MIN_LENGTH = 10
PASSWORD_MAX_LENGTH = 128
BACKUP_RETENTION_DAYS = 90
AUTO_BACKUP_HOUR = 3
MAX_MAIL_ATTACHMENT_BYTES = 20 * 1024 * 1024
_BACKUP_SCHEDULER_STARTED = False
_SYSTEM_UPDATE_LOCK = threading.Lock()
DEFAULT_USERS = (
    ("admin", "관리자", "admin", "admin1234"),
    ("user", "사용자", "user", "user1234"),
)
PERMISSION_DEFINITIONS = (
    ("ledger_delete", "대장 삭제", "통합관리대장/CS처리대장 선택 삭제"),
    ("notice_manage", "공지사항 관리", "공지사항 작성/수정/삭제"),
    ("ledger_edit", "대장 수정", "통합관리대장/CS처리대장 내용 저장"),
    ("excel_upload", "엑셀 업로드", "업로드/대량 등록 기능 사용"),
    ("excel_download", "엑셀 다운로드", "대장 및 변환 결과 다운로드"),
    ("cs_receive", "CS접수", "통합관리대장에서 CS 처리대장 접수"),
    ("mail_send", "메일 발송", "업체 CS 메일 발송"),
    ("import_shipment_manage", "수입제품 진행 관리", "수입제품 출고 진행 입력/완료 처리"),
    ("user_admin", "사용자 관리", "계정 추가/수정/권한 변경"),
    ("backup_manage", "백업 관리", "수동/자동 백업 파일 관리"),
    ("system_update", "시스템 업데이트", "GitHub 업데이트 확인/적용"),
    ("sales_report_manage", "매출현황 관리", "매출표 업로드 및 매출현황 관리"),
    ("leave_view", "연차 조회", "연차 내역 조회"),
    ("leave_approve", "연차 승인", "연차 신청 승인/반려"),
    ("leave_approve_team", "\uC5F0\uCC28 \uD300\uC7A5 \uC2B9\uC778", "\uC5F0\uCC28 \uC2E0\uCCAD \uD300\uC7A5 \uD655\uC778"),
    ("leave_approve_director", "\uC5F0\uCC28 \uC2E4\uC7A5 \uC2B9\uC778", "\uC5F0\uCC28 \uC2E0\uCCAD \uC2E4\uC7A5 \uD655\uC778"),
    ("leave_approve_ceo", "\uC5F0\uCC28 \uB300\uD45C \uC2B9\uC778", "\uC5F0\uCC28 \uC2E0\uCCAD \uB300\uD45C \uD655\uC778"),
    ("leave_director_override", "\uC5F0\uCC28 \uC2E4\uC7A5 \uC804\uACB0", "\uD300\uC7A5/\uB300\uD45C \uC9C0\uC5F0 \uC2DC \uC2E4\uC7A5 \uC804\uACB0 \uCC98\uB9AC"),
    ("leave_manage", "연차 관리", "연차 등록/수정/삭제"),
    ("crm_view", "CRM 조회", "CRM 거래처/업무 조회"),
    ("crm_manage", "CRM 관리", "CRM 거래처/업무 등록 및 수정"),
    ("crm_message_manage", "CRM 메신저 연동", "메신저 사용자 매핑 및 연동 로그 관리"),
)
ALL_PERMISSIONS = tuple(key for key, _, _ in PERMISSION_DEFINITIONS)
DEFAULT_ROLE_PERMISSIONS = {
    "admin": ALL_PERMISSIONS,
    "sub_admin": (
        "ledger_delete",
        "notice_manage",
        "ledger_edit",
        "excel_upload",
        "excel_download",
        "cs_receive",
        "mail_send",
        "import_shipment_manage",
        "leave_view",
        "leave_approve",
        "leave_approve_team",
        "leave_approve_director",
        "leave_director_override",
        "leave_manage",
        "crm_view",
        "crm_manage",
        "crm_message_manage",
    ),
    "user": ("ledger_edit", "excel_download", "cs_receive", "leave_view", "crm_view"),
}
LUCIDE_FALLBACK_JS = """
export function createIcons() {}
export const BriefcaseBusiness = {};
export const Home = {};
export const MessageCircle = {};
export const Info = {};
export const ChevronDown = {};
export const ChevronRight = {};
export const PlusSquare = {};
export const RefreshCw = {};
export const Ellipsis = {};
export const Headphones = {};
export const Package = {};
export const ClipboardCheck = {};
export const CircleDollarSign = {};
export const FileText = {};
export const FileSpreadsheet = {};
export const ClipboardList = {};
export const BarChart3 = {};
export const CopyCheck = {};
export const Bell = {};
export const Download = {};
export const Truck = {};
export const Mail = {};
export const Upload = {};
export const Database = {};
export const CalendarDays = {};
export const X = {};
export const Settings = {};
""".strip()


HTML = r"""<!doctype html>
<html lang="ko" data-theme="light">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>(주)소일브릿지 발주 업무자동화</title>
  <link rel="stylesheet" href="/static/workhub.css" />
  <link href="https://cdn.jsdelivr.net/npm/daisyui@5" rel="stylesheet" type="text/css" />
  <style>
    :root {
      --bg: #f5f7fb;
      --panel: #ffffff;
      --line: #e3e8f2;
      --text: #111827;
      --muted: #667085;
      --navy: #071a3b;
      --navy-2: #10285a;
      --blue: #2563eb;
      --blue-soft: #dbeafe;
      --green: #079455;
      --green-soft: #dcfae6;
      --orange: #d97706;
      --orange-soft: #fef0c7;
      --purple: #7c3aed;
      --purple-soft: #ede9fe;
      --red: #dc2626;
      --red-soft: #fee2e2;
      --sidebar-text-primary: #f8fbff;
      --sidebar-text-secondary: #d7e4ff;
      --sidebar-text-muted: #9fb0d3;
      --sidebar-text-subtle: #b8c7e6;
      --sidebar-accent: #93c5fd;
      --shadow: 0 10px 28px rgba(15, 23, 42, .08);
      font-family: Pretendard, Inter, "Noto Sans KR", "Malgun Gothic", Arial, sans-serif;
    }

    * { box-sizing: border-box; }
    html { height: 100%; }
    body {
      margin: 0;
      color: var(--text);
      background: var(--bg);
      letter-spacing: 0;
      height: 100%;
      overflow: hidden;
    }

    .app {
      min-height: 100vh;
      height: 100vh;
      display: grid;
      grid-template-columns: 248px minmax(0, 1fr);
      overflow: hidden;
    }
    body.standalone .app { grid-template-columns: 248px minmax(0, 1fr); }
    body.standalone .top-search,
    body.standalone .top-tools { display: none; }
    body.standalone .topbar { grid-template-columns: 1fr; }
    .sidebar {
      background: linear-gradient(180deg, var(--navy), #081430);
      color: var(--sidebar-text-primary);
      padding: 22px 16px;
      overflow-y: auto;
      scrollbar-width: thin;
      scrollbar-color: rgba(255,255,255,.28) transparent;
      border-right: 1px solid rgba(255,255,255,.08);
    }
    .brand-icon {
      width: 38px; height: 38px; border-radius: 9px;
      display: grid; place-items: center;
      background: linear-gradient(145deg, #2f6df6, #7c3aed);
      color: white;
      box-shadow: 0 8px 18px rgba(15, 35, 70, .18);
      margin: 0;
      flex: 0 0 auto;
    }
    .brand-icon svg { width: 21px; height: 21px; }
    .brand {
      display: flex;
      align-items: center;
      gap: 11px;
      margin-bottom: 14px;
      padding: 0 4px;
    }
    .brand-label { font-size: 18px; font-weight: 900; line-height: 1.32; margin: 0; color: var(--sidebar-text-primary); }
    .sidebar-search {
      position: relative;
      margin: 0 0 18px;
    }
    .sidebar-search svg {
      position: absolute;
      left: 10px;
      top: 50%;
      transform: translateY(-50%);
      width: 15px;
      height: 15px;
      color: var(--sidebar-text-muted);
      pointer-events: none;
    }
    .sidebar-search input {
      width: 100%;
      height: 36px;
      padding: 0 10px 0 32px;
      border: 1px solid rgba(255,255,255,.14);
      border-radius: 8px;
      background: rgba(255,255,255,.08);
      color: white;
      font-family: inherit;
      font-size: 12px;
      font-weight: 800;
      outline: none;
    }
    .sidebar-search input::placeholder { color: #9fb0d3; }
    .sidebar-search input:focus {
      border-color: rgba(147, 197, 253, .8);
      box-shadow: 0 0 0 2px rgba(59, 130, 246, .22);
    }
    .notice-board {
      min-height: 44px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      padding: 13px 15px;
      color: #344054;
      overflow: hidden;
      box-shadow: var(--shadow);
    }
    .notice-board-kicker {
      font-size: 11px;
      font-weight: 900;
      color: #155bc8;
      margin-bottom: 5px;
    }
    .notice-board-title {
      font-size: 16px;
      font-weight: 900;
      line-height: 1.32;
    }
    .notice-board-meta {
      font-size: 12px;
      color: #667085;
      margin-top: 3px;
    }
    .notice-board-body {
      font-size: 13px;
      line-height: 1.48;
      color: #475467;
      margin-top: 7px;
      max-height: 72px;
      overflow: hidden;
      white-space: pre-line;
    }
    .import-progress-card {
      overflow: hidden;
      background: white;
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .import-progress-head {
      min-height: 44px;
      padding: 0 14px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      border-bottom: 1px solid var(--line);
      background: #fbfcff;
    }
    .import-progress-title {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      border: 0;
      background: transparent;
      color: #111827;
      font-family: inherit;
      font-size: 14px;
      font-weight: 950;
      cursor: pointer;
    }
    .import-progress-title svg {
      width: 16px;
      height: 16px;
      color: #475467;
      transition: transform .16s ease;
    }
    .import-progress-card.open .import-progress-title svg { transform: rotate(90deg); }
    .import-progress-summary {
      color: #667085;
      font-size: 12px;
      font-weight: 800;
      white-space: nowrap;
    }
    .import-progress-actions {
      display: none;
      gap: 8px;
      align-items: center;
      padding: 10px 14px;
      border-bottom: 1px solid var(--line);
      background: white;
    }
    .import-progress-card.open .import-progress-actions {
      display: flex;
    }
    .import-table-wrap {
      overflow: auto;
    }
    .import-table {
      width: 100%;
      min-width: 980px;
      border-collapse: collapse;
      font-size: 12px;
    }
    .import-table th {
      height: 34px;
      padding: 0 8px;
      border: 1px solid #cbd5e1;
      background: #eef6ff;
      color: #1f2937;
      text-align: center;
      font-weight: 950;
      white-space: nowrap;
    }
    .import-table td {
      height: 34px;
      padding: 5px 8px;
      border: 1px solid #e2e8f0;
      text-align: center;
      font-weight: 700;
      white-space: nowrap;
      background: white;
    }
    .import-table td.left { text-align: left; }
    .import-table tr.completed td {
      background: #f3f4f6;
      color: #667085;
    }
    .import-empty {
      padding: 18px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 800;
      text-align: center;
    }
    .import-row-actions {
      display: none;
      gap: 6px;
      align-items: center;
    }
    .import-progress-card.open .import-row-actions { display: inline-flex; }
    .import-row-actions button {
      height: 28px;
      border: 1px solid #cbd5e1;
      border-radius: 6px;
      background: white;
      color: #1f2937;
      font-family: inherit;
      font-size: 12px;
      font-weight: 900;
      cursor: pointer;
    }
    .hidden-file-input {
      position: fixed;
      left: -9999px;
      top: -9999px;
      width: 1px;
      height: 1px;
      opacity: 0;
      pointer-events: none;
    }
    .nav-item, .nav-section, .app-add {
      display: flex; align-items: center; gap: 13px;
      min-height: 43px; padding: 0 12px; border-radius: 8px;
      font-size: 14px; font-weight: 750; color: var(--sidebar-text-secondary);
      margin-bottom: 6px;
      border: 0;
      background: transparent;
      width: 100%;
      font-family: inherit;
      text-align: left;
      cursor: pointer;
      transition: background .16s ease, color .16s ease, transform .16s ease, box-shadow .16s ease;
    }
    .nav-item.active {
      color: var(--sidebar-text-primary);
      background: rgba(72, 118, 255, .28);
      box-shadow: inset 0 0 0 1px rgba(255,255,255,.08);
    }
    .nav-item svg { width: 18px; height: 18px; flex: 0 0 auto; color: var(--sidebar-text-muted); }
    .nav-item .nav-label {
      display: flex;
      align-items: center;
      gap: 13px;
      min-width: 0;
    }
    .nav-item .nav-label span,
    .nav-item > span {
      color: var(--sidebar-text-secondary);
    }
    .nav-item:hover .nav-label span,
    .nav-item:hover > span,
    .nav-item.active .nav-label span,
    .nav-item.active > span {
      color: var(--sidebar-text-primary);
    }
    .nav-item:hover svg,
    .nav-item.active svg {
      color: var(--sidebar-accent);
    }
    .nav-item .nav-chevron {
      margin-left: auto;
      width: 16px;
      height: 16px;
      transition: transform .16s ease;
    }
    .nav-group.open .nav-chevron { transform: rotate(90deg); }
    .nav-submenu {
      display: none;
      margin: -2px 0 8px 31px;
      padding-left: 10px;
      border-left: 1px solid rgba(255,255,255,.14);
    }
    .nav-group.open .nav-submenu { display: grid; gap: 4px; }
    .nav-subitem {
      min-height: 34px;
      padding: 0 10px;
      border: 0;
      border-radius: 7px;
      background: transparent;
      color: var(--sidebar-text-subtle);
      font-family: inherit;
      font-size: 13px;
      font-weight: 750;
      text-align: left;
      cursor: pointer;
      transition: background .16s ease, color .16s ease, transform .16s ease, box-shadow .16s ease;
    }
    .nav-subitem:hover,
    .nav-item:hover {
      background: rgba(255,255,255,.08);
      color: var(--sidebar-text-primary);
      transform: translateX(1px);
    }
    .nav-item.active,
    .nav-subitem.active {
      background: rgba(72, 118, 255, .28);
      color: var(--sidebar-text-primary);
      box-shadow: inset 3px 0 0 rgba(147, 197, 253, .9);
    }
    .nav-section {
      min-height: auto;
      padding: 0 8px 8px;
      margin: 18px 0 0;
      color: var(--sidebar-text-muted);
      font-size: 12px;
      font-weight: 850;
    }
    .hash { font-size: 15px; width: 22px; text-align: center; color: var(--sidebar-accent); font-weight: 900; }
    .divider { height: 1px; background: rgba(255,255,255,.12); margin: 18px 6px; }

    main {
      min-width: 0;
      min-height: 0;
      height: 100vh;
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }
    .topbar {
      height: 74px;
      flex: 0 0 74px;
      display: grid;
      grid-template-columns: 1fr minmax(260px, 430px) auto;
      align-items: center;
      gap: 18px;
      padding: 0 22px;
      background: rgba(245, 247, 251, .88);
      border-bottom: 1px solid rgba(226, 232, 240, .9);
      backdrop-filter: blur(10px);
    }
    .title-wrap { display: grid; gap: 8px; }
    .title { display: flex; align-items: center; gap: 10px; font-size: 25px; font-weight: 900; line-height: 1.2; white-space: nowrap; word-break: keep-all; }
    .subtitle { display: none; }
    .top-actions { display: flex; gap: 12px; }
    .top-button {
      height: 38px; padding: 0 13px; border-radius: 8px; border: 1px solid var(--line);
      background: #fff; display: flex; align-items: center; gap: 8px;
      font-size: 13px; font-weight: 850; color: #344054;
    }
    .top-search {
      height: 42px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      display: flex;
      align-items: center;
      gap: 9px;
      padding: 0 14px;
      color: #98a2b3;
      font-size: 13px;
      font-weight: 700;
    }
    .top-tools {
      display: flex;
      align-items: center;
      gap: 10px;
    }
    .icon-button {
      width: 38px;
      height: 38px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      display: grid;
      place-items: center;
      color: #344054;
      cursor: default;
    }
    .icon-button svg,
    .top-search svg { width: 17px; height: 17px; }
    .user-chip {
      height: 40px;
      display: flex;
      align-items: center;
      gap: 9px;
      padding: 0 11px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      font-size: 13px;
      font-weight: 850;
      white-space: nowrap;
    }
    .logout-button {
      height: 40px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      padding: 0 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      color: #475467;
      font-size: 13px;
      font-weight: 850;
      text-decoration: none;
    }
    .avatar {
      width: 27px;
      height: 27px;
      border-radius: 50%;
      background: linear-gradient(145deg, #155bc8, #08a66c);
    }

    .content {
      flex: 1 1 auto;
      min-height: 0;
      overflow: auto;
      padding: 0 22px 24px;
      display: grid;
      align-content: start;
      gap: 16px;
    }
    .card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .summary { padding: 18px 21px 26px; }
    .section-title { margin: 0 0 26px; font-size: 24px; font-weight: 850; }
    .metrics { display: grid; grid-template-columns: repeat(4, 1fr); gap: 22px; align-items: center; }
    .metric { display: grid; grid-template-columns: 84px 1fr; gap: 18px; align-items: center; position: relative; }
    .metric:not(:last-child)::after {
      content: ""; position: absolute; right: -11px; top: 5px; height: 76px; width: 1px; background: #e0e4eb;
    }
    .metric-icon { width: 84px; height: 84px; border-radius: 17px; display: grid; place-items: center; }
    .metric-icon.red { background: var(--red-soft); color: var(--red); }
    .metric-icon.orange { background: var(--orange-soft); color: #7a430d; }
    .metric-icon.green { background: var(--green-soft); color: #0c5b34; }
    .metric-icon.blue { background: var(--blue-soft); color: var(--blue); }
    .metric-label { font-size: 17px; margin-bottom: 9px; }
    .metric-value { font-size: 40px; line-height: 1; font-weight: 800; color: var(--blue); }
    .metric-value.red { color: var(--red); }
    .metric-value.orange { color: #d16b0b; }
    .metric-value.green { color: var(--green); }

    .stat-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 14px;
    }
    .stat-card {
      padding: 18px 18px 17px;
      min-height: 116px;
      display: grid;
      grid-template-columns: 1fr 56px;
      gap: 10px;
      align-items: center;
    }
    .stat-label {
      color: #344054;
      font-size: 13px;
      font-weight: 850;
      margin-bottom: 10px;
    }
    .stat-value {
      font-size: 28px;
      line-height: 1;
      font-weight: 950;
    }
    .stat-trend {
      margin-top: 12px;
      font-size: 12px;
      color: var(--green);
      font-weight: 850;
    }
    .stat-trend.red { color: var(--red); }
    .stat-icon {
      width: 54px;
      height: 54px;
      border-radius: 16px;
      display: grid;
      place-items: center;
      font-size: 24px;
      font-weight: 950;
    }
    .stat-icon svg { width: 24px; height: 24px; }
    .stat-icon.blue { color: var(--blue); background: var(--blue-soft); }
    .stat-icon.green { color: var(--green); background: var(--green-soft); }
    .stat-icon.orange { color: var(--orange); background: var(--orange-soft); }
    .stat-icon.purple { color: var(--purple); background: var(--purple-soft); }

    .dashboard-card { padding: 0; }
    .dashboard-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      margin-bottom: 13px;
    }
    .dashboard-title {
      font-size: 17px;
      font-weight: 950;
    }
    .notice-template {
      display: grid;
      gap: 10px;
      padding: 14px;
    }
    .notice-template-grid {
      display: grid;
      grid-template-columns: 150px 1fr 160px;
      gap: 8px;
    }
    .notice-template input,
    .notice-template textarea {
      border: 1px solid #cbd5e1;
      border-radius: 7px;
      padding: 0 10px;
      font-family: inherit;
      font-size: 13px;
      background: white;
    }
    .notice-template input { height: 34px; }
    .notice-template textarea {
      min-height: 78px;
      padding-top: 9px;
      resize: vertical;
    }
    .notice-template-actions {
      display: flex;
      justify-content: flex-end;
      gap: 8px;
    }
    .notice-preview {
      border: 1px solid #e5e7eb;
      border-radius: 7px;
      padding: 10px 12px;
      background: #f8fafc;
      font-size: 13px;
      line-height: 1.5;
      color: #344054;
    }
    .notice-preview strong {
      display: block;
      color: #111827;
      font-size: 14px;
      margin-bottom: 3px;
    }
    .message-placeholder {
      display: none;
      padding: 28px;
      border: 1px solid var(--border);
      border-radius: 8px;
      background: #f8fafc;
      color: var(--muted);
      line-height: 1.7;
    }
    .message-placeholder strong {
      display: block;
      color: var(--ink);
      font-size: 18px;
      margin-bottom: 8px;
    }
    .notice-popup-backdrop {
      position: fixed;
      inset: 0;
      display: none;
      place-items: center;
      background: rgba(15, 23, 42, .28);
      z-index: 70;
      padding: 18px;
    }
    .notice-popup-backdrop.open { display: grid; }
    .notice-popup {
      width: min(680px, 100%);
      border-radius: 10px;
      border: 1px solid #d7dce5;
      background: white;
      box-shadow: 0 22px 50px rgba(15, 23, 42, .24);
    }
    .notice-popup-head {
      height: 52px;
      padding: 0 16px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      border-bottom: 1px solid #e5e7eb;
      font-weight: 950;
    }
    .action-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; }
    .action-card {
      min-height: 150px;
      padding: 16px;
      display: grid;
      grid-template-columns: 46px minmax(0, 1fr);
      gap: 12px;
      align-items: start;
      transition: border-color .16s ease, box-shadow .16s ease, transform .16s ease;
    }
    .action-card:hover {
      border-color: #b7c1d1;
      box-shadow: 0 12px 28px rgba(15, 23, 42, .08);
      transform: translateY(-1px);
    }
    .action-main { min-width: 0; }
    .action-top { display: contents; }
    .action-icon { width: 42px; height: 42px; border-radius: 10px; display: grid; place-items: center; flex: 0 0 auto; }
    .action-icon svg { width: 20px; height: 20px; }
    .action-icon.blue { background: var(--blue-soft); color: var(--blue); }
    .action-icon.green { background: var(--green-soft); color: var(--green); }
    .action-icon.orange { background: var(--orange-soft); color: var(--orange); }
    .action-icon.purple { background: var(--purple-soft); color: var(--purple); }
    .action-kicker {
      height: 23px;
      padding: 0 12px;
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      font-size: 11px;
      font-weight: 850;
      background: #f3f6fb;
      color: #566174;
      white-space: nowrap;
      margin-bottom: 9px;
    }
    .action-kicker.blue { background: var(--blue-soft); color: var(--blue); }
    .action-kicker.green { background: var(--green-soft); color: var(--green); }
    .action-kicker.orange { background: var(--orange-soft); color: var(--orange); }
    .action-kicker.purple { background: var(--purple-soft); color: var(--purple); }
    .action-title { font-size: 15px; font-weight: 950; margin: 0 0 6px; }
    .action-sub { min-height: 38px; color: var(--muted); font-size: 12px; line-height: 1.45; margin: 0 0 12px; }
    .action-button {
      height: 31px;
      min-width: 88px;
      padding: 0 12px;
      border-radius: 7px;
      border: 1px solid #cbd5e1;
      background: white;
      color: #1f2937;
      font-size: 12px;
      font-weight: 900;
      cursor: pointer;
    }
    .action-button.green { color: #1f2937; border-color: #cbd5e1; background: white; }
    .action-button.orange { color: #1f2937; border-color: #cbd5e1; background: white; }
    .action-button.purple { color: #1f2937; border-color: #cbd5e1; background: white; }
    .action-button:hover { filter: brightness(.985); }

    .lower { display: grid; grid-template-columns: 2fr 1fr; gap: 34px; }
    .panel { padding: 18px 22px 20px; min-height: 270px; }
    .list-item { display: grid; grid-template-columns: 57px 1fr auto; gap: 17px; align-items: center; padding: 10px 0 18px; border-bottom: 1px solid #e4e7ed; }
    .small-icon { width: 57px; height: 57px; border-radius: 10px; display: grid; place-items: center; }
    .small-icon.blue { background: var(--blue-soft); color: var(--blue); }
    .small-icon.green { background: var(--green-soft); color: var(--green); }
    .small-icon.purple { background: var(--purple-soft); color: var(--purple); }
    .item-title { font-size: 18px; font-weight: 800; margin-bottom: 4px; }
    .item-sub { font-size: 16px; color: var(--muted); }
    .more { width: 100%; height: 53px; border: 1px solid #d7dce5; border-radius: 8px; background: white; font-size: 17px; color: #343b46; margin-top: 12px; }
    .status { padding: 8px 13px; border-radius: 7px; background: #dff5e7; color: #0a6c3a; font-weight: 800; }
    .status.warn { background: #ffe1e1; color: #d20c1c; }

    .workhub-modal-backdrop {
      position: fixed; inset: 0; display: none; place-items: center;
      background: rgba(249, 250, 252, .42);
      z-index: 2000;
    }
    .workhub-modal-backdrop.open { display: grid; }
    .workhub-modal-backdrop.open,
    .workhub-modal-backdrop.open > .workhub-modal {
      visibility: visible !important;
    }
    .workhub-modal {
      display: block;
      width: min(620px, calc(100vw - 38px));
      max-height: calc(100vh - 38px);
      overflow-y: auto;
      background: white; border: 1px solid #b7bdc8; border-radius: 12px;
      box-shadow: var(--shadow); padding: 24px 28px 26px;
      position: relative;
      color: #1a2230;
      opacity: 1;
      pointer-events: auto;
    }
    .safe-number-dialog-backdrop {
      position: fixed;
      inset: 0;
      display: none;
      place-items: center;
      z-index: 3000;
      background: rgba(15, 23, 42, .34);
      padding: 18px;
    }
    .safe-number-dialog-backdrop.open { display: grid; }
    .safe-number-dialog {
      width: min(640px, 100%);
      max-height: calc(100vh - 36px);
      overflow: auto;
      border: 1px solid #b7bdc8;
      border-radius: 12px;
      background: white;
      box-shadow: 0 18px 46px rgba(15, 23, 42, .18);
      padding: 22px 24px 24px;
      color: #1a2230;
    }
    .safe-number-dialog-title {
      margin: 0 0 8px;
      font-size: 20px;
      font-weight: 950;
    }
    .safe-number-dialog-description {
      margin: 0 0 14px;
      color: #475467;
      font-size: 14px;
      line-height: 1.55;
      font-weight: 750;
    }
    .safe-number-dialog-preview {
      max-height: 280px;
      overflow: auto;
      margin: 0;
      padding: 13px 14px;
      border: 1px solid #d7dce5;
      border-radius: 8px;
      background: #f8fafc;
      color: #1f2937;
      font-family: "Malgun Gothic", "Noto Sans KR", sans-serif;
      font-size: 13px;
      line-height: 1.55;
      white-space: pre-wrap;
    }
    .safe-number-dialog-actions {
      display: flex;
      justify-content: flex-end;
      gap: 10px;
      margin-top: 18px;
    }
    .import-correction-dialog {
      width: min(920px, 100%);
    }
    .import-correction-list {
      display: grid;
      gap: 14px;
      max-height: min(58vh, 560px);
      overflow: auto;
      padding-right: 4px;
    }
    .import-correction-row {
      border: 1px solid #d7dce5;
      border-radius: 8px;
      background: #fbfcff;
      padding: 12px;
    }
    .import-correction-row-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 10px;
      color: #344054;
      font-size: 13px;
      font-weight: 900;
    }
    .import-correction-summary {
      color: #667085;
      font-size: 12px;
      font-weight: 800;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }
    .import-correction-fields {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 10px;
    }
    .import-correction-field label {
      display: block;
      margin-bottom: 5px;
      color: #344054;
      font-size: 12px;
      font-weight: 900;
    }
    .import-correction-field input {
      width: 100%;
      height: 36px;
      border: 1px solid #cbd5e1;
      border-radius: 8px;
      padding: 0 10px;
      font-family: inherit;
      font-size: 13px;
      font-weight: 750;
      background: white;
    }
    .import-correction-message {
      margin-top: 5px;
      color: #b42318;
      font-size: 11px;
      font-weight: 800;
      line-height: 1.35;
    }
    .workhub-modal.ledger-modal {
      width: calc(100vw - 18px);
      height: calc(100vh - 18px);
      max-height: calc(100vh - 18px);
      padding: 18px 18px 20px;
      overflow: hidden;
    }
    .workhub-modal.ledger-modal #uploadForm {
      height: calc(100% - 54px);
      display: flex;
      flex-direction: column;
      min-height: 0;
    }
    .workhub-modal-head { display: flex; align-items: center; justify-content: space-between; margin-bottom: 24px; }
    .modal-title { font-size: 25px; font-weight: 850; color: #1a2230; }
    .close { border: 0; background: transparent; color: #3f4650; cursor: pointer; padding: 4px; }
    .field-label { display: block; font-size: 18px; font-weight: 750; margin-bottom: 10px; color: #1a2230; }
    .dropzone {
      border: 1px dashed #9aa4b2; border-radius: 8px; background: #fbfcff;
      padding: 22px; min-height: 112px; display: grid; gap: 8px; align-content: center;
      cursor: pointer;
      transition: border-color .15s ease, background .15s ease, box-shadow .15s ease;
    }
    .dropzone.dragover {
      border-color: var(--blue);
      background: #eef5ff;
      box-shadow: 0 0 0 3px rgba(21, 91, 200, .12);
    }
    .drop-main { font-size: 17px; font-weight: 750; color: #1a2230; }
    .drop-sub { font-size: 14px; color: var(--muted); }
    .workhub-modal-backdrop.open .workhub-modal,
    .workhub-modal-backdrop.open .workhub-modal .modal-title,
    .workhub-modal-backdrop.open .workhub-modal .field-label,
    .workhub-modal-backdrop.open .workhub-modal .drop-main {
      color: #1a2230 !important;
    }
    .workhub-modal-backdrop.open .workhub-modal .drop-sub {
      color: #667085 !important;
    }
    input[type="file"] { display: none; }
    .options { margin-top: 16px; display: flex; gap: 12px; align-items: center; }
    select {
      height: 41px; border: 1px solid #aab2bf; border-radius: 7px; padding: 0 12px;
      font-size: 15px; background: white; min-width: 180px;
    }
    .modal-actions { display: flex; justify-content: flex-end; gap: 14px; margin-top: 24px; }
    .btn {
      height: 50px; min-width: 106px; border-radius: 8px; border: 1px solid #b9c0ca;
      background: white; font-size: 18px; font-weight: 760; cursor: pointer;
    }
    .btn.primary { background: linear-gradient(180deg, #08a66c, #047a4d); border-color: #08794f; color: white; }
    .btn.blue { background: linear-gradient(180deg, #1f73e8, #145bc8); border-color: #145bc8; color: white; }
    .result {
      margin-top: 18px; display: none;
    }
    .result.open { display: block; }
    textarea {
      width: 100%; height: 240px; resize: vertical; border: 1px solid #ccd3dd;
      border-radius: 8px; padding: 14px; font-size: 15px; line-height: 1.55;
      font-family: "Malgun Gothic", "Noto Sans KR", sans-serif;
    }
    .result-actions { display: flex; gap: 10px; justify-content: flex-end; margin-top: 10px; }
    .notice { margin-top: 10px; min-height: 20px; color: var(--muted); font-size: 14px; }
    .admin-panel {
      display: grid;
      gap: 14px;
      padding: 16px;
      background: white;
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .admin-card {
      display: grid;
      gap: 12px;
      padding: 14px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fbfcff;
    }
    #userAdminWorkspace.sales-report-only > .workspace-head,
    #userAdminWorkspace.sales-report-only .workspace-actions,
    #userAdminWorkspace.sales-report-only .admin-panel > .admin-card:not(#salesReportUploadCard),
    #userAdminWorkspace.sales-report-only .admin-panel > .admin-form,
    #userAdminWorkspace.sales-report-only .admin-panel > .permission-grid,
    #userAdminWorkspace.sales-report-only .admin-panel > .admin-message,
    #userAdminWorkspace.sales-report-only .admin-panel > .admin-table-wrap {
      display: none;
    }
    #userAdminWorkspace:not(.sales-report-only) #salesReportUploadCard {
      display: none;
    }
    .sales-dashboard {
      display: grid;
      gap: 12px;
      margin-top: 2px;
    }
    .sales-kpi-grid {
      display: grid;
      grid-template-columns: repeat(6, minmax(0, 1fr));
      gap: 8px;
    }
    .sales-kpi {
      min-height: 88px;
      padding: 10px;
      border: 1px solid #d8e0ec;
      border-radius: 8px;
      background: white;
    }
    .sales-kpi.primary {
      border: 2px solid #2563eb;
    }
    .sales-kpi.warning {
      border-color: #fed7aa;
      background: #fffaf0;
    }
    .sales-kpi.danger {
      border-color: #fecaca;
      background: #fff7f7;
    }
    .sales-kpi-label {
      font-size: 12px;
      color: #64748b;
      font-weight: 950;
    }
    .sales-kpi-value {
      margin-top: 6px;
      font-size: 20px;
      font-weight: 950;
      color: #0f172a;
    }
    .sales-kpi-note {
      margin-top: 4px;
      font-size: 11px;
      color: #64748b;
      font-weight: 800;
    }
    .sales-dashboard-grid {
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
      gap: 10px;
    }
    .sales-panel {
      overflow: hidden;
      border: 1px solid #d8e0ec;
      border-radius: 8px;
      background: white;
    }
    .sales-panel-head {
      min-height: 40px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      padding: 0 12px;
      border-bottom: 1px solid #e2e8f0;
      background: #f8fafc;
      font-size: 13px;
      font-weight: 950;
    }
    .sales-panel-head span {
      color: #64748b;
      font-size: 11px;
      font-weight: 850;
    }
    .sales-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
    }
    .sales-table th {
      height: 32px;
      padding: 0 8px;
      background: #f1f5f9;
      color: #475569;
      border-bottom: 1px solid #e2e8f0;
      text-align: right;
      font-weight: 950;
      white-space: nowrap;
    }
    .sales-table th:first-child,
    .sales-table td:first-child {
      text-align: left;
    }
    .sales-table td {
      height: 32px;
      padding: 5px 8px;
      border-bottom: 1px solid #eef2f7;
      text-align: right;
      font-weight: 750;
      white-space: nowrap;
    }
    .sales-table .empty {
      text-align: center;
      color: #667085;
      height: 46px;
    }
    .sales-positive { color: #047857; }
    .sales-negative { color: #b91c1c; }
    @media (max-width: 1280px) {
      .sales-kpi-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .sales-dashboard-grid { grid-template-columns: 1fr; }
    }
    .admin-section-title {
      font-size: 14px;
      font-weight: 950;
      color: #111827;
    }
    .admin-form {
      display: grid;
      grid-template-columns: 1.1fr 1.1fr .8fr 1fr auto auto;
      gap: 10px;
      align-items: end;
    }
    .admin-form label {
      display: grid;
      gap: 6px;
      font-size: 12px;
      font-weight: 850;
      color: #344054;
    }
    .admin-form input,
    .admin-form select {
      height: 34px;
      border: 1px solid #cfd6e2;
      border-radius: 7px;
      padding: 0 9px;
      font-size: 13px;
      font-weight: 700;
      background: white;
    }
    .admin-check {
      height: 34px;
      display: inline-flex;
      align-items: center;
      gap: 6px;
      font-size: 13px;
      font-weight: 850;
      color: #344054;
    }
    .admin-check input { width: 16px; height: 16px; }
    .permission-grid {
      display: grid;
      grid-template-columns: repeat(5, minmax(0, 1fr));
      gap: 8px;
      padding: 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #f8fafc;
    }
    .permission-item {
      min-height: 42px;
      display: flex;
      align-items: flex-start;
      gap: 7px;
      padding: 8px;
      border: 1px solid #e5e7eb;
      border-radius: 7px;
      background: white;
      font-size: 12px;
      font-weight: 850;
      color: #344054;
    }
    .permission-item input { width: 15px; height: 15px; margin-top: 1px; }
    .permission-item small {
      display: block;
      margin-top: 3px;
      color: #667085;
      font-size: 11px;
      font-weight: 650;
      line-height: 1.35;
      white-space: normal;
    }
    .admin-table-wrap {
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
    }
    .admin-table {
      width: 100%;
      min-width: 900px;
      border-collapse: collapse;
      font-size: 13px;
    }
    .admin-table th {
      height: 34px;
      padding: 0 10px;
      background: #eef6ff;
      border-bottom: 1px solid var(--line);
      text-align: left;
      font-weight: 900;
    }
    .admin-table td {
      height: 38px;
      padding: 4px 10px;
      border-bottom: 1px solid #edf1f7;
      white-space: nowrap;
    }
    .admin-table tr:last-child td { border-bottom: 0; }
    .admin-action {
      height: 30px;
      border: 1px solid #cfd6e2;
      border-radius: 7px;
      background: white;
      font-size: 12px;
      font-weight: 850;
      cursor: pointer;
    }
    .admin-message { min-height: 20px; color: var(--muted); font-size: 13px; font-weight: 750; }
    .permission-hidden { display: none !important; }
    .leave-panel { display: grid; gap: 14px; }
    .leave-summary-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; }
    .leave-summary-card {
      min-height: 86px;
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .leave-summary-card span { display: block; color: var(--muted); font-size: 13px; font-weight: 850; }
    .leave-summary-card strong { display: block; margin-top: 8px; font-size: 28px; line-height: 1; }
    .leave-summary-card.accent strong { color: var(--green); }
    .leave-tabs { display: flex; flex-wrap: wrap; gap: 8px; }
    .leave-tab {
      height: 36px;
      padding: 0 14px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      color: #344054;
      font-size: 13px;
      font-weight: 900;
      cursor: pointer;
    }
    .leave-tab.active { border-color: var(--blue); color: var(--blue); background: #eff6ff; }
    .leave-tab-panel { display: none; }
    .leave-tab-panel.active { display: block; }
    .leave-grid { display: grid; gap: 14px; }
    .leave-grid.two { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    .leave-card {
      min-width: 0;
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .leave-card.narrow { max-width: 720px; }
    .leave-card-title { margin-bottom: 12px; font-size: 16px; font-weight: 950; }
    .leave-form { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 10px; align-items: end; }
    .leave-form label { display: grid; gap: 6px; font-size: 12px; font-weight: 900; color: #344054; }
    .leave-form label.wide { grid-column: 1 / -1; }
    .leave-form input,
    .leave-form select,
    .leave-form textarea {
      width: 100%;
      min-height: 36px;
      border: 1px solid #cfd6e2;
      border-radius: 7px;
      padding: 8px 10px;
      font-size: 13px;
      font-weight: 700;
      background: white;
    }
    .leave-form textarea { min-height: 92px; resize: vertical; }
    .leave-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    .leave-table th {
      height: 34px;
      padding: 0 10px;
      background: #eef6ff;
      border-bottom: 1px solid var(--line);
      text-align: left;
      font-weight: 900;
      white-space: nowrap;
    }
    .leave-table td {
      min-height: 38px;
      padding: 8px 10px;
      border-bottom: 1px solid #edf1f7;
      vertical-align: top;
    }
    .leave-table tr:last-child td { border-bottom: 0; }
    .leave-action-row { display: flex; gap: 6px; }
    .leave-action {
      height: 30px;
      border: 1px solid #cfd6e2;
      border-radius: 7px;
      background: white;
      font-size: 12px;
      font-weight: 850;
      cursor: pointer;
    }
    .leave-action.approve { border-color: #0b8f55; color: #067647; }
    .leave-action.reject { border-color: #f4a7a7; color: #b42318; }
    .leave-comment-input {
      width: 100%;
      min-width: 160px;
      margin-bottom: 6px;
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 8px 10px;
      font-size: 12px;
      font-weight: 750;
    }
    .leave-notification-list {
      display: grid;
      gap: 6px;
      padding: 10px 12px;
      border: 1px solid #dbe7ff;
      border-radius: 8px;
      background: #f8fbff;
      color: #1f3b68;
      font-size: 12px;
      font-weight: 800;
    }
    .leave-notification-list:empty { display: none; }
    .leave-message { min-height: 20px; color: var(--muted); font-size: 13px; font-weight: 750; }
    .backup-panel { display: grid; gap: 14px; }
    .backup-summary-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; }
    .backup-summary-card {
      min-height: 86px;
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .backup-summary-card span { display: block; color: var(--muted); font-size: 13px; font-weight: 850; }
    .backup-summary-card strong {
      display: block;
      margin-top: 8px;
      font-size: 18px;
      line-height: 1.35;
      word-break: break-all;
    }
    .backup-card {
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .backup-note {
      margin: 0 0 12px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 750;
      line-height: 1.6;
    }
    .backup-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    .backup-table th {
      height: 34px;
      padding: 0 10px;
      background: #eef6ff;
      border-bottom: 1px solid var(--line);
      text-align: left;
      font-weight: 900;
      white-space: nowrap;
    }
    .backup-table td {
      height: 40px;
      padding: 6px 10px;
      border-bottom: 1px solid #edf1f7;
      vertical-align: middle;
      white-space: nowrap;
    }
    .backup-table tr:last-child td { border-bottom: 0; }
    .backup-actions { display: flex; gap: 6px; }
    .backup-action {
      height: 30px;
      padding: 0 10px;
      border: 1px solid #cfd6e2;
      border-radius: 7px;
      background: white;
      font-size: 12px;
      font-weight: 850;
      cursor: pointer;
    }
    .backup-action.danger { border-color: #f4a7a7; color: #b42318; }
    .backup-message { min-height: 20px; color: var(--muted); font-size: 13px; font-weight: 750; }
    .backup-restore-row {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
      margin: 0 0 12px;
    }
    .backup-restore-row input { display: none; }
    #backupRestoreInput { display: none; }
    .system-panel { display: grid; gap: 14px; }
    .system-summary-grid { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; }
    .system-summary-card {
      min-height: 92px;
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .system-summary-card span { display: block; color: var(--muted); font-size: 13px; font-weight: 850; }
    .system-summary-card strong {
      display: block;
      margin-top: 8px;
      font-size: 17px;
      line-height: 1.35;
      word-break: break-all;
    }
    .system-card {
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .system-note {
      margin: 0 0 12px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 750;
      line-height: 1.6;
    }
    .system-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    .system-table th {
      height: 34px;
      padding: 0 10px;
      background: #eef6ff;
      border-bottom: 1px solid var(--line);
      text-align: left;
      font-weight: 900;
      white-space: nowrap;
    }
    .system-table td {
      min-height: 40px;
      padding: 8px 10px;
      border-bottom: 1px solid #edf1f7;
      vertical-align: top;
      white-space: nowrap;
    }
    .system-table tr:last-child td { border-bottom: 0; }
    .system-message { min-height: 20px; color: var(--muted); font-size: 13px; font-weight: 750; }
    @media (max-width: 1100px) {
      .admin-form { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .permission-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .leave-grid.two,
      .leave-summary-grid,
      .backup-summary-grid,
      .system-summary-grid { grid-template-columns: 1fr; }
    }
    .vehicle-fields { display: none; }
    .cs-fields,
    .stock-notice-fields { display: none; }
    .ledger-fields { display: none; }
    .management-fields { display: none; }
    .ledger-cs-popup-head { display: none; }
    .workhub-modal.ledger-modal .cs-fields.ledger-cs-popup {
      position: absolute;
      z-index: 35;
      top: 70px;
      right: 22px;
      bottom: 72px;
      display: block !important;
      width: min(560px, calc(100vw - 62px));
      overflow: auto;
      padding: 18px;
      border: 1px solid #9aa4b2;
      border-radius: 10px;
      background: white;
      box-shadow: 0 18px 44px rgba(15, 23, 42, .28);
    }
    .workhub-modal.ledger-modal .cs-fields.ledger-cs-popup .ledger-cs-popup-head {
      position: sticky;
      top: -18px;
      z-index: 2;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin: -18px -18px 14px;
      padding: 15px 18px;
      border-bottom: 1px solid #d7dce5;
      background: white;
    }
    .ledger-cs-popup-title {
      font-size: 20px;
      font-weight: 850;
    }
    .ledger-cs-popup-close {
      height: 34px;
      min-width: 64px;
      border: 1px solid #aab2bf;
      border-radius: 6px;
      background: white;
      font-family: inherit;
      font-weight: 800;
      cursor: pointer;
    }
    .text-field { margin-top: 14px; }
    .text-field input,
    .text-field select,
    .text-field textarea {
      width: 100%;
      border: 1px solid #aab2bf;
      border-radius: 7px;
      font-size: 16px;
      font-family: inherit;
      background: #fff;
    }
    .text-field input,
    .text-field select {
      height: 48px;
      padding: 0 13px;
    }
    .text-field textarea {
      min-height: 78px;
      padding: 12px 13px;
      resize: vertical;
      line-height: 1.45;
    }
    .vendor-picker-button {
      width: 100%;
      min-height: 44px;
      padding: 0 14px;
      border: 1px solid #aab2bf;
      border-radius: 7px;
      background: #fff;
      color: #1a2230;
      font-family: inherit;
      font-size: 14px;
      font-weight: 850;
      text-align: left;
      cursor: pointer;
    }
    .vendor-picker-selected {
      margin-top: 8px;
      color: #667085;
      font-size: 12px;
      font-weight: 800;
    }
    .vendor-picker-tree {
      margin-top: 10px;
      max-height: 240px;
      overflow: auto;
      border: 1px solid #d8e0ee;
      border-radius: 8px;
      background: #fbfcff;
      padding: 10px;
    }
    .vendor-picker-group + .vendor-picker-group { margin-top: 10px; }
    .vendor-picker-group-title {
      margin-bottom: 6px;
      color: #344054;
      font-size: 12px;
      font-weight: 950;
    }
    .vendor-picker-option {
      width: 100%;
      min-height: 36px;
      margin-top: 4px;
      padding: 7px 10px;
      border: 1px solid #e2e8f0;
      border-radius: 7px;
      background: #fff;
      color: #1f2937;
      font-family: inherit;
      font-size: 13px;
      font-weight: 800;
      text-align: left;
      cursor: pointer;
    }
    .vendor-picker-option:hover {
      border-color: #93c5fd;
      background: #eff6ff;
    }
    .vendor-picker-empty {
      padding: 10px;
      color: #667085;
      font-size: 13px;
      font-weight: 800;
    }
    .product-table {
      margin-top: 12px;
      display: grid;
      gap: 8px;
      max-height: 280px;
      overflow: auto;
      padding-right: 4px;
    }
    .product-row {
      display: grid;
      grid-template-columns: 1fr 95px 95px;
      gap: 9px;
    }
    .product-row input {
      height: 42px;
      border: 1px solid #aab2bf;
      border-radius: 7px;
      padding: 0 11px;
      font-size: 15px;
      font-family: inherit;
    }
    .add-row {
      margin-top: 10px;
      height: 40px;
      border: 1px solid #a084c8;
      color: var(--purple);
      background: #fdfbff;
      border-radius: 7px;
      font-size: 13px;
      font-weight: 750;
      cursor: pointer;
    }
    .checkbox-field {
      display: flex;
      align-items: center;
      gap: 8px;
      margin-top: 12px;
      color: #323b48;
      font-size: 15px;
      font-weight: 650;
    }
    .checkbox-field input { width: 17px; height: 17px; }
    .cs-toolbar {
      display: flex;
      gap: 10px;
      margin-top: 12px;
    }
    .cs-save-button {
      height: 40px;
      border: 1px solid #155bc8;
      color: #155bc8;
      background: #f4f8ff;
      border-radius: 7px;
      font-weight: 750;
      cursor: pointer;
      padding: 0 14px;
    }
    .cs-case-list {
      margin-top: 16px;
      border: 1px solid #d7dce5;
      border-radius: 8px;
      overflow: hidden;
      background: #fbfcff;
    }
    .cs-case-head {
      padding: 10px 12px;
      font-size: 15px;
      font-weight: 850;
      border-bottom: 1px solid #d7dce5;
      background: #f3f6fb;
    }
    .cs-case-item {
      padding: 10px 12px;
      border-bottom: 1px solid #e5e9f0;
      font-size: 14px;
      line-height: 1.45;
      color: #253041;
    }
    .cs-case-item:last-child { border-bottom: 0; }
    .cs-case-meta {
      color: #687385;
      font-size: 13px;
      margin-top: 3px;
    }
    .ledger-toolbar {
      display: flex;
      flex-wrap: wrap;
      gap: 7px;
      margin-bottom: 10px;
      align-items: center;
    }
    .ledger-toolbar input { flex: 1 1 280px; min-width: 220px; }
    .ledger-toolbar input,
    .ledger-toolbar select {
      height: 34px;
      border: 1px solid #aab2bf;
      border-radius: 7px;
      padding: 0 9px;
      font-size: 12px;
      font-family: inherit;
      background: white;
    }
    .ledger-toolbar select { flex: 0 0 auto; max-width: 160px; }
    .ledger-toolbar .btn {
      height: 34px;
      min-width: auto;
      padding: 0 11px;
      border-radius: 7px;
      font-size: 12px;
    }
    .ledger-import-button {
      height: 34px;
      min-width: 86px;
      border: 1px solid #0d6ddf;
      border-radius: 7px;
      background: #eef6ff;
      color: #0d56c2;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 6px;
      font-size: 12px;
      font-weight: 850;
      cursor: pointer;
    }
    .ledger-import-button svg {
      width: 14px;
      height: 14px;
    }
    .ledger-count {
      height: 34px;
      display: inline-flex;
      align-items: center;
      padding: 0 8px;
      border: 1px solid #d7dce5;
      border-radius: 7px;
      background: #f8fafc;
      color: #475467;
      font-size: 12px;
      font-weight: 800;
      white-space: nowrap;
    }
    .ledger-import-button input { display: none; }
    .cell-edit-bar {
      display: none;
      grid-template-columns: minmax(150px, auto) minmax(280px, 1fr) auto auto;
      gap: 7px;
      align-items: center;
      margin: -2px 0 10px;
      padding: 8px;
      border: 1px solid #cfd6e2;
      border-radius: 8px;
      background: #f8fafc;
    }
    .cell-edit-bar.open { display: grid; }
    .cell-edit-label {
      color: #344054;
      font-size: 12px;
      font-weight: 900;
      white-space: nowrap;
    }
    .cell-edit-control {
      width: 100%;
      min-height: 34px;
      border: 1px solid #98a2b3;
      border-radius: 7px;
      padding: 6px 9px;
      background: white;
      color: #111827;
      font-family: inherit;
      font-size: 12px;
      font-weight: 750;
    }
    textarea.cell-edit-control {
      min-height: 64px;
      resize: vertical;
      line-height: 1.45;
    }
    .cell-edit-button {
      height: 34px;
      padding: 0 12px;
      border: 1px solid #aab2bf;
      border-radius: 7px;
      background: white;
      font-family: inherit;
      font-size: 12px;
      font-weight: 900;
      cursor: pointer;
    }
    .cell-edit-button.apply {
      border-color: #087a46;
      background: #087a46;
      color: white;
    }
    .management-month-tabs {
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      margin-top: 8px;
      padding: 8px 0 0;
    }
    .management-month-tab {
      height: 30px;
      min-width: 54px;
      padding: 0 10px;
      border: 1px solid #cfd6e2;
      border-radius: 7px;
      background: white;
      color: #344054;
      font-size: 12px;
      font-weight: 850;
      cursor: pointer;
    }
    .management-month-tab.active {
      border-color: #145bc8;
      background: #eef6ff;
      color: #145bc8;
    }
    .management-download-group {
      display: inline-flex;
      flex-wrap: wrap;
      gap: 5px;
      align-items: center;
    }
    .download-menu-wrap {
      position: relative;
      display: inline-flex;
    }
    .download-menu {
      position: absolute;
      top: calc(100% + 5px);
      right: 0;
      z-index: 45;
      display: none;
      width: 174px;
      padding: 7px;
      border: 1px solid #98a2b3;
      border-radius: 7px;
      background: white;
      box-shadow: 0 14px 32px rgba(15, 23, 42, .20);
    }
    .download-menu.open { display: grid; gap: 3px; }
    .download-menu button {
      width: 100%;
      min-height: 29px;
      border: 0;
      border-bottom: 1px solid #eef1f5;
      background: white;
      padding: 5px 7px;
      text-align: left;
      color: #1f2937;
      font-family: inherit;
      font-size: 12px;
      font-weight: 800;
      cursor: pointer;
    }
    .download-menu button:last-child { border-bottom: 0; }
    .download-menu button:hover { background: #eef6ff; }
    .ledger-wrap {
      border: 1px solid #d7dce5;
      border-radius: 8px;
      overflow-x: scroll;
      overflow-y: auto;
      max-height: calc(100vh - 230px);
      max-width: 100%;
      background: white;
      scrollbar-gutter: stable both-edges;
    }
    .workhub-modal.ledger-modal.ledger-view .ledger-fields {
      display: flex !important;
      flex-direction: column;
      flex: 1;
      min-height: 0;
    }
    .workhub-modal.ledger-modal.management-view .management-fields {
      display: flex !important;
      flex-direction: column;
      flex: 1;
      min-height: 0;
    }
    .workhub-modal.ledger-modal .ledger-wrap {
      flex: 1;
      min-height: 0;
      max-height: none;
    }
    .workhub-modal.ledger-modal .management-wrap {
      flex: 1;
      min-height: 0;
      max-height: none;
    }
    .ledger-table {
      width: 100%;
      min-width: 2600px;
      border-collapse: collapse;
      font-size: 12px;
    }
    .ledger-table th {
      position: sticky;
      top: 0;
      z-index: 1;
      background: #8ecf45;
      border-bottom: 1px solid #6baa2d;
      color: #111827;
      font-weight: 850;
      padding: 5px 6px;
      text-align: center;
      white-space: nowrap;
      vertical-align: bottom;
      line-height: 1.18;
    }
    .ledger-table th.has-filter {
      padding-right: 26px;
    }
    .ledger-th-title {
      display: block;
      line-height: 1.2;
    }
    .ledger-filter-trigger {
      position: absolute;
      right: 5px;
      top: 50%;
      transform: translateY(-50%);
      width: 19px;
      height: 19px;
      border: 1px solid rgba(17, 24, 39, .42);
      border-radius: 3px;
      background: rgba(255,255,255,.88);
      color: #111827;
      font-size: 11px;
      line-height: 1;
      cursor: pointer;
      padding: 0;
    }
    .ledger-filter-trigger.active {
      background: #155bc8;
      border-color: #0f4aaa;
      color: white;
    }
    .ledger-filter-popover {
      position: fixed;
      z-index: 40;
      display: none;
      width: 260px;
      max-height: 390px;
      padding: 10px;
      border: 1px solid #98a2b3;
      border-radius: 7px;
      background: white;
      box-shadow: 0 14px 32px rgba(15, 23, 42, .24);
    }
    .ledger-filter-popover.open { display: block; }
    .ledger-filter-title {
      font-size: 13px;
      font-weight: 850;
      margin-bottom: 8px;
    }
    .ledger-filter-search {
      width: 100%;
      height: 34px;
      border: 1px solid #aab2bf;
      border-radius: 5px;
      padding: 0 8px;
      font-family: inherit;
      font-size: 13px;
      margin-bottom: 8px;
    }
    .ledger-filter-option-list {
      max-height: 230px;
      overflow: auto;
      border: 1px solid #e2e6ee;
      border-radius: 5px;
      margin-bottom: 9px;
    }
    .ledger-filter-option {
      width: 100%;
      min-height: 28px;
      border: 0;
      border-bottom: 1px solid #eef1f5;
      background: white;
      padding: 5px 7px;
      text-align: left;
      font-family: inherit;
      font-size: 12px;
      cursor: pointer;
    }
    .ledger-filter-option:hover { background: #eef6ff; }
    .ledger-filter-actions {
      display: flex;
      justify-content: flex-end;
      gap: 7px;
    }
    .ledger-filter-actions button {
      height: 30px;
      border-radius: 5px;
      border: 1px solid #aab2bf;
      background: white;
      font-family: inherit;
      font-weight: 800;
      cursor: pointer;
      padding: 0 10px;
    }
    .ledger-filter-actions .apply {
      border-color: #087a46;
      background: #087a46;
      color: white;
    }
    .ledger-table th.invoice-head {
      background: #f3b21d;
      border-bottom-color: #c98e10;
    }
    .ledger-table td {
      border-bottom: 1px solid #e6eaf0;
      padding: 4px 6px;
      vertical-align: middle;
      text-align: center;
      color: #1f2937;
      font-size: 11px;
      line-height: 1.22;
    }
    .ledger-table tr.completed-cs td {
      background: #fff8d8;
    }
    .ledger-table tr.row-dirty td {
      box-shadow: inset 0 0 0 9999px rgba(37, 99, 235, .035);
    }
    .ledger-table tr.management-duplicate td {
      background: var(--duplicate-row-color, #eef6ff);
    }
    .ledger-table td.left { text-align: left; }
    .ledger-table td.editable-cell {
      cursor: pointer;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
      max-width: 260px;
    }
    .ledger-table td.editable-cell.left {
      max-width: 360px;
    }
    .ledger-table td.editable-cell:hover {
      background: #eef6ff;
    }
    .ledger-table td.editable-cell.selected-cell {
      outline: 2px solid #155bc8;
      outline-offset: -2px;
      background: #eef6ff;
    }
    .ledger-status {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 64px;
      height: 26px;
      border-radius: 999px;
      background: #eaf2ff;
      color: #155bc8;
      font-weight: 800;
      font-size: 12px;
    }
    .ledger-edit,
    .ledger-status-select {
      width: 100%;
      min-width: 118px;
      height: 28px;
      border: 1px solid #aab2bf;
      border-radius: 6px;
      padding: 0 6px;
      font-size: 11px;
      font-family: inherit;
      background: white;
    }
    .ledger-status-select { min-width: 128px; }
    .ledger-save {
      height: 28px;
      min-width: 48px;
      border: 1px solid #087a46;
      border-radius: 6px;
      background: #eefaf3;
      color: #087a46;
      font-weight: 850;
      font-size: 11px;
      cursor: pointer;
    }
    .ledger-check {
      width: 16px;
      height: 16px;
      accent-color: #155bc8;
      cursor: pointer;
    }
    .management-edit {
      width: 100%;
      min-width: 86px;
      height: 26px;
      border: 1px solid transparent;
      border-radius: 5px;
      padding: 0 5px;
      background: rgba(255,255,255,.72);
      font-family: inherit;
      font-size: 11px;
    }
    .management-edit:focus {
      border-color: #2563eb;
      background: white;
      outline: none;
      box-shadow: 0 0 0 2px rgba(37, 99, 235, .12);
    }
    .management-edit.wide { min-width: 260px; }
    .management-cs-button {
      height: 28px;
      min-width: 58px;
      border: 1px solid #155bc8;
      border-radius: 6px;
      background: #eef6ff;
      color: #155bc8;
      font-size: 11px;
      font-weight: 850;
      cursor: pointer;
    }
    .management-cs-button:disabled {
      border-color: #7a5a00;
      background: rgba(255,255,255,.46);
      color: #5f4300;
      cursor: default;
    }
    .ledger-table tr.management-cs-received td {
      background: #ffc000 !important;
    }
    .management-cs-received .management-edit {
      background: rgba(255,255,255,.48);
    }
    .workspace-view {
      display: none;
      flex-direction: column;
      gap: 12px;
      min-height: 0;
      flex: 1 1 auto;
      height: auto;
      overflow: hidden;
      padding: 0 22px 24px;
    }
    .workspace-view.active { display: flex; }
    .order-exec-panel {
      display: grid;
      gap: 14px;
      padding: 18px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .order-exec-summary {
      display: grid;
      gap: 7px;
      max-width: 780px;
    }
    .order-exec-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
    }
    .order-exec-card {
      display: grid;
      gap: 12px;
      align-content: start;
      min-height: 210px;
      padding: 16px;
      border: 1px solid #dfe5ec;
      border-radius: 8px;
      background: #fbfcff;
      cursor: default;
    }
    .order-exec-card:hover {
      border-color: #dfe5ec;
      background: #f5f9ff;
    }
    .order-exec-card-head {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
    }
    .order-exec-title-line {
      display: flex;
      align-items: center;
      gap: 10px;
      margin-bottom: 6px;
    }
    .order-exec-card-icon {
      width: 34px;
      height: 34px;
      display: grid;
      place-items: center;
      flex: 0 0 auto;
      border: 1px solid #d5dce8;
      border-radius: 8px;
      background: white;
      color: #155bc8;
    }
    .order-exec-card-icon svg {
      width: 18px;
      height: 18px;
    }
    .order-exec-card-icon::before {
      content: attr(data-icon);
      font-size: 18px;
      line-height: 1;
    }
    .order-exec-kicker {
      color: #155bc8;
      font-size: 12px;
      font-weight: 950;
    }
    .order-exec-title {
      font-size: 20px;
      font-weight: 950;
      line-height: 1.32;
    }
    .order-exec-description {
      color: #475467;
      font-size: 14px;
      font-weight: 750;
      line-height: 1.55;
    }
    .order-exec-steps {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 10px;
      margin: 0;
      padding: 0;
      list-style: none;
    }
    .order-exec-steps li {
      min-height: 68px;
      padding: 12px;
      border: 1px solid #dfe5ec;
      border-radius: 8px;
      background: #fbfcff;
      color: #344054;
      font-size: 13px;
      font-weight: 800;
      line-height: 1.45;
    }
    .order-exec-actions {
      display: flex;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
    }
    .order-exec-note {
      color: #667085;
      font-size: 12px;
      font-weight: 750;
    }
    .order-download-panel {
      display: grid;
      gap: 12px;
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
    }
    .order-download-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
    }
    .order-download-title {
      font-size: 15px;
      font-weight: 950;
      color: #111827;
    }
    .order-download-list {
      display: grid;
      gap: 8px;
    }
    .order-download-item {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 12px;
      align-items: center;
      padding: 10px 12px;
      border: 1px solid #dfe5ec;
      border-radius: 8px;
      background: #fbfcff;
    }
    .order-download-name {
      font-size: 13px;
      font-weight: 900;
      color: #1f2937;
      overflow-wrap: anywhere;
    }
    .order-download-meta {
      margin-top: 4px;
      font-size: 12px;
      font-weight: 750;
      color: #667085;
    }
    .order-download-empty {
      min-height: 52px;
      display: grid;
      place-items: center;
      border: 1px dashed #cbd5e1;
      border-radius: 8px;
      color: #667085;
      font-size: 13px;
      font-weight: 800;
      background: #fbfcff;
    }
    .shared-file-grid {
      display: grid;
      grid-template-columns: minmax(280px, 420px) minmax(0, 1fr);
      gap: 14px;
      min-height: 0;
    }
    .shared-file-panel {
      display: grid;
      gap: 12px;
      align-content: start;
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      box-shadow: var(--shadow);
      min-width: 0;
    }
    .shared-file-panel.full {
      min-height: 360px;
    }
    .shared-file-message {
      min-height: 20px;
      color: #475467;
      font-size: 13px;
      font-weight: 800;
    }
    .shared-file-table-wrap {
      overflow: auto;
      border: 1px solid #dfe5ec;
      border-radius: 8px;
      background: white;
    }
    .shared-file-table {
      width: 100%;
      min-width: 720px;
      border-collapse: collapse;
      font-size: 13px;
    }
    .shared-file-table th,
    .shared-file-table td {
      min-height: 42px;
      padding: 10px 12px;
      border-bottom: 1px solid #edf1f7;
      text-align: left;
      vertical-align: middle;
    }
    .shared-file-table th {
      background: #f8fafc;
      color: #344054;
      font-weight: 950;
      white-space: nowrap;
    }
    .shared-file-table td {
      color: #1f2937;
      font-weight: 750;
    }
    .shared-file-table .empty {
      text-align: center;
      color: #667085;
      font-weight: 850;
    }
    .shared-file-actions {
      display: inline-flex;
      gap: 7px;
      align-items: center;
      white-space: nowrap;
    }
    @media (max-width: 980px) {
      .shared-file-grid { grid-template-columns: 1fr; }
    }
    .workspace-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      flex: 0 0 auto;
      min-height: 48px;
    }
    .workspace-title {
      font-size: 18px;
      font-weight: 950;
      white-space: nowrap;
      word-break: keep-all;
    }
    .workspace-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
      justify-content: flex-end;
    }
    .workspace-button {
      height: 34px;
      border: 1px solid #cbd5e1;
      border-radius: 7px;
      background: white;
      color: #1f2937;
      font-family: inherit;
      font-size: 12px;
      font-weight: 900;
      padding: 0 12px;
      cursor: pointer;
      white-space: nowrap;
      word-break: keep-all;
    }
    .workspace-button.danger {
      border-color: #f1a7a7;
      background: #fff1f1;
      color: #b42318;
    }
    .workspace-mount {
      flex: 1;
      min-height: 0;
      display: flex;
      flex-direction: column;
    }
    .workspace-view.active .ledger-fields,
    .workspace-view.active .management-fields {
      display: flex !important;
      flex-direction: column;
      flex: 1;
      min-height: 0;
      overflow: hidden;
    }
    .workspace-view.active .ledger-wrap {
      flex: 1 1 auto;
      min-height: 0;
      max-height: none;
      overflow: auto;
    }
    .workspace-view.active .ledger-toolbar,
    .workspace-view.active .management-month-tabs {
      flex: 0 0 auto;
    }

    .crm-tabs {
      display: flex;
      gap: 8px;
      flex: 0 0 auto;
      border-bottom: 1px solid var(--line);
      overflow-x: auto;
      overflow-y: hidden;
    }
    .crm-tab {
      flex: 0 0 auto;
      height: 34px;
      padding: 0 13px;
      border: 0;
      border-bottom: 2px solid transparent;
      background: transparent;
      color: #475467;
      font-family: inherit;
      font-size: 12px;
      font-weight: 950;
      cursor: pointer;
      white-space: nowrap;
      word-break: keep-all;
    }
    .crm-tab.active {
      color: var(--blue);
      border-bottom-color: var(--blue);
    }
    .crm-message {
      display: none;
      min-height: 34px;
      align-items: center;
      padding: 0 12px;
      border: 1px solid #cbd5e1;
      border-radius: 7px;
      background: #f8fafc;
      color: #344054;
      font-size: 12px;
      font-weight: 850;
    }
    .crm-message.open { display: flex; }
    .crm-message.error {
      border-color: #fecaca;
      background: #fff1f1;
      color: #b42318;
    }
    .company-portal {
      display: flex;
      flex-direction: column;
      gap: 12px;
    }
    .company-tabs {
      display: flex;
      gap: 8px;
      border-bottom: 1px solid var(--line);
      overflow-x: auto;
      overflow-y: hidden;
    }
    .company-tab {
      flex: 0 0 auto;
      min-height: 36px;
      padding: 0 12px;
      border: 0;
      border-bottom: 2px solid transparent;
      background: transparent;
      color: #475467;
      font-family: inherit;
      font-size: 12px;
      font-weight: 950;
      cursor: pointer;
      white-space: nowrap;
      word-break: keep-all;
    }
    .company-tab.active {
      color: var(--blue);
      border-bottom-color: var(--blue);
    }
    .company-panel { display: none; }
    .company-panel.active {
      display: flex;
      flex-direction: column;
      gap: 12px;
    }
    .dashboard-import-card .import-table {
      min-width: 720px;
    }
    .company-grid {
      display: grid;
      grid-template-columns: minmax(0, 1.1fr) minmax(320px, .9fr);
      gap: 12px;
      align-items: stretch;
    }
    .company-rule-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
    }
    .company-card {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      overflow: hidden;
    }
    .company-card-head {
      min-height: 42px;
      padding: 0 14px;
      border-bottom: 1px solid var(--line);
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      font-size: 13px;
      font-weight: 950;
      background: #fbfcff;
    }
    .company-card-body {
      padding: 14px;
      color: #344054;
      font-size: 12px;
      font-weight: 750;
      line-height: 1.55;
    }
    .company-notice {
      margin: 0;
      min-height: 210px;
    }
    .company-mini-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      margin-top: 12px;
    }
    .company-mini-grid > div {
      min-height: 64px;
      padding: 10px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #f8fafc;
    }
    .company-mini-grid span {
      display: block;
      margin-bottom: 6px;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .company-mini-grid strong {
      color: #111827;
      font-size: 13px;
      font-weight: 950;
    }
    .company-task-list {
      display: flex;
      flex-direction: column;
      gap: 8px;
      padding: 12px;
      color: #667085;
      font-size: 12px;
      font-weight: 800;
    }
    .company-task-item {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      align-items: center;
      padding: 10px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #fff;
    }
    .company-task-title {
      color: #111827;
      font-size: 13px;
      font-weight: 950;
      line-height: 1.35;
    }
    .company-task-meta {
      margin-top: 4px;
      color: #667085;
      font-size: 11px;
      font-weight: 800;
    }
    .company-calendar-shell {
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(300px, 360px);
      gap: 12px;
      align-items: start;
    }
    .dashboard-calendar-panel {
      margin-top: 10px;
    }
    .dashboard-calendar-panel .company-calendar-shell {
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
      align-items: stretch;
    }
    .dashboard-calendar-card .company-calendar-grid {
      min-height: 360px;
    }
    .dashboard-calendar-card .calendar-day {
      min-height: 60px;
      padding: 6px;
      gap: 4px;
    }
    .dashboard-calendar-card .calendar-event {
      min-height: 18px;
      padding: 3px 5px;
      font-size: 10px;
    }
    .dashboard-sales-panel {
      display: grid;
      gap: 12px;
      align-content: stretch;
      grid-template-rows: minmax(0, 1fr);
    }
    .dashboard-sales-panel > .company-card {
      height: 100%;
    }
    .dashboard-sales-panel .company-card-body {
      display: grid;
      align-content: start;
      min-height: 0;
    }
    .dashboard-sales-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
    }
    .dashboard-sales-metric {
      min-height: 72px;
      padding: 12px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #f8fafc;
    }
    .dashboard-sales-metric span {
      display: block;
      margin-bottom: 8px;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .dashboard-sales-metric strong {
      color: #111827;
      font-size: 18px;
      font-weight: 950;
      line-height: 1.25;
    }
    .dashboard-sales-placeholder {
      margin-top: 12px;
      padding: 13px 14px;
      border: 1px dashed #cbd5e1;
      border-radius: 8px;
      background: #fbfcff;
      color: #667085;
      font-size: 12px;
      font-weight: 850;
      line-height: 1.5;
      text-align: center;
    }
    .company-calendar-toolbar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      background: #fbfcff;
    }
    .company-calendar-title {
      color: #111827;
      font-size: 18px;
      font-weight: 950;
      letter-spacing: 0;
    }
    .company-calendar-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      justify-content: flex-end;
    }
    .company-calendar-legend {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      padding: 10px 14px;
      border-bottom: 1px solid var(--line);
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .calendar-legend-dot {
      width: 9px;
      height: 9px;
      border-radius: 999px;
      display: inline-block;
      margin-right: 5px;
      vertical-align: -1px;
    }
    .calendar-legend-dot.project { background: #155bc8; }
    .calendar-legend-dot.task { background: #0b8f55; }
    .calendar-legend-dot.leave { background: #c2410c; }
    .calendar-legend-dot.pending { background: #9333ea; }
    .company-calendar-weekdays,
    .company-calendar-grid {
      display: grid;
      grid-template-columns: repeat(7, minmax(0, 1fr));
    }
    .company-calendar-weekdays {
      border-bottom: 1px solid var(--line);
      background: #f8fafc;
      color: #667085;
      font-size: 11px;
      font-weight: 950;
      text-align: center;
    }
    .company-calendar-weekdays div {
      padding: 9px 4px;
      border-right: 1px solid #eef2f7;
    }
    .company-calendar-weekdays div:last-child { border-right: 0; }
    .company-calendar-grid {
      min-height: 620px;
      background: #eef2f7;
      gap: 1px;
    }
    .calendar-day {
      min-height: 104px;
      padding: 8px;
      border: 0;
      background: #fff;
      display: flex;
      flex-direction: column;
      gap: 6px;
      text-align: left;
      font-family: inherit;
      cursor: pointer;
    }
    .calendar-day.other-month {
      background: #f8fafc;
      color: #98a2b3;
    }
    .calendar-day.today {
      box-shadow: inset 0 0 0 2px rgba(21, 91, 200, .32);
    }
    .calendar-day.selected {
      box-shadow: inset 0 0 0 2px #155bc8;
    }
    .calendar-day:focus-visible,
    .calendar-event:focus-visible,
    .company-calendar-side .company-task-item:focus-visible {
      outline: 3px solid rgba(21, 91, 200, .35);
      outline-offset: -2px;
    }
    .calendar-date {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 6px;
      color: #344054;
      font-size: 12px;
      font-weight: 950;
    }
    .calendar-date-count {
      min-width: 20px;
      height: 20px;
      padding: 0 6px;
      border-radius: 999px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      background: #eef2ff;
      color: #155bc8;
      font-size: 10px;
      font-weight: 950;
    }
    .calendar-event-list {
      display: grid;
      gap: 4px;
      min-width: 0;
    }
    .calendar-event {
      min-height: 22px;
      padding: 4px 6px;
      border: 1px solid transparent;
      border-radius: 6px;
      overflow: hidden;
      color: #1f2937;
      font-size: 11px;
      font-weight: 850;
      line-height: 1.2;
      white-space: nowrap;
      text-overflow: ellipsis;
      cursor: pointer;
    }
    .calendar-event.project {
      border-color: #bfdbfe;
      background: #eff6ff;
      color: #155bc8;
    }
    .calendar-event.task {
      border-color: #bbf7d0;
      background: #f0fdf4;
      color: #067647;
    }
    .calendar-event.leave {
      border-color: #fed7aa;
      background: #fff7ed;
      color: #c2410c;
    }
    .calendar-event.pending {
      border-color: #ddd6fe;
      background: #f5f3ff;
      color: #7e22ce;
    }
    .calendar-more {
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .company-calendar-side {
      display: grid;
      gap: 12px;
    }
    .calendar-side-date {
      color: #111827;
      font-size: 16px;
      font-weight: 950;
    }
    .calendar-side-list {
      display: grid;
      gap: 8px;
      margin-top: 12px;
    }
    .calendar-empty {
      padding: 18px;
      border: 1px dashed #cbd5e1;
      border-radius: 8px;
      background: #f8fafc;
      color: #667085;
      font-size: 12px;
      font-weight: 850;
      text-align: center;
    }
    .company-quick-links {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 12px;
    }
    .internal-chat {
      display: grid;
      grid-template-columns: 240px minmax(0, 1fr);
      min-height: 520px;
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
      background: white;
    }
    .internal-chat-rooms {
      display: flex;
      flex-direction: column;
      gap: 8px;
      padding: 12px;
      border-right: 1px solid var(--line);
      background: #f8fafc;
      overflow: auto;
    }
    .internal-chat-room {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      width: 100%;
      min-height: 38px;
      padding: 8px 10px;
      border: 1px solid transparent;
      border-radius: 8px;
      background: transparent;
      color: #344054;
      cursor: pointer;
      font-family: inherit;
      font-size: 12px;
      font-weight: 900;
      text-align: left;
    }
    .internal-chat-room:hover,
    .internal-chat-room.active {
      border-color: #bfdbfe;
      background: white;
      color: #1d4ed8;
    }
    .internal-chat-room small {
      color: #667085;
      font-size: 10px;
      font-weight: 850;
    }
    .internal-chat-main {
      display: grid;
      grid-template-rows: auto minmax(360px, 1fr) auto;
      min-width: 0;
    }
    .internal-chat-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      min-height: 58px;
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      background: white;
    }
    .internal-chat-title {
      color: #111827;
      font-size: 14px;
      font-weight: 950;
    }
    .internal-chat-hint {
      color: #667085;
      font-size: 11px;
      font-weight: 800;
    }
    .internal-chat-list {
      display: flex;
      flex-direction: column;
      gap: 10px;
      padding: 14px;
      overflow: auto;
      background: #f8fafc;
    }
    .internal-chat-empty {
      margin: auto;
      color: #667085;
      font-size: 12px;
      font-weight: 850;
      text-align: center;
    }
    .internal-message {
      max-width: min(680px, 82%);
      padding: 10px 12px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: white;
      box-shadow: 0 8px 18px rgba(15, 23, 42, .04);
    }
    .internal-message.mine {
      align-self: flex-end;
      border-color: #bfdbfe;
      background: #eff6ff;
    }
    .internal-message.command-ok {
      border-color: #bbf7d0;
      background: #f0fdf4;
    }
    .internal-message.command-error {
      border-color: #fecaca;
      background: #fef2f2;
    }
    .internal-message-meta {
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      margin-bottom: 5px;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .internal-message-name {
      color: #111827;
      font-weight: 950;
    }
    .internal-message-body {
      color: #1f2937;
      font-size: 13px;
      font-weight: 750;
      line-height: 1.5;
      white-space: pre-wrap;
      word-break: break-word;
    }
    .internal-chat-form {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      padding: 12px;
      border-top: 1px solid var(--line);
      background: white;
    }
    .internal-chat-form textarea {
      min-height: 54px;
      max-height: 130px;
      resize: vertical;
      border: 1px solid #cbd5e1;
      border-radius: 8px;
      padding: 10px;
      font-family: inherit;
      font-size: 13px;
      font-weight: 750;
      line-height: 1.4;
      outline: none;
    }
    .internal-chat-form textarea:focus {
      border-color: #2563eb;
      box-shadow: 0 0 0 2px rgba(37, 99, 235, .12);
    }
    .crm-staff-list {
      display: grid;
      gap: 10px;
    }
    .crm-staff-row {
      display: grid;
      grid-template-columns: minmax(180px, 1fr) repeat(3, minmax(72px, auto)) minmax(220px, 1.2fr);
      gap: 10px;
      align-items: center;
      padding: 12px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #fff;
    }
    .crm-staff-person {
      display: flex;
      align-items: center;
      gap: 10px;
      min-width: 0;
    }
    .crm-staff-avatar {
      display: grid;
      place-items: center;
      width: 36px;
      height: 36px;
      border-radius: 8px;
      background: #eff6ff;
      color: #1d4ed8;
      font-size: 12px;
      font-weight: 950;
      flex: 0 0 auto;
    }
    .crm-staff-name {
      color: #111827;
      font-size: 13px;
      font-weight: 950;
    }
    .crm-staff-role,
    .crm-staff-latest {
      color: #667085;
      font-size: 11px;
      font-weight: 800;
      line-height: 1.45;
    }
    .crm-staff-metric {
      text-align: center;
    }
    .crm-staff-metric span {
      display: block;
      color: #667085;
      font-size: 10px;
      font-weight: 850;
    }
    .crm-staff-metric strong {
      display: block;
      margin-top: 4px;
      color: #111827;
      font-size: 16px;
      font-weight: 950;
    }
    .internal-chat-side {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 8px;
    }
    .internal-chat-side > div {
      min-height: 64px;
      padding: 10px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #f8fafc;
    }
    .internal-chat-side span {
      display: block;
      margin-bottom: 6px;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .internal-chat-side strong {
      color: #111827;
      font-size: 13px;
      font-weight: 950;
    }
    .company-staff-layout {
      display: grid;
      grid-template-columns: minmax(0, 1.45fr) minmax(280px, .55fr);
      gap: 12px;
      align-items: stretch;
    }
    .company-org {
      padding: 14px;
      overflow: auto;
    }
    .company-org-tree {
      display: flex;
      flex-direction: column;
      gap: 14px;
      min-width: 680px;
    }
    .company-org-level {
      display: flex;
      justify-content: center;
      gap: 12px;
      position: relative;
    }
    .company-org-level + .company-org-level::before {
      content: "";
      position: absolute;
      top: -14px;
      left: 50%;
      width: 1px;
      height: 14px;
      background: #d0d5dd;
    }
    .company-org-level.staff {
      justify-content: stretch;
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    }
    .company-person-card {
      min-height: 132px;
      padding: 12px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #fff;
      display: flex;
      flex-direction: column;
      gap: 10px;
      box-shadow: 0 8px 20px rgba(15, 23, 42, .04);
    }
    .company-person-card.lead {
      max-width: 260px;
      border-color: #bfdbfe;
      background: #eff6ff;
    }
    .company-person-card.me {
      border-color: #2563eb;
      box-shadow: 0 0 0 2px rgba(37, 99, 235, .12);
    }
    .company-person-top {
      display: flex;
      align-items: center;
      gap: 10px;
      min-width: 0;
    }
    .company-person-avatar {
      flex: 0 0 auto;
      width: 38px;
      height: 38px;
      border-radius: 50%;
      display: grid;
      place-items: center;
      background: #111827;
      color: white;
      font-size: 13px;
      font-weight: 950;
    }
    .company-person-name {
      min-width: 0;
      color: #111827;
      font-size: 13px;
      font-weight: 950;
      line-height: 1.3;
      word-break: keep-all;
    }
    .company-person-role {
      margin-top: 2px;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .company-person-meta {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 6px;
    }
    .company-person-meta span {
      min-height: 46px;
      padding: 7px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #f8fafc;
      color: #667085;
      font-size: 10px;
      font-weight: 850;
      line-height: 1.2;
    }
    .company-person-meta strong {
      display: block;
      margin-top: 4px;
      color: #111827;
      font-size: 13px;
      font-weight: 950;
    }
    .company-org-empty {
      padding: 18px;
      border: 1px dashed #cbd5e1;
      border-radius: 8px;
      color: #667085;
      font-size: 12px;
      font-weight: 850;
      text-align: center;
    }
    .crm-panel {
      display: none;
      min-height: 0;
      flex: 1;
      overflow: auto;
      padding-bottom: 4px;
    }
    .crm-panel.active {
      display: flex;
      flex-direction: column;
      gap: 12px;
    }
    .crm-panel.active > .crm-card,
    .crm-panel.active > .crm-toolbar,
    .crm-panel.active > .crm-task-board-stats {
      flex: 0 0 auto;
    }
    .crm-stat-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
    }
    .crm-stat {
      min-height: 86px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      padding: 14px;
    }
    .crm-stat-label {
      font-size: 12px;
      color: #667085;
      font-weight: 850;
      margin-bottom: 9px;
    }
    .crm-stat-value {
      font-size: 24px;
      line-height: 1;
      font-weight: 950;
    }
    .crm-project-card {
      overflow: hidden;
    }
    .crm-project-tracker {
      display: grid;
      gap: 8px;
    }
    .crm-project-row {
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #fff;
      padding: 12px;
      display: grid;
      grid-template-columns: minmax(220px, 1fr) minmax(180px, .8fr) minmax(150px, .6fr);
      gap: 12px;
      align-items: center;
      cursor: pointer;
    }
    .crm-project-main {
      min-width: 0;
      display: grid;
      gap: 6px;
    }
    .crm-project-title {
      color: #111827;
      font-size: 14px;
      font-weight: 950;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }
    .crm-project-meta {
      color: #667085;
      font-size: 12px;
      font-weight: 850;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }
    .crm-project-progress {
      display: grid;
      gap: 6px;
    }
    .crm-project-progress-top {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      color: #344054;
      font-size: 12px;
      font-weight: 950;
    }
    .crm-project-bar {
      height: 9px;
      border-radius: 999px;
      background: #eef2f7;
      overflow: hidden;
    }
    .crm-project-bar span {
      display: block;
      height: 100%;
      border-radius: inherit;
      background: linear-gradient(90deg, #155bc8, #16a34a);
    }
    .crm-project-metrics {
      display: flex;
      flex-wrap: wrap;
      justify-content: flex-end;
      gap: 6px;
    }
    .crm-project-pill {
      min-height: 24px;
      border: 1px solid #d8dee9;
      border-radius: 999px;
      padding: 4px 8px;
      color: #344054;
      background: #f8fafc;
      font-size: 11px;
      line-height: 1.2;
      font-weight: 900;
      white-space: nowrap;
    }
    .crm-project-pill.danger {
      color: #b42318;
      border-color: #fecaca;
      background: #fff1f2;
    }
    .crm-project-pill.today {
      color: #155bc8;
      border-color: #bfdbfe;
      background: #eff6ff;
    }
    .crm-project-empty {
      border: 1px dashed #cbd5e1;
      border-radius: 8px;
      padding: 18px;
      color: #667085;
      text-align: center;
      font-size: 12px;
      font-weight: 850;
    }
    .crm-grid-2 {
      display: grid;
      grid-template-columns: minmax(0, 1.1fr) minmax(0, .9fr);
      gap: 12px;
    }
    .crm-card {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      overflow: hidden;
    }
    .crm-card-head {
      min-height: 42px;
      padding: 0 14px;
      border-bottom: 1px solid var(--line);
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      font-size: 13px;
      font-weight: 950;
      background: #fbfcff;
    }
    .crm-card-body { padding: 12px; }
    .crm-task-form.collapsed .crm-card-body,
    .crm-task-form.collapsed #crmTaskSave {
      display: none;
    }
    .crm-toolbar {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
      padding: 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
    }
    .sr-only {
      position: absolute;
      width: 1px;
      height: 1px;
      padding: 0;
      margin: -1px;
      overflow: hidden;
      clip: rect(0, 0, 0, 0);
      white-space: nowrap;
      border: 0;
    }
    .crm-tab:focus-visible,
    .crm-mini-button:focus-visible,
    .crm-input:focus-visible,
    .crm-select:focus-visible,
    .crm-textarea:focus-visible,
    .crm-task-card:focus-visible,
    .crm-project-row:focus-visible,
    .crm-filter-check input:focus-visible {
      outline: 3px solid rgba(21, 91, 200, .35);
      outline-offset: 2px;
    }
    .crm-task-toolbar {
      display: grid;
      grid-template-columns: minmax(150px, .9fr) minmax(180px, 1.2fr) auto auto auto;
      align-items: center;
    }
    .crm-task-toolbar .crm-input,
    .crm-task-toolbar .crm-select {
      width: 100%;
      min-width: 0;
    }
    .crm-advanced-filters {
      grid-column: 1 / -1;
      display: grid;
      grid-template-columns: minmax(140px, .9fr) auto auto repeat(5, minmax(128px, .8fr)) auto;
      gap: 8px;
      align-items: center;
    }
    .crm-advanced-filters[hidden] { display: none; }
    .crm-filter-check {
      min-height: 32px;
      display: inline-flex;
      align-items: center;
      gap: 7px;
      padding: 0 9px;
      border: 1px solid #cbd5e1;
      border-radius: 7px;
      background: #fbfcff;
      color: #344054;
      font-size: 12px;
      font-weight: 900;
      white-space: nowrap;
    }
    .crm-filter-check input {
      width: 15px;
      height: 15px;
    }
    .crm-view-strip {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
      padding: 10px 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fbfcff;
    }
    .crm-view-pill {
      min-height: 30px;
      border: 1px solid #d8dee9;
      border-radius: 7px;
      background: white;
      color: #344054;
      padding: 0 10px;
      font-family: inherit;
      font-size: 12px;
      font-weight: 900;
      cursor: pointer;
    }
    .crm-view-pill.active {
      border-color: #155bc8;
      background: #eef6ff;
      color: #155bc8;
    }
    .crm-input,
    .crm-select,
    .crm-textarea {
      min-height: 32px;
      border: 1px solid #cbd5e1;
      border-radius: 7px;
      padding: 0 10px;
      background: white;
      color: #111827;
      font-family: inherit;
      font-size: 12px;
      font-weight: 700;
    }
    .crm-input { min-width: 160px; }
    .crm-select { min-width: 128px; }
    .crm-textarea {
      height: 68px;
      min-height: 68px;
      padding-top: 8px;
      resize: vertical;
    }
    .crm-form-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 8px;
    }
    .crm-form-grid .wide { grid-column: span 2; }
    .crm-form-grid .full { grid-column: 1 / -1; }
    .crm-table-wrap {
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
    }
    .crm-table {
      width: 100%;
      min-width: 980px;
      border-collapse: collapse;
      font-size: 12px;
    }
    .crm-table th {
      height: 34px;
      padding: 0 9px;
      border-bottom: 1px solid #d8dee9;
      background: #f8fafc;
      color: #344054;
      text-align: left;
      font-weight: 950;
      white-space: nowrap;
    }
    .crm-table td {
      height: 38px;
      padding: 6px 9px;
      border-bottom: 1px solid #eef2f7;
      color: #1f2937;
      font-weight: 750;
      vertical-align: middle;
      white-space: nowrap;
    }
    .crm-table td.left {
      white-space: normal;
      min-width: 220px;
    }
    .crm-status {
      display: inline-flex;
      align-items: center;
      min-height: 24px;
      padding: 0 9px;
      border-radius: 7px;
      font-size: 11px;
      font-weight: 950;
    }
    .crm-status.wait { color: #92400e; background: #fef3c7; }
    .crm-status.progress { color: #155bc8; background: #dbeafe; }
    .crm-status.done { color: #067647; background: #dcfae6; }
    .crm-status.hold { color: #b42318; background: #fee2e2; }
    .crm-mini-actions {
      display: inline-flex;
      gap: 5px;
      align-items: center;
    }
    .crm-mini-button {
      min-height: 27px;
      border: 1px solid #cbd5e1;
      border-radius: 6px;
      background: white;
      color: #1f2937;
      font-family: inherit;
      font-size: 11px;
      font-weight: 900;
      padding: 0 8px;
      cursor: pointer;
    }
    .crm-mini-button.primary {
      border-color: #155bc8;
      background: #eef6ff;
      color: #155bc8;
    }
    .crm-task-board-stats {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
    }
    .crm-board-stat {
      min-height: 70px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      padding: 12px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
    }
    .crm-board-stat span {
      display: block;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
      margin-bottom: 6px;
    }
    .crm-board-stat strong {
      color: #111827;
      font-size: 22px;
      line-height: 1;
      font-weight: 950;
    }
    .crm-board-stat i {
      width: 34px;
      height: 34px;
      border-radius: 8px;
      display: grid;
      place-items: center;
      background: #eef6ff;
      color: #155bc8;
      font-style: normal;
      font-size: 12px;
      font-weight: 950;
    }
    .crm-task-layout {
      display: grid;
      grid-template-columns: minmax(0, 1fr) 340px;
      gap: 12px;
      min-height: 500px;
      align-items: stretch;
    }
    .crm-kanban {
      min-width: 0;
      display: grid;
      grid-template-columns: repeat(4, minmax(185px, 1fr));
      gap: 10px;
      overflow: auto;
      padding-bottom: 2px;
    }
    .crm-kanban-column {
      min-height: 460px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fbfcff;
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }
    .crm-kanban-head {
      min-height: 42px;
      padding: 0 12px;
      border-bottom: 1px solid var(--line);
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      color: #111827;
      font-size: 13px;
      font-weight: 950;
    }
    .crm-kanban-count {
      min-width: 26px;
      height: 24px;
      border-radius: 7px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      background: #eef2ff;
      color: #155bc8;
      font-size: 11px;
      font-weight: 950;
    }
    .crm-kanban-list {
      flex: 1;
      min-height: 0;
      overflow: auto;
      padding: 10px;
      display: grid;
      gap: 10px;
      align-content: start;
    }
    .crm-task-card {
      border: 1px solid #d8dee9;
      border-radius: 8px;
      background: white;
      padding: 12px;
      cursor: pointer;
      display: grid;
      gap: 8px;
      box-shadow: 0 8px 18px rgba(15, 23, 42, .04);
    }
    .crm-task-card.active {
      border-color: #155bc8;
      box-shadow: 0 0 0 2px rgba(21, 91, 200, .12), 0 10px 22px rgba(15, 23, 42, .08);
    }
    .crm-task-card-top,
    .crm-task-card-meta,
    .crm-task-card-actions {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      min-width: 0;
    }
    .crm-task-card-title {
      color: #111827;
      font-size: 13px;
      font-weight: 950;
      line-height: 1.35;
      word-break: keep-all;
    }
    .crm-task-card-sub {
      color: #667085;
      font-size: 11px;
      font-weight: 800;
      line-height: 1.35;
      word-break: keep-all;
    }
    .crm-task-card-meta {
      color: #475467;
      font-size: 11px;
      font-weight: 850;
    }
    .crm-priority {
      display: inline-flex;
      align-items: center;
      min-height: 22px;
      padding: 0 8px;
      border-radius: 7px;
      font-size: 11px;
      font-weight: 950;
      white-space: nowrap;
    }
    .crm-priority.high { background: #fee2e2; color: #b42318; }
    .crm-priority.normal { background: #eef2ff; color: #155bc8; }
    .crm-priority.low { background: #dcfae6; color: #067647; }
    .crm-due-text {
      color: #92400e;
      font-size: 11px;
      font-weight: 950;
      white-space: nowrap;
    }
    .crm-kanban-empty {
      min-height: 72px;
      border: 1px dashed #cbd5e1;
      border-radius: 8px;
      display: grid;
      place-items: center;
      color: #98a2b3;
      font-size: 12px;
      font-weight: 850;
      text-align: center;
      padding: 12px;
    }
    .crm-task-detail {
      min-width: 0;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      overflow: hidden;
      display: flex;
      flex-direction: column;
      align-self: stretch;
    }
    .crm-task-detail-empty {
      min-height: 220px;
      padding: 18px;
      display: grid;
      place-items: center;
      color: #667085;
      font-size: 12px;
      font-weight: 850;
      text-align: center;
      line-height: 1.5;
    }
    .crm-task-detail-inner {
      padding: 16px;
      display: grid;
      gap: 14px;
      overflow: auto;
    }
    .crm-task-detail-kicker {
      color: #155bc8;
      font-size: 11px;
      font-weight: 950;
    }
    .crm-task-detail-title {
      margin-top: 6px;
      color: #111827;
      font-size: 20px;
      line-height: 1.3;
      font-weight: 950;
      word-break: keep-all;
    }
    .crm-task-detail-desc {
      color: #475467;
      font-size: 12px;
      font-weight: 800;
      line-height: 1.5;
      word-break: keep-all;
    }
    .crm-detail-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
    }
    .crm-detail-cell {
      min-height: 66px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #f8fafc;
      padding: 10px;
    }
    .crm-detail-cell span {
      display: block;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
      margin-bottom: 5px;
    }
    .crm-detail-cell strong {
      color: #111827;
      font-size: 13px;
      line-height: 1.3;
      font-weight: 950;
      word-break: break-word;
    }
    .crm-detail-actions {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
    }
    .crm-detail-actions .crm-mini-button {
      min-height: 34px;
    }
    .crm-detail-comment-form {
      display: grid;
      gap: 8px;
      padding-top: 2px;
    }
    .crm-detail-comment-form label,
    .crm-timeline-title {
      color: #344054;
      font-size: 12px;
      font-weight: 950;
    }
    .crm-detail-comment-form textarea {
      width: 100%;
      min-height: 72px;
    }
    .crm-timeline {
      display: grid;
      gap: 8px;
      margin: 0;
      padding: 0;
      list-style: none;
    }
    .crm-timeline-item {
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #fbfcff;
      padding: 10px;
    }
    .crm-timeline-head {
      display: flex;
      justify-content: space-between;
      gap: 8px;
      color: #344054;
      font-size: 11px;
      font-weight: 950;
    }
    .crm-timeline-body {
      margin-top: 5px;
      color: #475467;
      font-size: 12px;
      font-weight: 800;
      line-height: 1.45;
      word-break: keep-all;
    }
    @media (max-width: 1280px) {
      .crm-task-layout { grid-template-columns: 1fr; }
      .crm-task-detail { min-height: 320px; }
      .company-grid { grid-template-columns: 1fr; }
      .company-calendar-shell { grid-template-columns: 1fr; }
      .company-staff-layout { grid-template-columns: 1fr; }
      .crm-task-toolbar { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .crm-advanced-filters { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .crm-project-row { grid-template-columns: 1fr; }
      .crm-project-metrics { justify-content: flex-start; }
    }
    @media (max-width: 980px) {
      .crm-task-board-stats { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .crm-kanban { grid-template-columns: repeat(2, minmax(220px, 1fr)); }
      .crm-grid-2 { grid-template-columns: 1fr; }
      .order-exec-grid { grid-template-columns: 1fr; }
      .order-exec-steps { grid-template-columns: 1fr; }
      .internal-chat { grid-template-columns: 1fr; }
      .internal-chat-rooms { border-right: 0; border-bottom: 1px solid var(--line); max-height: 210px; }
      .crm-staff-row { grid-template-columns: 1fr 1fr 1fr; }
      .crm-staff-person,
      .crm-staff-latest { grid-column: 1 / -1; }
      .company-rule-grid,
      .company-mini-grid,
      .internal-chat-side { grid-template-columns: 1fr; }
      .company-org-tree { min-width: 560px; }
      .company-calendar-grid { min-height: 520px; }
      .calendar-day { min-height: 92px; padding: 6px; }
      .internal-chat-form { grid-template-columns: 1fr; }
      .crm-task-toolbar { grid-template-columns: 1fr; }
      .crm-advanced-filters { grid-template-columns: 1fr; }
    }
    .crm-help {
      margin: 0;
      color: #667085;
      font-size: 12px;
      line-height: 1.5;
      font-weight: 750;
    }
    .crm-webhook-setup {
      display: flex;
      flex-direction: column;
      gap: 8px;
    }
    .crm-token {
      display: grid;
      grid-template-columns: 160px minmax(0, 1fr) auto;
      gap: 8px;
      align-items: center;
      min-height: 36px;
      font-size: 12px;
      font-weight: 850;
      color: #344054;
    }
    .crm-token > span:first-child {
      white-space: nowrap;
    }
    .crm-token code {
      min-width: 0;
      padding: 8px 10px;
      border: 1px solid #e5e7eb;
      border-radius: 7px;
      background: #f8fafc;
      color: #111827;
      word-break: break-all;
    }
    .crm-token-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
    }
    .crm-sample-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 10px;
      margin-top: 4px;
    }
    .crm-sample-title {
      margin-bottom: 6px;
      color: #344054;
      font-size: 12px;
      font-weight: 950;
    }
    .crm-code-block {
      min-height: 132px;
      max-height: 220px;
      overflow: auto;
      margin: 0;
      padding: 10px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #0f172a;
      color: #e5eefb;
      font-family: Consolas, "Courier New", monospace;
      font-size: 11px;
      line-height: 1.45;
      white-space: pre-wrap;
      word-break: break-word;
    }
    .focus-widget-backdrop {
      position: fixed;
      inset: 0;
      z-index: 110;
      display: none;
      place-items: center;
      padding: 22px;
      background: rgba(15, 23, 42, .45);
    }
    .focus-widget-backdrop.open {
      display: grid;
    }
    .focus-widget {
      width: min(1040px, 100%);
      max-height: min(860px, calc(100vh - 44px));
      border: 1px solid #d8dee9;
      border-radius: 8px;
      background: white;
      box-shadow: 0 24px 80px rgba(15, 23, 42, .24);
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }
    .focus-widget-head {
      min-height: 58px;
      padding: 12px 16px;
      border-bottom: 1px solid var(--line);
      background: #fbfcff;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 14px;
    }
    .focus-widget-kicker {
      color: #155bc8;
      font-size: 11px;
      font-weight: 950;
    }
    .focus-widget-title {
      margin-top: 4px;
      color: #111827;
      font-size: 20px;
      line-height: 1.25;
      font-weight: 950;
      word-break: keep-all;
    }
    .focus-widget-subtitle {
      color: #667085;
      font-size: 12px;
      font-weight: 850;
    }
    .focus-widget-close {
      flex: 0 0 auto;
      width: 34px;
      height: 34px;
      border: 1px solid #cbd5e1;
      border-radius: 7px;
      background: white;
      display: grid;
      place-items: center;
      cursor: pointer;
    }
    .focus-widget-body {
      min-height: 0;
      overflow: auto;
      padding: 16px;
      display: grid;
      gap: 14px;
    }
    .focus-widget-section {
      display: grid;
      gap: 10px;
    }
    .focus-widget-section-title {
      color: #344054;
      font-size: 12px;
      font-weight: 950;
    }
    .focus-widget-text {
      margin: 0;
      color: #475467;
      font-size: 14px;
      line-height: 1.65;
      font-weight: 750;
      white-space: pre-wrap;
      word-break: keep-all;
    }
    .focus-widget-grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 10px;
    }
    .focus-widget-metric {
      min-height: 74px;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      background: #f8fafc;
      padding: 12px;
    }
    .focus-widget-metric span {
      display: block;
      color: #667085;
      font-size: 11px;
      font-weight: 850;
      margin-bottom: 8px;
    }
    .focus-widget-metric strong {
      color: #111827;
      font-size: 18px;
      line-height: 1.25;
      font-weight: 950;
      word-break: break-word;
    }
    .focus-widget-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }
    .focus-widget-table {
      width: 100%;
      min-width: 620px;
      border-collapse: collapse;
      font-size: 12px;
    }
    .focus-widget-table th,
    .focus-widget-table td {
      padding: 9px;
      border-bottom: 1px solid #eef2f7;
      text-align: left;
      vertical-align: top;
    }
    .focus-widget-table th {
      background: #f8fafc;
      color: #344054;
      font-weight: 950;
      white-space: nowrap;
    }
    .focus-widget-table td {
      color: #1f2937;
      font-weight: 750;
    }
    .focus-widget-table-wrap {
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .notice-board,
    .company-person-card,
    .crm-staff-row,
    .company-task-item,
    .company-panel[data-company-panel="rules"] .company-card {
      cursor: pointer;
    }
    .notice-board:focus-visible,
    .company-person-card:focus-visible,
    .crm-staff-row:focus-visible,
    .company-task-item:focus-visible,
    .company-panel[data-company-panel="rules"] .company-card:focus-visible,
    .focus-widget-close:focus-visible {
      outline: 3px solid rgba(21, 91, 200, .35);
      outline-offset: 2px;
    }
    @media (max-width: 760px) {
      .focus-widget-backdrop {
        padding: 10px;
      }
      .focus-widget {
        max-height: calc(100vh - 20px);
      }
      .focus-widget-head {
        align-items: flex-start;
      }
      .focus-widget-title {
        font-size: 17px;
      }
      .focus-widget-grid {
        grid-template-columns: 1fr;
      }
    }

    /* Compact typography pass: keep controls usable while reducing visual bulk. */
    .brand-label { font-size: 16px; }
    .nav-item, .nav-section, .app-add { font-size: 13px; }
    .nav-subitem { font-size: 12px; }
    .title { font-size: 22px; }
    .top-button,
    .top-search,
    .user-chip,
    .logout-button { font-size: 12px; }
    .notice-board-title { font-size: 15px; }
    .notice-board-body,
    .import-empty,
    .notice-preview,
    .backup-note,
    .system-note,
    .system-message,
    .backup-message,
    .leave-message,
    .admin-message { font-size: 12px; }
    .dashboard-title { font-size: 15px; }
    .action-title { font-size: 14px; }
    .action-sub { min-height: 34px; font-size: 11px; }
    .modal-title { font-size: 22px; }
    .field-label { font-size: 15px; }
    .drop-main { font-size: 15px; }
    .drop-sub { font-size: 12px; }
    select,
    textarea,
    .text-field input,
    .text-field select,
    .text-field textarea { font-size: 13px; }
    .btn { font-size: 14px; }
    .notice { font-size: 12px; }
    .notice-template input,
    .notice-template textarea,
    .admin-form input,
    .admin-form select,
    .admin-check,
    .leave-form input,
    .leave-form select,
    .leave-form textarea,
    .backup-summary-card span,
    .system-summary-card span,
    .leave-summary-card span { font-size: 12px; }
    .permission-item,
    .admin-table,
    .leave-table,
    .system-table,
    .import-table { font-size: 12px; }
    .backup-summary-card strong { font-size: 16px; }
    .leave-summary-card strong { font-size: 25px; }
    .leave-card-title { font-size: 15px; }
    .ledger-cs-popup-title { font-size: 18px; }
    .product-row input,
    .checkbox-field,
    .cs-case-head { font-size: 13px; }
    .cs-case-item { font-size: 12px; }
    .cs-case-meta { font-size: 11px; }
    .ledger-toolbar input,
    .ledger-toolbar select,
    .ledger-toolbar .btn,
    .ledger-count,
    .management-month-tab { font-size: 11px; }
    .ledger-table { font-size: 11px; }
    .ledger-filter-trigger,
    .ledger-table td,
    .ledger-edit,
    .ledger-status-select,
    .ledger-save,
    .management-edit,
    .management-cs-button { font-size: 10px; }
    .ledger-filter-title,
    .ledger-filter-search { font-size: 12px; }
    .ledger-filter-option,
    .workspace-button { font-size: 11px; }
    .workspace-title { font-size: 16px; }
    .ledger-table th {
      padding: 3px 5px;
      line-height: 1.05;
      height: 26px;
      vertical-align: middle;
    }
    .ledger-table th.has-filter {
      padding-right: 22px;
    }
    .ledger-th-title {
      line-height: 1.08;
    }
    .ledger-filter-trigger {
      width: 16px;
      height: 16px;
      right: 4px;
      font-size: 9px;
    }
    .ledger-table td {
      padding: 2px 4px;
      line-height: 1.1;
    }
    .ledger-edit,
    .ledger-status-select {
      height: 23px;
      min-width: 104px;
      padding: 0 5px;
      border-radius: 5px;
    }
    .ledger-status-select {
      min-width: 112px;
    }
    .management-edit {
      height: 22px;
      min-width: 78px;
      padding: 0 4px;
    }
    .management-edit.wide {
      min-width: 220px;
    }
    .management-cs-button,
    .ledger-save {
      height: 24px;
    }

    @media (max-width: 1180px) {
      .app,
      body.standalone .app { grid-template-columns: 76px minmax(0, 1fr); }
      .brand { justify-content: center; padding: 0; }
      .sidebar-search { display: none; }
      .brand-label, .nav-item > span, .nav-label span, .nav-section, .nav-submenu { display: none; }
      .nav-item { justify-content: center; padding: 0; }
      .stat-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .action-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 760px) {
      .app,
      body.standalone .app { grid-template-columns: 1fr; }
      .sidebar { display: none; }
      .topbar {
        height: auto;
        padding: 16px 12px;
        grid-template-columns: 1fr;
      }
      .top-tools { display: none; }
      .content { padding: 0 12px 18px; }
      .title { font-size: 22px; }
      .stat-grid,
      .action-grid { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <div class="app">
    <aside class="sidebar">
      <div class="brand">
        <div class="brand-icon"><i data-lucide="briefcase-business"></i></div>
        <div class="brand-label">(주)소일브릿지<br>업무자동화</div>
      </div>
      <label class="sidebar-search" for="sidebarSearchInput">
        <i data-lucide="search"></i>
        <input id="sidebarSearchInput" name="workhub-menu-search" type="search" placeholder="메뉴 검색" autocomplete="off" autocapitalize="off" autocorrect="off" spellcheck="false" />
      </label>
      <div class="nav-section">MAIN</div>
      <div class="nav-group open" id="companyNavGroup">
        <button class="nav-item active" id="companyNavToggle" type="button" data-view="dashboard" data-company-tab="notice">
          <span class="nav-label"><i data-lucide="home"></i> <span>회사 포털</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem active" type="button" data-view="dashboard" data-company-tab="notice">공지사항</button>
          <button class="nav-subitem" type="button" data-view="dashboard" data-company-tab="calendar">캘린더</button>
          <button class="nav-subitem" type="button" data-view="dashboard" data-company-tab="rules">사규/가이드</button>
          <button class="nav-subitem" type="button" data-view="dashboard" data-company-tab="staff">직원 대시보드</button>
          <button class="nav-subitem" type="button" data-view="dashboard" data-company-tab="chat">사내 메신저</button>
        </div>
      </div>
      <div class="nav-group" id="importNavGroup">
        <button class="nav-item" id="importNavToggle" type="button" data-open="import">
          <span class="nav-label"><i data-lucide="truck"></i> <span>수출입 업무</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem" id="importShipmentInputOpen" type="button">수입제품 출고 진행 입력</button>
        </div>
      </div>
      <div class="nav-group" id="orderNavGroup">
        <button class="nav-item" id="orderNavToggle" type="button" data-open="order">
          <span class="nav-label"><i data-lucide="clipboard-list"></i> <span>발주업무</span></span>
        </button>
      </div>
      <div class="nav-group" id="managementNavGroup">
        <button class="nav-item" id="managementNavToggle" type="button" data-open="management">
          <span class="nav-label"><i data-lucide="database"></i> <span>통합관리대장 관리</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem" id="managementImportOpen" type="button" data-management-import-mode="daily">통합관리대장 일일 추가 업로드</button>
        </div>
      </div>
      <div class="nav-group" id="ledgerNavGroup">
        <button class="nav-item" id="ledgerNavToggle" type="button" data-open="ledger">
          <span class="nav-label"><i data-lucide="clipboard-check"></i> <span>CS 처리대장</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem" id="ledgerImportOpen" type="button" data-ledger-import-mode="daily">CS처리대장 일일 추가 업로드</button>
          <button class="nav-subitem" type="button" data-mail-popup="cs">CS처리 요청</button>
        </div>
      </div>
      <div class="nav-group" id="crmNavGroup">
        <button class="nav-item" id="crmNavToggle" type="button" data-open="crm" data-crm-nav-tab="dashboard">
          <span class="nav-label"><i data-lucide="message-circle"></i> <span>업무관리</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem" type="button" data-open="crm" data-crm-nav-tab="dashboard">업무 현황</button>
          <button class="nav-subitem" type="button" data-open="crm" data-crm-nav-tab="mine">내 업무</button>
          <button class="nav-subitem" type="button" data-open="crm" data-crm-nav-tab="tasks">업무보드</button>
          <button class="nav-subitem" type="button" data-open="crm" data-crm-nav-tab="accounts">직원 현황</button>
          <button class="nav-subitem" type="button" data-open="crm" data-crm-nav-tab="messages">메신저 연동</button>
        </div>
      </div>
      <div class="nav-group" id="distributionMailNavGroup">
        <button class="nav-item" id="distributionMailNavToggle" type="button">
          <span class="nav-label"><i data-lucide="mail"></i> <span>유통사 업무관련 메일 발송</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem" type="button" data-mail-popup="cs">CS 요청</button>
          <button class="nav-subitem" type="button" data-mail-popup="stock">입고 및 품절 공지</button>
        </div>
      </div>
      __LEAVE_NAV__
      __SALES_REPORT_NAV__
      __ADMIN_TOOLS_NAV__
      <div class="nav-section">보조 도구</div>
      <button class="nav-item" type="button" data-open="fileLibrary"><i data-lucide="download"></i> <span>업무 파일 자료실</span></button>
    </aside>

    <main>
      <header class="topbar">
        <div class="title-wrap">
          <div class="title">회사 포털 <i data-lucide="chevron-down"></i></div>
          <p class="subtitle">공지사항, 사규, 직원 현황을 한 곳에서 확인합니다.</p>
        </div>
        <div class="top-search"><i data-lucide="file-text"></i> 파일명, 수령인, 송장번호, CS내용 검색</div>
        <div class="top-tools">
          <button class="icon-button" type="button"><i data-lucide="bell"></i></button>
          <button class="icon-button" type="button"><i data-lucide="refresh-cw"></i></button>
          <div class="user-chip"><span class="avatar"></span><span>__USER_DISPLAY__</span></div>
          <a class="logout-button" href="/logout">로그아웃</a>
        </div>
      </header>

      <section class="content company-portal" id="dashboardContent">
        <div class="company-tabs">
          <button class="company-tab active" type="button" data-company-tab="notice">공지사항</button>
          <button class="company-tab" type="button" data-company-tab="calendar">캘린더</button>
          <button class="company-tab" type="button" data-company-tab="rules">사규/가이드</button>
          <button class="company-tab" type="button" data-company-tab="staff">직원 대시보드</button>
          <button class="company-tab" type="button" data-company-tab="chat">사내 메신저</button>
        </div>

        <section class="company-panel active" data-company-panel="notice">
          <div class="company-grid">
            <section class="notice-board company-notice" id="sidebarNoticePreview" role="button" tabindex="0" aria-label="공지사항 크게 보기">
              <div class="notice-board-kicker">금일 공지사항</div>
              <div class="notice-board-title">등록된 공지 없음</div>
              <div class="notice-board-body">공지사항 입력 버튼을 눌러 내용을 입력해주세요.</div>
            </section>
            <article class="company-card">
              <div class="company-card-head">
                <span>공지 관리</span>
                <button class="workspace-button" id="noticeInputOpen" type="button">공지사항 입력</button>
              </div>
              <div class="company-card-body">
                <p>오늘 공유해야 할 출고 마감, 업체 회신 필요 건, 내부 전달사항을 이곳에서 관리합니다.</p>
                <div class="company-mini-grid">
                  <div><span>저장 방식</span><strong>브라우저 localStorage</strong></div>
                  <div><span>권한</span><strong>공지사항 관리</strong></div>
                </div>
              </div>
            </article>
          </div>
        </section>

        <section class="import-progress-card dashboard-import-card open" id="dashboardImportScheduleCard">
          <div class="import-progress-head">
            <button class="import-progress-title" type="button" id="dashboardImportScheduleOpen">
              <i data-lucide="package"></i>
              <span>수입제품 입고 일정</span>
            </button>
            <div class="import-progress-summary" id="dashboardImportScheduleSummary">진행 0건</div>
          </div>
          <div class="import-progress-actions">
            <button class="workspace-button" type="button" id="dashboardImportScheduleRefresh">새로고침</button>
            <button class="workspace-button" type="button" data-view="import">전체 보기</button>
          </div>
          <div class="import-table-wrap">
            <table class="import-table">
              <thead>
                <tr>
                  <th>입항일</th>
                  <th>품명</th>
                  <th>수량</th>
                  <th>진행상태</th>
                  <th>입고/반입 일정</th>
                </tr>
              </thead>
              <tbody id="dashboardImportScheduleBody">
                <tr><td colspan="5"><div class="import-empty">수입제품 입고 일정을 불러오는 중입니다.</div></td></tr>
              </tbody>
            </table>
          </div>
        </section>

        <section class="company-panel active dashboard-calendar-panel" data-company-panel="calendar">
          <div class="company-calendar-shell">
            <article class="company-card dashboard-calendar-card">
              <div class="company-calendar-toolbar">
                <div class="company-calendar-title" id="companyCalendarTitle">캘린더</div>
                <div class="company-calendar-actions">
                  <button class="crm-mini-button" type="button" id="companyCalendarPrev" aria-label="이전 달"><i data-lucide="chevron-left"></i></button>
                  <button class="crm-mini-button" type="button" id="companyCalendarToday">오늘</button>
                  <button class="crm-mini-button" type="button" id="companyCalendarNext" aria-label="다음 달"><i data-lucide="chevron-right"></i></button>
                  <button class="crm-mini-button" type="button" id="companyCalendarRefresh">새로고침</button>
                </div>
              </div>
              <div class="company-calendar-legend" aria-label="캘린더 항목 구분">
                <span><i class="calendar-legend-dot project"></i>프로젝트</span>
                <span><i class="calendar-legend-dot task"></i>업무 마감</span>
                <span><i class="calendar-legend-dot leave"></i>연차</span>
                <span><i class="calendar-legend-dot pending"></i>승인대기</span>
              </div>
              <div class="company-calendar-weekdays" aria-hidden="true">
                <div>월</div><div>화</div><div>수</div><div>목</div><div>금</div><div>토</div><div>일</div>
              </div>
              <div class="company-calendar-grid" id="companyCalendarGrid" role="grid" aria-label="회사 일정 월간 캘린더">
                <div class="calendar-empty">캘린더를 불러오는 중입니다.</div>
              </div>
            </article>
            <aside class="dashboard-sales-panel" id="dashboardSalesPanel">
              <article class="company-card">
                <div class="company-card-head"><span>매출 현황</span><span>연동 대기</span></div>
                <div class="company-card-body">
                  <div class="dashboard-sales-grid">
                    <div class="dashboard-sales-metric"><span>오늘 매출</span><strong>-</strong></div>
                    <div class="dashboard-sales-metric"><span>이번 달 매출</span><strong>-</strong></div>
                    <div class="dashboard-sales-metric"><span>주문 건수</span><strong>-</strong></div>
                    <div class="dashboard-sales-metric"><span>평균 객단가</span><strong>-</strong></div>
                  </div>
                  <div class="dashboard-sales-placeholder">발주모아 매출 데이터 연결 대기 중</div>
                </div>
              </article>
            </aside>
          </div>
        </section>

        <section class="company-panel" data-company-panel="rules">
          <div class="company-rule-grid">
            <article class="company-card">
              <div class="company-card-head"><span>근무/휴가</span></div>
              <div class="company-card-body">출퇴근, 연차 신청, 반차/시간 단위 사용 기준을 정리할 영역입니다.</div>
            </article>
            <article class="company-card">
              <div class="company-card-head"><span>비용/정산</span></div>
              <div class="company-card-body">법인카드, 영수증, 물류비, 기타 비용 승인 기준을 정리할 영역입니다.</div>
            </article>
            <article class="company-card">
              <div class="company-card-head"><span>보안/계정</span></div>
              <div class="company-card-body">업무 계정, 비밀번호, 파일 공유, 외부 전달 주의사항을 정리할 영역입니다.</div>
            </article>
            <article class="company-card">
              <div class="company-card-head"><span>업무 처리 기준</span></div>
              <div class="company-card-body">발주, CS, 거래처 응대, 업무 지시 처리 기준을 정리할 영역입니다.</div>
            </article>
          </div>
        </section>

        <section class="company-panel" data-company-panel="staff">
          <div class="company-staff-layout">
            <article class="company-card">
              <div class="company-card-head"><span>직원 조직도</span><button class="crm-mini-button" type="button" id="companyStaffRefresh">새로고침</button></div>
              <div class="company-org" id="companyOrgBody">
                <div class="company-org-empty">직원 조직도를 불러오는 중입니다.</div>
              </div>
            </article>
            <article class="company-card">
              <div class="company-card-head"><span>내 자리</span><button class="crm-mini-button" type="button" data-open="crm" data-crm-nav-tab="mine">전체 보기</button></div>
              <div class="company-card-body">
                <div class="company-mini-grid">
                  <div><span>오늘 마감</span><strong id="companyStaffDueToday">0건</strong></div>
                  <div><span>공지</span><strong id="companyStaffNoticeTitle">등록 전</strong></div>
                </div>
                <div class="company-task-list" id="companyStaffTaskBody">내 업무를 불러오는 중입니다.</div>
                <div class="company-quick-links">
                  <button class="workspace-button" type="button" data-open="crm" data-crm-nav-tab="tasks">업무보드</button>
                  <button class="workspace-button" type="button" data-open="management">통합관리대장</button>
                  <button class="workspace-button" type="button" data-open="ledger">CS 처리대장</button>
                </div>
              </div>
            </article>
          </div>
        </section>

        <section class="company-panel" data-company-panel="chat">
          <div class="company-grid">
            <article class="company-card">
              <div class="company-card-head"><span>사내 메신저</span><button class="crm-mini-button" type="button" id="internalChatRefresh">새로고침</button></div>
              <div class="internal-chat">
                <div class="internal-chat-rooms" id="internalChatRoomList">
                  <button class="internal-chat-room active" type="button" data-chat-room="global"><span>전체방</span><small>공지/공유</small></button>
                  <div class="internal-chat-empty">직원 목록을 불러오는 중입니다.</div>
                </div>
                <div class="internal-chat-main">
                  <div class="internal-chat-head">
                    <div>
                      <div class="internal-chat-title" id="internalChatTitle">전체방</div>
                      <div class="internal-chat-hint" id="internalChatHint">/업무 @직원 업무내용 / 기한 으로 바로 업무 지시 가능</div>
                    </div>
                  </div>
                  <div class="internal-chat-list" id="internalChatList">
                    <div class="internal-chat-empty">메시지를 불러오는 중입니다.</div>
                  </div>
                  <form class="internal-chat-form" id="internalChatForm">
                    <textarea id="internalChatBody" placeholder="/업무 @관리자 샘플 업무 / 오늘 5시 또는 일반 메시지를 입력해줘."></textarea>
                    <button class="crm-mini-button primary" type="submit">보내기</button>
                  </form>
                </div>
              </div>
            </article>
            <article class="company-card">
              <div class="company-card-head"><span>메신저 방향</span></div>
              <div class="company-card-body">
                <p class="crm-help">전체 공유는 전체방, 특정 지시는 직원 DM에서 처리합니다. 어느 방에서든 /업무 @직원 업무내용 / 기한 형식으로 바로 업무를 만들 수 있습니다.</p>
                <div class="internal-chat-side">
                  <div><span>채널</span><strong>전체방 + 직원 DM</strong></div>
                  <div><span>저장</span><strong>앱 DB</strong></div>
                  <div><span>업무화</span><strong>/업무 명령</strong></div>
                </div>
                <div class="company-quick-links">
                  <button class="workspace-button" type="button" data-open="crm" data-crm-nav-tab="tasks">업무보드</button>
                  <button class="workspace-button" type="button" data-view="dashboard" data-company-tab="staff">조직도</button>
                </div>
              </div>
            </article>
          </div>
        </section>
      </section>

      <section class="workspace-view" id="orderWorkspace">
        <div class="workspace-head">
          <div class="workspace-title" id="orderWorkspaceTitle">발주업무</div>
        </div>
        <div class="order-exec-panel">
          <div class="order-exec-summary">
            <div class="order-exec-kicker">발주업무 실행</div>
            <div class="order-exec-title" id="orderWorkspacePanelTitle">작업을 선택해주세요.</div>
            <div class="order-exec-description" id="orderWorkspaceDescription">아래 5가지 작업 중 필요한 항목의 실행 버튼을 누르면 기존 드롭/업로드 실행창이 열립니다.</div>
          </div>
          <div class="order-exec-grid" id="orderWorkspaceCards">
            <article class="order-exec-card" data-order-card="delivery">
              <div class="order-exec-card-head">
                <div>
                  <div class="order-exec-title-line">
                    <span class="order-exec-card-icon" aria-hidden="true" data-icon="📦"></span>
                    <div class="order-exec-title">개별 택배건 정리</div>
                  </div>
                  <div class="order-exec-description">주소일브릿지 엑셀을 업로드해 수령자별 택배건 정리 텍스트를 생성합니다.</div>
                </div>
                <button class="workspace-button" type="button" data-order-execute="delivery">실행</button>
              </div>
              <ul class="order-exec-steps">
                <li>주소일브릿지 엑셀 파일을 선택합니다.</li>
                <li>정렬 기준을 선택합니다.</li>
                <li>생성 버튼을 눌러 정리 텍스트를 확인합니다.</li>
              </ul>
            </article>
            <article class="order-exec-card" data-order-card="invoice">
              <div class="order-exec-card-head">
                <div>
                  <div class="order-exec-title-line">
                    <span class="order-exec-card-icon" aria-hidden="true" data-icon="🔎"></span>
                    <div class="order-exec-title">송장번호 추출</div>
                  </div>
                  <div class="order-exec-description">출고송장 엑셀에서 수하인별 송장번호를 추출해 엑셀로 다운로드합니다.</div>
                </div>
                <button class="workspace-button" type="button" data-order-execute="invoice">실행</button>
              </div>
              <ul class="order-exec-steps">
                <li>출고송장 엑셀 파일을 선택합니다.</li>
                <li>엑셀 생성 버튼을 누릅니다.</li>
                <li>생성된 송장번호 엑셀을 다운로드합니다.</li>
              </ul>
            </article>
            <article class="order-exec-card" data-order-card="lotte">
              <div class="order-exec-card-head">
                <div>
                  <div class="order-exec-title-line">
                    <span class="order-exec-card-icon" aria-hidden="true" data-icon="▦"></span>
                    <div class="order-exec-title">롯데택배 발주서 변환</div>
                  </div>
                  <div class="order-exec-description">주소일브릿지 원본을 롯데택배 발주서 양식으로 변환합니다.</div>
                </div>
                <button class="workspace-button" type="button" data-order-execute="lotte">실행</button>
              </div>
              <ul class="order-exec-steps">
                <li>주소일브릿지 원본 엑셀을 선택합니다.</li>
                <li>엑셀 생성 버튼을 누릅니다.</li>
                <li>변환된 롯데택배 발주서를 다운로드합니다.</li>
              </ul>
            </article>
            <article class="order-exec-card" data-order-card="salesVendor">
              <div class="order-exec-card-head">
                <div>
                  <div class="order-exec-title-line">
                    <span class="order-exec-card-icon" aria-hidden="true" data-icon="₩"></span>
                    <div class="order-exec-title">매입/매출별 테이터 정리(feat. 얼마에요)</div>
                  </div>
                  <div class="order-exec-description">주소일브릿지 원본을 매출처별 시트와 매입처 요약 형식으로 정리합니다.</div>
                </div>
                <button class="workspace-button" type="button" data-order-execute="salesVendor">실행</button>
              </div>
              <ul class="order-exec-steps">
                <li>주소일브릿지 원본 엑셀을 선택합니다.</li>
                <li>엑셀 생성 버튼을 누릅니다.</li>
                <li>매출처별 정리 파일을 다운로드합니다.</li>
              </ul>
            </article>
            <article class="order-exec-card" data-order-card="vehicle">
              <div class="order-exec-card-head">
                <div>
                  <div class="order-exec-title-line">
                    <span class="order-exec-card-icon" aria-hidden="true" data-icon="🚚"></span>
                    <div class="order-exec-title">차량인수증</div>
                  </div>
                  <div class="order-exec-description">공급받는자, 제품, 납품장소, 담당자 정보를 입력해 차량인수증을 생성합니다.</div>
                </div>
                <button class="workspace-button" type="button" data-order-execute="vehicle">실행</button>
              </div>
              <ul class="order-exec-steps">
                <li>공급받는자와 제품 정보를 입력합니다.</li>
                <li>납품장소와 담당자명을 입력합니다.</li>
                <li>인수증 생성 버튼으로 엑셀을 다운로드합니다.</li>
              </ul>
            </article>
          </div>
        </div>
        <div class="order-download-panel" id="orderRecentDownloads">
          <div class="order-download-head">
            <div>
              <div class="order-exec-kicker">추가 다운로드</div>
              <div class="order-download-title">최근 출력 파일</div>
            </div>
            <button class="workspace-button" type="button" id="orderDownloadRefresh">새로고침</button>
          </div>
          <div class="order-download-list" id="orderDownloadList">
            <div class="order-download-empty">최근 출력된 파일이 없습니다.</div>
          </div>
        </div>
      </section>

      <section class="workspace-view" id="fileLibraryWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">업무 파일 자료실</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="sharedFileRefresh">새로고침</button>
          </div>
        </div>
        <div class="shared-file-grid">
          <section class="shared-file-panel" id="sharedFileUploadPanel">
            <div class="order-exec-kicker">파일 올리기</div>
            <div class="order-download-title">업무에 필요한 파일을 보관합니다</div>
            <label class="dropzone" for="sharedFileInput">
              <span class="drop-main" id="sharedFileDropMain">업무 파일을 선택해주세요.</span>
              <span class="drop-sub">엑셀, 문서, PDF, 압축파일 등 필요한 파일을 올릴 수 있습니다.</span>
              <input id="sharedFileInput" name="shared_file" type="file" />
            </label>
            <button class="workspace-button" type="button" id="sharedFileUpload">파일 올리기</button>
            <div class="shared-file-message" id="sharedFileMessage"></div>
          </section>
          <section class="shared-file-panel full">
            <div class="order-download-head">
              <div>
                <div class="order-exec-kicker">다운로드</div>
                <div class="order-download-title">저장된 업무 파일</div>
              </div>
            </div>
            <div class="shared-file-table-wrap">
              <table class="shared-file-table">
                <thead>
                  <tr><th>파일명</th><th>크기</th><th>등록자</th><th>등록일</th><th>작업</th></tr>
                </thead>
                <tbody id="sharedFileBody">
                  <tr><td class="empty" colspan="5">저장된 업무 파일이 없습니다.</td></tr>
                </tbody>
              </table>
            </div>
          </section>
        </div>
      </section>

      <section class="workspace-view" id="importWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">수출입 업무</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="importShipmentWorkspaceOpen">수입제품 출고 진행 입력</button>
            <button class="workspace-button" type="button" id="importShipmentRefresh">새로고침</button>
          </div>
        </div>
        <section class="import-progress-card open" id="importProgressCard">
          <div class="import-progress-head">
            <button class="import-progress-title" type="button" id="importShipmentTreeToggle">
              <i data-lucide="chevron-right"></i>
              <span>수입제품 출고 진행 상황</span>
            </button>
            <div class="import-progress-summary" id="importShipmentSummary">진행 0건</div>
          </div>
          <div class="import-table-wrap">
            <table class="import-table">
              <thead>
                <tr>
                  <th>출항일</th>
                  <th>입항일</th>
                  <th>선적항</th>
                  <th>도착항</th>
                  <th>제품명</th>
                  <th>수량</th>
                  <th>HBL NO.</th>
                  <th>SIZE</th>
                  <th>진행상황</th>
                  <th>프리타임</th>
                  <th>입고예정일</th>
                </tr>
              </thead>
              <tbody id="importShipmentBody">
                <tr><td colspan="11"><div class="import-empty">등록된 수입제품 출고 진행 건이 없습니다.</div></td></tr>
              </tbody>
            </table>
          </div>
        </section>
      </section>
      <section class="workspace-view" id="managementWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">통합관리대장 관리</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="managementSaveAll">해당 내용 저장</button>
            <button class="workspace-button danger" type="button" id="managementDeleteSelected">선택 주문 삭제</button>
            <button class="workspace-button" type="button" data-open-window="management">새창으로 열기</button>
          </div>
        </div>
        <div class="workspace-mount" id="managementWorkspaceMount"></div>
      </section>
      <section class="workspace-view" id="ledgerWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">CS 처리대장</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="ledgerSaveAll">해당 내용 저장</button>
            <button class="workspace-button danger" type="button" id="ledgerDeleteSelected">선택 주문 삭제</button>
            <button class="workspace-button" type="button" data-open-window="ledger">새창으로 열기</button>
          </div>
        </div>
        <div class="workspace-mount" id="ledgerWorkspaceMount"></div>
      </section>
      <section class="workspace-view" id="crmWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">업무 현황</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="crmRefresh">새로고침</button>
            <button class="workspace-button" type="button" id="crmAccountQuick">직원 현황</button>
            <button class="workspace-button" type="button" id="crmTaskQuick">업무 등록</button>
            <button class="workspace-button" type="button" data-open-window="crm">새창으로 열기</button>
          </div>
        </div>
        <div class="crm-tabs" role="tablist" aria-label="CRM 메뉴">
          <button class="crm-tab active" type="button" role="tab" id="crmTabDashboard" aria-selected="true" aria-controls="crmPanelDashboard" tabindex="0" data-crm-tab="dashboard">업무 현황</button>
          <button class="crm-tab" type="button" role="tab" id="crmTabMine" aria-selected="false" aria-controls="crmPanelMine" tabindex="-1" data-crm-tab="mine">내 업무</button>
          <button class="crm-tab" type="button" role="tab" id="crmTabTasks" aria-selected="false" aria-controls="crmPanelTasks" tabindex="-1" data-crm-tab="tasks">업무보드</button>
          <button class="crm-tab" type="button" role="tab" id="crmTabAccounts" aria-selected="false" aria-controls="crmPanelAccounts" tabindex="-1" data-crm-tab="accounts">직원 현황</button>
          <button class="crm-tab" type="button" role="tab" id="crmMessagesTab" aria-selected="false" aria-controls="crmPanelMessages" tabindex="-1" data-crm-tab="messages">연동 로그</button>
        </div>
        <div class="crm-message" id="crmMessage" role="status" aria-live="polite"></div>

        <section class="crm-panel active" role="tabpanel" id="crmPanelDashboard" aria-labelledby="crmTabDashboard" tabindex="0" data-crm-panel="dashboard">
          <div class="crm-stat-grid">
            <article class="crm-stat"><div class="crm-stat-label">활성 직원</div><div class="crm-stat-value" id="crmStatAccounts">0</div></article>
            <article class="crm-stat"><div class="crm-stat-label">진행 업무</div><div class="crm-stat-value" id="crmStatOpenTasks">0</div></article>
            <article class="crm-stat"><div class="crm-stat-label">오늘 마감</div><div class="crm-stat-value" id="crmStatDueToday">0</div></article>
            <article class="crm-stat"><div class="crm-stat-label">지연 업무</div><div class="crm-stat-value" id="crmStatOverdue">0</div></article>
          </div>
          <article class="crm-card crm-project-card">
            <div class="crm-card-head"><span>프로젝트별 진행상황</span><button class="crm-mini-button" type="button" data-crm-go="tasks">업무보드 보기</button></div>
            <div class="crm-card-body">
              <div class="crm-project-tracker" id="crmProjectProgressBody">
                <div class="crm-project-empty">프로젝트 진행상황을 불러오는 중입니다.</div>
              </div>
            </div>
          </article>
          <div class="crm-grid-2">
            <article class="crm-card">
              <div class="crm-card-head"><span>우선 처리 업무</span><button class="crm-mini-button" type="button" data-crm-go="tasks">전체 보기</button></div>
              <div class="crm-table-wrap">
                <table class="crm-table">
                  <thead><tr><th>번호</th><th>업무 구분</th><th>업무</th><th>담당자</th><th>기한</th><th>상태</th></tr></thead>
                  <tbody id="crmPriorityTaskBody"></tbody>
                </table>
              </div>
            </article>
            <article class="crm-card">
              <div class="crm-card-head"><span>최근 메신저 처리</span><button class="crm-mini-button" type="button" data-crm-go="messages">로그 보기</button></div>
              <div class="crm-table-wrap">
                <table class="crm-table">
                  <thead><tr><th>일시</th><th>발신자</th><th>결과</th><th>내용</th></tr></thead>
                  <tbody id="crmRecentMessageBody"></tbody>
                </table>
              </div>
            </article>
          </div>
        </section>

        <section class="crm-panel" role="tabpanel" id="crmPanelMine" aria-labelledby="crmTabMine" tabindex="0" data-crm-panel="mine">
          <div class="crm-task-board-stats" id="crmMineStats"></div>
          <div class="crm-table-wrap">
            <table class="crm-table">
              <thead><tr><th>번호</th><th>업무 구분</th><th>업무</th><th>마감</th><th>상태</th><th>우선순위</th><th>처리</th></tr></thead>
              <tbody id="crmMineTaskBody"></tbody>
            </table>
          </div>
        </section>

        <section class="crm-panel" role="tabpanel" id="crmPanelAccounts" aria-labelledby="crmTabAccounts" tabindex="0" data-crm-panel="accounts">
          <article class="crm-card">
            <div class="crm-card-head"><span>직원별 업무 현황</span><button class="crm-mini-button" type="button" id="crmStaffRefresh">새로고침</button></div>
            <div class="crm-card-body">
              <div class="crm-staff-list" id="crmStaffBody">
                <div class="internal-chat-empty">직원 업무 현황을 불러오는 중입니다.</div>
              </div>
            </div>
          </article>
          <div hidden>
          <form class="crm-card" id="crmAccountForm">
            <div class="crm-card-head"><span>거래처 등록/수정</span><button class="crm-mini-button primary" type="submit">저장</button></div>
            <div class="crm-card-body crm-form-grid">
              <input id="crmAccountId" type="hidden" />
              <input class="crm-input" id="crmAccountName" type="text" placeholder="거래처명" />
              <input class="crm-input" id="crmAccountType" type="text" placeholder="구분 예) 판매사" />
              <input class="crm-input" id="crmAccountContact" type="text" placeholder="담당자/대표" />
              <input class="crm-input" id="crmAccountPhone" type="text" placeholder="연락처" />
              <input class="crm-input" id="crmAccountEmail" type="email" placeholder="이메일" />
              <input class="crm-input wide" id="crmAccountMemo" type="text" placeholder="메모" />
              <button class="crm-mini-button" type="button" id="crmAccountReset">초기화</button>
            </div>
          </form>
          <div class="crm-toolbar">
            <input class="crm-input" id="crmAccountSearch" type="text" placeholder="거래처, 담당자, 메모 검색" />
            <button class="crm-mini-button primary" type="button" id="crmAccountSearchButton">조회</button>
          </div>
          <div class="crm-table-wrap">
            <table class="crm-table">
              <thead><tr><th>거래처</th><th>구분</th><th>담당자</th><th>연락처</th><th>이메일</th><th>진행 업무</th><th>관리</th></tr></thead>
              <tbody id="crmAccountBody"></tbody>
            </table>
          </div>
          </div>
        </section>

        <section class="crm-panel" role="tabpanel" id="crmPanelTasks" aria-labelledby="crmTabTasks" tabindex="0" data-crm-panel="tasks">
          <form class="crm-card crm-task-form collapsed" id="crmTaskForm">
            <div class="crm-card-head">
              <span>업무 등록/수정</span>
              <span class="crm-mini-actions">
                <button class="crm-mini-button" type="button" id="crmTaskFormToggle">입력 열기</button>
                <button class="crm-mini-button primary" type="submit" id="crmTaskSave">저장</button>
              </span>
            </div>
            <div class="crm-card-body crm-form-grid">
              <input id="crmTaskId" type="hidden" />
              <select class="crm-select" id="crmTaskAccount"></select>
              <input class="crm-input wide" id="crmTaskAccountName" type="text" placeholder="선택 업무 구분/거래처명 (선택)" />
              <input class="crm-input wide" id="crmTaskTitle" type="text" placeholder="업무 제목" />
              <select class="crm-select" id="crmTaskAssignee"></select>
              <input class="crm-input" id="crmTaskDue" type="text" placeholder="기한 예) 오늘 5시" />
              <select class="crm-select" id="crmTaskPriority"><option>보통</option><option>높음</option><option>낮음</option></select>
              <select class="crm-select" id="crmTaskStatus"><option>대기</option><option>진행중</option><option>완료</option><option>보류</option></select>
              <textarea class="crm-textarea full" id="crmTaskDescription" placeholder="상세 내용"></textarea>
              <button class="crm-mini-button" type="button" id="crmTaskReset">초기화</button>
            </div>
          </form>
          <div class="crm-view-strip" id="crmTaskPresetList" aria-label="업무 빠른 보기"></div>
          <div class="crm-toolbar crm-task-toolbar" aria-label="업무 필터">
            <label class="sr-only" for="crmTaskViewSelect">저장뷰</label>
            <select class="crm-select" id="crmTaskViewSelect"><option value="">저장뷰 선택</option></select>
            <label class="sr-only" for="crmTaskSearch">업무 검색</label>
            <input class="crm-input" id="crmTaskSearch" type="search" placeholder="업무, 직원, 번호 검색" />
            <button class="crm-mini-button primary" type="button" id="crmTaskSearchButton">조회</button>
            <button class="crm-mini-button" type="button" id="crmTaskAdvancedToggle" aria-expanded="false" aria-controls="crmAdvancedFilters">고급 필터</button>
            <button class="crm-mini-button" type="button" id="crmTaskFilterReset">초기화</button>
            <div class="crm-advanced-filters" id="crmAdvancedFilters" hidden>
            <label class="sr-only" for="crmTaskViewName">저장뷰 이름</label>
            <input class="crm-input" id="crmTaskViewName" type="text" placeholder="저장뷰 이름" />
            <button class="crm-mini-button" type="button" id="crmTaskViewSave">현재 보기 저장</button>
            <button class="crm-mini-button" type="button" id="crmTaskViewDelete">저장뷰 삭제</button>
            <label class="sr-only" for="crmTaskStatusFilter">상태</label>
            <select class="crm-select" id="crmTaskStatusFilter"><option value="">상태 전체</option><option>대기</option><option>진행중</option><option>완료</option><option>보류</option></select>
            <label class="sr-only" for="crmTaskAssigneeFilter">담당자</label>
            <select class="crm-select" id="crmTaskAssigneeFilter"><option value="">담당자 전체</option></select>
            <label class="sr-only" for="crmTaskPriorityFilter">우선순위</label>
            <select class="crm-select" id="crmTaskPriorityFilter"><option value="">우선순위 전체</option><option>높음</option><option>보통</option><option>낮음</option></select>
            <label class="sr-only" for="crmTaskDueFilter">마감</label>
            <select class="crm-select" id="crmTaskDueFilter"><option value="">마감 전체</option><option value="today">오늘 마감</option><option value="overdue">지연</option><option value="upcoming">향후 마감</option><option value="none">기한 없음</option></select>
            <label class="sr-only" for="crmTaskSourceFilter">출처</label>
            <select class="crm-select" id="crmTaskSourceFilter"><option value="">출처 전체</option><option value="app">직접 등록</option><option value="internal_message">사내 메신저</option><option value="messenger">외부 메신저</option></select>
            <label class="sr-only" for="crmTaskSort">정렬</label>
            <select class="crm-select" id="crmTaskSort"><option value="smart">추천순</option><option value="due">마감순</option><option value="updated">최근 수정순</option></select>
            <label class="crm-filter-check"><input type="checkbox" id="crmTaskOpenOnly" checked /> 미완료만</label>
            </div>
          </div>
          <div class="crm-task-board-stats" id="crmTaskBoardStats"></div>
          <div class="crm-task-layout">
            <div class="crm-kanban" id="crmTaskBody"></div>
            <aside class="crm-task-detail" id="crmTaskDetail">
              <div class="crm-task-detail-empty">업무 카드를 선택하면 상세 정보와 처리 버튼이 표시됩니다.</div>
            </aside>
          </div>
        </section>

        <section class="crm-panel" role="tabpanel" id="crmPanelMessages" aria-labelledby="crmMessagesTab" tabindex="0" data-crm-panel="messages">
          <article class="crm-card">
            <div class="crm-card-head"><span>카카오 공식 연동 준비</span></div>
            <div class="crm-card-body crm-webhook-setup">
              <p class="crm-help">카카오 챗봇 스킬 또는 공통 웹훅에서 아래 엔드포인트로 POST 요청을 보내면 CRM 업무에 반영됩니다. 실제 연결 시 공개 HTTPS 도메인과 헤더 토큰이 필요합니다.</p>
              <div class="crm-token"><span>전체 수신 URL</span><code id="crmWebhookUrl">도메인 설정 후 표시됩니다.</code><button class="crm-mini-button" type="button" id="crmWebhookUrlCopy">복사</button></div>
              <div class="crm-token"><span>헤더 이름</span><code id="crmWebhookHeader">X-Workhub-Webhook-Token</code><button class="crm-mini-button" type="button" id="crmWebhookHeaderCopy">복사</button></div>
              <div class="crm-token"><span>현재 토큰</span><code id="crmWebhookToken">권한이 있으면 표시됩니다.</code><button class="crm-mini-button" type="button" id="crmWebhookTokenCopy">복사</button></div>
              <div class="crm-token"><span>토큰 관리</span><div class="crm-token-actions"><button class="crm-mini-button primary" type="button" id="crmWebhookTokenRotate">토큰 재발급</button><span>재발급 즉시 이전 토큰은 사용할 수 없습니다.</span></div></div>
              <div class="crm-sample-grid">
                <div>
                  <div class="crm-sample-title">카카오 스킬 테스트 payload</div>
                  <pre class="crm-code-block" id="crmWebhookSamplePayload"></pre>
                </div>
                <div>
                  <div class="crm-sample-title">curl 테스트</div>
                  <pre class="crm-code-block" id="crmWebhookCurl"></pre>
                </div>
              </div>
            </div>
          </article>
          <form class="crm-card" id="crmMessengerForm">
            <div class="crm-card-head"><span>메신저 사용자 매핑</span><button class="crm-mini-button primary" type="submit">저장</button></div>
            <div class="crm-card-body crm-form-grid">
              <select class="crm-select" id="crmMessengerPlatform"><option value="kakao">kakao</option><option value="generic">generic</option></select>
              <input class="crm-input wide" id="crmMessengerSenderKey" type="text" placeholder="카카오/메신저 사용자 키" />
              <input class="crm-input" id="crmMessengerDisplayName" type="text" placeholder="표시 이름" />
              <select class="crm-select" id="crmMessengerUser"></select>
            </div>
          </form>
          <div class="crm-grid-2">
            <article class="crm-card">
              <div class="crm-card-head"><span>등록 직원 매핑</span></div>
              <div class="crm-table-wrap">
                <table class="crm-table">
                  <thead><tr><th>플랫폼</th><th>사용자 키</th><th>표시 이름</th><th>Workhub 사용자</th></tr></thead>
                  <tbody id="crmMessengerUserBody"></tbody>
                </table>
              </div>
            </article>
            <article class="crm-card">
              <div class="crm-card-head"><span>연동 로그</span></div>
              <div class="crm-table-wrap">
                <table class="crm-table">
                  <thead><tr><th>일시</th><th>플랫폼</th><th>발신자</th><th>결과</th><th>내용</th><th>오류</th><th>관리</th></tr></thead>
                  <tbody id="crmMessageEventBody"></tbody>
                </table>
              </div>
            </article>
          </div>
        </section>
      </section>
      __LEAVE_WORKSPACE__
      __ADMIN_WORKSPACE__
      __BACKUP_WORKSPACE__
      __SYSTEM_WORKSPACE__
    </main>
  </div>

  <div class="notice-popup-backdrop" id="noticePopup">
    <div class="notice-popup" role="dialog" aria-modal="true">
      <div class="notice-popup-head">
        <span>공지사항 입력</span>
        <button class="close" id="noticePopupClose" type="button" aria-label="닫기"><i data-lucide="x"></i></button>
      </div>
      <div class="notice-template">
        <div class="notice-template-grid">
          <input id="noticeDateInput" type="date" />
          <input id="noticeTitleInput" type="text" placeholder="공지 제목" />
          <input id="noticeOwnerInput" type="text" placeholder="담당자" />
        </div>
        <textarea id="noticeBodyInput" placeholder="공지 내용을 입력해주세요. 예) 금일 출고 마감 시간 / 택배 특이사항 / 업체 회신 필요 건"></textarea>
        <div class="notice-template-actions">
          <button class="workspace-button" type="button" id="noticeClearButton">초기화</button>
          <button class="workspace-button" type="button" id="noticeSaveButton">공지 저장</button>
        </div>
        <div class="notice-preview" id="noticePreview">
          <strong>저장된 공지사항이 없습니다.</strong>
          공지사항을 입력하고 저장하면 이곳에서 미리 볼 수 있습니다.
        </div>
      </div>
    </div>
  </div>

  <div class="focus-widget-backdrop" id="focusWidget" aria-hidden="true">
    <section class="focus-widget" role="dialog" aria-modal="true" aria-labelledby="focusWidgetTitle">
      <div class="focus-widget-head">
        <div>
          <div class="focus-widget-kicker" id="focusWidgetKicker">크게 보기</div>
          <div class="focus-widget-title" id="focusWidgetTitle">상세 보기</div>
          <div class="focus-widget-subtitle" id="focusWidgetSubtitle"></div>
        </div>
        <button class="focus-widget-close" id="focusWidgetClose" type="button" aria-label="닫기"><i data-lucide="x"></i></button>
      </div>
      <div class="focus-widget-body" id="focusWidgetBody"></div>
    </section>
  </div>

  <div class="notice-popup-backdrop" id="importShipmentPopup">
    <div class="notice-popup" role="dialog" aria-modal="true">
      <div class="notice-popup-head">
        <span>수입제품 출고 진행 입력</span>
        <button class="close" id="importShipmentClose" type="button" aria-label="닫기"><i data-lucide="x"></i></button>
      </div>
      <div class="notice-template">
        <input id="importShipmentId" type="hidden" />
        <div class="notice-template-grid">
          <input id="importDepartureDate" type="text" placeholder="출항일 예) 6/10" />
          <input id="importArrivalDate" type="text" placeholder="입항일 예) 6/13" />
          <input id="importLoadingPort" type="text" placeholder="선적항" />
        </div>
        <div class="notice-template-grid">
          <input id="importArrivalPort" type="text" placeholder="도착항" />
          <input id="importItem" type="text" placeholder="제품명" />
          <input id="importQuantity" type="text" placeholder="수량" />
        </div>
        <div class="notice-template-grid">
          <input id="importHblNo" type="text" placeholder="HBL NO." />
          <input id="importSize" type="text" placeholder="SIZE" />
          <input id="importProgressStatus" type="text" placeholder="진행상황" />
        </div>
        <div class="notice-template-grid">
          <input id="importFreeTime" type="text" placeholder="프리타임 예) 7일" />
          <input id="importWarehouseDueDate" type="text" placeholder="입고예정일" />
        </div>
        <div class="notice-template-actions">
          <button class="workspace-button" type="button" id="importShipmentReset">초기화</button>
          <button class="workspace-button" type="button" id="importShipmentSave">저장</button>
        </div>
      </div>
    </div>
  </div>

  <div class="workhub-modal-backdrop" id="modal">
    <div class="workhub-modal" role="dialog" aria-modal="true">
      <div class="workhub-modal-head">
        <div class="modal-title" id="modalTitle">파일 업로드</div>
        <button class="close" id="closeModal" aria-label="닫기"><i data-lucide="x"></i></button>
      </div>
      <form id="uploadForm">
        <label class="field-label" id="fileLabel">엑셀 파일 선택</label>
        <label class="dropzone" for="fileInput">
          <span class="drop-main" id="dropMain">파일을 선택하거나 여기에 올려주세요.</span>
          <span class="drop-sub" id="dropSub">xlsx 파일만 처리합니다.</span>
          <input id="fileInput" name="file" type="file" accept=".xlsx,.xlsm" />
        </label>
        <div id="templateUpload" style="display:none; margin-top:16px;">
          <label class="field-label" for="templateInput">양식 파일 선택</label>
          <label class="dropzone" for="templateInput">
            <span class="drop-main" id="templateDropMain">롯데택배 발주서 양식을 선택해주세요.</span>
            <span class="drop-sub">출력 파일은 이 양식의 서식을 따라갑니다.</span>
            <input id="templateInput" name="template" type="file" accept=".xlsx,.xlsm" />
          </label>
        </div>
        <div class="options" id="deliveryOptions">
          <label for="sortMode">정렬</label>
          <select id="sortMode" name="sort">
            <option value="name">상품명순</option>
            <option value="count">건수 많은 순</option>
            <option value="first">엑셀 순서</option>
          </select>
        </div>
        <div class="message-placeholder" id="messagePlaceholder">
          <strong id="messagePlaceholderTitle">메시지 창</strong>
          <span id="messagePlaceholderBody">해당 업무 메시지 UI는 다음 단계에서 생성합니다.</span>
        </div>
        <div class="vehicle-fields" id="vehicleFields">
          <div class="text-field">
            <label class="field-label" for="receiptTypeSelect">차량인수증 타입</label>
            <select id="receiptTypeSelect" name="receipt_type">
              <option value="일반">일반</option>
              <option value="모드니 전용">모드니 전용</option>
            </select>
          </div>
          <div class="text-field">
            <label class="field-label" for="supplierInput">공급받는자</label>
            <input id="supplierInput" name="supplier" type="text" placeholder="예) 모드니" />
          </div>
          <div class="text-field">
            <label class="field-label" for="receiptDateInput">일자</label>
            <input id="receiptDateInput" name="receipt_date" type="date" />
          </div>
          <div class="text-field">
            <label class="field-label" for="freightPaymentSelect">운임비용</label>
            <select id="freightPaymentSelect" name="freight_payment">
              <option value="선불">선불</option>
              <option value="후불">후불</option>
            </select>
          </div>
          <label class="field-label" style="margin-top:16px;">제품명 / 수량 / 입수량</label>
          <div class="product-table" id="productTable"></div>
          <button class="add-row" id="addProductRow" type="button">제품 줄 추가</button>
          <div class="text-field">
            <label class="field-label" for="requestNoteInput">요청사항</label>
            <textarea id="requestNoteInput" name="request_note" placeholder="예) 파손주의 / 현장 사진 문자 요청"></textarea>
          </div>
          <div class="text-field">
            <label class="field-label" for="deliveryPlaceInput">납품장소</label>
            <input id="deliveryPlaceInput" name="delivery_place" type="text" placeholder="예) 모드니 물류센터" />
          </div>
          <div class="text-field">
            <label class="field-label" for="managerInput">담당자명</label>
            <input id="managerInput" name="manager" type="text" placeholder="예) 홍길동 / 010-0000-0000" />
          </div>
        </div>
        <div class="cs-fields" id="csFields">
          <div class="ledger-cs-popup-head">
            <strong class="ledger-cs-popup-title">CS 추가</strong>
            <button class="ledger-cs-popup-close" id="ledgerCsPopupClose" type="button">닫기</button>
          </div>
          <div class="text-field">
            <label class="field-label" for="vendorContactSelect">업체 선택</label>
            <select id="vendorContactSelect">
              <option value="">업체를 선택해주세요</option>
            </select>
          </div>
          <div class="text-field">
            <label class="field-label" for="vendorTypeSelect">업체 구분</label>
            <select id="vendorTypeSelect">
              <option value="purchase">매입처</option>
              <option value="sales">매출처</option>
            </select>
          </div>
          <div class="text-field">
            <label class="field-label" for="recipientEmailInput">받는 업체 메일</label>
            <input id="recipientEmailInput" name="recipient_email" type="email" placeholder="예) vendor@example.com" />
          </div>
          <div class="text-field">
            <label class="field-label" for="vendorNameInput">업체명</label>
            <input id="vendorNameInput" name="vendor_name" type="text" placeholder="예) 키친쿡" />
          </div>
          <button class="add-row" id="saveVendorContact" type="button">업체 메일 저장/업데이트</button>
          <div class="text-field">
            <label class="field-label" for="csOriginInput">원출고일 및 원송장번호</label>
            <input id="csOriginInput" name="cs_origin" type="text" placeholder="예) 2026-06-13 / 1234567890" />
          </div>
          <div class="text-field">
            <label class="field-label" for="csProductInput">상품명</label>
            <input id="csProductInput" name="cs_product" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="csReceiverInput">수령인</label>
            <input id="csReceiverInput" name="cs_receiver" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="csPhoneInput">수령인 연락처</label>
            <input id="csPhoneInput" name="cs_phone" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="csAddressInput">수령인 주소</label>
            <input id="csAddressInput" name="cs_address" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="csTypeInput">CS타입</label>
            <select id="csTypeInput" name="cs_type">
              <option value="">선택</option>
              <option value="변심반품">변심반품</option>
              <option value="불량반품">불량반품</option>
              <option value="불량교환">불량교환</option>
              <option value="불량재출고(미회수)">불량재출고(미회수)</option>
              <option value="오출고(오배송)">오출고(오배송)</option>
            </select>
          </div>
          <div class="text-field">
            <label class="field-label" for="csContentInput">CS내용</label>
            <textarea id="csContentInput" name="cs_content" placeholder="예) 파손 / 오배송 / 누락 / 반품 접수 요청"></textarea>
          </div>
          <div class="text-field">
            <label class="field-label" for="csAttachmentInput">첨부파일(이미지/영상)</label>
            <input id="csAttachmentInput" name="cs_attachments" type="file" accept="image/*,video/*" multiple />
            <div class="hint-line" id="csAttachmentSummary">첨부파일 없음</div>
          </div>
          <div class="text-field">
            <label class="field-label" for="csSubjectInput">메일 제목</label>
            <input id="csSubjectInput" name="subject" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="csBodyInput">요청 내용</label>
            <textarea id="csBodyInput" name="body"></textarea>
          </div>
          <div class="cs-toolbar">
            <button class="cs-save-button" id="saveCsCase" type="button">CS건 DB 저장</button>
          </div>
          <div class="cs-case-list">
            <div class="cs-case-head">최근 저장 CS</div>
            <div id="csCaseList"></div>
          </div>
        </div>
        <div class="stock-notice-fields" id="stockNoticeFields">
          <div class="text-field">
            <label class="field-label" for="stockVendorPickerButton">업체 선택</label>
            <button class="vendor-picker-button" id="stockVendorPickerButton" type="button">업체를 선택해주세요</button>
            <div class="vendor-picker-selected" id="stockSelectedVendorLabel">선택된 업체 없음</div>
            <div class="vendor-picker-tree" id="stockVendorTree" hidden></div>
            <input id="stockVendorTypeSelect" type="hidden" value="purchase" />
            <input id="stockRecipientEmailInput" type="hidden" />
            <input id="stockVendorNameInput" type="hidden" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockNoticeDateInput">기준일자</label>
            <input id="stockNoticeDateInput" type="date" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockInboundProductInput">입고 품명(모델명)</label>
            <input id="stockInboundProductInput" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockInboundScheduleInput">입고 일정</label>
            <input id="stockInboundScheduleInput" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockOutboundAvailableInput">출고 가능 일정</label>
            <input id="stockOutboundAvailableInput" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockInboundNoteInput">입고 특이사항</label>
            <textarea id="stockInboundNoteInput"></textarea>
          </div>
          <div class="text-field">
            <label class="field-label" for="stockSoldoutProductInput">품절/단종 품명(모델명)</label>
            <input id="stockSoldoutProductInput" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockOutboundBlockedInput">출고 불가 일정</label>
            <input id="stockOutboundBlockedInput" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockRestockScheduleInput">재입고 일정</label>
            <input id="stockRestockScheduleInput" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockSoldoutNoteInput">품절 특이사항</label>
            <textarea id="stockSoldoutNoteInput"></textarea>
          </div>
          <div class="text-field">
            <label class="field-label" for="stockSubjectInput">메일 제목</label>
            <input id="stockSubjectInput" type="text" />
          </div>
          <div class="text-field">
            <label class="field-label" for="stockBodyInput">공지 내용</label>
            <textarea id="stockBodyInput"></textarea>
          </div>
        </div>
        <div class="ledger-fields" id="ledgerFields">
          <div class="ledger-toolbar">
            <input id="ledgerSearchInput" type="text" placeholder="업체명, 수령인, 상품명, 원송장, CS내용 검색" />
            <select id="ledgerStatusFilter">
              <option value="">상태 전체</option>
              <option value="회수지시">회수지시</option>
              <option value="회수 완료">회수 완료</option>
              <option value="재발송 완료">재발송 완료</option>
              <option value="전체 처리완료">전체 처리완료</option>
            </select>
            <button class="btn blue" id="ledgerRefresh" type="button">조회</button>
            <select id="ledgerPageSize">
              <option value="100">100개씩 보기</option>
              <option value="500">500개씩 보기</option>
              <option value="1000">1,000개씩 보기</option>
              <option value="2000">2,000개씩 보기</option>
              <option value="5000">5,000개씩 보기</option>
            </select>
            <div class="download-menu-wrap">
              <button class="btn blue" id="ledgerDownloadMenuButton" type="button">다운로드 선택</button>
              <div class="download-menu" id="ledgerDownloadMenu">
                <button type="button" data-ledger-download="all">전체 다운로드</button>
                <button type="button" data-ledger-download="selected">선택 다운로드</button>
              </div>
            </div>
            <button class="btn blue" type="button" data-ledger-import-mode="daily">일일 추가 업로드</button>
            <button class="btn danger" type="button" data-ledger-import-mode="replace">전체 데이터 교체 업로드</button>
            <button class="btn primary" id="ledgerAddCs" type="button">CS 추가</button>
          </div>
          <div class="cell-edit-bar" id="ledgerCellEditBar">
            <div class="cell-edit-label" id="ledgerCellEditLabel">셀 선택</div>
            <div id="ledgerCellEditMount"></div>
            <button class="cell-edit-button apply" id="ledgerCellApply" type="button">적용</button>
            <button class="cell-edit-button" id="ledgerCellCancel" type="button">취소</button>
          </div>
          <input class="hidden-file-input" id="ledgerImportInput" name="ledger_import" type="file" accept=".xlsx,.xlsm" />
          <div class="ledger-wrap">
            <table class="ledger-table">
              <thead>
                <tr>
                  <th>선택</th>
                  <th class="has-filter"><span class="ledger-th-title">날짜</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="occurred_at" data-label="날짜">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">매출거래처</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="sales_vendor" data-label="매출거래처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">매입거래처</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="purchase_vendor" data-label="매입거래처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">처리진행상태</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="status" data-label="처리진행상태">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">완료일</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="completed_at" data-label="완료일">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">처리내용</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="cs_type" data-label="처리내용">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">C/S 내용</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="cs_content" data-label="C/S 내용">▼</button></th>
                  <th class="invoice-head has-filter"><span class="ledger-th-title">재발송운송장번호</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="reship_invoice" data-label="재발송운송장번호">▼</button></th>
                  <th class="invoice-head has-filter"><span class="ledger-th-title">회수운송장번호</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="return_invoice" data-label="회수운송장번호">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">주문일자</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="order_date" data-label="주문일자">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">출고일</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="ship_date" data-label="출고일">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">주문자</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="orderer_name" data-label="주문자">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">연락처</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="orderer_phone" data-label="주문자 연락처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">수령자</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="receiver_name" data-label="수령자">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">연락처</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="receiver_phone" data-label="수령자 연락처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">제품명</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="product_name" data-label="제품명">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">수량</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="quantity" data-label="수량">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">상세주소</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="receiver_address" data-label="상세주소">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">택배사</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="courier" data-label="택배사">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">송장번호</span><button class="ledger-filter-trigger" type="button" data-ledger-filter-button="original_invoice" data-label="송장번호">▼</button></th>
                </tr>
              </thead>
              <tbody id="ledgerBody"></tbody>
            </table>
          </div>
        </div>
        <div class="management-fields" id="managementFields">
          <div class="ledger-toolbar">
            <input id="managementSearchInput" type="text" placeholder="거래처, 수령자, 상품명, 주소, 송장번호 검색" />
            <select id="managementYearFilter" hidden>
              <option value="">년도 선택</option>
            </select>
            <select id="managementMonthFilter">
              <option value="">전체 선택</option>
            </select>
            <button class="btn blue" id="managementRefresh" type="button">조회</button>
            <select id="managementPageSize">
              <option value="100">100개씩 보기</option>
              <option value="500" selected>500개씩 보기</option>
              <option value="1000">1,000개씩 보기</option>
              <option value="2000">2,000개씩 보기</option>
              <option value="5000">5,000개씩 보기</option>
            </select>
            <div class="download-menu-wrap">
              <button class="btn blue" id="managementDownloadMenuButton" type="button">다운로드 선택</button>
              <div class="download-menu" id="managementDownloadMenu">
                <button type="button" data-management-download="all">전체 다운로드</button>
                <button type="button" data-management-download="year">년별 다운로드</button>
                <button type="button" data-management-download="month">월별 다운로드</button>
                <button type="button" data-management-download="selected">선택 다운로드</button>
              </div>
            </div>
            <button class="btn blue" type="button" data-management-import-mode="daily">일일 추가 업로드</button>
            <button class="btn danger" type="button" data-management-import-mode="replace">전체 데이터 교체 업로드</button>
          </div>
          <div class="cell-edit-bar" id="managementCellEditBar">
            <div class="cell-edit-label" id="managementCellEditLabel">셀 선택</div>
            <div id="managementCellEditMount"></div>
            <button class="cell-edit-button apply" id="managementCellApply" type="button">적용</button>
            <button class="cell-edit-button" id="managementCellCancel" type="button">취소</button>
          </div>
          <input class="hidden-file-input" id="managementImportInput" name="management_import" type="file" accept=".xlsx,.xlsm" />
          <div class="ledger-wrap management-wrap">
            <table class="ledger-table">
              <thead>
                <tr>
                  <th><input class="ledger-check" id="managementSelectAll" type="checkbox" title="전체 선택" /></th>
                  <th class="has-filter"><span class="ledger-th-title">주문일자</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="order_date" data-label="주문일자">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">출고일</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="ship_date" data-label="출고일">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">매입거래처</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="purchase_vendor" data-label="매입거래처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">매출거래처</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="sales_vendor" data-label="매출거래처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">거래구분</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="transaction_type" data-label="거래구분">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">장부입력확인</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="ledger_checked" data-label="장부입력확인">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">주문자</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="orderer_name" data-label="주문자">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">발신자연락처</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="sender_phone" data-label="발신자연락처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">수령자</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="receiver_name" data-label="수령자">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">수령자연락처</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="receiver_phone" data-label="수령자연락처">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">제품명</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="product_name" data-label="제품명">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">수량</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="quantity" data-label="수량">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">상세주소</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="receiver_address" data-label="상세주소">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">택배사</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="courier" data-label="택배사">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">운송장번호</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="invoice_number" data-label="운송장번호">▼</button></th>
                  <th class="has-filter"><span class="ledger-th-title">특이사항</span><button class="ledger-filter-trigger" type="button" data-management-filter-button="memo" data-label="특이사항">▼</button></th>
                  <th>CS접수</th>
                </tr>
              </thead>
              <tbody id="managementBody"></tbody>
            </table>
          </div>
          <div class="management-month-tabs" id="managementMonthTabs"></div>
        </div>
        <div class="ledger-filter-popover" id="ledgerFilterPopover">
          <div class="ledger-filter-title" id="ledgerFilterTitle">필터</div>
          <input class="ledger-filter-search" id="ledgerFilterSearch" type="text" placeholder="검색어 입력" />
          <div class="ledger-filter-option-list" id="ledgerFilterOptions"></div>
          <div class="ledger-filter-actions">
            <button type="button" id="ledgerFilterClear">전체</button>
            <button class="apply" type="button" id="ledgerFilterApply">적용</button>
          </div>
        </div>
        <div class="notice" id="notice"></div>
        <div class="modal-actions">
          <button class="btn" type="button" id="cancel">취소</button>
          <button class="btn primary" id="submitButton" type="submit">생성</button>
        </div>
      </form>
      <div class="result" id="result">
        <textarea id="resultText" readonly></textarea>
        <div class="result-actions">
          <button class="btn" type="button" id="copyResult">복사</button>
          <button class="btn blue" type="button" id="downloadText">텍스트 저장</button>
        </div>
      </div>
    </div>
  </div>

  <div class="safe-number-dialog-backdrop" id="safeNumberPackageDialog" aria-hidden="true">
    <div class="safe-number-dialog" role="dialog" aria-modal="true" aria-labelledby="safeNumberPackageTitle">
      <h2 class="safe-number-dialog-title" id="safeNumberPackageTitle">안심번호 합포 후보 확인</h2>
      <p class="safe-number-dialog-description">
        동일한 수령자명과 주소인데 연락처만 다른 건입니다. 합포장으로 출력할지 확인해주세요.
      </p>
      <pre class="safe-number-dialog-preview" id="safeNumberPackagePreview"></pre>
      <div class="safe-number-dialog-actions">
        <button class="btn" type="button" id="safeNumberPackageReject">개별건으로 출력</button>
        <button class="btn blue" type="button" id="safeNumberPackageApprove">합포장으로 출력</button>
      </div>
    </div>
  </div>

  <div class="safe-number-dialog-backdrop" id="importWarningDialog" aria-hidden="true">
    <div class="safe-number-dialog" role="dialog" aria-modal="true" aria-labelledby="importWarningTitle">
      <h2 class="safe-number-dialog-title" id="importWarningTitle">업로드 확인</h2>
      <p class="safe-number-dialog-description" id="importWarningDescription"></p>
      <pre class="safe-number-dialog-preview" id="importWarningPreview"></pre>
      <div class="safe-number-dialog-actions">
        <button class="btn" id="importWarningCancel" type="button">취소</button>
        <button class="btn primary" id="importWarningProceed" type="button">진행</button>
      </div>
    </div>
  </div>

  <div class="safe-number-dialog-backdrop" id="importCorrectionDialog" aria-hidden="true">
    <div class="safe-number-dialog import-correction-dialog" role="dialog" aria-modal="true" aria-labelledby="importCorrectionTitle">
      <h2 class="safe-number-dialog-title" id="importCorrectionTitle">업로드 전 형식 수정</h2>
      <p class="safe-number-dialog-description" id="importCorrectionDescription"></p>
      <div class="import-correction-list" id="importCorrectionList"></div>
      <div class="safe-number-dialog-actions">
        <button class="btn" id="importCorrectionCancel" type="button">취소</button>
        <button class="btn primary" id="importCorrectionApply" type="button">수정 후 적용</button>
      </div>
    </div>
  </div>

  <script type="module">
    import { createIcons, BriefcaseBusiness, Home, MessageCircle, Info, ChevronDown, ChevronRight, PlusSquare, RefreshCw, Ellipsis, Headphones, Package, ClipboardCheck, CircleDollarSign, FileText, FileSpreadsheet, ClipboardList, BarChart3, CopyCheck, Bell, Download, Truck, Mail, Upload, Database, CalendarDays, X, Settings } from "/lucide/dist/esm/lucide.js";
    createIcons({ icons: { BriefcaseBusiness, Home, MessageCircle, Info, ChevronDown, ChevronRight, PlusSquare, RefreshCw, Ellipsis, Headphones, Package, ClipboardCheck, CircleDollarSign, FileText, FileSpreadsheet, ClipboardList, BarChart3, CopyCheck, Bell, Download, Truck, Mail, Upload, Database, CalendarDays, X, Settings, "package": Package, "file-text": FileText, "file-spreadsheet": FileSpreadsheet, "truck": Truck } });
    function applyDaisyUiClasses() {
      document.querySelectorAll(".workspace-button, .crm-mini-button, .action-button, .logout-button, .top-button").forEach((element) => {
        element.classList.add("btn", "btn-sm");
      });
      document.querySelectorAll(".icon-button").forEach((element) => {
        element.classList.add("btn", "btn-square", "btn-sm", "btn-ghost");
      });
      document.querySelectorAll(".crm-card, .company-card, .notice-board, .import-progress-card, .card").forEach((element) => {
        element.classList.add("card");
      });
      document.querySelectorAll(".crm-input, .notice-template input, .text-field input, .ledger-toolbar input, .ledger-edit, .management-edit").forEach((element) => {
        element.classList.add("input", "input-sm");
      });
      document.querySelectorAll(".crm-select, .notice-template select, .text-field select, .ledger-toolbar select, .ledger-status-select").forEach((element) => {
        element.classList.add("select", "select-sm");
      });
      document.querySelectorAll(".crm-textarea, .notice-template textarea, .text-field textarea").forEach((element) => {
        element.classList.add("textarea", "textarea-sm");
      });
      document.querySelectorAll(".crm-table, .import-table, .ledger-table").forEach((element) => {
        element.classList.add("table", "table-sm");
      });
      document.querySelectorAll(".crm-status, .crm-priority").forEach((element) => {
        element.classList.add("badge", "badge-sm");
      });
      document.querySelectorAll(".crm-tabs, .company-tabs").forEach((element) => {
        element.classList.add("tabs", "tabs-border");
      });
      document.querySelectorAll(".crm-tab, .company-tab").forEach((element) => {
        element.classList.add("tab");
      });
    }
    applyDaisyUiClasses();
    if (new URLSearchParams(window.location.search).get("standalone") === "1") {
      document.body.classList.add("standalone");
    }
    const currentUser = __CURRENT_USER__;
    const currentUserPermissions = new Set(__USER_PERMISSIONS__);
    const permissionLabels = __PERMISSION_LABELS__;
    const sidebar = document.querySelector(".sidebar");
    const sidebarSearchInput = document.querySelector("#sidebarSearchInput");
    let sidebarSearchUserTyped = false;
    const sidebarSearchAutofillValues = new Set([
      String(currentUser.username || "").trim().toLowerCase(),
      String(currentUser.display_name || "").trim().toLowerCase(),
    ].filter(Boolean));

    function can(permission) {
      return currentUserPermissions.has(permission);
    }

    function permissionLabel(permission) {
      return permissionLabels[permission] || permission;
    }

    function setHidden(element, hidden) {
      if (element) element.classList.toggle("permission-hidden", Boolean(hidden));
    }

    function applyStaticPermissions() {
      setHidden(document.querySelector("#noticeInputOpen"), !can("notice_manage"));
      setHidden(managementDeleteSelected, !can("ledger_delete"));
      setHidden(ledgerDeleteSelected, !can("ledger_delete"));
      setHidden(managementSaveAll, !can("ledger_edit"));
      setHidden(ledgerSaveAll, !can("ledger_edit"));
      setHidden(ledgerAddCs, !can("ledger_edit"));
      setHidden(saveCsCaseButton, !can("ledger_edit"));
      setHidden(ledgerDownloadMenuButton, !can("excel_download"));
      setHidden(managementDownloadMenuButton, !can("excel_download"));
      document.querySelectorAll('[data-ledger-import-mode="daily"], [data-management-import-mode="daily"]').forEach((button) => {
        setHidden(button, !can("excel_upload"));
      });
      document.querySelectorAll('[data-ledger-import-mode="replace"], [data-management-import-mode="replace"]').forEach((button) => {
        setHidden(button, currentUser.role !== "admin");
      });
      setHidden(document.querySelector("label[for='vendorContactsFileInput']"), !can("excel_upload"));
      setHidden(saveVendorContactButton, !can("mail_send"));
      setHidden(document.querySelector("#distributionMailNavGroup"), !can("mail_send"));
      document.querySelectorAll("[data-mail-popup]").forEach((button) => setHidden(button, !can("mail_send")));
      document.querySelectorAll('[data-open="cs"]').forEach((button) => setHidden(button, !can("mail_send")));
      setHidden(sharedFileUploadPanel, currentUser.role !== "admin");
      setHidden(importShipmentInputOpen, !can("import_shipment_manage"));
      setHidden(importShipmentWorkspaceOpen, !can("import_shipment_manage"));
      setHidden(dashboardImportScheduleOpen, !can("import_shipment_manage"));
      document.querySelectorAll('[data-open="crm"]').forEach((button) => setHidden(button, !can("crm_view")));
      setHidden(document.querySelector('.nav-subitem[data-crm-nav-tab="messages"]'), !can("crm_message_manage"));
      setHidden(crmAccountQuick, !can("crm_view"));
      setHidden(crmTaskQuick, !can("crm_manage"));
      setHidden(crmAccountForm, !can("crm_manage"));
      setHidden(crmTaskForm, !can("crm_manage"));
      setHidden(crmMessagesTab, !can("crm_message_manage"));
      setHidden(crmMessengerForm, !can("crm_message_manage"));
      if (!can("notice_manage")) {
        setHidden(noticeSaveButton, true);
        setHidden(noticeClearButton, true);
      }
    }

    const modal = document.querySelector("#modal");
    const modalTitle = document.querySelector("#modalTitle");
    const uploadForm = document.querySelector("#uploadForm");
    const fileLabel = document.querySelector("#fileLabel");
    const fileInput = document.querySelector("#fileInput");
    const templateInput = document.querySelector("#templateInput");
    const dropMain = document.querySelector("#dropMain");
    const dropSub = document.querySelector("#dropSub");
    const templateUpload = document.querySelector("#templateUpload");
    const templateDropMain = document.querySelector("#templateDropMain");
    const deliveryOptions = document.querySelector("#deliveryOptions");
    const messagePlaceholder = document.querySelector("#messagePlaceholder");
    const messagePlaceholderTitle = document.querySelector("#messagePlaceholderTitle");
    const messagePlaceholderBody = document.querySelector("#messagePlaceholderBody");
    const vehicleFields = document.querySelector("#vehicleFields");
    const csFields = document.querySelector("#csFields");
    const ledgerFields = document.querySelector("#ledgerFields");
    const managementFields = document.querySelector("#managementFields");
    const ledgerCellEditBar = document.querySelector("#ledgerCellEditBar");
    const ledgerCellEditLabel = document.querySelector("#ledgerCellEditLabel");
    const ledgerCellEditMount = document.querySelector("#ledgerCellEditMount");
    const ledgerCellApply = document.querySelector("#ledgerCellApply");
    const ledgerCellCancel = document.querySelector("#ledgerCellCancel");
    const managementCellEditBar = document.querySelector("#managementCellEditBar");
    const managementCellEditLabel = document.querySelector("#managementCellEditLabel");
    const managementCellEditMount = document.querySelector("#managementCellEditMount");
    const managementCellApply = document.querySelector("#managementCellApply");
    const managementCellCancel = document.querySelector("#managementCellCancel");
    const productTable = document.querySelector("#productTable");
    const receiptTypeSelect = document.querySelector("#receiptTypeSelect");
    const supplierInput = document.querySelector("#supplierInput");
    const receiptDateInput = document.querySelector("#receiptDateInput");
    const freightPaymentSelect = document.querySelector("#freightPaymentSelect");
    const requestNoteInput = document.querySelector("#requestNoteInput");
    const deliveryPlaceInput = document.querySelector("#deliveryPlaceInput");
    const managerInput = document.querySelector("#managerInput");
    const vendorContactSelect = document.querySelector("#vendorContactSelect");
    const vendorContactsFileInput = document.querySelector("#vendorContactsFileInput");
    const vendorContactsDropMain = document.querySelector("#vendorContactsDropMain");
    const salesReportFileInput = document.querySelector("#salesReportFileInput");
    const salesReportUploadMessage = document.querySelector("#salesReportUploadMessage");
    const salesReportRecentList = document.querySelector("#salesReportRecentList");
    const salesReportKpiGrid = document.querySelector("#salesReportKpiGrid");
    const salesReportDailyBody = document.querySelector("#salesReportDailyBody");
    const salesReportSellerBody = document.querySelector("#salesReportSellerBody");
    const salesReportProductBody = document.querySelector("#salesReportProductBody");
    const salesReportReviewBody = document.querySelector("#salesReportReviewBody");
    const vendorTypeSelect = document.querySelector("#vendorTypeSelect");
    const recipientEmailInput = document.querySelector("#recipientEmailInput");
    const vendorNameInput = document.querySelector("#vendorNameInput");
    const saveVendorContactButton = document.querySelector("#saveVendorContact");
    const csOriginInput = document.querySelector("#csOriginInput");
    const csProductInput = document.querySelector("#csProductInput");
    const csReceiverInput = document.querySelector("#csReceiverInput");
    const csPhoneInput = document.querySelector("#csPhoneInput");
    const csAddressInput = document.querySelector("#csAddressInput");
    const csTypeInput = document.querySelector("#csTypeInput");
    const csContentInput = document.querySelector("#csContentInput");
    const csAttachmentInput = document.querySelector("#csAttachmentInput");
    const csAttachmentSummary = document.querySelector("#csAttachmentSummary");
    const csSubjectInput = document.querySelector("#csSubjectInput");
    const csBodyInput = document.querySelector("#csBodyInput");
    const saveCsCaseButton = document.querySelector("#saveCsCase");
    const csCaseList = document.querySelector("#csCaseList");
    const stockNoticeFields = document.querySelector("#stockNoticeFields");
    const stockVendorPickerButton = document.querySelector("#stockVendorPickerButton");
    const stockVendorTree = document.querySelector("#stockVendorTree");
    const stockSelectedVendorLabel = document.querySelector("#stockSelectedVendorLabel");
    const stockVendorTypeSelect = document.querySelector("#stockVendorTypeSelect");
    const stockRecipientEmailInput = document.querySelector("#stockRecipientEmailInput");
    const stockVendorNameInput = document.querySelector("#stockVendorNameInput");
    const stockNoticeDateInput = document.querySelector("#stockNoticeDateInput");
    const stockInboundProductInput = document.querySelector("#stockInboundProductInput");
    const stockInboundScheduleInput = document.querySelector("#stockInboundScheduleInput");
    const stockOutboundAvailableInput = document.querySelector("#stockOutboundAvailableInput");
    const stockInboundNoteInput = document.querySelector("#stockInboundNoteInput");
    const stockSoldoutProductInput = document.querySelector("#stockSoldoutProductInput");
    const stockOutboundBlockedInput = document.querySelector("#stockOutboundBlockedInput");
    const stockRestockScheduleInput = document.querySelector("#stockRestockScheduleInput");
    const stockSoldoutNoteInput = document.querySelector("#stockSoldoutNoteInput");
    const stockSubjectInput = document.querySelector("#stockSubjectInput");
    const stockBodyInput = document.querySelector("#stockBodyInput");
    const ledgerCsPopupClose = document.querySelector("#ledgerCsPopupClose");
    const ledgerSearchInput = document.querySelector("#ledgerSearchInput");
    const ledgerStatusFilter = document.querySelector("#ledgerStatusFilter");
    const ledgerYearFilter = document.querySelector("#ledgerYearFilter");
    const ledgerMonthFilter = document.querySelector("#ledgerMonthFilter");
    const ledgerRefresh = document.querySelector("#ledgerRefresh");
    const ledgerPageSize = document.querySelector("#ledgerPageSize");
    const ledgerDownloadMenuButton = document.querySelector("#ledgerDownloadMenuButton");
    const ledgerDownloadMenu = document.querySelector("#ledgerDownloadMenu");
    const ledgerAddCs = document.querySelector("#ledgerAddCs");
    const ledgerBody = document.querySelector("#ledgerBody");
    const ledgerImportInput = document.querySelector("#ledgerImportInput");
    const ledgerImportDropMain = null;
    const ledgerImportOpen = document.querySelector("#ledgerImportOpen");
    const managementSearchInput = document.querySelector("#managementSearchInput");
    const managementYearFilter = document.querySelector("#managementYearFilter");
    const managementMonthFilter = document.querySelector("#managementMonthFilter");
    const managementRefresh = document.querySelector("#managementRefresh");
    const managementPageSize = document.querySelector("#managementPageSize");
    const managementDownloadMenuButton = document.querySelector("#managementDownloadMenuButton");
    const managementDownloadMenu = document.querySelector("#managementDownloadMenu");
    const managementImportInput = document.querySelector("#managementImportInput");
    const managementImportDropMain = null;
    const managementImportOpen = document.querySelector("#managementImportOpen");
    const managementBody = document.querySelector("#managementBody");
    const managementMonthTabs = document.querySelector("#managementMonthTabs");
    const managementSelectAll = document.querySelector("#managementSelectAll");
    const managementSaveAll = document.querySelector("#managementSaveAll");
    const ledgerSaveAll = document.querySelector("#ledgerSaveAll");
    const managementDeleteSelected = document.querySelector("#managementDeleteSelected");
    const ledgerDeleteSelected = document.querySelector("#ledgerDeleteSelected");
    const ledgerFilterButtons = Array.from(document.querySelectorAll("[data-ledger-filter-button]"));
    const managementFilterButtons = Array.from(document.querySelectorAll("[data-management-filter-button]"));
    const ledgerFilterPopover = document.querySelector("#ledgerFilterPopover");
    const ledgerFilterTitle = document.querySelector("#ledgerFilterTitle");
    const ledgerFilterSearch = document.querySelector("#ledgerFilterSearch");
    const ledgerFilterOptions = document.querySelector("#ledgerFilterOptions");
    const ledgerFilterClear = document.querySelector("#ledgerFilterClear");
    const ledgerFilterApply = document.querySelector("#ledgerFilterApply");
    const result = document.querySelector("#result");
    const resultText = document.querySelector("#resultText");
    const notice = document.querySelector("#notice");
    const submitButton = document.querySelector("#submitButton");
    const safeNumberPackageDialog = document.querySelector("#safeNumberPackageDialog");
    const safeNumberPackagePreview = document.querySelector("#safeNumberPackagePreview");
    const safeNumberPackageApprove = document.querySelector("#safeNumberPackageApprove");
    const safeNumberPackageReject = document.querySelector("#safeNumberPackageReject");
    const importWarningDialog = document.querySelector("#importWarningDialog");
    const importWarningTitle = document.querySelector("#importWarningTitle");
    const importWarningDescription = document.querySelector("#importWarningDescription");
    const importWarningPreview = document.querySelector("#importWarningPreview");
    const importWarningCancel = document.querySelector("#importWarningCancel");
    const importWarningProceed = document.querySelector("#importWarningProceed");
    const importCorrectionDialog = document.querySelector("#importCorrectionDialog");
    const importCorrectionTitle = document.querySelector("#importCorrectionTitle");
    const importCorrectionDescription = document.querySelector("#importCorrectionDescription");
    const importCorrectionList = document.querySelector("#importCorrectionList");
    const importCorrectionCancel = document.querySelector("#importCorrectionCancel");
    const importCorrectionApply = document.querySelector("#importCorrectionApply");
    const pageTitle = document.querySelector(".title");
    const dashboardContent = document.querySelector("#dashboardContent");
    const companyTabs = Array.from(document.querySelectorAll(".company-tab"));
    const companyNavTabs = Array.from(document.querySelectorAll(".nav-subitem[data-company-tab]"));
    const companyPanels = Array.from(document.querySelectorAll("[data-company-panel]"));
    const companyOrgBody = document.querySelector("#companyOrgBody");
    const companyStaffRefresh = document.querySelector("#companyStaffRefresh");
    const companyStaffTaskBody = document.querySelector("#companyStaffTaskBody");
    const companyStaffDueToday = document.querySelector("#companyStaffDueToday");
    const companyStaffNoticeTitle = document.querySelector("#companyStaffNoticeTitle");
    const companyCalendarTitle = document.querySelector("#companyCalendarTitle");
    const companyCalendarGrid = document.querySelector("#companyCalendarGrid");
    const companyCalendarPrev = document.querySelector("#companyCalendarPrev");
    const companyCalendarToday = document.querySelector("#companyCalendarToday");
    const companyCalendarNext = document.querySelector("#companyCalendarNext");
    const companyCalendarRefresh = document.querySelector("#companyCalendarRefresh");
    const companyCalendarSelectedDate = document.querySelector("#companyCalendarSelectedDate");
    const companyCalendarSelectedList = document.querySelector("#companyCalendarSelectedList");
    const companyCalendarProjectCount = document.querySelector("#companyCalendarProjectCount");
    const companyCalendarTaskCount = document.querySelector("#companyCalendarTaskCount");
    const companyCalendarLeaveCount = document.querySelector("#companyCalendarLeaveCount");
    const companyCalendarRiskCount = document.querySelector("#companyCalendarRiskCount");
    const internalChatRoomList = document.querySelector("#internalChatRoomList");
    const internalChatTitle = document.querySelector("#internalChatTitle");
    const internalChatHint = document.querySelector("#internalChatHint");
    const internalChatList = document.querySelector("#internalChatList");
    const internalChatForm = document.querySelector("#internalChatForm");
    const internalChatBody = document.querySelector("#internalChatBody");
    const internalChatRefresh = document.querySelector("#internalChatRefresh");
    const noticeDateInput = document.querySelector("#noticeDateInput");
    const noticeTitleInput = document.querySelector("#noticeTitleInput");
    const noticeOwnerInput = document.querySelector("#noticeOwnerInput");
    const noticeBodyInput = document.querySelector("#noticeBodyInput");
    const noticeSaveButton = document.querySelector("#noticeSaveButton");
    const noticeClearButton = document.querySelector("#noticeClearButton");
    const noticePreview = document.querySelector("#noticePreview");
    const sidebarNoticePreview = document.querySelector("#sidebarNoticePreview");
    const noticePopup = document.querySelector("#noticePopup");
    const noticePopupClose = document.querySelector("#noticePopupClose");
    const focusWidget = document.querySelector("#focusWidget");
    const focusWidgetKicker = document.querySelector("#focusWidgetKicker");
    const focusWidgetTitle = document.querySelector("#focusWidgetTitle");
    const focusWidgetSubtitle = document.querySelector("#focusWidgetSubtitle");
    const focusWidgetBody = document.querySelector("#focusWidgetBody");
    const focusWidgetClose = document.querySelector("#focusWidgetClose");
    const importShipmentBody = document.querySelector("#importShipmentBody");
    const importProgressCard = document.querySelector("#importProgressCard");
    const importShipmentTreeToggle = document.querySelector("#importShipmentTreeToggle");
    const importShipmentSummary = document.querySelector("#importShipmentSummary");
    const importShipmentRefresh = document.querySelector("#importShipmentRefresh");
    const importShipmentInputOpen = document.querySelector("#importShipmentInputOpen");
    const importShipmentWorkspaceOpen = document.querySelector("#importShipmentWorkspaceOpen");
    const dashboardImportScheduleBody = document.querySelector("#dashboardImportScheduleBody");
    const dashboardImportScheduleSummary = document.querySelector("#dashboardImportScheduleSummary");
    const dashboardImportScheduleRefresh = document.querySelector("#dashboardImportScheduleRefresh");
    const dashboardImportScheduleOpen = document.querySelector("#dashboardImportScheduleOpen");
    const importShipmentPopup = document.querySelector("#importShipmentPopup");
    const importShipmentClose = document.querySelector("#importShipmentClose");
    const importShipmentSave = document.querySelector("#importShipmentSave");
    const importShipmentReset = document.querySelector("#importShipmentReset");
    const importShipmentId = document.querySelector("#importShipmentId");
    const importDepartureDate = document.querySelector("#importDepartureDate");
    const importArrivalDate = document.querySelector("#importArrivalDate");
    const importLoadingPort = document.querySelector("#importLoadingPort");
    const importArrivalPort = document.querySelector("#importArrivalPort");
    const importItem = document.querySelector("#importItem");
    const importQuantity = document.querySelector("#importQuantity");
    const importHblNo = document.querySelector("#importHblNo");
    const importSize = document.querySelector("#importSize");
    const importProgressStatus = document.querySelector("#importProgressStatus");
    const importFreeTime = document.querySelector("#importFreeTime");
    const importWarehouseDueDate = document.querySelector("#importWarehouseDueDate");
    const managementWorkspace = document.querySelector("#managementWorkspace");
    const ledgerWorkspace = document.querySelector("#ledgerWorkspace");
    const importWorkspace = document.querySelector("#importWorkspace");
    const orderWorkspace = document.querySelector("#orderWorkspace");
    const orderWorkspaceTitle = document.querySelector("#orderWorkspaceTitle");
    const orderWorkspacePanelTitle = document.querySelector("#orderWorkspacePanelTitle");
    const orderWorkspaceDescription = document.querySelector("#orderWorkspaceDescription");
    const orderWorkspaceCards = document.querySelector("#orderWorkspaceCards");
    const orderDownloadList = document.querySelector("#orderDownloadList");
    const orderDownloadRefresh = document.querySelector("#orderDownloadRefresh");
    const fileLibraryWorkspace = document.querySelector("#fileLibraryWorkspace");
    const sharedFileUploadPanel = document.querySelector("#sharedFileUploadPanel");
    const sharedFileInput = document.querySelector("#sharedFileInput");
    const sharedFileDropMain = document.querySelector("#sharedFileDropMain");
    const sharedFileUpload = document.querySelector("#sharedFileUpload");
    const sharedFileRefresh = document.querySelector("#sharedFileRefresh");
    const sharedFileBody = document.querySelector("#sharedFileBody");
    const sharedFileMessage = document.querySelector("#sharedFileMessage");
    const leaveWorkspace = document.querySelector("#leaveWorkspace");
    const userAdminWorkspace = document.querySelector("#userAdminWorkspace");
    const backupWorkspace = document.querySelector("#backupWorkspace");
    const managementWorkspaceMount = document.querySelector("#managementWorkspaceMount");
    const ledgerWorkspaceMount = document.querySelector("#ledgerWorkspaceMount");
    const userAdminRefresh = document.querySelector("#userAdminRefresh");
    const userAdminId = document.querySelector("#userAdminId");
    const userAdminUsername = document.querySelector("#userAdminUsername");
    const userAdminDisplayName = document.querySelector("#userAdminDisplayName");
    const userAdminRole = document.querySelector("#userAdminRole");
    const userAdminPassword = document.querySelector("#userAdminPassword");
    const userAdminActive = document.querySelector("#userAdminActive");
    const userAdminSave = document.querySelector("#userAdminSave");
    const userAdminBody = document.querySelector("#userAdminBody");
    const userAdminMessage = document.querySelector("#userAdminMessage");
    const adminNaverEmailInput = document.querySelector("#adminNaverEmailInput");
    const adminNaverPasswordInput = document.querySelector("#adminNaverPasswordInput");
    const adminSaveMailCredentials = document.querySelector("#adminSaveMailCredentials");
    const adminMailSettingsSave = document.querySelector("#adminMailSettingsSave");
    const adminMailSettingsMessage = document.querySelector("#adminMailSettingsMessage");
    const adminSmtpPort = document.querySelector("#adminSmtpPort");
    const adminSmtpSecurity = document.querySelector("#adminSmtpSecurity");
    const adminBulkBatchSize = document.querySelector("#adminBulkBatchSize");
    const adminBulkSendInterval = document.querySelector("#adminBulkSendInterval");
    const adminBulkBatchPause = document.querySelector("#adminBulkBatchPause");
    const adminBulkTestRecipient = document.querySelector("#adminBulkTestRecipient");
    const adminMailTechnicalSave = document.querySelector("#adminMailTechnicalSave");
    const adminMailTestSend = document.querySelector("#adminMailTestSend");
    const adminMailTechnicalMessage = document.querySelector("#adminMailTechnicalMessage");
    const userAdminPermissionChecks = Array.from(document.querySelectorAll("[data-permission-check]"));
    const leaveRefresh = document.querySelector("#leaveRefresh");
    const leaveTotalDays = document.querySelector("#leaveTotalDays");
    const leaveUsedDays = document.querySelector("#leaveUsedDays");
    const leaveReservedDays = document.querySelector("#leaveReservedDays");
    const leaveRemainingDays = document.querySelector("#leaveRemainingDays");
    const leaveMessage = document.querySelector("#leaveMessage");
    const leaveNotificationList = document.querySelector("#leaveNotificationList");
    const leaveBalanceBody = document.querySelector("#leaveBalanceBody");
    const leaveHistoryBody = document.querySelector("#leaveHistoryBody");
    const leaveApprovalBody = document.querySelector("#leaveApprovalBody");
    const leaveTypeSelect = document.querySelector("#leaveTypeSelect");
    const leaveUnitSelect = document.querySelector("#leaveUnitSelect");
    const leaveStartDate = document.querySelector("#leaveStartDate");
    const leaveEndDate = document.querySelector("#leaveEndDate");
    const leaveReasonInput = document.querySelector("#leaveReasonInput");
    const leaveRequestSubmit = document.querySelector("#leaveRequestSubmit");
    const leaveAdminUserSelect = document.querySelector("#leaveAdminUserSelect");
    const leaveAdminTotalInput = document.querySelector("#leaveAdminTotalInput");
    const leaveAdminUsedInput = document.querySelector("#leaveAdminUsedInput");
    const leaveBalanceSave = document.querySelector("#leaveBalanceSave");
    const leaveAccrualApply = document.querySelector("#leaveAccrualApply");
    const leaveUsageUserSelect = document.querySelector("#leaveUsageUserSelect");
    const leaveUsageDatesInput = document.querySelector("#leaveUsageDatesInput");
    const leaveUsageNoteInput = document.querySelector("#leaveUsageNoteInput");
    const leaveUsageSave = document.querySelector("#leaveUsageSave");
    const leaveHolidayDateInput = document.querySelector("#leaveHolidayDateInput");
    const leaveHolidayNameInput = document.querySelector("#leaveHolidayNameInput");
    const leaveHolidaySubstituteInput = document.querySelector("#leaveHolidaySubstituteInput");
    const leaveHolidaySave = document.querySelector("#leaveHolidaySave");
    const leaveAdminBalanceBody = document.querySelector("#leaveAdminBalanceBody");
    const leaveTabs = Array.from(document.querySelectorAll("[data-leave-tab]"));
    const backupRefresh = document.querySelector("#backupRefresh");
    const backupCreate = document.querySelector("#backupCreate");
    const backupCreateSelected = document.querySelector("#backupCreateSelected");
    const backupSettingsSave = document.querySelector("#backupSettingsSave");
    const backupAutoEnabled = document.querySelector("#backupAutoEnabled");
    const backupAutoHour = document.querySelector("#backupAutoHour");
    const backupRetentionDays = document.querySelector("#backupRetentionDays");
    const backupDirInput = document.querySelector("#backupDirInput");
    const backupExternalEnabled = document.querySelector("#backupExternalEnabled");
    const backupRcloneExecutable = document.querySelector("#backupRcloneExecutable");
    const backupRcloneRemote = document.querySelector("#backupRcloneRemote");
    const backupRclonePath = document.querySelector("#backupRclonePath");
    const backupExternalStatus = document.querySelector("#backupExternalStatus");
    const backupAutoState = document.querySelector("#backupAutoState");
    const backupRetentionState = document.querySelector("#backupRetentionState");
    const backupRestoreInput = document.querySelector("#backupRestoreInput");
    const backupPath = document.querySelector("#backupPath");
    const backupBody = document.querySelector("#backupBody");
    const backupMessage = document.querySelector("#backupMessage");
    const systemUpdateWorkspace = document.querySelector("#systemUpdateWorkspace");
    const systemUpdateRefresh = document.querySelector("#systemUpdateRefresh");
    const systemUpdateCheck = document.querySelector("#systemUpdateCheck");
    const systemUpdateApply = document.querySelector("#systemUpdateApply");
    const systemUpdateSource = document.querySelector("#systemUpdateSource");
    const systemUpdateBranch = document.querySelector("#systemUpdateBranch");
    const systemUpdateCurrent = document.querySelector("#systemUpdateCurrent");
    const systemUpdateState = document.querySelector("#systemUpdateState");
    const systemUpdateMessage = document.querySelector("#systemUpdateMessage");
    const systemUpdateHistoryBody = document.querySelector("#systemUpdateHistoryBody");
    const crmWorkspace = document.querySelector("#crmWorkspace");
    const crmRefresh = document.querySelector("#crmRefresh");
    const crmAccountQuick = document.querySelector("#crmAccountQuick");
    const crmTaskQuick = document.querySelector("#crmTaskQuick");
    const crmTabs = Array.from(document.querySelectorAll("[data-crm-tab]"));
    const crmPanels = Array.from(document.querySelectorAll("[data-crm-panel]"));
    const crmMessagesTab = document.querySelector("#crmMessagesTab");
    const crmMessage = document.querySelector("#crmMessage");
    const crmStatAccounts = document.querySelector("#crmStatAccounts");
    const crmStatOpenTasks = document.querySelector("#crmStatOpenTasks");
    const crmStatDueToday = document.querySelector("#crmStatDueToday");
    const crmStatOverdue = document.querySelector("#crmStatOverdue");
    const crmProjectProgressBody = document.querySelector("#crmProjectProgressBody");
    const crmPriorityTaskBody = document.querySelector("#crmPriorityTaskBody");
    const crmRecentMessageBody = document.querySelector("#crmRecentMessageBody");
    const crmAccountForm = document.querySelector("#crmAccountForm");
    const crmAccountId = document.querySelector("#crmAccountId");
    const crmAccountName = document.querySelector("#crmAccountName");
    const crmAccountType = document.querySelector("#crmAccountType");
    const crmAccountContact = document.querySelector("#crmAccountContact");
    const crmAccountPhone = document.querySelector("#crmAccountPhone");
    const crmAccountEmail = document.querySelector("#crmAccountEmail");
    const crmAccountMemo = document.querySelector("#crmAccountMemo");
    const crmAccountReset = document.querySelector("#crmAccountReset");
    const crmAccountSearch = document.querySelector("#crmAccountSearch");
    const crmAccountSearchButton = document.querySelector("#crmAccountSearchButton");
    const crmAccountBody = document.querySelector("#crmAccountBody");
    const crmStaffRefresh = document.querySelector("#crmStaffRefresh");
    const crmStaffBody = document.querySelector("#crmStaffBody");
    const crmTaskForm = document.querySelector("#crmTaskForm");
    const crmTaskId = document.querySelector("#crmTaskId");
    const crmTaskAccount = document.querySelector("#crmTaskAccount");
    const crmTaskAccountName = document.querySelector("#crmTaskAccountName");
    const crmTaskTitle = document.querySelector("#crmTaskTitle");
    const crmTaskAssignee = document.querySelector("#crmTaskAssignee");
    const crmTaskDue = document.querySelector("#crmTaskDue");
    const crmTaskPriority = document.querySelector("#crmTaskPriority");
    const crmTaskStatus = document.querySelector("#crmTaskStatus");
    const crmTaskDescription = document.querySelector("#crmTaskDescription");
    const crmTaskReset = document.querySelector("#crmTaskReset");
    const crmTaskFormToggle = document.querySelector("#crmTaskFormToggle");
    const crmTaskPresetList = document.querySelector("#crmTaskPresetList");
    const crmTaskViewSelect = document.querySelector("#crmTaskViewSelect");
    const crmTaskViewName = document.querySelector("#crmTaskViewName");
    const crmTaskViewSave = document.querySelector("#crmTaskViewSave");
    const crmTaskViewDelete = document.querySelector("#crmTaskViewDelete");
    const crmTaskSearch = document.querySelector("#crmTaskSearch");
    const crmTaskStatusFilter = document.querySelector("#crmTaskStatusFilter");
    const crmTaskAssigneeFilter = document.querySelector("#crmTaskAssigneeFilter");
    const crmTaskPriorityFilter = document.querySelector("#crmTaskPriorityFilter");
    const crmTaskDueFilter = document.querySelector("#crmTaskDueFilter");
    const crmTaskSourceFilter = document.querySelector("#crmTaskSourceFilter");
    const crmTaskSort = document.querySelector("#crmTaskSort");
    const crmTaskOpenOnly = document.querySelector("#crmTaskOpenOnly");
    const crmTaskSearchButton = document.querySelector("#crmTaskSearchButton");
    const crmTaskAdvancedToggle = document.querySelector("#crmTaskAdvancedToggle");
    const crmAdvancedFilters = document.querySelector("#crmAdvancedFilters");
    const crmTaskFilterReset = document.querySelector("#crmTaskFilterReset");
    const crmTaskBoardStats = document.querySelector("#crmTaskBoardStats");
    const crmTaskBody = document.querySelector("#crmTaskBody");
    const crmTaskDetail = document.querySelector("#crmTaskDetail");
    const crmMineStats = document.querySelector("#crmMineStats");
    const crmMineTaskBody = document.querySelector("#crmMineTaskBody");
    const crmWebhookUrl = document.querySelector("#crmWebhookUrl");
    const crmWebhookUrlCopy = document.querySelector("#crmWebhookUrlCopy");
    const crmWebhookHeader = document.querySelector("#crmWebhookHeader");
    const crmWebhookHeaderCopy = document.querySelector("#crmWebhookHeaderCopy");
    const crmWebhookToken = document.querySelector("#crmWebhookToken");
    const crmWebhookTokenCopy = document.querySelector("#crmWebhookTokenCopy");
    const crmWebhookTokenRotate = document.querySelector("#crmWebhookTokenRotate");
    const crmWebhookSamplePayload = document.querySelector("#crmWebhookSamplePayload");
    const crmWebhookCurl = document.querySelector("#crmWebhookCurl");
    const crmMessengerForm = document.querySelector("#crmMessengerForm");
    const crmMessengerPlatform = document.querySelector("#crmMessengerPlatform");
    const crmMessengerSenderKey = document.querySelector("#crmMessengerSenderKey");
    const crmMessengerDisplayName = document.querySelector("#crmMessengerDisplayName");
    const crmMessengerUser = document.querySelector("#crmMessengerUser");
    const crmMessengerUserBody = document.querySelector("#crmMessengerUserBody");
    const crmMessageEventBody = document.querySelector("#crmMessageEventBody");
    let currentMode = "dashboard";
    let currentOrderMode = "delivery";
    let vendorContacts = [];
    let cachedMailSettings = {};
    let activeCsCaseId = "";
    let ledgerCases = [];
    let managementRecords = [];
    let ledgerImportMode = "daily";
    let managementImportMode = "daily";
    let managementPeriods = [];
    let importShipments = [];
    let userAccounts = [];
    let crmAccounts = [];
    let crmTasks = [];
    let crmMineTasks = [];
    let crmProjectProgress = [];
    let crmUsers = [];
    let internalChatUsers = [];
    let internalChatRoom = { type: "global", userId: "" };
    let companyActiveTab = "notice";
    let companyCalendarMonth = todayString().slice(0, 7);
    let companyCalendarSelectedDay = todayString();
    let companyCalendarEvents = [];
    let companyCalendarSummary = {};
    let crmActiveTab = "dashboard";
    let crmSelectedTaskId = "";
    const CRM_TASK_STATUSES = ["대기", "진행중", "보류", "완료"];
    const CRM_TASK_BUILTIN_VIEWS = [
      { id: "open", name: "전체 미완료", filters: { open_only: "1", sort: "smart" } },
      { id: "today", name: "오늘 마감", filters: { due: "today", open_only: "1", sort: "due" } },
      { id: "overdue", name: "지연", filters: { due: "overdue", open_only: "1", sort: "due" } },
      { id: "mine", name: "내 업무", filters: { assignee_user_id: String(currentUser.id || ""), open_only: "1", sort: "due" } },
      { id: "messages", name: "메신저 지시", filters: { source: "internal_message", open_only: "1", sort: "updated" } },
    ];
    let crmSavedViews = [];
    let crmActiveTaskViewId = "builtin:open";
    const crmTaskComments = {};
    const crmTaskCommentLoads = {};
    let companyStaffPayloadCache = null;
    let companyStaffTaskCache = [];
    let crmStaffPayloadCache = null;
    let crmStaffTaskCache = [];
    let focusWidgetLastFocus = null;
    let focusWidgetTaskId = "";
    let focusWidgetEmployeeId = "";
    let activeLedgerFilterField = "";
    let activeManagementFilterField = "";
    const activeCellEditors = {
      ledger: null,
      management: null,
    };
    const ledgerFilters = {};
    const managementFilters = {};
    let isBulkSaving = false;

    if (managementWorkspaceMount && managementFields) managementWorkspaceMount.appendChild(managementFields);
    if (ledgerWorkspaceMount && ledgerFields) ledgerWorkspaceMount.appendChild(ledgerFields);
    if (ledgerFilterPopover) document.body.appendChild(ledgerFilterPopover);
    if (ledgerYearFilter && ledgerMonthFilter) fillPeriodSelects(ledgerYearFilter, ledgerMonthFilter);
    document.querySelectorAll('.company-panel[data-company-panel="rules"] .company-card').forEach((card) => {
      const label = card.querySelector(".company-card-head span")?.textContent?.trim() || "사규 카드";
      card.setAttribute("role", "button");
      card.setAttribute("tabindex", "0");
      card.setAttribute("aria-label", `${label} 크게 보기`);
    });
    renderManagementPeriodControls();
    applyStaticPermissions();
    loadNoticeTemplate();
    loadImportShipments();

    function addProductRow(productName = "", quantity = "", packQuantity = "") {
      const row = document.createElement("div");
      row.className = "product-row";
      row.innerHTML = `
        <input class="product-name" type="text" placeholder="제품명" value="${productName}">
        <input class="product-quantity" type="text" placeholder="수량" value="${quantity}">
        <input class="product-pack-quantity" type="text" placeholder="입수량" value="${packQuantity}">
      `;
      productTable.appendChild(row);
    }

    function defaultProductRowCount() {
      return receiptTypeSelect.value === "모드니 전용" ? 15 : 5;
    }

    function resetProductRows() {
      productTable.innerHTML = "";
      for (let i = 0; i < defaultProductRowCount(); i += 1) addProductRow();
    }

    function todayString() {
      const now = new Date();
      const offset = now.getTimezoneOffset() * 60000;
      return new Date(now.getTime() - offset).toISOString().slice(0, 10);
    }

    function fillPeriodSelects(yearSelect, monthSelect) {
      const currentYear = new Date().getFullYear();
      const startYear = 2023;
      for (let year = currentYear + 1; year >= startYear; year -= 1) {
        const option = document.createElement("option");
        option.value = String(year);
        option.textContent = `${year}년`;
        yearSelect.appendChild(option);
      }
      for (let month = 1; month <= 12; month += 1) {
        const option = document.createElement("option");
        option.value = String(month).padStart(2, "0");
        option.textContent = `${month}월`;
        monthSelect.appendChild(option);
      }
    }

    function ensureYearForMonth(yearSelect, monthSelect) {
      if (monthSelect.value && !yearSelect.value) yearSelect.value = String(new Date().getFullYear());
    }

    function managementPeriodYears() {
      return Array.from(new Set(managementPeriods.map((period) => String(period.year || "")).filter(Boolean)))
        .sort((left, right) => right.localeCompare(left));
    }

    function managementMonthsForYear(year) {
      return managementPeriods
        .filter((period) => String(period.year || "") === String(year || ""))
        .map((period) => String(period.month || "").padStart(2, "0"))
        .filter(Boolean)
        .sort();
    }

    function selectedManagementPeriod() {
      const selectedOption = managementMonthFilter?.selectedOptions?.[0];
      const optionYear = selectedOption?.dataset?.year || "";
      const optionMonth = selectedOption?.dataset?.month || "";
      if (optionYear && optionMonth) return { year: optionYear, month: optionMonth };
      const value = managementMonthFilter?.value || "";
      if (!value) return { year: "", month: "" };
      const match = value.match(/^(\d{4})-(\d{2})$/);
      if (match) return { year: match[1], month: match[2] };
      return { year: managementYearFilter?.value || "", month: "" };
    }

    function setManagementPeriod(year, month) {
      const normalizedYear = String(year || "");
      const normalizedMonth = String(month || "").padStart(2, "0");
      managementYearFilter.value = normalizedYear;
      managementMonthFilter.value = normalizedYear && normalizedMonth ? `${normalizedYear}-${normalizedMonth}` : "";
    }

    function renderManagementPeriodControls() {
      if (!managementYearFilter || !managementMonthFilter) return;
      const years = managementPeriodYears();
      const previousPeriod = selectedManagementPeriod();
      managementYearFilter.innerHTML = `<option value="">년도 선택</option>${years.map((year) => (
        `<option value="${escapeHtml(year)}">${escapeHtml(year)}년</option>`
      )).join("")}`;
      if (previousPeriod.year && years.includes(previousPeriod.year)) managementYearFilter.value = previousPeriod.year;
      else managementYearFilter.value = "";

      const periodOptions = managementPeriods.map((period) => {
        const year = String(period.year || "");
        const month = String(period.month || "").padStart(2, "0");
        return `<option value="${escapeHtml(`${year}-${month}`)}" data-year="${escapeHtml(year)}" data-month="${escapeHtml(month)}">${escapeHtml(year)}년 ${Number(month)}월</option>`;
      }).join("");
      managementMonthFilter.innerHTML = `<option value="">전체 선택</option>${periodOptions}`;
      if (previousPeriod.year && previousPeriod.month) {
        const value = `${previousPeriod.year}-${previousPeriod.month}`;
        managementMonthFilter.value = Array.from(managementMonthFilter.options).some((option) => option.value === value) ? value : "";
      } else {
        managementMonthFilter.value = "";
      }

      renderManagementMonthTabs();
    }

    function renderManagementMonthTabs() {
      if (!managementMonthTabs) return;
      const activePeriod = selectedManagementPeriod();
      const activeMonth = activePeriod.month || "";
      const activeYear = activePeriod.year || managementPeriodYears()[0] || "";
      const months = activeYear ? managementMonthsForYear(activeYear) : [];
      managementMonthTabs.innerHTML = months.length ? months.map((month) => {
        const active = month === activeMonth ? " active" : "";
        return `<button class="management-month-tab${active}" type="button" data-management-year="${escapeHtml(activeYear)}" data-management-month="${escapeHtml(month)}">${escapeHtml(activeYear)}년 ${Number(month)}월</button>`;
      }).join("") : `<button class="management-month-tab" type="button" disabled>생성된 월 없음</button>`;
    }

    function roleText(role) {
      if (role === "admin") return "관리자";
      if (role === "sub_admin") return "부관리자";
      return "사용자";
    }

    function resetUserAdminForm() {
      if (!userAdminId) return;
      userAdminId.value = "";
      userAdminUsername.value = "";
      userAdminUsername.disabled = false;
      userAdminDisplayName.value = "";
      userAdminRole.value = "user";
      userAdminPassword.value = "";
      userAdminActive.checked = true;
      userAdminPermissionChecks.forEach((checkbox) => {
        checkbox.checked = ["ledger_edit", "excel_download", "cs_receive", "leave_view", "crm_view"].includes(checkbox.value);
      });
      userAdminMessage.textContent = "신규 사용자는 10자 이상 비밀번호를 입력한 뒤 저장하세요. 가입 요청 사용자는 활성화하면 로그인할 수 있습니다.";
    }

    function syncPermissionChecksForRole() {
      if (!userAdminRole) return;
      if (userAdminRole.value === "admin") {
        userAdminPermissionChecks.forEach((checkbox) => {
          checkbox.checked = true;
        });
      } else if (userAdminRole.value === "sub_admin") {
        const defaults = new Set([
          "ledger_delete",
          "notice_manage",
          "ledger_edit",
          "excel_upload",
          "excel_download",
          "cs_receive",
          "mail_send",
          "import_shipment_manage",
          "leave_view",
          "leave_approve",
          "leave_manage",
          "crm_view",
          "crm_manage",
          "crm_message_manage",
        ]);
        userAdminPermissionChecks.forEach((checkbox) => {
          checkbox.checked = defaults.has(checkbox.value);
        });
      } else {
        const defaults = new Set(["ledger_edit", "excel_download", "cs_receive", "leave_view", "crm_view"]);
        userAdminPermissionChecks.forEach((checkbox) => {
          checkbox.checked = defaults.has(checkbox.value);
        });
      }
    }

    function renderUserAccounts() {
      if (!userAdminBody) return;
      if (!userAccounts.length) {
        userAdminBody.innerHTML = `<tr><td colspan="8">등록된 사용자가 없습니다.</td></tr>`;
        return;
      }
      userAdminBody.innerHTML = userAccounts.map((user) => `
        <tr data-user-id="${user.id}">
          <td>${escapeHtml(user.username)}</td>
          <td>${escapeHtml(user.display_name)}</td>
          <td>${roleText(user.role)}</td>
          <td>${(user.permissions || []).map((permission) => permissionLabel(permission)).join(", ")}</td>
          <td>${user.active ? "사용" : (user.approved_at ? "중지" : "승인대기")}</td>
          <td>${escapeHtml(user.created_at || "")}</td>
          <td>${escapeHtml(user.last_login_at || "없음")}</td>
          <td><button class="admin-action" type="button" data-user-edit="${user.id}">${user.active ? "수정" : "승인/수정"}</button></td>
        </tr>
      `).join("");
    }

    function editUserAccount(userId) {
      const user = userAccounts.find((item) => String(item.id) === String(userId));
      if (!user || !userAdminId) return;
      userAdminId.value = user.id;
      userAdminUsername.value = user.username || "";
      userAdminUsername.disabled = false;
      userAdminDisplayName.value = user.display_name || "";
      userAdminRole.value = user.role || "user";
      userAdminPassword.value = "";
      userAdminActive.checked = Boolean(user.active);
      const permissions = new Set(user.permissions || []);
      userAdminPermissionChecks.forEach((checkbox) => {
        checkbox.checked = permissions.has(checkbox.value);
      });
      userAdminMessage.textContent = `${user.username} 계정 수정 중입니다. ${user.active ? "비밀번호는 변경할 때만 입력하세요." : "사용을 체크하고 저장하면 승인됩니다."}`;
      userAdminUsername?.focus();
    }

    async function loadUserAccounts() {
      if (!userAdminBody) return;
      userAdminMessage.textContent = "사용자 목록을 불러오는 중입니다.";
      try {
        const response = await fetch("/api/users");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "사용자 목록을 불러오지 못했습니다.");
        userAccounts = data.users || [];
        renderUserAccounts();
        if (!userAdminId.value) resetUserAdminForm();
      } catch (error) {
        userAdminMessage.textContent = error.message;
      }
    }

    async function saveUserAccount() {
      if (!userAdminSave) return;
      const payload = {
        id: userAdminId.value,
        username: userAdminUsername.value.trim(),
        display_name: userAdminDisplayName.value.trim(),
        role: userAdminRole.value,
        password: userAdminPassword.value,
        active: userAdminActive.checked,
        permissions: userAdminPermissionChecks.filter((checkbox) => checkbox.checked).map((checkbox) => checkbox.value),
      };
      userAdminMessage.textContent = "사용자 계정을 저장하는 중입니다.";
      userAdminSave.disabled = true;
      try {
        const response = await fetch("/api/users-save", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "사용자 계정을 저장하지 못했습니다.");
        userAccounts = data.users || [];
        renderUserAccounts();
        resetUserAdminForm();
        userAdminMessage.textContent = data.message || "사용자 계정을 저장했습니다.";
      } catch (error) {
        userAdminMessage.textContent = error.message;
      } finally {
        userAdminSave.disabled = false;
      }
    }

    function backupSizeText(size) {
      const number = Number(size || 0);
      if (number >= 1024 * 1024) return `${(number / 1024 / 1024).toFixed(1)} MB`;
      if (number >= 1024) return `${(number / 1024).toFixed(1)} KB`;
      return `${number} B`;
    }

    function renderBackups(data) {
      if (!backupBody) return;
      const settings = data.settings || {};
      backupPath.textContent = data.backup_dir || "-";
      if (backupAutoState) {
        backupAutoState.textContent = settings.auto_enabled === false
          ? "사용 안 함"
          : `매일 ${String(settings.auto_hour ?? data.auto_backup_hour ?? 3).padStart(2, "0")}:00`;
      }
      if (backupRetentionState) backupRetentionState.textContent = `최근 ${settings.retention_days || data.retention_days || 90}일`;
      if (backupAutoEnabled) backupAutoEnabled.checked = settings.auto_enabled !== false;
      if (backupAutoHour) backupAutoHour.value = String(settings.auto_hour ?? data.auto_backup_hour ?? 3);
      if (backupRetentionDays) backupRetentionDays.value = String(settings.retention_days || data.retention_days || 90);
      if (backupDirInput) backupDirInput.value = settings.backup_dir || data.backup_dir || "";
      if (backupExternalEnabled) backupExternalEnabled.checked = settings.external_enabled === true;
      if (backupRcloneExecutable) backupRcloneExecutable.value = settings.rclone_executable || "rclone";
      if (backupRcloneRemote) backupRcloneRemote.value = settings.rclone_remote || "";
      if (backupRclonePath) backupRclonePath.value = settings.rclone_path || "";
      if (backupExternalStatus) {
        const status = settings.last_external_status || (settings.external_enabled ? "대기" : "사용 안 함");
        const uploadedAt = settings.last_external_uploaded_at ? ` / ${settings.last_external_uploaded_at}` : "";
        const target = settings.last_external_target ? ` / ${settings.last_external_target}` : "";
        const message = settings.last_external_message ? ` / ${settings.last_external_message}` : "";
        backupExternalStatus.textContent = `Google Drive 업로드 상태: ${status}${uploadedAt}${target}${message}`;
      }
      const backups = data.backups || [];
      if (!backups.length) {
        backupBody.innerHTML = `<tr><td colspan="4">생성된 백업 파일이 없습니다.</td></tr>`;
        return;
      }
      backupBody.innerHTML = backups.map((backup) => `
        <tr>
          <td>${escapeHtml(backup.name)}</td>
          <td>${escapeHtml(backup.created_at || "")}</td>
          <td>${backupSizeText(backup.size)}</td>
          <td>
            <div class="backup-actions">
              <button class="backup-action" type="button" data-backup-restore="${escapeHtml(backup.name)}">복원</button>
              <button class="backup-action" type="button" data-backup-download="${escapeHtml(backup.name)}">다운로드</button>
              <button class="backup-action danger" type="button" data-backup-delete="${escapeHtml(backup.name)}">삭제</button>
            </div>
          </td>
        </tr>
      `).join("");
    }

    async function loadBackups() {
      if (!backupWorkspace) return;
      backupMessage.textContent = "백업 목록을 불러오는 중입니다.";
      try {
        const response = await fetch("/api/backups");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "백업 목록을 불러오지 못했습니다.");
        renderBackups(data);
        backupMessage.textContent = data.last_backup ? `마지막 백업: ${data.last_backup}` : "아직 백업이 없습니다.";
      } catch (error) {
        backupMessage.textContent = error.message;
      }
    }

    async function createBackupNow() {
      if (!backupCreate) return;
      backupCreate.disabled = true;
      backupMessage.textContent = "백업 파일을 생성하는 중입니다.";
      try {
        const response = await fetch("/api/backup-create", { method: "POST" });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "백업 생성에 실패했습니다.");
        const external = data.backup?.external_backup;
        backupMessage.textContent = external && external.status !== "disabled"
          ? `${data.message || "백업 파일을 생성했습니다."} Google Drive 업로드: ${external.status}`
          : data.message || "백업 파일을 생성했습니다.";
        await loadBackups();
      } catch (error) {
        backupMessage.textContent = error.message;
      } finally {
        backupCreate.disabled = false;
      }
    }

    function downloadBackup(name) {
      const url = `/api/backup-download?name=${encodeURIComponent(name)}`;
      window.location.href = url;
    }

    async function deleteBackup(name) {
      if (!confirm(`${name} 백업 파일을 삭제할까요?`)) return;
      backupMessage.textContent = "백업 파일을 삭제하는 중입니다.";
      try {
        const response = await fetch("/api/backup-delete", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ name }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "백업 삭제에 실패했습니다.");
        backupMessage.textContent = data.message || "백업 파일을 삭제했습니다.";
        await loadBackups();
      } catch (error) {
        backupMessage.textContent = error.message;
      }
    }

    function restoreCompleteMessage(data) {
      return `${data.message || "백업 데이터 복원이 완료되었습니다."} 프로그램을 다시 실행하거나 로그아웃 후 다시 로그인해주세요.`;
    }

    async function restoreBackupByName(name) {
      if (!confirm(`${name} 백업 데이터로 현재 업무 데이터를 복원할까요?\n\n복원 전 현재 데이터는 자동으로 예비 백업됩니다.`)) return;
      backupMessage.textContent = "백업 데이터를 복원하는 중입니다.";
      try {
        const response = await fetch("/api/backup-restore", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ name }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "백업 복원에 실패했습니다.");
        backupMessage.textContent = restoreCompleteMessage(data);
      } catch (error) {
        backupMessage.textContent = error.message;
      }
    }

    async function restoreBackupFromUpload() {
      if (!backupRestoreInput || !backupRestoreInput.files[0]) return;
      const file = backupRestoreInput.files[0];
      if (!confirm(`${file.name} 파일로 현재 업무 데이터를 복원할까요?\n\n복원 전 현재 데이터는 자동으로 예비 백업됩니다.`)) {
        backupRestoreInput.value = "";
        return;
      }
      const formData = new FormData();
      formData.append("file", file);
      backupMessage.textContent = "업로드한 백업 데이터를 복원하는 중입니다.";
      try {
        const response = await fetch("/api/backup-restore-upload", {
          method: "POST",
          body: formData,
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "백업 복원에 실패했습니다.");
        backupMessage.textContent = restoreCompleteMessage(data);
      } catch (error) {
        backupMessage.textContent = error.message;
      } finally {
        backupRestoreInput.value = "";
      }
    }

    function renderSystemUpdate(data) {
      if (!systemUpdateWorkspace) return;
      const status = data.status || {};
      systemUpdateSource.textContent = status.source_dir || "Git 저장소 연결 없음";
      systemUpdateBranch.textContent = status.branch || "-";
      systemUpdateCurrent.textContent = status.current_short || "-";
      if (!status.available) {
        systemUpdateState.textContent = status.message || "Git 저장소 연결 없음";
      } else if (status.behind > 0) {
        systemUpdateState.textContent = `${status.behind}개 업데이트 가능`;
      } else {
        systemUpdateState.textContent = "최신 상태";
      }
      systemUpdateMessage.textContent = status.message || "";
      const history = data.history || [];
      systemUpdateHistoryBody.innerHTML = history.length
        ? history.map((row) => `
          <tr>
            <td>${escapeHtml(row.created_at || "")}</td>
            <td>${escapeHtml(row.action || "")}</td>
            <td>${escapeHtml(row.status || "")}</td>
            <td>${escapeHtml(row.before_commit || "")}</td>
            <td>${escapeHtml(row.after_commit || "")}</td>
            <td>${escapeHtml(row.backup_name || "")}</td>
            <td>${escapeHtml(row.message || "")}</td>
          </tr>
        `).join("")
        : `<tr><td colspan="7">업데이트 이력이 없습니다.</td></tr>`;
      if (systemUpdateApply) systemUpdateApply.disabled = !status.available || Boolean(status.dirty);
      if (systemUpdateCheck) systemUpdateCheck.disabled = !status.available;
    }

    async function loadSystemUpdateStatus() {
      if (!systemUpdateWorkspace) return;
      systemUpdateMessage.textContent = "업데이트 정보를 불러오는 중입니다.";
      try {
        const response = await fetch("/api/system-update");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "업데이트 정보를 불러오지 못했습니다.");
        renderSystemUpdate(data);
      } catch (error) {
        systemUpdateMessage.textContent = error.message;
      }
    }

    async function checkSystemUpdate() {
      if (!systemUpdateCheck) return;
      systemUpdateCheck.disabled = true;
      systemUpdateMessage.textContent = "GitHub 업데이트를 확인하는 중입니다.";
      try {
        const response = await fetch("/api/system-update-check", { method: "POST" });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "업데이트 확인에 실패했습니다.");
        renderSystemUpdate(data);
      } catch (error) {
        systemUpdateMessage.textContent = error.message;
      } finally {
        systemUpdateCheck.disabled = false;
      }
    }

    async function applySystemUpdate() {
      if (!confirm("업데이트 적용 전 현재 데이터를 자동 백업합니다.\nGitHub 최신 코드를 적용할까요?")) return;
      systemUpdateApply.disabled = true;
      systemUpdateMessage.textContent = "업데이트 적용 중입니다. 창을 닫지 마세요.";
      try {
        const response = await fetch("/api/system-update-apply", { method: "POST" });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "업데이트 적용에 실패했습니다.");
        renderSystemUpdate(data);
        systemUpdateMessage.textContent = `${data.message || "업데이트 적용이 완료되었습니다."} 프로그램/나스 서비스를 다시 시작해주세요.`;
      } catch (error) {
        systemUpdateMessage.textContent = error.message;
      } finally {
        systemUpdateApply.disabled = false;
      }
    }

    function dayText(value) {
      const number = Number(value || 0);
      return `${Number.isInteger(number) ? number : number.toFixed(1)}일`;
    }

    function renderLeaveSelectOptions(select, rows, valueField = "id", labelField = "name") {
      if (!select) return;
      select.innerHTML = rows.map((row) => `<option value="${row[valueField]}">${escapeHtml(row[labelField])}</option>`).join("");
    }

    function setLeaveTab(tabName) {
      leaveTabs.forEach((button) => button.classList.toggle("active", button.dataset.leaveTab === tabName));
      ["mine", "request", "approvals", "admin"].forEach((name) => {
        const panel = document.querySelector(`#leaveTab${name[0].toUpperCase()}${name.slice(1)}`);
        if (panel) panel.classList.toggle("active", name === tabName);
      });
    }

    function renderLeaveData(data) {
      if (!leaveWorkspace) return;
      leaveTotalDays.textContent = dayText(data.summary?.total_days);
      leaveUsedDays.textContent = dayText(data.summary?.used_days);
      if (leaveReservedDays) leaveReservedDays.textContent = dayText(data.summary?.reserved_days);
      leaveRemainingDays.textContent = dayText(data.summary?.remaining_days);
      if (leaveNotificationList) {
        leaveNotificationList.innerHTML = (data.notifications || []).slice(0, 5).map((item) => `
          <div>${escapeHtml(item.message)} <small>${escapeHtml(item.created_at || "")}</small></div>
        `).join("");
      }
      leaveBalanceBody.innerHTML = (data.balances || []).length
        ? data.balances.map((row) => `
          <tr><td>${escapeHtml(row.name)}</td><td>${dayText(row.total_days)}</td><td>${dayText(row.used_days)} / 예약 ${dayText(row.reserved_days)}</td><td><strong>${dayText(row.remaining_days)}</strong></td></tr>
        `).join("")
        : `<tr><td colspan="4">연차 기준이 아직 설정되지 않았습니다.</td></tr>`;
      leaveHistoryBody.innerHTML = (data.requests || []).length
        ? data.requests.map((row) => `
          <tr>
            <td>${escapeHtml(row.start_date)}${row.start_date === row.end_date ? "" : ` ~ ${escapeHtml(row.end_date)}`}</td>
            <td>${escapeHtml(row.unit_label)}</td>
            <td>${dayText(row.requested_days)}</td>
            <td>${escapeHtml(row.status_label)}${row.status === "PENDING" ? ` ? ${escapeHtml(row.approval_step_label)}` : ""}</td>
            <td>${escapeHtml(row.reason)}${row.status === "PENDING" ? `<div class="leave-action-row"><input class="leave-comment-input" data-leave-cancel-reason="${row.id}" placeholder="\uCDE8\uC18C \uC0AC\uC720" /><button class="leave-action reject" type="button" data-leave-cancel="${row.id}">\uCDE8\uC18C</button></div>` : ""}</td>
          </tr>
        `).join("")
        : `<tr><td colspan="5">연차 사용/신청 이력이 없습니다.</td></tr>`;
      renderLeaveSelectOptions(leaveTypeSelect, data.leave_types || []);
      const userOptions = (data.users || []).map((user) => ({
        id: user.id,
        name: `${user.display_name} (${user.username})`,
      }));
      renderLeaveSelectOptions(leaveAdminUserSelect, userOptions);
      renderLeaveSelectOptions(leaveUsageUserSelect, userOptions);
      leaveApprovalBody.innerHTML = (data.pending_requests || []).length
        ? data.pending_requests.map((row) => `
          <tr>
            <td>${escapeHtml(row.requester)}</td>
            <td>${escapeHtml(row.start_date)}${row.start_date === row.end_date ? "" : ` ~ ${escapeHtml(row.end_date)}`}</td>
            <td>${escapeHtml(row.unit_label)}</td>
            <td>${dayText(row.requested_days)}</td>
            <td>${escapeHtml(row.reason)}</td>
            <td>
              <input class="leave-comment-input" data-leave-comment="${row.id}" placeholder="\uC2B9\uC778/\uBC18\uB824 \uCF54\uBA58\uD2B8" />
              <div class="leave-action-row">
                <button class="leave-action approve" type="button" data-leave-decision="approve" data-leave-id="${row.id}">\uC2B9\uC778</button>
                <button class="leave-action reject" type="button" data-leave-decision="reject" data-leave-id="${row.id}">\uBC18\uB824</button>
                ${data.can_override ? `<button class="leave-action approve" type="button" data-leave-decision="override" data-leave-id="${row.id}">\uC2E4\uC7A5 \uC804\uACB0</button>` : ""}
              </div>
            </td>
          </tr>
        `).join("")
        : `<tr><td colspan="6">승인 대기 중인 연차 신청이 없습니다.</td></tr>`;
      leaveAdminBalanceBody.innerHTML = (data.admin_balances || []).length
        ? data.admin_balances.map((row) => `
          <tr><td>${escapeHtml(row.display_name)}</td><td>${escapeHtml(row.username)}</td><td>${dayText(row.total_days)}</td><td>${dayText(row.used_days)}</td><td><strong>${dayText(row.remaining_days)}</strong></td></tr>
        `).join("")
        : `<tr><td colspan="5">직원 연차 현황이 없습니다.</td></tr>`;
      document.querySelector('[data-leave-tab="approvals"]')?.classList.toggle("permission-hidden", !data.can_approve);
      document.querySelector('[data-leave-tab="admin"]')?.classList.toggle("permission-hidden", !data.can_manage);
    }

    async function loadLeaveData() {
      if (!leaveWorkspace) return;
      leaveMessage.textContent = "연차 정보를 불러오는 중입니다.";
      try {
        const response = await fetch("/api/leaves");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "연차 정보를 불러오지 못했습니다.");
        renderLeaveData(data);
        leaveMessage.textContent = "";
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    async function submitLeaveRequest() {
      leaveMessage.textContent = "연차 신청을 저장하는 중입니다.";
      try {
        const payload = {
          leave_type_id: leaveTypeSelect.value,
          unit: leaveUnitSelect.value,
          start_date: leaveStartDate.value,
          end_date: leaveEndDate.value,
          reason: leaveReasonInput.value.trim(),
        };
        const response = await fetch("/api/leave-request", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "연차 신청에 실패했습니다.");
        leaveReasonInput.value = "";
        leaveMessage.textContent = data.message || "연차 신청이 저장되었습니다.";
        await loadLeaveData();
        if (companyActiveTab === "calendar") await loadCompanyCalendar().catch(() => {});
        setLeaveTab("mine");
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    function syncHalfDayDates() {
      if (!leaveUnitSelect || !leaveStartDate || !leaveEndDate) return;
      const isHalfDay = leaveUnitSelect.value === "HALF_DAY";
      if (isHalfDay && leaveStartDate.value) {
        leaveEndDate.value = leaveStartDate.value;
      }
      leaveEndDate.disabled = isHalfDay;
    }

    async function decideLeaveRequest(requestId, decision) {
      const commentInput = document.querySelector(`[data-leave-comment="${requestId}"]`);
      const comment = commentInput?.value?.trim() || (decision === "reject" ? "\uBC18\uB824" : "");
      leaveMessage.textContent = "연차 신청을 처리하는 중입니다.";
      try {
        const response = await fetch("/api/leave-decision", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ request_id: requestId, decision, comment }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "연차 신청 처리에 실패했습니다.");
        leaveMessage.textContent = data.message || "처리되었습니다.";
        await loadLeaveData();
        if (companyActiveTab === "calendar") await loadCompanyCalendar().catch(() => {});
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    async function cancelLeaveRequest(requestId) {
      const reasonInput = document.querySelector(`[data-leave-cancel-reason="${requestId}"]`);
      const reason = reasonInput?.value?.trim() || "\uC2E0\uCCAD\uC790 \uCDE8\uC18C";
      leaveMessage.textContent = "\uC5F0\uCC28 \uC2E0\uCCAD\uC744 \uCDE8\uC18C\uD558\uB294 \uC911\uC785\uB2C8\uB2E4.";
      try {
        const response = await fetch("/api/leave-cancel", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ request_id: requestId, reason }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "\uC5F0\uCC28 \uC790\uB3D9 \uBC1C\uC0DD\uC744 \uCC98\uB9AC\uD588\uC2B5\uB2C8\uB2E4.");
        leaveMessage.textContent = data.message || "\uC5F0\uCC28 \uC2E0\uCCAD\uC744 \uCDE8\uC18C\uD588\uC2B5\uB2C8\uB2E4.";
        await loadLeaveData();
        if (companyActiveTab === "calendar") await loadCompanyCalendar().catch(() => {});
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    async function applyLeaveAccrual() {
      const year = new Date().getFullYear();
      leaveMessage.textContent = `${year}\uB144 \uC5F0\uCC28\uB97C \uC790\uB3D9 \uBC1C\uC0DD\uD558\uB294 \uC911\uC785\uB2C8\uB2E4.`;
      try {
        const response = await fetch("/api/leave-accrual", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ year, default_days: 15 }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "\uC5F0\uCC28 \uC790\uB3D9 \uBC1C\uC0DD\uC744 \uCC98\uB9AC\uD588\uC2B5\uB2C8\uB2E4.");
        leaveMessage.textContent = data.message || "\uC5F0\uCC28 \uC790\uB3D9 \uBC1C\uC0DD\uC744 \uCC98\uB9AC\uD588\uC2B5\uB2C8\uB2E4.";
        await loadLeaveData();
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    async function saveLeaveBalance() {
      leaveMessage.textContent = "직원 연차 기준을 저장하는 중입니다.";
      try {
        const response = await fetch("/api/leave-balance", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({
            user_id: leaveAdminUserSelect.value,
            total_days: leaveAdminTotalInput.value,
            used_days: leaveAdminUsedInput.value,
          }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "연차 기준 저장에 실패했습니다.");
        leaveMessage.textContent = data.message || "연차 기준을 저장했습니다.";
        await loadLeaveData();
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    async function saveHistoricalLeaveUsage() {
      leaveMessage.textContent = "기존 사용 연차를 등록하는 중입니다.";
      try {
        const response = await fetch("/api/leave-historical", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({
            user_id: leaveUsageUserSelect.value,
            usage_dates: leaveUsageDatesInput.value,
            note: leaveUsageNoteInput.value,
          }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "기존 사용 연차 등록에 실패했습니다.");
        leaveUsageDatesInput.value = "";
        leaveUsageNoteInput.value = "";
        leaveMessage.textContent = data.message || "기존 사용 연차를 등록했습니다.";
        await loadLeaveData();
        if (companyActiveTab === "calendar") await loadCompanyCalendar().catch(() => {});
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    function loadNoticeTemplate() {
      let saved = {};
      try {
        saved = JSON.parse(localStorage.getItem("workhub_notice_template") || "{}");
      } catch {
        saved = {};
      }
      noticeDateInput.value = saved.date || todayString();
      noticeTitleInput.value = saved.title || "";
      noticeOwnerInput.value = saved.owner || "";
      noticeBodyInput.value = saved.body || "";
      renderNoticePreview();
    }

    function noticePayload() {
      return {
        date: noticeDateInput.value || todayString(),
        title: noticeTitleInput.value.trim(),
        owner: noticeOwnerInput.value.trim(),
        body: noticeBodyInput.value.trim(),
      };
    }

    function renderNoticePreview() {
      const payload = noticePayload();
      if (!payload.title && !payload.body) {
        noticePreview.innerHTML = `<strong>저장된 공지사항이 없습니다.</strong>공지사항을 입력하고 저장하면 이곳에서 미리 볼 수 있습니다.`;
        sidebarNoticePreview.innerHTML = `
          <div class="notice-board-kicker">금일 공지사항</div>
          <div class="notice-board-title">등록된 공지 없음</div>
          <div class="notice-board-body">공지사항 입력 버튼을 눌러 내용을 입력해주세요.</div>
        `;
        return;
      }
      const meta = [shortKoreanDate(payload.date), payload.owner ? `담당 ${escapeHtml(payload.owner)}` : ""].filter(Boolean).join(" / ");
      noticePreview.innerHTML = `
        <strong>${escapeHtml(payload.title || "제목 없음")}</strong>
        <div>${escapeHtml(meta)}</div>
        <div>${escapeHtml(payload.body).replaceAll("\n", "<br>")}</div>
      `;
      sidebarNoticePreview.innerHTML = `
        <div class="notice-board-kicker">금일 공지사항</div>
        <div class="notice-board-title">${escapeHtml(payload.title || "제목 없음")}</div>
        <div class="notice-board-meta">${escapeHtml(meta)}</div>
        <div class="notice-board-body">${escapeHtml(payload.body || "내용 없음")}</div>
      `;
    }

    function openNoticePopup() {
      if (!can("notice_manage")) {
        notice.textContent = "공지사항 관리 권한이 없습니다.";
        return;
      }
      loadNoticeTemplate();
      noticePopup.classList.add("open");
      setTimeout(() => noticeTitleInput?.focus(), 0);
    }

    function closeNoticePopup() {
      noticePopup.classList.remove("open");
    }

    function saveNoticeTemplate() {
      if (!can("notice_manage")) {
        notice.textContent = "공지사항 관리 권한이 없습니다.";
        return;
      }
      const payload = noticePayload();
      localStorage.setItem("workhub_notice_template", JSON.stringify(payload));
      renderNoticePreview();
      if (companyStaffNoticeTitle) companyStaffNoticeTitle.textContent = payload.title || "등록 전";
      notice.textContent = "공지사항을 저장했습니다.";
      closeNoticePopup();
    }

    function clearNoticeTemplate() {
      if (!can("notice_manage")) {
        notice.textContent = "공지사항 관리 권한이 없습니다.";
        return;
      }
      localStorage.removeItem("workhub_notice_template");
      noticeTitleInput.value = "";
      noticeOwnerInput.value = "";
      noticeBodyInput.value = "";
      noticeDateInput.value = todayString();
      renderNoticePreview();
      if (companyStaffNoticeTitle) companyStaffNoticeTitle.textContent = "등록 전";
      notice.textContent = "공지사항 입력 내용을 초기화했습니다.";
    }

    function resetImportShipmentForm(record = null) {
      importShipmentId.value = record?.id || "";
      importDepartureDate.value = record?.departure_date || "";
      importArrivalDate.value = record?.arrival_date || "";
      importLoadingPort.value = record?.loading_port || "";
      importArrivalPort.value = record?.arrival_port || "";
      importItem.value = record?.item || "";
      importQuantity.value = record?.quantity || "";
      importHblNo.value = record?.hbl_no || "";
      importSize.value = record?.size || "";
      importProgressStatus.value = record?.progress_status || "";
      importFreeTime.value = record?.free_time || "";
      importWarehouseDueDate.value = record?.warehouse_due_date || "";
    }

    function importShipmentPayload() {
      return {
        id: importShipmentId.value,
        departure_date: importDepartureDate.value.trim(),
        arrival_date: importArrivalDate.value.trim(),
        loading_port: importLoadingPort.value.trim(),
        arrival_port: importArrivalPort.value.trim(),
        shipper: "",
        item: importItem.value.trim(),
        quantity: importQuantity.value.trim(),
        vessel_name: "",
        hbl_no: importHblNo.value.trim(),
        size: importSize.value.trim(),
        progress_status: importProgressStatus.value.trim(),
        free_time: importFreeTime.value.trim(),
        warehouse_due_date: importWarehouseDueDate.value.trim(),
      };
    }

    function openImportShipmentPopup(record = null) {
      if (!can("import_shipment_manage")) {
        notice.textContent = "수입제품 진행 관리 권한이 없습니다.";
        return;
      }
      resetImportShipmentForm(record);
      importShipmentPopup.classList.add("open");
      setTimeout(() => importDepartureDate?.focus(), 0);
    }

    function closeImportShipmentPopup() {
      importShipmentPopup.classList.remove("open");
    }

    function renderDashboardImportSchedule() {
      if (!dashboardImportScheduleBody || !dashboardImportScheduleSummary) return;
      const activeRecords = importShipments.filter((record) => !record.completed_at);
      dashboardImportScheduleSummary.textContent = `진행 ${activeRecords.length}건`;
      if (!activeRecords.length) {
        dashboardImportScheduleBody.innerHTML = `<tr><td colspan="5"><div class="import-empty">등록된 수입제품 입고 일정이 없습니다.</div></td></tr>`;
        return;
      }
      dashboardImportScheduleBody.innerHTML = activeRecords.slice(0, 6).map((record) => `
        <tr>
          <td>${escapeHtml(record.arrival_date || "-")}</td>
          <td class="left">${escapeHtml(record.item || "-")}</td>
          <td>${escapeHtml(record.quantity || "-")}</td>
          <td>${escapeHtml(record.progress_status || "진행중")}</td>
          <td>${escapeHtml(record.warehouse_due_date || "-")}</td>
        </tr>
      `).join("");
    }

    function renderImportShipments() {
      const activeCount = importShipments.filter((record) => !record.completed_at).length;
      const doneCount = importShipments.length - activeCount;
      importShipmentSummary.textContent = `진행 ${activeCount}건 / 완료 ${doneCount}건`;
      renderDashboardImportSchedule();
      if (!importShipments.length) {
        importShipmentBody.innerHTML = `<tr><td colspan="11"><div class="import-empty">등록된 수입제품 출고 진행 건이 없습니다.</div></td></tr>`;
        return;
      }
      importShipmentBody.innerHTML = "";
      importShipments.forEach((record) => {
        const row = document.createElement("tr");
        if (record.completed_at) row.classList.add("completed");
        const progressCell = can("import_shipment_manage")
          ? `<td>
              <span>${escapeHtml(record.completed_at ? "완료" : record.progress_status)}</span>
              <span class="import-row-actions">
                <button type="button" data-import-edit="${record.id}">수정</button>
                ${record.completed_at ? "" : `<button type="button" data-import-complete="${record.id}">완료</button>`}
              </span>
            </td>`
          : `<td>${escapeHtml(record.completed_at ? "완료" : record.progress_status)}</td>`;
        row.innerHTML = `
          <td>${escapeHtml(record.departure_date)}</td>
          <td>${escapeHtml(record.arrival_date)}</td>
          <td>${escapeHtml(record.loading_port)}</td>
          <td>${escapeHtml(record.arrival_port)}</td>
          <td class="left">${escapeHtml(record.item)}</td>
          <td>${escapeHtml(record.quantity)}</td>
          <td>${escapeHtml(record.hbl_no)}</td>
          <td>${escapeHtml(record.size)}</td>
          ${progressCell}
          <td>${escapeHtml(record.free_time)}</td>
          <td>${escapeHtml(record.warehouse_due_date)}</td>
        `;
        importShipmentBody.appendChild(row);
      });
    }

    async function loadImportShipments() {
      try {
        const response = await fetch("/api/import-shipments");
        if (!response.ok) throw new Error("수입제품 진행 상황을 불러오지 못했습니다.");
        const data = await response.json();
        importShipments = data.shipments || [];
        renderImportShipments();
      } catch (error) {
        importShipments = [];
        renderImportShipments();
        notice.textContent = error.message;
      }
    }

    async function saveImportShipment() {
      if (!can("import_shipment_manage")) {
        notice.textContent = "수입제품 진행 관리 권한이 없습니다.";
        return;
      }
      const payload = importShipmentPayload();
      const hasContent = Object.entries(payload).some(([key, value]) => key !== "id" && String(value || "").trim());
      if (!hasContent) {
        notice.textContent = "수입제품 출고 진행 내용을 입력해주세요.";
        return;
      }
      const response = await fetch("/api/import-shipment-save", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "수입제품 출고 진행 저장에 실패했습니다.");
      notice.textContent = data.message || "수입제품 출고 진행 상황을 저장했습니다.";
      closeImportShipmentPopup();
      await loadImportShipments();
    }

    async function completeImportShipment(id) {
      if (!can("import_shipment_manage")) {
        notice.textContent = "수입제품 진행 관리 권한이 없습니다.";
        return;
      }
      const response = await fetch("/api/import-shipment-complete", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ id }),
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "수입제품 출고 진행 완료 처리에 실패했습니다.");
      notice.textContent = data.message || "완료 처리했습니다.";
      await loadImportShipments();
    }

    function collectVehiclePayload() {
      const items = [...productTable.querySelectorAll(".product-row")].map((row) => ({
        product_name: row.querySelector(".product-name").value.trim(),
        quantity: row.querySelector(".product-quantity").value.trim(),
        pack_quantity: row.querySelector(".product-pack-quantity").value.trim(),
      })).filter((item) => item.product_name || item.quantity || item.pack_quantity);
      return {
        receipt_type: receiptTypeSelect.value,
        supplier: supplierInput.value.trim(),
        receipt_date: receiptDateInput.value,
        freight_payment: freightPaymentSelect.value,
        request_note: requestNoteInput.value.trim(),
        delivery_place: deliveryPlaceInput.value.trim(),
        manager: managerInput.value.trim(),
        items,
      };
    }

    function defaultCsSubject(vendorName = "") {
      return `[CS 요청] ${vendorName ? `${vendorName} ` : ""}확인 부탁드립니다`;
    }

    function defaultCsBody() {
      return `안녕하세요. (주)소일브릿지 입니다.\n\n\n\n- 원출고일  및 원송장번호 : ${csOriginInput.value.trim()}\n\n\n\n- 상품명 : ${csProductInput.value.trim()} \n\n- 수령인 : ${csReceiverInput.value.trim()}\n\n- 수령인 연락처 : ${csPhoneInput.value.trim()}\n\n- 수령인 주소 : ${csAddressInput.value.trim()}\n\n\n\n- CS내용 : ${csContentInput.value.trim()}\n\n \n\nCS건을 보내드립니다.\n\n \n\n★반품 접수 후 일주일 이상 회신 없으실 경우 자체 환불 및 정산 반영 예정이오니, 처리 결과 꼭 회신 부탁드립니다.\n\n `;
    }

    function refreshCsBody() {
      csBodyInput.value = defaultCsBody();
    }

    function refreshStockNoticeBody() {
      if (stockBodyInput) stockBodyInput.value = defaultStockNoticeBody();
    }

    function defaultStockContactInfo() {
      return {
        managerName: cachedMailSettings.stock_manager_name || currentUser.display_name || currentUser.username || "",
        managerPhone: cachedMailSettings.stock_manager_phone || "",
        senderEmail: cachedMailSettings.naver_email || "",
      };
    }

    function defaultStockNoticeBody() {
      const value = (input) => input?.value.trim() || "";
      const contact = defaultStockContactInfo();
      return `안녕하세요. (주)소일브릿지 입니다.

제품 입고 및 품절 현황 안내드립니다.

■ 기준일자: ${stockNoticeDateInput?.value || ""}

■ 제품 입고 안내

▶품명(모델명) : ${value(stockInboundProductInput)}

▶입고 일정 : ${value(stockInboundScheduleInput)}

▶출고 가능 일정 : ${value(stockOutboundAvailableInput)}

▶특이사항 : ${value(stockInboundNoteInput)}

■ 제품 일시 품절(단종) 안내

▶품명(모델명) : ${value(stockSoldoutProductInput)}

▶출고 불가 일정 : ${value(stockOutboundBlockedInput)}

▶재입고 일정 : ${value(stockRestockScheduleInput)}

▶특이사항 : ${value(stockSoldoutNoteInput)}

업무 진행 시 참고 부탁드리며, 확인이 필요한 내용이 있으시면 회신 부탁드립니다.

감사합니다.

(주)소일브릿지
담당자: ${contact.managerName}
연락처: ${contact.managerPhone}
이메일: ${contact.senderEmail}`;
    }

    function handleLoginRequiredResponse(response, messageTarget) {
      if (response.status === 401) {
        if (messageTarget) {
          messageTarget.textContent = "로그인이 만료되었습니다. 다시 로그인해주세요.";
        }
        window.setTimeout(() => {
          window.location.href = "/login";
        }, 700);
        return true;
      }
      return false;
    }

    function backupSettingsPayload() {
      return {
        backup_dir: backupDirInput?.value?.trim() || "",
        auto_enabled: Boolean(backupAutoEnabled?.checked),
        auto_hour: Number(backupAutoHour?.value || 3),
        retention_days: Number(backupRetentionDays?.value || 90),
        external_enabled: Boolean(backupExternalEnabled?.checked),
        external_type: "rclone",
        rclone_executable: backupRcloneExecutable?.value?.trim() || "rclone",
        rclone_remote: backupRcloneRemote?.value?.trim() || "",
        rclone_path: backupRclonePath?.value?.trim() || "",
      };
    }

    async function saveBackupSettings() {
      if (!backupSettingsSave) return;
      backupSettingsSave.disabled = true;
      backupMessage.textContent = "백업 설정을 저장하는 중입니다.";
      try {
        const response = await fetch("/api/backup-settings", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(backupSettingsPayload()),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "백업 설정 저장에 실패했습니다.");
        backupMessage.textContent = data.message || "백업 설정을 저장했습니다.";
        await loadBackups();
      } catch (error) {
        backupMessage.textContent = error.message;
      } finally {
        backupSettingsSave.disabled = false;
      }
    }

    async function createBackupAtSelectedPath() {
      if (!backupCreateSelected) return;
      const backupDir = backupDirInput?.value?.trim() || "";
      if (!backupDir) {
        backupMessage.textContent = "지정 백업을 만들 폴더 경로를 입력해주세요.";
        return;
      }
      backupCreateSelected.disabled = true;
      backupMessage.textContent = "지정 위치로 백업 파일을 생성하는 중입니다.";
      try {
        const response = await fetch("/api/backup-create", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ backup_dir: backupDir }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "지정 위치 백업 생성에 실패했습니다.");
        const external = data.backup?.external_backup;
        backupMessage.textContent = external && external.status !== "disabled"
          ? `${data.message || "지정 위치에 백업 파일을 생성했습니다."} Google Drive 업로드: ${external.status}`
          : data.message || "지정 위치에 백업 파일을 생성했습니다.";
        await loadBackups();
      } catch (error) {
        backupMessage.textContent = error.message;
      } finally {
        backupCreateSelected.disabled = false;
      }
    }

    async function loadMailSettings() {
      try {
        const response = await fetch("/api/mail-settings", { credentials: "same-origin" });
        if (handleLoginRequiredResponse(response, adminMailSettingsMessage || adminMailTechnicalMessage)) return {};
        if (!response.ok) return;
        const data = await response.json();
        cachedMailSettings = data || {};
        if (adminNaverEmailInput) adminNaverEmailInput.value = data.naver_email || "";
        if (adminNaverPasswordInput) {
          adminNaverPasswordInput.value = "";
          adminNaverPasswordInput.placeholder = data.has_password ? "저장된 비밀번호 사용" : "저장된 비밀번호가 없으면 입력";
        }
        if (adminSmtpPort) adminSmtpPort.value = String(data.smtp_port || 465);
        if (adminSmtpSecurity) adminSmtpSecurity.value = data.smtp_security || "ssl";
        if (adminBulkBatchSize) adminBulkBatchSize.value = data.bulk_batch_size || 20;
        if (adminBulkSendInterval) adminBulkSendInterval.value = data.bulk_send_interval_seconds || 15;
        if (adminBulkBatchPause) adminBulkBatchPause.value = data.bulk_batch_pause_minutes ?? 5;
        if (adminBulkTestRecipient) adminBulkTestRecipient.value = data.bulk_test_recipient || "";
        return data;
      } catch {
        // 저장된 메일 설정이 없어도 다른 화면 사용은 계속 가능합니다.
        cachedMailSettings = {};
      }
    }

    async function loadAdminMailSettings() {
      if (!adminNaverEmailInput) return;
      adminMailSettingsMessage.textContent = "메일 기본정보를 불러오는 중입니다.";
      const data = await loadMailSettings();
      adminMailSettingsMessage.textContent = data?.has_password ? "저장된 네이버 메일 비밀번호가 있습니다." : "저장된 네이버 메일 비밀번호가 없습니다.";
    }

    async function saveAdminMailSettings() {
      if (!adminMailSettingsSave) return;
      const payload = {
        naver_email: adminNaverEmailInput.value.trim(),
        naver_password: adminNaverPasswordInput.value,
        save_credentials: adminSaveMailCredentials.checked,
      };
      if (!payload.naver_email) {
        adminMailSettingsMessage.textContent = "네이버 메일 아이디를 입력해주세요.";
        return;
      }
      adminMailSettingsSave.disabled = true;
      adminMailSettingsMessage.textContent = "메일 기본정보를 저장하는 중입니다.";
      try {
        const response = await fetch("/api/mail-settings", {
          method: "POST",
          credentials: "same-origin",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        if (handleLoginRequiredResponse(response, adminMailSettingsMessage)) return;
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "메일 기본정보 저장에 실패했습니다.");
        adminNaverPasswordInput.value = "";
        adminNaverPasswordInput.placeholder = data.has_password ? "저장된 비밀번호 사용" : "저장된 비밀번호가 없으면 입력";
        adminMailSettingsMessage.textContent = data.message || "메일 기본정보를 저장했습니다.";
      } catch (error) {
        adminMailSettingsMessage.textContent = error.message;
      } finally {
        adminMailSettingsSave.disabled = false;
      }
    }

    function adminMailTechnicalPayload() {
      const smtpPort = adminSmtpPort?.value || "465";
      return {
        naver_email: adminNaverEmailInput?.value.trim() || "",
        naver_password: "",
        save_credentials: false,
        smtp_port: smtpPort,
        smtp_security: adminSmtpSecurity?.value || (smtpPort === "587" ? "tls" : "ssl"),
        bulk_batch_size: adminBulkBatchSize?.value || "20",
        bulk_send_interval_seconds: adminBulkSendInterval?.value || "15",
        bulk_batch_pause_minutes: adminBulkBatchPause?.value || "5",
        bulk_test_recipient: adminBulkTestRecipient?.value.trim() || "",
      };
    }

    async function saveAdminMailTechnicalSettings() {
      if (!adminMailTechnicalSave) return;
      const payload = adminMailTechnicalPayload();
      if (!payload.naver_email) {
        adminMailTechnicalMessage.textContent = "네이버 메일 아이디를 먼저 입력해주세요.";
        return;
      }
      adminMailTechnicalSave.disabled = true;
      adminMailTechnicalMessage.textContent = "단체메일 기술 설정을 저장하는 중입니다.";
      try {
        const response = await fetch("/api/mail-settings", {
          method: "POST",
          credentials: "same-origin",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        if (handleLoginRequiredResponse(response, adminMailTechnicalMessage)) return;
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "단체메일 기술 설정 저장에 실패했습니다.");
        if (adminSmtpPort) adminSmtpPort.value = String(data.smtp_port || payload.smtp_port);
        if (adminSmtpSecurity) adminSmtpSecurity.value = data.smtp_security || payload.smtp_security;
        adminMailTechnicalMessage.textContent = "단체메일 기술 설정을 저장했습니다.";
      } catch (error) {
        adminMailTechnicalMessage.textContent = error.message;
      } finally {
        adminMailTechnicalSave.disabled = false;
      }
    }

    async function sendAdminMailTestMessage() {
      if (!adminMailTestSend) return;
      const recipient = adminBulkTestRecipient?.value.trim() || "";
      if (!recipient) {
        adminMailTechnicalMessage.textContent = "테스트 수신 메일 주소를 입력해주세요.";
        return;
      }
      adminMailTestSend.disabled = true;
      adminMailTechnicalMessage.textContent = "테스트 메일을 발송하는 중입니다.";
      try {
        const response = await fetch("/api/mail-test", {
          method: "POST",
          credentials: "same-origin",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ recipient_email: recipient }),
        });
        if (handleLoginRequiredResponse(response, adminMailTechnicalMessage)) return;
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "테스트 메일 발송에 실패했습니다.");
        adminMailTechnicalMessage.textContent = data.message || "테스트 메일을 발송했습니다.";
      } catch (error) {
        adminMailTechnicalMessage.textContent = error.message;
      } finally {
        adminMailTestSend.disabled = false;
      }
    }

    function renderVendorContacts() {
      vendorContactSelect.innerHTML = "";
      const emptyOption = document.createElement("option");
      emptyOption.value = "";
      emptyOption.textContent = vendorContacts.length ? "업체를 선택해주세요" : "저장된 업체가 없습니다";
      vendorContactSelect.appendChild(emptyOption);

      vendorContacts.forEach((contact) => {
        const option = document.createElement("option");
        option.value = `${contact.vendor_type || "purchase"}::${contact.vendor_name}`;
        option.textContent = `[${contact.vendor_type_label || "매입처"}] ${contact.vendor_name} / ${contact.email}`;
        vendorContactSelect.appendChild(option);
      });
      renderStockVendorContacts();
    }

    function renderStockVendorContacts() {
      if (!stockVendorTree) return;
      if (!vendorContacts.length) {
        stockVendorTree.innerHTML = '<div class="vendor-picker-empty">저장된 업체 메일 리스트가 없습니다.</div>';
        return;
      }
      const groups = [
        ["purchase", "매입처"],
        ["sales", "매출처"],
      ];
      stockVendorTree.innerHTML = groups.map(([type, label]) => {
        const items = vendorContacts.filter((contact) => (contact.vendor_type || "purchase") === type);
        if (!items.length) return "";
        return `
          <div class="vendor-picker-group">
            <div class="vendor-picker-group-title">${label}</div>
            ${items.map((contact) => `
              <button class="vendor-picker-option" type="button" data-stock-vendor-type="${escapeHtml(contact.vendor_type || "purchase")}" data-stock-vendor-name="${escapeHtml(contact.vendor_name)}">
                ${escapeHtml(contact.vendor_name)} / ${escapeHtml(contact.email)}
              </button>
            `).join("")}
          </div>
        `;
      }).join("") || '<div class="vendor-picker-empty">저장된 업체 메일 리스트가 없습니다.</div>';
    }

    async function loadVendorContacts() {
      try {
        const response = await fetch("/api/vendor-contacts");
        if (!response.ok) return;
        const data = await response.json();
        vendorContacts = data.contacts || [];
        renderVendorContacts();
      } catch {
        vendorContacts = [];
        renderVendorContacts();
      }
    }

    async function saveCompanyHolidayFromLeave() {
      leaveMessage.textContent = "\uD734\uC77C \uC815\uBCF4\uB97C \uC800\uC7A5\uD558\uB294 \uC911\uC785\uB2C8\uB2E4.";
      try {
        const response = await fetch("/api/company-holiday", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({
            holiday_date: leaveHolidayDateInput.value,
            name: leaveHolidayNameInput.value.trim(),
            is_substitute: Boolean(leaveHolidaySubstituteInput.checked),
          }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "\uD734\uC77C \uC800\uC7A5\uC5D0 \uC2E4\uD328\uD588\uC2B5\uB2C8\uB2E4.");
        leaveHolidayNameInput.value = "";
        leaveHolidaySubstituteInput.checked = false;
        leaveMessage.textContent = data.message || "\uD734\uC77C\uC744 \uC800\uC7A5\uD588\uC2B5\uB2C8\uB2E4.";
      } catch (error) {
        leaveMessage.textContent = error.message;
      }
    }

    function applySelectedVendor() {
      const [selectedType, selectedName] = vendorContactSelect.value.split("::");
      const selected = vendorContacts.find((contact) => (
        (contact.vendor_type || "purchase") === selectedType && contact.vendor_name === selectedName
      ));
      if (!selected) return;
      vendorTypeSelect.value = selected.vendor_type || "purchase";
      vendorNameInput.value = selected.vendor_name;
      recipientEmailInput.value = selected.email;
      csSubjectInput.value = currentMode === "mail-stock" ? "입고 및 품절 공지" : defaultCsSubject(selected.vendor_name);
    }

    function setSelectedStockVendor(selected) {
      if (!selected) {
        stockVendorTypeSelect.value = "purchase";
        stockVendorNameInput.value = "";
        stockRecipientEmailInput.value = "";
        if (stockVendorPickerButton) stockVendorPickerButton.textContent = "업체를 선택해주세요";
        if (stockSelectedVendorLabel) stockSelectedVendorLabel.textContent = "선택된 업체 없음";
        return;
      }
      stockVendorTypeSelect.value = selected.vendor_type || "purchase";
      stockVendorNameInput.value = selected.vendor_name;
      stockRecipientEmailInput.value = selected.email;
      if (stockVendorPickerButton) stockVendorPickerButton.textContent = selected.vendor_name;
      if (stockSelectedVendorLabel) {
        const typeLabel = selected.vendor_type_label || (selected.vendor_type === "sales" ? "매출처" : "매입처");
        stockSelectedVendorLabel.textContent = `[${typeLabel}] ${selected.vendor_name} / ${selected.email}`;
      }
    }

    function toggleStockVendorTree() {
      if (!stockVendorTree) return;
      stockVendorTree.hidden = !stockVendorTree.hidden;
    }

    function applySelectedStockVendorFromTree(button) {
      const selectedType = button?.dataset.stockVendorType || "";
      const selectedName = button?.dataset.stockVendorName || "";
      const selected = vendorContacts.find((contact) => (
        (contact.vendor_type || "purchase") === selectedType && contact.vendor_name === selectedName
      ));
      if (!selected) return;
      setSelectedStockVendor(selected);
      if (stockVendorTree) stockVendorTree.hidden = true;
    }

    function applySelectedStockVendor() {
      const selected = vendorContacts.find((contact) => (
        (contact.vendor_type || "purchase") === stockVendorTypeSelect.value && contact.vendor_name === stockVendorNameInput.value
      ));
      if (selected) setSelectedStockVendor(selected);
    }

    async function saveCurrentVendorContact() {
      const vendorName = vendorNameInput.value.trim();
      const email = recipientEmailInput.value.trim();
      const vendorType = vendorTypeSelect.value || "purchase";
      if (!vendorName || !email) {
        notice.textContent = "업체명과 받는 업체 메일을 입력해주세요.";
        return;
      }
      try {
        saveVendorContactButton.disabled = true;
        const response = await fetch("/api/vendor-contact", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ vendor_type: vendorType, vendor_name: vendorName, email }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "업체 메일 저장에 실패했습니다.");
        vendorContacts = data.contacts || [];
        renderVendorContacts();
        vendorContactSelect.value = `${vendorType}::${vendorName}`;
        notice.textContent = "업체 메일 주소를 저장했습니다.";
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        saveVendorContactButton.disabled = false;
      }
    }

    async function uploadVendorContactsWorkbook() {
      if (!vendorContactsFileInput) return;
      const file = vendorContactsFileInput.files[0];
      if (!file) return;
      if (vendorContactsDropMain) vendorContactsDropMain.textContent = file.name;
      const formData = new FormData();
      formData.append("file", file);
      notice.textContent = "업체 메일 주소록을 저장 중입니다.";
      try {
        const response = await fetch("/api/vendor-contacts-import", {
          method: "POST",
          body: formData,
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "업체 메일 주소록 저장에 실패했습니다.");
        vendorContacts = data.contacts || [];
        renderVendorContacts();
        notice.textContent = data.message || "업체 메일 주소록을 저장했습니다.";
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        vendorContactsFileInput.value = "";
      }
    }

    function renderSalesReportUploads(files = []) {
      if (!salesReportRecentList) return;
      const recent = files.slice(0, 5);
      if (!recent.length) {
        salesReportRecentList.textContent = "업로드된 매출표가 없습니다.";
        return;
      }
      salesReportRecentList.innerHTML = recent
        .map((file) => `${escapeHtml(file.original_name || "")} · ${escapeHtml(file.uploaded_at || "")}`)
        .join("<br />");
    }

    async function loadSalesReportUploads() {
      if (!salesReportRecentList) return;
      try {
        const response = await fetch("/api/sales-report-uploads");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "매출표 업로드 목록을 불러오지 못했습니다.");
        renderSalesReportUploads(data.files || []);
      } catch (error) {
        salesReportRecentList.textContent = error.message;
      }
    }

    function formatSalesNumber(value) {
      const number = Number(value || 0);
      return number.toLocaleString("ko-KR");
    }

    function formatSalesPercent(value) {
      const number = Number(value || 0);
      return `${number.toLocaleString("ko-KR", { maximumFractionDigits: 1 })}%`;
    }

    function salesAmountClass(value) {
      const number = Number(value || 0);
      if (number > 0) return "sales-positive";
      if (number < 0) return "sales-negative";
      return "";
    }

    function renderSalesEmpty(tbody, colspan, message) {
      if (!tbody) return;
      tbody.innerHTML = `<tr><td class="empty" colspan="${colspan}">${escapeHtml(message)}</td></tr>`;
    }

    function renderSalesReportDashboard(data) {
      if (!salesReportKpiGrid) return;
      const today = data.today || {};
      const yesterday = data.yesterday || {};
      const comparison = data.comparison || {};
      const month = data.month || {};
      const sellerTotal = data.seller_total || {};
      const consistency = data.consistency || {};
      salesReportKpiGrid.innerHTML = [
        ["오늘 손익 매출", formatSalesNumber(today.profit_sales_amount), `수량 ${formatSalesNumber(today.quantity)}`, "primary"],
        ["어제 손익 매출", formatSalesNumber(yesterday.profit_sales_amount), `수량 ${formatSalesNumber(yesterday.quantity)}`, ""],
        ["전일 대비", formatSalesPercent(comparison.profit_sales_amount_delta_rate), formatSalesNumber(comparison.profit_sales_amount_delta), Number(comparison.profit_sales_amount_delta || 0) < 0 ? "danger" : ""],
        ["월 누적 매출", formatSalesNumber(month.profit_sales_amount), data.period || "", ""],
        ["매출처별 합계", formatSalesNumber(sellerTotal.profit_sales_amount), `판매사 수량 ${formatSalesNumber(sellerTotal.quantity)}`, ""],
        ["파일 검증", formatSalesNumber(consistency.difference), consistency.ok ? "차이 없음" : "확인 필요", consistency.ok ? "" : "warning"],
      ].map(([label, value, note, variant]) => `
        <div class="sales-kpi ${variant}">
          <div class="sales-kpi-label">${escapeHtml(label)}</div>
          <div class="sales-kpi-value ${label === "전일 대비" ? salesAmountClass(comparison.profit_sales_amount_delta) : ""}">${escapeHtml(value)}</div>
          <div class="sales-kpi-note">${escapeHtml(note)}</div>
        </div>
      `).join("");

      const dailyRows = data.daily_rows || [];
      if (dailyRows.length) {
        salesReportDailyBody.innerHTML = dailyRows.map((row) => `
          <tr>
            <td>${escapeHtml(row.label || row.report_date || "")}</td>
            <td>${formatSalesNumber(row.quantity)}</td>
            <td>${formatSalesNumber(row.profit_sales_amount)}</td>
            <td>${formatSalesNumber(row.sales_total)}</td>
            <td class="${salesAmountClass(row.profit_margin)}">${formatSalesNumber(row.profit_margin)}</td>
          </tr>
        `).join("");
      } else {
        renderSalesEmpty(salesReportDailyBody, 5, "날짜별 매출 통계 파일을 업로드해주세요.");
      }

      const sellerRows = data.seller_top || [];
      if (sellerRows.length) {
        salesReportSellerBody.innerHTML = sellerRows.map((row) => `
          <tr>
            <td>${escapeHtml(row.name || "")}</td>
            <td>${formatSalesNumber(row.quantity)}</td>
            <td>${formatSalesNumber(row.profit_sales_amount)}</td>
            <td class="${salesAmountClass(row.profit_margin)}">${formatSalesNumber(row.profit_margin)}</td>
          </tr>
        `).join("");
      } else {
        renderSalesEmpty(salesReportSellerBody, 4, "매출처별 파일을 업로드해주세요.");
      }

      const productRows = data.product_top || [];
      if (productRows.length) {
        salesReportProductBody.innerHTML = productRows.map((row) => `
          <tr>
            <td>${escapeHtml(row.name || "")}</td>
            <td>${formatSalesNumber(row.quantity)}</td>
            <td>${formatSalesNumber(row.profit_sales_amount)}</td>
            <td class="${salesAmountClass(row.profit_margin)}">${formatSalesNumber(row.profit_margin)}</td>
          </tr>
        `).join("");
      } else {
        renderSalesEmpty(salesReportProductBody, 4, "상품별 Statistics_Good 파일을 업로드해주세요.");
      }

      const purchaseRows = data.supplier_purchase_totals || [];
      if (purchaseRows.length) {
        salesReportReviewBody.innerHTML = purchaseRows.slice(0, 10).map((row) => `
          <tr>
            <td>${escapeHtml(row.name || "")}</td>
            <td>${formatSalesNumber(row.purchase_total)}</td>
            <td>${formatSalesNumber(row.quantity)}</td>
          </tr>
        `).join("");
      } else {
        renderSalesEmpty(salesReportReviewBody, 3, "공급사별 매입금액 파일을 업로드해주세요.");
      }
    }

    async function loadSalesReportDashboard() {
      if (!salesReportKpiGrid) return;
      try {
        const response = await fetch("/api/sales-report-dashboard");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "매출현황을 불러오지 못했습니다.");
        renderSalesReportDashboard(data);
      } catch (error) {
        salesReportKpiGrid.innerHTML = `<div class="admin-message">${escapeHtml(error.message)}</div>`;
        renderSalesEmpty(salesReportDailyBody, 5, "매출현황을 불러오지 못했습니다.");
        renderSalesEmpty(salesReportSellerBody, 4, "매출현황을 불러오지 못했습니다.");
        renderSalesEmpty(salesReportProductBody, 4, "매출현황을 불러오지 못했습니다.");
        renderSalesEmpty(salesReportReviewBody, 3, "매출현황을 불러오지 못했습니다.");
      }
    }

    async function uploadSalesReportWorkbook() {
      if (!salesReportFileInput) return;
      const file = salesReportFileInput.files[0];
      if (!file) return;
      const formData = new FormData();
      formData.append("file", file);
      if (salesReportUploadMessage) salesReportUploadMessage.textContent = "매출표를 업로드하는 중입니다.";
      try {
        const response = await fetch("/api/sales-report-upload", {
          method: "POST",
          body: formData,
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "매출표 업로드에 실패했습니다.");
        renderSalesReportUploads(data.files || []);
        loadSalesReportDashboard();
        if (salesReportUploadMessage) salesReportUploadMessage.textContent = data.message || "매출표를 저장했습니다.";
      } catch (error) {
        if (salesReportUploadMessage) salesReportUploadMessage.textContent = error.message;
      } finally {
        salesReportFileInput.value = "";
      }
    }

    function openSalesReportUploadPicker() {
      if (!salesReportFileInput || !can("sales_report_manage")) return;
      salesReportFileInput.click();
    }

    function collectCsPayload() {
      return {
        case_id: activeCsCaseId,
        vendor_type: vendorTypeSelect.value || "purchase",
        recipient_email: recipientEmailInput.value.trim(),
        vendor_name: vendorNameInput.value.trim(),
        cs_origin: csOriginInput.value.trim(),
        cs_product: csProductInput.value.trim(),
        cs_receiver: csReceiverInput.value.trim(),
        cs_phone: csPhoneInput.value.trim(),
        cs_address: csAddressInput.value.trim(),
        cs_type: csTypeInput.value.trim(),
        cs_content: csContentInput.value.trim(),
        subject: csSubjectInput.value.trim(),
        body: csBodyInput.value.trim(),
      };
    }

    function updateCsAttachmentSummary() {
      if (!csAttachmentInput || !csAttachmentSummary) return;
      const files = Array.from(csAttachmentInput.files || []);
      if (!files.length) {
        csAttachmentSummary.textContent = "첨부파일 없음";
        return;
      }
      const totalSize = files.reduce((sum, file) => sum + file.size, 0);
      const sizeMb = Math.ceil((totalSize / 1024 / 1024) * 10) / 10;
      csAttachmentSummary.textContent = `${files.length}개 첨부 선택 / 약 ${sizeMb}MB`;
    }

    function appendCsMailPayload(formData, payload) {
      formData.append("payload", JSON.stringify(payload));
      Array.from(csAttachmentInput?.files || []).forEach((file, index) => {
        formData.append(`cs_attachment_${index + 1}`, file, file.name);
      });
      return formData;
    }

    function collectStockNoticePayload() {
      return {
        vendor_type: stockVendorTypeSelect?.value || "purchase",
        recipient_email: stockRecipientEmailInput?.value.trim() || "",
        vendor_name: stockVendorNameInput?.value.trim() || "",
        subject: stockSubjectInput?.value.trim() || "",
        body: stockBodyInput?.value.trim() || "",
        save_credentials: true,
      };
    }

    function renderCsCases(cases) {
      csCaseList.innerHTML = "";
      if (!cases || cases.length === 0) {
        const empty = document.createElement("div");
        empty.className = "cs-case-item";
        empty.textContent = "저장된 CS건이 없습니다.";
        csCaseList.appendChild(empty);
        return;
      }
      cases.forEach((csCase) => {
        const item = document.createElement("div");
        item.className = "cs-case-item";
        const title = [
          `#${csCase.id}`,
          csCase.status || "접수",
          csCase.vendor_name || "업체 미입력",
          csCase.product_name || "상품 미입력",
        ].join(" · ");
        const meta = [
          csCase.receiver_name || "수령인 미입력",
          csCase.receiver_phone || "",
          csCase.original_invoice ? `원송장 ${csCase.original_invoice}` : "",
          csCase.mail_sent_at ? `메일 ${csCase.mail_sent_at}` : "",
        ].filter(Boolean).join(" / ");
        item.innerHTML = `<strong>${title}</strong><div class="cs-case-meta">${meta}</div>`;
        csCaseList.appendChild(item);
      });
    }

    function escapeHtml(value) {
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");
    }

    function crmStatusClass(status) {
      if (status === "완료") return "done";
      if (status === "진행중") return "progress";
      if (status === "보류") return "hold";
      return "wait";
    }

    function crmStatusBadge(status) {
      const label = status || "대기";
      return `<span class="crm-status badge badge-sm ${crmStatusClass(label)}">${escapeHtml(label)}</span>`;
    }

    function setCrmMessage(text, isError = false) {
      if (!crmMessage) return;
      crmMessage.textContent = text || "";
      crmMessage.classList.toggle("open", Boolean(text));
      crmMessage.classList.toggle("error", Boolean(isError));
    }

    async function copyCrmText(text, label = "값") {
      const value = String(text || "").trim();
      if (!value) {
        setCrmMessage(`${label}이 비어 있습니다.`, true);
        return;
      }
      try {
        await navigator.clipboard.writeText(value);
        setCrmMessage(`${label}을 복사했습니다.`);
      } catch (error) {
        setCrmMessage(`${label}: ${value}`);
      }
    }

    async function crmFetchJson(url, options = {}) {
      const response = await fetch(url, options);
      const data = await response.json().catch(() => ({}));
      if (!response.ok) throw new Error(data.error || "CRM 요청 처리에 실패했습니다.");
      return data;
    }

    function syncCompanyNavState() {
      companyTabs.forEach((button) => button.classList.toggle("active", button.dataset.companyTab === companyActiveTab));
      companyNavTabs.forEach((button) => button.classList.toggle("active", button.dataset.companyTab === companyActiveTab));
    }

    function setCompanyTab(tabName) {
      companyActiveTab = tabName || "notice";
      companyPanels.forEach((panel) => {
        const isActivePanel = panel.dataset.companyPanel === companyActiveTab
          || (companyActiveTab === "notice" && panel.dataset.companyPanel === "calendar");
        panel.classList.toggle("active", isActivePanel);
      });
      syncCompanyNavState();
      if (companyActiveTab === "staff") {
        loadCompanyStaffDashboard().catch(() => {});
      } else if (companyActiveTab === "notice") {
        loadDashboardEntryData().catch(() => {});
      } else if (companyActiveTab === "calendar") {
        loadCompanyCalendar().catch(() => {
          if (companyCalendarGrid) companyCalendarGrid.innerHTML = `<div class="calendar-empty">캘린더를 불러오지 못했습니다.</div>`;
        });
      } else if (companyActiveTab === "chat") {
        loadInternalChatUsers()
          .then(() => loadInternalMessages())
          .catch(() => {
          if (internalChatList) internalChatList.innerHTML = `<div class="internal-chat-empty">메시지를 불러오지 못했습니다.</div>`;
        });
      }
    }

    function parseLocalDate(value) {
      const match = String(value || "").match(/^(\d{4})-(\d{2})-(\d{2})$/);
      if (!match) return null;
      return new Date(Number(match[1]), Number(match[2]) - 1, Number(match[3]));
    }

    function localDateString(dateValue) {
      const year = dateValue.getFullYear();
      const month = String(dateValue.getMonth() + 1).padStart(2, "0");
      const day = String(dateValue.getDate()).padStart(2, "0");
      return `${year}-${month}-${day}`;
    }

    function monthTitle(monthText) {
      const [year, month] = String(monthText || todayString().slice(0, 7)).split("-");
      return `${year}년 ${Number(month)}월`;
    }

    function shiftCalendarMonth(delta) {
      const base = parseLocalDate(`${companyCalendarMonth}-01`) || parseLocalDate(`${todayString().slice(0, 7)}-01`);
      base.setMonth(base.getMonth() + delta);
      companyCalendarMonth = localDateString(base).slice(0, 7);
      companyCalendarSelectedDay = localDateString(base);
      return loadCompanyCalendar();
    }

    function calendarEventLabel(event) {
      if (event.type === "project") return `프로젝트 · ${event.subtitle || ""}`;
      if (event.type === "leave") return `연차 · ${event.subtitle || ""}`;
      if (event.type === "pending") return `승인대기 · ${event.subtitle || ""}`;
      return `업무 · ${event.subtitle || ""}`;
    }

    function eventsForDay(dayText) {
      return companyCalendarEvents.filter((event) => event.date === dayText);
    }

    function renderCalendarEvent(event, compact = true) {
      const title = compact ? event.title : `${event.title} · ${event.subtitle || ""}`;
      return `
        <button class="calendar-event ${escapeHtml(event.type || "task")}" type="button" data-calendar-event-id="${escapeHtml(event.id)}" aria-label="${escapeHtml(calendarEventLabel(event))}">
          ${escapeHtml(title || "일정")}
        </button>
      `;
    }

    function renderCompanyCalendarSelectedDay() {
      if (!companyCalendarSelectedDate || !companyCalendarSelectedList) return;
      const selected = parseLocalDate(companyCalendarSelectedDay);
      companyCalendarSelectedDate.textContent = selected ? shortKoreanDate(companyCalendarSelectedDay) : companyCalendarSelectedDay;
      const items = eventsForDay(companyCalendarSelectedDay);
      if (!items.length) {
        companyCalendarSelectedList.innerHTML = `<div class="calendar-empty">선택한 날짜에 표시할 일정이 없습니다.</div>`;
        return;
      }
      companyCalendarSelectedList.innerHTML = items.map((event) => `
        <div class="company-task-item" role="button" tabindex="0" data-calendar-event-id="${escapeHtml(event.id)}" aria-label="${escapeHtml(calendarEventLabel(event))}">
          <div>
            <div class="company-task-title">${escapeHtml(event.title || "일정")}</div>
            <div class="company-task-meta">${escapeHtml(calendarEventLabel(event))}</div>
          </div>
          <span class="calendar-event ${escapeHtml(event.type || "task")}">${escapeHtml(event.type === "project" ? "프로젝트" : event.type === "task" ? "업무" : event.type === "pending" ? "대기" : "연차")}</span>
        </div>
      `).join("");
    }

    function renderCompanyCalendar(payload = {}) {
      if (!companyCalendarGrid) return;
      companyCalendarMonth = payload.month || companyCalendarMonth;
      companyCalendarEvents = payload.events || [];
      companyCalendarSummary = payload.summary || companyCalendarSummary || {};
      const monthStart = parseLocalDate(`${companyCalendarMonth}-01`) || parseLocalDate(`${todayString().slice(0, 7)}-01`);
      const gridStart = new Date(monthStart);
      gridStart.setDate(1 - ((monthStart.getDay() + 6) % 7));
      const today = todayString();
      if (companyCalendarTitle) companyCalendarTitle.textContent = monthTitle(companyCalendarMonth);
      const summary = companyCalendarSummary;
      if (companyCalendarProjectCount) companyCalendarProjectCount.textContent = `${summary.project || 0}건`;
      if (companyCalendarTaskCount) companyCalendarTaskCount.textContent = `${summary.task || 0}건`;
      if (companyCalendarLeaveCount) companyCalendarLeaveCount.textContent = `${summary.leave || 0}건`;
      if (companyCalendarRiskCount) companyCalendarRiskCount.textContent = `${summary.risk || 0}건`;
      const cells = [];
      for (let index = 0; index < 42; index += 1) {
        const day = new Date(gridStart);
        day.setDate(gridStart.getDate() + index);
        const dayText = localDateString(day);
        const events = eventsForDay(dayText);
        const visibleEvents = events.slice(0, 4);
        const classes = [
          "calendar-day",
          day.getMonth() === monthStart.getMonth() ? "" : "other-month",
          dayText === today ? "today" : "",
          dayText === companyCalendarSelectedDay ? "selected" : "",
        ].filter(Boolean).join(" ");
        cells.push(`
          <div class="${classes}" role="gridcell" tabindex="0" data-calendar-day="${escapeHtml(dayText)}" aria-label="${escapeHtml(`${shortKoreanDate(dayText)} 일정 ${events.length}건`)}">
            <span class="calendar-date">
              <span>${escapeHtml(day.getDate())}</span>
              ${events.length ? `<span class="calendar-date-count">${escapeHtml(events.length)}</span>` : ""}
            </span>
            <span class="calendar-event-list">
              ${visibleEvents.map((event) => renderCalendarEvent(event)).join("")}
              ${events.length > visibleEvents.length ? `<span class="calendar-more">+${escapeHtml(events.length - visibleEvents.length)} 더보기</span>` : ""}
            </span>
          </div>
        `);
      }
      companyCalendarGrid.innerHTML = cells.join("");
      renderCompanyCalendarSelectedDay();
    }

    async function loadCompanyCalendar() {
      if (!companyCalendarGrid) return;
      if (!can("crm_view")) {
        companyCalendarGrid.innerHTML = `<div class="calendar-empty">CRM 조회 권한이 없어 캘린더를 볼 수 없습니다.</div>`;
        return;
      }
      companyCalendarGrid.innerHTML = `<div class="calendar-empty">캘린더를 불러오는 중입니다.</div>`;
      const data = await crmFetchJson(`/api/company-calendar-events?month=${encodeURIComponent(companyCalendarMonth)}`);
      renderCompanyCalendar(data);
    }

    async function loadDashboardEntryData() {
      const tasks = [loadImportShipments()];
      tasks.push(loadCompanyCalendar().catch(() => {
        if (companyCalendarGrid) companyCalendarGrid.innerHTML = `<div class="calendar-empty">캘린더를 불러오지 못했습니다.</div>`;
      }));
      await Promise.allSettled(tasks);
    }

    function openCalendarEventWidget(eventId) {
      const event = companyCalendarEvents.find((item) => String(item.id) === String(eventId));
      if (!event) return;
      if (event.type === "task" && event.task_id) {
        openCrmTaskWidget(event.task_id).catch((error) => setCrmMessage(error.message, true));
        return;
      }
      const stateText = event.type === "project"
        ? `${event.progress || 0}% 완료`
        : event.type === "pending"
          ? "승인대기"
          : event.status || "승인";
      openFocusWidget({
        kicker: event.type === "project" ? "프로젝트 일정" : event.type === "pending" ? "승인대기 연차" : "연차 일정",
        title: event.title || "일정",
        subtitle: [shortKoreanDate(event.date), event.subtitle].filter(Boolean).join(" · "),
        body: `
          <div class="focus-widget-grid">
            ${focusWidgetMetric("구분", event.type === "project" ? "프로젝트" : event.type === "pending" ? "승인대기" : "연차")}
            ${focusWidgetMetric("일자", shortKoreanDate(event.date))}
            ${focusWidgetMetric("상태", stateText)}
          </div>
          <section class="focus-widget-section">
            <div class="focus-widget-section-title">상세</div>
            <p class="focus-widget-text">${escapeHtml(event.reason || event.assignees || event.subtitle || "상세 내용이 없습니다.")}</p>
          </section>
        `,
      });
    }

    function renderCompanyStaffTasks(tasks) {
      if (!companyStaffTaskBody) return;
      companyStaffTaskCache = tasks || [];
      const openTasks = (tasks || []).filter((task) => task.status !== "완료");
      const today = todayString();
      const dueToday = openTasks.filter((task) => crmDueDateOnly(task) === today).length;
      if (companyStaffDueToday) companyStaffDueToday.textContent = `${dueToday}건`;
      if (!openTasks.length) {
        companyStaffTaskBody.innerHTML = `<div>현재 배정된 미완료 업무가 없습니다.</div>`;
        return;
      }
      companyStaffTaskBody.innerHTML = openTasks.slice(0, 5).map((task) => `
        <div class="company-task-item" role="button" tabindex="0" aria-label="${escapeHtml(`${task.public_id || ""} ${task.title || ""} 크게 보기`)}" data-company-task-id="${escapeHtml(task.id)}">
          <div>
            <div class="company-task-title">${escapeHtml(task.title)}</div>
            <div class="company-task-meta">${escapeHtml(task.public_id || "")} · ${escapeHtml(crmTaskContextLabel(task))} · ${escapeHtml(crmDueDateText(task))}</div>
          </div>
          ${crmStatusBadge(task.status)}
        </div>
      `).join("");
    }

    function personInitials(user) {
      const label = (user.display_name || user.username || "?").trim();
      const compact = label.replace(/\s+/g, "");
      if (!compact) return "?";
      return Array.from(compact).slice(0, 2).join("");
    }

    function renderCompanyOrg(payload) {
      if (!companyOrgBody) return;
      const staff = payload.staff || [];
      if (!staff.length) {
        companyOrgBody.innerHTML = `<div class="company-org-empty">표시할 직원 계정이 없습니다.</div>`;
        return;
      }
      const leads = staff.filter((user) => user.role === "admin");
      const subLeads = staff.filter((user) => user.role === "sub_admin");
      const members = staff.filter((user) => !["admin", "sub_admin"].includes(user.role));
      const renderPerson = (user, lead = false) => `
        <article class="company-person-card${lead ? " lead" : ""}${String(user.id) === String(currentUser.id) ? " me" : ""}" role="button" tabindex="0" aria-label="${escapeHtml(`${user.display_name || user.username} 직원 크게 보기`)}" data-company-person-card="${escapeHtml(user.id)}">
          <div class="company-person-top">
            <div class="company-person-avatar">${escapeHtml(personInitials(user))}</div>
            <div>
              <div class="company-person-name">${escapeHtml(user.display_name || user.username)}</div>
              <div class="company-person-role">${escapeHtml(user.team_label || roleText(user.role))} · ${escapeHtml(user.username)}</div>
            </div>
          </div>
          <div class="company-person-meta">
            <span>진행<strong>${escapeHtml(user.open_tasks || 0)}</strong></span>
            <span>오늘<strong>${escapeHtml(user.due_today || 0)}</strong></span>
            <span>지연<strong>${escapeHtml(user.overdue || 0)}</strong></span>
          </div>
        </article>
      `;
      companyOrgBody.innerHTML = `
        <div class="company-org-tree">
          <div class="company-org-level lead">
            ${(leads.length ? leads : staff.slice(0, 1)).map((user) => renderPerson(user, true)).join("")}
          </div>
          ${subLeads.length ? `<div class="company-org-level staff">${subLeads.map((user) => renderPerson(user, true)).join("")}</div>` : ""}
          <div class="company-org-level staff">
            ${(members.length ? members : staff.filter((user) => !leads.includes(user) && !subLeads.includes(user))).map((user) => renderPerson(user)).join("")}
          </div>
        </div>
      `;
    }

    async function loadCompanyStaffDashboard() {
      let saved = {};
      try {
        saved = JSON.parse(localStorage.getItem("workhub_notice_template") || "{}");
      } catch {
        saved = {};
      }
      if (companyStaffNoticeTitle) companyStaffNoticeTitle.textContent = saved.title || "등록 전";
      if (!can("crm_view")) {
        renderCompanyStaffTasks([]);
        if (companyOrgBody) companyOrgBody.innerHTML = `<div class="company-org-empty">CRM 조회 권한이 없어 조직도 업무 현황을 볼 수 없습니다.</div>`;
        return;
      }
      const [staffData, myTaskData] = await Promise.all([
        crmFetchJson("/api/company-staff-dashboard"),
        crmFetchJson("/api/crm-tasks?mine=1&limit=6"),
      ]);
      companyStaffPayloadCache = staffData;
      renderCompanyOrg(staffData);
      renderCompanyStaffTasks(myTaskData.tasks || []);
    }

    function internalChatRoomLabel() {
      if (internalChatRoom.type === "dm") {
        const user = internalChatUsers.find((item) => String(item.id) === String(internalChatRoom.userId));
        return user ? `${user.display_name || user.username} DM` : "직원 DM";
      }
      return "전체방";
    }

    function renderInternalChatRooms() {
      if (!internalChatRoomList) return;
      const globalActive = internalChatRoom.type === "global";
      const rows = [`<button class="internal-chat-room${globalActive ? " active" : ""}" type="button" data-chat-room="global"><span>전체방</span><small>공지/공유</small></button>`];
      internalChatUsers.forEach((user) => {
        if (String(user.id) === String(currentUser.id)) return;
        const active = internalChatRoom.type === "dm" && String(internalChatRoom.userId) === String(user.id);
        rows.push(`
          <button class="internal-chat-room${active ? " active" : ""}" type="button" data-chat-room="dm" data-chat-user-id="${escapeHtml(user.id)}">
            <span>${escapeHtml(user.display_name || user.username)}</span>
            <small>${escapeHtml(roleText(user.role))}</small>
          </button>
        `);
      });
      internalChatRoomList.innerHTML = rows.join("");
      if (internalChatTitle) internalChatTitle.textContent = internalChatRoomLabel();
      if (internalChatHint) {
        internalChatHint.textContent = internalChatRoom.type === "dm"
          ? "DM에서도 /업무 @직원 업무내용 / 기한 으로 업무 지시 가능"
          : "전체방 공유와 /업무 @직원 업무내용 / 기한 명령을 지원";
      }
    }

    async function loadInternalChatUsers() {
      if (!can("crm_view")) {
        internalChatUsers = [];
        renderInternalChatRooms();
        return;
      }
      const data = await crmFetchJson("/api/company-staff-dashboard");
      internalChatUsers = data.staff || [];
      renderInternalChatRooms();
    }

    function internalMessageClass(message) {
      if (message.command_error) return " command-error";
      if (message.command_result) return " command-ok";
      return "";
    }

    function renderInternalMessages(messages) {
      if (!internalChatList) return;
      const rows = messages || [];
      if (!rows.length) {
        internalChatList.innerHTML = `<div class="internal-chat-empty">아직 메시지가 없습니다. 첫 공유나 DM을 남겨줘.</div>`;
        return;
      }
      internalChatList.innerHTML = rows.map((message) => {
        const isMine = String(message.user_id) === String(currentUser.id);
        return `
          <article class="internal-message${isMine ? " mine" : ""}${internalMessageClass(message)}">
            <div class="internal-message-meta">
              <span class="internal-message-name">${escapeHtml(message.display_name || message.username || "직원")}</span>
              <span>${escapeHtml(message.created_at || "")}</span>
            </div>
            <div class="internal-message-body">${escapeHtml(message.body || "")}</div>
            ${message.command_result ? `<div class="internal-message-meta"><span>${escapeHtml(message.command_result)}</span></div>` : ""}
            ${message.command_error ? `<div class="internal-message-meta"><span>${escapeHtml(message.command_error)}</span></div>` : ""}
          </article>
        `;
      }).join("");
      internalChatList.scrollTop = internalChatList.scrollHeight;
    }

    async function loadInternalMessages() {
      renderInternalChatRooms();
      const params = new URLSearchParams({ limit: "100", room: internalChatRoom.type });
      if (internalChatRoom.type === "dm" && internalChatRoom.userId) params.set("user_id", internalChatRoom.userId);
      const data = await crmFetchJson(`/api/internal-messages?${params.toString()}`);
      renderInternalMessages(data.messages || []);
    }

    async function setInternalChatRoom(type, userId = "") {
      internalChatRoom = { type: type === "dm" ? "dm" : "global", userId: type === "dm" ? String(userId || "") : "" };
      renderInternalChatRooms();
      await loadInternalMessages();
      internalChatBody?.focus();
    }

    async function sendInternalMessage(event) {
      event.preventDefault();
      if (!internalChatBody) return;
      const body = internalChatBody.value.trim();
      if (!body) {
        internalChatBody?.focus();
        return;
      }
      await crmFetchJson("/api/internal-message-save", {
        method: "POST",
        body: JSON.stringify({
          body,
          room_type: internalChatRoom.type,
          recipient_user_id: internalChatRoom.type === "dm" ? internalChatRoom.userId : "",
        }),
      });
      internalChatBody.value = "";
      await loadInternalMessages();
      if (body.startsWith("/업무")) {
        await Promise.all([
          loadCrmTasks().catch(() => {}),
          loadCrmMineTasks().catch(() => {}),
          loadCrmStaffDashboard().catch(() => {}),
          loadCompanyStaffDashboard().catch(() => {}),
          loadCompanyCalendar().catch(() => {}),
          loadCrmDashboard().catch(() => {}),
        ]);
      }
      internalChatBody?.focus();
    }

    function setCrmTab(tabName) {
      if (tabName === "messages" && !can("crm_message_manage")) tabName = "dashboard";
      crmActiveTab = tabName;
      crmTabs.forEach((button) => {
        const active = button.dataset.crmTab === tabName;
        button.classList.toggle("active", active);
        button.setAttribute("aria-selected", active ? "true" : "false");
        button.tabIndex = active ? 0 : -1;
      });
      crmPanels.forEach((panel) => {
        const active = panel.dataset.crmPanel === tabName;
        panel.classList.toggle("active", active);
        panel.hidden = !active;
      });
      syncCrmNavState();
      if (tabName === "accounts") {
        loadCrmStaffDashboard().catch((error) => setCrmMessage(error.message, true));
      } else if (tabName === "mine") {
        loadCrmMineTasks().catch((error) => setCrmMessage(error.message, true));
      } else if (tabName === "tasks") {
        loadCrmTasks().catch((error) => setCrmMessage(error.message, true));
      } else if (tabName === "messages") {
        loadCrmMessenger().catch((error) => setCrmMessage(error.message, true));
      } else {
        loadCrmDashboard().catch((error) => setCrmMessage(error.message, true));
      }
    }

    function crmTaskFilterDefaults() {
      return { q: "", status: "", assignee_user_id: "", priority: "", due: "", source: "", open_only: "1", sort: "smart" };
    }

    function normalizeCrmTaskFilters(filters = {}) {
      const defaults = crmTaskFilterDefaults();
      const normalized = { ...defaults };
      Object.keys(defaults).forEach((key) => {
        if (Object.prototype.hasOwnProperty.call(filters, key)) normalized[key] = String(filters[key] || "");
      });
      normalized.open_only = normalized.open_only === "1" || normalized.open_only === "true" ? "1" : "";
      if (normalized.status === "완료") normalized.open_only = "";
      if (!["smart", "due", "updated"].includes(normalized.sort)) normalized.sort = "smart";
      return normalized;
    }

    function readCrmTaskFilters() {
      return normalizeCrmTaskFilters({
        q: crmTaskSearch?.value.trim() || "",
        status: crmTaskStatusFilter?.value || "",
        assignee_user_id: crmTaskAssigneeFilter?.value || "",
        priority: crmTaskPriorityFilter?.value || "",
        due: crmTaskDueFilter?.value || "",
        source: crmTaskSourceFilter?.value || "",
        open_only: crmTaskOpenOnly?.checked ? "1" : "",
        sort: crmTaskSort?.value || "smart",
      });
    }

    function writeCrmTaskFilters(filters = {}) {
      const normalized = normalizeCrmTaskFilters(filters);
      if (crmTaskSearch) crmTaskSearch.value = normalized.q;
      if (crmTaskStatusFilter) crmTaskStatusFilter.value = normalized.status;
      if (crmTaskAssigneeFilter) crmTaskAssigneeFilter.value = normalized.assignee_user_id;
      if (crmTaskPriorityFilter) crmTaskPriorityFilter.value = normalized.priority;
      if (crmTaskDueFilter) crmTaskDueFilter.value = normalized.due;
      if (crmTaskSourceFilter) crmTaskSourceFilter.value = normalized.source;
      if (crmTaskOpenOnly) crmTaskOpenOnly.checked = normalized.open_only === "1";
      if (crmTaskSort) crmTaskSort.value = normalized.sort;
    }

    function crmBuiltinTaskView(viewId) {
      const id = String(viewId || "").replace(/^builtin:/, "");
      return CRM_TASK_BUILTIN_VIEWS.find((view) => view.id === id);
    }

    function crmSavedTaskView(viewId) {
      const id = String(viewId || "").replace(/^saved:/, "");
      return crmSavedViews.find((view) => String(view.id) === id);
    }

    function renderCrmSavedViews() {
      if (!crmTaskViewSelect || !crmTaskPresetList) return;
      const savedOptions = crmSavedViews.map((view) => (
        `<option value="saved:${escapeHtml(view.id)}">${escapeHtml(view.name)}</option>`
      )).join("");
      crmTaskViewSelect.innerHTML = [
        `<option value="">저장뷰 선택</option>`,
        `<optgroup label="기본 보기">${CRM_TASK_BUILTIN_VIEWS.map((view) => (
          `<option value="builtin:${escapeHtml(view.id)}">${escapeHtml(view.name)}</option>`
        )).join("")}</optgroup>`,
        savedOptions ? `<optgroup label="내 저장뷰">${savedOptions}</optgroup>` : "",
      ].join("");
      crmTaskViewSelect.value = Array.from(crmTaskViewSelect.options).some((option) => option.value === crmActiveTaskViewId) ? crmActiveTaskViewId : "";
      crmTaskPresetList.innerHTML = CRM_TASK_BUILTIN_VIEWS.map((view) => {
        const active = crmActiveTaskViewId === `builtin:${view.id}`;
        return `<button class="crm-view-pill${active ? " active" : ""}" type="button" data-crm-task-view="builtin:${escapeHtml(view.id)}">${escapeHtml(view.name)}</button>`;
      }).join("");
      if (crmTaskViewName && document.activeElement !== crmTaskViewName) {
        const selectedSaved = crmSavedTaskView(crmActiveTaskViewId);
        crmTaskViewName.value = selectedSaved?.name || "";
      }
      if (crmTaskViewDelete) crmTaskViewDelete.disabled = !crmActiveTaskViewId.startsWith("saved:");
    }

    function applyCrmTaskView(viewId, load = true) {
      const viewKey = String(viewId || "");
      const builtin = crmBuiltinTaskView(viewKey);
      const saved = crmSavedTaskView(viewKey);
      if (!builtin && !saved) return;
      crmActiveTaskViewId = builtin ? `builtin:${builtin.id}` : `saved:${saved.id}`;
      writeCrmTaskFilters(builtin ? builtin.filters : saved.filters || {});
      renderCrmSavedViews();
      if (load) loadCrmTasks().catch((error) => setCrmMessage(error.message, true));
    }

    function markCrmTaskFiltersDirty() {
      crmActiveTaskViewId = "";
      renderCrmSavedViews();
    }

    async function loadCrmSavedViews() {
      if (!can("crm_view")) return;
      const data = await crmFetchJson("/api/crm-saved-views?scope=tasks");
      crmSavedViews = data.views || [];
      renderCrmSavedViews();
    }

    async function saveCurrentCrmTaskView() {
      const filters = readCrmTaskFilters();
      const selected = crmSavedTaskView(crmActiveTaskViewId);
      const name = String(crmTaskViewName?.value || selected?.name || "").trim();
      if (!name) {
        setCrmMessage("저장뷰 이름을 입력해줘.", true);
        crmTaskViewName?.focus();
        return;
      }
      const data = await crmFetchJson("/api/crm-saved-view-save", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({
          id: selected?.id || "",
          scope: "tasks",
          name: name.trim(),
          filters,
          sort_key: filters.sort,
        }),
      });
      crmSavedViews = data.views || [];
      crmActiveTaskViewId = `saved:${data.view_id}`;
      if (crmTaskViewName) crmTaskViewName.value = name.trim();
      renderCrmSavedViews();
      setCrmMessage(data.message || "저장뷰를 저장했습니다.");
    }

    async function deleteCurrentCrmTaskView() {
      const selected = crmSavedTaskView(crmActiveTaskViewId);
      if (!selected) {
        setCrmMessage("삭제할 내 저장뷰를 선택해줘.", true);
        return;
      }
      if (!confirm(`'${selected.name}' 저장뷰를 삭제할까요?`)) return;
      const data = await crmFetchJson("/api/crm-saved-view-delete", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ id: selected.id, scope: "tasks" }),
      });
      crmSavedViews = data.views || [];
      crmActiveTaskViewId = "builtin:open";
      writeCrmTaskFilters(crmBuiltinTaskView("open")?.filters || {});
      renderCrmSavedViews();
      setCrmMessage(data.message || "저장뷰를 삭제했습니다.");
      await loadCrmTasks();
    }

    function renderCrmUserOptions() {
      const options = [`<option value="">담당자 선택</option>`].concat(
        crmUsers.map((user) => `<option value="${escapeHtml(user.id)}">${escapeHtml(user.display_name || user.username)}</option>`)
      ).join("");
      if (crmTaskAssignee) crmTaskAssignee.innerHTML = options;
      if (crmMessengerUser) crmMessengerUser.innerHTML = options;
      if (crmTaskAssigneeFilter) {
        const current = crmTaskAssigneeFilter.value;
        crmTaskAssigneeFilter.innerHTML = [`<option value="">담당자 전체</option>`].concat(
          crmUsers.map((user) => `<option value="${escapeHtml(user.id)}">${escapeHtml(user.display_name || user.username)}</option>`)
        ).join("");
        crmTaskAssigneeFilter.value = Array.from(crmTaskAssigneeFilter.options).some((option) => option.value === current) ? current : "";
      }
    }

    function renderCrmAccountOptions() {
      if (!crmTaskAccount) return;
      crmTaskAccount.innerHTML = [`<option value="">업무 구분/거래처 선택</option>`].concat(
        crmAccounts.map((account) => `<option value="${escapeHtml(account.id)}">${escapeHtml(account.name)}</option>`)
      ).join("");
    }

    function projectProgressLabel(project) {
      if ((project.overdue || 0) > 0) return "지연 확인";
      if ((project.due_today || 0) > 0) return "오늘 마감";
      if ((project.open_tasks || 0) === 0) return "완료";
      if ((project.progress_tasks || 0) > 0) return "진행중";
      return "대기";
    }

    function projectProgressPercent(project) {
      const value = Number(project.progress_percent || 0);
      return Math.max(0, Math.min(100, Number.isFinite(value) ? value : 0));
    }

    function renderCrmProjectProgress(projects) {
      if (!crmProjectProgressBody) return;
      if (!projects.length) {
        crmProjectProgressBody.innerHTML = `<div class="crm-project-empty">등록된 프로젝트 업무가 없습니다.</div>`;
        return;
      }
      crmProjectProgressBody.innerHTML = projects.map((project) => {
        const percent = projectProgressPercent(project);
        const assignees = String(project.assignee_names || "").split(",").map((name) => name.trim()).filter(Boolean);
        const ownerText = assignees.length ? assignees.slice(0, 4).join(", ") : "담당자 미정";
        const label = projectProgressLabel(project);
        return `
          <article class="crm-project-row" role="button" tabindex="0" aria-label="${escapeHtml(`${project.project_name || "프로젝트"} 진행상황 크게 보기`)}" data-crm-project-key="${escapeHtml(project.project_key || "")}">
            <div class="crm-project-main">
              <div class="crm-project-title">${escapeHtml(project.project_name || "프로젝트 미지정")}</div>
              <div class="crm-project-meta">${escapeHtml(ownerText)} · 다음 마감 ${escapeHtml(project.next_due_at || "없음")}</div>
            </div>
            <div class="crm-project-progress" role="progressbar" aria-valuemin="0" aria-valuemax="100" aria-valuenow="${escapeHtml(percent)}" aria-label="${escapeHtml(project.project_name || "프로젝트")} 완료율">
              <div class="crm-project-progress-top"><span>${escapeHtml(label)}</span><strong>${escapeHtml(percent)}%</strong></div>
              <div class="crm-project-bar"><span style="width: ${escapeHtml(percent)}%"></span></div>
            </div>
            <div class="crm-project-metrics">
              <span class="crm-project-pill">${escapeHtml(project.open_tasks || 0)}건 진행</span>
              <span class="crm-project-pill">${escapeHtml(project.completed_tasks || 0)}/${escapeHtml(project.total_tasks || 0)} 완료</span>
              ${(project.due_today || 0) > 0 ? `<span class="crm-project-pill today">오늘 ${escapeHtml(project.due_today)}건</span>` : ""}
              ${(project.overdue || 0) > 0 ? `<span class="crm-project-pill danger">지연 ${escapeHtml(project.overdue)}건</span>` : ""}
            </div>
          </article>
        `;
      }).join("");
    }

    function renderCrmDashboard(data) {
      const stats = data.stats || {};
      crmStatAccounts.textContent = crmUsers.length || stats.accounts || 0;
      crmStatOpenTasks.textContent = stats.open_tasks || 0;
      crmStatDueToday.textContent = stats.due_today || 0;
      crmStatOverdue.textContent = stats.overdue || 0;
      crmProjectProgress = data.project_progress || [];
      renderCrmProjectProgress(crmProjectProgress);
      const tasks = data.priority_tasks || [];
      crmPriorityTaskBody.innerHTML = tasks.length ? tasks.map((task) => `
        <tr>
          <td>${escapeHtml(task.public_id)}</td>
          <td>${escapeHtml(crmTaskContextLabel(task))}</td>
          <td class="left">${escapeHtml(task.title)}</td>
          <td>${escapeHtml(task.assignee_name)}</td>
          <td>${escapeHtml(task.due_at)}</td>
          <td>${crmStatusBadge(task.status)}</td>
        </tr>
      `).join("") : `<tr><td colspan="6">우선 처리할 CRM 업무가 없습니다.</td></tr>`;
      const events = data.recent_events || [];
      crmRecentMessageBody.innerHTML = events.length ? events.map((event) => `
        <tr>
          <td>${escapeHtml(event.created_at)}</td>
          <td>${escapeHtml(event.sender_name || event.sender_key)}</td>
          <td>${escapeHtml(event.result)}</td>
          <td class="left">${escapeHtml(event.text)}</td>
        </tr>
      `).join("") : `<tr><td colspan="4">최근 메신저 처리 로그가 없습니다.</td></tr>`;
    }

    async function loadCrmDashboard() {
      if (!can("crm_view")) return;
      const data = await crmFetchJson("/api/crm-dashboard");
      renderCrmDashboard(data);
    }

    function resetCrmAccountForm() {
      crmAccountId.value = "";
      crmAccountName.value = "";
      crmAccountType.value = "";
      crmAccountContact.value = "";
      crmAccountPhone.value = "";
      crmAccountEmail.value = "";
      crmAccountMemo.value = "";
    }

    function fillCrmAccountForm(account) {
      setCrmTab("accounts");
      crmAccountId.value = account.id || "";
      crmAccountName.value = account.name || "";
      crmAccountType.value = account.account_type || "";
      crmAccountContact.value = account.contact_name || "";
      crmAccountPhone.value = account.phone || "";
      crmAccountEmail.value = account.email || "";
      crmAccountMemo.value = account.memo || "";
      crmAccountName?.focus();
    }

    function renderCrmAccounts() {
      renderCrmAccountOptions();
      crmAccountBody.innerHTML = crmAccounts.length ? crmAccounts.map((account) => `
        <tr>
          <td class="left"><strong>${escapeHtml(account.name)}</strong><div>${escapeHtml(account.memo)}</div></td>
          <td>${escapeHtml(account.account_type)}</td>
          <td>${escapeHtml(account.contact_name)}</td>
          <td>${escapeHtml(account.phone)}</td>
          <td>${escapeHtml(account.email)}</td>
          <td>${escapeHtml(account.open_task_count || 0)}건</td>
          <td><button class="crm-mini-button" type="button" data-crm-account-edit="${escapeHtml(account.id)}">수정</button></td>
        </tr>
      `).join("") : `<tr><td colspan="7">등록된 CRM 거래처가 없습니다.</td></tr>`;
    }

    function latestTaskForUser(tasks, userId) {
      return (tasks || []).find((task) => String(task.assignee_user_id || "") === String(userId || "") && task.status !== "완료");
    }

    function renderCrmStaffDashboard(payload, tasks) {
      if (!crmStaffBody) return;
      crmStaffPayloadCache = payload || null;
      crmStaffTaskCache = tasks || [];
      const staff = payload?.staff || [];
      const rows = tasks || [];
      if (!staff.length) {
        crmStaffBody.innerHTML = `<div class="internal-chat-empty">표시할 직원 계정이 없습니다.</div>`;
        return;
      }
      crmStaffBody.innerHTML = staff.map((user) => {
        const latest = latestTaskForUser(rows, user.id);
        return `
          <article class="crm-staff-row" role="button" tabindex="0" aria-label="${escapeHtml(`${user.display_name || user.username} 직원 업무 크게 보기`)}" data-crm-staff-card="${escapeHtml(user.id)}">
            <div class="crm-staff-person">
              <div class="crm-staff-avatar">${escapeHtml(personInitials(user))}</div>
              <div>
                <div class="crm-staff-name">${escapeHtml(user.display_name || user.username)}</div>
                <div class="crm-staff-role">${escapeHtml(roleText(user.role))} · ${escapeHtml(user.username)}</div>
              </div>
            </div>
            <div class="crm-staff-metric"><span>진행</span><strong>${escapeHtml(user.open_tasks || 0)}</strong></div>
            <div class="crm-staff-metric"><span>오늘</span><strong>${escapeHtml(user.due_today || 0)}</strong></div>
            <div class="crm-staff-metric"><span>지연</span><strong>${escapeHtml(user.overdue || 0)}</strong></div>
            <div class="crm-staff-latest">${latest ? `${escapeHtml(latest.public_id || "")} · ${escapeHtml(latest.title || "")} · ${escapeHtml(crmDueDateText(latest))}` : "최근 미완료 업무 없음"}</div>
          </article>
        `;
      }).join("");
    }

    async function loadCrmStaffDashboard() {
      if (!can("crm_view")) return;
      const [staffPayload, taskPayload] = await Promise.all([
        crmFetchJson("/api/company-staff-dashboard"),
        crmFetchJson("/api/crm-tasks?limit=2000"),
      ]);
      crmUsers = staffPayload.staff || crmUsers;
      renderCrmUserOptions();
      renderCrmStaffDashboard(staffPayload, taskPayload.tasks || []);
    }

    async function loadCrmAccounts() {
      if (!can("crm_view")) return;
      const params = new URLSearchParams();
      if (crmAccountSearch?.value) params.set("q", crmAccountSearch.value.trim());
      const data = await crmFetchJson(`/api/crm-accounts?${params.toString()}`);
      crmAccounts = data.accounts || [];
      renderCrmAccounts();
    }

    async function saveCrmAccountForm(event) {
      event.preventDefault();
      if (!can("crm_manage")) return;
      const payload = {
        id: crmAccountId.value,
        name: crmAccountName.value.trim(),
        account_type: crmAccountType.value.trim(),
        contact_name: crmAccountContact.value.trim(),
        phone: crmAccountPhone.value.trim(),
        email: crmAccountEmail.value.trim(),
        memo: crmAccountMemo.value.trim(),
      };
      const data = await crmFetchJson("/api/crm-account-save", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      setCrmMessage(data.message || "거래처를 저장했습니다.");
      resetCrmAccountForm();
      await loadCrmAccounts();
      await loadCrmDashboard();
    }

    function setCrmTaskFormOpen(open) {
      if (!crmTaskForm) return;
      crmTaskForm.classList.toggle("collapsed", !open);
      if (crmTaskFormToggle) crmTaskFormToggle.textContent = open ? "입력 닫기" : "입력 열기";
    }

    function resetCrmTaskForm() {
      crmTaskId.value = "";
      crmTaskAccount.value = "";
      crmTaskAccountName.value = "";
      crmTaskTitle.value = "";
      crmTaskAssignee.value = "";
      crmTaskDue.value = "";
      crmTaskPriority.value = "보통";
      crmTaskStatus.value = "대기";
      crmTaskDescription.value = "";
    }

    function fillCrmTaskForm(task) {
      setCrmTaskFormOpen(true);
      setCrmTab("tasks");
      crmTaskId.value = task.id || "";
      crmTaskAccount.value = task.account_id || "";
      crmTaskAccountName.value = task.account_name || "";
      crmTaskTitle.value = task.title || "";
      crmTaskAssignee.value = task.assignee_user_id || "";
      crmTaskDue.value = task.due_at || "";
      crmTaskPriority.value = task.priority || "보통";
      crmTaskStatus.value = task.status || "대기";
      crmTaskDescription.value = task.description || "";
      crmTaskTitle?.focus();
    }

    function crmPriorityClass(priority) {
      if (priority === "높음") return "high";
      if (priority === "낮음") return "low";
      return "normal";
    }

    function crmPriorityBadge(priority) {
      const label = priority || "보통";
      return `<span class="crm-priority badge badge-sm ${crmPriorityClass(label)}">${escapeHtml(label)}</span>`;
    }

    function crmDueDateText(task) {
      return task?.due_at || "기한 없음";
    }

    function crmTaskContextLabel(task) {
      return task?.account_name || "직원 지시 업무";
    }

    function crmSourceLabel(source) {
      const value = String(source || "app");
      if (value === "internal_message") return "사내 메신저";
      if (value.startsWith("messenger:")) return "외부 메신저";
      if (value === "app") return "직접 등록";
      return value;
    }

    function crmDueDateOnly(task) {
      const due = task?.due_at || "";
      return due.length >= 10 ? due.slice(0, 10) : "";
    }

    function renderCrmTaskBoardStats() {
      if (!crmTaskBoardStats) return;
      const today = todayString();
      const openTasks = crmTasks.filter((task) => task.status !== "완료");
      const dueToday = openTasks.filter((task) => crmDueDateOnly(task) === today).length;
      const overdue = openTasks.filter((task) => {
        const day = crmDueDateOnly(task);
        return day && day < today;
      }).length;
      const highPriority = openTasks.filter((task) => task.priority === "높음").length;
      const holdTasks = crmTasks.filter((task) => task.status === "보류").length;
      const stats = [
        ["진행 업무", openTasks.length, "OPEN"],
        ["오늘 마감", dueToday, "DUE"],
        ["지연 업무", overdue, "LATE"],
        ["높음/보류", `${highPriority}/${holdTasks}`, "HOT"],
      ];
      crmTaskBoardStats.innerHTML = stats.map(([label, value, icon]) => `
        <article class="crm-board-stat">
          <div><span>${escapeHtml(label)}</span><strong>${escapeHtml(value)}</strong></div>
          <i>${escapeHtml(icon)}</i>
        </article>
      `).join("");
    }

    function renderCrmTaskCard(task) {
      const isActive = String(task.id) === String(crmSelectedTaskId);
      const canEdit = can("crm_manage");
      return `
        <article class="crm-task-card ${isActive ? "active" : ""}" role="button" tabindex="0" aria-pressed="${isActive ? "true" : "false"}" aria-label="${escapeHtml(`${task.public_id || ""} ${task.title || ""}`)}" data-crm-task-card="${escapeHtml(task.id)}">
          <div class="crm-task-card-top">
            ${crmPriorityBadge(task.priority)}
            <span class="crm-status ${crmStatusClass(task.status)}">${escapeHtml(task.status || "대기")}</span>
          </div>
          <div>
            <div class="crm-task-card-title">${escapeHtml(task.title)}</div>
            <div class="crm-task-card-sub">${escapeHtml(crmTaskContextLabel(task))}</div>
          </div>
          <div class="crm-task-card-sub">${escapeHtml(task.description || "상세 내용이 없습니다.")}</div>
          <div class="crm-task-card-meta">
            <span>${escapeHtml(task.assignee_name || "담당자 미정")}</span>
            <span class="crm-due-text">${escapeHtml(crmDueDateText(task))}</span>
          </div>
          <div class="crm-task-card-actions">
            <span>${escapeHtml(task.public_id || "")}</span>
            <span class="crm-mini-actions">
              ${canEdit ? `<button class="crm-mini-button" type="button" data-crm-task-edit="${escapeHtml(task.id)}">수정</button>` : ""}
              <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="진행중">확인</button>
              <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="완료">완료</button>
            </span>
          </div>
        </article>
      `;
    }

    function crmTimelineTypeLabel(type) {
      if (type === "status") return "상태 변경";
      if (type === "messenger") return "메신저";
      return "댓글";
    }

    function renderCrmTaskTimeline(task) {
      const key = String(task.id || "");
      const comments = crmTaskComments[key];
      const loading = crmTaskCommentLoads[key];
      const creation = {
        label: "업무 생성",
        created_at: task.created_at || "",
        author_name: task.requester_name || "시스템",
        body: `${crmSourceLabel(task.source)}에서 생성됐습니다.`,
      };
      const commentItems = (comments || []).map((comment) => ({
        label: crmTimelineTypeLabel(comment.comment_type),
        created_at: comment.created_at || "",
        author_name: comment.author_name || "직원",
        body: comment.body || "",
      }));
      const items = commentItems.concat([creation]);
      if (!comments && loading) {
        items.unshift({
          label: "불러오는 중",
          created_at: "",
          author_name: "",
          body: "활동 이력을 불러오고 있습니다.",
        });
      } else if (comments && !comments.length) {
        items.unshift({
          label: "활동 없음",
          created_at: "",
          author_name: "",
          body: "아직 댓글이나 상태 변경 이력이 없습니다.",
        });
      }
      return `
        <div>
          <div class="crm-timeline-title">활동 이력</div>
          <ol class="crm-timeline" aria-label="업무 활동 이력">
            ${items.map((item) => `
              <li class="crm-timeline-item">
                <div class="crm-timeline-head">
                  <span>${escapeHtml(item.label)}${item.author_name ? ` · ${escapeHtml(item.author_name)}` : ""}</span>
                  <span>${escapeHtml(item.created_at)}</span>
                </div>
                <div class="crm-timeline-body">${escapeHtml(item.body)}</div>
              </li>
            `).join("")}
          </ol>
        </div>
      `;
    }

    function renderCrmTaskDetail(task) {
      if (!crmTaskDetail) return;
      if (!task) {
        crmTaskDetail.innerHTML = `<div class="crm-task-detail-empty">업무 카드를 선택하면 상세 정보와 처리 버튼이 표시됩니다.</div>`;
        return;
      }
      const canEdit = can("crm_manage");
      const commentId = `crmTaskCommentBody-${task.id}`;
      crmTaskDetail.innerHTML = `
        <div class="crm-task-detail-inner">
          <div>
            <div class="crm-task-detail-kicker">${escapeHtml(crmTaskContextLabel(task))} · ${escapeHtml(task.public_id || "")}</div>
            <div class="crm-task-detail-title">${escapeHtml(task.title)}</div>
          </div>
          <div class="crm-task-detail-desc">${escapeHtml(task.description || "상세 내용이 없습니다.")}</div>
          <div class="crm-detail-grid">
            <div class="crm-detail-cell"><span>담당자</span><strong>${escapeHtml(task.assignee_name || "미정")}</strong></div>
            <div class="crm-detail-cell"><span>요청자</span><strong>${escapeHtml(task.requester_name || "미정")}</strong></div>
            <div class="crm-detail-cell"><span>마감</span><strong>${escapeHtml(crmDueDateText(task))}</strong></div>
            <div class="crm-detail-cell"><span>상태</span><strong>${escapeHtml(task.status || "대기")}</strong></div>
            <div class="crm-detail-cell"><span>우선순위</span><strong>${escapeHtml(task.priority || "보통")}</strong></div>
            <div class="crm-detail-cell"><span>출처</span><strong>${escapeHtml(crmSourceLabel(task.source))}</strong></div>
          </div>
          <div class="crm-detail-actions">
            ${canEdit ? `<button class="crm-mini-button" type="button" data-crm-task-edit="${escapeHtml(task.id)}">수정</button>` : ""}
            <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="진행중">확인</button>
            <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="완료">완료</button>
            <button class="crm-mini-button" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="보류">보류</button>
            <button class="crm-mini-button" type="button" data-crm-task-comment="${escapeHtml(task.id)}">댓글</button>
          </div>
          <form class="crm-detail-comment-form" data-crm-comment-form="${escapeHtml(task.id)}">
            <label for="${escapeHtml(commentId)}">댓글</label>
            <textarea class="crm-textarea" id="${escapeHtml(commentId)}" data-crm-comment-body placeholder="처리 내용이나 다음 액션을 남겨줘."></textarea>
            <button class="crm-mini-button primary" type="submit">댓글 등록</button>
          </form>
          ${renderCrmTaskTimeline(task)}
        </div>
      `;
    }

    function focusCrmTaskCommentForm(taskId) {
      const escapedTaskId = window.CSS?.escape ? CSS.escape(String(taskId || "")) : String(taskId || "").replaceAll('"', '\\"');
      const textarea = crmTaskDetail?.querySelector(`[data-crm-comment-form="${escapedTaskId}"] [data-crm-comment-body]`);
      if (textarea) textarea.focus();
    }

    async function ensureCrmTaskComments(taskId, force = false) {
      const key = String(taskId || "");
      if (!key || (!force && crmTaskComments[key])) return;
      if (crmTaskCommentLoads[key]) return crmTaskCommentLoads[key];
      crmTaskCommentLoads[key] = crmFetchJson(`/api/crm-task-comments?task_id=${encodeURIComponent(key)}`)
        .then((data) => {
          crmTaskComments[key] = data.comments || [];
          delete crmTaskCommentLoads[key];
          const selected = crmTasks.find((task) => String(task.id) === String(crmSelectedTaskId));
          if (selected && String(selected.id) === key) renderCrmTaskDetail(selected);
        })
        .catch((error) => {
          delete crmTaskCommentLoads[key];
          setCrmMessage(error.message, true);
        });
      return crmTaskCommentLoads[key];
    }

    function selectCrmTask(taskId) {
      crmSelectedTaskId = String(taskId || "");
      renderCrmTasks();
      ensureCrmTaskComments(crmSelectedTaskId).catch(() => {});
    }

    function focusWidgetMetric(label, value, raw = false) {
      const text = value === undefined || value === null || value === "" ? "-" : value;
      return `
        <div class="focus-widget-metric">
          <span>${escapeHtml(label)}</span>
          <strong>${raw ? text : escapeHtml(text)}</strong>
        </div>
      `;
    }

    function openFocusWidget({ kicker = "크게 보기", title = "상세 보기", subtitle = "", body = "" } = {}) {
      if (!focusWidget || !focusWidgetBody) return;
      const wasOpen = focusWidget.classList.contains("open");
      if (!wasOpen) focusWidgetLastFocus = document.activeElement;
      focusWidgetKicker.textContent = kicker;
      focusWidgetTitle.textContent = title;
      focusWidgetSubtitle.textContent = subtitle || "";
      focusWidgetBody.innerHTML = body;
      focusWidget.classList.add("open");
      focusWidget.setAttribute("aria-hidden", "false");
      if (!wasOpen) setTimeout(() => focusWidgetClose?.focus(), 0);
    }

    function closeFocusWidget() {
      if (!focusWidget) return;
      focusWidget.classList.remove("open");
      focusWidget.setAttribute("aria-hidden", "true");
      focusWidgetBody.innerHTML = "";
      focusWidgetTaskId = "";
      focusWidgetEmployeeId = "";
      if (focusWidgetLastFocus && typeof focusWidgetLastFocus.focus === "function") {
        const returnFocusTarget = focusWidgetLastFocus;
        setTimeout(() => returnFocusTarget.focus(), 0);
      }
      focusWidgetLastFocus = null;
    }

    function focusWidgetTaskBody(task) {
      const canEdit = can("crm_manage");
      const commentId = `focusWidgetComment-${String(task.id || "").replace(/[^a-zA-Z0-9_-]/g, "-")}`;
      return `
        <div class="focus-widget-grid">
          ${focusWidgetMetric("상태", crmStatusBadge(task.status), true)}
          ${focusWidgetMetric("우선순위", crmPriorityBadge(task.priority), true)}
          ${focusWidgetMetric("마감", crmDueDateText(task))}
          ${focusWidgetMetric("담당자", task.assignee_name || "담당자 미정")}
          ${focusWidgetMetric("요청자", task.requester_name || "요청자 미정")}
          ${focusWidgetMetric("출처", crmSourceLabel(task.source))}
        </div>
        <section class="focus-widget-section">
          <div class="focus-widget-section-title">업무 내용</div>
          <p class="focus-widget-text">${escapeHtml(task.description || "상세 내용이 없습니다.")}</p>
        </section>
        <div class="focus-widget-actions">
          ${canEdit ? `<button class="crm-mini-button" type="button" data-crm-task-edit="${escapeHtml(task.id)}">수정</button>` : ""}
          <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="진행중">확인</button>
          <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="완료">완료</button>
          <button class="crm-mini-button" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="보류">보류</button>
        </div>
        <form class="crm-detail-comment-form" data-focus-widget-comment-form="${escapeHtml(task.id)}">
          <label for="${escapeHtml(commentId)}">댓글</label>
          <textarea class="crm-textarea" id="${escapeHtml(commentId)}" data-focus-widget-comment-body placeholder="처리 내용이나 다음 액션을 남겨줘."></textarea>
          <button class="crm-mini-button primary" type="submit">댓글 등록</button>
        </form>
        ${renderCrmTaskTimeline(task)}
      `;
    }

    function renderFocusWidgetTask(task) {
      openFocusWidget({
        kicker: "업무 크게 보기",
        title: task.title || "제목 없음",
        subtitle: [task.public_id || "", crmTaskContextLabel(task), crmSourceLabel(task.source)].filter(Boolean).join(" · "),
        body: focusWidgetTaskBody(task),
      });
    }

    async function openCrmTaskWidget(taskId) {
      const task = findCrmTaskById(taskId);
      if (!task) {
        openFocusWidget({
          kicker: "업무 크게 보기",
          title: "업무를 찾지 못했습니다",
          subtitle: "",
          body: `<p class="focus-widget-text">현재 화면에 로드된 업무 목록에서 해당 업무를 찾지 못했습니다. 업무보드를 새로고침한 뒤 다시 시도해줘.</p>`,
        });
        return;
      }
      focusWidgetTaskId = String(task.id || "");
      focusWidgetEmployeeId = "";
      renderFocusWidgetTask(task);
      await ensureCrmTaskComments(task.id);
      const refreshed = findCrmTaskById(task.id) || task;
      if (focusWidget?.classList.contains("open") && focusWidgetTaskId === String(task.id || "")) {
        renderFocusWidgetTask(refreshed);
      }
    }

    function findStaffUserById(userId) {
      const sources = [
        companyStaffPayloadCache?.staff || [],
        crmStaffPayloadCache?.staff || [],
        crmUsers || [],
        internalChatUsers || [],
      ];
      for (const source of sources) {
        const found = source.find((user) => String(user.id) === String(userId));
        if (found) return found;
      }
      return null;
    }

    function tasksForAssignee(userId, rows) {
      return (rows || []).filter((task) => String(task.assignee_user_id || "") === String(userId || ""));
    }

    function focusWidgetEmployeeBody(user, tasks) {
      const openTasks = (tasks || []).filter((task) => task.status !== "완료");
      const today = todayString();
      const dueToday = openTasks.filter((task) => crmDueDateOnly(task) === today).length;
      const overdue = openTasks.filter((task) => {
        const day = crmDueDateOnly(task);
        return day && day < today;
      }).length;
      const rows = openTasks.slice(0, 20);
      return `
        <div class="focus-widget-grid">
          ${focusWidgetMetric("진행 업무", user.open_tasks ?? openTasks.length)}
          ${focusWidgetMetric("오늘 마감", user.due_today ?? dueToday)}
          ${focusWidgetMetric("지연 업무", user.overdue ?? overdue)}
          ${focusWidgetMetric("역할", user.team_label || roleText(user.role))}
          ${focusWidgetMetric("계정", user.username || "-")}
          ${focusWidgetMetric("상태", user.is_active === 0 ? "비활성" : "활성")}
        </div>
        <section class="focus-widget-section">
          <div class="focus-widget-section-title">최근 배정 업무</div>
          <div class="focus-widget-table-wrap">
            <table class="focus-widget-table">
              <thead><tr><th>번호</th><th>업무</th><th>마감</th><th>상태</th><th>우선순위</th><th>보기</th></tr></thead>
              <tbody>
                ${rows.length ? rows.map((task) => `
                  <tr>
                    <td>${escapeHtml(task.public_id || "")}</td>
                    <td>${escapeHtml(task.title || "")}<div>${escapeHtml(crmTaskContextLabel(task))}</div></td>
                    <td>${escapeHtml(crmDueDateText(task))}</td>
                    <td>${crmStatusBadge(task.status)}</td>
                    <td>${crmPriorityBadge(task.priority)}</td>
                    <td><button class="crm-mini-button" type="button" data-focus-open-task="${escapeHtml(task.id)}">열기</button></td>
                  </tr>
                `).join("") : `<tr><td colspan="6">현재 표시할 미완료 업무가 없습니다.</td></tr>`}
              </tbody>
            </table>
          </div>
        </section>
      `;
    }

    async function openEmployeeWidget(userId) {
      const user = findStaffUserById(userId);
      if (!user) {
        openFocusWidget({
          kicker: "직원 크게 보기",
          title: "직원을 찾지 못했습니다",
          body: `<p class="focus-widget-text">현재 화면에 로드된 직원 목록에서 해당 직원을 찾지 못했습니다.</p>`,
        });
        return;
      }
      focusWidgetEmployeeId = String(userId || "");
      focusWidgetTaskId = "";
      const cachedTasks = tasksForAssignee(userId, crmStaffTaskCache.concat(companyStaffTaskCache));
      openFocusWidget({
        kicker: "직원 크게 보기",
        title: user.display_name || user.username || "직원",
        subtitle: `${user.team_label || roleText(user.role)} · ${user.username || ""}`,
        body: focusWidgetEmployeeBody(user, cachedTasks),
      });
      if (!can("crm_view")) return;
      try {
        const data = await crmFetchJson(`/api/crm-tasks?assignee_user_id=${encodeURIComponent(userId)}&open_only=1&sort=due&limit=20`);
        const latestTasks = data.tasks || [];
        crmStaffTaskCache = crmStaffTaskCache
          .filter((task) => String(task.assignee_user_id || "") !== String(userId))
          .concat(latestTasks);
        if (focusWidget?.classList.contains("open") && focusWidgetEmployeeId === String(userId || "")) {
          openFocusWidget({
            kicker: "직원 크게 보기",
            title: user.display_name || user.username || "직원",
            subtitle: `${user.team_label || roleText(user.role)} · ${user.username || ""}`,
            body: focusWidgetEmployeeBody(user, latestTasks),
          });
        }
      } catch (error) {
        setCrmMessage(error.message, true);
      }
    }

    function findCrmProjectByKey(projectKey) {
      return crmProjectProgress.find((project) => String(project.project_key || "") === String(projectKey || ""));
    }

    function focusWidgetProjectBody(project) {
      const percent = projectProgressPercent(project);
      const tasks = project.tasks || [];
      const assignees = String(project.assignee_names || "").split(",").map((name) => name.trim()).filter(Boolean);
      return `
        <div class="focus-widget-grid">
          ${focusWidgetMetric("완료율", `${percent}%`)}
          ${focusWidgetMetric("진행 업무", `${project.open_tasks || 0}건`)}
          ${focusWidgetMetric("전체 업무", `${project.total_tasks || 0}건`)}
          ${focusWidgetMetric("오늘 마감", `${project.due_today || 0}건`)}
          ${focusWidgetMetric("지연 업무", `${project.overdue || 0}건`)}
          ${focusWidgetMetric("높은 우선순위", `${project.high_priority || 0}건`)}
        </div>
        <section class="focus-widget-section">
          <div class="focus-widget-section-title">프로젝트 상태</div>
          <div class="crm-project-progress" role="progressbar" aria-valuemin="0" aria-valuemax="100" aria-valuenow="${escapeHtml(percent)}" aria-label="${escapeHtml(project.project_name || "프로젝트")} 완료율">
            <div class="crm-project-progress-top"><span>${escapeHtml(projectProgressLabel(project))}</span><strong>${escapeHtml(percent)}%</strong></div>
            <div class="crm-project-bar"><span style="width: ${escapeHtml(percent)}%"></span></div>
          </div>
          <p class="focus-widget-text">담당자: ${escapeHtml(assignees.length ? assignees.join(", ") : "담당자 미정")}\n다음 마감: ${escapeHtml(project.next_due_at || "없음")}</p>
        </section>
        <div class="focus-widget-actions">
          <button class="crm-mini-button primary" type="button" data-focus-filter-project="${escapeHtml(project.project_key || "")}">업무보드에서 보기</button>
        </div>
        <section class="focus-widget-section">
          <div class="focus-widget-section-title">관련 업무</div>
          <div class="focus-widget-table-wrap">
            <table class="focus-widget-table">
              <thead><tr><th>번호</th><th>업무</th><th>담당자</th><th>마감</th><th>상태</th><th>우선순위</th><th>보기</th></tr></thead>
              <tbody>
                ${tasks.length ? tasks.map((task) => `
                  <tr>
                    <td>${escapeHtml(task.public_id || "")}</td>
                    <td>${escapeHtml(task.title || "")}<div>${escapeHtml(task.description || "")}</div></td>
                    <td>${escapeHtml(task.assignee_name || "미정")}</td>
                    <td>${escapeHtml(crmDueDateText(task))}</td>
                    <td>${crmStatusBadge(task.status)}</td>
                    <td>${crmPriorityBadge(task.priority)}</td>
                    <td><button class="crm-mini-button" type="button" data-focus-open-task="${escapeHtml(task.id)}">열기</button></td>
                  </tr>
                `).join("") : `<tr><td colspan="7">표시할 관련 업무가 없습니다.</td></tr>`}
              </tbody>
            </table>
          </div>
        </section>
      `;
    }

    function openCrmProjectWidget(projectKey) {
      const project = findCrmProjectByKey(projectKey);
      if (!project) {
        openFocusWidget({
          kicker: "프로젝트 추적기",
          title: "프로젝트를 찾지 못했습니다",
          body: `<p class="focus-widget-text">현재 대시보드에 로드된 프로젝트 목록에서 찾지 못했습니다. 대시보드를 새로고침한 뒤 다시 시도해줘.</p>`,
        });
        return;
      }
      focusWidgetTaskId = "";
      focusWidgetEmployeeId = "";
      openFocusWidget({
        kicker: "프로젝트 추적기",
        title: project.project_name || "프로젝트 미지정",
        subtitle: `${projectProgressLabel(project)} · ${project.open_tasks || 0}건 진행 · ${project.total_tasks || 0}건 전체`,
        body: focusWidgetProjectBody(project),
      });
    }

    function applyCrmProjectFilter(projectKey) {
      const project = findCrmProjectByKey(projectKey);
      if (!project) return;
      crmActiveTaskViewId = "";
      writeCrmTaskFilters({
        ...crmTaskFilterDefaults(),
        q: project.project_name || "",
        open_only: "",
        sort: "due",
      });
      renderCrmSavedViews();
      closeFocusWidget();
      setCrmTab("tasks");
      loadCrmTasks().catch((error) => setCrmMessage(error.message, true));
    }

    function openNoticeWidget() {
      const payload = noticePayload();
      focusWidgetTaskId = "";
      focusWidgetEmployeeId = "";
      const meta = [shortKoreanDate(payload.date), payload.owner ? `담당 ${payload.owner}` : ""].filter(Boolean).join(" / ");
      openFocusWidget({
        kicker: "공지사항",
        title: payload.title || "등록된 공지 없음",
        subtitle: meta,
        body: `
          <section class="focus-widget-section">
            <div class="focus-widget-section-title">공지 내용</div>
            <p class="focus-widget-text">${escapeHtml(payload.body || "공지사항 입력 버튼을 눌러 내용을 입력해주세요.")}</p>
          </section>
        `,
      });
    }

    function openCompanyCardWidget(card) {
      const title = card.querySelector(".company-card-head span")?.textContent?.trim() || "상세 보기";
      const body = card.querySelector(".company-card-body")?.textContent?.trim() || "표시할 내용이 없습니다.";
      focusWidgetTaskId = "";
      focusWidgetEmployeeId = "";
      openFocusWidget({
        kicker: "사규/가이드",
        title,
        body: `<p class="focus-widget-text">${escapeHtml(body)}</p>`,
      });
    }

    function isInteractiveTarget(target) {
      return Boolean(target?.closest("button, a, input, textarea, select, label, [contenteditable='true']"));
    }

    function isCardActivationKey(event) {
      return ["Enter", " "].includes(event.key);
    }

    function renderCrmTasks() {
      renderCrmTaskBoardStats();
      if (!crmTasks.length) {
        crmSelectedTaskId = "";
        crmTaskBody.innerHTML = CRM_TASK_STATUSES.map((status) => `
          <section class="crm-kanban-column">
            <div class="crm-kanban-head"><span>${escapeHtml(status)}</span><span class="crm-kanban-count">0</span></div>
            <div class="crm-kanban-list"><div class="crm-kanban-empty">조회된 업무가 없습니다.</div></div>
          </section>
        `).join("");
        renderCrmTaskDetail(null);
        return;
      }
      if (!crmTasks.some((task) => String(task.id) === String(crmSelectedTaskId))) {
        crmSelectedTaskId = String(crmTasks[0].id);
      }
      crmTaskBody.innerHTML = CRM_TASK_STATUSES.map((status) => {
        const items = crmTasks.filter((task) => (task.status || "대기") === status);
        return `
          <section class="crm-kanban-column">
            <div class="crm-kanban-head"><span>${escapeHtml(status)}</span><span class="crm-kanban-count">${escapeHtml(items.length)}</span></div>
            <div class="crm-kanban-list">
              ${items.length ? items.map(renderCrmTaskCard).join("") : `<div class="crm-kanban-empty">${escapeHtml(status)} 업무가 없습니다.</div>`}
            </div>
          </section>
        `;
      }).join("");
      const selectedTask = crmTasks.find((task) => String(task.id) === String(crmSelectedTaskId));
      renderCrmTaskDetail(selectedTask);
      if (selectedTask) ensureCrmTaskComments(selectedTask.id).catch(() => {});
    }

    function renderCrmMineTasks() {
      if (!crmMineTaskBody) return;
      const today = todayString();
      const openTasks = crmMineTasks.filter((task) => task.status !== "완료");
      const dueToday = openTasks.filter((task) => crmDueDateOnly(task) === today).length;
      const overdue = openTasks.filter((task) => {
        const day = crmDueDateOnly(task);
        return day && day < today;
      }).length;
      if (crmMineStats) {
        crmMineStats.innerHTML = [
          ["내 미완료", openTasks.length, "OPEN"],
          ["오늘 마감", dueToday, "DUE"],
          ["지연", overdue, "LATE"],
          ["전체 배정", crmMineTasks.length, "ALL"],
        ].map(([label, value, icon]) => `
          <article class="crm-board-stat">
            <div><span>${escapeHtml(label)}</span><strong>${escapeHtml(value)}</strong></div>
            <i>${escapeHtml(icon)}</i>
          </article>
        `).join("");
      }
      crmMineTaskBody.innerHTML = openTasks.length ? openTasks.map((task) => `
        <tr>
          <td>${escapeHtml(task.public_id || "")}</td>
          <td class="left">${escapeHtml(crmTaskContextLabel(task))}</td>
          <td class="left">${escapeHtml(task.title)}</td>
          <td>${escapeHtml(crmDueDateText(task))}</td>
          <td>${crmStatusBadge(task.status)}</td>
          <td>${crmPriorityBadge(task.priority)}</td>
          <td>
            <span class="crm-mini-actions">
              <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="진행중">확인</button>
              <button class="crm-mini-button primary" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="완료">완료</button>
              <button class="crm-mini-button" type="button" data-crm-task-status="${escapeHtml(task.id)}" data-status="보류">보류</button>
              <button class="crm-mini-button" type="button" data-crm-task-comment="${escapeHtml(task.id)}">댓글</button>
            </span>
          </td>
        </tr>
      `).join("") : `<tr><td colspan="7">내 미완료 업무가 없습니다.</td></tr>`;
    }

    async function loadCrmTasks() {
      if (!can("crm_view")) return;
      const filters = readCrmTaskFilters();
      const params = new URLSearchParams({ limit: "2000" });
      if (filters.q) params.set("q", filters.q);
      if (filters.status) params.set("status", filters.status);
      if (filters.assignee_user_id) params.set("assignee_user_id", filters.assignee_user_id);
      if (filters.priority) params.set("priority", filters.priority);
      if (filters.due) params.set("due", filters.due);
      if (filters.source) params.set("source", filters.source);
      if (filters.open_only) params.set("open_only", "1");
      if (filters.sort) params.set("sort", filters.sort);
      const data = await crmFetchJson(`/api/crm-tasks?${params.toString()}`);
      crmTasks = data.tasks || [];
      renderCrmTasks();
    }

    async function loadCrmMineTasks() {
      if (!can("crm_view")) return;
      const data = await crmFetchJson("/api/crm-tasks?mine=1&open_only=1&sort=due&limit=300");
      crmMineTasks = data.tasks || [];
      renderCrmMineTasks();
    }

    async function saveCrmTaskForm(event) {
      event.preventDefault();
      if (!can("crm_manage")) return;
      const selectedAccount = crmAccounts.find((account) => String(account.id) === String(crmTaskAccount.value));
      const selectedUser = crmUsers.find((user) => String(user.id) === String(crmTaskAssignee.value));
      const payload = {
        id: crmTaskId.value,
        account_id: crmTaskAccount.value,
        account_name: crmTaskAccountName.value.trim() || selectedAccount?.name || "",
        title: crmTaskTitle.value.trim(),
        assignee_user_id: crmTaskAssignee.value,
        assignee_name: selectedUser?.display_name || "",
        due_at: crmTaskDue.value.trim(),
        priority: crmTaskPriority.value,
        status: crmTaskStatus.value,
        description: crmTaskDescription.value.trim(),
      };
      const data = await crmFetchJson("/api/crm-task-save", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      setCrmMessage(data.message || "업무를 저장했습니다.");
      resetCrmTaskForm();
      setCrmTaskFormOpen(false);
      await loadCrmAccounts();
      await loadCrmTasks();
      if (crmActiveTab === "mine") await loadCrmMineTasks();
      if (companyActiveTab === "staff") await loadCompanyStaffDashboard().catch(() => {});
      if (companyActiveTab === "calendar") await loadCompanyCalendar().catch(() => {});
      await loadCrmDashboard();
    }

    async function updateCrmTaskStatus(taskId, status) {
      const data = await crmFetchJson("/api/crm-task-status", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ id: taskId, status }),
      });
      setCrmMessage(data.message || "업무 상태를 저장했습니다.");
      delete crmTaskComments[String(taskId || "")];
      await loadCrmTasks();
      await ensureCrmTaskComments(taskId, true);
      if (crmActiveTab === "mine") await loadCrmMineTasks();
      if (companyActiveTab === "staff") await loadCompanyStaffDashboard().catch(() => {});
      if (companyActiveTab === "calendar") await loadCompanyCalendar().catch(() => {});
      await loadCrmDashboard();
    }

    function openCrmTaskComment(taskId) {
      const task = findCrmTaskById(taskId);
      crmSelectedTaskId = String(taskId || "");
      if (crmActiveTab !== "tasks") {
        if (task) {
          crmActiveTaskViewId = "";
          writeCrmTaskFilters({
            ...crmTaskFilterDefaults(),
            q: task.public_id || task.title || "",
            open_only: task.status === "완료" ? "" : "1",
          });
          renderCrmSavedViews();
        }
        setCrmTab("tasks");
      }
      else renderCrmTasks();
      ensureCrmTaskComments(taskId).finally(() => setTimeout(() => focusCrmTaskCommentForm(taskId), 0));
    }

    async function addCrmTaskComment(taskId, body) {
      const text = String(body || "").trim();
      if (!text) {
        focusCrmTaskCommentForm(taskId);
        return;
      }
      const data = await crmFetchJson("/api/crm-task-comment", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ id: taskId, body: text }),
      });
      setCrmMessage(data.message || "댓글을 저장했습니다.");
      delete crmTaskComments[String(taskId || "")];
      await loadCrmTasks();
      await ensureCrmTaskComments(taskId, true);
      if (crmActiveTab === "mine") await loadCrmMineTasks();
      if (companyActiveTab === "staff") await loadCompanyStaffDashboard().catch(() => {});
      await loadCrmDashboard();
    }

    function renderCrmMessengerUsers(payload) {
      crmUsers = payload.users || crmUsers;
      renderCrmUserOptions();
      const mappings = payload.mappings || [];
      crmMessengerUserBody.innerHTML = mappings.length ? mappings.map((item) => `
        <tr>
          <td>${escapeHtml(item.platform)}</td>
          <td class="left">${escapeHtml(item.sender_key)}</td>
          <td>${escapeHtml(item.display_name)}</td>
          <td>${escapeHtml(item.workhub_display_name || item.username)}</td>
        </tr>
      `).join("") : `<tr><td colspan="4">등록된 메신저 사용자 매핑이 없습니다.</td></tr>`;
    }

    function crmWebhookSamplePayloadText() {
      return JSON.stringify({
        userRequest: {
          utterance: "도움말",
          user: {
            id: "kakao-test-user-key",
            properties: {
              nickname: "테스트 직원",
            },
          },
        },
      }, null, 2);
    }

    function renderCrmWebhookSetup(payload) {
      const webhook = payload.webhook || {};
      const webhookUrl = webhook.url || webhook.path || "/api/crm-messenger-webhook";
      const webhookToken = webhook.token || "";
      const samplePayload = crmWebhookSamplePayloadText();
      if (crmWebhookUrl) crmWebhookUrl.textContent = webhookUrl;
      if (crmWebhookToken) crmWebhookToken.textContent = webhookToken || "토큰을 불러오지 못했습니다.";
      if (crmWebhookSamplePayload) crmWebhookSamplePayload.textContent = samplePayload;
      if (crmWebhookCurl) {
        crmWebhookCurl.textContent = [
          `curl -X POST "${webhookUrl}"`,
          `  -H "Content-Type: application/json"`,
          `  -H "X-Workhub-Webhook-Token: ${webhookToken}"`,
          `  -d '${samplePayload.replaceAll("\n", "")}'`,
        ].join(" \\\n");
      }
    }

    function renderCrmMessageEvents(payload) {
      const events = payload.events || [];
      renderCrmWebhookSetup(payload);
      crmMessageEventBody.innerHTML = events.length ? events.map((event) => `
        <tr>
          <td>${escapeHtml(event.created_at)}</td>
          <td>${escapeHtml(event.platform)}</td>
          <td>${escapeHtml(event.sender_name || event.sender_key)}</td>
          <td>${escapeHtml(event.result)}</td>
          <td class="left">${escapeHtml(event.text)}</td>
          <td class="left">${escapeHtml(event.error)}</td>
          <td>
            <span class="crm-mini-actions">
              ${event.sender_key ? `<button class="crm-mini-button" type="button" data-crm-copy-text="${escapeHtml(event.sender_key)}">키 복사</button>` : ""}
              ${event.result === "거절" && event.sender_key ? `<button class="crm-mini-button primary" type="button" data-crm-map-sender="${escapeHtml(event.sender_key)}" data-platform="${escapeHtml(event.platform || "kakao")}" data-sender-name="${escapeHtml(event.sender_name || "")}">매핑</button>` : ""}
            </span>
          </td>
        </tr>
      `).join("") : `<tr><td colspan="7">메신저 연동 기록이 없습니다.</td></tr>`;
    }

    async function loadCrmMessenger() {
      if (!can("crm_message_manage")) return;
      const [usersPayload, eventsPayload] = await Promise.all([
        crmFetchJson("/api/crm-messenger-users"),
        crmFetchJson("/api/crm-message-events"),
      ]);
      renderCrmMessengerUsers(usersPayload);
      renderCrmMessageEvents(eventsPayload);
    }

    async function saveCrmMessengerForm(event) {
      event.preventDefault();
      if (!can("crm_message_manage")) return;
      const payload = {
        platform: crmMessengerPlatform.value,
        sender_key: crmMessengerSenderKey.value.trim(),
        display_name: crmMessengerDisplayName.value.trim(),
        user_id: crmMessengerUser.value,
      };
      const data = await crmFetchJson("/api/crm-messenger-user-save", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      setCrmMessage(data.message || "메신저 사용자 매핑을 저장했습니다.");
      crmMessengerSenderKey.value = "";
      crmMessengerDisplayName.value = "";
      await loadCrmMessenger();
    }

    async function rotateCrmWebhookToken() {
      if (!can("crm_message_manage")) return;
      if (!confirm("웹훅 토큰을 재발급할까요? 기존 카카오 스킬 헤더 값은 즉시 실패합니다.")) return;
      const data = await crmFetchJson("/api/crm-webhook-token-rotate", { method: "POST" });
      setCrmMessage(data.message || "웹훅 토큰을 재발급했습니다.");
      renderCrmWebhookSetup(data);
      await loadCrmMessenger();
    }

    async function loadCrmUsersForForms() {
      if (!can("crm_view")) return;
      const data = await crmFetchJson("/api/crm-messenger-users");
      crmUsers = data.users || [];
      renderCrmUserOptions();
      renderCrmMessengerUsers(data);
      if (crmStatAccounts) crmStatAccounts.textContent = crmUsers.length || 0;
    }

    async function loadCrmAll() {
      if (!can("crm_view")) return;
      setCrmMessage("");
      renderCrmSavedViews();
      await Promise.all([
        loadCrmDashboard(),
        loadCrmAccounts(),
        loadCrmTasks(),
        loadCrmSavedViews().catch(() => {}),
        loadCrmUsersForForms().catch(() => {}),
      ]);
      if (crmActiveTab === "mine") await loadCrmMineTasks();
      if (crmActiveTab === "messages" && can("crm_message_manage")) await loadCrmMessenger();
    }

    function todayStatusDate() {
      const now = new Date();
      return `${now.getMonth() + 1}/${now.getDate()}`;
    }

    function parseDateParts(value) {
      const text = String(value || "").trim();
      if (!text) return null;
      let match = text.match(/(\d{4})[-./년\s]+(\d{1,2})[-./월\s]+(\d{1,2})/);
      if (match) {
        return {
          year: match[1],
          month: String(Number(match[2])).padStart(2, "0"),
          day: String(Number(match[3])).padStart(2, "0"),
        };
      }
      match = text.match(/(\d{1,2})\s*월\s*(\d{1,2})\s*일?/);
      if (match) {
        return {
          year: "",
          month: String(Number(match[1])).padStart(2, "0"),
          day: String(Number(match[2])).padStart(2, "0"),
        };
      }
      match = text.match(/^(\d{1,2})[./-](\d{1,2})$/);
      if (match) {
        return {
          year: "",
          month: String(Number(match[1])).padStart(2, "0"),
          day: String(Number(match[2])).padStart(2, "0"),
        };
      }
      return null;
    }

    function shortKoreanDate(value) {
      const parts = parseDateParts(value);
      return parts ? `${Number(parts.month)}월 ${Number(parts.day)}일` : String(value || "");
    }

    function fullDateForSave(displayValue, rawValue) {
      const displayParts = parseDateParts(displayValue);
      if (!displayParts) return String(displayValue || "").trim();
      const rawParts = parseDateParts(rawValue);
      const year = displayParts.year || rawParts?.year || String(new Date().getFullYear());
      return `${year}-${displayParts.month}-${displayParts.day}`;
    }

    function ledgerStatusOptions(currentStatus = "") {
      const dateText = todayStatusDate();
      const statuses = [
        `회수지시(${dateText})`,
        `회수 완료(${dateText})`,
        `재발송 완료(${dateText})`,
        `재발송 완료(${dateText})/회수지시(${dateText})`,
        "전체 처리완료",
      ];
      const normalizedCurrent = currentStatus || statuses[0];
      return statuses.includes(normalizedCurrent) ? statuses : [normalizedCurrent, ...statuses];
    }

    const csTypeOptions = ["변심반품", "불량반품", "불량교환", "불량재출고(미회수)", "오출고(오배송)"];

    function selectOptions(options, currentValue = "") {
      const normalizedCurrent = currentValue || "";
      const optionList = normalizedCurrent && !options.includes(normalizedCurrent)
        ? [normalizedCurrent, ...options]
        : options;
      return optionList.map((option) => (
        `<option value="${escapeHtml(option)}" ${option === normalizedCurrent ? "selected" : ""}>${escapeHtml(option)}</option>`
      )).join("");
    }

    function isCompletedByValues(typeValue, statusValue) {
      const type = String(typeValue || "").replaceAll(" ", "").trim();
      const status = String(statusValue || "").replaceAll(" ", "").trim();
      if (status.includes("전체처리완료")) return true;
      if ((type === "변심반품" || type === "변신반품" || type === "불량반품") && status.includes("회수완료")) return true;
      if ((type === "불량교환" || type === "오출고(오배송)") && status.includes("전체처리완료")) return true;
      if (type === "불량재출고(미회수)" && status.includes("재발송완료")) return true;
      return false;
    }

    function isCompletedCsCase(csCase) {
      return isCompletedByValues(csCase.cs_type, csCase.status);
    }

    function fieldValue(element) {
      if (!element) return "";
      if ("value" in element) return String(element.value || "").trim();
      return String(element.dataset.value ?? element.textContent ?? "").trim();
    }

    function updateLedgerRowCompletion(row) {
      if (!row) return;
      const status = fieldValue(row.querySelector('[data-field="status"]'));
      const csType = fieldValue(row.querySelector('[data-field="cs_type"]'));
      row.classList.toggle("completed-cs", isCompletedByValues(csType, status));
    }

    function ledgerFieldValue(csCase, field) {
      if (field === "purchase_vendor") return csCase.purchase_vendor || csCase.vendor_name || "";
      if (field === "occurred_at") return csCase.occurred_at || csCase.created_at || "";
      if (field === "original_invoice") return csCase.original_invoice || csCase.original_info || "";
      return csCase[field] || "";
    }

    function matchesLedgerFilters(csCase) {
      const toolbarStatus = ledgerStatusFilter.value.trim().toLowerCase();
      if (toolbarStatus && !String(csCase.status || "").toLowerCase().includes(toolbarStatus)) return false;
      return Object.entries(ledgerFilters).every(([field, filterValue]) => {
        const value = String(filterValue || "").trim().toLowerCase();
        if (!value) return true;
        return String(ledgerFieldValue(csCase, field)).toLowerCase().includes(value);
      });
    }

    function applyLedgerFilters() {
      const filtered = ledgerCases.filter(matchesLedgerFilters);
      renderLedger(filtered);
      ledgerFilterButtons.forEach((button) => {
        const field = button.dataset.ledgerFilterButton;
        button.classList.toggle("active", Boolean(ledgerFilters[field]));
      });
      if (currentMode === "ledger") notice.textContent = `${filtered.length}건 조회되었습니다.`;
    }

    function managementFieldValue(record, field) {
      return record[field] || "";
    }

    function matchesManagementFilters(record) {
      return Object.entries(managementFilters).every(([field, filterValue]) => {
        const value = String(filterValue || "").trim().toLowerCase();
        if (!value) return true;
        return String(managementFieldValue(record, field)).toLowerCase().includes(value);
      });
    }

    function applyManagementFilters() {
      const filtered = managementRecords.filter(matchesManagementFilters);
      renderManagement(filtered);
      managementFilterButtons.forEach((button) => {
        const field = button.dataset.managementFilterButton;
        button.classList.toggle("active", Boolean(managementFilters[field]));
      });
      if (currentMode === "management") notice.textContent = `${filtered.length}건 조회되었습니다.`;
    }

    function renderLedgerFilterOptions(field, searchText = "") {
      const normalizedSearch = searchText.trim().toLowerCase();
      const values = Array.from(new Set(
        ledgerCases
          .map((csCase) => String(ledgerFieldValue(csCase, field) || "").trim())
          .filter(Boolean)
      )).sort((left, right) => left.localeCompare(right, "ko"));
      const filteredValues = values
        .filter((value) => !normalizedSearch || value.toLowerCase().includes(normalizedSearch))
        .slice(0, 220);
      ledgerFilterOptions.innerHTML = filteredValues.length
        ? filteredValues.map((value) => (
          `<button class="ledger-filter-option" type="button" data-filter-value="${escapeHtml(value)}">${escapeHtml(value)}</button>`
        )).join("")
        : `<button class="ledger-filter-option" type="button" disabled>표시할 값이 없습니다.</button>`;
    }

    function renderManagementFilterOptions(field, searchText = "") {
      const normalizedSearch = searchText.trim().toLowerCase();
      const values = Array.from(new Set(
        managementRecords
          .map((record) => String(managementFieldValue(record, field) || "").trim())
          .filter(Boolean)
      )).sort((left, right) => left.localeCompare(right, "ko"));
      const filteredValues = values
        .filter((value) => !normalizedSearch || value.toLowerCase().includes(normalizedSearch))
        .slice(0, 220);
      ledgerFilterOptions.innerHTML = filteredValues.length
        ? filteredValues.map((value) => (
          `<button class="ledger-filter-option" type="button" data-filter-value="${escapeHtml(value)}">${escapeHtml(value)}</button>`
        )).join("")
        : `<button class="ledger-filter-option" type="button" disabled>표시할 값이 없습니다.</button>`;
    }

    function openLedgerFilter(button) {
      activeManagementFilterField = "";
      activeLedgerFilterField = button.dataset.ledgerFilterButton || "";
      ledgerFilterTitle.textContent = `${button.dataset.label || "필터"} 필터`;
      ledgerFilterSearch.value = ledgerFilters[activeLedgerFilterField] || "";
      renderLedgerFilterOptions(activeLedgerFilterField, ledgerFilterSearch.value);
      const rect = button.getBoundingClientRect();
      ledgerFilterPopover.style.left = `${Math.min(rect.left, window.innerWidth - 280)}px`;
      ledgerFilterPopover.style.top = `${Math.min(rect.bottom + 6, window.innerHeight - 410)}px`;
      ledgerFilterPopover.classList.add("open");
      ledgerFilterSearch?.focus();
      ledgerFilterSearch.select();
    }

    function openManagementFilter(button) {
      activeLedgerFilterField = "";
      activeManagementFilterField = button.dataset.managementFilterButton || "";
      ledgerFilterTitle.textContent = `${button.dataset.label || "필터"} 필터`;
      ledgerFilterSearch.value = managementFilters[activeManagementFilterField] || "";
      renderManagementFilterOptions(activeManagementFilterField, ledgerFilterSearch.value);
      const rect = button.getBoundingClientRect();
      ledgerFilterPopover.style.left = `${Math.min(rect.left, window.innerWidth - 280)}px`;
      ledgerFilterPopover.style.top = `${Math.min(rect.bottom + 6, window.innerHeight - 410)}px`;
      ledgerFilterPopover.classList.add("open");
      ledgerFilterSearch?.focus();
      ledgerFilterSearch.select();
    }

    function closeLedgerFilter() {
      ledgerFilterPopover.classList.remove("open");
      activeLedgerFilterField = "";
      activeManagementFilterField = "";
    }

    function markRowDirty(row, dirty = true) {
      if (!row || !row.dataset) return;
      if (dirty) {
        row.dataset.dirty = "1";
        row.classList.add("row-dirty");
      } else {
        delete row.dataset.dirty;
        row.classList.remove("row-dirty");
      }
    }

    function applyRowPermissions(row) {
      if (!row) return;
      if (!can("ledger_edit")) {
        row.querySelectorAll("[data-field], [data-management-field]").forEach((input) => {
          input.disabled = true;
        });
        row.querySelectorAll(".editable-cell").forEach((cell) => {
          cell.dataset.readonly = "1";
        });
      }
      row.querySelectorAll(".management-cs-button").forEach((button) => {
        setHidden(button, !can("cs_receive"));
      });
    }

    function cellDisplayValue(value, options = {}) {
      const raw = String(value || "");
      if (options.date) return shortKoreanDate(raw);
      return raw;
    }

    function editableCell({ scope, field, label, value, align = "", date = false, input = "text", options = [] }) {
      const displayValue = cellDisplayValue(value, { date });
      const className = `editable-cell ${align}`.trim();
      const dataAttr = scope === "management" ? "data-management-field" : "data-field";
      const optionData = options.length ? ` data-options="${escapeHtml(JSON.stringify(options))}"` : "";
      return `<td class="${className}" ${dataAttr}="${escapeHtml(field)}" data-label="${escapeHtml(label)}" data-value="${escapeHtml(value)}" data-input="${escapeHtml(input)}"${date ? ` data-raw-date="${escapeHtml(value)}" data-date="1"` : ""}${optionData}>${escapeHtml(displayValue)}</td>`;
    }

    function selectedEditorParts(scope) {
      if (scope === "management") {
        return {
          bar: managementCellEditBar,
          label: managementCellEditLabel,
          mount: managementCellEditMount,
        };
      }
      return {
        bar: ledgerCellEditBar,
        label: ledgerCellEditLabel,
        mount: ledgerCellEditMount,
      };
    }

    function closeCellEditor(scope) {
      const selected = activeCellEditors[scope];
      if (selected?.cell) selected.cell.classList.remove("selected-cell");
      activeCellEditors[scope] = null;
      const parts = selectedEditorParts(scope);
      parts.bar?.classList.remove("open");
      if (parts.mount) parts.mount.innerHTML = "";
      if (parts.label) parts.label.textContent = "셀 선택";
    }

    function editorOptionsFromCell(cell) {
      try {
        return JSON.parse(cell.dataset.options || "[]");
      } catch {
        return [];
      }
    }

    function createCellEditorControl(cell) {
      const inputType = cell.dataset.input || "text";
      const currentValue = cell.dataset.value || cell.textContent.trim();
      if (inputType === "select") {
        const select = document.createElement("select");
        select.className = "cell-edit-control";
        editorOptionsFromCell(cell).forEach((option) => {
          const item = document.createElement("option");
          item.value = option;
          item.textContent = option || "선택";
          if (option === currentValue) item.selected = true;
          select.appendChild(item);
        });
        return select;
      }
      if (inputType === "textarea") {
        const textarea = document.createElement("textarea");
        textarea.className = "cell-edit-control";
        textarea.value = currentValue;
        return textarea;
      }
      const input = document.createElement("input");
      input.className = "cell-edit-control";
      input.type = inputType === "date" ? "date" : "text";
      input.value = currentValue;
      return input;
    }

    function openCellEditor(scope, cell) {
      if (!cell || cell.dataset.readonly === "1" || !can("ledger_edit")) return;
      closeCellEditor(scope);
      closeLedgerFilter();
      const row = cell.closest("tr");
      if (!row) return;
      const parts = selectedEditorParts(scope);
      const label = cell.dataset.label || cell.dataset.field || cell.dataset.managementField || "선택 셀";
      const rowHint = scope === "management"
        ? [row.querySelector('[data-management-field="order_date"]')?.textContent.trim(), row.querySelector('[data-management-field="receiver_name"]')?.textContent.trim(), row.querySelector('[data-management-field="product_name"]')?.textContent.trim()].filter(Boolean).join(" / ")
        : [textFromCell(row, 1), textFromCell(row, 14), textFromCell(row, 16)].filter(Boolean).join(" / ");
      const control = createCellEditorControl(cell);
      parts.mount.innerHTML = "";
      parts.mount.appendChild(control);
      parts.label.textContent = rowHint ? `${label} · ${rowHint}` : label;
      parts.bar.classList.add("open");
      cell.classList.add("selected-cell");
      activeCellEditors[scope] = { cell, control };
      setTimeout(() => {
        control?.focus();
        if (control.select) control.select();
      }, 0);
    }

    function applyCellEditor(scope) {
      const selected = activeCellEditors[scope];
      if (!selected?.cell || !selected.control) return;
      const { cell, control } = selected;
      const row = cell.closest("tr");
      const value = control.value || "";
      cell.dataset.value = value;
      if (cell.dataset.date === "1") cell.dataset.rawDate = value;
      cell.textContent = cellDisplayValue(value, { date: cell.dataset.date === "1" });
      markRowDirty(row, true);
      const checkbox = row?.querySelector("[data-row-check]");
      if (checkbox) checkbox.checked = true;
      if (scope === "ledger") updateLedgerRowCompletion(row);
      if (activeLedgerFilterField || activeManagementFilterField) refreshActiveFilterOptions();
      notice.textContent = `${cell.dataset.label || "선택 셀"} 값을 반영했습니다. 저장하려면 체크된 항목 저장 버튼을 눌러주세요.`;
    }

    function dirtyRows(container, rowSelector) {
      return Array.from(container.querySelectorAll(`${rowSelector}[data-dirty="1"]`));
    }

    function selectedRows(container, rowSelector) {
      return Array.from(container.querySelectorAll(rowSelector))
        .filter((row) => row.querySelector("[data-row-check]")?.checked);
    }

    function setLedgerFilter(value) {
      if (!activeLedgerFilterField) return;
      const normalized = String(value || "").trim();
      if (normalized) ledgerFilters[activeLedgerFilterField] = normalized;
      else delete ledgerFilters[activeLedgerFilterField];
      applyLedgerFilters();
      closeLedgerFilter();
    }

    function setManagementFilter(value) {
      if (!activeManagementFilterField) return;
      const normalized = String(value || "").trim();
      if (normalized) managementFilters[activeManagementFilterField] = normalized;
      else delete managementFilters[activeManagementFilterField];
      applyManagementFilters();
      closeLedgerFilter();
    }

    function setActivePopoverFilter(value) {
      if (activeManagementFilterField) setManagementFilter(value);
      else setLedgerFilter(value);
    }

    function refreshActiveFilterOptions() {
      if (activeManagementFilterField) {
        renderManagementFilterOptions(activeManagementFilterField, ledgerFilterSearch.value);
      } else {
        renderLedgerFilterOptions(activeLedgerFilterField, ledgerFilterSearch.value);
      }
    }

    function resetCsFormInputs() {
      vendorContactSelect.value = "";
      vendorTypeSelect.value = "purchase";
      recipientEmailInput.value = "";
      vendorNameInput.value = "";
      if (csAttachmentInput) csAttachmentInput.value = "";
      updateCsAttachmentSummary();
      csOriginInput.value = "";
      csProductInput.value = "";
      csReceiverInput.value = "";
      csPhoneInput.value = "";
      csAddressInput.value = "";
      csTypeInput.value = "";
      csContentInput.value = "";
      csSubjectInput.value = defaultCsSubject();
      refreshCsBody();
      setSelectedStockVendor(null);
      if (stockVendorTree) stockVendorTree.hidden = true;
      if (stockNoticeDateInput) stockNoticeDateInput.value = todayString();
      [
        stockInboundProductInput,
        stockInboundScheduleInput,
        stockOutboundAvailableInput,
        stockInboundNoteInput,
        stockSoldoutProductInput,
        stockOutboundBlockedInput,
        stockRestockScheduleInput,
        stockSoldoutNoteInput,
      ].forEach((input) => {
        if (input) input.value = "";
      });
      if (stockSubjectInput) stockSubjectInput.value = "입고 및 품절 공지";
      refreshStockNoticeBody();
      activeCsCaseId = "";
    }

    function openLedgerCsPopup() {
      closeLedgerFilter();
      resetCsFormInputs();
      csFields.classList.add("ledger-cs-popup");
      csFields.style.display = "block";
      loadVendorContacts();
      loadCsCases();
      notice.textContent = "새 CS 내용을 입력한 뒤 CS건 DB 저장을 눌러주세요.";
      setTimeout(() => vendorNameInput?.focus(), 0);
    }

    function closeLedgerCsPopup() {
      csFields.classList.remove("ledger-cs-popup");
      if (currentMode === "ledger") csFields.style.display = "none";
    }

    function renderLedger(cases) {
      closeCellEditor("ledger");
      ledgerBody.innerHTML = "";
      if (!cases || cases.length === 0) {
        const row = document.createElement("tr");
        row.innerHTML = `<td colspan="21">조회된 CS건이 없습니다.</td>`;
        ledgerBody.appendChild(row);
        return;
      }
      cases.forEach((csCase) => {
        const row = document.createElement("tr");
        row.dataset.caseId = csCase.id;
        const statusValue = csCase.status || ledgerStatusOptions()[0];
        const statusSelectOptions = ledgerStatusOptions(statusValue);
        const csTypeSelectOptions = ["", ...csTypeOptions];
        if (isCompletedCsCase(csCase)) row.classList.add("completed-cs");
        row.innerHTML = `
          <td><input class="ledger-check" type="checkbox" data-row-check /></td>
          <td>${escapeHtml(csCase.occurred_at || csCase.created_at)}</td>
          <td>${escapeHtml(csCase.sales_vendor)}</td>
          <td>${escapeHtml(csCase.purchase_vendor || csCase.vendor_name)}</td>
          ${editableCell({ scope: "ledger", field: "status", label: "처리진행상태", value: statusValue, input: "select", options: statusSelectOptions })}
          <td>${escapeHtml(csCase.completed_at)}</td>
          ${editableCell({ scope: "ledger", field: "cs_type", label: "처리내용", value: csCase.cs_type, input: "select", options: csTypeSelectOptions })}
          <td class="left">${escapeHtml(csCase.cs_content)}</td>
          ${editableCell({ scope: "ledger", field: "reship_invoice", label: "재발송운송장번호", value: csCase.reship_invoice })}
          ${editableCell({ scope: "ledger", field: "return_invoice", label: "회수운송장번호", value: csCase.return_invoice })}
          <td data-full-date="${escapeHtml(csCase.order_date)}">${escapeHtml(shortKoreanDate(csCase.order_date))}</td>
          <td data-full-date="${escapeHtml(csCase.ship_date)}">${escapeHtml(shortKoreanDate(csCase.ship_date))}</td>
          <td>${escapeHtml(csCase.orderer_name)}</td>
          <td>${escapeHtml(csCase.orderer_phone)}</td>
          <td>${escapeHtml(csCase.receiver_name)}</td>
          <td>${escapeHtml(csCase.receiver_phone)}</td>
          <td class="left">${escapeHtml(csCase.product_name)}</td>
          <td>${escapeHtml(csCase.quantity)}</td>
          <td class="left">${escapeHtml(csCase.receiver_address)}</td>
          <td>${escapeHtml(csCase.courier)}</td>
          <td>${escapeHtml(csCase.original_invoice || csCase.original_info)}</td>
        `;
        applyRowPermissions(row);
        ledgerBody.appendChild(row);
      });
    }

    async function loadCsCases() {
      try {
        const response = await fetch("/api/cs-cases");
        if (!response.ok) return;
        const data = await response.json();
        renderCsCases(data.cases || []);
      } catch {
        renderCsCases([]);
      }
    }

    async function loadLedgerCases() {
      const query = ledgerSearchInput.value.trim();
      const params = new URLSearchParams({ limit: ledgerPageSize.value || "100" });
      if (query) params.set("q", query);
      if (ledgerYearFilter?.value) params.set("year", ledgerYearFilter.value);
      if (ledgerMonthFilter?.value) params.set("month", ledgerMonthFilter.value);
      const url = `/api/cs-cases?${params.toString()}`;
      try {
        const response = await fetch(url);
        if (!response.ok) return;
        const data = await response.json();
        ledgerCases = data.cases || [];
        applyLedgerFilters();
      } catch {
        ledgerCases = [];
        renderLedger([]);
        notice.textContent = "CS 처리대장을 불러오지 못했습니다.";
      }
    }

    function importModeLabel(mode) {
      return mode === "replace" ? "전체 데이터 교체 업로드" : "일일 추가 업로드";
    }

    function importPreviewText(preview) {
      const lines = [
        `전체 행: ${preview.total || 0}건`,
        `추가 예정: ${preview.insertable || 0}건`,
        `이미 DB에 있는 중복: ${preview.duplicate_existing || 0}건`,
        `파일 안 중복: ${preview.duplicate_in_file || 0}건`,
      ];
      const duplicates = Array.isArray(preview.duplicates) ? preview.duplicates : [];
      if (duplicates.length) {
        lines.push("", "중복 예시");
        duplicates.slice(0, 10).forEach((item) => {
          lines.push(`- ${item.row || ""}행 · ${item.reason || "중복"} · ${item.summary || ""}`);
        });
      }
      return lines.join("\n");
    }

    function requestImportWarningApproval({ title, description, previewText, proceedLabel = "진행" }) {
      if (!importWarningDialog || !importWarningTitle || !importWarningDescription || !importWarningPreview || !importWarningCancel || !importWarningProceed) {
        return Promise.resolve(false);
      }
      importWarningTitle.textContent = title;
      importWarningDescription.textContent = description;
      importWarningPreview.textContent = previewText;
      importWarningProceed.textContent = proceedLabel;
      importWarningDialog.classList.add("open");
      importWarningDialog.setAttribute("aria-hidden", "false");
      return new Promise((resolve) => {
        const finish = (approved) => {
          importWarningDialog.classList.remove("open");
          importWarningDialog.setAttribute("aria-hidden", "true");
          importWarningCancel.removeEventListener("click", cancel);
          importWarningProceed.removeEventListener("click", proceed);
          importWarningDialog.removeEventListener("click", backdropCancel);
          resolve(approved);
        };
        const cancel = () => finish(false);
        const proceed = () => finish(true);
        const backdropCancel = (event) => {
          if (event.target === importWarningDialog) finish(false);
        };
        importWarningCancel.addEventListener("click", cancel);
        importWarningProceed.addEventListener("click", proceed);
        importWarningDialog.addEventListener("click", backdropCancel);
        importWarningProceed.focus();
      });
    }

    async function previewLedgerImport(file, mode) {
      const formData = new FormData();
      formData.append("file", file);
      formData.append("mode", mode);
      const response = await fetch("/api/cs-cases-import-preview", {
        method: "POST",
        body: formData,
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "CS 처리대장 업로드 검토에 실패했습니다.");
      return data;
    }

    async function previewManagementImport(file, mode) {
      const formData = new FormData();
      formData.append("file", file);
      formData.append("mode", mode);
      const response = await fetch("/api/management-import-preview", {
        method: "POST",
        body: formData,
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "통합관리대장 업로드 검토에 실패했습니다.");
      return data;
    }

    function correctionFieldsForRow(row) {
      const fields = new Map();
      (row.issues || []).forEach((issue) => {
        if (!issue.field) return;
        fields.set(issue.field, {
          field: issue.field,
          label: issue.label || issue.field,
          inputType: issue.input_type || "text",
          message: issue.message || "",
        });
      });
      return Array.from(fields.values());
    }

    function renderImportCorrectionRows(rows) {
      if (!importCorrectionList) return;
      importCorrectionList.innerHTML = rows.map((row, rowIndex) => {
        const fields = correctionFieldsForRow(row);
        const fieldHtml = fields.map((field) => {
          const value = row.record?.[field.field] || "";
          const inputType = field.inputType === "number" ? "number" : "text";
          return `
            <div class="import-correction-field">
              <label for="importCorrection_${rowIndex}_${escapeHtml(field.field)}">${escapeHtml(field.label)}</label>
              <input
                id="importCorrection_${rowIndex}_${escapeHtml(field.field)}"
                type="${inputType}"
                value="${escapeHtml(value)}"
                data-correction-row="${rowIndex}"
                data-correction-field="${escapeHtml(field.field)}"
              />
              <div class="import-correction-message">${escapeHtml(field.message)}</div>
            </div>
          `;
        }).join("");
        return `
          <section class="import-correction-row" data-correction-card="${rowIndex}">
            <div class="import-correction-row-head">
              <span>${escapeHtml(row.source_sheet || "")} ${escapeHtml(row.row || "")}행</span>
              <span class="import-correction-summary">${escapeHtml(row.summary || "수정 후 적용할 데이터")}</span>
            </div>
            <div class="import-correction-fields">${fieldHtml}</div>
          </section>
        `;
      }).join("");
    }

    function requestImportCorrectionApproval({ ledgerName, preview }) {
      const rows = Array.isArray(preview?.invalid_rows) ? preview.invalid_rows : [];
      if (!rows.length) return Promise.resolve([]);
      if (!importCorrectionDialog || !importCorrectionTitle || !importCorrectionDescription || !importCorrectionList || !importCorrectionCancel || !importCorrectionApply) {
        return Promise.resolve(null);
      }
      importCorrectionTitle.textContent = `${ledgerName} 업로드 전 형식 수정`;
      importCorrectionDescription.textContent = `기존 양식과 맞지 않는 행 ${rows.length}건이 있습니다. 텍스트/숫자 값을 수정한 뒤 적용해주세요.`;
      renderImportCorrectionRows(rows);
      importCorrectionDialog.classList.add("open");
      importCorrectionDialog.setAttribute("aria-hidden", "false");
      return new Promise((resolve) => {
        const finish = (payload) => {
          importCorrectionDialog.classList.remove("open");
          importCorrectionDialog.setAttribute("aria-hidden", "true");
          importCorrectionCancel.removeEventListener("click", cancel);
          importCorrectionApply.removeEventListener("click", apply);
          importCorrectionDialog.removeEventListener("click", backdropCancel);
          resolve(payload);
        };
        const cancel = () => finish(null);
        const apply = () => {
          const corrections = rows.map((row, rowIndex) => {
            const correction = {
              source_sheet: row.source_sheet || row.record?.source_sheet || "",
              source_row: row.source_row || row.record?.source_row || row.row || "",
            };
            importCorrectionList.querySelectorAll(`[data-correction-row="${rowIndex}"]`).forEach((input) => {
              correction[input.dataset.correctionField] = input.value.trim();
            });
            return correction;
          });
          finish(corrections);
        };
        const backdropCancel = (event) => {
          if (event.target === importCorrectionDialog) finish(null);
        };
        importCorrectionCancel.addEventListener("click", cancel);
        importCorrectionApply.addEventListener("click", apply);
        importCorrectionDialog.addEventListener("click", backdropCancel);
        const firstInput = importCorrectionList.querySelector("input");
        if (firstInput) firstInput.focus();
      });
    }

    async function confirmImportIfNeeded({ ledgerName, mode, preview }) {
      if (mode === "replace") {
        return requestImportWarningApproval({
          title: `${ledgerName} 전체 데이터 교체 업로드`,
          description: "현재 저장된 전체 데이터를 삭제하고 선택한 엑셀 파일 내용으로 다시 저장합니다. 승인된 관리자만 진행할 수 있습니다.",
          previewText: "기존 데이터는 교체 후 되돌릴 수 없습니다. 필요한 경우 먼저 백업을 만들어주세요.",
          proceedLabel: "전체 교체 진행",
        });
      }
      if (preview?.has_duplicates) {
        return requestImportWarningApproval({
          title: `${ledgerName} 중복 업로드 경고`,
          description: "이미 올라간 것으로 보이는 데이터가 있습니다. 진행하면 중복 건은 제외하고 새 데이터만 추가합니다.",
          previewText: importPreviewText(preview),
          proceedLabel: "새 데이터만 추가",
        });
      }
      return true;
    }

    async function uploadLedgerWorkbook() {
      const file = ledgerImportInput.files[0];
      if (!file) return;
      const mode = ledgerImportMode || "daily";
      notice.textContent = `CS 처리대장 ${importModeLabel(mode)} 검토 중입니다.`;
      try {
        const preview = await previewLedgerImport(file, mode);
        const corrections = await requestImportCorrectionApproval({ ledgerName: "CS 처리대장", preview });
        if (corrections === null) {
          notice.textContent = "CS 처리대장 업로드를 취소했습니다.";
          return;
        }
        const approved = await confirmImportIfNeeded({ ledgerName: "CS 처리대장", mode, preview });
        if (!approved) {
          notice.textContent = "CS 처리대장 업로드를 취소했습니다.";
          return;
        }
        const formData = new FormData();
        formData.append("file", file);
        formData.append("mode", mode);
        if (corrections.length) formData.append("corrections", JSON.stringify(corrections));
        notice.textContent = `CS 처리대장 ${importModeLabel(mode)} 진행 중입니다.`;
        const response = await fetch("/api/cs-cases-import", {
          method: "POST",
          body: formData,
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "CS 처리대장 업로드에 실패했습니다.");
        notice.textContent = data.message || "CS 처리대장 데이터를 업로드했습니다.";
        if (currentMode !== "ledger") {
          showWorkspace("ledger");
        } else {
          await loadLedgerCases();
        }
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        ledgerImportInput.value = "";
        ledgerImportMode = "daily";
      }
    }

    function renderManagement(records) {
      closeCellEditor("management");
      managementBody.innerHTML = "";
      if (managementSelectAll) managementSelectAll.checked = false;
      if (!records || records.length === 0) {
        const row = document.createElement("tr");
        row.innerHTML = `<td colspan="18">조회된 통합관리대장 데이터가 없습니다.</td>`;
        managementBody.appendChild(row);
        return;
      }
      const duplicateColors = [
        "#fff7d6",
        "#e8f7ee",
        "#e8f1ff",
        "#f4eaff",
        "#ffecef",
        "#e8faf8",
        "#fff0df",
        "#eef2ff",
      ];
      const duplicateCounts = new Map();
      const duplicateColorByKey = new Map();
      records.forEach((record) => {
        const dateKey = String(record.order_date || record.ship_date || "").trim();
        const invoiceKey = String(record.invoice_number || "").trim();
        if (!dateKey || !invoiceKey) return;
        const key = `${dateKey}||${invoiceKey}`;
        duplicateCounts.set(key, (duplicateCounts.get(key) || 0) + 1);
      });
      let duplicateGroupIndex = 0;
      records.forEach((record) => {
        const row = document.createElement("tr");
        row.dataset.recordId = record.id;
        const dateKey = String(record.order_date || record.ship_date || "").trim();
        const invoiceKey = String(record.invoice_number || "").trim();
        const duplicateKey = dateKey && invoiceKey ? `${dateKey}||${invoiceKey}` : "";
        if (duplicateKey && duplicateCounts.get(duplicateKey) > 1) {
          if (!duplicateColorByKey.has(duplicateKey)) {
            duplicateColorByKey.set(
              duplicateKey,
              duplicateColors[duplicateGroupIndex % duplicateColors.length]
            );
            duplicateGroupIndex += 1;
          }
          row.classList.add("management-duplicate");
          row.style.setProperty("--duplicate-row-color", duplicateColorByKey.get(duplicateKey));
        }
        const csReceived = Boolean(record.cs_received_at);
        if (csReceived) row.classList.add("management-cs-received");
        row.innerHTML = `
          <td><input class="ledger-check" type="checkbox" data-row-check /></td>
          ${editableCell({ scope: "management", field: "order_date", label: "주문일자", value: record.order_date, date: true, input: "date" })}
          ${editableCell({ scope: "management", field: "ship_date", label: "출고일", value: record.ship_date, date: true, input: "date" })}
          ${editableCell({ scope: "management", field: "purchase_vendor", label: "매입거래처", value: record.purchase_vendor })}
          ${editableCell({ scope: "management", field: "sales_vendor", label: "매출거래처", value: record.sales_vendor })}
          ${editableCell({ scope: "management", field: "transaction_type", label: "거래구분", value: record.transaction_type })}
          ${editableCell({ scope: "management", field: "ledger_checked", label: "장부입력확인", value: record.ledger_checked })}
          ${editableCell({ scope: "management", field: "orderer_name", label: "주문자", value: record.orderer_name })}
          ${editableCell({ scope: "management", field: "sender_phone", label: "발신자연락처", value: record.sender_phone })}
          ${editableCell({ scope: "management", field: "receiver_name", label: "수령자", value: record.receiver_name })}
          ${editableCell({ scope: "management", field: "receiver_phone", label: "수령자연락처", value: record.receiver_phone })}
          ${editableCell({ scope: "management", field: "product_name", label: "제품명", value: record.product_name, align: "left", input: "textarea" })}
          ${editableCell({ scope: "management", field: "quantity", label: "수량", value: record.quantity })}
          ${editableCell({ scope: "management", field: "receiver_address", label: "상세주소", value: record.receiver_address, align: "left", input: "textarea" })}
          ${editableCell({ scope: "management", field: "courier", label: "택배사", value: record.courier })}
          ${editableCell({ scope: "management", field: "invoice_number", label: "운송장번호", value: record.invoice_number })}
          ${editableCell({ scope: "management", field: "memo", label: "특이사항", value: record.memo, align: "left", input: "textarea" })}
          <td><button class="management-cs-button" type="button" ${csReceived ? "disabled" : ""}>${csReceived ? "접수완료" : "CS접수"}</button></td>
        `;
        applyRowPermissions(row);
        managementBody.appendChild(row);
      });
    }

    async function loadManagementRecords() {
      const query = managementSearchInput.value.trim();
      const params = new URLSearchParams({ limit: managementPageSize.value || "500" });
      const period = selectedManagementPeriod();
      if (query) params.set("q", query);
      if (period.year) params.set("year", period.year);
      if (period.month) params.set("month", period.month);
      renderManagementMonthTabs();
      try {
        const response = await fetch(`/api/management-records?${params.toString()}`);
        if (!response.ok) return;
        const data = await response.json();
        managementRecords = data.records || [];
        applyManagementFilters();
      } catch {
        managementRecords = [];
        applyManagementFilters();
        notice.textContent = "통합관리대장을 불러오지 못했습니다.";
      }
    }

    async function loadManagementPeriods() {
      if (!managementYearFilter || !managementMonthFilter) return;
      try {
        const response = await fetch("/api/management-periods");
        if (!response.ok) return;
        const data = await response.json();
        managementPeriods = data.periods || [];
        renderManagementPeriodControls();
      } catch {
        managementPeriods = [];
        renderManagementPeriodControls();
      }
    }

    async function loadManagementWorkspaceData() {
      await loadManagementPeriods();
      await loadManagementRecords();
    }

    async function uploadManagementWorkbook() {
      const file = managementImportInput.files[0];
      if (!file) return;
      const mode = managementImportMode || "daily";
      notice.textContent = `통합관리대장 ${importModeLabel(mode)} 검토 중입니다.`;
      try {
        const preview = await previewManagementImport(file, mode);
        const corrections = await requestImportCorrectionApproval({ ledgerName: "통합관리대장", preview });
        if (corrections === null) {
          notice.textContent = "통합관리대장 업로드를 취소했습니다.";
          return;
        }
        const approved = await confirmImportIfNeeded({ ledgerName: "통합관리대장", mode, preview });
        if (!approved) {
          notice.textContent = "통합관리대장 업로드를 취소했습니다.";
          return;
        }
        const formData = new FormData();
        formData.append("file", file);
        formData.append("mode", mode);
        if (corrections.length) formData.append("corrections", JSON.stringify(corrections));
        notice.textContent = `통합관리대장 ${importModeLabel(mode)} 진행 중입니다.`;
        const response = await fetch("/api/management-import", {
          method: "POST",
          body: formData,
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "통합관리대장 업로드에 실패했습니다.");
        notice.textContent = data.message || "통합관리대장 데이터를 업로드했습니다.";
        if (currentMode !== "management") {
          showWorkspace("management");
        } else {
          await loadManagementWorkspaceData();
        }
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        managementImportInput.value = "";
        managementImportMode = "daily";
      }
    }

    function collectManagementRow(row) {
      const payload = { id: row.dataset.recordId };
      row.querySelectorAll("[data-management-field]").forEach((cell) => {
        const field = cell.dataset.managementField;
        const value = fieldValue(cell);
        payload[field] = field === "order_date" || field === "ship_date"
          ? fullDateForSave(value, cell.dataset.rawDate || "")
          : value;
      });
      return payload;
    }

    async function saveManagementPayload(payload) {
      const response = await fetch("/api/management-record-update", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "통합관리대장 저장에 실패했습니다.");
      return data;
    }

    function updateManagementRecordCache(payload) {
      const record = managementRecords.find((item) => String(item.id) === String(payload.id));
      if (!record) return;
      Object.entries(payload).forEach(([key, value]) => {
        if (key !== "id") record[key] = value;
      });
    }

    async function saveManagementRow(button) {
      const row = button.closest("tr");
      if (!row) return;
      const payload = collectManagementRow(row);
      try {
        button.disabled = true;
        const data = await saveManagementPayload(payload);
        notice.textContent = data.message || "통합관리대장 행을 저장했습니다.";
        updateManagementRecordCache(payload);
        markRowDirty(row, false);
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        button.disabled = false;
      }
    }

    async function openVendorCsMailPrompt(prompt) {
      if (!prompt?.enabled) return;
      const payload = prompt.payload || {};
      openModal("cs");
      await loadVendorContacts();
      activeCsCaseId = String(prompt.case_id || payload.case_id || "");
      vendorTypeSelect.value = payload.vendor_type || "purchase";
      vendorNameInput.value = payload.vendor_name || prompt.vendor_name || "";
      recipientEmailInput.value = payload.recipient_email || prompt.recipient_email || "";
      csOriginInput.value = payload.cs_origin || "";
      csProductInput.value = payload.cs_product || "";
      csReceiverInput.value = payload.cs_receiver || "";
      csPhoneInput.value = payload.cs_phone || "";
      csAddressInput.value = payload.cs_address || "";
      csTypeInput.value = payload.cs_type || "";
      csContentInput.value = payload.cs_content || "";
      csSubjectInput.value = payload.subject || defaultCsSubject(vendorNameInput.value.trim());
      csBodyInput.value = payload.body || defaultCsBody();
      vendorContactSelect.value = `${vendorTypeSelect.value}::${vendorNameInput.value}`;
      if (!recipientEmailInput.value.trim()) {
        notice.textContent = "매입처 메일 주소가 없습니다. 주소록 업로드 또는 직접 입력 후 발송해주세요.";
      } else {
        notice.textContent = "매입처 CS 요청 메일 내용을 확인한 뒤 전송해주세요.";
      }
    }

    async function receiveManagementCs(button) {
      const row = button.closest("tr");
      if (!row) return;
      const payload = collectManagementRow(row);
      try {
        button.disabled = true;
        const saveResponse = await fetch("/api/management-record-update", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        const saveData = await saveResponse.json();
        if (!saveResponse.ok) throw new Error(saveData.error || "통합관리대장 저장에 실패했습니다.");
        const response = await fetch("/api/management-to-cs", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ id: payload.id }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "CS 처리대장 접수에 실패했습니다.");
        notice.textContent = data.message || "CS 처리대장에 접수했습니다.";
        showWorkspace("ledger");
        if (data.mail_prompt?.enabled) await openVendorCsMailPrompt(data.mail_prompt);
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        button.disabled = false;
      }
    }

    function collectLedgerRow(row) {
      return {
        id: row.dataset.caseId,
        status: fieldValue(row.querySelector('[data-field="status"]')),
        cs_type: fieldValue(row.querySelector('[data-field="cs_type"]')),
        return_invoice: fieldValue(row.querySelector('[data-field="return_invoice"]')),
        reship_invoice: fieldValue(row.querySelector('[data-field="reship_invoice"]')),
      };
    }

    async function saveLedgerPayload(payload) {
      const response = await fetch("/api/cs-case-update", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "CS 처리내용 저장에 실패했습니다.");
      return data;
    }

    function updateLedgerCaseCache(payload) {
      const savedCase = ledgerCases.find((item) => String(item.id) === String(payload.id));
      if (!savedCase) return;
      savedCase.status = payload.status;
      savedCase.cs_type = payload.cs_type;
      savedCase.return_invoice = payload.return_invoice;
      savedCase.reship_invoice = payload.reship_invoice;
    }

    async function saveLedgerRow(button) {
      const row = button.closest("tr");
      if (!row) return;
      const payload = collectLedgerRow(row);
      try {
        button.disabled = true;
        const data = await saveLedgerPayload(payload);
        notice.textContent = data.message || "CS 처리내용을 저장했습니다.";
        updateLedgerCaseCache(payload);
        updateLedgerRowCompletion(row);
        markRowDirty(row, false);
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        button.disabled = false;
      }
    }

    async function saveVisibleManagementRows({ silent = false, selectedOnly = false } = {}) {
      const rows = selectedOnly
        ? selectedRows(managementBody, "tr[data-record-id]")
        : dirtyRows(managementBody, "tr[data-record-id]");
      if (rows.length === 0) {
        if (!silent) notice.textContent = selectedOnly ? "체크된 통합관리대장 행이 없습니다." : "저장할 변경 내용이 없습니다.";
        return 0;
      }
      for (const row of rows) {
        const payload = collectManagementRow(row);
        await saveManagementPayload(payload);
        updateManagementRecordCache(payload);
        markRowDirty(row, false);
        const checkbox = row.querySelector("[data-row-check]");
        if (checkbox) checkbox.checked = false;
      }
      if (!silent) notice.textContent = `통합관리대장 ${rows.length}건 저장 완료`;
      return rows.length;
    }

    async function saveVisibleLedgerRows({ silent = false, selectedOnly = false } = {}) {
      const rows = selectedOnly
        ? selectedRows(ledgerBody, "tr[data-case-id]")
        : dirtyRows(ledgerBody, "tr[data-case-id]");
      if (rows.length === 0) {
        if (!silent) notice.textContent = selectedOnly ? "체크된 CS 처리대장 행이 없습니다." : "저장할 변경 내용이 없습니다.";
        return 0;
      }
      for (const row of rows) {
        const payload = collectLedgerRow(row);
        await saveLedgerPayload(payload);
        updateLedgerCaseCache(payload);
        updateLedgerRowCompletion(row);
        markRowDirty(row, false);
        const checkbox = row.querySelector("[data-row-check]");
        if (checkbox) checkbox.checked = false;
      }
      if (!silent) notice.textContent = `CS 처리대장 ${rows.length}건 저장 완료`;
      return rows.length;
    }

    async function saveCurrentWorkspaceRows({ silent = false, mode = currentMode, selectedOnly = false } = {}) {
      if (isBulkSaving) return 0;
      if (mode !== "management" && mode !== "ledger") return 0;
      try {
        isBulkSaving = true;
        if (mode === "management") return await saveVisibleManagementRows({ silent, selectedOnly });
        return await saveVisibleLedgerRows({ silent, selectedOnly });
      } catch (error) {
        if (!silent) notice.textContent = error.message;
        return 0;
      } finally {
        isBulkSaving = false;
      }
    }

    function selectedIds(container, rowSelector, idName) {
      return selectedRows(container, rowSelector)
        .map((row) => row.dataset[idName])
        .filter(Boolean);
    }

    async function deleteSelectedRows(mode) {
      const isManagement = mode === "management";
      const ids = isManagement
        ? selectedIds(managementBody, "tr[data-record-id]", "recordId")
        : selectedIds(ledgerBody, "tr[data-case-id]", "caseId");
      if (ids.length === 0) {
        notice.textContent = isManagement ? "삭제할 통합관리대장 행을 체크해주세요." : "삭제할 CS 처리대장 행을 체크해주세요.";
        return;
      }
      const label = isManagement ? "통합관리대장" : "CS 처리대장";
      if (!window.confirm(`${label} 선택 주문 ${ids.length}건을 삭제할까요?`)) return;
      const endpoint = isManagement ? "/api/management-records-delete" : "/api/cs-cases-delete";
      try {
        const response = await fetch(endpoint, {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ ids }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "선택 주문 삭제에 실패했습니다.");
        notice.textContent = data.message || `${label} 선택 주문 ${ids.length}건을 삭제했습니다.`;
        if (isManagement) await loadManagementWorkspaceData();
        else await loadLedgerCases();
      } catch (error) {
        notice.textContent = error.message;
      }
    }

    function collectManagementExportRows() {
      return Array.from(managementBody.querySelectorAll("tr[data-record-id]")).map((row) => collectManagementRow(row));
    }

    function collectSelectedManagementExportRows() {
      return selectedRows(managementBody, "tr[data-record-id]").map((row) => collectManagementRow(row));
    }

    function textFromCell(row, index) {
      return row.children[index]?.textContent.trim() || "";
    }

    function fullDateFromCell(row, index) {
      return row.children[index]?.dataset.fullDate || textFromCell(row, index);
    }

    function collectLedgerExportRows() {
      return Array.from(ledgerBody.querySelectorAll("tr[data-case-id]")).map((row) => ({
        occurred_at: textFromCell(row, 1),
        sales_vendor: textFromCell(row, 2),
        purchase_vendor: textFromCell(row, 3),
        status: fieldValue(row.querySelector('[data-field="status"]')),
        completed_at: textFromCell(row, 5),
        cs_type: fieldValue(row.querySelector('[data-field="cs_type"]')),
        cs_content: textFromCell(row, 7),
        reship_invoice: fieldValue(row.querySelector('[data-field="reship_invoice"]')),
        return_invoice: fieldValue(row.querySelector('[data-field="return_invoice"]')),
        order_date: fullDateFromCell(row, 10),
        ship_date: fullDateFromCell(row, 11),
        orderer_name: textFromCell(row, 12),
        orderer_phone: textFromCell(row, 13),
        receiver_name: textFromCell(row, 14),
        receiver_phone: textFromCell(row, 15),
        product_name: textFromCell(row, 16),
        quantity: textFromCell(row, 17),
        receiver_address: textFromCell(row, 18),
        courier: textFromCell(row, 19),
        original_invoice: textFromCell(row, 20),
      }));
    }

    async function downloadExcel(endpoint, rows, fallbackName, button) {
      if (!rows.length) {
        notice.textContent = "다운로드할 데이터가 없습니다.";
        return;
      }
      try {
        button.disabled = true;
        const response = await fetch(endpoint, {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ rows }),
        });
        if (!response.ok) {
          const data = await response.json();
          throw new Error(data.error || "엑셀 다운로드에 실패했습니다.");
        }
        await downloadWorkbookResponse(response, fallbackName);
        notice.textContent = "엑셀 다운로드가 시작되었습니다.";
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        button.disabled = false;
      }
    }

    async function downloadManagementExcel(scope, button) {
      const payload = { scope };
      let fallbackName = "통합관리대장.xlsx";
      if (scope === "selected") {
        payload.rows = collectSelectedManagementExportRows();
        if (!payload.rows.length) {
          notice.textContent = "다운로드할 행을 체크해주세요.";
          return;
        }
        fallbackName = "통합관리대장_선택.xlsx";
      } else if (scope === "month") {
        const period = selectedManagementPeriod();
        if (!period.year || !period.month) {
          notice.textContent = "월별 다운로드는 년도와 월을 선택해주세요.";
          return;
        }
        payload.year = period.year;
        payload.month = period.month;
        fallbackName = `통합관리대장_${payload.year}년_${Number(payload.month)}월.xlsx`;
      } else if (scope === "year") {
        const period = selectedManagementPeriod();
        if (!period.year) {
          notice.textContent = "년별 다운로드는 년도를 선택해주세요.";
          return;
        }
        payload.year = period.year;
        fallbackName = `통합관리대장_${payload.year}년.xlsx`;
      } else {
        fallbackName = "통합관리대장_전체.xlsx";
      }

      try {
        button.disabled = true;
        const response = await fetch("/api/management-export", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        if (!response.ok) {
          const data = await response.json();
          throw new Error(data.error || "통합관리대장 엑셀 다운로드에 실패했습니다.");
        }
        await downloadWorkbookResponse(response, fallbackName);
        notice.textContent = "통합관리대장 엑셀 다운로드가 시작되었습니다.";
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        button.disabled = false;
      }
    }

    async function downloadLedgerExcel(scope, button) {
      const payload = { scope };
      let fallbackName = "CS처리대장.xlsx";
      if (scope === "selected") {
        payload.rows = collectLedgerExportRows();
        if (!payload.rows.length) {
          notice.textContent = "다운로드할 행을 체크해주세요.";
          return;
        }
        fallbackName = "CS처리대장_선택.xlsx";
      } else if (scope === "month") {
        const year = ledgerYearFilter?.value || "";
        const month = ledgerMonthFilter?.value || "";
        if (!year || !month) {
          notice.textContent = "월별 다운로드는 년도와 월을 선택해주세요.";
          return;
        }
        payload.year = year;
        payload.month = month;
        fallbackName = `CS처리대장_${year}년_${Number(month)}월.xlsx`;
      } else if (scope === "year") {
        const year = ledgerYearFilter?.value || "";
        if (!year) {
          notice.textContent = "년별 다운로드는 년도를 선택해주세요.";
          return;
        }
        payload.year = year;
        fallbackName = `CS처리대장_${year}년.xlsx`;
      } else {
        fallbackName = "CS처리대장_전체.xlsx";
      }
      try {
        button.disabled = true;
        const response = await fetch("/api/cs-cases-export", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        if (!response.ok) {
          const data = await response.json();
          throw new Error(data.error || "CS 처리대장 엑셀 다운로드에 실패했습니다.");
        }
        await downloadWorkbookResponse(response, fallbackName);
        notice.textContent = "CS 처리대장 엑셀 다운로드가 시작되었습니다.";
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        button.disabled = false;
      }
    }

    async function saveCurrentCsCase() {
      refreshCsBody();
      const payload = collectCsPayload();
      if (!payload.vendor_name && !payload.cs_receiver && !payload.cs_product && !payload.cs_content) {
        notice.textContent = "업체명, 수령인, 상품명, CS내용 중 하나 이상 입력해주세요.";
        return;
      }
      try {
        saveCsCaseButton.disabled = true;
        const response = await fetch("/api/cs-case", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "CS건 저장에 실패했습니다.");
        notice.textContent = data.message || "CS건을 DB에 저장했습니다.";
        if (currentMode === "ledger") {
          closeLedgerCsPopup();
          await loadLedgerCases();
        } else {
          await loadCsCases();
        }
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        saveCsCaseButton.disabled = false;
      }
    }

    function filenameFromResponse(response, fallback) {
      const disposition = response.headers.get("Content-Disposition") || "";
      const utf8Match = disposition.match(/filename\*=UTF-8''([^;]+)/i);
      if (utf8Match) {
        try {
          return decodeURIComponent(utf8Match[1]);
        } catch {
          return utf8Match[1];
        }
      }
      const asciiMatch = disposition.match(/filename="?([^";]+)"?/i);
      return asciiMatch ? asciiMatch[1] : fallback;
    }

    async function downloadWorkbookResponse(response, fallbackName) {
      const blob = await response.blob();
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.href = url;
      link.download = filenameFromResponse(response, fallbackName);
      link.style.display = "none";
      document.body.appendChild(link);
      link.click();
      link.remove();
      window.setTimeout(() => URL.revokeObjectURL(url), 1000);
    }

    function safeNumberConfirmMessage(candidates) {
      const preview = candidates.slice(0, 5).map((candidate, index) => {
        const phones = Array.isArray(candidate.phones) ? candidate.phones.join(", ") : "";
        const items = Array.isArray(candidate.items) ? candidate.items.slice(0, 4).join(" + ") : "";
        const extra = Array.isArray(candidate.items) && candidate.items.length > 4 ? ` 외 ${candidate.items.length - 4}건` : "";
        return `${index + 1}. ${candidate.name || "수령자 미확인"} / ${phones}\n   ${candidate.address || "주소 미확인"}\n   ${items}${extra}`;
      }).join("\n\n");
      const hiddenCount = candidates.length > 5 ? `\n\n외 ${candidates.length - 5}개 후보가 더 있습니다.` : "";
      return `안심번호 합포 후보가 있습니다.\n\n동일한 수령자명과 주소인데 연락처만 다른 건입니다.\n합포장으로 출력 버튼을 누르면 합포장으로 출력하고, 개별건으로 출력 버튼을 누르면 개별건으로 출력합니다.\n\n${preview}${hiddenCount}`;
    }

    function requestSafeNumberPackageApproval(candidates) {
      if (!safeNumberPackageDialog || !safeNumberPackagePreview || !safeNumberPackageApprove || !safeNumberPackageReject) {
        return Promise.resolve(false);
      }
      safeNumberPackagePreview.textContent = safeNumberConfirmMessage(candidates);
      safeNumberPackageDialog.classList.add("open");
      safeNumberPackageDialog.setAttribute("aria-hidden", "false");
      safeNumberPackageApprove.focus();
      return new Promise((resolve) => {
        const finish = (approved) => {
          safeNumberPackageDialog.classList.remove("open");
          safeNumberPackageDialog.setAttribute("aria-hidden", "true");
          safeNumberPackageApprove.removeEventListener("click", approve);
          safeNumberPackageReject.removeEventListener("click", reject);
          safeNumberPackageDialog.removeEventListener("click", backdropReject);
          document.removeEventListener("keydown", escapeReject);
          resolve(approved);
        };
        const approve = () => finish(true);
        const reject = () => finish(false);
        const backdropReject = (event) => {
          if (event.target === safeNumberPackageDialog) finish(false);
        };
        const escapeReject = (event) => {
          if (event.key === "Escape") finish(false);
        };
        safeNumberPackageApprove.addEventListener("click", approve);
        safeNumberPackageReject.addEventListener("click", reject);
        safeNumberPackageDialog.addEventListener("click", backdropReject);
        document.addEventListener("keydown", escapeReject);
      });
    }

    function orderDownloadSizeLabel(size) {
      const bytes = Number(size || 0);
      if (!bytes) return "";
      if (bytes >= 1024 * 1024) return `${(bytes / 1024 / 1024).toFixed(1)}MB`;
      return `${Math.max(1, Math.round(bytes / 1024))}KB`;
    }

    function renderOrderDownloads(downloads) {
      if (!orderDownloadList) return;
      if (!downloads.length) {
        orderDownloadList.innerHTML = '<div class="order-download-empty">최근 출력된 파일이 없습니다.</div>';
        return;
      }
      orderDownloadList.innerHTML = downloads.map((item) => {
        const sizeLabel = orderDownloadSizeLabel(item.size);
        const meta = [item.workflow, item.created_at, sizeLabel].filter(Boolean).join(" · ");
        return `
          <div class="order-download-item">
            <div>
              <div class="order-download-name">${escapeHtml(item.filename)}</div>
              <div class="order-download-meta">${escapeHtml(meta)}</div>
            </div>
            <button class="workspace-button" type="button" data-order-download-id="${escapeHtml(item.id)}">다운로드</button>
          </div>
        `;
      }).join("");
    }

    async function loadOrderDownloads() {
      if (!orderDownloadList) return;
      try {
        const response = await fetch("/api/order-downloads");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "최근 출력 파일을 불러오지 못했습니다.");
        renderOrderDownloads(data.downloads || []);
      } catch (error) {
        orderDownloadList.innerHTML = `<div class="order-download-empty">${escapeHtml(error.message)}</div>`;
      }
    }

    async function downloadSavedOrderFile(downloadId) {
      const response = await fetch(`/api/order-download?id=${encodeURIComponent(downloadId)}`);
      if (!response.ok) {
        const data = await response.json();
        throw new Error(data.error || "파일을 다운로드하지 못했습니다.");
      }
      await downloadWorkbookResponse(response, "발주업무_출력.xlsx");
    }

    function sharedFileSizeLabel(size) {
      const bytes = Number(size || 0);
      if (!bytes) return "0KB";
      if (bytes >= 1024 * 1024) return `${(bytes / 1024 / 1024).toFixed(1)}MB`;
      return `${Math.max(1, Math.round(bytes / 1024))}KB`;
    }

    function renderSharedFiles(files) {
      if (!sharedFileBody) return;
      if (!files.length) {
        sharedFileBody.innerHTML = '<tr><td class="empty" colspan="5">저장된 업무 파일이 없습니다.</td></tr>';
        return;
      }
      sharedFileBody.innerHTML = files.map((item) => {
        const deleteButton = currentUser.role === "admin"
          ? `<button class="workspace-button danger" type="button" data-shared-file-delete="${escapeHtml(item.id)}">삭제</button>`
          : "";
        return `
          <tr>
            <td>${escapeHtml(item.original_name)}</td>
            <td>${escapeHtml(sharedFileSizeLabel(item.size))}</td>
            <td>${escapeHtml(item.uploaded_by || "-")}</td>
            <td>${escapeHtml(item.uploaded_at || "")}</td>
            <td>
              <span class="shared-file-actions">
                <button class="workspace-button" type="button" data-shared-file-download="${escapeHtml(item.id)}">다운로드</button>
                ${deleteButton}
              </span>
            </td>
          </tr>
        `;
      }).join("");
    }

    async function loadSharedFiles() {
      if (!sharedFileBody) return;
      try {
        const response = await fetch("/api/shared-files");
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "업무 파일 목록을 불러오지 못했습니다.");
        renderSharedFiles(data.files || []);
      } catch (error) {
        sharedFileBody.innerHTML = `<tr><td class="empty" colspan="5">${escapeHtml(error.message)}</td></tr>`;
      }
    }

    async function uploadSharedFile() {
      if (!sharedFileInput || !sharedFileUpload) return;
      const file = sharedFileInput.files[0];
      if (!file) {
        if (sharedFileMessage) sharedFileMessage.textContent = "올릴 파일을 선택해주세요.";
        return;
      }
      const formData = new FormData();
      formData.append("file", file);
      sharedFileUpload.disabled = true;
      if (sharedFileMessage) sharedFileMessage.textContent = "파일을 올리는 중입니다.";
      try {
        const response = await fetch("/api/shared-file-upload", { method: "POST", body: formData });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "파일 업로드에 실패했습니다.");
        if (sharedFileMessage) sharedFileMessage.textContent = data.message || "파일을 저장했습니다.";
        sharedFileInput.value = "";
        if (sharedFileDropMain) sharedFileDropMain.textContent = "업무 파일을 선택해주세요.";
        renderSharedFiles(data.files || []);
      } catch (error) {
        if (sharedFileMessage) sharedFileMessage.textContent = error.message;
      } finally {
        sharedFileUpload.disabled = false;
      }
    }

    async function downloadSharedFile(fileId) {
      const response = await fetch(`/api/shared-file-download?id=${encodeURIComponent(fileId)}`);
      if (!response.ok) {
        const data = await response.json();
        throw new Error(data.error || "업무 파일을 다운로드하지 못했습니다.");
      }
      await downloadWorkbookResponse(response, "업무파일");
    }

    async function deleteSharedFile(fileId) {
      const response = await fetch("/api/shared-file-delete", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ id: fileId }),
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "업무 파일을 삭제하지 못했습니다.");
      if (sharedFileMessage) sharedFileMessage.textContent = data.message || "파일을 삭제했습니다.";
      renderSharedFiles(data.files || []);
    }

    const mailPopupTitles = {
      cs: "CS 요청",
      stock: "입고 및 품절 공지",
    };

    function openMailMessagePopup(type) {
      const title = mailPopupTitles[type] || "유통사 업무관련 메일 발송";
      openModal(type === "stock" ? "mail-stock" : "cs");
      modalTitle.textContent = title;
    }

    function openModal(mode) {
      currentMode = mode;
      closeLedgerCsPopup();
      modal.classList.add("open");
      const modalPanel = modal.querySelector(".workhub-modal");
      modal.style.visibility = "visible";
      modalPanel.style.visibility = "visible";
      modalPanel.style.color = "#1a2230";
      modalPanel.classList.toggle("ledger-modal", mode === "ledger" || mode === "management");
      modalPanel.classList.toggle("ledger-view", mode === "ledger");
      modalPanel.classList.toggle("management-view", mode === "management");
      result.classList.remove("open");
      resultText.value = "";
      fileInput.value = "";
      templateInput.value = "";
      receiptTypeSelect.value = "일반";
      supplierInput.value = "";
      receiptDateInput.value = todayString();
      freightPaymentSelect.value = "선불";
      requestNoteInput.value = "";
      deliveryPlaceInput.value = "";
      managerInput.value = "";
      vendorContactSelect.value = "";
      vendorTypeSelect.value = "purchase";
      recipientEmailInput.value = "";
      vendorNameInput.value = "";
      csOriginInput.value = "";
      csProductInput.value = "";
      csReceiverInput.value = "";
      csPhoneInput.value = "";
      csAddressInput.value = "";
      csTypeInput.value = "";
      csContentInput.value = "";
      csSubjectInput.value = defaultCsSubject();
      refreshCsBody();
      activeCsCaseId = "";
      resetProductRows();
      notice.textContent = "";
      dropMain.textContent = "파일을 선택하거나 여기에 올려주세요.";
      templateDropMain.textContent = "롯데택배 발주서 양식을 선택해주세요.";
      messagePlaceholder.style.display = "none";
      if (stockNoticeFields) stockNoticeFields.style.display = "none";
      if (mode === "delivery") {
        modalTitle.textContent = "개별 택배건 정리";
        fileLabel.textContent = "주소일브릿지 엑셀 선택";
        submitButton.textContent = "생성";
        submitButton.className = "btn primary";
        deliveryOptions.style.display = "flex";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        dropSub.textContent = "주소일브릿지 엑셀을 업로드하면 전달용 텍스트를 만듭니다.";
      } else if (mode === "invoice") {
        modalTitle.textContent = "송장번호 추출";
        fileLabel.textContent = "출고송장 엑셀 선택";
        submitButton.textContent = "엑셀 생성";
        submitButton.className = "btn blue";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        dropSub.textContent = "출고송장 엑셀을 업로드하면 수하인별 송장번호 엑셀을 다운로드합니다.";
      } else if (mode === "lotte") {
        modalTitle.textContent = "롯데택배 발주서 변환";
        fileLabel.textContent = "주소일브릿지 원본 선택";
        submitButton.textContent = "엑셀 생성";
        submitButton.className = "btn primary";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        dropSub.textContent = "주소일브릿지 원본을 업로드하면 지정된 롯데택배 발주서 양식으로 출력합니다.";
      } else if (mode === "salesVendor") {
        modalTitle.textContent = "매입/매출별 테이터 정리(feat. 얼마에요)";
        fileLabel.textContent = "주소일브릿지 원본 선택";
        submitButton.textContent = "엑셀 생성";
        submitButton.className = "btn primary";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        dropSub.textContent = "주소일브릿지 원본을 업로드하면 매출처별 정리 엑셀을 다운로드합니다.";
      } else if (mode === "vehicle") {
        modalTitle.textContent = "차량인수증 생성";
        submitButton.textContent = "인수증 생성";
        submitButton.className = "btn primary";
        document.querySelector("label[for='fileInput']").style.display = "none";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "block";
        csFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
      } else if (mode === "cs") {
        modalTitle.textContent = "업체 CS 요청";
        submitButton.textContent = "메일 전송";
        submitButton.className = "btn primary";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "block";
        if (stockNoticeFields) stockNoticeFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        loadVendorContacts();
        loadCsCases();
      } else if (mode === "mail-stock") {
        modalTitle.textContent = "입고 및 품절 공지";
        submitButton.textContent = "공지 메일 발송";
        submitButton.className = "btn primary";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        stockNoticeFields.style.display = "block";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        loadMailSettings().then(refreshStockNoticeBody);
        refreshStockNoticeBody();
        loadVendorContacts();
      } else if (mode.startsWith("mail-")) {
        submitButton.textContent = "닫기";
        submitButton.className = "btn";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "block";
        fileInput.required = false;
        templateInput.required = false;
      } else if (mode === "ledger") {
        modalTitle.textContent = "CS 처리대장";
        submitButton.textContent = "닫기";
        submitButton.className = "btn";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        ledgerFields.style.display = "block";
        managementFields.style.display = "none";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        ledgerSearchInput.value = "";
        ledgerStatusFilter.value = "";
        if (ledgerYearFilter) ledgerYearFilter.value = "";
        if (ledgerMonthFilter) ledgerMonthFilter.value = "";
        ledgerImportInput.value = "";
        Object.keys(ledgerFilters).forEach((key) => delete ledgerFilters[key]);
        closeLedgerFilter();
        loadLedgerCases();
      } else {
        modalTitle.textContent = "통합관리대장 관리";
        submitButton.textContent = "닫기";
        submitButton.className = "btn";
        deliveryOptions.style.display = "none";
        templateUpload.style.display = "none";
        vehicleFields.style.display = "none";
        csFields.style.display = "none";
        ledgerFields.style.display = "none";
        managementFields.style.display = "block";
        messagePlaceholder.style.display = "none";
        fileInput.required = false;
        templateInput.required = false;
        managementSearchInput.value = "";
        managementYearFilter.value = "";
        managementMonthFilter.value = "";
        managementPageSize.value = "500";
        managementImportInput.value = "";
        Object.keys(managementFilters).forEach((key) => delete managementFilters[key]);
        closeLedgerFilter();
        loadManagementWorkspaceData();
      }

      const fileDrop = document.querySelector("label[for='fileInput']");
      fileDrop.style.display = mode === "vehicle" || mode === "cs" || mode === "ledger" || mode === "management" || mode.startsWith("mail-") ? "none" : "grid";
      fileLabel.style.display = mode === "vehicle" || mode === "cs" || mode === "ledger" || mode === "management" || mode.startsWith("mail-") ? "none" : "block";
    }

    function closeModal() {
      closeLedgerCsPopup();
      closeLedgerFilter();
      modal.classList.remove("open");
    }

    function setPageTitle(text) {
      if (!pageTitle) return;
      const firstNode = pageTitle.childNodes[0];
      if (firstNode && firstNode.nodeType === Node.TEXT_NODE) {
        firstNode.nodeValue = `${text} `;
      } else {
        pageTitle.prepend(document.createTextNode(`${text} `));
      }
    }

    function syncCrmNavState() {
      document.querySelectorAll("[data-crm-nav-tab]").forEach((item) => {
        if (item.classList.contains("nav-subitem")) {
          item.classList.toggle("active", item.dataset.crmNavTab === crmActiveTab);
        }
      });
    }

    function setActiveNav(mode) {
      document.querySelectorAll(".nav-item, .nav-subitem").forEach((item) => item.classList.remove("active"));
      if (mode === "crm") {
        document.querySelector("#crmNavToggle")?.classList.add("active");
        document.querySelector("#crmNavGroup")?.classList.add("open");
        syncCrmNavState();
        return;
      }
      if (mode === "dashboard") {
        document.querySelector("#companyNavToggle")?.classList.add("active");
        document.querySelector("#companyNavGroup")?.classList.add("open");
        syncCompanyNavState();
        return;
      }
      if (mode === "import") {
        document.querySelector("#importNavToggle")?.classList.add("active");
        document.querySelector("#importNavGroup")?.classList.add("open");
        return;
      }
      if (mode === "order") {
        document.querySelector("#orderNavToggle")?.classList.add("active");
        return;
      }
      if (mode === "management") {
        document.querySelector("#managementNavToggle")?.classList.add("active");
        document.querySelector("#managementNavGroup")?.classList.add("open");
        return;
      }
      if (mode === "ledger") {
        document.querySelector("#ledgerNavToggle")?.classList.add("active");
        document.querySelector("#ledgerNavGroup")?.classList.add("open");
        return;
      }
      if (mode === "salesReport") {
        document.querySelector("#salesReportNavToggle")?.classList.add("active");
        document.querySelector("#salesReportNavGroup")?.classList.add("open");
        document.querySelector('#salesReportNavGroup [data-open="salesReport"]')?.classList.add("active");
        return;
      }
      const selector = `[data-open="${mode}"]`;
      const activeItem = document.querySelector(selector);
      if (activeItem) activeItem.classList.add("active");
    }

    function showWorkspace(mode) {
      closeModal();
      if (mode === "userAdmin" && !userAdminWorkspace) mode = "dashboard";
      if (mode === "salesReport" && (!userAdminWorkspace || !can("sales_report_manage"))) mode = "dashboard";
      if (mode === "leave" && !leaveWorkspace) mode = "dashboard";
      if (mode === "backup" && !backupWorkspace) mode = "dashboard";
      if (mode === "systemUpdate" && !systemUpdateWorkspace) mode = "dashboard";
      if (mode === "fileLibrary" && !fileLibraryWorkspace) mode = "dashboard";
      if (mode === "crm" && !can("crm_view")) mode = "dashboard";
      currentMode = mode;
      const showImport = mode === "import";
      const showManagement = mode === "management";
      const showLedger = mode === "ledger";
      const showCrm = mode === "crm";
      const showLeave = mode === "leave";
      const showFileLibrary = mode === "fileLibrary" && Boolean(fileLibraryWorkspace);
      const showUserAdmin = mode === "userAdmin" && Boolean(userAdminWorkspace);
      const showSalesReport = mode === "salesReport" && Boolean(userAdminWorkspace);
      const showBackup = mode === "backup" && Boolean(backupWorkspace);
      const showSystemUpdate = mode === "systemUpdate" && Boolean(systemUpdateWorkspace);
      dashboardContent.style.display = mode === "dashboard" ? "" : "none";
      if (importWorkspace) importWorkspace.classList.toggle("active", showImport);
      if (orderWorkspace) orderWorkspace.classList.toggle("active", false);
      if (fileLibraryWorkspace) fileLibraryWorkspace.classList.toggle("active", showFileLibrary);
      managementWorkspace.classList.toggle("active", showManagement);
      ledgerWorkspace.classList.toggle("active", showLedger);
      crmWorkspace.classList.toggle("active", showCrm);
      if (leaveWorkspace) leaveWorkspace.classList.toggle("active", showLeave);
      if (userAdminWorkspace) {
        userAdminWorkspace.classList.toggle("active", showUserAdmin || showSalesReport);
        userAdminWorkspace.classList.toggle("sales-report-only", showSalesReport);
      }
      if (backupWorkspace) backupWorkspace.classList.toggle("active", showBackup);
      if (systemUpdateWorkspace) systemUpdateWorkspace.classList.toggle("active", showSystemUpdate);
      setActiveNav(mode);
      if (showManagement) {
        setPageTitle("통합관리대장 관리");
        managementSearchInput.value = "";
        managementYearFilter.value = "";
        managementMonthFilter.value = "";
        managementPageSize.value = "500";
        managementImportInput.value = "";
        closeLedgerFilter();
        loadManagementWorkspaceData();
      } else if (showLedger) {
        setPageTitle("CS 처리대장");
        ledgerSearchInput.value = "";
        ledgerStatusFilter.value = "";
        ledgerImportInput.value = "";
        Object.keys(ledgerFilters).forEach((key) => delete ledgerFilters[key]);
        closeLedgerFilter();
        loadLedgerCases();
      } else if (showCrm) {
        setPageTitle("업무관리");
        closeLedgerFilter();
        loadCrmAll().catch((error) => setCrmMessage(error.message, true));
      } else if (showImport) {
        setPageTitle("수출입 업무");
        closeLedgerFilter();
        loadImportShipments();
      } else if (showFileLibrary) {
        setPageTitle("업무 파일 자료실");
        closeLedgerFilter();
        loadSharedFiles();
      } else if (showLeave) {
        setPageTitle(leaveWorkspace.querySelector(".workspace-title")?.textContent || "연차");
        closeLedgerFilter();
        loadLeaveData();
      } else if (showUserAdmin) {
        setPageTitle("권한설정");
        closeLedgerFilter();
        resetUserAdminForm();
        loadAdminMailSettings();
        loadUserAccounts();
      } else if (showSalesReport) {
        setPageTitle("매출현황");
        closeLedgerFilter();
        loadSalesReportUploads();
        loadSalesReportDashboard();
      } else if (showBackup) {
        setPageTitle("백업 관리");
        closeLedgerFilter();
        loadBackups();
      } else if (showSystemUpdate) {
        setPageTitle("업데이트 관리");
        closeLedgerFilter();
        loadSystemUpdateStatus();
      } else {
        currentMode = "dashboard";
        setPageTitle("회사 포털");
        closeLedgerFilter();
        if (companyActiveTab === "staff") loadCompanyStaffDashboard().catch(() => {});
        else if (companyActiveTab === "notice") loadDashboardEntryData().catch(() => {});
        else if (companyActiveTab === "calendar") loadCompanyCalendar().catch(() => {});
      }
    }

    function openWorkspaceWindow(mode) {
      const url = new URL(window.location.href);
      url.searchParams.set("view", mode);
      url.searchParams.set("standalone", "1");
      window.open(url.toString(), "_blank", "width=1480,height=920");
    }

    function applySidebarSearch() {
      let rawQuery = (sidebarSearchInput?.value || "").trim();
      if (sidebarSearchAutofillValues.has(rawQuery.toLowerCase())) {
        sidebarSearchInput.value = "";
        rawQuery = "";
      }
      const query = rawQuery.toLowerCase();
      const sidebar = document.querySelector(".sidebar");
      if (!sidebar) return;
      const directItems = Array.from(sidebar.querySelectorAll(":scope > .nav-item"));
      const groups = Array.from(sidebar.querySelectorAll(":scope > .nav-group"));
      directItems.forEach((item) => { item.style.display = ""; });
      groups.forEach((group) => {
        group.style.display = "";
        group.querySelectorAll(".nav-subitem").forEach((subitem) => { subitem.style.display = ""; });
      });
      if (!query) return;
      directItems.forEach((item) => {
        const text = item.innerText.toLowerCase();
        item.style.display = text.includes(query) ? "" : "none";
      });
      groups.forEach((group) => {
        if (group.classList.contains("permission-hidden")) return;
        const toggle = group.querySelector(".nav-item");
        const groupText = (toggle?.innerText || "").toLowerCase();
        let hasMatch = groupText.includes(query);
        group.querySelectorAll(".nav-subitem").forEach((subitem) => {
          const matched = subitem.innerText.toLowerCase().includes(query) || hasMatch;
          subitem.style.display = matched ? "" : "none";
          hasMatch = hasMatch || matched;
        });
        group.style.display = hasMatch ? "" : "none";
        if (hasMatch) group.classList.add("open");
      });
    }
    function guardSidebarSearchAutofill() {
      const rawQuery = (sidebarSearchInput?.value || "").trim().toLowerCase();
      if (!sidebarSearchInput || !sidebarSearchAutofillValues.has(rawQuery)) return;
      sidebarSearchInput.value = "";
      applySidebarSearch();
    }
    function scheduleSidebarSearchAutofillGuard() {
      [0, 50, 150, 500, 1000, 2000, 4000].forEach((delay) => {
        window.setTimeout(guardSidebarSearchAutofill, delay);
      });
    }
    if (sidebarSearchInput) {
      sidebarSearchInput.value = "";
      applySidebarSearch();
      scheduleSidebarSearchAutofillGuard();
      window.setInterval(guardSidebarSearchAutofill, 500);
    }

    const ORDER_MODAL_MODES = new Set(["delivery", "invoice", "lotte", "salesVendor", "vehicle"]);
    const ORDER_MODAL_TITLES = {
      delivery: "개별 택배건 정리",
      invoice: "송장번호 추출",
      lotte: "롯데택배 발주서 변환",
      salesVendor: "매입/매출별 테이터 정리(feat. 얼마에요)",
      vehicle: "차량인수증",
    };
    const ORDER_WORKFLOWS = {
      delivery: {
        title: "개별 택배건 정리",
        description: "주소일브릿지 엑셀을 업로드해 수령자별 택배건 정리 텍스트를 생성합니다.",
        action: "개별 택배건 정리 실행",
        note: "결과는 실행창 안의 텍스트 영역에 표시됩니다.",
        steps: ["주소일브릿지 엑셀 파일을 선택합니다.", "정렬 기준을 선택합니다.", "생성 버튼을 눌러 정리 텍스트를 확인합니다."],
      },
      invoice: {
        title: "송장번호 추출",
        description: "출고송장 엑셀에서 수하인별 송장번호를 추출해 엑셀로 다운로드합니다.",
        action: "송장번호 추출 실행",
        note: "동일 송장번호도 합치지 않고 원본 건수 기준으로 유지됩니다.",
        steps: ["출고송장 엑셀 파일을 선택합니다.", "엑셀 생성 버튼을 누릅니다.", "생성된 송장번호 엑셀을 다운로드합니다."],
      },
      lotte: {
        title: "롯데택배 발주서 변환",
        description: "주소일브릿지 원본을 롯데택배 발주서 양식으로 변환합니다.",
        action: "롯데택배 발주서 변환 실행",
        note: "지정된 롯데택배 템플릿 형식으로 엑셀이 생성됩니다.",
        steps: ["주소일브릿지 원본 엑셀을 선택합니다.", "엑셀 생성 버튼을 누릅니다.", "변환된 롯데택배 발주서를 다운로드합니다."],
      },
      salesVendor: {
        title: "매입/매출별 테이터 정리(feat. 얼마에요)",
        description: "주소일브릿지 원본을 매출처별 시트와 매입처 요약 형식으로 정리합니다.",
        action: "매입/매출별 정리 실행",
        note: "파일명은 통합관리대장 양식과 같은 날짜 표기 방식으로 생성됩니다.",
        steps: ["주소일브릿지 원본 엑셀을 선택합니다.", "엑셀 생성 버튼을 누릅니다.", "매출처별 정리 파일을 다운로드합니다."],
      },
      vehicle: {
        title: "차량인수증",
        description: "공급받는자, 제품, 납품장소, 담당자 정보를 입력해 차량인수증을 생성합니다.",
        action: "차량인수증 생성 실행",
        note: "비고란은 검정색으로 출력되며, 지정 전화 안내 문구는 기존 서식대로 유지됩니다.",
        steps: ["공급받는자와 제품 정보를 입력합니다.", "납품장소와 담당자명을 입력합니다.", "인수증 생성 버튼으로 엑셀을 다운로드합니다."],
      },
    };
    function showOrderWorkspace(mode) {
      if (ORDER_WORKFLOWS[mode]) currentOrderMode = mode;
      closeLedgerFilter();
      dashboardContent.style.display = "none";
      if (orderWorkspace) orderWorkspace.classList.toggle("active", true);
      if (importWorkspace) importWorkspace.classList.toggle("active", false);
      if (fileLibraryWorkspace) fileLibraryWorkspace.classList.toggle("active", false);
      managementWorkspace.classList.toggle("active", false);
      ledgerWorkspace.classList.toggle("active", false);
      crmWorkspace.classList.toggle("active", false);
      if (leaveWorkspace) leaveWorkspace.classList.toggle("active", false);
      if (userAdminWorkspace) userAdminWorkspace.classList.toggle("active", false);
      if (backupWorkspace) backupWorkspace.classList.toggle("active", false);
      if (systemUpdateWorkspace) systemUpdateWorkspace.classList.toggle("active", false);
      setActiveNav("order");
      setPageTitle("발주업무");
      if (orderWorkspaceTitle) orderWorkspaceTitle.textContent = "발주업무";
      if (orderWorkspacePanelTitle) orderWorkspacePanelTitle.textContent = "작업을 선택해주세요.";
      if (orderWorkspaceDescription) orderWorkspaceDescription.textContent = "아래 5가지 작업 중 필요한 항목의 실행 버튼을 누르면 기존 드롭/업로드 실행창이 열립니다.";
      loadOrderDownloads();
    }
    function openOrderModal(mode) {
      currentOrderMode = ORDER_WORKFLOWS[mode] ? mode : "delivery";
      showOrderWorkspace(currentOrderMode);
      setPageTitle(ORDER_MODAL_TITLES[currentOrderMode] || "발주업무");
      openModal(currentOrderMode);
    }
    sidebar.addEventListener("click", (event) => {
      const button = event.target.closest("[data-open]");
      if (!button || !sidebar.contains(button)) return;
      const mode = button.dataset.open;
      if (!ORDER_MODAL_MODES.has(mode)) return;
      event.preventDefault();
      event.stopImmediatePropagation();
      scheduleSidebarSearchAutofillGuard();
      openOrderModal(mode);
    }, true);

    document.querySelectorAll("[data-open]").forEach((button) => {
      button.addEventListener("click", () => {
        scheduleSidebarSearchAutofillGuard();
        const mode = button.dataset.open;
        if (mode === "crm") {
          const crmGroup = document.querySelector("#crmNavGroup");
          const wasCrmGroupOpen = Boolean(crmGroup?.classList.contains("open"));
          showWorkspace("crm");
          if (button.dataset.crmNavTab) setCrmTab(button.dataset.crmNavTab);
          if (button.id === "crmNavToggle") {
            crmGroup?.classList.toggle("open", !wasCrmGroupOpen);
          } else {
            crmGroup?.classList.add("open");
          }
          return;
        }
        if (mode === "order") {
          showOrderWorkspace();
          return;
        }
        if (mode === "management" || mode === "ledger" || mode === "crm" || mode === "import" || mode === "fileLibrary" || mode === "userAdmin" || mode === "salesReport" || mode === "leave" || mode === "backup" || mode === "systemUpdate") {
          showWorkspace(mode);
          if (mode === "salesReport" && button.closest("#salesReportNavGroup")) {
            openSalesReportUploadPicker();
          }
          return;
        }
        openModal(mode);
      });
    });
    document.querySelectorAll("[data-view]").forEach((button) => {
      button.addEventListener("click", () => {
        const companyGroup = document.querySelector("#companyNavGroup");
        showWorkspace(button.dataset.view);
        if (button.dataset.companyTab) setCompanyTab(button.dataset.companyTab);
        if (button.id === "companyNavToggle") {
          companyGroup?.classList.add("open");
        } else if (button.dataset.companyTab) {
          companyGroup?.classList.add("open");
        }
      });
    });
    document.querySelectorAll("[data-open-window]").forEach((button) => {
      button.addEventListener("click", () => openWorkspaceWindow(button.dataset.openWindow));
    });
    document.addEventListener("click", (event) => {
      const button = event.target.closest("[data-order-execute]");
      if (!button) return;
      event.preventDefault();
      openOrderModal(button.dataset.orderExecute);
    });
    orderDownloadRefresh?.addEventListener("click", loadOrderDownloads);
    orderDownloadList?.addEventListener("click", async (event) => {
      const button = event.target.closest("[data-order-download-id]");
      if (!button) return;
      event.preventDefault();
      button.disabled = true;
      try {
        await downloadSavedOrderFile(button.dataset.orderDownloadId);
      } catch (error) {
        alert(error.message);
      } finally {
        button.disabled = false;
      }
    });
    sharedFileRefresh?.addEventListener("click", loadSharedFiles);
    sharedFileUpload?.addEventListener("click", uploadSharedFile);
    sharedFileBody?.addEventListener("click", async (event) => {
      const downloadButton = event.target.closest("[data-shared-file-download]");
      const deleteButton = event.target.closest("[data-shared-file-delete]");
      const button = downloadButton || deleteButton;
      if (!button) return;
      event.preventDefault();
      button.disabled = true;
      try {
        if (downloadButton) {
          await downloadSharedFile(downloadButton.dataset.sharedFileDownload);
        } else if (deleteButton) {
          if (!confirm("선택한 업무 파일을 삭제할까요?")) return;
          await deleteSharedFile(deleteButton.dataset.sharedFileDelete);
        }
      } catch (error) {
        alert(error.message);
      } finally {
        button.disabled = false;
      }
    });
    crmTabs.forEach((button) => {
      button.addEventListener("click", () => setCrmTab(button.dataset.crmTab));
      button.addEventListener("keydown", (event) => {
        const keys = ["ArrowLeft", "ArrowRight", "Home", "End"];
        if (!keys.includes(event.key)) return;
        event.preventDefault();
        const availableTabs = crmTabs.filter((tab) => !tab.classList.contains("permission-hidden"));
        const currentIndex = availableTabs.indexOf(button);
        let nextIndex = currentIndex;
        if (event.key === "ArrowRight") nextIndex = (currentIndex + 1) % availableTabs.length;
        if (event.key === "ArrowLeft") nextIndex = (currentIndex - 1 + availableTabs.length) % availableTabs.length;
        if (event.key === "Home") nextIndex = 0;
        if (event.key === "End") nextIndex = availableTabs.length - 1;
        const nextTab = availableTabs[nextIndex];
        if (nextTab) {
          setCrmTab(nextTab.dataset.crmTab);
          nextTab?.focus();
        }
      });
    });
    companyTabs.forEach((button) => {
      button.addEventListener("click", () => setCompanyTab(button.dataset.companyTab));
    });
    companyCalendarPrev?.addEventListener("click", () => shiftCalendarMonth(-1).catch(() => {
      if (companyCalendarGrid) companyCalendarGrid.innerHTML = `<div class="calendar-empty">이전 달 캘린더를 불러오지 못했습니다.</div>`;
    }));
    companyCalendarNext?.addEventListener("click", () => shiftCalendarMonth(1).catch(() => {
      if (companyCalendarGrid) companyCalendarGrid.innerHTML = `<div class="calendar-empty">다음 달 캘린더를 불러오지 못했습니다.</div>`;
    }));
    companyCalendarToday?.addEventListener("click", () => {
      companyCalendarMonth = todayString().slice(0, 7);
      companyCalendarSelectedDay = todayString();
      loadCompanyCalendar().catch(() => {
        if (companyCalendarGrid) companyCalendarGrid.innerHTML = `<div class="calendar-empty">오늘 캘린더를 불러오지 못했습니다.</div>`;
      });
    });
    companyCalendarRefresh?.addEventListener("click", () => loadCompanyCalendar().catch(() => {
      if (companyCalendarGrid) companyCalendarGrid.innerHTML = `<div class="calendar-empty">캘린더를 새로고침하지 못했습니다.</div>`;
    }));
    companyCalendarGrid?.addEventListener("click", (event) => {
      const eventButton = event.target.closest("[data-calendar-event-id]");
      if (eventButton) {
        event.stopPropagation();
        openCalendarEventWidget(eventButton.dataset.calendarEventId);
        return;
      }
      const dayCell = event.target.closest("[data-calendar-day]");
      if (!dayCell) return;
      companyCalendarSelectedDay = dayCell.dataset.calendarDay;
      renderCompanyCalendar({ month: companyCalendarMonth, events: companyCalendarEvents, summary: companyCalendarSummary });
    });
    companyCalendarGrid?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      const eventButton = event.target.closest("[data-calendar-event-id]");
      if (eventButton) {
        event.preventDefault();
        openCalendarEventWidget(eventButton.dataset.calendarEventId);
        return;
      }
      const dayCell = event.target.closest("[data-calendar-day]");
      if (!dayCell) return;
      event.preventDefault();
      companyCalendarSelectedDay = dayCell.dataset.calendarDay;
      renderCompanyCalendar({ month: companyCalendarMonth, events: companyCalendarEvents, summary: companyCalendarSummary });
    });
    companyCalendarSelectedList?.addEventListener("click", (event) => {
      const row = event.target.closest("[data-calendar-event-id]");
      if (!row) return;
      openCalendarEventWidget(row.dataset.calendarEventId);
    });
    companyCalendarSelectedList?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      const row = event.target.closest("[data-calendar-event-id]");
      if (!row) return;
      event.preventDefault();
      openCalendarEventWidget(row.dataset.calendarEventId);
    });
    if (companyStaffRefresh) {
      companyStaffRefresh.addEventListener("click", () => loadCompanyStaffDashboard().catch((error) => {
        if (companyOrgBody) companyOrgBody.innerHTML = `<div class="company-org-empty">${escapeHtml(error.message)}</div>`;
      }));
    }
    sidebarNoticePreview?.addEventListener("click", () => openNoticeWidget());
    sidebarNoticePreview?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      event.preventDefault();
      openNoticeWidget();
    });
    noticePreview?.addEventListener("click", () => openNoticeWidget());
    companyOrgBody?.addEventListener("click", (event) => {
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest("[data-company-person-card]");
      if (!card) return;
      openEmployeeWidget(card.dataset.companyPersonCard).catch((error) => setCrmMessage(error.message, true));
    });
    companyOrgBody?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest("[data-company-person-card]");
      if (!card) return;
      event.preventDefault();
      openEmployeeWidget(card.dataset.companyPersonCard).catch((error) => setCrmMessage(error.message, true));
    });
    companyStaffTaskBody?.addEventListener("click", (event) => {
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest("[data-company-task-id]");
      if (!card) return;
      openCrmTaskWidget(card.dataset.companyTaskId).catch((error) => setCrmMessage(error.message, true));
    });
    companyStaffTaskBody?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest("[data-company-task-id]");
      if (!card) return;
      event.preventDefault();
      openCrmTaskWidget(card.dataset.companyTaskId).catch((error) => setCrmMessage(error.message, true));
    });
    dashboardContent?.addEventListener("click", (event) => {
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest('.company-panel[data-company-panel="rules"] .company-card');
      if (!card) return;
      openCompanyCardWidget(card);
    });
    dashboardContent?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest('.company-panel[data-company-panel="rules"] .company-card');
      if (!card) return;
      event.preventDefault();
      openCompanyCardWidget(card);
    });
    if (internalChatForm) {
      internalChatForm.addEventListener("submit", (event) => {
        sendInternalMessage(event).catch(() => {
          if (internalChatList) internalChatList.innerHTML = `<div class="internal-chat-empty">메시지를 저장하지 못했습니다.</div>`;
        });
      });
    }
    if (internalChatRoomList) {
      internalChatRoomList.addEventListener("click", (event) => {
        const button = event.target.closest("[data-chat-room]");
        if (!button) return;
        setInternalChatRoom(button.dataset.chatRoom, button.dataset.chatUserId || "").catch(() => {
          if (internalChatList) internalChatList.innerHTML = `<div class="internal-chat-empty">메시지를 불러오지 못했습니다.</div>`;
        });
      });
    }
    if (internalChatRefresh) {
      internalChatRefresh.addEventListener("click", () => loadInternalChatUsers().then(() => loadInternalMessages()).catch(() => {
        if (internalChatList) internalChatList.innerHTML = `<div class="internal-chat-empty">메시지를 불러오지 못했습니다.</div>`;
      }));
    }
    if (sidebarSearchInput) {
      sidebarSearchInput.addEventListener("keydown", () => {
        sidebarSearchUserTyped = true;
      });
      sidebarSearchInput.addEventListener("paste", () => {
        sidebarSearchUserTyped = true;
      });
      sidebarSearchInput.addEventListener("search", () => {
        if (!sidebarSearchInput.value) sidebarSearchUserTyped = false;
      });
      sidebarSearchInput.addEventListener("input", () => {
        const rawQuery = (sidebarSearchInput.value || "").trim().toLowerCase();
        if (document.activeElement === sidebarSearchInput && !sidebarSearchAutofillValues.has(rawQuery)) {
          sidebarSearchUserTyped = true;
        }
        applySidebarSearch();
      });
    }
    document.querySelectorAll("[data-crm-go]").forEach((button) => {
      button.addEventListener("click", () => setCrmTab(button.dataset.crmGo));
    });
    crmProjectProgressBody?.addEventListener("click", (event) => {
      if (isInteractiveTarget(event.target)) return;
      const row = event.target.closest("[data-crm-project-key]");
      if (!row) return;
      openCrmProjectWidget(row.dataset.crmProjectKey);
    });
    crmProjectProgressBody?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      if (isInteractiveTarget(event.target)) return;
      const row = event.target.closest("[data-crm-project-key]");
      if (!row) return;
      event.preventDefault();
      openCrmProjectWidget(row.dataset.crmProjectKey);
    });
    crmRefresh.addEventListener("click", () => loadCrmAll().catch((error) => setCrmMessage(error.message, true)));
    crmAccountQuick.addEventListener("click", () => {
      setCrmTab("accounts");
    });
    crmStaffRefresh?.addEventListener("click", () => loadCrmStaffDashboard().catch((error) => setCrmMessage(error.message, true)));
    crmStaffBody?.addEventListener("click", (event) => {
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest("[data-crm-staff-card]");
      if (!card) return;
      openEmployeeWidget(card.dataset.crmStaffCard).catch((error) => setCrmMessage(error.message, true));
    });
    crmStaffBody?.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      if (isInteractiveTarget(event.target)) return;
      const card = event.target.closest("[data-crm-staff-card]");
      if (!card) return;
      event.preventDefault();
      openEmployeeWidget(card.dataset.crmStaffCard).catch((error) => setCrmMessage(error.message, true));
    });
    crmTaskQuick.addEventListener("click", () => {
      resetCrmTaskForm();
      setCrmTaskFormOpen(true);
      setCrmTab("tasks");
      crmTaskTitle?.focus();
    });
    crmAccountForm.addEventListener("submit", (event) => {
      saveCrmAccountForm(event).catch((error) => setCrmMessage(error.message, true));
    });
    crmAccountReset.addEventListener("click", resetCrmAccountForm);
    crmAccountSearchButton.addEventListener("click", () => loadCrmAccounts().catch((error) => setCrmMessage(error.message, true)));
    crmAccountSearch.addEventListener("keydown", (event) => {
      if (event.key === "Enter") {
        event.preventDefault();
        loadCrmAccounts().catch((error) => setCrmMessage(error.message, true));
      }
    });
    crmAccountBody.addEventListener("click", (event) => {
      const editButton = event.target.closest("[data-crm-account-edit]");
      if (!editButton) return;
      const account = crmAccounts.find((item) => String(item.id) === String(editButton.dataset.crmAccountEdit));
      if (account) fillCrmAccountForm(account);
    });
    crmTaskAccount.addEventListener("change", () => {
      const account = crmAccounts.find((item) => String(item.id) === String(crmTaskAccount.value));
      if (account) crmTaskAccountName.value = account.name;
    });
    crmTaskForm.addEventListener("submit", (event) => {
      saveCrmTaskForm(event).catch((error) => setCrmMessage(error.message, true));
    });
    crmTaskReset.addEventListener("click", resetCrmTaskForm);
    crmTaskFormToggle?.addEventListener("click", () => {
      setCrmTaskFormOpen(crmTaskForm.classList.contains("collapsed"));
    });
    crmTaskPresetList?.addEventListener("click", (event) => {
      const button = event.target.closest("[data-crm-task-view]");
      if (!button) return;
      applyCrmTaskView(button.dataset.crmTaskView);
    });
    crmTaskViewSelect?.addEventListener("change", () => {
      if (crmTaskViewSelect.value) applyCrmTaskView(crmTaskViewSelect.value);
    });
    crmTaskViewSave?.addEventListener("click", () => {
      saveCurrentCrmTaskView().catch((error) => setCrmMessage(error.message, true));
    });
    crmTaskViewDelete?.addEventListener("click", () => {
      deleteCurrentCrmTaskView().catch((error) => setCrmMessage(error.message, true));
    });
    crmTaskSearchButton.addEventListener("click", () => {
      markCrmTaskFiltersDirty();
      loadCrmTasks().catch((error) => setCrmMessage(error.message, true));
    });
    crmTaskAdvancedToggle?.addEventListener("click", () => {
      const open = Boolean(crmAdvancedFilters?.hidden);
      if (crmAdvancedFilters) crmAdvancedFilters.hidden = !open;
      crmTaskAdvancedToggle.setAttribute("aria-expanded", open ? "true" : "false");
      crmTaskAdvancedToggle.textContent = open ? "필터 닫기" : "고급 필터";
    });
    crmTaskSearch.addEventListener("keydown", (event) => {
      if (event.key === "Enter") {
        event.preventDefault();
        markCrmTaskFiltersDirty();
        loadCrmTasks().catch((error) => setCrmMessage(error.message, true));
      }
    });
    crmTaskSearch?.addEventListener("input", markCrmTaskFiltersDirty);
    [
      crmTaskStatusFilter,
      crmTaskAssigneeFilter,
      crmTaskPriorityFilter,
      crmTaskDueFilter,
      crmTaskSourceFilter,
      crmTaskSort,
      crmTaskOpenOnly,
    ].forEach((control) => {
      control?.addEventListener("change", () => {
        if (crmTaskStatusFilter?.value === "완료" && crmTaskOpenOnly) crmTaskOpenOnly.checked = false;
        markCrmTaskFiltersDirty();
        loadCrmTasks().catch((error) => setCrmMessage(error.message, true));
      });
    });
    crmTaskFilterReset?.addEventListener("click", () => {
      crmActiveTaskViewId = "builtin:open";
      writeCrmTaskFilters(crmBuiltinTaskView("open")?.filters || {});
      renderCrmSavedViews();
      loadCrmTasks().catch((error) => setCrmMessage(error.message, true));
    });
    function findCrmTaskById(taskId) {
      return crmTasks.find((item) => String(item.id) === String(taskId))
        || crmMineTasks.find((item) => String(item.id) === String(taskId))
        || companyStaffTaskCache.find((item) => String(item.id) === String(taskId))
        || crmStaffTaskCache.find((item) => String(item.id) === String(taskId))
        || crmProjectProgress.flatMap((project) => project.tasks || []).find((item) => String(item.id) === String(taskId));
    }
    function handleCrmTaskAction(event) {
      const editButton = event.target.closest("[data-crm-task-edit]");
      const statusButton = event.target.closest("[data-crm-task-status]");
      const commentButton = event.target.closest("[data-crm-task-comment]");
      if (editButton) {
        const task = findCrmTaskById(editButton.dataset.crmTaskEdit);
        if (task) fillCrmTaskForm(task);
        return true;
      }
      if (statusButton) {
        crmSelectedTaskId = String(statusButton.dataset.crmTaskStatus || crmSelectedTaskId);
        updateCrmTaskStatus(statusButton.dataset.crmTaskStatus, statusButton.dataset.status).catch((error) => setCrmMessage(error.message, true));
        return true;
      }
      if (commentButton) {
        openCrmTaskComment(commentButton.dataset.crmTaskComment);
        return true;
      }
      return false;
    }
    crmTaskBody.addEventListener("click", (event) => {
      if (handleCrmTaskAction(event)) return;
      const card = event.target.closest("[data-crm-task-card]");
      if (card) {
        selectCrmTask(card.dataset.crmTaskCard);
        openCrmTaskWidget(card.dataset.crmTaskCard).catch((error) => setCrmMessage(error.message, true));
      }
    });
    crmTaskBody.addEventListener("keydown", (event) => {
      if (!isCardActivationKey(event)) return;
      if (event.target.closest("button")) return;
      const card = event.target.closest("[data-crm-task-card]");
      if (!card) return;
      event.preventDefault();
      selectCrmTask(card.dataset.crmTaskCard);
      openCrmTaskWidget(card.dataset.crmTaskCard).catch((error) => setCrmMessage(error.message, true));
    });
    focusWidgetClose?.addEventListener("click", closeFocusWidget);
    focusWidget?.addEventListener("click", (event) => {
      if (event.target === focusWidget) closeFocusWidget();
    });
    focusWidgetBody?.addEventListener("click", (event) => {
      const openTaskButton = event.target.closest("[data-focus-open-task]");
      if (openTaskButton) {
        openCrmTaskWidget(openTaskButton.dataset.focusOpenTask).catch((error) => setCrmMessage(error.message, true));
        return;
      }
      const filterProjectButton = event.target.closest("[data-focus-filter-project]");
      if (filterProjectButton) {
        applyCrmProjectFilter(filterProjectButton.dataset.focusFilterProject);
        return;
      }
      if (handleCrmTaskAction(event)) closeFocusWidget();
    });
    focusWidgetBody?.addEventListener("submit", (event) => {
      const form = event.target.closest("[data-focus-widget-comment-form]");
      if (!form) return;
      event.preventDefault();
      const textarea = form.querySelector("[data-focus-widget-comment-body]");
      addCrmTaskComment(form.dataset.focusWidgetCommentForm, textarea?.value || "")
        .then(() => openCrmTaskWidget(form.dataset.focusWidgetCommentForm))
        .catch((error) => setCrmMessage(error.message, true));
    });
    crmMineTaskBody?.addEventListener("click", (event) => {
      handleCrmTaskAction(event);
    });
    crmTaskDetail?.addEventListener("click", (event) => {
      handleCrmTaskAction(event);
    });
    crmTaskDetail?.addEventListener("submit", (event) => {
      const form = event.target.closest("[data-crm-comment-form]");
      if (!form) return;
      event.preventDefault();
      const textarea = form.querySelector("[data-crm-comment-body]");
      addCrmTaskComment(form.dataset.crmCommentForm, textarea?.value || "")
        .then(() => {
          if (textarea) textarea.value = "";
        })
        .catch((error) => setCrmMessage(error.message, true));
    });
    crmMessengerForm.addEventListener("submit", (event) => {
      saveCrmMessengerForm(event).catch((error) => setCrmMessage(error.message, true));
    });
    crmWebhookUrlCopy?.addEventListener("click", () => copyCrmText(crmWebhookUrl?.textContent, "수신 URL"));
    crmWebhookHeaderCopy?.addEventListener("click", () => copyCrmText(crmWebhookHeader?.textContent, "헤더 이름"));
    crmWebhookTokenCopy?.addEventListener("click", () => copyCrmText(crmWebhookToken?.textContent, "웹훅 토큰"));
    crmWebhookTokenRotate?.addEventListener("click", () => {
      rotateCrmWebhookToken().catch((error) => setCrmMessage(error.message, true));
    });
    crmMessageEventBody?.addEventListener("click", (event) => {
      const copyButton = event.target.closest("[data-crm-copy-text]");
      if (copyButton) {
        copyCrmText(copyButton.dataset.crmCopyText, "발신자 키");
        return;
      }
      const mapButton = event.target.closest("[data-crm-map-sender]");
      if (!mapButton) return;
      crmMessengerPlatform.value = mapButton.dataset.platform || "kakao";
      crmMessengerSenderKey.value = mapButton.dataset.crmMapSender || "";
      crmMessengerDisplayName.value = mapButton.dataset.senderName || "";
      crmMessengerUser?.focus();
      setCrmMessage("Workhub 사용자를 선택한 뒤 저장하면 이 발신자가 등록 직원으로 매핑됩니다.");
    });
    document.querySelector("#orderNavToggle").addEventListener("click", () => {
      document.querySelector("#orderNavGroup").classList.toggle("open");
    });
    document.querySelector("#managementNavToggle")?.addEventListener("click", () => {
      document.querySelector("#managementNavGroup")?.classList.add("open");
    });
    document.querySelector("#ledgerNavToggle")?.addEventListener("click", () => {
      document.querySelector("#ledgerNavGroup")?.classList.add("open");
    });
    document.querySelector("#distributionMailNavToggle")?.addEventListener("click", () => {
      document.querySelector("#distributionMailNavGroup")?.classList.toggle("open");
    });
    document.querySelectorAll("[data-mail-popup]").forEach((button) => {
      button.addEventListener("click", () => openMailMessagePopup(button.dataset.mailPopup));
    });
    csAttachmentInput?.addEventListener("change", updateCsAttachmentSummary);
    document.querySelector("#adminNavToggle")?.addEventListener("click", () => {
      document.querySelector("#adminNavGroup")?.classList.toggle("open");
    });
    document.querySelector("#salesReportNavToggle")?.addEventListener("click", () => {
      document.querySelector("#salesReportNavGroup")?.classList.add("open");
      showWorkspace("salesReport");
    });
    document.querySelector("#noticeInputOpen").addEventListener("click", openNoticePopup);
    importShipmentInputOpen.addEventListener("click", () => {
      showWorkspace("import");
      openImportShipmentPopup();
    });
    function openManagementImport(mode = "daily") {
      managementImportMode = mode;
      managementImportInput.value = "";
      showWorkspace("management");
      managementImportInput.click();
    }

    function openLedgerImport(mode = "daily") {
      ledgerImportMode = mode;
      ledgerImportInput.value = "";
      showWorkspace("ledger");
      ledgerImportInput.click();
    }

    document.querySelectorAll("[data-management-import-mode]").forEach((button) => {
      button.addEventListener("click", () => openManagementImport(button.dataset.managementImportMode || "daily"));
    });
    document.querySelectorAll("[data-ledger-import-mode]").forEach((button) => {
      button.addEventListener("click", () => openLedgerImport(button.dataset.ledgerImportMode || "daily"));
    });
    noticePopupClose.addEventListener("click", closeNoticePopup);
    noticePopup.addEventListener("click", (event) => {
      if (event.target === noticePopup) closeNoticePopup();
    });
    document.addEventListener("keydown", (event) => {
      if (event.key === "Escape" && focusWidget?.classList.contains("open")) {
        closeFocusWidget();
      }
    });
    importShipmentRefresh.addEventListener("click", loadImportShipments);
    dashboardImportScheduleRefresh?.addEventListener("click", loadImportShipments);
    dashboardImportScheduleOpen?.addEventListener("click", () => {
      showWorkspace("import");
      openImportShipmentPopup();
    });
    importShipmentWorkspaceOpen?.addEventListener("click", () => openImportShipmentPopup());
    importShipmentTreeToggle.addEventListener("click", () => {
      importProgressCard.classList.toggle("open");
    });
    importShipmentClose.addEventListener("click", closeImportShipmentPopup);
    importShipmentReset.addEventListener("click", () => resetImportShipmentForm());
    importShipmentSave.addEventListener("click", () => {
      saveImportShipment().catch((error) => {
        notice.textContent = error.message;
      });
    });
    importShipmentPopup.addEventListener("click", (event) => {
      if (event.target === importShipmentPopup) closeImportShipmentPopup();
    });
    importShipmentBody.addEventListener("click", (event) => {
      const editButton = event.target.closest("[data-import-edit]");
      const completeButton = event.target.closest("[data-import-complete]");
      if (editButton) {
        const record = importShipments.find((item) => String(item.id) === String(editButton.dataset.importEdit));
        if (record) openImportShipmentPopup(record);
      }
      if (completeButton) {
        completeImportShipment(completeButton.dataset.importComplete).catch((error) => {
          notice.textContent = error.message;
        });
      }
    });
    document.querySelector("#closeModal").addEventListener("click", closeModal);
    document.querySelector("#cancel").addEventListener("click", closeModal);
    modal.addEventListener("click", (event) => {
      if (event.target === modal) closeModal();
    });
    fileInput.addEventListener("change", () => {
      dropMain.textContent = fileInput.files[0] ? fileInput.files[0].name : "파일을 선택하거나 여기에 올려주세요.";
    });
    templateInput.addEventListener("change", () => {
      templateDropMain.textContent = templateInput.files[0] ? templateInput.files[0].name : "롯데택배 발주서 양식을 선택해주세요.";
    });

    function setupDropzone(dropzone, input, label, fallbackText) {
      if (!dropzone || !input || !label) return;
      ["dragenter", "dragover"].forEach((eventName) => {
        dropzone.addEventListener(eventName, (event) => {
          event.preventDefault();
          event.stopPropagation();
          dropzone.classList.add("dragover");
        });
      });
      ["dragleave", "drop"].forEach((eventName) => {
        dropzone.addEventListener(eventName, (event) => {
          event.preventDefault();
          event.stopPropagation();
          dropzone.classList.remove("dragover");
        });
      });
      dropzone.addEventListener("drop", (event) => {
        const files = event.dataTransfer.files;
        if (!files || files.length === 0) return;
        input.files = files;
        label.textContent = files[0] ? files[0].name : fallbackText;
        input.dispatchEvent(new Event("change", { bubbles: true }));
      });
    }

    setupDropzone(
      document.querySelector("label[for='fileInput']"),
      fileInput,
      dropMain,
      "파일을 선택하거나 여기에 올려주세요."
    );
    setupDropzone(
      document.querySelector("label[for='templateInput']"),
      templateInput,
      templateDropMain,
      "롯데택배 발주서 양식을 선택해주세요."
    );
    setupDropzone(
      document.querySelector("label[for='vendorContactsFileInput']"),
      vendorContactsFileInput,
      vendorContactsDropMain,
      "업체구분/업체명/메일주소 엑셀을 선택해주세요."
    );
    setupDropzone(
      document.querySelector("label[for='sharedFileInput']"),
      sharedFileInput,
      sharedFileDropMain,
      "업무 파일을 선택해주세요."
    );
    document.querySelector("#addProductRow").addEventListener("click", () => addProductRow());
    noticeSaveButton.addEventListener("click", saveNoticeTemplate);
    noticeClearButton.addEventListener("click", clearNoticeTemplate);
    [noticeDateInput, noticeTitleInput, noticeOwnerInput, noticeBodyInput]
      .forEach((input) => input.addEventListener("input", renderNoticePreview));
    receiptTypeSelect.addEventListener("change", resetProductRows);
    vendorContactSelect.addEventListener("change", applySelectedVendor);
    saveVendorContactButton.addEventListener("click", saveCurrentVendorContact);
    if (vendorContactsFileInput) vendorContactsFileInput.addEventListener("change", uploadVendorContactsWorkbook);
    if (salesReportFileInput) salesReportFileInput.addEventListener("change", uploadSalesReportWorkbook);
    saveCsCaseButton.addEventListener("click", saveCurrentCsCase);
    ledgerRefresh.addEventListener("click", loadLedgerCases);
    ledgerStatusFilter.addEventListener("change", applyLedgerFilters);
    ledgerFilterButtons.forEach((button) => {
      button.addEventListener("click", (event) => {
        event.stopPropagation();
        openLedgerFilter(button);
      });
    });
    managementFilterButtons.forEach((button) => {
      button.addEventListener("click", (event) => {
        event.stopPropagation();
        openManagementFilter(button);
      });
    });
    ledgerFilterSearch.addEventListener("input", () => {
      refreshActiveFilterOptions();
    });
    ledgerFilterSearch.addEventListener("keydown", (event) => {
      if (event.key === "Enter") setActivePopoverFilter(ledgerFilterSearch.value);
      if (event.key === "Escape") closeLedgerFilter();
    });
    ledgerFilterApply.addEventListener("click", () => setActivePopoverFilter(ledgerFilterSearch.value));
    ledgerFilterClear.addEventListener("click", () => setActivePopoverFilter(""));
    ledgerFilterOptions.addEventListener("click", (event) => {
      const option = event.target.closest("[data-filter-value]");
      if (option) setActivePopoverFilter(option.dataset.filterValue || "");
    });
    document.addEventListener("click", (event) => {
      if (
        !ledgerFilterPopover.contains(event.target)
        && !event.target.closest("[data-ledger-filter-button]")
        && !event.target.closest("[data-management-filter-button]")
      ) {
        closeLedgerFilter();
      }
    });
    ledgerAddCs.addEventListener("click", openLedgerCsPopup);
    ledgerCsPopupClose.addEventListener("click", closeLedgerCsPopup);
    ledgerImportInput.addEventListener("change", uploadLedgerWorkbook);
    managementRefresh.addEventListener("click", loadManagementRecords);
    managementImportInput.addEventListener("change", uploadManagementWorkbook);
    managementSaveAll.addEventListener("click", () => saveCurrentWorkspaceRows({ mode: "management", selectedOnly: true }));
    ledgerSaveAll.addEventListener("click", () => saveCurrentWorkspaceRows({ mode: "ledger", selectedOnly: true }));
    managementDeleteSelected.addEventListener("click", () => deleteSelectedRows("management"));
    ledgerDeleteSelected.addEventListener("click", () => deleteSelectedRows("ledger"));
    if (userAdminRefresh) userAdminRefresh.addEventListener("click", loadUserAccounts);
    if (userAdminSave) userAdminSave.addEventListener("click", saveUserAccount);
    if (adminMailSettingsSave) adminMailSettingsSave.addEventListener("click", saveAdminMailSettings);
    if (adminMailTechnicalSave) adminMailTechnicalSave.addEventListener("click", saveAdminMailTechnicalSettings);
    if (adminMailTestSend) adminMailTestSend.addEventListener("click", sendAdminMailTestMessage);
    if (adminSmtpPort) {
      adminSmtpPort.addEventListener("change", () => {
        if (adminSmtpSecurity) adminSmtpSecurity.value = adminSmtpPort.value === "587" ? "tls" : "ssl";
      });
    }
    if (userAdminRole) userAdminRole.addEventListener("change", syncPermissionChecksForRole);
    if (backupRefresh) backupRefresh.addEventListener("click", loadBackups);
    if (backupCreate) backupCreate.addEventListener("click", createBackupNow);
    if (backupSettingsSave) backupSettingsSave.addEventListener("click", saveBackupSettings);
    if (backupCreateSelected) backupCreateSelected.addEventListener("click", createBackupAtSelectedPath);
    if (backupRestoreInput) backupRestoreInput.addEventListener("change", restoreBackupFromUpload);
    if (systemUpdateRefresh) systemUpdateRefresh.addEventListener("click", loadSystemUpdateStatus);
    if (systemUpdateCheck) systemUpdateCheck.addEventListener("click", checkSystemUpdate);
    if (systemUpdateApply) systemUpdateApply.addEventListener("click", applySystemUpdate);
    if (backupBody) {
      backupBody.addEventListener("click", (event) => {
        const restoreButton = event.target.closest("[data-backup-restore]");
        const downloadButton = event.target.closest("[data-backup-download]");
        const deleteButton = event.target.closest("[data-backup-delete]");
        if (restoreButton) restoreBackupByName(restoreButton.dataset.backupRestore);
        if (downloadButton) downloadBackup(downloadButton.dataset.backupDownload);
        if (deleteButton) deleteBackup(deleteButton.dataset.backupDelete);
      });
    }
    if (userAdminBody) {
      userAdminBody.addEventListener("click", (event) => {
        const editButton = event.target.closest("[data-user-edit]");
        if (editButton) editUserAccount(editButton.dataset.userEdit);
      });
    }
    leaveTabs.forEach((button) => {
      button.addEventListener("click", () => setLeaveTab(button.dataset.leaveTab));
    });
    if (leaveRefresh) leaveRefresh.addEventListener("click", loadLeaveData);
    if (leaveRequestSubmit) leaveRequestSubmit.addEventListener("click", submitLeaveRequest);
    if (leaveUnitSelect) leaveUnitSelect.addEventListener("change", syncHalfDayDates);
    if (leaveStartDate) leaveStartDate.addEventListener("change", syncHalfDayDates);
    if (leaveBalanceSave) leaveBalanceSave.addEventListener("click", saveLeaveBalance);
    if (leaveAccrualApply) leaveAccrualApply.addEventListener("click", applyLeaveAccrual);
    if (leaveUsageSave) leaveUsageSave.addEventListener("click", saveHistoricalLeaveUsage);
    if (leaveHolidaySave) leaveHolidaySave.addEventListener("click", saveCompanyHolidayFromLeave);
    if (leaveApprovalBody) {
      leaveApprovalBody.addEventListener("click", (event) => {
        const button = event.target.closest("[data-leave-decision]");
        if (button) decideLeaveRequest(button.dataset.leaveId, button.dataset.leaveDecision);
      });
    }
    if (leaveHistoryBody) {
      leaveHistoryBody.addEventListener("click", (event) => {
        const button = event.target.closest("[data-leave-cancel]");
        if (button) cancelLeaveRequest(button.dataset.leaveCancel);
      });
    }
    managementPageSize.addEventListener("change", loadManagementRecords);
    ledgerPageSize.addEventListener("change", loadLedgerCases);
    ledgerYearFilter?.addEventListener("change", loadLedgerCases);
    managementYearFilter.addEventListener("change", () => {
      renderManagementPeriodControls();
    });
    ledgerMonthFilter?.addEventListener("change", () => {
      ensureYearForMonth(ledgerYearFilter, ledgerMonthFilter);
      loadLedgerCases();
    });
    managementMonthFilter.addEventListener("change", () => {
      const period = selectedManagementPeriod();
      managementYearFilter.value = period.year || "";
      renderManagementMonthTabs();
    });
    if (managementMonthTabs) {
      managementMonthTabs.addEventListener("click", (event) => {
        const button = event.target.closest("[data-management-month]");
        if (!button) return;
        setManagementPeriod(button.dataset.managementYear || "", button.dataset.managementMonth || "");
        renderManagementPeriodControls();
        loadManagementRecords();
      });
    }
    if (managementSelectAll) {
      managementSelectAll.addEventListener("change", () => {
        managementBody.querySelectorAll("tr[data-record-id] [data-row-check]").forEach((checkbox) => {
          checkbox.checked = managementSelectAll.checked;
        });
      });
    }
    function closeDownloadMenus() {
      ledgerDownloadMenu?.classList.remove("open");
      managementDownloadMenu?.classList.remove("open");
    }
    ledgerDownloadMenuButton.addEventListener("click", (event) => {
      event.stopPropagation();
      managementDownloadMenu.classList.remove("open");
      ledgerDownloadMenu.classList.toggle("open");
    });
    managementDownloadMenuButton.addEventListener("click", (event) => {
      event.stopPropagation();
      ledgerDownloadMenu.classList.remove("open");
      managementDownloadMenu.classList.toggle("open");
    });
    ledgerDownloadMenu.addEventListener("click", (event) => {
      const option = event.target.closest("[data-ledger-download]");
      if (!option) return;
      closeDownloadMenus();
      downloadLedgerExcel(option.dataset.ledgerDownload, ledgerDownloadMenuButton);
    });
    managementDownloadMenu.addEventListener("click", (event) => {
      const option = event.target.closest("[data-management-download]");
      if (!option) return;
      closeDownloadMenus();
      downloadManagementExcel(option.dataset.managementDownload, managementDownloadMenuButton);
    });
    document.addEventListener("click", (event) => {
      if (!event.target.closest(".download-menu-wrap")) closeDownloadMenus();
    });
    managementBody.addEventListener("click", (event) => {
      const editableCell = event.target.closest(".editable-cell[data-management-field]");
      if (editableCell) {
        openCellEditor("management", editableCell);
        return;
      }
      if (event.target.closest("[data-row-check]") && managementSelectAll) {
        const checks = Array.from(managementBody.querySelectorAll("tr[data-record-id] [data-row-check]"));
        managementSelectAll.checked = checks.length > 0 && checks.every((checkbox) => checkbox.checked);
      }
      const csButton = event.target.closest(".management-cs-button");
      if (csButton) receiveManagementCs(csButton);
    });
    ledgerBody.addEventListener("click", (event) => {
      const editableCell = event.target.closest(".editable-cell[data-field]");
      if (editableCell) {
        openCellEditor("ledger", editableCell);
      }
    });
    [ledgerCellApply, managementCellApply].forEach((button) => {
      button?.addEventListener("click", () => applyCellEditor(button === ledgerCellApply ? "ledger" : "management"));
    });
    [ledgerCellCancel, managementCellCancel].forEach((button) => {
      button?.addEventListener("click", () => closeCellEditor(button === ledgerCellCancel ? "ledger" : "management"));
    });
    [ledgerCellEditMount, managementCellEditMount].forEach((mount) => {
      mount?.addEventListener("keydown", (event) => {
        if (event.key !== "Enter") return;
        if (event.target.tagName === "TEXTAREA" && !event.ctrlKey) return;
        event.preventDefault();
        applyCellEditor(mount === ledgerCellEditMount ? "ledger" : "management");
      });
    });
    ledgerSearchInput.addEventListener("keydown", (event) => {
      if (event.key === "Enter") {
        event.preventDefault();
        loadLedgerCases();
      }
    });
    managementSearchInput.addEventListener("keydown", (event) => {
      if (event.key === "Enter") {
        event.preventDefault();
        loadManagementRecords();
      }
    });
    vendorNameInput.addEventListener("input", () => {
      csSubjectInput.value = currentMode === "mail-stock" ? "입고 및 품절 공지" : defaultCsSubject(vendorNameInput.value.trim());
    });
    [csOriginInput, csProductInput, csReceiverInput, csPhoneInput, csAddressInput, csContentInput]
      .forEach((input) => input.addEventListener("input", refreshCsBody));
    stockVendorPickerButton?.addEventListener("click", toggleStockVendorTree);
    stockVendorTree?.addEventListener("click", (event) => {
      const button = event.target.closest("[data-stock-vendor-name]");
      if (button) applySelectedStockVendorFromTree(button);
    });
    [
      stockNoticeDateInput,
      stockInboundProductInput,
      stockInboundScheduleInput,
      stockOutboundAvailableInput,
      stockInboundNoteInput,
      stockSoldoutProductInput,
      stockOutboundBlockedInput,
      stockRestockScheduleInput,
      stockSoldoutNoteInput,
    ].forEach((input) => input?.addEventListener("input", refreshStockNoticeBody));
    stockNoticeDateInput?.addEventListener("change", refreshStockNoticeBody);

    setInterval(() => {
      saveCurrentWorkspaceRows({ silent: true });
    }, 10 * 60 * 1000);

    uploadForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      const formData = new FormData(uploadForm);
      notice.textContent = "처리 중입니다.";
      submitButton.disabled = true;
      try {
        if (currentMode === "ledger" || currentMode === "management") {
          closeModal();
        } else if (currentMode === "cs") {
          refreshCsBody();
          const payload = collectCsPayload();
          if (!payload.recipient_email || !payload.subject || !payload.body) {
            throw new Error("받는 업체 메일, 제목, 요청 내용을 입력해주세요.");
          }
          const csMailFormData = appendCsMailPayload(new FormData(), payload);
          const response = await fetch("/api/cs-mail", {
            method: "POST",
            body: csMailFormData,
          });
          const data = await response.json();
          if (!response.ok) throw new Error(data.error || "메일 전송에 실패했습니다.");
          notice.textContent = data.message || "메일 전송이 완료되었습니다.";
          activeCsCaseId = "";
          if (csAttachmentInput) csAttachmentInput.value = "";
          updateCsAttachmentSummary();
          await loadCsCases();
        } else if (currentMode === "mail-stock") {
          refreshStockNoticeBody();
          const payload = collectStockNoticePayload();
          if (!payload.recipient_email || !payload.subject || !payload.body) {
            throw new Error("받는 업체 메일, 제목, 공지 내용을 입력해주세요.");
          }
          const response = await fetch("/api/mail-send", {
            method: "POST",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify(payload),
          });
          const data = await response.json();
          if (!response.ok) throw new Error(data.error || "공지 메일 발송에 실패했습니다.");
          notice.textContent = data.message || "공지 메일 발송이 완료되었습니다.";
        } else if (currentMode === "vehicle") {
          const payload = collectVehiclePayload();
          if (!payload.supplier || !payload.delivery_place || !payload.manager || payload.items.length === 0) {
            throw new Error("공급받는자, 제품, 납품장소, 담당자명을 입력해주세요.");
          }
          const response = await fetch("/api/vehicle-receipt", {
            method: "POST",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify(payload),
          });
          if (!response.ok) {
            const data = await response.json();
            throw new Error(data.error || "처리에 실패했습니다.");
          }
          await downloadWorkbookResponse(response, "차량인수증.xlsx");
          loadOrderDownloads();
          notice.textContent = "차량인수증 다운로드가 시작되었습니다.";
        } else if (currentMode === "delivery") {
          if (!fileInput.files[0]) throw new Error("주소일브릿지 엑셀 파일을 선택해주세요.");
          const response = await fetch("/api/delivery-summary", { method: "POST", body: formData });
          const data = await response.json();
          if (!response.ok) throw new Error(data.error || "처리에 실패했습니다.");
          const safeNumberCandidates = Array.isArray(data.safe_number_candidates) ? data.safe_number_candidates : [];
          let outputText = data.text;
          let safeNumberNotice = "";
          if (safeNumberCandidates.length && data.approved_text && data.approved_text !== data.text) {
            const approvedSafePackages = await requestSafeNumberPackageApproval(safeNumberCandidates);
            if (approvedSafePackages) {
              outputText = data.approved_text;
              safeNumberNotice = ` 안심번호 합포 후보 ${safeNumberCandidates.length}건을 합포장으로 반영했습니다.`;
            } else {
              safeNumberNotice = ` 안심번호 합포 후보 ${safeNumberCandidates.length}건은 개별건으로 유지했습니다.`;
            }
          }
          resultText.value = outputText;
          result.classList.add("open");
          notice.textContent = `${data.line_count}개 묶음이 생성되었습니다.${safeNumberNotice}`;
        } else if (currentMode === "salesVendor") {
          if (!fileInput.files[0]) throw new Error("주소일브릿지 원본 엑셀 파일을 선택해주세요.");
          const response = await fetch("/api/sales-vendor-summary", { method: "POST", body: formData });
          if (!response.ok) {
            const data = await response.json();
            throw new Error(data.error || "처리에 실패했습니다.");
          }
          await downloadWorkbookResponse(response, "주소일브릿지_매출처별_정리.xlsx");
          loadOrderDownloads();
          notice.textContent = "매출처별 정리 엑셀 다운로드가 시작되었습니다.";
        } else {
          if (!fileInput.files[0]) {
            throw new Error(currentMode === "invoice" ? "출고송장 엑셀 파일을 선택해주세요." : "주소일브릿지 원본 엑셀 파일을 선택해주세요.");
          }
          const endpoint = currentMode === "invoice" ? "/api/invoice-export" : "/api/lotte-order-form";
          const response = await fetch(endpoint, { method: "POST", body: formData });
          if (!response.ok) {
            const data = await response.json();
            throw new Error(data.error || "처리에 실패했습니다.");
          }
          await downloadWorkbookResponse(
            response,
            currentMode === "invoice" ? "송장번호_추출.xlsx" : "롯데택배_발주서.xlsx"
          );
          loadOrderDownloads();
          notice.textContent = "엑셀 파일 다운로드가 시작되었습니다.";
        }
      } catch (error) {
        notice.textContent = error.message;
      } finally {
        submitButton.disabled = false;
      }
    });

    document.querySelector("#copyResult").addEventListener("click", async () => {
      await navigator.clipboard.writeText(resultText.value);
      notice.textContent = "텍스트를 복사했습니다.";
    });

    document.querySelector("#downloadText").addEventListener("click", () => {
      const blob = new Blob([resultText.value], { type: "text/plain;charset=utf-8" });
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.href = url;
      link.download = "개별_택배건_요약.txt";
      document.body.appendChild(link);
      link.click();
      link.remove();
      URL.revokeObjectURL(url);
    });

    const initialView = new URLSearchParams(window.location.search).get("view");
    showWorkspace(["management", "ledger", "crm", "import", "fileLibrary", "leave", "userAdmin", "salesReport", "backup", "systemUpdate"].includes(initialView) ? initialView : "dashboard");
  </script>
</body>
</html>
"""

LOGIN_HTML = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>(주)소일브릿지 업무자동화 로그인</title>
  <link rel="stylesheet" href="/static/workhub.css" />
  <style>
    :root {
      --bg: #f5f7fb;
      --panel: #ffffff;
      --line: #d8dee9;
      --text: #111827;
      --muted: #667085;
      --blue: #2563eb;
      --green: #079455;
      font-family: Pretendard, Inter, "Noto Sans KR", "Malgun Gothic", Arial, sans-serif;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      display: grid;
      place-items: center;
      background:
        radial-gradient(circle at 20% 20%, rgba(37, 99, 235, .13), transparent 28%),
        radial-gradient(circle at 80% 10%, rgba(7, 148, 85, .12), transparent 25%),
        var(--bg);
      color: var(--text);
    }
    .login-shell {
      width: min(760px, calc(100vw - 32px));
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 14px;
      box-shadow: 0 22px 60px rgba(15, 23, 42, .16);
      padding: 34px 32px 30px;
    }
    .brand {
      display: flex;
      align-items: center;
      gap: 12px;
      margin-bottom: 26px;
    }
    .brand-mark {
      width: 44px;
      height: 44px;
      border-radius: 12px;
      display: grid;
      place-items: center;
      background: linear-gradient(145deg, #2563eb, #079455);
      color: white;
      font-weight: 900;
    }
    .brand-title { font-size: 20px; font-weight: 900; line-height: 1.32; }
    .brand-sub { margin-top: 4px; color: var(--muted); font-size: 13px; font-weight: 700; }
    h1 { margin: 0 0 8px; font-size: 26px; }
    .lead { margin: 0 0 24px; color: var(--muted); font-size: 14px; line-height: 1.55; }
    label { display: block; margin: 14px 0 7px; font-size: 13px; font-weight: 850; color: #344054; }
    input {
      width: 100%;
      height: 44px;
      border: 1px solid #cfd6e2;
      border-radius: 9px;
      padding: 0 13px;
      font-size: 15px;
      font-weight: 700;
      outline: none;
    }
    input:focus {
      border-color: var(--blue);
      box-shadow: 0 0 0 3px rgba(37, 99, 235, .13);
    }
    .error {
      margin: 0 0 14px;
      padding: 11px 12px;
      border-radius: 8px;
      background: #fee2e2;
      color: #b42318;
      font-size: 13px;
      font-weight: 850;
    }
    .success {
      margin: 0 0 14px;
      padding: 11px 12px;
      border-radius: 8px;
      background: #dcfce7;
      color: #047857;
      font-size: 13px;
      font-weight: 850;
    }
    .login-grid {
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
      gap: 24px;
      align-items: start;
    }
    .login-panel {
      min-width: 0;
    }
    .register-panel {
      min-width: 0;
      padding: 18px;
      border: 1px solid #e5e7eb;
      border-radius: 10px;
      background: #fbfcff;
    }
    .register-panel h2 {
      margin: 0 0 8px;
      font-size: 18px;
      line-height: 1.25;
    }
    button {
      width: 100%;
      height: 46px;
      margin-top: 22px;
      border: 0;
      border-radius: 9px;
      background: linear-gradient(135deg, #2563eb, #079455);
      color: white;
      font-size: 15px;
      font-weight: 900;
      cursor: pointer;
    }
    .hint {
      margin-top: 18px;
      padding: 12px;
      border-radius: 8px;
      background: #f8fafc;
      border: 1px solid #e5e7eb;
      color: #475467;
      font-size: 12px;
      line-height: 1.55;
    }
    .hint strong { color: #111827; }
    @media (max-width: 720px) {
      .login-grid { grid-template-columns: 1fr; }
      .login-shell { padding: 26px 22px 24px; }
    }
  </style>
</head>
<body>
  <main class="login-shell">
    <div class="brand">
      <div class="brand-mark">SB</div>
      <div>
        <div class="brand-title">(주)소일브릿지<br>업무자동화</div>
        <div class="brand-sub">관리자 / 부관리자 / 사용자 로그인</div>
      </div>
    </div>
    <div class="success" style="display: __LOGIN_MESSAGE_DISPLAY__;">__LOGIN_MESSAGE__</div>
    <div class="login-grid">
      <section class="login-panel">
        <h1>로그인</h1>
        <p class="lead">업무 화면을 사용하려면 승인된 계정으로 로그인해주세요.</p>
        <div class="error" style="display: __LOGIN_ERROR_DISPLAY__;">__LOGIN_ERROR__</div>
        <form method="post" action="/login">
          <label for="username">아이디</label>
          <input id="username" name="username" type="text" autocomplete="username" autofocus />
          <label for="password">비밀번호</label>
          <input id="password" name="password" type="password" autocomplete="current-password" />
          <button type="submit">로그인</button>
        </form>
        <div class="hint">
          로그인 실패가 반복되면 잠시 제한됩니다.<br>
          계정은 관리자 승인 후 사용할 수 있습니다.
        </div>
      </section>
      <section class="register-panel">
        <h2>계정 등록 요청</h2>
        <p class="lead">신규 직원은 요청을 남기고, 관리자가 승인하면 로그인할 수 있습니다.</p>
        <div class="error" style="display: __REGISTER_ERROR_DISPLAY__;">__REGISTER_ERROR__</div>
        <form method="post" action="/register">
          <label for="registerUsername">아이디</label>
          <input id="registerUsername" name="username" type="text" autocomplete="username" placeholder="영문/숫자 3~32자" />
          <label for="registerDisplayName">표시 이름</label>
          <input id="registerDisplayName" name="display_name" type="text" autocomplete="name" placeholder="예) 홍길동" />
          <label for="registerPassword">비밀번호</label>
          <input id="registerPassword" name="password" type="password" autocomplete="new-password" placeholder="10자 이상" />
          <label for="registerPasswordConfirm">비밀번호 확인</label>
          <input id="registerPasswordConfirm" name="password_confirm" type="password" autocomplete="new-password" />
          <button type="submit">등록 요청</button>
        </form>
        <div class="hint">
          비밀번호는 10자 이상, 기본/반복/아이디 포함 패턴은 사용할 수 없습니다.
        </div>
      </section>
    </div>
  </main>
</body>
</html>
"""

ADMIN_TOOLS_NAV_HTML = r"""
      <div class="nav-group" id="adminNavGroup">
        <button class="nav-item" id="adminNavToggle" type="button">
          <span class="nav-label"><i data-lucide="settings"></i> <span>관리자</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem" type="button" data-management-import-mode="replace">통합관리대장 전체 데이터 교체 업로드</button>
          <button class="nav-subitem" type="button" data-ledger-import-mode="replace">CS처리대장 전체 데이터 교체 업로드</button>
          <button class="nav-subitem" type="button" data-open="systemUpdate">업데이트 관리</button>
          <button class="nav-subitem" type="button" data-open="backup">백업 관리</button>
          <button class="nav-subitem" type="button" data-open="userAdmin">권한설정</button>
        </div>
      </div>
"""

SALES_REPORT_NAV_HTML = r"""
      <div class="nav-group" id="salesReportNavGroup">
        <button class="nav-item" id="salesReportNavToggle" type="button">
          <span class="nav-label"><i data-lucide="bar-chart-3"></i> <span>매출현황 및 관리</span></span>
          <i class="nav-chevron" data-lucide="chevron-right"></i>
        </button>
        <div class="nav-submenu">
          <button class="nav-subitem" type="button" data-open="salesReport">매출표 업로드</button>
        </div>
      </div>
"""

LEAVE_NAV_HTML = r"""
      <button class="nav-item" type="button" data-open="leave"><i data-lucide="calendar-days"></i> <span>__LEAVE_TITLE__</span></button>
"""

LEAVE_WORKSPACE_HTML = r"""
      <section class="workspace-view" id="leaveWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">__LEAVE_TITLE__</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="leaveRefresh">새로고침</button>
          </div>
        </div>
        <div class="workspace-mount">
          <div class="leave-panel">
            <div class="leave-summary-grid">
              <article class="leave-summary-card">
                <span>시작 연차</span>
                <strong id="leaveTotalDays">0일</strong>
              </article>
              <article class="leave-summary-card">
                <span>사용 연차</span>
                <strong id="leaveUsedDays">0일</strong>
              </article>
              <article class="leave-summary-card">
                <span>승인대기 예약</span>
                <strong id="leaveReservedDays">0일</strong>
              </article>
              <article class="leave-summary-card accent">
                <span>남은 연차</span>
                <strong id="leaveRemainingDays">0일</strong>
              </article>
            </div>
            <div class="leave-tabs">
              <button class="leave-tab active" type="button" data-leave-tab="mine">내 연차</button>
              <button class="leave-tab" type="button" data-leave-tab="request">연차 신청</button>
              <button class="leave-tab" type="button" data-leave-tab="approvals">승인 대기</button>
              <button class="leave-tab" type="button" data-leave-tab="admin">직원별 관리</button>
            </div>
            <div class="leave-message" id="leaveMessage"></div>
            <div class="leave-notification-list" id="leaveNotificationList"></div>
            <section class="leave-tab-panel active" id="leaveTabMine">
              <div class="leave-grid two">
                <div class="leave-card">
                  <div class="leave-card-title">연차 잔여 현황</div>
                  <table class="leave-table">
                    <thead><tr><th>유형</th><th>시작</th><th>사용</th><th>잔여</th></tr></thead>
                    <tbody id="leaveBalanceBody"></tbody>
                  </table>
                </div>
                <div class="leave-card">
                  <div class="leave-card-title">연차 사용/신청 이력</div>
                  <table class="leave-table">
                    <thead><tr><th>기간</th><th>구분</th><th>수량</th><th>상태</th><th>사유</th></tr></thead>
                    <tbody id="leaveHistoryBody"></tbody>
                  </table>
                </div>
              </div>
            </section>
            <section class="leave-tab-panel" id="leaveTabRequest">
              <div class="leave-card narrow">
                <div class="leave-card-title">연차 신청</div>
                <div class="leave-form">
                  <label>휴가 유형<select id="leaveTypeSelect"></select></label>
                  <label>단위<select id="leaveUnitSelect"><option value="FULL_DAY">연차</option><option value="HALF_DAY">반차</option></select></label>
                  <label>시작일<input id="leaveStartDate" type="date" /></label>
                  <label>종료일<input id="leaveEndDate" type="date" /></label>
                  <label class="wide">사유<textarea id="leaveReasonInput" placeholder="연차 사유를 입력해주세요."></textarea></label>
                  <button class="workspace-button" type="button" id="leaveRequestSubmit">신청하기</button>
                </div>
              </div>
            </section>
            <section class="leave-tab-panel" id="leaveTabApprovals">
              <div class="leave-card">
                <div class="leave-card-title">승인 대기</div>
                <table class="leave-table">
                  <thead><tr><th>신청자</th><th>기간</th><th>구분</th><th>수량</th><th>사유</th><th>처리</th></tr></thead>
                  <tbody id="leaveApprovalBody"></tbody>
                </table>
              </div>
            </section>
            <section class="leave-tab-panel" id="leaveTabAdmin">
              <div class="leave-grid two">
                <div class="leave-card">
                  <div class="leave-card-title">직원 연차 기준 설정</div>
                  <div class="leave-form">
                    <label>직원<select id="leaveAdminUserSelect"></select></label>
                    <label>시작 연차<input id="leaveAdminTotalInput" type="number" min="0" step="0.5" value="10" /></label>
                    <label>기존 사용 연차<input id="leaveAdminUsedInput" type="number" min="0" step="0.5" value="0" /></label>
                    <button class="workspace-button" type="button" id="leaveBalanceSave">연차 기준 저장</button>
                    <button class="workspace-button" type="button" id="leaveAccrualApply">올해 연차 자동 발생</button>
                  </div>
                </div>
                <div class="leave-card">
                  <div class="leave-card-title">기존 사용 연차 날짜별 입력</div>
                  <div class="leave-form">
                    <label>직원<select id="leaveUsageUserSelect"></select></label>
                    <label class="wide">사용 일자 목록<textarea id="leaveUsageDatesInput" placeholder="2026-01-03&#10;2026-02-14 반차&#10;2026-03-05 0.5"></textarea></label>
                    <label class="wide">메모<input id="leaveUsageNoteInput" type="text" placeholder="예: 시스템 도입 전 사용분" /></label>
                    <button class="workspace-button" type="button" id="leaveUsageSave">사용 일자 등록</button>
                  </div>
                </div>
              </div>
                <div class="leave-card">
                  <div class="leave-card-title">&#44277;&#55092;&#51068;/&#45824;&#52404;&#44277;&#55092;&#51068; &#44288;&#47532;</div>
                  <div class="leave-form">
                    <label>&#45216;&#51676;<input id="leaveHolidayDateInput" type="date" /></label>
                    <label>&#51060;&#47492;<input id="leaveHolidayNameInput" type="text" placeholder="&#50696;: &#45824;&#52404;&#44277;&#55092;&#51068;" /></label>
                    <label class="admin-check"><input id="leaveHolidaySubstituteInput" type="checkbox" /> &#45824;&#52404;&#44277;&#55092;&#51068;</label>
                    <button class="workspace-button" type="button" id="leaveHolidaySave">&#55092;&#51068; &#51200;&#51109;</button>
                  </div>
                </div>
              <div class="leave-card">
                <div class="leave-card-title">직원별 연차 현황</div>
                <table class="leave-table">
                  <thead><tr><th>직원</th><th>아이디</th><th>시작</th><th>사용</th><th>잔여</th></tr></thead>
                  <tbody id="leaveAdminBalanceBody"></tbody>
                </table>
              </div>
            </section>
          </div>
        </div>
      </section>
"""

ADMIN_WORKSPACE_HTML = r"""
      <section class="workspace-view" id="userAdminWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">권한설정</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="userAdminRefresh">새로고침</button>
          </div>
        </div>
        <div class="workspace-mount">
          <div class="admin-panel">
            <div class="admin-card">
              <div class="admin-section-title">네이버 메일 연동</div>
              <div class="admin-form">
                <label>네이버 메일 아이디
                  <input id="adminNaverEmailInput" type="text" placeholder="예) soilbridge@naver.com" />
                </label>
                <label>네이버 메일 비밀번호
                  <input id="adminNaverPasswordInput" type="password" placeholder="저장된 비밀번호가 없으면 입력" autocomplete="new-password" />
                </label>
                <label class="admin-check"><input id="adminSaveMailCredentials" type="checkbox" checked /> 아이디/비밀번호 저장</label>
                <button class="workspace-button" type="button" id="adminMailSettingsSave">메일 기본정보 저장</button>
              </div>
              <div class="admin-message" id="adminMailSettingsMessage"></div>
            </div>
            <div class="admin-card">
              <div class="admin-section-title">단체메일 기술 설정</div>
              <div class="admin-form">
                <label>SMTP 포트
                  <select id="adminSmtpPort">
                    <option value="465">465 - SSL</option>
                    <option value="587">587 - TLS</option>
                  </select>
                </label>
                <label>보안 방식
                  <select id="adminSmtpSecurity">
                    <option value="ssl">SSL</option>
                    <option value="tls">TLS</option>
                  </select>
                </label>
                <label>묶음 발송 수
                  <input id="adminBulkBatchSize" type="number" min="1" max="100" value="20" />
                </label>
                <label>발송 간격(초)
                  <input id="adminBulkSendInterval" type="number" min="5" max="600" value="15" />
                </label>
                <label>묶음 사이 대기(분)
                  <input id="adminBulkBatchPause" type="number" min="0" max="120" value="5" />
                </label>
                <label>테스트 수신 메일
                  <input id="adminBulkTestRecipient" type="email" placeholder="예) test@example.com" />
                </label>
                <button class="workspace-button" type="button" id="adminMailTechnicalSave">단체메일 기술 설정 저장</button>
                <button class="workspace-button ghost" type="button" id="adminMailTestSend">연동 테스트 메일 발송</button>
              </div>
              <div class="admin-message" id="adminMailTechnicalMessage">저장된 네이버 메일 계정으로 1건만 발송합니다.</div>
            </div>
            <div class="admin-card">
              <div class="admin-section-title">업체 메일 주소록</div>
              <label class="dropzone" for="vendorContactsFileInput">
                <span class="drop-main" id="vendorContactsDropMain">업체구분/업체명/메일주소 엑셀을 선택해주세요.</span>
                <span class="drop-sub">헤더 예시: 거래처구분, 업체명, 메일주소</span>
                <input id="vendorContactsFileInput" name="vendor_contacts" type="file" accept=".xlsx,.xlsm" />
              </label>
              <div class="admin-message">매입처/매출처 업체 메일 주소록 엑셀을 업로드하면 DB에 저장됩니다.</div>
            </div>
            <div class="admin-card" id="salesReportUploadCard">
              <div class="admin-section-title">매출현황</div>
              <input id="salesReportFileInput" name="sales_report" type="file" accept=".xlsx,.xlsm,.xls,.csv" hidden />
              <div class="sales-dashboard" id="salesReportDashboard">
                <div class="sales-kpi-grid" id="salesReportKpiGrid"></div>
                <div class="sales-dashboard-grid">
                  <div class="sales-panel">
                    <div class="sales-panel-head">일자별 매출 흐름 <span>매출 통계.xlsx</span></div>
                    <table class="sales-table">
                      <thead><tr><th>일자</th><th>수량</th><th>손익매출</th><th>판매합계</th><th>손익마진</th></tr></thead>
                      <tbody id="salesReportDailyBody"></tbody>
                    </table>
                  </div>
                  <div class="sales-panel">
                    <div class="sales-panel-head">매출처별 TOP <span>매출처별.xlsx</span></div>
                    <table class="sales-table">
                      <thead><tr><th>판매사</th><th>수량</th><th>손익매출</th><th>마진</th></tr></thead>
                      <tbody id="salesReportSellerBody"></tbody>
                    </table>
                  </div>
                  <div class="sales-panel">
                    <div class="sales-panel-head">상품별 TOP <span>Statistics_Good</span></div>
                    <table class="sales-table">
                      <thead><tr><th>상품명</th><th>수량</th><th>손익매출</th><th>마진</th></tr></thead>
                      <tbody id="salesReportProductBody"></tbody>
                    </table>
                  </div>
                  <div class="sales-panel">
                    <div class="sales-panel-head">업체별 총 매입금액 <span>공급사별</span></div>
                    <table class="sales-table">
                      <thead><tr><th>업체</th><th>총 매입금액</th><th>수량</th></tr></thead>
                      <tbody id="salesReportReviewBody"></tbody>
                    </table>
                  </div>
                </div>
              </div>
            </div>
            <div class="admin-form">
              <input id="userAdminId" type="hidden" />
              <label>아이디
                <input id="userAdminUsername" type="text" placeholder="예) hong" />
              </label>
              <label>표시 이름
                <input id="userAdminDisplayName" type="text" placeholder="예) 홍길동" />
              </label>
              <label>권한
                <select id="userAdminRole">
                  <option value="admin">관리자</option>
                  <option value="sub_admin">부관리자</option>
                  <option value="user">사용자</option>
                </select>
              </label>
              <label>비밀번호
                <input id="userAdminPassword" type="password" placeholder="신규/변경 시 입력" />
              </label>
              <label class="admin-check"><input id="userAdminActive" type="checkbox" checked /> 사용/승인</label>
              <button class="workspace-button" type="button" id="userAdminSave">저장</button>
            </div>
            <div class="permission-grid" id="userAdminPermissions">
              __PERMISSION_CHECKBOXES__
            </div>
            <div class="admin-message" id="userAdminMessage"></div>
            <div class="admin-message">가입 요청 계정은 기본적으로 승인대기 상태입니다. 내용을 확인한 뒤 사용을 체크하고 저장하면 로그인할 수 있습니다.</div>
            <div class="admin-table-wrap">
              <table class="admin-table">
                <thead>
                  <tr>
                    <th>아이디</th>
                    <th>표시 이름</th>
                    <th>권한</th>
                    <th>세부권한</th>
                    <th>상태</th>
                    <th>생성일</th>
                    <th>마지막 로그인</th>
                    <th>수정</th>
                  </tr>
                </thead>
                <tbody id="userAdminBody"></tbody>
              </table>
            </div>
          </div>
        </div>
      </section>
"""

BACKUP_WORKSPACE_HTML = r"""
      <section class="workspace-view" id="backupWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">백업 관리</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="backupRefresh">새로고침</button>
            <button class="workspace-button" type="button" id="backupCreate">지금 백업하기</button>
            <label class="workspace-button" for="backupRestoreInput">백업 파일 복원</label>
            <input id="backupRestoreInput" type="file" accept=".zip" />
          </div>
        </div>
        <div class="workspace-mount">
          <div class="backup-panel">
            <div class="backup-summary-grid">
              <article class="backup-summary-card">
                <span>자동 백업</span>
                <strong id="backupAutoState">매일 03:00</strong>
              </article>
              <article class="backup-summary-card">
                <span>보관 기준</span>
                <strong id="backupRetentionState">최근 90일</strong>
              </article>
              <article class="backup-summary-card">
                <span>백업 위치</span>
                <strong id="backupPath">-</strong>
              </article>
            </div>
            <div class="backup-card">
              <div class="admin-section-title">자동백업 / 지정 백업 설정</div>
              <div class="admin-form">
                <label class="admin-check"><input id="backupAutoEnabled" type="checkbox" checked /> 자동 백업 사용</label>
                <label>자동 백업 시간
                  <input id="backupAutoHour" type="number" min="0" max="23" value="3" />
                </label>
                <label>보관 기간(일)
                  <input id="backupRetentionDays" type="number" min="1" max="3650" value="90" />
                </label>
                <label>백업 저장 위치
                  <input id="backupDirInput" type="text" placeholder="예) G:\내 드라이브\Soillbridge\Workhub_Backup" />
                </label>
                <button class="workspace-button" type="button" id="backupSettingsSave">백업 설정 저장</button>
                <button class="workspace-button" type="button" id="backupCreateSelected">지정 위치로 지금 백업</button>
              </div>
            </div>
            <div class="backup-card">
              <div class="admin-section-title">Google Drive 외부 백업 연동(rclone)</div>
              <div class="admin-form">
                <label class="admin-check"><input id="backupExternalEnabled" type="checkbox" /> 백업 후 Google Drive 업로드 사용</label>
                <label>rclone 실행 파일
                  <input id="backupRcloneExecutable" type="text" placeholder="예) rclone 또는 /usr/bin/rclone" />
                </label>
                <label>rclone 원격 이름
                  <input id="backupRcloneRemote" type="text" placeholder="예) gdrive" />
                </label>
                <label>Google Drive 저장 폴더
                  <input id="backupRclonePath" type="text" placeholder="예) Soillbridge/Workhub_Backup" />
                </label>
              </div>
              <div class="backup-message" id="backupExternalStatus">Google Drive 업로드 상태: 사용 안 함</div>
            </div>
            <div class="backup-card">
              <p class="backup-note">
                백업에는 업무 DB, 메일 설정, 업체 주소록, 암호화 키가 포함됩니다. VPS 운영 시에는 서버 내부 백업 폴더를 기본으로 두고, 필요하면 rclone/Google Drive 동기화 대상 폴더를 지정하면 됩니다.
                복원 전에는 현재 데이터를 자동으로 한 번 더 백업합니다.
              </p>
              <div class="backup-message" id="backupMessage"></div>
              <table class="backup-table">
                <thead>
                  <tr>
                    <th>백업 파일</th>
                    <th>생성일</th>
                    <th>크기</th>
                    <th>관리</th>
                  </tr>
                </thead>
                <tbody id="backupBody"></tbody>
              </table>
            </div>
          </div>
        </div>
      </section>
"""

SYSTEM_WORKSPACE_HTML = r"""
      <section class="workspace-view" id="systemUpdateWorkspace">
        <div class="workspace-head">
          <div class="workspace-title">업데이트 관리</div>
          <div class="workspace-actions">
            <button class="workspace-button" type="button" id="systemUpdateRefresh">새로고침</button>
            <button class="workspace-button" type="button" id="systemUpdateCheck">업데이트 확인</button>
            <button class="workspace-button" type="button" id="systemUpdateApply">업데이트 적용</button>
          </div>
        </div>
        <div class="workspace-mount">
          <div class="system-panel">
            <div class="system-summary-grid">
              <article class="system-summary-card">
                <span>실행 위치</span>
                <strong id="systemUpdateSource">-</strong>
              </article>
              <article class="system-summary-card">
                <span>브랜치</span>
                <strong id="systemUpdateBranch">-</strong>
              </article>
              <article class="system-summary-card">
                <span>현재 버전</span>
                <strong id="systemUpdateCurrent">-</strong>
              </article>
              <article class="system-summary-card">
                <span>업데이트 상태</span>
                <strong id="systemUpdateState">-</strong>
              </article>
            </div>
            <div class="system-card">
              <p class="system-note">
                업데이트 적용 시 현재 업무 데이터가 자동 백업된 뒤 GitHub 최신 코드가 적용됩니다. 적용 후에는 프로그램/나스 서비스를 다시 시작해주세요.
              </p>
              <div class="system-message" id="systemUpdateMessage"></div>
              <table class="system-table">
                <thead>
                  <tr>
                    <th>일시</th>
                    <th>작업</th>
                    <th>결과</th>
                    <th>이전 버전</th>
                    <th>적용 버전</th>
                    <th>백업</th>
                    <th>메시지</th>
                  </tr>
                </thead>
                <tbody id="systemUpdateHistoryBody"></tbody>
              </table>
            </div>
          </div>
        </div>
      </section>
"""


def safe_filename(filename: str) -> str:
    filename = Path(filename).name
    filename = re.sub(r'[<>:"/\\|?*]+', "_", filename)
    return filename or "upload.xlsx"


def original_uploaded_filename(filename: str) -> str:
    return re.sub(r"^\d{10,}_", "", filename)


def parse_multipart(headers, rfile) -> dict[str, tuple[str, bytes] | str]:
    content_type = headers.get("Content-Type", "")
    boundary_match = re.search(r"boundary=(.+)", content_type)
    if not boundary_match:
        raise ValueError("업로드 형식이 올바르지 않습니다.")

    boundary = boundary_match.group(1).strip().strip('"').encode()
    length = int(headers.get("Content-Length", "0"))
    body = rfile.read(length)
    fields: dict[str, tuple[str, bytes] | str] = {}

    for part in body.split(b"--" + boundary):
        part = part.strip()
        if not part or part == b"--":
            continue
        if part.endswith(b"--"):
            part = part[:-2].strip()
        if b"\r\n\r\n" not in part:
            continue

        raw_headers, data = part.split(b"\r\n\r\n", 1)
        data = data.rstrip(b"\r\n")
        disposition = ""
        for line in raw_headers.decode("utf-8", errors="ignore").split("\r\n"):
            if line.lower().startswith("content-disposition:"):
                disposition = line
                break

        name_match = re.search(r'name="([^"]+)"', disposition)
        if not name_match:
            continue
        field_name = name_match.group(1)
        filename_match = re.search(r'filename="([^"]*)"', disposition)
        if filename_match:
            fields[field_name] = (safe_filename(filename_match.group(1)), data)
        else:
            fields[field_name] = data.decode("utf-8", errors="ignore")

    return fields


def collect_mail_attachments(fields: dict[str, tuple[str, bytes] | str]) -> list[dict[str, object]]:
    attachments: list[dict[str, object]] = []
    total_size = 0
    for field_name, value in fields.items():
        if not field_name.startswith("cs_attachment_") or not isinstance(value, tuple):
            continue
        filename, data = value
        if not filename or not data:
            continue
        total_size += len(data)
        if total_size > MAX_MAIL_ATTACHMENT_BYTES:
            raise ValueError("첨부파일 총 용량은 20MB 이하로 업로드해주세요.")
        content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
        attachments.append({
            "filename": filename,
            "data": data,
            "content_type": content_type,
        })
    return attachments


def save_uploaded_file(fields: dict[str, tuple[str, bytes] | str], field_name: str = "file") -> Path:
    uploaded = fields.get(field_name)
    if not isinstance(uploaded, tuple):
        raise ValueError("업로드된 파일이 없습니다.")

    filename, data = uploaded
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("xlsx 또는 xlsm 파일만 업로드할 수 있습니다.")

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    path = UPLOAD_DIR / f"{int(time.time() * 1000)}_{filename}"
    path.write_bytes(data)
    return path


def save_uploaded_shared_file(fields: dict[str, tuple[str, bytes] | str], field_name: str = "file") -> Path:
    uploaded = fields.get(field_name)
    if not isinstance(uploaded, tuple):
        raise ValueError("업로드된 업무 파일이 없습니다.")

    filename, data = uploaded
    filename = safe_filename(filename)
    if not filename or not data:
        raise ValueError("업무 파일을 다시 선택해주세요.")

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    path = UPLOAD_DIR / f"{int(time.time() * 1000)}_{filename}"
    path.write_bytes(data)
    return path


def save_uploaded_sales_report_file(fields: dict[str, tuple[str, bytes] | str], field_name: str = "file") -> Path:
    uploaded = fields.get(field_name)
    if not isinstance(uploaded, tuple):
        raise ValueError("업로드된 매출표 파일이 없습니다.")

    filename, data = uploaded
    filename = safe_filename(filename)
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls", ".csv")):
        raise ValueError("매출표는 xlsx, xlsm, xls, csv 파일만 업로드할 수 있습니다.")
    if not data:
        raise ValueError("매출표 파일을 다시 선택해주세요.")

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    path = UPLOAD_DIR / f"{int(time.time() * 1000)}_{filename}"
    path.write_bytes(data)
    return path


def save_uploaded_backup_zip(fields: dict[str, tuple[str, bytes] | str], field_name: str = "file") -> Path:
    uploaded = fields.get(field_name)
    if not isinstance(uploaded, tuple):
        raise ValueError("업로드된 백업 파일이 없습니다.")

    filename, data = uploaded
    if not filename.lower().endswith(".zip"):
        raise ValueError("zip 백업 파일만 업로드할 수 있습니다.")

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    path = UPLOAD_DIR / f"{int(time.time() * 1000)}_{safe_filename(filename)}"
    path.write_bytes(data)
    return path


def parse_receipt_date(value: object) -> date:
    if isinstance(value, str) and value:
        try:
            return date.fromisoformat(value)
        except ValueError:
            pass
    return date.today()


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _safe_order_download_filename(filename: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", Path(filename).name).strip(" .")
    return cleaned or "order-output.xlsx"


def _load_order_download_history() -> list[dict[str, str | int]]:
    if not ORDER_DOWNLOAD_HISTORY_PATH.exists():
        return []
    try:
        data = json.loads(ORDER_DOWNLOAD_HISTORY_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    return [item for item in data if isinstance(item, dict)]


def _save_order_download_history(history: list[dict[str, str | int]]) -> None:
    ORDER_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    ORDER_DOWNLOAD_HISTORY_PATH.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def _public_order_download(item: dict[str, str | int]) -> dict[str, str | int]:
    return {
        "id": str(item.get("id", "")),
        "filename": str(item.get("filename", "")),
        "workflow": str(item.get("workflow", "")),
        "created_at": str(item.get("created_at", "")),
        "size": int(item.get("size", 0) or 0),
    }


def list_order_downloads() -> list[dict[str, str | int]]:
    ORDER_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    active: list[dict[str, str | int]] = []
    removed: list[dict[str, str | int]] = []
    base_dir = ORDER_DOWNLOAD_DIR.resolve()
    for item in _load_order_download_history():
        stored_name = str(item.get("stored_name", ""))
        if not stored_name:
            continue
        target = (ORDER_DOWNLOAD_DIR / stored_name).resolve()
        if base_dir in target.parents and target.is_file():
            active.append(item)
        else:
            removed.append(item)

    keep = active[:ORDER_DOWNLOAD_LIMIT]
    stale = active[ORDER_DOWNLOAD_LIMIT:] + removed
    for item in stale:
        stored_name = str(item.get("stored_name", ""))
        if not stored_name:
            continue
        target = (ORDER_DOWNLOAD_DIR / stored_name).resolve()
        if base_dir in target.parents and target.exists():
            target.unlink(missing_ok=True)

    if len(keep) != len(active) or removed:
        _save_order_download_history(keep)
    return [_public_order_download(item) for item in keep]


def register_order_download(output_path: str | Path, workflow: str) -> dict[str, str | int]:
    source = Path(output_path)
    if not source.is_file():
        raise FileNotFoundError("발주업무 출력 파일을 찾지 못했습니다.")

    ORDER_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    download_id = secrets.token_hex(10)
    safe_filename = _safe_order_download_filename(source.name)
    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{download_id}_{safe_filename}"
    target = ORDER_DOWNLOAD_DIR / stored_name
    target.write_bytes(source.read_bytes())

    item: dict[str, str | int] = {
        "id": download_id,
        "filename": source.name,
        "stored_name": stored_name,
        "workflow": workflow,
        "created_at": now_text(),
        "size": target.stat().st_size,
    }
    history = [item] + _load_order_download_history()
    _save_order_download_history(history[:ORDER_DOWNLOAD_LIMIT])

    base_dir = ORDER_DOWNLOAD_DIR.resolve()
    for stale in history[ORDER_DOWNLOAD_LIMIT:]:
        stale_name = str(stale.get("stored_name", ""))
        stale_path = (ORDER_DOWNLOAD_DIR / stale_name).resolve()
        if base_dir in stale_path.parents and stale_path.exists():
            stale_path.unlink(missing_ok=True)
    return _public_order_download(item)


def order_download_path(download_id: str) -> Path:
    base_dir = ORDER_DOWNLOAD_DIR.resolve()
    for item in _load_order_download_history():
        if str(item.get("id", "")) != str(download_id):
            continue
        stored_name = str(item.get("stored_name", ""))
        target = (ORDER_DOWNLOAD_DIR / stored_name).resolve()
        if base_dir not in target.parents or not target.is_file():
            raise FileNotFoundError("출력 파일을 찾지 못했습니다.")
        return target
    raise FileNotFoundError("출력 파일을 찾지 못했습니다.")


def order_download_filename(download_id: str) -> str:
    for item in _load_order_download_history():
        if str(item.get("id", "")) == str(download_id):
            return str(item.get("filename", "")) or "발주업무_출력.xlsx"
    raise FileNotFoundError("출력 파일을 찾지 못했습니다.")


def _safe_shared_filename(filename: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", Path(filename).name).strip(" .")
    return cleaned or "workhub-file"


def _shared_file_public(row: sqlite3.Row | dict) -> dict[str, str | int]:
    return {
        "id": int(row["id"]),
        "original_name": str(row["original_name"]),
        "size": int(row["size"] or 0),
        "uploaded_by": str(row["uploaded_by"] or ""),
        "uploaded_at": str(row["uploaded_at"] or ""),
    }


def list_shared_files() -> list[dict[str, str | int]]:
    init_db()
    SHARED_FILE_DIR.mkdir(parents=True, exist_ok=True)
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT id, original_name, size, uploaded_by, uploaded_at
              FROM shared_files
             ORDER BY id DESC
            """
        ).fetchall()
        return [_shared_file_public(row) for row in rows]
    finally:
        connection.close()


def save_shared_file(source_path: str | Path, original_name: str, uploaded_by: str = "") -> dict[str, str | int]:
    source = Path(source_path)
    if not source.is_file():
        raise FileNotFoundError("올릴 파일을 찾지 못했습니다.")

    init_db()
    SHARED_FILE_DIR.mkdir(parents=True, exist_ok=True)
    safe_original = _safe_shared_filename(original_name or source.name)
    file_id = secrets.token_hex(10)
    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file_id}_{safe_original}"
    target = SHARED_FILE_DIR / stored_name
    target.write_bytes(source.read_bytes())
    uploaded_at = now_text()

    connection = connect_db()
    try:
        cursor = connection.execute(
            """
            INSERT INTO shared_files (stored_name, original_name, size, uploaded_by, uploaded_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (stored_name, safe_original, target.stat().st_size, str(uploaded_by or ""), uploaded_at),
        )
        connection.commit()
        row_id = int(cursor.lastrowid)
    finally:
        connection.close()
    return {
        "id": row_id,
        "original_name": safe_original,
        "size": target.stat().st_size,
        "uploaded_by": str(uploaded_by or ""),
        "uploaded_at": uploaded_at,
    }


def shared_file_download_info(file_id: object) -> tuple[Path, dict[str, str | int]]:
    init_db()
    try:
        numeric_id = int(file_id)
    except (TypeError, ValueError) as exc:
        raise FileNotFoundError("업무 파일을 찾지 못했습니다.") from exc

    connection = connect_db()
    try:
        row = connection.execute(
            """
            SELECT id, stored_name, original_name, size, uploaded_by, uploaded_at
              FROM shared_files
             WHERE id = ?
            """,
            (numeric_id,),
        ).fetchone()
    finally:
        connection.close()
    if not row:
        raise FileNotFoundError("업무 파일을 찾지 못했습니다.")

    base_dir = SHARED_FILE_DIR.resolve()
    target = (SHARED_FILE_DIR / str(row["stored_name"])).resolve()
    if base_dir not in target.parents or not target.is_file():
        raise FileNotFoundError("업무 파일을 찾지 못했습니다.")
    return target, _shared_file_public(row)


def delete_shared_file(file_id: object) -> None:
    path, _ = shared_file_download_info(file_id)
    numeric_id = int(file_id)
    connection = connect_db()
    try:
        connection.execute("DELETE FROM shared_files WHERE id = ?", (numeric_id,))
        connection.commit()
    finally:
        connection.close()
    path.unlink(missing_ok=True)


def _sales_report_public(row: sqlite3.Row | dict) -> dict[str, str | int]:
    return {
        "id": int(row["id"]),
        "original_name": str(row["original_name"]),
        "size": int(row["size"] or 0),
        "uploaded_by": str(row["uploaded_by"] or ""),
        "uploaded_at": str(row["uploaded_at"] or ""),
    }


def list_sales_report_uploads(limit: int = 5) -> list[dict[str, str | int]]:
    init_db()
    SALES_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    safe_limit = max(1, min(int(limit or 5), 20))
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT id, original_name, size, uploaded_by, uploaded_at
              FROM sales_report_uploads
             ORDER BY id DESC
             LIMIT ?
            """,
            (safe_limit,),
        ).fetchall()
        return [_sales_report_public(row) for row in rows]
    finally:
        connection.close()


class _SalesReportTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "tr":
            self._current_row = []
        elif tag.lower() in {"td", "th"} and self._current_row is not None:
            self._current_cell = []

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        lowered = tag.lower()
        if lowered in {"td", "th"} and self._current_row is not None and self._current_cell is not None:
            self._current_row.append("".join(self._current_cell).strip())
            self._current_cell = None
        elif lowered == "tr" and self._current_row is not None:
            if any(str(cell).strip() for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None


def _sales_report_text(value: object) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def _sales_report_number(value: object) -> float:
    text = _sales_report_text(value).replace(",", "").replace("%", "")
    if not text or text.lower() == "nan":
        return 0.0
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def _sales_report_int(value: object) -> int:
    return int(round(_sales_report_number(value)))


def _sales_report_percent(value: object) -> float:
    text = _sales_report_text(value)
    number = _sales_report_number(text)
    if "%" in text:
        return round(number / 100, 6)
    return round(number, 6)


def _sales_report_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    match = re.search(r"(20\d{2})[-./](\d{1,2})[-./](\d{1,2})", _sales_report_text(value))
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def _sales_report_period(value: object) -> str:
    parsed = _sales_report_date(value)
    if parsed:
        return parsed[:7]
    match = re.search(r"(20\d{2})[-./](\d{1,2})", _sales_report_text(value))
    if match:
        year, month = match.groups()
        return f"{int(year):04d}-{int(month):02d}"
    return ""


def _sales_report_rows_from_xlsx(path: str | Path) -> list[list[object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows: list[list[object]] = []
        for row in worksheet.iter_rows(values_only=True):
            values = list(row)
            if any(_sales_report_text(value) for value in values):
                rows.append(values)
        return rows
    finally:
        workbook.close()


def _sales_report_xlsx_sheet_title(path: str | Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        return str(workbook.worksheets[0].title or "")
    finally:
        workbook.close()


def _sales_report_rows_from_html_table(path: str | Path) -> list[list[str]]:
    data = Path(path).read_bytes()
    text = ""
    for encoding in ("utf-8", "euc-kr", "cp949"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        text = data.decode("utf-8", errors="ignore")
    parser = _SalesReportTableParser()
    parser.feed(text)
    return parser.rows


def _sales_report_table(path: str | Path) -> tuple[list[str], list[list[object]]]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".xls" and source.read_bytes()[:32].lstrip().lower().startswith(b"<!doctype"):
        rows = _sales_report_rows_from_html_table(source)
    else:
        rows = _sales_report_rows_from_xlsx(source)
    if not rows:
        return [], []
    headers = [_sales_report_text(value) for value in rows[0]]
    return headers, rows[1:]


def detect_sales_report_type(path: str | Path, original_name: str = "") -> str:
    try:
        headers, _ = _sales_report_table(path)
    except Exception:
        return ""
    header_set = {_sales_report_text(header) for header in headers}
    if "일자" in header_set:
        return "daily"
    if "판매사" in header_set:
        return "seller"
    if "공급사" in header_set:
        return "supplier"
    if "상품코드" in header_set and "상품명" in header_set:
        return "product"
    lowered_name = str(original_name or Path(path).name).lower()
    if "statistics_good" in lowered_name:
        return "product"
    return ""


def _sales_report_row_dict(headers: list[str], row: list[object]) -> dict[str, object]:
    return {header: row[index] if index < len(row) else "" for index, header in enumerate(headers)}


def _sales_report_value(row: dict[str, object], key: str) -> object:
    return row.get(key, "")


def _parse_daily_sales_report(headers: list[str], rows: list[list[object]]) -> dict[str, object]:
    parsed_rows: list[dict[str, object]] = []
    for raw in rows:
        row = _sales_report_row_dict(headers, raw)
        report_date = _sales_report_date(_sales_report_value(row, "일자"))
        if not report_date:
            continue
        parsed_rows.append({
            "report_date": report_date,
            "period": report_date[:7],
            "label": _sales_report_text(_sales_report_value(row, "일자")),
            "quantity": _sales_report_int(_sales_report_value(row, "판매-수량")),
            "sales_amount": _sales_report_int(_sales_report_value(row, "판매-금액")),
            "supply_amount": _sales_report_int(_sales_report_value(row, "판매-공급금액")),
            "sales_total": _sales_report_int(_sales_report_value(row, "판매-판매합계")),
            "supply_total": _sales_report_int(_sales_report_value(row, "판매-공급합계")),
            "sales_margin": _sales_report_int(_sales_report_value(row, "판매-마진")),
            "cs_margin": _sales_report_int(_sales_report_value(row, "CS-마진")),
            "profit_quantity_sales": _sales_report_int(_sales_report_value(row, "손익-수량 판매사기준")),
            "profit_quantity_supply": _sales_report_int(_sales_report_value(row, "손익-수량 공급사기준")),
            "profit_sales_amount": _sales_report_int(_sales_report_value(row, "손익-판매금액")),
            "profit_supply_amount": _sales_report_int(_sales_report_value(row, "손익-공급금액")),
            "profit_sales_margin": _sales_report_int(_sales_report_value(row, "손익-판매마진")),
            "profit_shipping_sales": _sales_report_int(_sales_report_value(row, "손익-판매배송비")),
            "profit_shipping_supply": _sales_report_int(_sales_report_value(row, "손익-공급배송비")),
            "profit_shipping": _sales_report_int(_sales_report_value(row, "손익-배송비")),
            "profit_margin": _sales_report_int(_sales_report_value(row, "손익-마진")),
            "margin_rate": _sales_report_percent(_sales_report_value(row, "손익-마진율")),
        })
    report_date = parsed_rows[0]["report_date"] if parsed_rows else ""
    period = str(report_date)[:7] if report_date else ""
    return {"report_type": "daily", "report_date": report_date, "period": period, "rows": parsed_rows}


def _parse_seller_sales_report(headers: list[str], rows: list[list[object]], original_name: str = "") -> dict[str, object]:
    parsed_rows: list[dict[str, object]] = []
    for raw in rows:
        row = _sales_report_row_dict(headers, raw)
        name = _sales_report_text(_sales_report_value(row, "판매사"))
        if not name:
            continue
        parsed_rows.append({
            "name": name,
            "quantity": _sales_report_int(_sales_report_value(row, "판매-수량")),
            "sales_amount": _sales_report_int(_sales_report_value(row, "판매-금액")),
            "supply_amount": _sales_report_int(_sales_report_value(row, "판매-공급금액")),
            "sales_total": _sales_report_int(_sales_report_value(row, "판매-판매합계")),
            "supply_total": _sales_report_int(_sales_report_value(row, "판매-공급합계")),
            "sales_margin": _sales_report_int(_sales_report_value(row, "판매-마진")),
            "cs_amount": _sales_report_int(_sales_report_value(row, "CS-금액")),
            "cs_supply_amount": _sales_report_int(_sales_report_value(row, "CS-공급금액")),
            "cs_margin": _sales_report_int(_sales_report_value(row, "CS-마진")),
            "profit_quantity_sales": _sales_report_int(_sales_report_value(row, "손익-수량 판매사기준")),
            "profit_quantity_supply": _sales_report_int(_sales_report_value(row, "손익-수량 공급사기준")),
            "profit_sales_amount": _sales_report_int(_sales_report_value(row, "손익-판매금액")),
            "profit_supply_amount": _sales_report_int(_sales_report_value(row, "손익-공급금액")),
            "profit_sales_margin": _sales_report_int(_sales_report_value(row, "손익-판매마진")),
            "profit_shipping": _sales_report_int(_sales_report_value(row, "손익-배송비")),
            "profit_margin": _sales_report_int(_sales_report_value(row, "손익-마진")),
            "margin_rate": _sales_report_percent(_sales_report_value(row, "손익-마진율")),
        })
    period = _sales_report_period(original_name)
    return {"report_type": "seller", "report_date": _sales_report_date(original_name), "period": period, "rows": parsed_rows}


def _parse_supplier_sales_report(headers: list[str], rows: list[list[object]], original_name: str = "") -> dict[str, object]:
    parsed_rows: list[dict[str, object]] = []
    for raw in rows:
        row = _sales_report_row_dict(headers, raw)
        name = _sales_report_text(_sales_report_value(row, "공급사"))
        if not name:
            continue
        parsed_rows.append({
            "name": name,
            "quantity": _sales_report_int(_sales_report_value(row, "판매-수량")),
            "sales_amount": _sales_report_int(_sales_report_value(row, "판매-금액")),
            "supply_amount": _sales_report_int(_sales_report_value(row, "판매-공급금액")),
            "sales_total": _sales_report_int(_sales_report_value(row, "판매-판매합계")),
            "supply_total": _sales_report_int(_sales_report_value(row, "판매-공급합계")),
            "sales_margin": _sales_report_int(_sales_report_value(row, "판매-마진")),
            "cs_amount": _sales_report_int(_sales_report_value(row, "CS-금액")),
            "cs_supply_amount": _sales_report_int(_sales_report_value(row, "CS-공급금액")),
            "cs_margin": _sales_report_int(_sales_report_value(row, "CS-마진")),
            "profit_quantity_sales": _sales_report_int(_sales_report_value(row, "손익-수량 판매사기준")),
            "profit_quantity_supply": _sales_report_int(_sales_report_value(row, "손익-수량 공급사기준")),
            "profit_sales_amount": _sales_report_int(_sales_report_value(row, "손익-판매금액")),
            "profit_supply_amount": _sales_report_int(_sales_report_value(row, "손익-공급금액")),
            "profit_sales_margin": _sales_report_int(_sales_report_value(row, "손익-판매마진")),
            "profit_shipping": _sales_report_int(_sales_report_value(row, "손익-배송비")),
            "profit_margin": _sales_report_int(_sales_report_value(row, "손익-마진")),
            "margin_rate": _sales_report_percent(_sales_report_value(row, "손익-마진율")),
        })
    return {
        "report_type": "supplier",
        "report_date": _sales_report_date(original_name),
        "period": _sales_report_period(original_name),
        "rows": parsed_rows,
    }


def _parse_product_sales_report(headers: list[str], rows: list[list[object]], original_name: str = "") -> dict[str, object]:
    parsed_rows: list[dict[str, object]] = []
    for raw in rows:
        row = _sales_report_row_dict(headers, raw)
        code = _sales_report_text(_sales_report_value(row, "상품코드"))
        name = _sales_report_text(_sales_report_value(row, "상품명"))
        if not code and not name:
            continue
        parsed_rows.append({
            "code": code,
            "name": name,
            "quantity": _sales_report_int(_sales_report_value(row, "판매-수량")),
            "sales_amount": _sales_report_int(_sales_report_value(row, "판매-금액")),
            "supply_amount": _sales_report_int(_sales_report_value(row, "판매-공급금액")),
            "sales_margin": _sales_report_int(_sales_report_value(row, "판매-마진")),
            "cs_amount": _sales_report_int(_sales_report_value(row, "CS-금액")),
            "cs_supply_amount": _sales_report_int(_sales_report_value(row, "CS-공급금액")),
            "cs_margin": _sales_report_int(_sales_report_value(row, "CS-마진")),
            "profit_quantity_sales": _sales_report_int(_sales_report_value(row, "손익-수량 판매사기준")),
            "profit_quantity_supply": _sales_report_int(_sales_report_value(row, "손익-수량 공급사기준")),
            "profit_sales_amount": _sales_report_int(_sales_report_value(row, "손익-판매금액")),
            "profit_supply_amount": _sales_report_int(_sales_report_value(row, "손익-공급금액")),
            "profit_margin": _sales_report_int(_sales_report_value(row, "손익-마진")),
            "margin_rate": _sales_report_percent(_sales_report_value(row, "손익-마진율")),
        })
    return {
        "report_type": "product",
        "report_date": _sales_report_date(original_name),
        "period": _sales_report_period(original_name),
        "rows": parsed_rows,
    }


def parse_sales_report_file(path: str | Path, original_name: str = "") -> dict[str, object]:
    headers, rows = _sales_report_table(path)
    report_type = detect_sales_report_type(path, original_name)
    if report_type == "daily":
        return _parse_daily_sales_report(headers, rows)
    if report_type == "seller":
        parsed = _parse_seller_sales_report(headers, rows, original_name or Path(path).name)
        if not parsed.get("period") and Path(path).suffix.lower() in {".xlsx", ".xlsm"}:
            parsed["period"] = _sales_report_period(_sales_report_xlsx_sheet_title(path))
        return parsed
    if report_type == "supplier":
        return _parse_supplier_sales_report(headers, rows, original_name or Path(path).name)
    if report_type == "product":
        return _parse_product_sales_report(headers, rows, original_name or Path(path).name)
    raise ValueError("지원하는 매출표 형식이 아닙니다.")


def _sales_row_sum(rows: list[dict[str, object]], key: str) -> int:
    return int(sum(_sales_report_int(row.get(key, 0)) for row in rows))


def save_sales_report_snapshot(file_id: int, parsed: dict[str, object]) -> None:
    report_type = str(parsed.get("report_type") or "")
    report_date = str(parsed.get("report_date") or "")
    period = str(parsed.get("period") or report_date[:7] or "")
    rows = [row for row in parsed.get("rows", []) if isinstance(row, dict)]
    connection = connect_db()
    try:
        if report_type == "daily":
            for row in rows:
                connection.execute(
                    """
                    INSERT INTO sales_report_daily_rows (
                        report_date, period, file_id, label, quantity, sales_amount, supply_amount,
                        sales_total, supply_total, sales_margin, cs_margin, profit_quantity_sales,
                        profit_quantity_supply, profit_sales_amount, profit_supply_amount,
                        profit_sales_margin, profit_shipping_sales, profit_shipping_supply,
                        profit_shipping, profit_margin, margin_rate
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(report_date) DO UPDATE SET
                        period = excluded.period,
                        file_id = excluded.file_id,
                        label = excluded.label,
                        quantity = excluded.quantity,
                        sales_amount = excluded.sales_amount,
                        supply_amount = excluded.supply_amount,
                        sales_total = excluded.sales_total,
                        supply_total = excluded.supply_total,
                        sales_margin = excluded.sales_margin,
                        cs_margin = excluded.cs_margin,
                        profit_quantity_sales = excluded.profit_quantity_sales,
                        profit_quantity_supply = excluded.profit_quantity_supply,
                        profit_sales_amount = excluded.profit_sales_amount,
                        profit_supply_amount = excluded.profit_supply_amount,
                        profit_sales_margin = excluded.profit_sales_margin,
                        profit_shipping_sales = excluded.profit_shipping_sales,
                        profit_shipping_supply = excluded.profit_shipping_supply,
                        profit_shipping = excluded.profit_shipping,
                        profit_margin = excluded.profit_margin,
                        margin_rate = excluded.margin_rate
                    """,
                    (
                        row.get("report_date"),
                        row.get("period"),
                        file_id,
                        row.get("label", ""),
                        row.get("quantity", 0),
                        row.get("sales_amount", 0),
                        row.get("supply_amount", 0),
                        row.get("sales_total", 0),
                        row.get("supply_total", 0),
                        row.get("sales_margin", 0),
                        row.get("cs_margin", 0),
                        row.get("profit_quantity_sales", 0),
                        row.get("profit_quantity_supply", 0),
                        row.get("profit_sales_amount", 0),
                        row.get("profit_supply_amount", 0),
                        row.get("profit_sales_margin", 0),
                        row.get("profit_shipping_sales", 0),
                        row.get("profit_shipping_supply", 0),
                        row.get("profit_shipping", 0),
                        row.get("profit_margin", 0),
                        row.get("margin_rate", 0),
                    ),
                )
        elif report_type == "seller":
            connection.execute("DELETE FROM sales_report_seller_rows WHERE period = ?", (period,))
            for row in rows:
                connection.execute(
                    """
                    INSERT INTO sales_report_seller_rows (
                        period, report_date, file_id, seller_name, quantity, sales_amount, supply_amount,
                        sales_total, supply_total, sales_margin, cs_amount, cs_supply_amount, cs_margin,
                        profit_quantity_sales, profit_quantity_supply, profit_sales_amount, profit_supply_amount,
                        profit_sales_margin, profit_shipping, profit_margin, margin_rate
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        period,
                        report_date,
                        file_id,
                        row.get("name", ""),
                        row.get("quantity", 0),
                        row.get("sales_amount", 0),
                        row.get("supply_amount", 0),
                        row.get("sales_total", 0),
                        row.get("supply_total", 0),
                        row.get("sales_margin", 0),
                        row.get("cs_amount", 0),
                        row.get("cs_supply_amount", 0),
                        row.get("cs_margin", 0),
                        row.get("profit_quantity_sales", 0),
                        row.get("profit_quantity_supply", 0),
                        row.get("profit_sales_amount", 0),
                        row.get("profit_supply_amount", 0),
                        row.get("profit_sales_margin", 0),
                        row.get("profit_shipping", 0),
                        row.get("profit_margin", 0),
                        row.get("margin_rate", 0),
                    ),
                )
        elif report_type == "supplier":
            connection.execute("DELETE FROM sales_report_supplier_rows WHERE period = ?", (period,))
            for row in rows:
                connection.execute(
                    """
                    INSERT INTO sales_report_supplier_rows (
                        period, report_date, file_id, supplier_name, quantity, sales_amount, supply_amount,
                        sales_total, supply_total, sales_margin, cs_amount, cs_supply_amount, cs_margin,
                        profit_quantity_sales, profit_quantity_supply, profit_sales_amount, profit_supply_amount,
                        profit_sales_margin, profit_shipping, profit_margin, margin_rate
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        period,
                        report_date,
                        file_id,
                        row.get("name", ""),
                        row.get("quantity", 0),
                        row.get("sales_amount", 0),
                        row.get("supply_amount", 0),
                        row.get("sales_total", 0),
                        row.get("supply_total", 0),
                        row.get("sales_margin", 0),
                        row.get("cs_amount", 0),
                        row.get("cs_supply_amount", 0),
                        row.get("cs_margin", 0),
                        row.get("profit_quantity_sales", 0),
                        row.get("profit_quantity_supply", 0),
                        row.get("profit_sales_amount", 0),
                        row.get("profit_supply_amount", 0),
                        row.get("profit_sales_margin", 0),
                        row.get("profit_shipping", 0),
                        row.get("profit_margin", 0),
                        row.get("margin_rate", 0),
                    ),
                )
        elif report_type == "product":
            if report_date:
                connection.execute("DELETE FROM sales_report_product_rows WHERE report_date = ?", (report_date,))
            else:
                connection.execute("DELETE FROM sales_report_product_rows WHERE file_id = ?", (file_id,))
            for row in rows:
                connection.execute(
                    """
                    INSERT INTO sales_report_product_rows (
                        period, report_date, file_id, product_code, product_name, quantity, sales_amount,
                        supply_amount, sales_margin, cs_amount, cs_supply_amount, cs_margin,
                        profit_quantity_sales, profit_quantity_supply, profit_sales_amount, profit_supply_amount,
                        profit_margin, margin_rate
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        period,
                        report_date,
                        file_id,
                        row.get("code", ""),
                        row.get("name", ""),
                        row.get("quantity", 0),
                        row.get("sales_amount", 0),
                        row.get("supply_amount", 0),
                        row.get("sales_margin", 0),
                        row.get("cs_amount", 0),
                        row.get("cs_supply_amount", 0),
                        row.get("cs_margin", 0),
                        row.get("profit_quantity_sales", 0),
                        row.get("profit_quantity_supply", 0),
                        row.get("profit_sales_amount", 0),
                        row.get("profit_supply_amount", 0),
                        row.get("profit_margin", 0),
                        row.get("margin_rate", 0),
                    ),
                )
        connection.commit()
    finally:
        connection.close()


def _sales_report_daily_public(row: sqlite3.Row | None) -> dict[str, object]:
    if row is None:
        return {}
    return {
        "report_date": str(row["report_date"] or ""),
        "label": str(row["label"] or ""),
        "quantity": int(row["quantity"] or 0),
        "sales_amount": int(row["sales_amount"] or 0),
        "sales_total": int(row["sales_total"] or 0),
        "profit_sales_amount": int(row["profit_sales_amount"] or 0),
        "profit_supply_amount": int(row["profit_supply_amount"] or 0),
        "profit_margin": int(row["profit_margin"] or 0),
        "margin_rate": float(row["margin_rate"] or 0),
    }


def _sales_report_named_public(row: sqlite3.Row) -> dict[str, object]:
    name = row["name"] if "name" in row.keys() else row[0]
    return {
        "name": str(name or ""),
        "quantity": int(row["quantity"] or 0),
        "sales_amount": int(row["sales_amount"] or 0),
        "profit_sales_amount": int(row["profit_sales_amount"] or 0),
        "profit_margin": int(row["profit_margin"] or 0),
        "margin_rate": float(row["margin_rate"] or 0),
        "cs_amount": int(row["cs_amount"] or 0) if "cs_amount" in row.keys() else 0,
        "cs_margin": int(row["cs_margin"] or 0) if "cs_margin" in row.keys() else 0,
    }


def _sales_report_purchase_public(row: sqlite3.Row) -> dict[str, object]:
    return {
        "name": str(row["name"] or ""),
        "quantity": int(row["quantity"] or 0),
        "purchase_total": int(row["purchase_total"] or 0),
    }


def sales_report_dashboard_payload(period: str = "", report_date: str = "") -> dict[str, object]:
    init_db()
    connection = connect_db()
    try:
        selected_date = report_date or ""
        if not selected_date:
            row = connection.execute("SELECT MAX(report_date) AS report_date FROM sales_report_daily_rows").fetchone()
            selected_date = str(row["report_date"] or "") if row else ""
        selected_period = period or selected_date[:7]
        if not selected_period:
            row = connection.execute(
                """
                SELECT period
                  FROM (
                        SELECT period FROM sales_report_daily_rows WHERE period != ''
                        UNION ALL
                        SELECT period FROM sales_report_seller_rows WHERE period != ''
                        UNION ALL
                        SELECT period FROM sales_report_supplier_rows WHERE period != ''
                        UNION ALL
                        SELECT period FROM sales_report_product_rows WHERE period != ''
                       )
                 ORDER BY period DESC
                 LIMIT 1
                """
            ).fetchone()
            selected_period = str(row["period"] or "") if row else ""
        if not selected_date and selected_period:
            row = connection.execute(
                "SELECT MAX(report_date) AS report_date FROM sales_report_daily_rows WHERE period = ?",
                (selected_period,),
            ).fetchone()
            selected_date = str(row["report_date"] or "") if row else ""
        if not selected_date and selected_period:
            row = connection.execute(
                "SELECT MAX(report_date) AS report_date FROM sales_report_product_rows WHERE period = ? AND report_date != ''",
                (selected_period,),
            ).fetchone()
            selected_date = str(row["report_date"] or "") if row else ""
        today = connection.execute(
            "SELECT * FROM sales_report_daily_rows WHERE report_date = ?",
            (selected_date,),
        ).fetchone()
        yesterday = None
        if selected_date:
            previous_date = (date.fromisoformat(selected_date) - timedelta(days=1)).isoformat()
            yesterday = connection.execute(
                "SELECT * FROM sales_report_daily_rows WHERE report_date = ?",
                (previous_date,),
            ).fetchone()
        month = connection.execute(
            """
            SELECT COALESCE(SUM(quantity), 0) AS quantity,
                   COALESCE(SUM(sales_amount), 0) AS sales_amount,
                   COALESCE(SUM(sales_total), 0) AS sales_total,
                   COALESCE(SUM(profit_sales_amount), 0) AS profit_sales_amount,
                   COALESCE(SUM(profit_supply_amount), 0) AS profit_supply_amount,
                   COALESCE(SUM(profit_margin), 0) AS profit_margin
              FROM sales_report_daily_rows
             WHERE period = ?
            """,
            (selected_period,),
        ).fetchone()
        seller_total = connection.execute(
            """
            SELECT COALESCE(SUM(quantity), 0) AS quantity,
                   COALESCE(SUM(sales_amount), 0) AS sales_amount,
                   COALESCE(SUM(profit_sales_amount), 0) AS profit_sales_amount,
                   COALESCE(SUM(profit_margin), 0) AS profit_margin
              FROM sales_report_seller_rows
             WHERE period = ?
            """,
            (selected_period,),
        ).fetchone()
        daily_rows = connection.execute(
            """
            SELECT report_date, label, quantity, sales_amount, sales_total, profit_sales_amount,
                   profit_supply_amount, profit_margin, margin_rate
              FROM sales_report_daily_rows
             WHERE period = ?
             ORDER BY report_date DESC
             LIMIT 10
            """,
            (selected_period,),
        ).fetchall()
        seller_rows = connection.execute(
            """
            SELECT seller_name AS name, quantity, sales_amount, profit_sales_amount,
                   profit_margin, margin_rate, cs_amount, cs_margin
              FROM sales_report_seller_rows
             WHERE period = ?
             ORDER BY profit_sales_amount DESC
             LIMIT 10
            """,
            (selected_period,),
        ).fetchall()
        product_rows = connection.execute(
            """
            SELECT product_name AS name, quantity, sales_amount, profit_sales_amount,
                   profit_margin, margin_rate, cs_amount, cs_margin
              FROM sales_report_product_rows
             WHERE (? = '' OR report_date = ? OR period = ?)
             ORDER BY profit_sales_amount DESC
             LIMIT 10
            """,
            (selected_date, selected_date, selected_period),
        ).fetchall()
        supplier_purchase_rows = connection.execute(
            """
            SELECT supplier_name AS name,
                   COALESCE(SUM(quantity), 0) AS quantity,
                   COALESCE(SUM(supply_total), 0) AS purchase_total
              FROM sales_report_supplier_rows
             WHERE period = ?
             GROUP BY supplier_name
             ORDER BY purchase_total DESC
             LIMIT 10
            """,
            (selected_period,),
        ).fetchall()
        review_rows = connection.execute(
            """
            SELECT 'supplier' AS kind, supplier_name AS name, cs_amount, cs_margin,
                   profit_sales_amount, profit_margin
              FROM sales_report_supplier_rows
             WHERE period = ? AND (cs_amount != 0 OR cs_margin != 0 OR profit_margin < 0)
            UNION ALL
            SELECT 'seller' AS kind, seller_name AS name, cs_amount, cs_margin,
                   profit_sales_amount, profit_margin
              FROM sales_report_seller_rows
             WHERE period = ? AND (cs_amount != 0 OR cs_margin != 0 OR profit_margin < 0)
            UNION ALL
            SELECT 'product' AS kind, product_name AS name, cs_amount, cs_margin,
                   profit_sales_amount, profit_margin
              FROM sales_report_product_rows
             WHERE (? = '' OR report_date = ? OR period = ?)
               AND (cs_amount != 0 OR cs_margin != 0 OR profit_margin < 0)
             LIMIT 10
            """,
            (selected_period, selected_period, selected_date, selected_date, selected_period),
        ).fetchall()
    finally:
        connection.close()

    today_public = _sales_report_daily_public(today)
    yesterday_public = _sales_report_daily_public(yesterday)
    today_amount = int(today_public.get("profit_sales_amount", 0) or 0)
    yesterday_amount = int(yesterday_public.get("profit_sales_amount", 0) or 0)
    delta = today_amount - yesterday_amount
    delta_rate = round((delta / yesterday_amount) * 100, 1) if yesterday_amount else 0
    month_total = {
        "quantity": int(month["quantity"] or 0),
        "sales_amount": int(month["sales_amount"] or 0),
        "sales_total": int(month["sales_total"] or 0),
        "profit_sales_amount": int(month["profit_sales_amount"] or 0),
        "profit_supply_amount": int(month["profit_supply_amount"] or 0),
        "profit_margin": int(month["profit_margin"] or 0),
    }
    seller_total_public = {
        "quantity": int(seller_total["quantity"] or 0),
        "sales_amount": int(seller_total["sales_amount"] or 0),
        "profit_sales_amount": int(seller_total["profit_sales_amount"] or 0),
        "profit_margin": int(seller_total["profit_margin"] or 0),
    }
    difference = month_total["profit_sales_amount"] - seller_total_public["profit_sales_amount"]
    return {
        "period": selected_period,
        "selected_date": selected_date,
        "today": today_public,
        "yesterday": yesterday_public,
        "comparison": {
            "profit_sales_amount_delta": delta,
            "profit_sales_amount_delta_rate": delta_rate,
        },
        "month": month_total,
        "seller_total": seller_total_public,
        "consistency": {
            "difference": difference,
            "ok": difference == 0,
        },
        "daily_rows": [_sales_report_daily_public(row) for row in daily_rows],
        "seller_top": [_sales_report_named_public(row) for row in seller_rows],
        "product_top": [_sales_report_named_public(row) for row in product_rows],
        "supplier_purchase_totals": [_sales_report_purchase_public(row) for row in supplier_purchase_rows],
        "reviews": [
            {
                "kind": str(row["kind"] or ""),
                "name": str(row["name"] or ""),
                "cs_amount": int(row["cs_amount"] or 0),
                "cs_margin": int(row["cs_margin"] or 0),
                "profit_sales_amount": int(row["profit_sales_amount"] or 0),
                "profit_margin": int(row["profit_margin"] or 0),
            }
            for row in review_rows
        ],
    }


def save_sales_report_file(source_path: str | Path, original_name: str, uploaded_by: str = "") -> dict[str, str | int]:
    source = Path(source_path)
    if not source.is_file():
        raise FileNotFoundError("매출표 파일을 찾지 못했습니다.")

    init_db()
    SALES_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    safe_original = _safe_shared_filename(original_name or source.name)
    file_id = secrets.token_hex(10)
    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file_id}_{safe_original}"
    target = SALES_REPORT_DIR / stored_name
    target.write_bytes(source.read_bytes())
    uploaded_at = now_text()
    report_type = detect_sales_report_type(target, safe_original)
    parsed_report: dict[str, object] | None = None
    report_date = ""
    period = ""
    if report_type:
        parsed_report = parse_sales_report_file(target, safe_original)
        report_date = str(parsed_report.get("report_date") or "")
        period = str(parsed_report.get("period") or report_date[:7] or "")

    connection = connect_db()
    try:
        cursor = connection.execute(
            """
            INSERT INTO sales_report_uploads (
                stored_name, original_name, size, uploaded_by, uploaded_at, report_type, report_date, period
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                stored_name,
                safe_original,
                target.stat().st_size,
                str(uploaded_by or ""),
                uploaded_at,
                report_type,
                report_date,
                period,
            ),
        )
        connection.commit()
        row_id = int(cursor.lastrowid)
    finally:
        connection.close()
    if parsed_report:
        save_sales_report_snapshot(row_id, parsed_report)
    return {
        "id": row_id,
        "original_name": safe_original,
        "size": target.stat().st_size,
        "uploaded_by": str(uploaded_by or ""),
        "uploaded_at": uploaded_at,
        "report_type": report_type,
        "report_date": report_date,
        "period": period,
    }


def crm_webhook_token() -> str:
    env_token = os.environ.get("WORKHUB_CRM_WEBHOOK_TOKEN", "").strip()
    return env_token or ensure_webhook_token(CRM_WEBHOOK_TOKEN_PATH)


def rotate_crm_webhook_token() -> str:
    if os.environ.get("WORKHUB_CRM_WEBHOOK_TOKEN", "").strip():
        raise ValueError("환경변수 WORKHUB_CRM_WEBHOOK_TOKEN 사용 중에는 화면에서 토큰을 재발급할 수 없습니다.")
    return rotate_webhook_token(CRM_WEBHOOK_TOKEN_PATH)


def crm_public_base_url(handler: BaseHTTPRequestHandler | None = None) -> str:
    configured = os.environ.get("WORKHUB_PUBLIC_BASE_URL", "").strip().rstrip("/")
    if configured:
        return configured
    if not handler:
        return ""
    host = handler.headers.get("X-Forwarded-Host") or handler.headers.get("Host", "")
    if not host:
        return ""
    proto = handler.headers.get("X-Forwarded-Proto")
    if not proto:
        proto = "https" if isinstance(handler.request, ssl.SSLSocket) else "http"
    return f"{proto}://{host}".rstrip("/")


def crm_webhook_public_url(handler: BaseHTTPRequestHandler | None = None) -> str:
    base_url = crm_public_base_url(handler)
    return f"{base_url}/api/crm-messenger-webhook" if base_url else "/api/crm-messenger-webhook"


def clean_payload_text(payload: dict, key: str) -> str:
    return str(payload.get(key, "") or "").strip()


def default_permissions_for_role(role: str) -> list[str]:
    return list(DEFAULT_ROLE_PERMISSIONS.get(role, DEFAULT_ROLE_PERMISSIONS["user"]))


def normalize_permissions(value: object, role: str = "user") -> list[str]:
    if role == "admin":
        return list(ALL_PERMISSIONS)
    if isinstance(value, str):
        try:
            value = json.loads(value) if value else []
        except json.JSONDecodeError:
            value = []
    if not isinstance(value, list):
        value = default_permissions_for_role(role)
    allowed = {str(item) for item in value if str(item) in ALL_PERMISSIONS}
    return [key for key in ALL_PERMISSIONS if key in allowed]


def permissions_html() -> str:
    return "\n".join(
        f"""<label class="permission-item"><input type="checkbox" value="{key}" data-permission-check /> <span>{label}<small>{description}</small></span></label>"""
        for key, label, description in PERMISSION_DEFINITIONS
    )


def role_label(role: str) -> str:
    if role == "admin":
        return "관리자"
    if role == "sub_admin":
        return "부관리자"
    return "사용자"


def security_time_text(timestamp: float | int | None) -> str:
    if not timestamp:
        return ""
    return datetime.fromtimestamp(float(timestamp)).strftime("%Y-%m-%d %H:%M:%S")


def normalize_username(value: object) -> str:
    return str(value or "").strip()


def validate_username(username: str) -> None:
    if not re.fullmatch(r"[A-Za-z0-9_.-]{3,32}", username):
        raise ValueError("아이디는 영문/숫자/._- 조합 3~32자로 입력해주세요.")


def validate_password_policy(password: str, username: str, display_name: str = "") -> None:
    if len(password) < PASSWORD_MIN_LENGTH:
        raise ValueError(f"비밀번호는 {PASSWORD_MIN_LENGTH}자 이상으로 입력해주세요.")
    if len(password) > PASSWORD_MAX_LENGTH:
        raise ValueError(f"비밀번호는 {PASSWORD_MAX_LENGTH}자 이내로 입력해주세요.")
    lowered = password.lower()
    blocked = {
        "admin1234",
        "user1234",
        "password",
        "password123",
        "qwer1234",
        "qwerasdf",
        "1234567890",
        "soillbridge",
        "workhub1234",
    }
    if lowered in blocked:
        raise ValueError("초기/추측 쉬운 비밀번호는 사용할 수 없습니다.")
    for value in (username, display_name):
        compact = str(value or "").strip().lower().replace(" ", "")
        if compact and len(compact) >= 3 and compact in lowered.replace(" ", ""):
            raise ValueError("비밀번호에 아이디나 표시 이름을 그대로 넣을 수 없습니다.")
    if len(set(password)) < 4:
        raise ValueError("반복 문자만 있는 비밀번호는 사용할 수 없습니다.")


def login_attempt_key(username: str, ip_address: str) -> str:
    return f"{normalize_username(username).lower()}|{ip_address or 'local'}"


def render_app_html(user: dict[str, str]) -> str:
    display_name = user.get("display_name") or user.get("username") or "사용자"
    role = role_label(user.get("role", "user"))
    display = display_name if display_name == role else f"{display_name} · {role}"
    permissions = normalize_permissions(user.get("permissions", []), user.get("role", "user"))
    is_admin = user.get("role") == "admin"
    leave_enabled = any(permission in permissions for permission in ("leave_view", "leave_approve", "leave_manage"))
    leave_title = "연차 관리 및 신청" if any(permission in permissions for permission in ("leave_approve", "leave_manage")) else "연차 신청 및 확인"
    leave_nav = LEAVE_NAV_HTML.replace("__LEAVE_TITLE__", leave_title) if leave_enabled else ""
    leave_workspace = LEAVE_WORKSPACE_HTML.replace("__LEAVE_TITLE__", leave_title) if leave_enabled else ""
    sales_report_nav = SALES_REPORT_NAV_HTML if is_admin and "sales_report_manage" in permissions else ""
    admin_tools_nav = ADMIN_TOOLS_NAV_HTML if is_admin else ""
    admin_workspace = ADMIN_WORKSPACE_HTML.replace("__PERMISSION_CHECKBOXES__", permissions_html()) if is_admin else ""
    backup_workspace = BACKUP_WORKSPACE_HTML if is_admin else ""
    system_workspace = SYSTEM_WORKSPACE_HTML if is_admin else ""
    return (
        HTML
        .replace("__USER_DISPLAY__", html_escape(display))
        .replace("__CURRENT_USER__", json.dumps({
            "id": int(user.get("id") or 0),
            "username": user.get("username", ""),
            "display_name": display_name,
            "role": user.get("role", "user"),
        }, ensure_ascii=False))
        .replace("__LEAVE_NAV__", leave_nav)
        .replace("__SALES_REPORT_NAV__", sales_report_nav)
        .replace("__LEAVE_WORKSPACE__", leave_workspace)
        .replace("__ADMIN_TOOLS_NAV__", admin_tools_nav)
        .replace("__ADMIN_WORKSPACE__", admin_workspace)
        .replace("__BACKUP_WORKSPACE__", backup_workspace)
        .replace("__SYSTEM_WORKSPACE__", system_workspace)
        .replace("__USER_PERMISSIONS__", json.dumps(permissions, ensure_ascii=False))
        .replace("__PERMISSION_LABELS__", json.dumps({key: label for key, label, _ in PERMISSION_DEFINITIONS}, ensure_ascii=False))
    )


def render_login_html(show_error: bool = False, login_error: str = "", message: str = "", register_error: str = "") -> str:
    return (
        LOGIN_HTML
        .replace("__LOGIN_ERROR_DISPLAY__", "block" if show_error or login_error else "none")
        .replace("__LOGIN_ERROR__", html_escape(login_error or "아이디 또는 비밀번호가 올바르지 않습니다."))
        .replace("__LOGIN_MESSAGE_DISPLAY__", "block" if message else "none")
        .replace("__LOGIN_MESSAGE__", html_escape(message))
        .replace("__REGISTER_ERROR_DISPLAY__", "block" if register_error else "none")
        .replace("__REGISTER_ERROR__", html_escape(register_error))
    )


def password_hash(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    iterations = 260000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
    return f"pbkdf2_sha256${iterations}${salt}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations_text, salt, expected = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        iterations = int(iterations_text)
        digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
        return hmac.compare_digest(digest.hex(), expected)
    except ValueError:
        return False


def token_digest(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def parse_iso_date(value: object) -> date:
    try:
        return date.fromisoformat(str(value or "").strip())
    except ValueError as exc:
        raise ValueError("날짜 형식이 올바르지 않습니다.") from exc


def is_half_step(value: float) -> bool:
    return abs(value * 2 - round(value * 2)) < 0.000001


def clean_leave_days(value: object, label: str) -> float:
    try:
        days = float(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"{label}는 숫자로 입력해주세요.") from exc
    if days < 0 or not is_half_step(days):
        raise ValueError(f"{label}는 0 이상, 0.5일 단위로 입력해주세요.")
    return round(days, 2)


def company_holiday_dates(start: date, end: date) -> set[str]:
    connection = connect_db()
    try:
        rows = connection.execute(
            "SELECT holiday_date FROM company_holidays WHERE holiday_date BETWEEN ? AND ?",
            (start.isoformat(), end.isoformat()),
        ).fetchall()
        return {str(row["holiday_date"]) for row in rows}
    except sqlite3.OperationalError:
        return set()
    finally:
        connection.close()


def save_company_holiday(holiday_date: str, name: str, is_substitute: bool = False) -> None:
    init_db()
    parsed_date = parse_iso_date(holiday_date)
    clean_name = str(name or "").strip() or "\uD734\uC77C"
    now = now_text()
    connection = connect_db()
    try:
        connection.execute(
            """
            INSERT INTO company_holidays (holiday_date, name, is_substitute, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(holiday_date) DO UPDATE SET
                name = excluded.name,
                is_substitute = excluded.is_substitute,
                updated_at = excluded.updated_at
            """,
            (parsed_date.isoformat(), clean_name, 1 if is_substitute else 0, now, now),
        )
        connection.commit()
    finally:
        connection.close()


def calculate_leave_days(start: date, end: date, unit: str, excluded_dates: set[str] | None = None) -> float:
    if start > end:
        raise ValueError("\uC2DC\uC791\uC77C\uC740 \uC885\uB8CC\uC77C\uBCF4\uB2E4 \uB2A6\uC744 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
    excluded_dates = excluded_dates if excluded_dates is not None else company_holiday_dates(start, end)
    if unit == "HALF_DAY":
        if start != end:
            raise ValueError("\uBC18\uCC28\uB294 \uD558\uB8E8\uB9CC \uC2E0\uCCAD\uD560 \uC218 \uC788\uC2B5\uB2C8\uB2E4.")
        if start.weekday() >= 5 or start.isoformat() in excluded_dates:
            raise ValueError("\uC8FC\uB9D0 \uB610\uB294 \uACF5\uD734\uC77C\uC5D0\uB294 \uBC18\uCC28\uB97C \uC2E0\uCCAD\uD560 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
        return 0.5
    days = 0
    cursor = start
    while cursor <= end:
        if cursor.weekday() < 5 and cursor.isoformat() not in excluded_dates:
            days += 1
        cursor = date.fromordinal(cursor.toordinal() + 1)
    if days <= 0:
        raise ValueError("\uC2E0\uCCAD \uAC00\uB2A5\uD55C \uADFC\uBB34\uC77C\uC774 \uC5C6\uC2B5\uB2C8\uB2E4.")
    return float(days)


def leave_status_label(status: str) -> str:
    return {
        "PENDING": "승인대기",
        "APPROVED": "승인완료",
        "REJECTED": "반려",
        "CANCELED": "취소",
        "HISTORICAL": "기존사용",
    }.get(status, status)


def connect_db() -> sqlite3.Connection:
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def normalize_vendor_type(value: object) -> str:
    text = re.sub(r"[\s/_-]+", "", str(value or "").strip().lower())
    if text in {"매출", "매출처", "매출거래처", "판매처", "sales", "seller", "customer"}:
        return "sales"
    return "purchase"


def vendor_type_label(vendor_type: object) -> str:
    return "매출처" if normalize_vendor_type(vendor_type) == "sales" else "매입처"


def normalize_company_key(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\(주\)|㈜|주식회사|[\s\(\)\[\]{}./_-]+", "", text)
    return text


def init_db() -> None:
    connection = connect_db()
    try:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                display_name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                permissions TEXT,
                password_hash TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        user_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(users)").fetchall()
        }
        if "permissions" not in user_columns:
            connection.execute("ALTER TABLE users ADD COLUMN permissions TEXT")
        user_extra_columns = {
            "last_login_at": "TEXT",
            "password_changed_at": "TEXT",
            "created_by": "INTEGER",
            "approved_by": "INTEGER",
            "approved_at": "TEXT",
        }
        for column, column_type in user_extra_columns.items():
            if column not in user_columns:
                connection.execute(f"ALTER TABLE users ADD COLUMN {column} {column_type}")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS shared_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stored_name TEXT NOT NULL UNIQUE,
                original_name TEXT NOT NULL,
                size INTEGER NOT NULL DEFAULT 0,
                uploaded_by TEXT,
                uploaded_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_report_uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stored_name TEXT NOT NULL UNIQUE,
                original_name TEXT NOT NULL,
                size INTEGER NOT NULL DEFAULT 0,
                uploaded_by TEXT,
                uploaded_at TEXT NOT NULL
            )
            """
        )
        sales_upload_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(sales_report_uploads)").fetchall()
        }
        for column, column_type in {
            "report_type": "TEXT",
            "report_date": "TEXT",
            "period": "TEXT",
        }.items():
            if column not in sales_upload_columns:
                connection.execute(f"ALTER TABLE sales_report_uploads ADD COLUMN {column} {column_type}")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_report_daily_rows (
                report_date TEXT PRIMARY KEY,
                period TEXT NOT NULL,
                file_id INTEGER NOT NULL,
                label TEXT,
                quantity INTEGER NOT NULL DEFAULT 0,
                sales_amount INTEGER NOT NULL DEFAULT 0,
                supply_amount INTEGER NOT NULL DEFAULT 0,
                sales_total INTEGER NOT NULL DEFAULT 0,
                supply_total INTEGER NOT NULL DEFAULT 0,
                sales_margin INTEGER NOT NULL DEFAULT 0,
                cs_margin INTEGER NOT NULL DEFAULT 0,
                profit_quantity_sales INTEGER NOT NULL DEFAULT 0,
                profit_quantity_supply INTEGER NOT NULL DEFAULT 0,
                profit_sales_amount INTEGER NOT NULL DEFAULT 0,
                profit_supply_amount INTEGER NOT NULL DEFAULT 0,
                profit_sales_margin INTEGER NOT NULL DEFAULT 0,
                profit_shipping_sales INTEGER NOT NULL DEFAULT 0,
                profit_shipping_supply INTEGER NOT NULL DEFAULT 0,
                profit_shipping INTEGER NOT NULL DEFAULT 0,
                profit_margin INTEGER NOT NULL DEFAULT 0,
                margin_rate REAL NOT NULL DEFAULT 0
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_report_seller_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period TEXT NOT NULL,
                report_date TEXT,
                file_id INTEGER NOT NULL,
                seller_name TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0,
                sales_amount INTEGER NOT NULL DEFAULT 0,
                supply_amount INTEGER NOT NULL DEFAULT 0,
                sales_total INTEGER NOT NULL DEFAULT 0,
                supply_total INTEGER NOT NULL DEFAULT 0,
                sales_margin INTEGER NOT NULL DEFAULT 0,
                cs_amount INTEGER NOT NULL DEFAULT 0,
                cs_supply_amount INTEGER NOT NULL DEFAULT 0,
                cs_margin INTEGER NOT NULL DEFAULT 0,
                profit_quantity_sales INTEGER NOT NULL DEFAULT 0,
                profit_quantity_supply INTEGER NOT NULL DEFAULT 0,
                profit_sales_amount INTEGER NOT NULL DEFAULT 0,
                profit_supply_amount INTEGER NOT NULL DEFAULT 0,
                profit_sales_margin INTEGER NOT NULL DEFAULT 0,
                profit_shipping INTEGER NOT NULL DEFAULT 0,
                profit_margin INTEGER NOT NULL DEFAULT 0,
                margin_rate REAL NOT NULL DEFAULT 0
            )
            """
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_sales_report_seller_period ON sales_report_seller_rows(period)")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_report_supplier_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period TEXT NOT NULL,
                report_date TEXT,
                file_id INTEGER NOT NULL,
                supplier_name TEXT,
                quantity INTEGER DEFAULT 0,
                sales_amount INTEGER DEFAULT 0,
                supply_amount INTEGER DEFAULT 0,
                sales_total INTEGER DEFAULT 0,
                supply_total INTEGER DEFAULT 0,
                sales_margin INTEGER DEFAULT 0,
                cs_amount INTEGER DEFAULT 0,
                cs_supply_amount INTEGER DEFAULT 0,
                cs_margin INTEGER DEFAULT 0,
                profit_quantity_sales INTEGER DEFAULT 0,
                profit_quantity_supply INTEGER DEFAULT 0,
                profit_sales_amount INTEGER DEFAULT 0,
                profit_supply_amount INTEGER DEFAULT 0,
                profit_sales_margin INTEGER DEFAULT 0,
                profit_shipping INTEGER DEFAULT 0,
                profit_margin INTEGER DEFAULT 0,
                margin_rate REAL DEFAULT 0
            )
            """
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_sales_report_supplier_period ON sales_report_supplier_rows(period)")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_report_product_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period TEXT,
                report_date TEXT,
                file_id INTEGER NOT NULL,
                product_code TEXT,
                product_name TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0,
                sales_amount INTEGER NOT NULL DEFAULT 0,
                supply_amount INTEGER NOT NULL DEFAULT 0,
                sales_margin INTEGER NOT NULL DEFAULT 0,
                cs_amount INTEGER NOT NULL DEFAULT 0,
                cs_supply_amount INTEGER NOT NULL DEFAULT 0,
                cs_margin INTEGER NOT NULL DEFAULT 0,
                profit_quantity_sales INTEGER NOT NULL DEFAULT 0,
                profit_quantity_supply INTEGER NOT NULL DEFAULT 0,
                profit_sales_amount INTEGER NOT NULL DEFAULT 0,
                profit_supply_amount INTEGER NOT NULL DEFAULT 0,
                profit_margin INTEGER NOT NULL DEFAULT 0,
                margin_rate REAL NOT NULL DEFAULT 0
            )
            """
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_sales_report_product_date ON sales_report_product_rows(report_date)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_sales_report_product_period ON sales_report_product_rows(period)")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS login_sessions (
                token_hash TEXT PRIMARY KEY,
                username TEXT NOT NULL,
                created_at REAL NOT NULL,
                expires_at REAL NOT NULL,
                last_seen_at REAL,
                absolute_expires_at REAL
            )
            """
        )
        session_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(login_sessions)").fetchall()
        }
        if "last_seen_at" not in session_columns:
            connection.execute("ALTER TABLE login_sessions ADD COLUMN last_seen_at REAL")
        if "absolute_expires_at" not in session_columns:
            connection.execute("ALTER TABLE login_sessions ADD COLUMN absolute_expires_at REAL")
        connection.execute(
            """
            UPDATE login_sessions
               SET last_seen_at = COALESCE(last_seen_at, created_at),
                   absolute_expires_at = COALESCE(absolute_expires_at, created_at + ?)
             WHERE last_seen_at IS NULL OR absolute_expires_at IS NULL
            """,
            (SESSION_SECONDS,),
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS login_attempts (
                identifier TEXT PRIMARY KEY,
                failed_count INTEGER NOT NULL DEFAULT 0,
                first_failed_at REAL NOT NULL,
                last_failed_at REAL NOT NULL,
                locked_until REAL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS internal_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                recipient_user_id INTEGER,
                room_type TEXT NOT NULL DEFAULT 'global',
                body TEXT NOT NULL,
                created_at TEXT NOT NULL,
                task_id INTEGER,
                command_result TEXT,
                command_error TEXT
            )
            """
        )
        internal_message_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(internal_messages)").fetchall()
        }
        if "recipient_user_id" not in internal_message_columns:
            connection.execute("ALTER TABLE internal_messages ADD COLUMN recipient_user_id INTEGER")
        if "room_type" not in internal_message_columns:
            connection.execute("ALTER TABLE internal_messages ADD COLUMN room_type TEXT NOT NULL DEFAULT 'global'")
        if "task_id" not in internal_message_columns:
            connection.execute("ALTER TABLE internal_messages ADD COLUMN task_id INTEGER")
        if "command_result" not in internal_message_columns:
            connection.execute("ALTER TABLE internal_messages ADD COLUMN command_result TEXT")
        if "command_error" not in internal_message_columns:
            connection.execute("ALTER TABLE internal_messages ADD COLUMN command_error TEXT")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_internal_messages_created ON internal_messages(created_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_internal_messages_room ON internal_messages(room_type, recipient_user_id, user_id)")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS system_update_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                before_commit TEXT,
                after_commit TEXT,
                status TEXT NOT NULL,
                message TEXT,
                backup_name TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        now = now_text()
        for username, display_name, role, default_password in DEFAULT_USERS:
            exists = connection.execute("SELECT id, permissions FROM users WHERE username = ?", (username,)).fetchone()
            if not exists:
                connection.execute(
                    """
                    INSERT INTO users (username, display_name, role, permissions, password_hash, active, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        username,
                        display_name,
                        role,
                        json.dumps(default_permissions_for_role(role), ensure_ascii=False),
                        password_hash(default_password),
                        now,
                        now,
                    ),
                )
            elif not exists["permissions"]:
                connection.execute(
                    "UPDATE users SET permissions = ?, updated_at = ? WHERE username = ?",
                    (json.dumps(default_permissions_for_role(role), ensure_ascii=False), now, username),
                )
        for row in connection.execute("SELECT username, role FROM users WHERE permissions IS NULL OR permissions = ''").fetchall():
            connection.execute(
                "UPDATE users SET permissions = ?, updated_at = ? WHERE username = ?",
                (
                    json.dumps(default_permissions_for_role(row["role"]), ensure_ascii=False),
                    now,
                    row["username"],
                ),
            )
        for row in connection.execute("SELECT username, role, permissions FROM users WHERE role != 'admin'").fetchall():
            permissions = normalize_permissions(row["permissions"], row["role"])
            if "crm_view" not in permissions:
                permissions.append("crm_view")
                ordered_permissions = [key for key in ALL_PERMISSIONS if key in set(permissions)]
                connection.execute(
                    "UPDATE users SET permissions = ?, updated_at = ? WHERE username = ?",
                    (json.dumps(ordered_permissions, ensure_ascii=False), now, row["username"]),
                )
        connection.execute("DELETE FROM login_sessions WHERE expires_at < ? OR COALESCE(absolute_expires_at, expires_at) < ?", (time.time(), time.time()))
        connection.execute("DELETE FROM login_attempts WHERE last_failed_at < ? AND COALESCE(locked_until, 0) < ?", (time.time() - 24 * 60 * 60, time.time()))
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS cs_cases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '접수',
                vendor_name TEXT,
                vendor_email TEXT,
                original_info TEXT,
                original_invoice TEXT,
                product_name TEXT,
                orderer_name TEXT,
                orderer_phone TEXT,
                receiver_name TEXT,
                receiver_phone TEXT,
                receiver_address TEXT,
                cs_type TEXT,
                cs_content TEXT,
                return_invoice TEXT,
                reship_invoice TEXT,
                mail_subject TEXT,
                mail_body TEXT,
                mail_sent_at TEXT,
                source_file TEXT,
                source_sheet TEXT,
                source_row INTEGER,
                occurred_at TEXT,
                completed_at TEXT,
                order_date TEXT,
                ship_date TEXT,
                sales_vendor TEXT,
                purchase_vendor TEXT,
                courier TEXT,
                quantity TEXT
            )
            """
        )
        existing_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(cs_cases)").fetchall()
        }
        extra_columns = {
            "source_file": "TEXT",
            "source_sheet": "TEXT",
            "source_row": "INTEGER",
            "occurred_at": "TEXT",
            "completed_at": "TEXT",
            "order_date": "TEXT",
            "ship_date": "TEXT",
            "orderer_name": "TEXT",
            "orderer_phone": "TEXT",
            "sales_vendor": "TEXT",
            "purchase_vendor": "TEXT",
            "courier": "TEXT",
            "quantity": "TEXT",
            "cs_type": "TEXT",
        }
        for column, column_type in extra_columns.items():
            if column not in existing_columns:
                connection.execute(f"ALTER TABLE cs_cases ADD COLUMN {column} {column_type}")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_cs_cases_status ON cs_cases(status)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_cs_cases_original_invoice ON cs_cases(original_invoice)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_cs_cases_receiver_phone ON cs_cases(receiver_phone)")
        connection.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_cs_cases_source ON cs_cases(source_file, source_sheet, source_row)")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS vendor_contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vendor_type TEXT NOT NULL DEFAULT 'purchase',
                vendor_name TEXT NOT NULL,
                email TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(vendor_type, vendor_name)
            )
            """
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_vendor_contacts_name ON vendor_contacts(vendor_name)")
        if VENDOR_CONTACTS_PATH.exists():
            existing_count = connection.execute("SELECT COUNT(*) AS count FROM vendor_contacts").fetchone()["count"]
            if int(existing_count or 0) == 0:
                try:
                    raw_contacts = json.loads(VENDOR_CONTACTS_PATH.read_text(encoding="utf-8"))
                except (OSError, json.JSONDecodeError):
                    raw_contacts = []
                if isinstance(raw_contacts, list):
                    for raw in raw_contacts:
                        if not isinstance(raw, dict):
                            continue
                        vendor_name = str(raw.get("vendor_name", "")).strip()
                        email = str(raw.get("email", "")).strip()
                        vendor_type = normalize_vendor_type(str(raw.get("vendor_type", "") or "purchase"))
                        if vendor_name and email and "@" in email:
                            connection.execute(
                                """
                                INSERT OR IGNORE INTO vendor_contacts
                                    (vendor_type, vendor_name, email, created_at, updated_at)
                                VALUES (?, ?, ?, ?, ?)
                                """,
                                (vendor_type, vendor_name, email, now, now),
                            )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS management_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                source_file TEXT NOT NULL,
                source_sheet TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                purchase_vendor TEXT,
                sales_vendor TEXT,
                transaction_type TEXT,
                ledger_checked TEXT,
                order_date TEXT,
                ship_date TEXT,
                orderer_name TEXT,
                sender_phone TEXT,
                receiver_name TEXT,
                receiver_phone TEXT,
                product_name TEXT,
                quantity TEXT,
                receiver_address TEXT,
                courier TEXT,
                invoice_number TEXT,
                memo TEXT,
                order_item_id TEXT,
                product_code TEXT,
                order_number TEXT,
                customer_option TEXT,
                cs_received_at TEXT
            )
            """
        )
        management_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(management_records)").fetchall()
        }
        management_extra_columns = {
            "order_item_id": "TEXT",
            "product_code": "TEXT",
            "order_number": "TEXT",
            "customer_option": "TEXT",
            "cs_received_at": "TEXT",
        }
        for column, column_type in management_extra_columns.items():
            if column not in management_columns:
                connection.execute(f"ALTER TABLE management_records ADD COLUMN {column} {column_type}")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_management_invoice ON management_records(invoice_number)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_management_receiver_phone ON management_records(receiver_phone)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_management_order_date ON management_records(order_date)")
        connection.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_management_source ON management_records(source_file, source_sheet, source_row)")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS import_shipments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                departure_date TEXT,
                arrival_date TEXT,
                loading_port TEXT,
                arrival_port TEXT,
                shipper TEXT,
                item TEXT,
                quantity TEXT,
                vessel_name TEXT,
                hbl_no TEXT,
                size TEXT,
                progress_status TEXT,
                free_time TEXT,
                warehouse_due_date TEXT,
                completed_at TEXT
            )
            """
        )
        import_shipment_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(import_shipments)").fetchall()
        }
        if "quantity" not in import_shipment_columns:
            connection.execute("ALTER TABLE import_shipments ADD COLUMN quantity TEXT")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS leave_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                is_paid INTEGER NOT NULL DEFAULT 1,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS leave_balances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                leave_type_id INTEGER NOT NULL,
                total_days REAL NOT NULL DEFAULT 0,
                used_days REAL NOT NULL DEFAULT 0,
                remaining_days REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(user_id, leave_type_id)
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS leave_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                leave_type_id INTEGER NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                unit TEXT NOT NULL,
                requested_days REAL NOT NULL,
                reason TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'PENDING',
                rejected_reason TEXT,
                decided_by INTEGER,
                finalized_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS leave_balance_ledger (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                balance_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                actor_id INTEGER,
                delta_days REAL NOT NULL,
                reason TEXT NOT NULL,
                request_id INTEGER,
                created_at TEXT NOT NULL
            )
            """
        )
        leave_balance_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(leave_balances)").fetchall()
        }
        if "reserved_days" not in leave_balance_columns:
            connection.execute("ALTER TABLE leave_balances ADD COLUMN reserved_days REAL NOT NULL DEFAULT 0")
        if "accrual_year" not in leave_balance_columns:
            connection.execute("ALTER TABLE leave_balances ADD COLUMN accrual_year INTEGER")
        leave_request_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(leave_requests)").fetchall()
        }
        leave_request_extra_columns = {
            "approval_step": "TEXT NOT NULL DEFAULT 'TEAM_LEAD'",
            "team_status": "TEXT NOT NULL DEFAULT 'PENDING'",
            "team_decided_by": "INTEGER",
            "team_decided_at": "TEXT",
            "team_comment": "TEXT",
            "director_status": "TEXT NOT NULL DEFAULT 'WAITING'",
            "director_decided_by": "INTEGER",
            "director_decided_at": "TEXT",
            "director_comment": "TEXT",
            "ceo_status": "TEXT NOT NULL DEFAULT 'WAITING'",
            "ceo_decided_by": "INTEGER",
            "ceo_decided_at": "TEXT",
            "ceo_comment": "TEXT",
            "cancel_reason": "TEXT",
            "canceled_at": "TEXT",
            "canceled_by": "INTEGER",
        }
        for column, column_type in leave_request_extra_columns.items():
            if column not in leave_request_columns:
                connection.execute(f"ALTER TABLE leave_requests ADD COLUMN {column} {column_type}")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS company_holidays (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                holiday_date TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                is_substitute INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS leave_notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                request_id INTEGER,
                notification_type TEXT NOT NULL,
                message TEXT NOT NULL,
                is_read INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_company_holidays_date ON company_holidays(holiday_date)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_leave_notifications_user ON leave_notifications(user_id, is_read)")
        timestamp = now_text()
        for code, name in (("annual", "연차"), ("special", "특별휴가"), ("unpaid", "무급휴가"), ("sick", "병가")):
            connection.execute(
                """
                INSERT OR IGNORE INTO leave_types (code, name, is_paid, is_active, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?)
                """,
                (code, name, 0 if code == "unpaid" else 1, timestamp, timestamp),
            )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_leave_requests_user ON leave_requests(user_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_leave_requests_status ON leave_requests(status)")
        init_crm_db(connection)
        connection.commit()
    finally:
        connection.close()


def login_lock_status(username: str, ip_address: str) -> tuple[bool, str]:
    init_db()
    key = login_attempt_key(username, ip_address)
    now = time.time()
    connection = connect_db()
    try:
        row = connection.execute(
            "SELECT locked_until FROM login_attempts WHERE identifier = ?",
            (key,),
        ).fetchone()
        if not row or not row["locked_until"] or float(row["locked_until"]) <= now:
            return False, ""
        minutes = max(1, int((float(row["locked_until"]) - now + 59) // 60))
        return True, f"로그인 시도가 잠시 제한됐습니다. 약 {minutes}분 후 다시 시도해주세요."
    finally:
        connection.close()


def record_login_failure(username: str, ip_address: str) -> None:
    init_db()
    key = login_attempt_key(username, ip_address)
    now = time.time()
    connection = connect_db()
    try:
        row = connection.execute(
            "SELECT failed_count, first_failed_at FROM login_attempts WHERE identifier = ?",
            (key,),
        ).fetchone()
        if not row or now - float(row["first_failed_at"]) > LOGIN_FAILURE_WINDOW_SECONDS:
            failed_count = 1
            first_failed_at = now
        else:
            failed_count = int(row["failed_count"] or 0) + 1
            first_failed_at = float(row["first_failed_at"])
        locked_until = now + LOGIN_LOCK_SECONDS if failed_count >= LOGIN_MAX_FAILURES else None
        connection.execute(
            """
            INSERT INTO login_attempts (identifier, failed_count, first_failed_at, last_failed_at, locked_until)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(identifier) DO UPDATE SET
                failed_count = excluded.failed_count,
                first_failed_at = excluded.first_failed_at,
                last_failed_at = excluded.last_failed_at,
                locked_until = excluded.locked_until
            """,
            (key, failed_count, first_failed_at, now, locked_until),
        )
        connection.commit()
    finally:
        connection.close()


def clear_login_failures(username: str, ip_address: str) -> None:
    init_db()
    key = login_attempt_key(username, ip_address)
    connection = connect_db()
    try:
        connection.execute("DELETE FROM login_attempts WHERE identifier = ?", (key,))
        connection.commit()
    finally:
        connection.close()


def authenticate_user(username: str, password: str) -> dict[str, str] | None:
    init_db()
    normalized = normalize_username(username)
    if not normalized or not password:
        return None
    connection = connect_db()
    try:
        row = connection.execute(
            """
            SELECT id, username, display_name, role, permissions, password_hash
              FROM users
             WHERE username = ? AND active = 1
            """,
            (normalized,),
        ).fetchone()
        if not row or not verify_password(password, row["password_hash"]):
            return None
        connection.execute("UPDATE users SET last_login_at = ?, updated_at = ? WHERE id = ?", (now_text(), now_text(), int(row["id"])))
        connection.commit()
        return {
            "id": str(row["id"]),
            "username": row["username"],
            "display_name": row["display_name"],
            "role": row["role"],
            "permissions": normalize_permissions(row["permissions"], row["role"]),
        }
    finally:
        connection.close()


def create_login_session(username: str) -> str:
    init_db()
    token = secrets.token_urlsafe(32)
    now = time.time()
    idle_expires_at = now + SESSION_IDLE_SECONDS
    absolute_expires_at = now + SESSION_SECONDS
    connection = connect_db()
    try:
        connection.execute(
            """
            INSERT INTO login_sessions (token_hash, username, created_at, expires_at, last_seen_at, absolute_expires_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (token_digest(token), username, now, min(idle_expires_at, absolute_expires_at), now, absolute_expires_at),
        )
        connection.commit()
        return token
    finally:
        connection.close()


def current_user_from_token(token: str) -> dict[str, str] | None:
    if not token:
        return None
    init_db()
    connection = connect_db()
    try:
        row = connection.execute(
            """
            SELECT users.id, users.username, users.display_name, users.role, users.permissions,
                   login_sessions.expires_at,
                   COALESCE(login_sessions.last_seen_at, login_sessions.created_at) AS last_seen_at,
                   COALESCE(login_sessions.absolute_expires_at, login_sessions.created_at + ?) AS absolute_expires_at
              FROM login_sessions
              JOIN users ON users.username = login_sessions.username
             WHERE login_sessions.token_hash = ?
               AND users.active = 1
            """,
            (SESSION_SECONDS, token_digest(token)),
        ).fetchone()
        if not row:
            return None
        now = time.time()
        if float(row["expires_at"]) < now or float(row["absolute_expires_at"]) < now:
            connection.execute("DELETE FROM login_sessions WHERE token_hash = ?", (token_digest(token),))
            connection.commit()
            return None
        next_expires_at = min(now + SESSION_IDLE_SECONDS, float(row["absolute_expires_at"]))
        connection.execute(
            "UPDATE login_sessions SET last_seen_at = ?, expires_at = ? WHERE token_hash = ?",
            (now, next_expires_at, token_digest(token)),
        )
        connection.commit()
        return {
            "id": str(row["id"]),
            "username": row["username"],
            "display_name": row["display_name"],
            "role": row["role"],
            "permissions": normalize_permissions(row["permissions"], row["role"]),
        }
    finally:
        connection.close()


def delete_login_session(token: str) -> None:
    if not token:
        return
    init_db()
    connection = connect_db()
    try:
        connection.execute("DELETE FROM login_sessions WHERE token_hash = ?", (token_digest(token),))
        connection.commit()
    finally:
        connection.close()


def list_users() -> list[dict[str, str | int]]:
    init_db()
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT id, username, display_name, role, permissions, active, created_at, updated_at,
                   last_login_at, password_changed_at, approved_at
              FROM users
             ORDER BY CASE role WHEN 'admin' THEN 0 WHEN 'sub_admin' THEN 1 ELSE 2 END,
                      username COLLATE NOCASE
            """
        ).fetchall()
        users = []
        for row in rows:
            item = dict(row)
            item["permissions"] = normalize_permissions(item.get("permissions"), str(item.get("role", "user")))
            users.append(item)
        return users
    finally:
        connection.close()


def company_staff_dashboard_payload(current_user: dict[str, str]) -> dict:
    init_db()
    connection = connect_db()
    try:
        today = date.today().isoformat()
        rows = connection.execute(
            """
            SELECT users.id,
                   users.username,
                   users.display_name,
                   users.role,
                   COUNT(tasks.id) AS task_count,
                   SUM(CASE WHEN tasks.status != '완료' THEN 1 ELSE 0 END) AS open_tasks,
                   SUM(CASE WHEN tasks.status != '완료' AND substr(tasks.due_at, 1, 10) = ? THEN 1 ELSE 0 END) AS due_today,
                   SUM(CASE WHEN tasks.status != '완료'
                              AND length(tasks.due_at) >= 10
                              AND substr(tasks.due_at, 1, 10) < ? THEN 1 ELSE 0 END) AS overdue
              FROM users
              LEFT JOIN crm_tasks tasks ON tasks.assignee_user_id = users.id
             WHERE users.active = 1
             GROUP BY users.id
             ORDER BY CASE users.role WHEN 'admin' THEN 0 WHEN 'sub_admin' THEN 1 ELSE 2 END,
                      users.display_name COLLATE NOCASE,
                      users.username COLLATE NOCASE
            """,
            (today, today),
        ).fetchall()
        staff = []
        for row in rows:
            item = dict(row)
            role = str(item.get("role") or "user")
            item["team_label"] = "운영 리드" if role == "admin" else ("부운영 리드" if role == "sub_admin" else "업무 담당")
            item["task_count"] = int(item.get("task_count") or 0)
            item["open_tasks"] = int(item.get("open_tasks") or 0)
            item["due_today"] = int(item.get("due_today") or 0)
            item["overdue"] = int(item.get("overdue") or 0)
            staff.append(item)
        return {
            "current_user_id": int(current_user.get("id") or 0),
            "staff": staff,
        }
    finally:
        connection.close()


def month_bounds(month_text: str) -> tuple[date, date]:
    today = date.today()
    match = re.match(r"^(\d{4})-(\d{2})$", str(month_text or "").strip())
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if not 1 <= month <= 12:
            raise ValueError("캘린더 월 형식이 올바르지 않습니다.")
        start = date(year, month, 1)
    else:
        start = date(today.year, today.month, 1)
    next_month = date(start.year + (1 if start.month == 12 else 0), 1 if start.month == 12 else start.month + 1, 1)
    return start, next_month - timedelta(days=1)


def company_calendar_payload(user: dict[str, str], month_text: str) -> dict:
    init_db()
    start, end = month_bounds(month_text)
    start_text = start.isoformat()
    end_text = end.isoformat()
    today_text = date.today().isoformat()
    can_see_leave = any(user_has_permission(user, permission) for permission in ("leave_view", "leave_approve", "leave_manage"))
    can_see_team_leave = user_has_permission(user, "leave_approve") or user_has_permission(user, "leave_manage")
    user_id = int(user.get("id") or 0)
    connection = connect_db()
    try:
        task_rows = connection.execute(
            """
            SELECT id, public_id, title, description, account_name, assignee_user_id, assignee_name,
                   requester_name, due_at, priority, status, source, updated_at
              FROM crm_tasks
             WHERE due_at IS NOT NULL
               AND due_at != ''
               AND substr(due_at, 1, 10) BETWEEN ? AND ?
             ORDER BY substr(due_at, 1, 10), priority = '높음' DESC, due_at
            """,
            (start_text, end_text),
        ).fetchall()
        project_rows = connection.execute(
            """
            SELECT COALESCE(NULLIF(account_name, ''), '직원 지시 업무') AS project_name,
                   MIN(substr(due_at, 1, 10)) AS event_date,
                   COUNT(*) AS total_tasks,
                   SUM(CASE WHEN status != '완료' THEN 1 ELSE 0 END) AS open_tasks,
                   SUM(CASE WHEN status = '완료' THEN 1 ELSE 0 END) AS completed_tasks,
                   SUM(CASE WHEN priority = '높음' AND status != '완료' THEN 1 ELSE 0 END) AS high_priority,
                   GROUP_CONCAT(DISTINCT NULLIF(assignee_name, '')) AS assignee_names
              FROM crm_tasks
             WHERE due_at IS NOT NULL
               AND due_at != ''
               AND substr(due_at, 1, 10) BETWEEN ? AND ?
             GROUP BY COALESCE(NULLIF(account_name, ''), '직원 지시 업무')
             ORDER BY event_date, open_tasks DESC
             LIMIT 80
            """,
            (start_text, end_text),
        ).fetchall()
        leave_rows = []
        if can_see_leave:
            leave_conditions = ["leave_requests.start_date <= ?", "leave_requests.end_date >= ?"]
            leave_params: list[object] = [end_text, start_text]
            if can_see_team_leave:
                leave_conditions.append("leave_requests.status IN ('APPROVED', 'PENDING')")
            else:
                leave_conditions.append("(leave_requests.status = 'APPROVED' OR leave_requests.user_id = ?)")
                leave_params.append(user_id)
            leave_rows = connection.execute(
                f"""
                SELECT leave_requests.*, leave_types.name AS leave_type_name, users.display_name
                  FROM leave_requests
                  JOIN leave_types ON leave_types.id = leave_requests.leave_type_id
                  JOIN users ON users.id = leave_requests.user_id
                 WHERE {' AND '.join(leave_conditions)}
                 ORDER BY leave_requests.start_date, users.display_name
                """,
                tuple(leave_params),
            ).fetchall()
    finally:
        connection.close()

    events: list[dict[str, object]] = []
    for row in project_rows:
        total_tasks = int(row["total_tasks"] or 0)
        completed_tasks = int(row["completed_tasks"] or 0)
        progress = round((completed_tasks / total_tasks) * 100) if total_tasks else 0
        events.append({
            "id": f"project:{row['project_name']}",
            "type": "project",
            "date": row["event_date"],
            "title": row["project_name"],
            "subtitle": f"{progress}% 완료 · 진행 {int(row['open_tasks'] or 0)}건",
            "project_name": row["project_name"],
            "assignees": row["assignee_names"] or "",
            "high_priority": int(row["high_priority"] or 0),
            "progress": progress,
        })
    for row in task_rows:
        due_day = str(row["due_at"] or "")[:10]
        risk = row["status"] != "완료" and due_day < today_text
        events.append({
            "id": f"task:{row['id']}",
            "type": "task",
            "date": due_day,
            "title": row["title"] or row["public_id"],
            "subtitle": f"{row['public_id']} · {row['assignee_name'] or '담당자 미정'} · {row['status']}",
            "task_id": row["id"],
            "public_id": row["public_id"],
            "priority": row["priority"],
            "status": row["status"],
            "risk": risk or row["priority"] == "높음",
        })
    for row in leave_rows:
        leave_start = max(parse_iso_date(row["start_date"]), start)
        leave_end = min(parse_iso_date(row["end_date"]), end)
        cursor = leave_start
        while cursor <= leave_end:
            status = row["status"] or "PENDING"
            events.append({
                "id": f"leave:{row['id']}:{cursor.isoformat()}",
                "type": "pending" if status == "PENDING" else "leave",
                "date": cursor.isoformat(),
                "title": f"{row['display_name']} {row['leave_type_name']}",
                "subtitle": f"{leave_status_label(status)} · {row['unit'] == 'HALF_DAY' and '반차' or '연차'}",
                "leave_id": row["id"],
                "user_id": row["user_id"],
                "status": status,
                "requested_days": round(float(row["requested_days"] or 0), 2),
                "reason": row["reason"] or "",
            })
            cursor += timedelta(days=1)
    summary = {
        "project": sum(1 for event in events if event["type"] == "project"),
        "task": sum(1 for event in events if event["type"] == "task"),
        "leave": sum(1 for event in events if event["type"] in {"leave", "pending"}),
        "risk": sum(1 for event in events if event.get("risk") or (event["type"] == "pending")),
    }
    return {
        "month": start.strftime("%Y-%m"),
        "start": start_text,
        "end": end_text,
        "today": today_text,
        "can_see_team_leave": can_see_team_leave,
        "summary": summary,
        "events": sorted(events, key=lambda item: (str(item["date"]), {"project": 0, "leave": 1, "pending": 2, "task": 3}.get(str(item["type"]), 9), str(item["title"]))),
    }


def list_internal_messages(
    limit: int = 100,
    room_type: str = "global",
    current_user_id: int = 0,
    other_user_id: int = 0,
) -> list[dict[str, str | int]]:
    init_db()
    safe_limit = max(1, min(int(limit), 300))
    room = "dm" if room_type == "dm" else "global"
    connection = connect_db()
    try:
        conditions = ["COALESCE(internal_messages.room_type, 'global') = ?"]
        params: list[object] = [room]
        if room == "dm":
            if not current_user_id or not other_user_id:
                return []
            conditions.append(
                "((internal_messages.user_id = ? AND internal_messages.recipient_user_id = ?) "
                "OR (internal_messages.user_id = ? AND internal_messages.recipient_user_id = ?))"
            )
            params.extend([current_user_id, other_user_id, other_user_id, current_user_id])
        where_sql = " AND ".join(conditions)
        rows = connection.execute(
            f"""
            SELECT internal_messages.id,
                   internal_messages.user_id,
                   internal_messages.recipient_user_id,
                   COALESCE(internal_messages.room_type, 'global') AS room_type,
                   internal_messages.body,
                   internal_messages.created_at,
                   internal_messages.task_id,
                   internal_messages.command_result,
                   internal_messages.command_error,
                   users.username,
                   users.display_name
              FROM internal_messages
              LEFT JOIN users ON users.id = internal_messages.user_id
             WHERE {where_sql}
             ORDER BY internal_messages.id DESC
             LIMIT ?
            """,
            tuple(params + [safe_limit]),
        ).fetchall()
        return [dict(row) for row in reversed(rows)]
    finally:
        connection.close()


def find_internal_command_assignee(connection: sqlite3.Connection, mention: str) -> dict[str, str | int]:
    mention_text = mention.strip().lstrip("@")
    if not mention_text:
        raise ValueError("@직원을 입력해주세요.")
    rows = connection.execute(
        """
        SELECT id, username, display_name
          FROM users
         WHERE active = 1
           AND (username = ? COLLATE NOCASE OR display_name = ? COLLATE NOCASE)
         ORDER BY id
        """,
        (mention_text, mention_text),
    ).fetchall()
    if not rows:
        raise ValueError(f"직원 '{mention_text}'를 찾지 못했습니다.")
    if len(rows) > 1:
        names = ", ".join(row["display_name"] or row["username"] for row in rows[:5])
        raise ValueError(f"'{mention_text}'에 해당하는 직원이 여러 명입니다: {names}")
    return dict(rows[0])


def parse_internal_task_command(body: str) -> tuple[str, str, str]:
    match = re.match(r"^/업무\s+@([^\s/]+)\s+(.+?)\s*/\s*(.+)$", body.strip())
    if not match:
        raise ValueError("업무 지시 형식: /업무 @직원 업무내용 / 기한")
    assignee_text, title, due_at = [part.strip() for part in match.groups()]
    if not title:
        raise ValueError("업무내용을 입력해주세요.")
    if not due_at:
        raise ValueError("기한을 입력해주세요.")
    return assignee_text, title, due_at


def task_public_id(task_id: int) -> str:
    connection = connect_db()
    try:
        row = connection.execute("SELECT public_id FROM crm_tasks WHERE id = ?", (task_id,)).fetchone()
        return str(row["public_id"] if row else f"TASK-{task_id:04d}")
    finally:
        connection.close()


def save_internal_message(payload: dict, user: dict[str, str]) -> int:
    body = clean_payload_text(payload, "body")
    if not body:
        raise ValueError("메시지 내용을 입력해주세요.")
    if len(body) > 2000:
        raise ValueError("메시지는 2,000자 이하로 입력해주세요.")
    room_type = "dm" if clean_payload_text(payload, "room_type") == "dm" else "global"
    recipient_user_id = int(payload.get("recipient_user_id") or 0) if room_type == "dm" else 0
    init_db()
    if room_type == "dm":
        validation_connection = connect_db()
        try:
            recipient = validation_connection.execute(
                "SELECT id FROM users WHERE id = ? AND active = 1",
                (recipient_user_id,),
            ).fetchone()
            if not recipient:
                raise ValueError("DM 대상 직원을 찾지 못했습니다.")
        finally:
            validation_connection.close()
    task_id: int | None = None
    command_result = ""
    command_error = ""
    if body.startswith("/업무"):
        try:
            assignee_text, title, due_at = parse_internal_task_command(body)
            assignee_connection = connect_db()
            try:
                assignee = find_internal_command_assignee(assignee_connection, assignee_text)
            finally:
                assignee_connection.close()
            task_id = save_crm_task(DB_PATH, {
                "account_id": "",
                "account_name": "",
                "title": title,
                "description": body,
                "assignee_user_id": assignee["id"],
                "assignee_name": assignee["display_name"] or assignee["username"],
                "due_at": due_at,
                "priority": "보통",
                "status": "대기",
                "source": "internal_message",
            }, user)
            public_id = task_public_id(task_id)
            command_result = f"{public_id} 업무가 등록됐습니다."
        except Exception as exc:  # noqa: BLE001
            command_error = str(exc)
    connection = connect_db()
    try:
        cursor = connection.execute(
            """
            INSERT INTO internal_messages
                (user_id, recipient_user_id, room_type, body, created_at, task_id, command_result, command_error)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(user.get("id") or 0),
                recipient_user_id or None,
                room_type,
                body,
                now_text(),
                task_id,
                command_result,
                command_error,
            ),
        )
        connection.commit()
        return int(cursor.lastrowid)
    finally:
        connection.close()


def save_user_account(payload: dict, actor: dict[str, str]) -> int:
    init_db()
    user_id = int(payload.get("id", 0) or 0)
    username = normalize_username(payload.get("username"))
    display_name = clean_payload_text(payload, "display_name")
    role = clean_payload_text(payload, "role") or "user"
    password = str(payload.get("password", "") or "")
    active = 1 if payload.get("active", True) else 0
    permissions = normalize_permissions(payload.get("permissions", []), role)
    if role not in {"admin", "sub_admin", "user"}:
        raise ValueError("권한은 관리자, 부관리자, 사용자 중 하나만 선택할 수 있습니다.")
    validate_username(username)
    if not display_name:
        display_name = username
    if not user_id and not password:
        raise ValueError("신규 사용자는 비밀번호를 입력해주세요.")
    if password:
        validate_password_policy(password, username, display_name)
    now = now_text()
    connection = connect_db()
    try:
        if user_id:
            existing = connection.execute("SELECT username, active FROM users WHERE id = ?", (user_id,)).fetchone()
            if not existing:
                raise ValueError("수정할 사용자를 찾지 못했습니다.")
            if existing["username"] == actor.get("username") and not active:
                raise ValueError("현재 로그인한 관리자 계정은 사용 중지할 수 없습니다.")
            columns = [
                "username = ?",
                "display_name = ?",
                "role = ?",
                "permissions = ?",
                "active = ?",
                "updated_at = ?",
            ]
            values: list[object] = [username, display_name, role, json.dumps(permissions, ensure_ascii=False), active, now]
            if password:
                columns.append("password_hash = ?")
                values.append(password_hash(password))
                columns.append("password_changed_at = ?")
                values.append(now)
            if active and not int(existing["active"] or 0):
                columns.append("approved_by = ?")
                columns.append("approved_at = ?")
                values.extend([int(actor.get("id") or 0) or None, now])
            values.append(user_id)
            connection.execute(f"UPDATE users SET {', '.join(columns)} WHERE id = ?", values)
            saved_id = user_id
        else:
            cursor = connection.execute(
                """
                INSERT INTO users
                    (username, display_name, role, permissions, password_hash, active,
                     created_at, updated_at, password_changed_at, created_by, approved_by, approved_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    username,
                    display_name,
                    role,
                    json.dumps(permissions, ensure_ascii=False),
                    password_hash(password),
                    active,
                    now,
                    now,
                    now,
                    int(actor.get("id") or 0) or None,
                    int(actor.get("id") or 0) if active else None,
                    now if active else None,
                ),
            )
            saved_id = int(cursor.lastrowid)
        connection.commit()
        return saved_id
    except sqlite3.IntegrityError as exc:
        raise ValueError("이미 사용 중인 아이디입니다.") from exc
    finally:
        connection.close()


def register_user_request(payload: dict[str, str]) -> int:
    init_db()
    username = normalize_username(payload.get("username"))
    display_name = clean_payload_text(payload, "display_name")
    password = str(payload.get("password", "") or "")
    password_confirm = str(payload.get("password_confirm", "") or "")
    validate_username(username)
    if not display_name:
        raise ValueError("표시 이름을 입력해주세요.")
    if password != password_confirm:
        raise ValueError("비밀번호 확인이 일치하지 않습니다.")
    validate_password_policy(password, username, display_name)
    now = now_text()
    permissions = default_permissions_for_role("user")
    connection = connect_db()
    try:
        existing = connection.execute("SELECT id, active FROM users WHERE username = ?", (username,)).fetchone()
        if existing:
            if not int(existing["active"] or 0):
                raise ValueError("이미 등록 요청된 아이디입니다. 관리자 승인을 기다려주세요.")
            raise ValueError("이미 사용 중인 아이디입니다.")
        cursor = connection.execute(
            """
            INSERT INTO users
                (username, display_name, role, permissions, password_hash, active,
                 created_at, updated_at, password_changed_at)
            VALUES (?, ?, 'user', ?, ?, 0, ?, ?, ?)
            """,
            (username, display_name, json.dumps(permissions, ensure_ascii=False), password_hash(password), now, now, now),
        )
        connection.commit()
        return int(cursor.lastrowid)
    finally:
        connection.close()


def user_has_permission(user: dict[str, str], permission: str) -> bool:
    return permission in normalize_permissions(user.get("permissions", []), user.get("role", "user"))


RESTORE_ALLOWED_FILES = {
    "config/workhub.db": DB_PATH,
    "config/mail_settings.json": MAIL_SETTINGS_PATH,
    "config/vendor_contacts.json": VENDOR_CONTACTS_PATH,
    "config/backup_settings.json": BACKUP_SETTINGS_PATH,
    "config/secret.key": SECRET_KEY_PATH,
    "config/crm_webhook_token.txt": CRM_WEBHOOK_TOKEN_PATH,
}

DEFAULT_BACKUP_SETTINGS = {
    "backup_dir": "",
    "auto_enabled": True,
    "auto_hour": AUTO_BACKUP_HOUR,
    "retention_days": BACKUP_RETENTION_DAYS,
    "external_enabled": False,
    "external_type": "rclone",
    "rclone_remote": "",
    "rclone_path": "",
    "rclone_executable": "rclone",
    "last_external_status": "",
    "last_external_message": "",
    "last_external_uploaded_at": "",
    "last_external_backup_name": "",
    "last_external_target": "",
}


def backup_default_dir() -> Path:
    return BACKUP_DIR


def clean_backup_dir(value: object) -> str:
    text = str(value or "").strip().strip('"')
    if not text:
        return ""
    path = Path(text).expanduser()
    if not path.is_absolute():
        path = RUNTIME_ROOT / path
    return str(path)


def normalize_backup_settings(payload: dict | None = None) -> dict[str, object]:
    source = payload or {}
    try:
        auto_hour = int(source.get("auto_hour", DEFAULT_BACKUP_SETTINGS["auto_hour"]))
    except (TypeError, ValueError):
        auto_hour = int(DEFAULT_BACKUP_SETTINGS["auto_hour"])
    auto_hour = min(max(auto_hour, 0), 23)
    try:
        retention_days = int(source.get("retention_days", DEFAULT_BACKUP_SETTINGS["retention_days"]))
    except (TypeError, ValueError):
        retention_days = int(DEFAULT_BACKUP_SETTINGS["retention_days"])
    retention_days = min(max(retention_days, 1), 3650)
    return {
        "backup_dir": clean_backup_dir(source.get("backup_dir", DEFAULT_BACKUP_SETTINGS["backup_dir"])),
        "auto_enabled": bool(source.get("auto_enabled", DEFAULT_BACKUP_SETTINGS["auto_enabled"])),
        "auto_hour": auto_hour,
        "retention_days": retention_days,
        "external_enabled": bool(source.get("external_enabled", DEFAULT_BACKUP_SETTINGS["external_enabled"])),
        "external_type": str(source.get("external_type") or DEFAULT_BACKUP_SETTINGS["external_type"]).strip() or "rclone",
        "rclone_remote": str(source.get("rclone_remote") or "").strip(),
        "rclone_path": str(source.get("rclone_path") or "").strip().strip("/"),
        "rclone_executable": str(source.get("rclone_executable") or DEFAULT_BACKUP_SETTINGS["rclone_executable"]).strip() or "rclone",
        "last_external_status": str(source.get("last_external_status") or "").strip(),
        "last_external_message": str(source.get("last_external_message") or "").strip(),
        "last_external_uploaded_at": str(source.get("last_external_uploaded_at") or "").strip(),
        "last_external_backup_name": str(source.get("last_external_backup_name") or "").strip(),
        "last_external_target": str(source.get("last_external_target") or "").strip(),
    }


def load_backup_settings() -> dict[str, object]:
    if not BACKUP_SETTINGS_PATH.exists():
        return normalize_backup_settings(DEFAULT_BACKUP_SETTINGS)
    try:
        raw = json.loads(BACKUP_SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        raw = {}
    if not isinstance(raw, dict):
        raw = {}
    merged = {**DEFAULT_BACKUP_SETTINGS, **raw}
    return normalize_backup_settings(merged)


def save_backup_settings(payload: dict) -> dict[str, object]:
    settings = normalize_backup_settings({**load_backup_settings(), **payload})
    target_dir = backup_dir_path(settings)
    target_dir.mkdir(parents=True, exist_ok=True)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_SETTINGS_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    return settings


def update_backup_external_status(result: dict[str, str]) -> dict[str, object]:
    settings = load_backup_settings()
    updated = {
        **settings,
        "last_external_status": result.get("status", ""),
        "last_external_message": result.get("message", ""),
        "last_external_uploaded_at": result.get("uploaded_at", ""),
        "last_external_backup_name": result.get("backup_name", ""),
        "last_external_target": result.get("target", ""),
    }
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_SETTINGS_PATH.write_text(json.dumps(normalize_backup_settings(updated), ensure_ascii=False, indent=2), encoding="utf-8")
    return normalize_backup_settings(updated)


def backup_dir_path(settings: dict[str, object] | None = None, backup_dir: str | Path | None = None) -> Path:
    if backup_dir:
        return Path(clean_backup_dir(str(backup_dir))).resolve()
    current = settings or load_backup_settings()
    configured = clean_backup_dir(current.get("backup_dir", ""))
    return Path(configured).resolve() if configured else backup_default_dir().resolve()


def rclone_target(settings: dict[str, object]) -> str:
    remote = str(settings.get("rclone_remote") or "").strip().rstrip(":")
    target_path = str(settings.get("rclone_path") or "").strip().strip("/")
    if not remote:
        raise ValueError("rclone 원격 이름을 입력해주세요.")
    return f"{remote}:{target_path}" if target_path else f"{remote}:"


def upload_backup_to_external_storage(backup_path: Path, settings: dict[str, object] | None = None, runner=None) -> dict[str, str]:
    current = settings or load_backup_settings()
    if not current.get("external_enabled"):
        return {
            "status": "disabled",
            "message": "외부 백업 업로드를 사용하지 않습니다.",
            "backup_name": backup_path.name,
            "target": "",
            "uploaded_at": "",
        }
    if str(current.get("external_type") or "rclone") != "rclone":
        return {
            "status": "fail",
            "message": "지원하지 않는 외부 백업 방식입니다.",
            "backup_name": backup_path.name,
            "target": "",
            "uploaded_at": now_text(),
        }
    try:
        target = rclone_target(current)
    except ValueError as exc:
        return {
            "status": "fail",
            "message": str(exc),
            "backup_name": backup_path.name,
            "target": "",
            "uploaded_at": now_text(),
        }
    command = [str(current.get("rclone_executable") or "rclone"), "copy", str(backup_path), target]
    run = runner or subprocess.run
    try:
        completed = run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=300,
            check=False,
        )
    except FileNotFoundError:
        return {
            "status": "fail",
            "message": "rclone 실행 파일을 찾지 못했습니다. VPS에 rclone 설치 및 PATH 설정을 확인해주세요.",
            "backup_name": backup_path.name,
            "target": target,
            "uploaded_at": now_text(),
        }
    except subprocess.TimeoutExpired:
        return {
            "status": "fail",
            "message": "rclone 업로드 시간이 초과되었습니다.",
            "backup_name": backup_path.name,
            "target": target,
            "uploaded_at": now_text(),
        }
    output = (getattr(completed, "stderr", "") or getattr(completed, "stdout", "") or "").strip()
    if int(getattr(completed, "returncode", 1)) != 0:
        return {
            "status": "fail",
            "message": output or "rclone 업로드에 실패했습니다.",
            "backup_name": backup_path.name,
            "target": target,
            "uploaded_at": now_text(),
        }
    return {
        "status": "success",
        "message": output or "Google Drive 업로드가 완료되었습니다.",
        "backup_name": backup_path.name,
        "target": target,
        "uploaded_at": now_text(),
    }


def valid_backup_name(name: str) -> bool:
    return bool(re.fullmatch(r"workhub_backup_\d{8}_\d{6}(?:_\d+)?\.zip", name))


def backup_path_from_name(name: str, backup_dir: str | Path | None = None) -> Path:
    name = Path(name).name
    if not valid_backup_name(name):
        raise ValueError("백업 파일명이 올바르지 않습니다.")
    root = backup_dir_path(backup_dir=backup_dir)
    path = (root / name).resolve()
    if root not in path.parents:
        raise ValueError("백업 파일 경로가 올바르지 않습니다.")
    return path


def backup_file_info(path: Path) -> dict[str, str | int]:
    stat = path.stat()
    return {
        "name": path.name,
        "size": stat.st_size,
        "created_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
    }


def list_backup_files(backup_dir: str | Path | None = None) -> list[dict[str, str | int]]:
    root = backup_dir_path(backup_dir=backup_dir)
    root.mkdir(parents=True, exist_ok=True)
    files = [path for path in root.glob("workhub_backup_*.zip") if path.is_file() and valid_backup_name(path.name)]
    files.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return [backup_file_info(path) for path in files]


def cleanup_backup_retention(backup_dir: str | Path | None = None, retention_days: int | None = None) -> None:
    settings = load_backup_settings()
    root = backup_dir_path(settings, backup_dir=backup_dir)
    root.mkdir(parents=True, exist_ok=True)
    days = retention_days if retention_days is not None else int(settings["retention_days"])
    cutoff = time.time() - (days * 24 * 60 * 60)
    for path in root.glob("workhub_backup_*.zip"):
        if path.is_file() and valid_backup_name(path.name) and path.stat().st_mtime < cutoff:
            path.unlink(missing_ok=True)


def safe_sqlite_copy(target_db: Path) -> None:
    init_db()
    source = connect_db()
    target = sqlite3.connect(target_db)
    try:
        source.backup(target)
        target.commit()
    finally:
        target.close()
        source.close()


def create_workhub_backup(reason: str = "manual", backup_dir: str | Path | None = None) -> dict[str, str | int]:
    settings = load_backup_settings()
    root = backup_dir_path(settings, backup_dir=backup_dir)
    root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = root / f"workhub_backup_{timestamp}.zip"
    counter = 1
    while output_path.exists():
        output_path = root / f"workhub_backup_{timestamp}_{counter}.zip"
        counter += 1
    with tempfile.TemporaryDirectory(prefix="workhub_backup_", dir=root) as temp_dir:
        temp_db = Path(temp_dir) / "workhub.db"
        safe_sqlite_copy(temp_db)
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.write(temp_db, "config/workhub.db")
            for path in (MAIL_SETTINGS_PATH, VENDOR_CONTACTS_PATH, BACKUP_SETTINGS_PATH, SECRET_KEY_PATH, CRM_WEBHOOK_TOKEN_PATH):
                if path.exists() and path.is_file():
                    archive.write(path, f"config/{path.name}")
            archive.writestr(
                "manifest.json",
                json.dumps(
                    {
                        "created_at": now_text(),
                        "reason": reason,
                        "data_dir": str(RUNTIME_ROOT),
                        "backup_dir": str(root),
                        "backup_retention_days": int(settings["retention_days"]),
                        "included": [
                            "config/workhub.db",
                            "config/mail_settings.json",
                            "config/vendor_contacts.json",
                            "config/backup_settings.json",
                            "config/secret.key",
                            "config/crm_webhook_token.txt",
                        ],
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
            )
    cleanup_backup_retention(backup_dir=root, retention_days=int(settings["retention_days"]))
    info = backup_file_info(output_path)
    external_result = upload_backup_to_external_storage(output_path, settings)
    info["external_backup"] = external_result
    if external_result["status"] != "disabled":
        update_backup_external_status(external_result)
    return info


def validate_restored_db(path: Path) -> None:
    connection = sqlite3.connect(path)
    try:
        result = connection.execute("PRAGMA quick_check").fetchone()
        if not result or str(result[0]).lower() != "ok":
            raise ValueError("백업 DB 파일 검증에 실패했습니다.")
    finally:
        connection.close()


def restore_workhub_backup(source_zip: Path) -> dict[str, object]:
    if not source_zip.exists() or not zipfile.is_zipfile(source_zip):
        raise ValueError("올바른 백업 zip 파일이 아닙니다.")

    pre_restore = create_workhub_backup("pre-restore")
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    restore_temp_dir = backup_dir_path()
    restore_temp_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="workhub_restore_", dir=restore_temp_dir) as temp_dir:
        temp_root = Path(temp_dir)
        extracted: dict[str, Path] = {}
        with zipfile.ZipFile(source_zip) as archive:
            names = {item.filename.replace("\\", "/").lstrip("/") for item in archive.infolist()}
            if "config/workhub.db" not in names:
                raise ValueError("백업 파일 안에 config/workhub.db가 없습니다.")
            for member in archive.infolist():
                member_name = member.filename.replace("\\", "/").lstrip("/")
                if member_name not in RESTORE_ALLOWED_FILES:
                    continue
                output = temp_root / member_name
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_bytes(archive.read(member))
                extracted[member_name] = output

        validate_restored_db(extracted["config/workhub.db"])

        for backup_name, target in RESTORE_ALLOWED_FILES.items():
            source = extracted.get(backup_name)
            if source:
                target.parent.mkdir(parents=True, exist_ok=True)
                source.replace(target)
            elif target.exists():
                target.unlink()

    return {
        "message": "백업 데이터 복원이 완료되었습니다.",
        "restored_from": source_zip.name,
        "pre_restore_backup": pre_restore,
    }


def has_backup_for_date(date_token: str) -> bool:
    root = backup_dir_path()
    root.mkdir(parents=True, exist_ok=True)
    return any(root.glob(f"workhub_backup_{date_token}_*.zip"))


def backup_scheduler_loop() -> None:
    while True:
        try:
            now = datetime.now()
            settings = load_backup_settings()
            if (
                settings.get("auto_enabled", True)
                and now.hour >= int(settings.get("auto_hour", AUTO_BACKUP_HOUR))
                and not has_backup_for_date(now.strftime("%Y%m%d"))
            ):
                created = create_workhub_backup("auto")
                print(f"자동 백업 완료: {created['name']}")
        except Exception as exc:  # noqa: BLE001
            print(f"자동 백업 실패: {exc}")
        time.sleep(5 * 60)


def start_backup_scheduler() -> None:
    global _BACKUP_SCHEDULER_STARTED
    if _BACKUP_SCHEDULER_STARTED:
        return
    _BACKUP_SCHEDULER_STARTED = True
    thread = threading.Thread(target=backup_scheduler_loop, name="workhub-backup-scheduler", daemon=True)
    thread.start()


def git_source_dir() -> Path | None:
    candidates: list[Path] = []
    env_source = os.environ.get("WORKHUB_SOURCE_DIR")
    if env_source:
        candidates.append(Path(env_source))
    candidates.extend([ROOT, RUNTIME_ROOT, Path.cwd()])
    for candidate in candidates:
        try:
            path = candidate.resolve()
        except OSError:
            continue
        if (path / ".git").exists():
            return path
    return None


def run_git(args: list[str], source_dir: Path, timeout: int = 30) -> tuple[int, str, str]:
    try:
        completed = subprocess.run(
            ["git", "-C", str(source_dir), *args],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
            check=False,
        )
        return completed.returncode, completed.stdout.strip(), completed.stderr.strip()
    except FileNotFoundError:
        return 127, "", "git 명령을 찾지 못했습니다."
    except subprocess.TimeoutExpired:
        return 124, "", "git 명령 시간이 초과되었습니다."


def git_output(args: list[str], source_dir: Path, default: str = "") -> str:
    code, stdout, _ = run_git(args, source_dir)
    return stdout if code == 0 else default


def system_update_status(fetch: bool = False) -> dict[str, object]:
    source_dir = git_source_dir()
    if not source_dir:
        return {
            "available": False,
            "source_dir": "",
            "branch": "",
            "current_commit": "",
            "current_short": "",
            "upstream_commit": "",
            "remote_url": "",
            "ahead": 0,
            "behind": 0,
            "dirty": False,
            "message": "현재 실행 위치가 Git 저장소가 아닙니다. 나스 배포 시 GitHub 저장소 폴더에서 실행하면 업데이트 기능을 사용할 수 있습니다.",
        }
    if fetch:
        run_git(["fetch", "--all", "--prune"], source_dir, timeout=60)
    current_commit = git_output(["rev-parse", "HEAD"], source_dir)
    current_short = git_output(["rev-parse", "--short", "HEAD"], source_dir)
    branch = git_output(["rev-parse", "--abbrev-ref", "HEAD"], source_dir)
    upstream = git_output(["rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{u}"], source_dir)
    upstream_commit = git_output(["rev-parse", "@{u}"], source_dir) if upstream else ""
    ahead = 0
    behind = 0
    if upstream:
        counts = git_output(["rev-list", "--left-right", "--count", "HEAD...@{u}"], source_dir)
        parts = counts.split()
        if len(parts) == 2 and all(part.isdigit() for part in parts):
            ahead = int(parts[0])
            behind = int(parts[1])
    dirty = bool(git_output(["status", "--porcelain"], source_dir))
    message = ""
    if not upstream:
        message = "원격 추적 브랜치가 설정되어 있지 않습니다."
    elif dirty:
        message = "저장되지 않은 코드 변경이 있어 자동 업데이트를 적용하지 않습니다."
    elif behind > 0:
        message = f"GitHub에 적용 가능한 업데이트 {behind}개가 있습니다."
    else:
        message = "최신 상태입니다."
    return {
        "available": bool(upstream),
        "source_dir": str(source_dir),
        "branch": branch,
        "current_commit": current_commit,
        "current_short": current_short,
        "upstream": upstream,
        "upstream_commit": upstream_commit,
        "remote_url": git_output(["config", "--get", "remote.origin.url"], source_dir),
        "ahead": ahead,
        "behind": behind,
        "dirty": dirty,
        "message": message,
    }


def list_system_update_history(limit: int = 30) -> list[dict[str, str | int]]:
    init_db()
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT id, action, before_commit, after_commit, status, message, backup_name, created_at
              FROM system_update_history
             ORDER BY id DESC
             LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def record_system_update(action: str, before_commit: str, after_commit: str, status: str, message: str, backup_name: str = "") -> None:
    init_db()
    connection = connect_db()
    try:
        connection.execute(
            """
            INSERT INTO system_update_history
                (action, before_commit, after_commit, status, message, backup_name, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (action, before_commit, after_commit, status, message, backup_name, now_text()),
        )
        connection.commit()
    finally:
        connection.close()


def system_update_payload(fetch: bool = False) -> dict[str, object]:
    return {
        "status": system_update_status(fetch=fetch),
        "history": list_system_update_history(),
    }


def apply_system_update() -> dict[str, object]:
    if not _SYSTEM_UPDATE_LOCK.acquire(blocking=False):
        raise ValueError("이미 업데이트가 진행 중입니다.")
    backup_name = ""
    before_commit = ""
    after_commit = ""
    try:
        status = system_update_status(fetch=True)
        before_commit = str(status.get("current_short") or status.get("current_commit") or "")
        if not status.get("available"):
            raise ValueError(str(status.get("message") or "업데이트 기능을 사용할 수 없습니다."))
        if status.get("dirty"):
            raise ValueError("저장되지 않은 코드 변경이 있어 업데이트를 중단했습니다.")
        if int(status.get("behind") or 0) <= 0:
            record_system_update("apply", before_commit, before_commit, "skip", "이미 최신 상태입니다.", "")
            payload = system_update_payload(fetch=False)
            payload["message"] = "이미 최신 상태입니다."
            return payload

        backup = create_workhub_backup("pre-update")
        backup_name = str(backup["name"])
        source_dir = Path(str(status["source_dir"]))
        code, stdout, stderr = run_git(["pull", "--ff-only"], source_dir, timeout=120)
        if code != 0:
            message = stderr or stdout or "git pull 실패"
            record_system_update("apply", before_commit, before_commit, "fail", message, backup_name)
            raise ValueError(message)
        new_status = system_update_status(fetch=False)
        after_commit = str(new_status.get("current_short") or new_status.get("current_commit") or "")
        message = stdout or "업데이트 적용 완료"
        record_system_update("apply", before_commit, after_commit, "success", message, backup_name)
        payload = system_update_payload(fetch=False)
        payload["message"] = "업데이트 적용이 완료되었습니다."
        payload["backup_name"] = backup_name
        payload["restart_required"] = True
        return payload
    finally:
        _SYSTEM_UPDATE_LOCK.release()


def get_leave_type_id(code: str = "annual") -> int:
    connection = connect_db()
    try:
        row = connection.execute("SELECT id FROM leave_types WHERE code = ?", (code,)).fetchone()
        if not row:
            raise ValueError("\uC5F0\uCC28 \uC720\uD615\uC744 \uCC3E\uC744 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
        return int(row["id"])
    finally:
        connection.close()


def ensure_leave_balance(connection: sqlite3.Connection, user_id: int, leave_type_id: int) -> sqlite3.Row:
    now = now_text()
    connection.execute(
        """
        INSERT OR IGNORE INTO leave_balances
            (user_id, leave_type_id, total_days, used_days, remaining_days, reserved_days, created_at, updated_at)
        VALUES (?, ?, 0, 0, 0, 0, ?, ?)
        """,
        (user_id, leave_type_id, now, now),
    )
    row = connection.execute(
        "SELECT * FROM leave_balances WHERE user_id = ? AND leave_type_id = ?",
        (user_id, leave_type_id),
    ).fetchone()
    if not row:
        raise ValueError("\uC5F0\uCC28 \uC794\uC5EC \uC815\uBCF4\uB97C \uCC3E\uC744 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
    return row


def balance_remaining(total_days: float, used_days: float, reserved_days: float) -> float:
    return round(max(0.0, float(total_days or 0) - float(used_days or 0) - float(reserved_days or 0)), 2)


def update_leave_balance_amounts(
    connection: sqlite3.Connection,
    balance_id: int,
    *,
    total_days: float | None = None,
    used_days: float | None = None,
    reserved_days: float | None = None,
    accrual_year: int | None = None,
) -> sqlite3.Row:
    current = connection.execute("SELECT * FROM leave_balances WHERE id = ?", (balance_id,)).fetchone()
    if not current:
        raise ValueError("\uC5F0\uCC28 \uC794\uC5EC \uC815\uBCF4\uB97C \uCC3E\uC744 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
    total = round(float(current["total_days"] if total_days is None else total_days), 2)
    used = round(float(current["used_days"] if used_days is None else used_days), 2)
    reserved = round(max(0.0, float(current["reserved_days"] if reserved_days is None else reserved_days)), 2)
    remaining = balance_remaining(total, used, reserved)
    if used + reserved > total:
        raise ValueError("\uC0AC\uC6A9/\uC608\uC57D \uC77C\uC218\uAC00 \uCD1D \uC5F0\uCC28\uB97C \uCD08\uACFC\uD569\uB2C8\uB2E4.")
    if accrual_year is None:
        connection.execute(
            "UPDATE leave_balances SET total_days = ?, used_days = ?, reserved_days = ?, remaining_days = ?, updated_at = ? WHERE id = ?",
            (total, used, reserved, remaining, now_text(), balance_id),
        )
    else:
        connection.execute(
            "UPDATE leave_balances SET total_days = ?, used_days = ?, reserved_days = ?, remaining_days = ?, accrual_year = ?, updated_at = ? WHERE id = ?",
            (total, used, reserved, remaining, accrual_year, now_text(), balance_id),
        )
    return connection.execute("SELECT * FROM leave_balances WHERE id = ?", (balance_id,)).fetchone()


def list_leave_types() -> list[dict[str, str | int]]:
    connection = connect_db()
    try:
        rows = connection.execute(
            "SELECT id, code, name, is_paid, is_active FROM leave_types WHERE is_active = 1 ORDER BY id"
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def list_active_users_for_leave() -> list[dict[str, str | int]]:
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT id, username, display_name
              FROM users
             WHERE active = 1
             ORDER BY display_name COLLATE NOCASE, username COLLATE NOCASE
            """
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def leave_balance_rows(connection: sqlite3.Connection, user_id: int) -> list[dict[str, str | float]]:
    annual_type_id = get_leave_type_id("annual")
    ensure_leave_balance(connection, user_id, annual_type_id)
    rows = connection.execute(
        """
        SELECT leave_types.code, leave_types.name, leave_balances.total_days,
               leave_balances.used_days, leave_balances.reserved_days, leave_balances.remaining_days
          FROM leave_balances
          JOIN leave_types ON leave_types.id = leave_balances.leave_type_id
         WHERE leave_balances.user_id = ?
         ORDER BY leave_types.id
        """,
        (user_id,),
    ).fetchall()
    return [
        {
            "code": row["code"],
            "name": row["name"],
            "total_days": round(float(row["total_days"] or 0), 2),
            "used_days": round(float(row["used_days"] or 0), 2),
            "reserved_days": round(float(row["reserved_days"] or 0), 2),
            "remaining_days": round(float(row["remaining_days"] or 0), 2),
        }
        for row in rows
    ]


def leave_approval_label(step: str) -> str:
    return {
        "TEAM_LEAD": "\uD300\uC7A5 \uD655\uC778",
        "DIRECTOR": "\uC2E4\uC7A5 \uD655\uC778",
        "CEO": "\uB300\uD45C \uD655\uC778",
        "COMPLETED": "\uC2B9\uC778 \uC644\uB8CC",
    }.get(step, step)


def leave_step_permission(step: str) -> str:
    return {
        "TEAM_LEAD": "leave_approve_team",
        "DIRECTOR": "leave_approve_director",
        "CEO": "leave_approve_ceo",
    }.get(step, "leave_approve")


def actor_can_approve_leave_step(actor: dict[str, str], step: str) -> bool:
    return (
        user_has_permission(actor, "leave_manage")
        or user_has_permission(actor, "leave_approve")
        or user_has_permission(actor, leave_step_permission(step))
    )


def actor_can_override_leave(actor: dict[str, str]) -> bool:
    return user_has_permission(actor, "leave_director_override") or user_has_permission(actor, "leave_manage")


def active_users_with_leave_permission(connection: sqlite3.Connection, permissions: list[str]) -> list[sqlite3.Row]:
    rows = connection.execute(
        "SELECT id, username, display_name, role, permissions FROM users WHERE active = 1"
    ).fetchall()
    matches = []
    for row in rows:
        normalized = normalize_permissions(row["permissions"], row["role"])
        if any(permission in normalized for permission in permissions):
            matches.append(row)
    return matches


def add_leave_notification(
    connection: sqlite3.Connection,
    user_id: int,
    request_id: int | None,
    notification_type: str,
    message: str,
) -> None:
    connection.execute(
        """
        INSERT INTO leave_notifications (user_id, request_id, notification_type, message, is_read, created_at)
        VALUES (?, ?, ?, ?, 0, ?)
        """,
        (user_id, request_id, notification_type, message, now_text()),
    )


def notify_leave_step(connection: sqlite3.Connection, request_id: int, step: str, requester_name: str) -> None:
    permission = leave_step_permission(step)
    permissions = [permission, "leave_manage"]
    if step == "TEAM_LEAD":
        permissions.append("leave_approve")
    if step == "DIRECTOR":
        permissions.append("leave_director_override")
    users = active_users_with_leave_permission(connection, permissions)
    message = f"{requester_name}\uB2D8\uC758 \uC5F0\uCC28 \uC2E0\uCCAD\uC774 {leave_approval_label(step)} \uC2B9\uC778\uC744 \uAE30\uB2E4\uB9BD\uB2C8\uB2E4."
    for user in users:
        add_leave_notification(connection, int(user["id"]), request_id, "approval_waiting", message)


def notify_leave_requester(connection: sqlite3.Connection, user_id: int, request_id: int, message: str) -> None:
    add_leave_notification(connection, user_id, request_id, "request_update", message)


def list_leave_notifications(user: dict[str, str]) -> list[dict[str, str | int]]:
    init_db()
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT id, request_id, notification_type, message, is_read, created_at
              FROM leave_notifications
             WHERE user_id = ?
             ORDER BY created_at DESC, id DESC
             LIMIT 30
            """,
            (int(user["id"]),),
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def list_leave_requests_for_user(connection: sqlite3.Connection, user_id: int) -> list[dict[str, str | float | int]]:
    rows = connection.execute(
        """
        SELECT leave_requests.*, leave_types.name AS leave_type_name, users.display_name
          FROM leave_requests
          JOIN leave_types ON leave_types.id = leave_requests.leave_type_id
          JOIN users ON users.id = leave_requests.user_id
         WHERE leave_requests.user_id = ?
         ORDER BY leave_requests.start_date DESC, leave_requests.id DESC
        """,
        (user_id,),
    ).fetchall()
    return [leave_request_dict(row) for row in rows]


def leave_request_dict(row: sqlite3.Row) -> dict[str, str | float | int]:
    approval_step = row["approval_step"] if "approval_step" in row.keys() else "COMPLETED"
    return {
        "id": row["id"],
        "user_id": row["user_id"],
        "requester": row["display_name"] if "display_name" in row.keys() else "",
        "leave_type_name": row["leave_type_name"] if "leave_type_name" in row.keys() else "\uC5F0\uCC28",
        "start_date": row["start_date"],
        "end_date": row["end_date"],
        "unit": row["unit"],
        "unit_label": "\uBC18\uCC28" if row["unit"] == "HALF_DAY" else "\uC885\uC77C",
        "requested_days": round(float(row["requested_days"] or 0), 2),
        "reason": row["reason"],
        "status": row["status"],
        "status_label": leave_status_label(row["status"]),
        "approval_step": approval_step,
        "approval_step_label": leave_approval_label(approval_step),
        "team_status": row["team_status"] if "team_status" in row.keys() else "APPROVED",
        "director_status": row["director_status"] if "director_status" in row.keys() else "APPROVED",
        "ceo_status": row["ceo_status"] if "ceo_status" in row.keys() else "APPROVED",
        "team_comment": row["team_comment"] or "" if "team_comment" in row.keys() else "",
        "director_comment": row["director_comment"] or "" if "director_comment" in row.keys() else "",
        "ceo_comment": row["ceo_comment"] or "" if "ceo_comment" in row.keys() else "",
        "rejected_reason": row["rejected_reason"] or "",
        "cancel_reason": row["cancel_reason"] or "" if "cancel_reason" in row.keys() else "",
        "created_at": row["created_at"],
    }


def list_pending_leave_requests(connection: sqlite3.Connection) -> list[dict[str, str | float | int]]:
    rows = connection.execute(
        """
        SELECT leave_requests.*, leave_types.name AS leave_type_name, users.display_name
          FROM leave_requests
          JOIN leave_types ON leave_types.id = leave_requests.leave_type_id
          JOIN users ON users.id = leave_requests.user_id
         WHERE leave_requests.status = 'PENDING'
         ORDER BY leave_requests.created_at ASC
        """
    ).fetchall()
    return [leave_request_dict(row) for row in rows]


def list_leave_admin_balances(connection: sqlite3.Connection) -> list[dict[str, str | float | int]]:
    annual_type_id = get_leave_type_id("annual")
    users = list_active_users_for_leave()
    rows = []
    for user in users:
        balance = ensure_leave_balance(connection, int(user["id"]), annual_type_id)
        rows.append({
            "user_id": user["id"],
            "username": user["username"],
            "display_name": user["display_name"],
            "total_days": round(float(balance["total_days"] or 0), 2),
            "used_days": round(float(balance["used_days"] or 0), 2),
            "reserved_days": round(float(balance["reserved_days"] or 0), 2),
            "remaining_days": round(float(balance["remaining_days"] or 0), 2),
        })
    return rows


def leave_payload(user: dict[str, str]) -> dict:
    user_id = int(user["id"])
    connection = connect_db()
    try:
        balances = leave_balance_rows(connection, user_id)
        annual = next((row for row in balances if row["code"] == "annual"), balances[0] if balances else {})
        can_approve = any(
            user_has_permission(user, permission)
            for permission in ("leave_approve", "leave_approve_team", "leave_approve_director", "leave_approve_ceo", "leave_director_override", "leave_manage")
        )
        payload = {
            "summary": {
                "total_days": annual.get("total_days", 0),
                "used_days": annual.get("used_days", 0),
                "reserved_days": annual.get("reserved_days", 0),
                "remaining_days": annual.get("remaining_days", 0),
            },
            "balances": balances,
            "requests": list_leave_requests_for_user(connection, user_id),
            "leave_types": list_leave_types(),
            "can_approve": can_approve,
            "can_override": actor_can_override_leave(user),
            "can_manage": user_has_permission(user, "leave_manage"),
            "pending_requests": [],
            "users": [],
            "admin_balances": [],
            "notifications": list_leave_notifications(user),
        }
        if payload["can_approve"]:
            payload["pending_requests"] = list_pending_leave_requests(connection)
        if payload["can_manage"]:
            payload["users"] = list_active_users_for_leave()
            payload["admin_balances"] = list_leave_admin_balances(connection)
        connection.commit()
        return payload
    finally:
        connection.close()


def create_leave_request(user: dict[str, str], payload: dict) -> int:
    init_db()
    user_id = int(user["id"])
    leave_type_id = int(payload.get("leave_type_id") or get_leave_type_id("annual"))
    start = parse_iso_date(payload.get("start_date"))
    end = parse_iso_date(payload.get("end_date"))
    unit = clean_payload_text(payload, "unit") or "FULL_DAY"
    if unit not in {"FULL_DAY", "HALF_DAY"}:
        raise ValueError("\uC5F0\uCC28 \uC720\uD615\uC774 \uBE44\uD65C\uC131\uD654\uB418\uC5C8\uC2B5\uB2C8\uB2E4.")
    requested_days = calculate_leave_days(start, end, unit)
    reason = clean_payload_text(payload, "reason")
    if not reason:
        raise ValueError("\uC5F0\uCC28 \uC2E0\uCCAD\uC744 \uCDE8\uC18C\uD588\uC2B5\uB2C8\uB2E4.")
    now = now_text()
    connection = connect_db()
    try:
        balance = ensure_leave_balance(connection, user_id, leave_type_id)
        if float(balance["remaining_days"] or 0) < requested_days:
            raise ValueError("\uC794\uC5EC \uC5F0\uCC28\uAC00 \uBD80\uC871\uD569\uB2C8\uB2E4.")
        cursor = connection.execute(
            """
            INSERT INTO leave_requests
                (user_id, leave_type_id, start_date, end_date, unit, requested_days, reason,
                 status, approval_step, team_status, director_status, ceo_status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'PENDING', 'TEAM_LEAD', 'PENDING', 'WAITING', 'WAITING', ?, ?)
            """,
            (user_id, leave_type_id, start.isoformat(), end.isoformat(), unit, requested_days, reason, now, now),
        )
        request_id = int(cursor.lastrowid)
        update_leave_balance_amounts(
            connection,
            int(balance["id"]),
            reserved_days=round(float(balance["reserved_days"] or 0) + requested_days, 2),
        )
        requester_name = str(user.get("display_name") or user.get("username") or "\uC2E0\uCCAD\uC790")
        notify_leave_step(connection, request_id, "TEAM_LEAD", requester_name)
        connection.commit()
        return request_id
    finally:
        connection.close()


def set_leave_step_decision(
    connection: sqlite3.Connection,
    request_id: int,
    step: str,
    status: str,
    actor_id: int,
    comment: str,
) -> None:
    prefix = {"TEAM_LEAD": "team", "DIRECTOR": "director", "CEO": "ceo"}[step]
    connection.execute(
        f"""
        UPDATE leave_requests
           SET {prefix}_status = ?, {prefix}_decided_by = ?, {prefix}_decided_at = ?, {prefix}_comment = ?, updated_at = ?
         WHERE id = ?
        """,
        (status, actor_id, now_text(), comment, now_text(), request_id),
    )


def release_leave_reservation(connection: sqlite3.Connection, row: sqlite3.Row) -> sqlite3.Row:
    balance = ensure_leave_balance(connection, int(row["user_id"]), int(row["leave_type_id"]))
    requested_days = float(row["requested_days"] or 0)
    return update_leave_balance_amounts(
        connection,
        int(balance["id"]),
        reserved_days=round(max(0.0, float(balance["reserved_days"] or 0) - requested_days), 2),
    )


def finalize_leave_approval(
    connection: sqlite3.Connection,
    row: sqlite3.Row,
    actor: dict[str, str],
    comment: str,
    overridden: bool = False,
) -> None:
    request_id = int(row["id"])
    now = now_text()
    balance = ensure_leave_balance(connection, int(row["user_id"]), int(row["leave_type_id"]))
    requested_days = float(row["requested_days"] or 0)
    if float(balance["reserved_days"] or 0) < requested_days and float(balance["remaining_days"] or 0) < requested_days:
        raise ValueError("\uC5F0\uCC28 \uC2E0\uCCAD\uC740 \uC2B9\uC778 \uB300\uAE30 \uC0C1\uD0DC\uC5D0\uC11C\uB9CC \uCC98\uB9AC\uD560 \uC218 \uC788\uC2B5\uB2C8\uB2E4.")
    used = round(float(balance["used_days"] or 0) + requested_days, 2)
    reserved = round(max(0.0, float(balance["reserved_days"] or 0) - requested_days), 2)
    updated_balance = update_leave_balance_amounts(connection, int(balance["id"]), used_days=used, reserved_days=reserved)
    connection.execute(
        """
        INSERT INTO leave_balance_ledger
            (balance_id, user_id, actor_id, delta_days, reason, request_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (updated_balance["id"], row["user_id"], int(actor["id"]), -requested_days, "\uC5F0\uCC28 \uC2B9\uC778 \uC644\uB8CC", request_id, now),
    )
    if overridden:
        for step in ("TEAM_LEAD", "DIRECTOR", "CEO"):
            prefix = {"TEAM_LEAD": "team", "DIRECTOR": "director", "CEO": "ceo"}[step]
            if row[f"{prefix}_status"] not in {"APPROVED", "REJECTED"}:
                connection.execute(
                    f"UPDATE leave_requests SET {prefix}_status = 'OVERRIDDEN', {prefix}_decided_by = ?, {prefix}_decided_at = ?, {prefix}_comment = ? WHERE id = ?",
                    (int(actor["id"]), now, comment, request_id),
                )
    connection.execute(
        """
        UPDATE leave_requests
           SET status = 'APPROVED', approval_step = 'COMPLETED', decided_by = ?, finalized_at = ?, updated_at = ?
         WHERE id = ?
        """,
        (int(actor["id"]), now, now, request_id),
    )
    notify_leave_requester(connection, int(row["user_id"]), request_id, "\uC5F0\uCC28 \uC2E0\uCCAD\uC774 \uCD5C\uC885 \uC2B9\uC778\uB418\uC5C8\uC2B5\uB2C8\uB2E4.")


def decide_leave_request(request_id: int, actor: dict[str, str], decision: str, comment: str = "") -> None:
    init_db()
    if decision not in {"approve", "reject", "override"}:
        raise ValueError("\uC5F0\uCC28 \uC2E0\uCCAD\uC744 \uCC3E\uC744 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
    clean_comment = str(comment or "").strip()
    connection = connect_db()
    try:
        row = connection.execute("SELECT * FROM leave_requests WHERE id = ?", (request_id,)).fetchone()
        if not row:
            raise ValueError("\uC774\uBBF8 \uCC98\uB9AC\uB41C \uC2E0\uCCAD\uC785\uB2C8\uB2E4.")
        if row["status"] != "PENDING":
            raise ValueError("\uC804\uACB0 \uAD8C\uD55C\uC774 \uC5C6\uC2B5\uB2C8\uB2E4.")
        step = str(row["approval_step"] or "TEAM_LEAD")
        if decision == "override":
            if not actor_can_override_leave(actor):
                raise ValueError("\uC2B9\uC778 \uAD8C\uD55C\uC774 \uC5C6\uC2B5\uB2C8\uB2E4.")
            finalize_leave_approval(connection, row, actor, clean_comment or "\uC2E4\uC7A5 \uC804\uACB0", overridden=True)
            connection.commit()
            return
        if not actor_can_approve_leave_step(actor, step):
            raise ValueError(f"{leave_approval_label(step)} \uAD8C\uD55C\uC774 \uC5C6\uC2B5\uB2C8\uB2E4.")
        if decision == "reject":
            set_leave_step_decision(connection, request_id, step, "REJECTED", int(actor["id"]), clean_comment or "\uBC18\uB824")
            release_leave_reservation(connection, row)
            now = now_text()
            connection.execute(
                """
                UPDATE leave_requests
                   SET status = 'REJECTED', approval_step = 'COMPLETED', rejected_reason = ?, decided_by = ?, finalized_at = ?, updated_at = ?
                 WHERE id = ?
                """,
                (clean_comment or "\uBC18\uB824", int(actor["id"]), now, now, request_id),
            )
            notify_leave_requester(connection, int(row["user_id"]), request_id, "\uC5F0\uCC28 \uC2E0\uCCAD\uC774 \uBC18\uB824\uB418\uC5C8\uC2B5\uB2C8\uB2E4.")
            connection.commit()
            return
        set_leave_step_decision(connection, request_id, step, "APPROVED", int(actor["id"]), clean_comment)
        if step == "TEAM_LEAD":
            connection.execute(
                "UPDATE leave_requests SET approval_step = 'DIRECTOR', director_status = 'PENDING', updated_at = ? WHERE id = ?",
                (now_text(), request_id),
            )
            notify_leave_step(connection, request_id, "DIRECTOR", row["display_name"] if "display_name" in row.keys() else "\uC2E0\uCCAD\uC790")
        elif step == "DIRECTOR":
            connection.execute(
                "UPDATE leave_requests SET approval_step = 'CEO', ceo_status = 'PENDING', updated_at = ? WHERE id = ?",
                (now_text(), request_id),
            )
            notify_leave_step(connection, request_id, "CEO", row["display_name"] if "display_name" in row.keys() else "\uC2E0\uCCAD\uC790")
        else:
            refreshed = connection.execute("SELECT * FROM leave_requests WHERE id = ?", (request_id,)).fetchone()
            finalize_leave_approval(connection, refreshed, actor, clean_comment)
        connection.commit()
    finally:
        connection.close()


def cancel_leave_request(request_id: int, actor: dict[str, str], reason: str = "") -> None:
    init_db()
    connection = connect_db()
    try:
        row = connection.execute("SELECT * FROM leave_requests WHERE id = ?", (request_id,)).fetchone()
        if not row:
            raise ValueError("\uC5F0\uCC28 \uC2E0\uCCAD\uC744 \uCC3E\uC744 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
        if row["status"] != "PENDING":
            raise ValueError("\uC2B9\uC778 \uB300\uAE30 \uC0C1\uD0DC\uC758 \uC2E0\uCCAD\uB9CC \uCDE8\uC18C\uD560 \uC218 \uC788\uC2B5\uB2C8\uB2E4.")
        if int(row["user_id"]) != int(actor["id"]) and not user_has_permission(actor, "leave_manage"):
            raise ValueError("\uBCF8\uC778 \uC2E0\uCCAD\uB9CC \uCDE8\uC18C\uD560 \uC218 \uC788\uC2B5\uB2C8\uB2E4.")
        release_leave_reservation(connection, row)
        now = now_text()
        clean_reason = str(reason or "").strip() or "\uC2E0\uCCAD\uC790 \uCDE8\uC18C"
        connection.execute(
            """
            UPDATE leave_requests
               SET status = 'CANCELED', approval_step = 'COMPLETED', cancel_reason = ?, canceled_by = ?, canceled_at = ?, updated_at = ?
             WHERE id = ?
            """,
            (clean_reason, int(actor["id"]), now, now, request_id),
        )
        notify_leave_requester(connection, int(row["user_id"]), request_id, "\uC5F0\uCC28 \uC2E0\uCCAD\uC774 \uCDE8\uC18C\uB418\uC5C8\uC2B5\uB2C8\uB2E4.")
        connection.commit()
    finally:
        connection.close()


def set_leave_balance(payload: dict, actor: dict[str, str]) -> None:
    init_db()
    user_id = int(payload.get("user_id") or 0)
    if not user_id:
        raise ValueError("\uC9C1\uC6D0\uC744 \uC120\uD0DD\uD574\uC8FC\uC138\uC694.")
    total = clean_leave_days(payload.get("total_days"), "\uCD1D \uC5F0\uCC28")
    used = clean_leave_days(payload.get("used_days"), "\uC0AC\uC6A9 \uC5F0\uCC28")
    if used > total:
        raise ValueError("\uC0AC\uC6A9 \uC5F0\uCC28\uB294 \uCD1D \uC5F0\uCC28\uBCF4\uB2E4 \uD074 \uC218 \uC5C6\uC2B5\uB2C8\uB2E4.")
    leave_type_id = get_leave_type_id("annual")
    now = now_text()
    connection = connect_db()
    try:
        balance = ensure_leave_balance(connection, user_id, leave_type_id)
        update_leave_balance_amounts(connection, int(balance["id"]), total_days=total, used_days=used, reserved_days=0)
        connection.execute(
            """
            INSERT INTO leave_balance_ledger
                (balance_id, user_id, actor_id, delta_days, reason, request_id, created_at)
            VALUES (?, ?, ?, ?, ?, NULL, ?)
            """,
            (balance["id"], user_id, int(actor["id"]), 0, f"\uC5F0\uCC28 \uC794\uC5EC \uC218\uC815: \uCD1D {total}\uC77C, \uC0AC\uC6A9 {used}\uC77C", now),
        )
        connection.commit()
    finally:
        connection.close()


def parse_historical_leave_lines(text: str) -> list[tuple[date, str, float]]:
    entries: list[tuple[date, str, float]] = []
    seen: set[tuple[str, float]] = set()
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = re.match("^(\\d{4}-\\d{2}-\\d{2})(?:[\\s,]+(\uBC18\uCC28|0\\.5|1|1\\.0|\uC885\uC77C|\uC804\uC77C))?$", line)
        if not match:
            raise ValueError(f"\uAE30\uC874 \uC0AC\uC6A9\uBD84 \uD615\uC2DD\uC744 \uD655\uC778\uD574\uC8FC\uC138\uC694: {line}")
        used_date = parse_iso_date(match.group(1))
        unit_text = match.group(2) or "\uC804\uC77C"
        unit = "HALF_DAY" if unit_text in {"\uBC18\uCC28", "0.5"} else "FULL_DAY"
        days = 0.5 if unit == "HALF_DAY" else 1.0
        key = (used_date.isoformat(), days)
        if key in seen:
            raise ValueError(f"\uC911\uBCF5\uB41C \uC0AC\uC6A9 \uC77C\uC790\uAC00 \uC788\uC2B5\uB2C8\uB2E4: {used_date.isoformat()}")
        seen.add(key)
        entries.append((used_date, unit, days))
    if not entries:
        raise ValueError("\uC0AC\uC6A9 \uC77C\uC790\uB97C 1\uAC74 \uC774\uC0C1 \uC785\uB825\uD574\uC8FC\uC138\uC694.")
    return entries


def add_historical_leave_usage(payload: dict, actor: dict[str, str]) -> int:
    init_db()
    user_id = int(payload.get("user_id") or 0)
    if not user_id:
        raise ValueError("\uC9C1\uC6D0\uC744 \uC120\uD0DD\uD574\uC8FC\uC138\uC694.")
    entries = parse_historical_leave_lines(str(payload.get("usage_dates", "") or ""))
    note = clean_payload_text(payload, "note")
    leave_type_id = get_leave_type_id("annual")
    now = now_text()
    connection = connect_db()
    try:
        balance = ensure_leave_balance(connection, user_id, leave_type_id)
        total_days = round(sum(entry[2] for entry in entries), 2)
        if float(balance["remaining_days"] or 0) < total_days:
            raise ValueError("\uC794\uC5EC \uC5F0\uCC28\uAC00 \uAE30\uC874 \uC0AC\uC6A9\uBD84\uBCF4\uB2E4 \uBD80\uC871\uD569\uB2C8\uB2E4.")
        for used_date, unit, days in entries:
            reason = f"\uAE30\uC874 \uC5F0\uCC28 \uC0AC\uC6A9 \uBC18\uC601{': ' + note if note else ''}"
            cursor = connection.execute(
                """
                INSERT INTO leave_requests
                    (user_id, leave_type_id, start_date, end_date, unit, requested_days, reason,
                     status, approval_step, team_status, director_status, ceo_status, decided_by, finalized_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'APPROVED', 'COMPLETED', 'APPROVED', 'APPROVED', 'APPROVED', ?, ?, ?, ?)
                """,
                (user_id, leave_type_id, used_date.isoformat(), used_date.isoformat(), unit, days, reason, int(actor["id"]), now, now, now),
            )
            connection.execute(
                """
                INSERT INTO leave_balance_ledger
                    (balance_id, user_id, actor_id, delta_days, reason, request_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (balance["id"], user_id, int(actor["id"]), -days, reason, int(cursor.lastrowid), now),
            )
        update_leave_balance_amounts(
            connection,
            int(balance["id"]),
            used_days=round(float(balance["used_days"] or 0) + total_days, 2),
        )
        connection.commit()
        return len(entries)
    finally:
        connection.close()


def apply_annual_leave_accrual(year: int | None = None, actor: dict[str, str] | None = None, default_days: float = 15.0) -> int:
    init_db()
    target_year = int(year or date.today().year)
    annual_type_id = get_leave_type_id("annual")
    actor_id = int((actor or {}).get("id") or 0) or None
    now = now_text()
    updated = 0
    connection = connect_db()
    try:
        users = connection.execute("SELECT id FROM users WHERE active = 1").fetchall()
        for user in users:
            balance = ensure_leave_balance(connection, int(user["id"]), annual_type_id)
            if int(balance["accrual_year"] or 0) == target_year:
                continue
            updated_balance = update_leave_balance_amounts(
                connection,
                int(balance["id"]),
                total_days=float(default_days),
                used_days=0,
                reserved_days=0,
                accrual_year=target_year,
            )
            connection.execute(
                """
                INSERT INTO leave_balance_ledger
                    (balance_id, user_id, actor_id, delta_days, reason, request_id, created_at)
                VALUES (?, ?, ?, ?, ?, NULL, ?)
                """,
                (updated_balance["id"], user["id"], actor_id, float(default_days), f"{target_year}\uB144 \uC5F0\uCC28 \uC790\uB3D9 \uBC1C\uC0DD", now),
            )
            add_leave_notification(connection, int(user["id"]), None, "accrual", f"{target_year}\uB144 \uC5F0\uCC28 {default_days:g}\uC77C\uC774 \uC790\uB3D9 \uBC1C\uC0DD\uB418\uC5C8\uC2B5\uB2C8\uB2E4.")
            updated += 1
        connection.commit()
        return updated
    finally:
        connection.close()


def extract_invoice_number(value: str) -> str:
    numbers = re.findall(r"\d{6,}", value.replace("-", "").replace(" ", ""))
    return numbers[-1] if numbers else ""


def normalize_compact(value: str) -> str:
    return re.sub(r"[\s/_-]+", "", str(value or "").strip())


def status_date_text(*values: object) -> str:
    for value in values:
        text = clean_cell(value)
        if not text:
            continue
        month_day = re.search(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일?", text)
        if month_day:
            return f"{int(month_day.group(1))}/{int(month_day.group(2))}"
        slash_day = re.search(r"(?<!\d)(\d{1,2})[./-](\d{1,2})(?!\d)", text)
        if slash_day:
            return f"{int(slash_day.group(1))}/{int(slash_day.group(2))}"
        iso_day = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
        if iso_day:
            return f"{int(iso_day.group(2))}/{int(iso_day.group(3))}"
    now = datetime.now()
    return f"{now.month}/{now.day}"


def normalize_cs_type_value(cs_type: str, cs_content: str, status: str = "") -> str:
    existing = clean_cell(cs_type)
    if existing in {"변심반품", "변신반품", "불량반품", "불량교환", "불량재출고(미회수)", "오출고(오배송)"}:
        return "변심반품" if existing == "변신반품" else existing

    text = normalize_compact(f"{existing} {cs_content} {status}")
    if not text:
        return ""

    if any(keyword in text for keyword in ["오배송", "오출고", "오발주", "착오", "중복발주", "중복출고", "이중발주", "이중출고", "주소오류", "주소오기재"]):
        return "오출고(오배송)"
    if any(keyword in text for keyword in ["출고취소", "출고전취소", "출소취소"]):
        return "오출고(오배송)"
    if any(keyword in text for keyword in ["변심", "단순변심"]):
        return "변심반품"
    if any(keyword in text for keyword in ["맞교환", "맞효관", "불량교환", "제품불량교환", "교환요청", "교환원", "교환진행", "교환"]):
        return "불량교환"
    if any(keyword in text for keyword in ["상품누락", "누락재발송", "누락재배송", "부분재발송", "부분재출고", "부족분재발송", "일부재발송", "일부분재배송", "미수령재배송", "재출고", "대체출고", "재발송요청", "배출고", "누락배송", "부분누락", "부분발송", "부분미발송", "미발송", "추가배송", "추가발송", "부품구매", "부품출고", "추가구매", "제품미수령", "화물추적", "배송추적", "배송확인", "운송장회신", "충전기분실", "재발송", "발송완료"]):
        return "불량재출고(미회수)"
    if any(keyword in text for keyword in ["불량반품", "제품불량반품", "부분반품", "반품회수", "불량", "파손배송", "파손불량", "환불", "회수완료", "회수요청", "코팅벗겨짐", "입고사유확인"]):
        return "불량반품"
    if "반품" in text:
        return "변심반품"
    return ""


def normalize_progress_status(status: str, completed_at: str = "", occurred_at: str = "") -> str:
    raw = clean_cell(status)
    text = normalize_compact(raw)
    date_text = status_date_text(raw, completed_at, occurred_at)

    has_return_complete = any(keyword in text for keyword in ["회수완료", "입고완료", "반품완료", "수거완료", "반송완료"])
    has_return_request = any(keyword in text for keyword in ["회수요청", "회수지시", "회수등록", "회수접수", "업체처리요청", "자동수거", "직접수거", "업체측회수", "회수"])
    has_reship_complete = any(keyword in text for keyword in ["재발송완료", "재출고완료", "발송완료", "출고완료", "재발송", "재출고", "배출고"])
    has_all_done = any(keyword in text for keyword in ["전체처리완료", "처리완료", "업체처리완료", "정산반영완료", "사고처리완료", "확인완료", "수령확인완료", "화물추적완료"])

    if has_all_done or (has_reship_complete and has_return_complete):
        return "전체 처리완료"
    if has_reship_complete and has_return_request:
        return f"재발송 완료({date_text})/회수지시({date_text})"
    if has_return_complete:
        return f"회수 완료({date_text})"
    if has_reship_complete or any(keyword in text for keyword in ["직접보내주심", "직접보냄", "업체에서직접보냄"]):
        return f"재발송 완료({date_text})"
    if has_return_request or any(keyword in text for keyword in ["반품", "회수철회", "반품철회"]):
        return f"회수지시({date_text})"
    return f"회수지시({date_text})"


def cs_case_from_payload(payload: dict, status: str = "접수", mail_sent: bool = False) -> dict[str, str]:
    original_info = clean_payload_text(payload, "cs_origin")
    return {
        "status": status,
        "vendor_name": clean_payload_text(payload, "vendor_name"),
        "vendor_email": clean_payload_text(payload, "recipient_email"),
        "original_info": original_info,
        "original_invoice": clean_payload_text(payload, "original_invoice") or extract_invoice_number(original_info),
        "product_name": clean_payload_text(payload, "cs_product"),
        "orderer_name": clean_payload_text(payload, "orderer_name"),
        "orderer_phone": clean_payload_text(payload, "orderer_phone"),
        "receiver_name": clean_payload_text(payload, "cs_receiver"),
        "receiver_phone": clean_payload_text(payload, "cs_phone"),
        "receiver_address": clean_payload_text(payload, "cs_address"),
        "cs_type": clean_payload_text(payload, "cs_type"),
        "cs_content": clean_payload_text(payload, "cs_content"),
        "return_invoice": clean_payload_text(payload, "return_invoice"),
        "reship_invoice": clean_payload_text(payload, "reship_invoice"),
        "mail_subject": clean_payload_text(payload, "subject"),
        "mail_body": clean_payload_text(payload, "body"),
        "mail_sent_at": now_text() if mail_sent else clean_payload_text(payload, "mail_sent_at"),
    }


def save_cs_case(payload: dict, status: str = "접수", mail_sent: bool = False) -> int:
    init_db()
    case = cs_case_from_payload(payload, status=status, mail_sent=mail_sent)
    timestamp = now_text()
    columns = [
        "created_at",
        "updated_at",
        "status",
        "vendor_name",
        "vendor_email",
        "original_info",
        "original_invoice",
        "product_name",
        "orderer_name",
        "orderer_phone",
        "receiver_name",
        "receiver_phone",
        "receiver_address",
        "cs_type",
        "cs_content",
        "return_invoice",
        "reship_invoice",
        "mail_subject",
        "mail_body",
        "mail_sent_at",
    ]
    values = {
        "created_at": timestamp,
        "updated_at": timestamp,
        **case,
    }
    placeholders = ", ".join("?" for _ in columns)
    connection = connect_db()
    try:
        cursor = connection.execute(
            f"INSERT INTO cs_cases ({', '.join(columns)}) VALUES ({placeholders})",
            [values[column] for column in columns],
        )
        connection.commit()
        return int(cursor.lastrowid)
    finally:
        connection.close()


def date_period_condition(fields: list[str], year: str = "", month: str = "") -> tuple[str, list[str]]:
    year = clean_cell(year)
    month = clean_cell(month).zfill(2) if clean_cell(month) else ""
    if not re.fullmatch(r"\d{4}", year):
        year = ""
    if not re.fullmatch(r"\d{2}", month) or not (1 <= int(month) <= 12):
        month = ""
    if not year and not month:
        return "", []
    if year and month:
        pattern = f"{year}-{month}%"
    elif year:
        pattern = f"{year}%"
    else:
        pattern = f"%-{month}-%"
    return "(" + " OR ".join(f"{field} LIKE ?" for field in fields) + ")", [pattern] * len(fields)


def list_cs_cases(query: str = "", status: str = "", limit: int = 20, year: str = "", month: str = "") -> list[dict[str, str | int]]:
    init_db()
    query = query.strip()
    status = status.strip()
    params: list[object] = []
    conditions: list[str] = []
    if query:
        pattern = f"%{query}%"
        conditions.append(
            """
            (
                vendor_name LIKE ?
                OR vendor_email LIKE ?
                OR sales_vendor LIKE ?
                OR purchase_vendor LIKE ?
                OR original_info LIKE ?
                OR original_invoice LIKE ?
                OR product_name LIKE ?
                OR orderer_name LIKE ?
                OR orderer_phone LIKE ?
                OR receiver_name LIKE ?
                OR receiver_phone LIKE ?
                OR receiver_address LIKE ?
                OR cs_type LIKE ?
                OR cs_content LIKE ?
                OR return_invoice LIKE ?
                OR reship_invoice LIKE ?
            )
            """
        )
        params = [pattern] * 16
    if status:
        conditions.append("status = ?")
        params.append(status)
    period_condition, period_params = date_period_condition(
        ["occurred_at", "order_date", "ship_date", "created_at"],
        year,
        month,
    )
    if period_condition:
        conditions.append(period_condition)
        params.extend(period_params)
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    params.append(limit)
    connection = connect_db()
    try:
        rows = connection.execute(
            f"""
            SELECT id, created_at, updated_at, status, vendor_name, vendor_email,
                   original_info, original_invoice, product_name, orderer_name, orderer_phone, receiver_name,
                   receiver_phone, receiver_address, cs_type, cs_content, return_invoice,
                   reship_invoice, mail_subject, mail_sent_at, occurred_at, completed_at, order_date,
                   ship_date, sales_vendor, purchase_vendor, courier, quantity
              FROM cs_cases
              {where}
             ORDER BY id DESC
             LIMIT ?
            """,
            params,
        ).fetchall()
    finally:
        connection.close()
    return [dict(row) for row in rows]


def update_cs_case(case_id: int, payload: dict) -> None:
    status = clean_payload_text(payload, "status")
    cs_type = clean_payload_text(payload, "cs_type")
    return_invoice = clean_payload_text(payload, "return_invoice")
    reship_invoice = clean_payload_text(payload, "reship_invoice")

    init_db()
    connection = connect_db()
    try:
        cursor = connection.execute(
            """
            UPDATE cs_cases
               SET status = COALESCE(NULLIF(?, ''), status),
                   cs_type = ?,
                   return_invoice = ?,
                   reship_invoice = ?,
                   updated_at = ?
             WHERE id = ?
            """,
            [status, cs_type, return_invoice, reship_invoice, now_text(), case_id],
        )
        connection.commit()
        if cursor.rowcount == 0:
            raise ValueError("수정할 CS건을 찾지 못했습니다.")
    finally:
        connection.close()


def delete_cs_cases(case_ids: list[int]) -> int:
    init_db()
    if not case_ids:
        return 0
    placeholders = ", ".join("?" for _ in case_ids)
    connection = connect_db()
    try:
        cursor = connection.execute(f"DELETE FROM cs_cases WHERE id IN ({placeholders})", case_ids)
        connection.commit()
        return int(cursor.rowcount or 0)
    finally:
        connection.close()


def management_query_conditions(query: str = "", year: str = "", month: str = "") -> tuple[str, list[object]]:
    query = query.strip()
    params: list[object] = []
    conditions: list[str] = []
    if query:
        pattern = f"%{query}%"
        conditions.append(
            """
            (
                purchase_vendor LIKE ?
                OR sales_vendor LIKE ?
                OR transaction_type LIKE ?
                OR ledger_checked LIKE ?
                OR order_date LIKE ?
                OR ship_date LIKE ?
                OR orderer_name LIKE ?
                OR sender_phone LIKE ?
                OR receiver_name LIKE ?
                OR receiver_phone LIKE ?
                OR product_name LIKE ?
                OR receiver_address LIKE ?
                OR courier LIKE ?
                OR invoice_number LIKE ?
                OR memo LIKE ?
                OR order_item_id LIKE ?
                OR product_code LIKE ?
                OR order_number LIKE ?
                OR customer_option LIKE ?
            )
            """
        )
        params = [pattern] * 19
    period_condition, period_params = date_period_condition(["order_date", "ship_date"], year, month)
    if period_condition:
        conditions.append(period_condition)
        params.extend(period_params)
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    return where, params

def list_management_records(query: str = "", limit: int | None = 300, year: str = "", month: str = "") -> list[dict[str, str | int]]:
    init_db()
    where, params = management_query_conditions(query, year, month)
    connection = connect_db()
    try:
        limit_sql = ""
        if limit is not None:
            limit_sql = "LIMIT ?"
            params.append(limit)
        rows = connection.execute(
            f"""
            SELECT id, created_at, source_file, source_sheet, source_row, purchase_vendor,
                   sales_vendor, transaction_type, ledger_checked, order_date, ship_date,
                   orderer_name, sender_phone, receiver_name, receiver_phone, product_name,
                   quantity, receiver_address, courier, invoice_number, memo, order_item_id,
                   product_code, order_number, customer_option, cs_received_at
              FROM management_records
              {where}
             ORDER BY order_date DESC, id DESC
             {limit_sql}
            """,
            params,
        ).fetchall()
    finally:
        connection.close()
    return [dict(row) for row in rows]

def list_management_records_by_ids(record_ids: list[int]) -> list[dict[str, str | int]]:
    init_db()
    if not record_ids:
        return []
    placeholders = ", ".join("?" for _ in record_ids)
    order_case = " ".join(f"WHEN {record_id} THEN {index}" for index, record_id in enumerate(record_ids))
    connection = connect_db()
    try:
        rows = connection.execute(
            f"""
            SELECT id, created_at, source_file, source_sheet, source_row, purchase_vendor,
                   sales_vendor, transaction_type, ledger_checked, order_date, ship_date,
                   orderer_name, sender_phone, receiver_name, receiver_phone, product_name,
                   quantity, receiver_address, courier, invoice_number, memo, order_item_id,
                   product_code, order_number, customer_option, cs_received_at
              FROM management_records
             WHERE id IN ({placeholders})
             ORDER BY CASE id {order_case} END
            """,
            record_ids,
        ).fetchall()
    finally:
        connection.close()
    return [dict(row) for row in rows]

def list_management_periods() -> list[dict[str, str | int]]:
    init_db()
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT order_date, ship_date
              FROM management_records
             WHERE COALESCE(order_date, '') <> '' OR COALESCE(ship_date, '') <> ''
            """
        ).fetchall()
    finally:
        connection.close()

    counts: dict[tuple[str, str], int] = {}
    for row in rows:
        year_month = extract_year_month(row["order_date"], row["ship_date"])
        if not year_month:
            continue
        counts[year_month] = counts.get(year_month, 0) + 1
    return [
        {"year": year, "month": month, "count": count}
        for (year, month), count in sorted(counts.items(), key=lambda item: (item[0][0], item[0][1]), reverse=True)
    ]


IMPORT_SHIPMENT_FIELDS = (
    "departure_date",
    "arrival_date",
    "loading_port",
    "arrival_port",
    "shipper",
    "item",
    "quantity",
    "vessel_name",
    "hbl_no",
    "size",
    "progress_status",
    "free_time",
    "warehouse_due_date",
)


def list_import_shipments() -> list[dict[str, str | int]]:
    init_db()
    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT id, created_at, updated_at, departure_date, arrival_date,
                   loading_port, arrival_port, shipper, item, quantity, vessel_name,
                   hbl_no, size, progress_status, free_time, warehouse_due_date, completed_at
              FROM import_shipments
             ORDER BY CASE WHEN completed_at IS NULL OR completed_at = '' THEN 0 ELSE 1 END,
                      id DESC
            """
        ).fetchall()
    finally:
        connection.close()
    return [dict(row) for row in rows]


def save_import_shipment(payload: dict) -> int:
    init_db()
    now = now_text()
    values = {field: clean_payload_text(payload, field) for field in IMPORT_SHIPMENT_FIELDS}
    if not any(values.values()):
        raise ValueError("수입제품 출고 진행 내용을 입력해주세요.")
    shipment_id = int(payload.get("id") or 0)
    connection = connect_db()
    try:
        if shipment_id:
            assignments = ", ".join(f"{field} = ?" for field in IMPORT_SHIPMENT_FIELDS)
            cursor = connection.execute(
                f"UPDATE import_shipments SET {assignments}, updated_at = ? WHERE id = ?",
                [values[field] for field in IMPORT_SHIPMENT_FIELDS] + [now, shipment_id],
            )
            if cursor.rowcount == 0:
                raise ValueError("수정할 수입제품 진행 건을 찾지 못했습니다.")
        else:
            columns = ["created_at", "updated_at", *IMPORT_SHIPMENT_FIELDS]
            placeholders = ", ".join("?" for _ in columns)
            cursor = connection.execute(
                f"INSERT INTO import_shipments ({', '.join(columns)}) VALUES ({placeholders})",
                [now, now] + [values[field] for field in IMPORT_SHIPMENT_FIELDS],
            )
            shipment_id = int(cursor.lastrowid)
        connection.commit()
        return shipment_id
    finally:
        connection.close()


def complete_import_shipment(shipment_id: int) -> None:
    if not shipment_id:
        raise ValueError("완료 처리할 수입제품 진행 건이 없습니다.")
    init_db()
    now = now_text()
    connection = connect_db()
    try:
        cursor = connection.execute(
            """
            UPDATE import_shipments
               SET progress_status = '완료',
                   completed_at = ?,
                   updated_at = ?
             WHERE id = ?
            """,
            (now, now, shipment_id),
        )
        connection.commit()
        if cursor.rowcount == 0:
            raise ValueError("완료 처리할 수입제품 진행 건을 찾지 못했습니다.")
    finally:
        connection.close()


MANAGEMENT_EDIT_COLUMNS = [
    "purchase_vendor",
    "sales_vendor",
    "transaction_type",
    "ledger_checked",
    "order_date",
    "ship_date",
    "orderer_name",
    "sender_phone",
    "receiver_name",
    "receiver_phone",
    "product_name",
    "quantity",
    "receiver_address",
    "courier",
    "invoice_number",
    "memo",
]


def get_management_record(record_id: int) -> dict[str, str | int]:
    init_db()
    connection = connect_db()
    try:
        row = connection.execute(
            """
            SELECT id, created_at, source_file, source_sheet, source_row, purchase_vendor,
                   sales_vendor, transaction_type, ledger_checked, order_date, ship_date,
                   orderer_name, sender_phone, receiver_name, receiver_phone, product_name,
                   quantity, receiver_address, courier, invoice_number, memo, order_item_id,
                   product_code, order_number, customer_option, cs_received_at
              FROM management_records
             WHERE id = ?
            """,
            [record_id],
        ).fetchone()
    finally:
        connection.close()
    if not row:
        raise ValueError("통합관리대장 행을 찾지 못했습니다.")
    return dict(row)

def update_management_record(record_id: int, payload: dict) -> None:
    init_db()
    values = [clean_payload_text(payload, column) for column in MANAGEMENT_EDIT_COLUMNS]
    assignments = ", ".join(f"{column} = ?" for column in MANAGEMENT_EDIT_COLUMNS)
    connection = connect_db()
    try:
        cursor = connection.execute(
            f"UPDATE management_records SET {assignments} WHERE id = ?",
            [*values, record_id],
        )
        connection.commit()
        if cursor.rowcount == 0:
            raise ValueError("수정할 통합관리대장 행을 찾지 못했습니다.")
    finally:
        connection.close()


def delete_management_records(record_ids: list[int]) -> int:
    init_db()
    if not record_ids:
        return 0
    placeholders = ", ".join("?" for _ in record_ids)
    connection = connect_db()
    try:
        cursor = connection.execute(f"DELETE FROM management_records WHERE id IN ({placeholders})", record_ids)
        connection.commit()
        return int(cursor.rowcount or 0)
    finally:
        connection.close()


def create_cs_case_from_management(record_id: int) -> int:
    record = get_management_record(record_id)
    timestamp = now_text()
    source_file = f"통합관리대장:{record.get('source_file', '')}"
    source_sheet = str(record.get("source_sheet", "") or "")
    source_row = int(record.get("source_row", 0) or 0)
    connection = connect_db()
    try:
        cursor = connection.execute(
            """
            INSERT OR IGNORE INTO cs_cases (
                created_at, updated_at, status, vendor_name, original_info, original_invoice,
                product_name, orderer_name, orderer_phone, receiver_name, receiver_phone,
                receiver_address, cs_type, cs_content, source_file, source_sheet, source_row,
                occurred_at, order_date, ship_date, sales_vendor, purchase_vendor, courier, quantity
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                timestamp,
                timestamp,
                "접수",
                record.get("purchase_vendor", ""),
                f"{record.get('ship_date', '')} / {record.get('invoice_number', '')}".strip(" /"),
                record.get("invoice_number", ""),
                record.get("product_name", ""),
                record.get("orderer_name", ""),
                record.get("sender_phone", ""),
                record.get("receiver_name", ""),
                record.get("receiver_phone", ""),
                record.get("receiver_address", ""),
                "",
                "",
                source_file,
                source_sheet,
                source_row,
                record.get("order_date", "") or record.get("ship_date", ""),
                record.get("order_date", ""),
                record.get("ship_date", ""),
                record.get("sales_vendor", ""),
                record.get("purchase_vendor", ""),
                record.get("courier", ""),
                record.get("quantity", ""),
            ],
        )
        if cursor.rowcount:
            case_id = int(cursor.lastrowid)
        else:
            existing = connection.execute(
                "SELECT id FROM cs_cases WHERE source_file = ? AND source_sheet = ? AND source_row = ?",
                [source_file, source_sheet, source_row],
            ).fetchone()
            case_id = int(existing["id"]) if existing else 0
        connection.execute(
            """
            UPDATE management_records
               SET cs_received_at = COALESCE(NULLIF(cs_received_at, ''), ?)
             WHERE id = ?
            """,
            [timestamp, record_id],
        )
        connection.commit()
    finally:
        connection.close()
    if not case_id:
        raise ValueError("CS 처리대장 접수에 실패했습니다.")
    return case_id


def normalized_header(value: object) -> str:
    return re.sub(r"[\s\r\n/]+", "", str(value or "")).strip().lower()


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\r\n", "\n").strip()
    return re.sub(r"[ \t]+", " ", text)


MANAGEMENT_EXPORT_COLUMNS = [
    ("sequence", "순서"),
    ("purchase_vendor", "매입거래처"),
    ("sales_vendor", "매출거래처"),
    ("transaction_type", "거래구분"),
    ("ledger_checked", "장부입력확인"),
    ("order_date", "주문일"),
    ("ship_date", "출고일"),
    ("orderer_name", "주문자"),
    ("sender_phone", "발신자연락처"),
    ("receiver_name", "수령자"),
    ("receiver_phone", "수령자연락처"),
    ("product_name", "제 품 명"),
    ("quantity", "수량"),
    ("receiver_address", "상 세 주 소"),
    ("courier", "택배사**"),
    ("invoice_number", "배송번호"),
    ("memo", "특이(요청)사항"),
    ("order_item_id", "주문상품고유번호"),
    ("product_code", "상품코드"),
    ("order_number", "주문번호"),
    ("customer_option", "고객선택옵션"),
]


MANAGEMENT_TEMPLATE_HEADER_ROW = 1
MANAGEMENT_TEMPLATE_DATA_ROW = 2


LEDGER_EXPORT_COLUMNS = [
    ("occurred_at", "날짜"),
    ("sales_vendor", "매출거래처"),
    ("purchase_vendor", "매입거래처"),
    ("status", "처리진행상태"),
    ("completed_at", "완료일"),
    ("cs_type", "처리내용"),
    ("cs_content", "C/S 내용"),
    ("reship_invoice", "재발송운송장번호"),
    ("return_invoice", "회수운송장번호"),
    ("order_date", "주문일자"),
    ("ship_date", "출고일"),
    ("orderer_name", "주문자"),
    ("orderer_phone", "연락처"),
    ("receiver_name", "수령자"),
    ("receiver_phone", "연락처"),
    ("product_name", "제품명"),
    ("quantity", "수량"),
    ("receiver_address", "상세주소"),
    ("courier", "택배사"),
    ("original_invoice", "송장번호"),
]


def workbook_bytes_from_rows(rows: list[dict], columns: list[tuple[str, str]], sheet_name: str) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    worksheet.append([header for _, header in columns])
    for row in rows:
        worksheet.append([clean_cell(row.get(key, "")) for key, _ in columns])
    for column_cells in worksheet.columns:
        max_length = max((len(clean_cell(cell.value)) for cell in column_cells), default=0)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def extract_year_month(*values: object) -> tuple[str, str] | None:
    for value in values:
        text = clean_cell(value)
        if not text:
            continue
        match = re.search(r"(\d{4})[-./년\s]+(\d{1,2})", text)
        if match:
            return match.group(1), f"{int(match.group(2)):02d}"
    return None


def grouped_management_rows(rows: list[dict]) -> list[tuple[str, list[dict]]]:
    groups: dict[str, list[dict]] = {}
    for row in rows:
        year_month = extract_year_month(row.get("order_date"), row.get("ship_date"))
        key = f"{year_month[0]}년 {year_month[1]}월" if year_month else "다운로드"
        groups.setdefault(key, []).append(row)
    return sorted(groups.items(), key=lambda item: item[0])


def cell_style_snapshot(cell) -> dict:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "number_format": cell.number_format,
    }


def apply_cell_style(cell, style: dict) -> None:
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.protection = copy(style["protection"])
    cell.number_format = style["number_format"]


def clear_management_template_sheet(worksheet, style_row: list[dict], row_height: float | None) -> None:
    if worksheet.max_row >= MANAGEMENT_TEMPLATE_DATA_ROW:
        worksheet.delete_rows(MANAGEMENT_TEMPLATE_DATA_ROW, worksheet.max_row - MANAGEMENT_TEMPLATE_DATA_ROW + 1)
    for column_index, (_, header) in enumerate(MANAGEMENT_EXPORT_COLUMNS, start=1):
        worksheet.cell(MANAGEMENT_TEMPLATE_HEADER_ROW, column_index, header)
    if row_height:
        worksheet.row_dimensions[MANAGEMENT_TEMPLATE_DATA_ROW].height = row_height
    for column_index, style in enumerate(style_row, start=1):
        apply_cell_style(worksheet.cell(MANAGEMENT_TEMPLATE_DATA_ROW, column_index), style)

def populate_management_template_sheet(worksheet, title: str, rows: list[dict], style_row: list[dict], row_height: float | None) -> None:
    worksheet.title = title[:31]
    for sequence, row in enumerate(rows, start=1):
        row_offset = MANAGEMENT_TEMPLATE_DATA_ROW + sequence - 1
        if row_height:
            worksheet.row_dimensions[row_offset].height = row_height
        for column_index, (key, _) in enumerate(MANAGEMENT_EXPORT_COLUMNS, start=1):
            cell = worksheet.cell(row_offset, column_index, management_export_cell_value(row, key, sequence))
            apply_cell_style(cell, style_row[column_index - 1])
    last_row = max(MANAGEMENT_TEMPLATE_HEADER_ROW, len(rows) + MANAGEMENT_TEMPLATE_DATA_ROW - 1)
    last_column = get_column_letter(len(MANAGEMENT_EXPORT_COLUMNS))
    worksheet.auto_filter.ref = f"A{MANAGEMENT_TEMPLATE_HEADER_ROW}:{last_column}{last_row}"
    worksheet.freeze_panes = f"A{MANAGEMENT_TEMPLATE_DATA_ROW}"

def management_export_cell_value(row: dict, key: str, sequence: int) -> str:
    if key == "sequence":
        return clean_cell(row.get(key, "")) or str(sequence)
    return clean_cell(row.get(key, ""))

def management_year_from_rows(rows: list[dict], fallback: str = "") -> str:
    for row in rows:
        year_month = extract_year_month(row.get("order_date"), row.get("ship_date"))
        if year_month:
            return year_month[0]
    return fallback

def management_template_filename_stem(
    rows: list[dict],
    payload: dict,
    now: datetime | None = None,
) -> str:
    now = now or datetime.now()
    requested_year = clean_payload_text(payload, "year")
    year = requested_year if re.fullmatch(r"\d{4}", requested_year) else management_year_from_rows(rows, str(now.year))
    return f"통합관리대장 양식 {year}년 {now.strftime('%Y%m%d')}"

def management_export_filename(filename_stem: str, payload: dict) -> str:
    scope = clean_payload_text(payload, "scope")
    if scope == "template":
        return f"{filename_stem}.xlsx"
    return f"{filename_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def management_workbook_bytes_from_template(rows: list[dict]) -> bytes:
    if not MANAGEMENT_EXPORT_TEMPLATE.exists():
        return workbook_bytes_from_rows(rows, MANAGEMENT_EXPORT_COLUMNS, "통합관리대장")
    workbook = load_workbook(MANAGEMENT_EXPORT_TEMPLATE)
    base_sheet = workbook.worksheets[0]
    max_columns = len(MANAGEMENT_EXPORT_COLUMNS)
    style_row = [
        cell_style_snapshot(base_sheet.cell(MANAGEMENT_TEMPLATE_DATA_ROW, column_index))
        for column_index in range(1, max_columns + 1)
    ]
    row_height = base_sheet.row_dimensions[MANAGEMENT_TEMPLATE_DATA_ROW].height
    for sheet in list(workbook.worksheets)[1:]:
        workbook.remove(sheet)
    clear_management_template_sheet(base_sheet, style_row, row_height)
    groups = grouped_management_rows(rows) or [("다운로드", rows)]
    sheets = [base_sheet]
    for _title, _rows in groups[1:]:
        sheets.append(workbook.copy_worksheet(base_sheet))
    for worksheet, (title, group_rows) in zip(sheets, groups, strict=False):
        populate_management_template_sheet(worksheet, title, group_rows, style_row, row_height)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()

def management_export_rows_from_payload(payload: dict) -> tuple[list[dict], str]:
    scope = clean_payload_text(payload, "scope") or "selected"
    if scope == "selected":
        rows = payload.get("rows", [])
        if isinstance(rows, list) and rows:
            return rows, "통합관리대장_선택"
        record_ids = [int(value) for value in payload.get("ids", []) if str(value).isdigit()]
        rows = list_management_records_by_ids(record_ids)
        return rows, "통합관리대장_선택"
    if scope == "month":
        year = clean_payload_text(payload, "year")
        month = clean_payload_text(payload, "month").zfill(2)
        if not re.fullmatch(r"\d{4}", year) or not re.fullmatch(r"\d{2}", month) or not (1 <= int(month) <= 12):
            raise ValueError("월별 다운로드는 년도와 월을 선택해주세요.")
        rows = list_management_records(limit=None, year=year, month=month)
        return rows, f"통합관리대장_{year}년_{int(month)}월"
    if scope == "year":
        year = clean_payload_text(payload, "year")
        if not re.fullmatch(r"\d{4}", year):
            raise ValueError("년별 다운로드는 년도를 선택해주세요.")
        rows = list_management_records(limit=None, year=year)
        return rows, f"통합관리대장_{year}년"
    if scope == "all":
        rows = list_management_records(limit=None)
        return rows, "통합관리대장_전체"
    rows = payload.get("rows", [])
    return rows if isinstance(rows, list) else [], "통합관리대장"


def ledger_export_rows_from_payload(payload: dict) -> tuple[list[dict], str]:
    scope = clean_payload_text(payload, "scope") or "selected"
    if scope == "selected":
        rows = payload.get("rows", [])
        return rows if isinstance(rows, list) else [], "CS처리대장_선택"
    if scope == "month":
        year = clean_payload_text(payload, "year")
        month = clean_payload_text(payload, "month").zfill(2)
        if not re.fullmatch(r"\d{4}", year) or not re.fullmatch(r"\d{2}", month) or not (1 <= int(month) <= 12):
            raise ValueError("월별 다운로드는 년도와 월을 선택해주세요.")
        rows = list_cs_cases(limit=50000, year=year, month=month)
        return rows, f"CS처리대장_{year}년_{int(month)}월"
    if scope == "year":
        year = clean_payload_text(payload, "year")
        if not re.fullmatch(r"\d{4}", year):
            raise ValueError("년별 다운로드는 년도를 선택해주세요.")
        rows = list_cs_cases(limit=50000, year=year)
        return rows, f"CS처리대장_{year}년"
    if scope == "all":
        rows = list_cs_cases(limit=50000)
        return rows, "CS처리대장_전체"
    rows = payload.get("rows", [])
    return rows if isinstance(rows, list) else [], "CS처리대장"


def find_header(headers: list[str], names: set[str]) -> int | None:
    for idx, header in enumerate(headers):
        if header in names:
            return idx
    return None


def find_header_contains(headers: list[str], *parts: str) -> int | None:
    for idx, header in enumerate(headers):
        if all(part in header for part in parts):
            return idx
    return None


def row_value(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return clean_cell(row[idx])


def contact_index_after(headers: list[str], anchor_idx: int | None, before_idx: int | None = None) -> int | None:
    contact_indexes = [idx for idx, header in enumerate(headers) if "연락처" in header]
    if not contact_indexes:
        return None
    if anchor_idx is not None:
        for idx in contact_indexes:
            if idx > anchor_idx and (before_idx is None or idx < before_idx):
                return idx
    return None


def receiver_phone_index(headers: list[str], receiver_idx: int | None) -> int | None:
    contact_indexes = [idx for idx, header in enumerate(headers) if "연락처" in header]
    if not contact_indexes:
        return None
    if receiver_idx is not None:
        for idx in contact_indexes:
            if idx > receiver_idx:
                return idx
    return contact_indexes[-1]


def management_header_indexes(headers: list[str]) -> dict[str, int | None]:
    return {
        "purchase_vendor": find_header(headers, {"매입거래처"}),
        "sales_vendor": find_header(headers, {"매출거래처"}),
        "transaction_type": find_header(headers, {"거래구분"}),
        "ledger_checked": find_header(headers, {"장부입력확인"}),
        "order_date": find_header(headers, {"주문일자", "주문일"}),
        "ship_date": find_header(headers, {"출고일"}),
        "orderer_name": find_header(headers, {"주문자"}),
        "sender_phone": find_header(headers, {"발신자연락처", "주문자연락처"}),
        "receiver_name": find_header(headers, {"수령자", "수취인", "수하인"}),
        "receiver_phone": find_header(headers, {"수령자연락처", "수취인연락처", "수하인연락처"}),
        "product_name": find_header(headers, {"제품명", "상품명", "품명"}),
        "quantity": find_header(headers, {"수량"}),
        "receiver_address": find_header(headers, {"상세주소", "주소"}),
        "courier": find_header_contains(headers, "택배사"),
        "invoice_number": find_header(headers, {"운송장번호", "송장번호", "배송번호"}),
        "memo": find_header(headers, {"특이사항", "특이(요청)사항", "배송메세지", "배송메시지", "비고"}),
        "order_item_id": find_header(headers, {"주문상품고유번호"}),
        "product_code": find_header(headers, {"상품코드"}),
        "order_number": find_header(headers, {"주문번호"}),
        "customer_option": find_header(headers, {"고객선택옵션"}),
    }

def find_management_header_row(worksheet, max_scan_rows: int = 10) -> tuple[int, dict[str, int | None]] | None:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_rows, max_col=80, values_only=True),
        start=1,
    ):
        headers = [normalized_header(value) for value in row]
        indexes = management_header_indexes(headers)
        if indexes["receiver_name"] is not None and indexes["product_name"] is not None:
            return row_number, indexes
    return None

MANAGEMENT_IMPORT_COLUMNS = [
    "created_at",
    "source_file",
    "source_sheet",
    "source_row",
    "purchase_vendor",
    "sales_vendor",
    "transaction_type",
    "ledger_checked",
    "order_date",
    "ship_date",
    "orderer_name",
    "sender_phone",
    "receiver_name",
    "receiver_phone",
    "product_name",
    "quantity",
    "receiver_address",
    "courier",
    "invoice_number",
    "memo",
    "order_item_id",
    "product_code",
    "order_number",
    "customer_option",
]
MANAGEMENT_IMPORT_EDITABLE_FIELDS = set(MANAGEMENT_IMPORT_COLUMNS) - {"created_at", "source_file", "source_sheet", "source_row"}


def normalize_duplicate_token(value: object) -> str:
    text = clean_cell(value)
    return re.sub(r"[\s\-()._/]+", "", text).lower()


def compact_duplicate_key(prefix: str, values: list[object]) -> str:
    tokens = [normalize_duplicate_token(value) for value in values]
    return f"{prefix}:" + "|".join(tokens)


def management_duplicate_key(record: dict[str, object]) -> str:
    order_item_id = normalize_duplicate_token(record.get("order_item_id"))
    if order_item_id:
        return f"management-order-item:{order_item_id}"
    order_number = record.get("order_number")
    invoice_number = record.get("invoice_number")
    product_identity = record.get("product_code") or record.get("product_name")
    if normalize_duplicate_token(order_number) or normalize_duplicate_token(invoice_number):
        return compact_duplicate_key(
            "management-order",
            [
                order_number,
                invoice_number,
                product_identity,
                record.get("quantity"),
                record.get("receiver_name"),
                record.get("receiver_phone"),
                record.get("receiver_address"),
            ],
        )
    return compact_duplicate_key(
        "management-fallback",
        [
            record.get("order_date"),
            record.get("ship_date"),
            record.get("product_name"),
            record.get("quantity"),
            record.get("receiver_name"),
            record.get("receiver_phone"),
            record.get("receiver_address"),
        ],
    )


def cs_duplicate_key(record: dict[str, object]) -> str:
    invoice = extract_invoice_number(record.get("original_invoice")) or record.get("original_invoice")
    return compact_duplicate_key(
        "cs-case",
        [
            record.get("occurred_at"),
            invoice,
            record.get("cs_type"),
            record.get("cs_content"),
            record.get("receiver_name"),
            record.get("receiver_phone"),
            record.get("product_name"),
        ],
    )


def duplicate_summary(record: dict[str, object], fields: list[str]) -> str:
    values = [clean_cell(record.get(field)) for field in fields]
    return " / ".join(value for value in values if value)[:160]


def digit_count(value: object) -> int:
    return len(re.sub(r"\D+", "", clean_cell(value)))


def is_quantity_text(value: object) -> bool:
    text = clean_cell(value)
    if not text:
        return False
    if not re.fullmatch(r"\d+(\.\d+)?", text):
        return False
    try:
        return float(text) > 0
    except ValueError:
        return False


def is_numeric_identifier(value: object, minimum_digits: int = 5) -> bool:
    text = clean_cell(value)
    if not text:
        return False
    return digit_count(text) >= minimum_digits and not re.search(r"[가-힣a-zA-Z]", text)


def import_issue(field: str, label: str, message: str, input_type: str = "text") -> dict[str, str]:
    return {
        "field": field,
        "label": label,
        "message": message,
        "input_type": input_type,
    }


def require_text_issue(record: dict[str, object], field: str, label: str) -> dict[str, str] | None:
    if clean_cell(record.get(field)):
        return None
    return import_issue(field, label, f"{label} 값을 입력해주세요.", "text")


def quantity_issue(record: dict[str, object], field: str = "quantity", label: str = "수량") -> dict[str, str] | None:
    if is_quantity_text(record.get(field)):
        return None
    return import_issue(field, label, f"{label}은 숫자로 입력해주세요.", "number")


def optional_numeric_issue(record: dict[str, object], field: str, label: str, minimum_digits: int = 5) -> dict[str, str] | None:
    value = clean_cell(record.get(field))
    if not value or is_numeric_identifier(value, minimum_digits):
        return None
    return import_issue(field, label, f"{label} 형식이 숫자와 맞지 않습니다.", "text")


def required_numeric_issue(record: dict[str, object], field: str, label: str, minimum_digits: int = 5) -> dict[str, str] | None:
    if is_numeric_identifier(record.get(field), minimum_digits):
        return None
    return import_issue(field, label, f"{label}을 숫자 형식으로 입력해주세요.", "text")


def management_import_issues(record: dict[str, object]) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    for field, label in (
        ("receiver_name", "수령자"),
        ("product_name", "제품명"),
        ("receiver_address", "상세주소"),
    ):
        issue = require_text_issue(record, field, label)
        if issue:
            issues.append(issue)
    issue = quantity_issue(record)
    if issue:
        issues.append(issue)
    for field, label, digits in (
        ("receiver_phone", "수령자연락처", 7),
        ("sender_phone", "발신자연락처", 7),
        ("invoice_number", "운송장번호", 5),
    ):
        issue = optional_numeric_issue(record, field, label, digits)
        if issue:
            issues.append(issue)
    return issues


def cs_import_issues(record: dict[str, object]) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    for field, label in (
        ("receiver_name", "수령자"),
        ("product_name", "제품명"),
        ("cs_content", "CS내용"),
    ):
        issue = require_text_issue(record, field, label)
        if issue:
            issues.append(issue)
    for issue in (
        required_numeric_issue(record, "original_invoice", "기존운송장번호", 5),
        required_numeric_issue(record, "receiver_phone", "수령자연락처", 7),
        quantity_issue(record),
    ):
        if issue:
            issues.append(issue)
    return issues


def import_row_identity(record: dict[str, object]) -> tuple[str, int]:
    try:
        row_number = int(record.get("source_row") or 0)
    except (TypeError, ValueError):
        row_number = 0
    return clean_cell(record.get("source_sheet")), row_number


def apply_import_corrections(records: list[dict[str, object]], corrections: list[dict[str, object]] | None, allowed_fields: set[str]) -> None:
    if not corrections:
        return
    correction_map: dict[tuple[str, int], dict[str, object]] = {}
    for correction in corrections:
        if not isinstance(correction, dict):
            continue
        key = import_row_identity(correction)
        if not key[0] or not key[1]:
            continue
        correction_map[key] = correction
    for record in records:
        correction = correction_map.get(import_row_identity(record))
        if not correction:
            continue
        for field in allowed_fields:
            if field in correction:
                record[field] = clean_cell(correction.get(field))


def invalid_import_rows(records: list[dict[str, object]], issue_func, summary_fields: list[str]) -> list[dict[str, object]]:
    invalid_rows = []
    for record in records:
        issues = issue_func(record)
        if not issues:
            continue
        invalid_rows.append({
            "source_sheet": record.get("source_sheet", ""),
            "source_row": record.get("source_row", ""),
            "row": record.get("source_row", ""),
            "summary": duplicate_summary(record, summary_fields),
            "record": {key: clean_cell(value) for key, value in record.items() if key not in {"created_at", "updated_at"}},
            "issues": issues,
        })
    return invalid_rows


def valid_import_records(records: list[dict[str, object]], issue_func) -> list[dict[str, object]]:
    return [record for record in records if not issue_func(record)]


def import_preview_payload(records: list[dict[str, object]], existing_keys: set[str], key_func, summary_fields: list[str], issue_func=None) -> dict[str, object]:
    invalid_rows = invalid_import_rows(records, issue_func, summary_fields) if issue_func else []
    records_for_duplicate_check = valid_import_records(records, issue_func) if issue_func else records
    seen: set[str] = set()
    duplicates: list[dict[str, object]] = []
    duplicate_existing = 0
    duplicate_in_file = 0
    insertable = 0
    for record in records_for_duplicate_check:
        key = key_func(record)
        if key in existing_keys:
            duplicate_existing += 1
            if len(duplicates) < 20:
                duplicates.append({
                    "row": record.get("source_row", ""),
                    "reason": "이미 DB에 있는 데이터",
                    "summary": duplicate_summary(record, summary_fields),
                })
            continue
        if key in seen:
            duplicate_in_file += 1
            if len(duplicates) < 20:
                duplicates.append({
                    "row": record.get("source_row", ""),
                    "reason": "업로드 파일 안의 중복 데이터",
                    "summary": duplicate_summary(record, summary_fields),
                })
            continue
        seen.add(key)
        insertable += 1
    skipped = duplicate_existing + duplicate_in_file
    return {
        "total": len(records),
        "insertable": insertable,
        "invalid_count": len(invalid_rows),
        "invalid_rows": invalid_rows,
        "duplicate_existing": duplicate_existing,
        "duplicate_in_file": duplicate_in_file,
        "skipped": skipped,
        "duplicates": duplicates,
        "has_duplicates": skipped > 0,
        "has_invalid_rows": len(invalid_rows) > 0,
    }


def parse_management_import_records(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    timestamp = now_text()
    source_file = original_uploaded_filename(path.name)
    records: list[dict[str, object]] = []
    try:
        for worksheet in workbook.worksheets:
            header_match = find_management_header_row(worksheet)
            if header_match is None:
                continue
            header_row_number, indexes = header_match
            rows = worksheet.iter_rows(min_row=header_row_number + 1, max_col=80, values_only=True)
            for excel_row_number, row in enumerate(rows, start=header_row_number + 1):
                record = {
                    "created_at": timestamp,
                    "source_file": source_file,
                    "source_sheet": worksheet.title,
                    "source_row": excel_row_number,
                    "purchase_vendor": row_value(row, indexes["purchase_vendor"]),
                    "sales_vendor": row_value(row, indexes["sales_vendor"]),
                    "transaction_type": row_value(row, indexes["transaction_type"]),
                    "ledger_checked": row_value(row, indexes["ledger_checked"]),
                    "order_date": row_value(row, indexes["order_date"]),
                    "ship_date": row_value(row, indexes["ship_date"]),
                    "orderer_name": row_value(row, indexes["orderer_name"]),
                    "sender_phone": row_value(row, indexes["sender_phone"]),
                    "receiver_name": row_value(row, indexes["receiver_name"]),
                    "receiver_phone": row_value(row, indexes["receiver_phone"]),
                    "product_name": row_value(row, indexes["product_name"]),
                    "quantity": row_value(row, indexes["quantity"]),
                    "receiver_address": row_value(row, indexes["receiver_address"]),
                    "courier": row_value(row, indexes["courier"]),
                    "invoice_number": row_value(row, indexes["invoice_number"]),
                    "memo": row_value(row, indexes["memo"]),
                    "order_item_id": row_value(row, indexes["order_item_id"]),
                    "product_code": row_value(row, indexes["product_code"]),
                    "order_number": row_value(row, indexes["order_number"]),
                    "customer_option": row_value(row, indexes["customer_option"]),
                }
                if not any(record[key] for key in ("receiver_name", "product_name", "invoice_number", "receiver_address")):
                    continue
                records.append(record)
    finally:
        workbook.close()

    return records


def existing_management_duplicate_keys(connection: sqlite3.Connection) -> set[str]:
    rows = connection.execute(
        """
        SELECT order_date, ship_date, receiver_name, receiver_phone, product_name,
               quantity, receiver_address, invoice_number, order_item_id,
               product_code, order_number
          FROM management_records
        """
    ).fetchall()
    return {management_duplicate_key(dict(row)) for row in rows}


def preview_management_import(path: Path) -> dict[str, object]:
    init_db()
    records = parse_management_import_records(path)
    connection = connect_db()
    try:
        existing_keys = existing_management_duplicate_keys(connection)
    finally:
        connection.close()
    return import_preview_payload(
        records,
        existing_keys,
        management_duplicate_key,
        ["receiver_name", "receiver_phone", "product_name", "quantity", "invoice_number", "order_number"],
        management_import_issues,
    )


def import_management_workbook(path: Path, mode: str = "daily", corrections: list[dict[str, object]] | None = None) -> tuple[int, int]:
    init_db()
    records = parse_management_import_records(path)
    apply_import_corrections(records, corrections, MANAGEMENT_IMPORT_EDITABLE_FIELDS)
    placeholders = ", ".join("?" for _ in MANAGEMENT_IMPORT_COLUMNS)
    insert_sql = f"INSERT OR IGNORE INTO management_records ({', '.join(MANAGEMENT_IMPORT_COLUMNS)}) VALUES ({placeholders})"
    connection = connect_db()
    inserted = 0
    skipped = 0
    try:
        valid_records = valid_import_records(records, management_import_issues)
        skipped += len(records) - len(valid_records)
        if mode == "replace":
            connection.execute("DELETE FROM management_records")
            import_records = valid_records
        else:
            existing_keys = existing_management_duplicate_keys(connection)
            seen_keys: set[str] = set()
            import_records = []
            for record in valid_records:
                key = management_duplicate_key(record)
                if key in existing_keys or key in seen_keys:
                    skipped += 1
                    continue
                seen_keys.add(key)
                import_records.append(record)
        for record in import_records:
            cursor = connection.execute(
                insert_sql,
                [record[column] for column in MANAGEMENT_IMPORT_COLUMNS],
            )
            if cursor.rowcount:
                inserted += 1
            else:
                skipped += 1
        connection.commit()
    finally:
        connection.close()
    return inserted, skipped


def parse_cs_case_import_records(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    timestamp = now_text()
    source_file = path.name
    records: list[dict[str, object]] = []
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(max_col=80, values_only=True)
            title_row = next(rows, None)
            header_row = next(rows, None)
            if not header_row:
                continue

            headers = [normalized_header(value) for value in header_row]
            date_idx = find_header(headers, {"발생일", "날짜"})
            sales_idx = find_header(headers, {"발생거래처", "매출거래처"})
            purchase_idx = find_header(headers, {"처리거래처", "매입거래처"})
            status_idx = find_header_contains(headers, "처리", "상태")
            completed_idx = find_header(headers, {"처리완료일", "완료일"})
            order_idx = find_header(headers, {"주문일자"})
            ship_idx = find_header(headers, {"출고일"})
            orderer_idx = find_header(headers, {"주문자", "주문인", "구매자"})
            receiver_idx = find_header(headers, {"수령자", "수취인", "수하인", "받는분", "받으시는분"})
            orderer_phone_idx = contact_index_after(headers, orderer_idx, receiver_idx)
            phone_idx = receiver_phone_index(headers, receiver_idx)
            product_idx = find_header(headers, {"제품명", "상품명", "품명"})
            quantity_idx = find_header(headers, {"수량"})
            address_idx = find_header_contains(headers, "상세", "주소")
            courier_idx = find_header(headers, {"택배사"})
            original_idx = find_header(headers, {"기존운송장번호", "송장번호"})
            request_idx = find_header(headers, {"요구내용", "처리내용"})
            cs_content_idx = find_header(headers, {"cs내용", "c/s내용"})
            return_idx = find_header(headers, {"회수운송장번호"})
            reship_idx = find_header(headers, {"재발송운송장번호", "재발송송장번호"})
            if address_idx is None and quantity_idx is not None and courier_idx is not None and quantity_idx + 1 < courier_idx:
                address_idx = quantity_idx + 1

            for excel_row_number, row in enumerate(rows, start=3):
                row = tuple(row)
                if not any(clean_cell(value) for value in row):
                    continue

                sales_vendor = row_value(row, sales_idx)
                purchase_vendor = row_value(row, purchase_idx)
                vendor_name = purchase_vendor or sales_vendor
                original_invoice = row_value(row, original_idx)
                original_info = " / ".join(
                    value for value in [row_value(row, ship_idx), original_invoice] if value
                )
                request_text = row_value(row, request_idx)
                cs_text = row_value(row, cs_content_idx)
                cs_content = cs_text or request_text
                raw_status = row_value(row, status_idx)
                cs_type = normalize_cs_type_value(request_text, cs_content, raw_status)
                product_name = row_value(row, product_idx)
                orderer_name = row_value(row, orderer_idx)
                orderer_phone = row_value(row, orderer_phone_idx)
                receiver_name = row_value(row, receiver_idx)
                receiver_phone = row_value(row, phone_idx)

                if not any([vendor_name, original_invoice, product_name, receiver_name, receiver_phone, cs_content]):
                    continue

                records.append({
                    "created_at": timestamp,
                    "updated_at": timestamp,
                    "status": normalize_progress_status(raw_status, row_value(row, completed_idx), row_value(row, date_idx)),
                    "vendor_name": vendor_name,
                    "vendor_email": "",
                    "original_info": original_info,
                    "original_invoice": extract_invoice_number(original_invoice) or original_invoice,
                    "product_name": product_name,
                    "orderer_name": orderer_name,
                    "orderer_phone": orderer_phone,
                    "receiver_name": receiver_name,
                    "receiver_phone": receiver_phone,
                    "receiver_address": row_value(row, address_idx),
                    "cs_type": cs_type,
                    "cs_content": cs_content,
                    "return_invoice": row_value(row, return_idx),
                    "reship_invoice": row_value(row, reship_idx),
                    "mail_subject": "",
                    "mail_body": "",
                    "mail_sent_at": "",
                    "source_file": source_file,
                    "source_sheet": worksheet.title,
                    "source_row": excel_row_number,
                    "occurred_at": row_value(row, date_idx),
                    "completed_at": row_value(row, completed_idx),
                    "order_date": row_value(row, order_idx),
                    "ship_date": row_value(row, ship_idx),
                    "sales_vendor": sales_vendor,
                    "purchase_vendor": purchase_vendor,
                    "courier": row_value(row, courier_idx),
                    "quantity": row_value(row, quantity_idx),
                })
    finally:
        workbook.close()

    return records


CS_IMPORT_COLUMNS = [
    "created_at",
    "updated_at",
    "status",
    "vendor_name",
    "vendor_email",
    "original_info",
    "original_invoice",
    "product_name",
    "orderer_name",
    "orderer_phone",
    "receiver_name",
    "receiver_phone",
    "receiver_address",
    "cs_type",
    "cs_content",
    "return_invoice",
    "reship_invoice",
    "mail_subject",
    "mail_body",
    "mail_sent_at",
    "source_file",
    "source_sheet",
    "source_row",
    "occurred_at",
    "completed_at",
    "order_date",
    "ship_date",
    "sales_vendor",
    "purchase_vendor",
    "courier",
    "quantity",
]
CS_IMPORT_EDITABLE_FIELDS = set(CS_IMPORT_COLUMNS) - {"created_at", "updated_at", "source_file", "source_sheet", "source_row"}


def existing_cs_duplicate_keys(connection: sqlite3.Connection) -> set[str]:
    rows = connection.execute(
        """
        SELECT occurred_at, original_invoice, receiver_name, receiver_phone,
               product_name, cs_type, cs_content
          FROM cs_cases
        """
    ).fetchall()
    return {cs_duplicate_key(dict(row)) for row in rows}


def preview_cs_cases_import(path: Path) -> dict[str, object]:
    init_db()
    records = parse_cs_case_import_records(path)
    connection = connect_db()
    try:
        existing_keys = existing_cs_duplicate_keys(connection)
    finally:
        connection.close()
    return import_preview_payload(
        records,
        existing_keys,
        cs_duplicate_key,
        ["occurred_at", "receiver_name", "product_name", "original_invoice", "cs_type", "cs_content"],
        cs_import_issues,
    )


def import_cs_cases_from_workbook(path: Path, mode: str = "daily", corrections: list[dict[str, object]] | None = None) -> tuple[int, int]:
    init_db()
    records = parse_cs_case_import_records(path)
    apply_import_corrections(records, corrections, CS_IMPORT_EDITABLE_FIELDS)
    placeholders = ", ".join("?" for _ in CS_IMPORT_COLUMNS)
    insert_sql = f"""
        INSERT OR IGNORE INTO cs_cases ({', '.join(CS_IMPORT_COLUMNS)})
        VALUES ({placeholders})
    """
    connection = connect_db()
    inserted = 0
    skipped = 0
    try:
        valid_records = valid_import_records(records, cs_import_issues)
        skipped += len(records) - len(valid_records)
        if mode == "replace":
            connection.execute("DELETE FROM cs_cases")
            import_records = valid_records
        else:
            existing_keys = existing_cs_duplicate_keys(connection)
            seen_keys: set[str] = set()
            import_records = []
            for record in valid_records:
                key = cs_duplicate_key(record)
                if key in existing_keys or key in seen_keys:
                    skipped += 1
                    continue
                seen_keys.add(key)
                import_records.append(record)
        for record in import_records:
            cursor = connection.execute(insert_sql, [record[column] for column in CS_IMPORT_COLUMNS])
            if cursor.rowcount:
                inserted += 1
            else:
                skipped += 1
        connection.commit()
    finally:
        connection.close()
    return inserted, skipped


def parse_import_corrections(fields: dict[str, tuple[str, bytes] | str]) -> list[dict[str, object]] | None:
    raw = fields.get("corrections")
    if not isinstance(raw, str) or not raw.strip():
        return None
    parsed = json.loads(raw)
    if not isinstance(parsed, list):
        raise ValueError("수정 데이터 형식이 올바르지 않습니다.")
    return [item for item in parsed if isinstance(item, dict)]


class DataBlob(ctypes.Structure):
    _fields_ = [
        ("cbData", ctypes.c_ulong),
        ("pbData", ctypes.POINTER(ctypes.c_char)),
    ]


def protect_text(value: str) -> str:
    if not value:
        return ""
    if os.name != "nt":
        return TOKEN_PREFIX_KEY + protect_text_with_secret(value)
    raw = value.encode("utf-8")
    in_buffer = ctypes.create_string_buffer(raw)
    in_blob = DataBlob(len(raw), ctypes.cast(in_buffer, ctypes.POINTER(ctypes.c_char)))
    out_blob = DataBlob()
    if not ctypes.windll.crypt32.CryptProtectData(
        ctypes.byref(in_blob),
        None,
        None,
        None,
        None,
        0,
        ctypes.byref(out_blob),
    ):
        raise OSError("메일 비밀번호 저장 암호화에 실패했습니다.")
    try:
        encrypted = ctypes.string_at(out_blob.pbData, out_blob.cbData)
        return TOKEN_PREFIX_DPAPI + base64.b64encode(encrypted).decode("ascii")
    finally:
        ctypes.windll.kernel32.LocalFree(out_blob.pbData)


def unprotect_text(value: str) -> str:
    if not value:
        return ""
    if value.startswith(TOKEN_PREFIX_KEY):
        return unprotect_text_with_secret(value.removeprefix(TOKEN_PREFIX_KEY))
    if value.startswith(TOKEN_PREFIX_DPAPI):
        value = value.removeprefix(TOKEN_PREFIX_DPAPI)
    if os.name != "nt":
        return unprotect_text_with_secret(value)
    raw = base64.b64decode(value.encode("ascii"))
    in_buffer = ctypes.create_string_buffer(raw)
    in_blob = DataBlob(len(raw), ctypes.cast(in_buffer, ctypes.POINTER(ctypes.c_char)))
    out_blob = DataBlob()
    if not ctypes.windll.crypt32.CryptUnprotectData(
        ctypes.byref(in_blob),
        None,
        None,
        None,
        None,
        0,
        ctypes.byref(out_blob),
    ):
        raise OSError("저장된 메일 비밀번호를 읽지 못했습니다.")
    try:
        decrypted = ctypes.string_at(out_blob.pbData, out_blob.cbData)
        return decrypted.decode("utf-8")
    finally:
        ctypes.windll.kernel32.LocalFree(out_blob.pbData)


def get_server_secret() -> bytes:
    env_secret = os.environ.get("WORKHUB_SECRET_KEY", "").strip()
    if env_secret:
        return hashlib.sha256(env_secret.encode("utf-8")).digest()

    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    if SECRET_KEY_PATH.exists():
        return base64.b64decode(SECRET_KEY_PATH.read_text(encoding="utf-8").strip())

    secret = secrets.token_bytes(32)
    SECRET_KEY_PATH.write_text(base64.b64encode(secret).decode("ascii"), encoding="utf-8")
    try:
        os.chmod(SECRET_KEY_PATH, 0o600)
    except OSError:
        pass
    return secret


def key_stream(secret: bytes, nonce: bytes, size: int) -> bytes:
    chunks: list[bytes] = []
    counter = 0
    while sum(len(chunk) for chunk in chunks) < size:
        chunks.append(hashlib.sha256(secret + nonce + counter.to_bytes(4, "big")).digest())
        counter += 1
    return b"".join(chunks)[:size]


def protect_text_with_secret(value: str) -> str:
    secret = get_server_secret()
    nonce = secrets.token_bytes(16)
    raw = value.encode("utf-8")
    encrypted = bytes(left ^ right for left, right in zip(raw, key_stream(secret, nonce, len(raw))))
    signature = hmac.new(secret, nonce + encrypted, hashlib.sha256).digest()
    return base64.b64encode(nonce + signature + encrypted).decode("ascii")


def unprotect_text_with_secret(value: str) -> str:
    secret = get_server_secret()
    raw = base64.b64decode(value.encode("ascii"))
    nonce = raw[:16]
    signature = raw[16:48]
    encrypted = raw[48:]
    expected = hmac.new(secret, nonce + encrypted, hashlib.sha256).digest()
    if not hmac.compare_digest(signature, expected):
        raise OSError("저장된 메일 비밀번호 검증에 실패했습니다.")
    decrypted = bytes(left ^ right for left, right in zip(encrypted, key_stream(secret, nonce, len(encrypted))))
    return decrypted.decode("utf-8")


def normalize_naver_email(value: object) -> str:
    email = str(value or "").strip()
    if email and "@" not in email:
        email = f"{email}@naver.com"
    return email


def clean_int_setting(value: object, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return min(max(parsed, minimum), maximum)


def normalize_mail_technical_settings(settings: dict | None = None) -> dict[str, str | int]:
    source = settings or {}
    smtp_port = clean_int_setting(
        source.get("smtp_port"),
        int(DEFAULT_MAIL_TECHNICAL_SETTINGS["smtp_port"]),
        1,
        65535,
    )
    smtp_security = str(source.get("smtp_security") or "").strip().lower()
    if smtp_security not in {"ssl", "tls"}:
        smtp_security = "tls" if smtp_port == 587 else "ssl"
    if smtp_security == "ssl" and smtp_port not in {465, 587}:
        smtp_port = 465
    if smtp_security == "tls" and smtp_port not in {465, 587}:
        smtp_port = 587

    return {
        "smtp_host": NAVER_SMTP_HOST,
        "smtp_port": smtp_port,
        "smtp_security": smtp_security,
        "bulk_batch_size": clean_int_setting(
            source.get("bulk_batch_size"),
            int(DEFAULT_MAIL_TECHNICAL_SETTINGS["bulk_batch_size"]),
            1,
            100,
        ),
        "bulk_send_interval_seconds": clean_int_setting(
            source.get("bulk_send_interval_seconds"),
            int(DEFAULT_MAIL_TECHNICAL_SETTINGS["bulk_send_interval_seconds"]),
            5,
            600,
        ),
        "bulk_batch_pause_minutes": clean_int_setting(
            source.get("bulk_batch_pause_minutes"),
            int(DEFAULT_MAIL_TECHNICAL_SETTINGS["bulk_batch_pause_minutes"]),
            0,
            120,
        ),
        "bulk_test_recipient": str(source.get("bulk_test_recipient") or "").strip(),
    }


def load_mail_settings(include_password: bool = False) -> dict[str, str | int | bool]:
    defaults = normalize_mail_technical_settings(DEFAULT_MAIL_TECHNICAL_SETTINGS)
    if not MAIL_SETTINGS_PATH.exists():
        return {"naver_email": "", "has_password": False, **defaults}
    try:
        settings = json.loads(MAIL_SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"naver_email": "", "has_password": False, **defaults}
    password_token = str(settings.get("password_token", ""))
    loaded: dict[str, str | int | bool] = {
        "naver_email": str(settings.get("naver_email", "")),
        "has_password": bool(password_token),
        **normalize_mail_technical_settings(settings),
    }
    if include_password and password_token:
        loaded["naver_password"] = unprotect_text(password_token)
    return loaded


def save_mail_settings(
    naver_email: str,
    naver_password: str | None = None,
    bulk_settings: dict | None = None,
) -> None:
    settings = load_mail_settings(include_password=False)
    settings["naver_email"] = naver_email
    settings.update(normalize_mail_technical_settings({**settings, **(bulk_settings or {})}))
    if naver_password:
        settings["password_token"] = protect_text(naver_password)
        settings["has_password"] = True
    else:
        existing = {}
        if MAIL_SETTINGS_PATH.exists():
            try:
                existing = json.loads(MAIL_SETTINGS_PATH.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                existing = {}
        if existing.get("password_token"):
            settings["password_token"] = existing["password_token"]
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    MAIL_SETTINGS_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def contact_from_row(row: sqlite3.Row) -> dict[str, str]:
    vendor_type = normalize_vendor_type(row["vendor_type"])
    return {
        "vendor_type": vendor_type,
        "vendor_type_label": vendor_type_label(vendor_type),
        "vendor_name": str(row["vendor_name"] or ""),
        "email": str(row["email"] or ""),
    }


def load_vendor_contacts(vendor_type: str = "") -> list[dict[str, str]]:
    init_db()
    normalized_type = normalize_vendor_type(vendor_type) if vendor_type else ""
    connection = connect_db()
    try:
        if normalized_type:
            rows = connection.execute(
                """
                SELECT vendor_type, vendor_name, email
                  FROM vendor_contacts
                 WHERE vendor_type = ?
                 ORDER BY vendor_type, vendor_name
                """,
                (normalized_type,),
            ).fetchall()
        else:
            rows = connection.execute(
                """
                SELECT vendor_type, vendor_name, email
                  FROM vendor_contacts
                 ORDER BY vendor_type, vendor_name
                """
            ).fetchall()
    finally:
        connection.close()
    return [contact_from_row(row) for row in rows]


def save_vendor_contact(vendor_name: str, email: str, vendor_type: str = "purchase") -> list[dict[str, str]]:
    vendor_name = vendor_name.strip()
    email = email.strip()
    vendor_type = normalize_vendor_type(vendor_type)
    if not vendor_name:
        raise ValueError("업체명을 입력해주세요.")
    if not email or "@" not in email:
        raise ValueError("업체 메일주소를 올바르게 입력해주세요.")

    init_db()
    timestamp = now_text()
    connection = connect_db()
    try:
        connection.execute(
            """
            INSERT INTO vendor_contacts (vendor_type, vendor_name, email, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(vendor_type, vendor_name)
            DO UPDATE SET email = excluded.email, updated_at = excluded.updated_at
            """,
            (vendor_type, vendor_name, email, timestamp, timestamp),
        )
        connection.commit()
    finally:
        connection.close()
    return load_vendor_contacts()


def save_vendor_contacts_bulk(new_contacts: list[dict[str, str]]) -> tuple[list[dict[str, str]], int]:
    init_db()
    timestamp = now_text()
    saved_count = 0
    connection = connect_db()
    try:
        for contact in new_contacts:
            vendor_name = str(contact.get("vendor_name", "")).strip()
            email = str(contact.get("email", "")).strip()
            vendor_type = normalize_vendor_type(contact.get("vendor_type", "purchase"))
            if not vendor_name or not email or "@" not in email:
                continue
            connection.execute(
                """
                INSERT INTO vendor_contacts (vendor_type, vendor_name, email, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(vendor_type, vendor_name)
                DO UPDATE SET email = excluded.email, updated_at = excluded.updated_at
                """,
                (vendor_type, vendor_name, email, timestamp, timestamp),
            )
            saved_count += 1
        connection.commit()
    finally:
        connection.close()
    return load_vendor_contacts(), saved_count


def header_text(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def import_vendor_contacts_from_workbook(path: Path) -> tuple[list[dict[str, str]], int]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("업체 메일 주소록 엑셀에 데이터가 없습니다.")

    header = [header_text(value) for value in rows[0]]
    vendor_headers = {"업체명", "거래처명", "회사명", "업체", "거래처", "vendor", "vendorname", "vendor_name", "company", "companyname"}
    email_headers = {"메일", "메일주소", "이메일", "email", "e-mail", "emailaddress"}
    type_headers = {"구분", "업체구분", "거래처구분", "유형", "분류", "매입매출", "매입/매출", "type", "vendortype", "vendor_type", "category"}
    vendor_idx = next((idx for idx, value in enumerate(header) if value in vendor_headers), None)
    email_idx = next((idx for idx, value in enumerate(header) if value in email_headers), None)
    type_idx = next((idx for idx, value in enumerate(header) if value in type_headers), None)

    data_rows = rows[1:] if vendor_idx is not None and email_idx is not None else rows
    if vendor_idx is None or email_idx is None:
        vendor_idx = 0
        email_idx = 1
        type_idx = None

    imported: list[dict[str, str]] = []
    for row in data_rows:
        vendor_name = str(row[vendor_idx] or "").strip() if len(row) > vendor_idx else ""
        email = str(row[email_idx] or "").strip() if len(row) > email_idx else ""
        vendor_type = normalize_vendor_type(row[type_idx]) if type_idx is not None and len(row) > type_idx else "purchase"
        if vendor_name and email and "@" in email:
            imported.append({"vendor_type": vendor_type, "vendor_name": vendor_name, "email": email})

    if not imported:
        raise ValueError("저장할 업체명/메일주소를 찾지 못했습니다. 엑셀에 업체명과 메일주소 열을 넣어주세요.")

    return save_vendor_contacts_bulk(imported)


def find_vendor_contact(vendor_name: str, vendor_type: str = "purchase") -> dict[str, str] | None:
    init_db()
    vendor_name = str(vendor_name or "").strip()
    if not vendor_name:
        return None
    normalized_type = normalize_vendor_type(vendor_type)
    connection = connect_db()
    try:
        row = connection.execute(
            """
            SELECT vendor_type, vendor_name, email
              FROM vendor_contacts
             WHERE vendor_type = ? AND vendor_name = ?
            """,
            (normalized_type, vendor_name),
        ).fetchone()
        if row:
            return contact_from_row(row)
        target_key = normalize_company_key(vendor_name)
        for candidate in connection.execute(
            """
            SELECT vendor_type, vendor_name, email
              FROM vendor_contacts
             WHERE vendor_type = ?
            """,
            (normalized_type,),
        ).fetchall():
            if normalize_company_key(candidate["vendor_name"]) == target_key:
                return contact_from_row(candidate)
    finally:
        connection.close()
    return None


def is_purchase_vendor_cs_target(vendor_name: str) -> bool:
    key = normalize_company_key(vendor_name)
    if not key:
        return False
    if "소일브릿지" in key and "본사" in key:
        return False
    return True


def get_cs_case(case_id: int) -> dict[str, str | int] | None:
    init_db()
    connection = connect_db()
    try:
        row = connection.execute(
            """
            SELECT id, created_at, updated_at, status, vendor_name, vendor_email,
                   original_info, original_invoice, product_name, orderer_name, orderer_phone, receiver_name,
                   receiver_phone, receiver_address, cs_type, cs_content, return_invoice,
                   reship_invoice, mail_subject, mail_body, mail_sent_at, occurred_at, completed_at, order_date,
                   ship_date, sales_vendor, purchase_vendor, courier, quantity
              FROM cs_cases
             WHERE id = ?
            """,
            (case_id,),
        ).fetchone()
    finally:
        connection.close()
    return dict(row) if row else None


def default_vendor_cs_subject(vendor_name: str) -> str:
    return f"[CS 요청] {vendor_name + ' ' if vendor_name else ''}확인 부탁드립니다"


def default_vendor_cs_body(case: dict[str, object]) -> str:
    return (
        "안녕하세요. (주)소일브릿지 입니다.\n\n"
        f"- 원출고일 및 원송장번호 : {case.get('original_info') or ''}\n\n"
        f"- 상품명 : {case.get('product_name') or ''}\n\n"
        f"- 수령인 : {case.get('receiver_name') or ''}\n\n"
        f"- 수령인 연락처 : {case.get('receiver_phone') or ''}\n\n"
        f"- 수령인 주소 : {case.get('receiver_address') or ''}\n\n"
        f"- CS내용 : {case.get('cs_content') or case.get('cs_type') or ''}\n\n"
        "CS건을 보내드립니다.\n\n"
        "입고 접수 후 일주일 이상 회신 없을 경우 자체 환불 및 정산 반영 예정이오니 처리 결과 회신 부탁드립니다.\n"
    )


def vendor_cs_mail_prompt(case: dict[str, object] | None) -> dict[str, object]:
    if not case:
        return {"enabled": False}
    vendor_name = str(case.get("purchase_vendor") or case.get("vendor_name") or "").strip()
    if not is_purchase_vendor_cs_target(vendor_name):
        return {"enabled": False}
    contact = find_vendor_contact(vendor_name, "purchase")
    recipient_email = contact["email"] if contact else ""
    payload = {
        "case_id": str(case.get("id") or ""),
        "vendor_type": "purchase",
        "vendor_name": vendor_name,
        "recipient_email": recipient_email,
        "cs_origin": str(case.get("original_info") or ""),
        "original_invoice": str(case.get("original_invoice") or ""),
        "cs_product": str(case.get("product_name") or ""),
        "cs_receiver": str(case.get("receiver_name") or ""),
        "cs_phone": str(case.get("receiver_phone") or ""),
        "cs_address": str(case.get("receiver_address") or ""),
        "cs_type": str(case.get("cs_type") or ""),
        "cs_content": str(case.get("cs_content") or ""),
        "subject": default_vendor_cs_subject(vendor_name),
        "body": default_vendor_cs_body(case),
    }
    return {
        "enabled": True,
        "case_id": case.get("id") or "",
        "vendor_name": vendor_name,
        "recipient_email": recipient_email,
        "missing_email": not bool(recipient_email),
        "payload": payload,
    }


def mark_cs_case_mail_sent(case_id: int, payload: dict) -> None:
    if not case_id:
        raise ValueError("메일 발송 처리할 CS건 ID가 없습니다.")
    init_db()
    case = cs_case_from_payload(payload, status="메일발송", mail_sent=True)
    connection = connect_db()
    try:
        cursor = connection.execute(
            """
            UPDATE cs_cases
               SET status = ?,
                   vendor_name = COALESCE(NULLIF(?, ''), vendor_name),
                   vendor_email = COALESCE(NULLIF(?, ''), vendor_email),
                   original_info = COALESCE(NULLIF(?, ''), original_info),
                   original_invoice = COALESCE(NULLIF(?, ''), original_invoice),
                   product_name = COALESCE(NULLIF(?, ''), product_name),
                   receiver_name = COALESCE(NULLIF(?, ''), receiver_name),
                   receiver_phone = COALESCE(NULLIF(?, ''), receiver_phone),
                   receiver_address = COALESCE(NULLIF(?, ''), receiver_address),
                   cs_type = COALESCE(NULLIF(?, ''), cs_type),
                   cs_content = COALESCE(NULLIF(?, ''), cs_content),
                   mail_subject = ?,
                   mail_body = ?,
                   mail_sent_at = ?,
                   updated_at = ?
             WHERE id = ?
            """,
            [
                case["status"],
                case["vendor_name"],
                case["vendor_email"],
                case["original_info"],
                case["original_invoice"],
                case["product_name"],
                case["receiver_name"],
                case["receiver_phone"],
                case["receiver_address"],
                case["cs_type"],
                case["cs_content"],
                case["mail_subject"],
                case["mail_body"],
                case["mail_sent_at"],
                now_text(),
                case_id,
            ],
        )
        connection.commit()
        if cursor.rowcount == 0:
            raise ValueError("메일 발송 처리할 CS건을 찾지 못했습니다.")
    finally:
        connection.close()


def send_cs_mail(payload: dict, attachments: list[dict[str, object]] | None = None) -> None:
    saved = load_mail_settings(include_password=True)
    naver_email = normalize_naver_email(payload.get("naver_email")) or str(saved.get("naver_email", ""))
    naver_password = str(payload.get("naver_password") or saved.get("naver_password", ""))
    recipient = str(payload.get("recipient_email", "")).strip()
    subject = str(payload.get("subject", "")).strip()
    body = str(payload.get("body", "")).strip()

    if not naver_email:
        raise ValueError("네이버 메일 아이디를 입력해주세요.")
    if not naver_password:
        raise ValueError("네이버 메일 비밀번호를 입력해주세요.")
    if not recipient:
        raise ValueError("받는 업체 메일을 입력해주세요.")
    if not subject:
        raise ValueError("메일 제목을 입력해주세요.")
    if not body:
        raise ValueError("메일 내용을 입력해주세요.")

    if payload.get("save_credentials", True):
        save_mail_settings(naver_email, naver_password)
    send_naver_mail(
        naver_email,
        naver_password,
        recipient,
        subject,
        body,
        smtp_port=int(saved.get("smtp_port") or NAVER_SMTP_PORT),
        smtp_security=str(saved.get("smtp_security") or "ssl"),
        attachments=attachments,
    )
    return

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = formataddr(("(주)소일브릿지", naver_email))
    message["To"] = recipient
    message.set_content(body)

    context = ssl.create_default_context()
    try:
        with smtplib.SMTP_SSL(NAVER_SMTP_HOST, NAVER_SMTP_PORT, context=context, timeout=20) as smtp:
            smtp.login(naver_email, naver_password)
            smtp.send_message(message)
    except smtplib.SMTPAuthenticationError as exc:
        raise ValueError("네이버 메일 로그인에 실패했습니다. 아이디/비밀번호와 네이버 메일 SMTP 사용 설정을 확인해주세요.") from exc


def send_general_mail(payload: dict, attachments: list[dict[str, object]] | None = None) -> None:
    send_cs_mail(payload, attachments=attachments)


def send_naver_mail(
    naver_email: str,
    naver_password: str,
    recipient: str,
    subject: str,
    body: str,
    smtp_port: int = NAVER_SMTP_PORT,
    smtp_security: str = "ssl",
    attachments: list[dict[str, object]] | None = None,
) -> None:
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = formataddr(("(주)소일브릿지", naver_email))
    message["To"] = recipient
    message.set_content(body)
    for attachment in attachments or []:
        filename = str(attachment.get("filename") or "attachment")
        data = attachment.get("data") or b""
        if not isinstance(data, (bytes, bytearray)):
            continue
        content_type = str(attachment.get("content_type") or "application/octet-stream")
        maintype, _, subtype = content_type.partition("/")
        if not maintype or not subtype:
            maintype, subtype = "application", "octet-stream"
        message.add_attachment(bytes(data), maintype=maintype, subtype=subtype, filename=filename)

    smtp_security = smtp_security if smtp_security in {"ssl", "tls"} else "ssl"
    context = ssl.create_default_context()
    try:
        if smtp_security == "tls":
            with smtplib.SMTP(NAVER_SMTP_HOST, smtp_port, timeout=20) as smtp:
                smtp.starttls(context=context)
                smtp.login(naver_email, naver_password)
                smtp.send_message(message)
        else:
            with smtplib.SMTP_SSL(NAVER_SMTP_HOST, smtp_port, context=context, timeout=20) as smtp:
                smtp.login(naver_email, naver_password)
                smtp.send_message(message)
    except smtplib.SMTPAuthenticationError as exc:
        raise ValueError("네이버 메일 로그인에 실패했습니다. 아이디/비밀번호와 네이버 메일 SMTP 사용 설정을 확인해주세요.") from exc


def send_mail_test(payload: dict) -> None:
    saved = load_mail_settings(include_password=True)
    naver_email = normalize_naver_email(saved.get("naver_email"))
    naver_password = str(saved.get("naver_password", ""))
    recipient = str(payload.get("recipient_email") or saved.get("bulk_test_recipient") or "").strip()
    if not naver_email:
        raise ValueError("네이버 메일 아이디를 입력해주세요.")
    if not naver_password:
        raise ValueError("네이버 메일 비밀번호를 입력해주세요.")
    if not recipient:
        raise ValueError("테스트 수신 메일 주소를 입력해주세요.")
    send_naver_mail(
        naver_email,
        naver_password,
        recipient,
        "[소일브릿지] 네이버 메일 SMTP 테스트",
        "소일브릿지 업무자동화 프로그램의 네이버 메일 SMTP 테스트 발송입니다.",
        smtp_port=int(saved.get("smtp_port") or NAVER_SMTP_PORT),
        smtp_security=str(saved.get("smtp_security") or "ssl"),
    )


class WorkhubHandler(BaseHTTPRequestHandler):
    def cookie_value(self, name: str) -> str:
        cookies = self.headers.get("Cookie", "")
        for part in cookies.split(";"):
            if "=" not in part:
                continue
            key, value = part.strip().split("=", 1)
            if key == name:
                return unquote(value)
        return ""

    def current_user(self) -> dict[str, str] | None:
        return current_user_from_token(self.cookie_value(SESSION_COOKIE_NAME))

    def client_ip(self) -> str:
        forwarded = self.headers.get("X-Forwarded-For", "").split(",", 1)[0].strip()
        return forwarded or str(self.client_address[0] if self.client_address else "local")

    def is_secure_request(self) -> bool:
        return isinstance(self.request, ssl.SSLSocket) or self.headers.get("X-Forwarded-Proto", "").lower() == "https"

    def send_redirect(self, location: str, status: int = 303) -> None:
        self.send_response(status)
        self.send_header("Location", location)
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "same-origin")
        self.end_headers()

    def set_session_cookie(self, token: str) -> None:
        secure = "; Secure" if self.is_secure_request() else ""
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE_NAME}={quote(token)}; Path=/; Max-Age={SESSION_SECONDS}; HttpOnly; SameSite=Lax{secure}",
        )

    def clear_session_cookie(self) -> None:
        secure = "; Secure" if self.is_secure_request() else ""
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE_NAME}=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax{secure}",
        )

    def require_permission(self, user: dict[str, str], permission: str, label: str) -> bool:
        if user_has_permission(user, permission):
            return True
        self.send_json({"error": f"{label} 권한이 없습니다."}, status=403)
        return False

    def require_admin(self, user: dict[str, str], label: str) -> bool:
        if user.get("role") == "admin":
            return True
        self.send_json({"error": f"{label}은 관리자만 사용할 수 있습니다."}, status=403)
        return False

    def do_GET(self) -> None:
        if self.path.startswith("/static/"):
            relative = unquote(self.path.removeprefix("/static/"))
            target = (STATIC_DIR / relative).resolve()
            if STATIC_DIR.resolve() in target.parents and target.is_file():
                content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
                self.send_bytes(target.read_bytes(), content_type)
                return

        if self.path.startswith("/lucide/"):
            relative = unquote(self.path.removeprefix("/lucide/"))
            target = (LUCIDE_DIR / relative).resolve()
            if LUCIDE_DIR.resolve() in target.parents and target.is_file():
                content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
                self.send_bytes(target.read_bytes(), content_type)
                return
            if relative == "dist/esm/lucide.js":
                self.send_bytes(LUCIDE_FALLBACK_JS.encode("utf-8"), "application/javascript; charset=utf-8")
                return

        if self.path.startswith("/login"):
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            message = ""
            login_error = ""
            error_code = params.get("error", [""])[0]
            if params.get("registered", [""])[0] == "1":
                message = "계정 등록 요청이 접수됐습니다. 관리자가 승인하면 로그인할 수 있습니다."
            if error_code == "locked":
                login_error = "로그인 시도가 잠시 제한됐습니다. 잠시 후 다시 시도해주세요."
            elif error_code:
                login_error = "아이디 또는 비밀번호가 올바르지 않습니다."
            self.send_bytes(
                render_login_html(show_error=bool(error_code), login_error=login_error, message=message).encode("utf-8"),
                "text/html; charset=utf-8",
            )
            return

        if self.path == "/logout":
            delete_login_session(self.cookie_value(SESSION_COOKIE_NAME))
            self.send_response(303)
            self.clear_session_cookie()
            self.send_header("Location", "/login")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            return

        user = self.current_user()
        if not user:
            if self.path.startswith("/api/"):
                self.send_json({"error": "로그인이 필요합니다."}, status=401)
                return
            self.send_redirect("/login")
            return

        if self.path == "/" or self.path.startswith("/?"):
            self.send_bytes(render_app_html(user).encode("utf-8"), "text/html; charset=utf-8")
            return

        if self.path == "/api/users":
            if not self.require_permission(user, "user_admin", "사용자 관리"):
                return
            self.send_json({"users": list_users()})
            return

        if self.path == "/api/backups":
            if not self.require_permission(user, "backup_manage", "백업 관리"):
                return
            settings = load_backup_settings()
            backups = list_backup_files()
            self.send_json({
                "backup_dir": str(backup_dir_path(settings)),
                "backups": backups,
                "last_backup": backups[0]["created_at"] if backups else "",
                "retention_days": settings["retention_days"],
                "auto_backup_hour": settings["auto_hour"],
                "settings": settings,
            })
            return

        if self.path == "/api/order-downloads":
            if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                return
            self.send_json({"downloads": list_order_downloads(), "limit": ORDER_DOWNLOAD_LIMIT})
            return

        if self.path == "/api/shared-files":
            self.send_json({"files": list_shared_files()})
            return

        if self.path == "/api/sales-report-uploads":
            if not self.require_permission(user, "sales_report_manage", "매출표 업로드"):
                return
            self.send_json({"files": list_sales_report_uploads()})
            return

        if self.path.startswith("/api/sales-report-dashboard"):
            if not self.require_permission(user, "sales_report_manage", "매출현황"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            self.send_json(sales_report_dashboard_payload(
                period=params.get("period", [""])[0],
                report_date=params.get("date", [""])[0],
            ))
            return

        if self.path.startswith("/api/shared-file-download"):
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            try:
                path, metadata = shared_file_download_info(params.get("id", [""])[0])
            except Exception as exc:  # noqa: BLE001
                self.send_json({"error": str(exc)}, status=404)
                return
            data = path.read_bytes()
            content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
            filename = quote(str(metadata.get("original_name") or path.name))
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        if self.path.startswith("/api/order-download"):
            if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            try:
                download_id = params.get("id", [""])[0]
                path = order_download_path(download_id)
                filename = order_download_filename(download_id)
            except Exception as exc:  # noqa: BLE001
                self.send_json({"error": str(exc)}, status=404)
                return
            data = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        if self.path == "/api/system-update":
            if not self.require_permission(user, "system_update", "시스템 업데이트"):
                return
            self.send_json(system_update_payload(fetch=False))
            return

        if self.path.startswith("/api/backup-download"):
            if not self.require_permission(user, "backup_manage", "백업 관리"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            try:
                path = backup_path_from_name(params.get("name", [""])[0])
                if not path.exists():
                    raise FileNotFoundError("백업 파일을 찾지 못했습니다.")
            except Exception as exc:  # noqa: BLE001
                self.send_json({"error": str(exc)}, status=400)
                return
            data = path.read_bytes()
            filename = quote(path.name)
            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        if self.path == "/api/leaves":
            if not any(user_has_permission(user, permission) for permission in ("leave_view", "leave_approve", "leave_manage")):
                self.send_json({"error": "연차 조회 권한이 없습니다."}, status=403)
                return
            self.send_json(leave_payload(user))
            return

        if self.path == "/api/mail-settings":
            if not self.require_permission(user, "mail_send", "메일 발송"):
                return
            self.send_json(load_mail_settings(include_password=False))
            return

        if self.path == "/api/vendor-contacts":
            if not self.require_permission(user, "mail_send", "메일 발송"):
                return
            self.send_json({"contacts": load_vendor_contacts()})
            return

        if self.path.startswith("/api/cs-cases"):
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            query = params.get("q", [""])[0]
            status = params.get("status", [""])[0]
            year = params.get("year", [""])[0]
            month = params.get("month", [""])[0]
            try:
                limit = min(max(int(params.get("limit", ["100"])[0]), 1), 5000)
            except ValueError:
                limit = 100
            self.send_json({"cases": list_cs_cases(query=query, status=status, limit=limit, year=year, month=month)})
            return

        if self.path.startswith("/api/management-records"):
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            query = params.get("q", [""])[0]
            year = params.get("year", [""])[0]
            month = params.get("month", [""])[0]
            try:
                limit = min(max(int(params.get("limit", ["100"])[0]), 1), 5000)
            except ValueError:
                limit = 100
            self.send_json({"records": list_management_records(query=query, limit=limit, year=year, month=month)})
            return

        if self.path == "/api/management-periods":
            self.send_json({"periods": list_management_periods()})
            return

        if self.path == "/api/company-staff-dashboard":
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            self.send_json(company_staff_dashboard_payload(user))
            return

        if self.path.startswith("/api/company-calendar-events"):
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            try:
                self.send_json(company_calendar_payload(user, params.get("month", [""])[0]))
            except ValueError as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if self.path.startswith("/api/internal-messages"):
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            try:
                limit = min(max(int(params.get("limit", ["100"])[0]), 1), 300)
            except ValueError:
                limit = 100
            try:
                other_user_id = int(params.get("user_id", ["0"])[0] or 0)
            except ValueError:
                other_user_id = 0
            self.send_json({
                "messages": list_internal_messages(
                    limit=limit,
                    room_type=params.get("room", ["global"])[0],
                    current_user_id=int(user.get("id") or 0),
                    other_user_id=other_user_id,
                )
            })
            return

        if self.path == "/api/crm-dashboard":
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            self.send_json(crm_dashboard_payload(DB_PATH))
            return

        if self.path.startswith("/api/crm-accounts"):
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            try:
                limit = min(max(int(params.get("limit", ["200"])[0]), 1), 1000)
            except ValueError:
                limit = 200
            self.send_json({"accounts": list_crm_accounts(DB_PATH, query=params.get("q", [""])[0], limit=limit)})
            return

        if self.path.startswith("/api/crm-tasks"):
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            try:
                limit = min(max(int(params.get("limit", ["300"])[0]), 1), 2000)
            except ValueError:
                limit = 300
            try:
                filter_assignee_user_id = int(params.get("assignee_user_id", ["0"])[0] or 0) or None
            except ValueError:
                filter_assignee_user_id = None
            if params.get("mine", [""])[0] == "1":
                filter_assignee_user_id = int(user.get("id") or 0)
            self.send_json({
                "tasks": list_crm_tasks(
                    DB_PATH,
                    query=params.get("q", [""])[0],
                    status=params.get("status", [""])[0],
                    assignee=params.get("assignee", [""])[0],
                    assignee_user_id=filter_assignee_user_id,
                    priority=params.get("priority", [""])[0],
                    due=params.get("due", [""])[0],
                    source=params.get("source", [""])[0],
                    open_only=params.get("open_only", [""])[0] == "1",
                    sort=params.get("sort", ["smart"])[0],
                    limit=limit,
                )
            })
            return

        if self.path.startswith("/api/crm-task-comments"):
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            self.send_json({
                "comments": list_crm_task_comments(
                    DB_PATH,
                    int(params.get("task_id", ["0"])[0] or 0),
                )
            })
            return

        if self.path.startswith("/api/crm-saved-views"):
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            parsed = urlsplit(self.path)
            params = parse_qs(parsed.query)
            self.send_json({
                "views": list_crm_saved_views(
                    DB_PATH,
                    int(user.get("id") or 0),
                    scope=params.get("scope", ["tasks"])[0],
                )
            })
            return

        if self.path == "/api/crm-message-events":
            if not self.require_permission(user, "crm_message_manage", "CRM 메신저 연동"):
                return
            self.send_json({
                "events": list_crm_message_events(DB_PATH),
                "webhook": {
                    "path": "/api/crm-messenger-webhook",
                    "url": crm_webhook_public_url(self),
                    "token": crm_webhook_token(),
                },
            })
            return

        if self.path == "/api/crm-messenger-users":
            if not self.require_permission(user, "crm_view", "CRM 조회"):
                return
            payload = list_crm_messenger_users(DB_PATH)
            if not user_has_permission(user, "crm_message_manage"):
                payload["mappings"] = []
            self.send_json(payload)
            return

        if self.path == "/api/import-shipments":
            self.send_json({"shipments": list_import_shipments()})
            return

        self.send_error(404)

    def do_POST(self) -> None:
        try:
            if self.path == "/login":
                length = int(self.headers.get("Content-Length", "0"))
                raw_body = self.rfile.read(length).decode("utf-8")
                payload = parse_qs(raw_body)
                username = payload.get("username", [""])[0]
                password = payload.get("password", [""])[0]
                locked, lock_message = login_lock_status(username, self.client_ip())
                if locked:
                    self.send_bytes(
                        render_login_html(show_error=True, login_error=lock_message).encode("utf-8"),
                        "text/html; charset=utf-8",
                        status=429,
                    )
                    return
                user = authenticate_user(
                    username,
                    password,
                )
                if not user:
                    record_login_failure(username, self.client_ip())
                    self.send_redirect("/login?error=1")
                    return
                clear_login_failures(username, self.client_ip())
                token = create_login_session(user["username"])
                self.send_response(303)
                self.set_session_cookie(token)
                self.send_header("Location", "/")
                self.send_header("Cache-Control", "no-store")
                self.send_header("X-Content-Type-Options", "nosniff")
                self.send_header("X-Frame-Options", "DENY")
                self.send_header("Referrer-Policy", "same-origin")
                self.end_headers()
                return

            if self.path == "/register":
                length = int(self.headers.get("Content-Length", "0"))
                raw_body = self.rfile.read(length).decode("utf-8")
                payload = {key: values[0] if values else "" for key, values in parse_qs(raw_body).items()}
                try:
                    register_user_request(payload)
                except ValueError as exc:
                    self.send_bytes(
                        render_login_html(register_error=str(exc)).encode("utf-8"),
                        "text/html; charset=utf-8",
                        status=400,
                    )
                    return
                self.send_redirect("/login?registered=1")
                return

            if self.path == "/api/crm-messenger-webhook":
                expected_token = crm_webhook_token()
                received_token = self.headers.get("X-Workhub-Webhook-Token", "")
                if not hmac.compare_digest(received_token, expected_token):
                    self.send_json({"error": "웹훅 토큰이 올바르지 않습니다."}, status=403)
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                self.send_json(handle_crm_messenger_webhook(DB_PATH, payload))
                return

            user = self.current_user()
            if not user:
                self.send_json({"error": "로그인이 필요합니다."}, status=401)
                return

            if self.path == "/api/internal-message-save":
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                message_id = save_internal_message(payload, user)
                self.send_json({"message": "메시지를 저장했습니다.", "message_id": message_id})
                return

            if self.path == "/api/crm-account-save":
                if not self.require_permission(user, "crm_manage", "CRM 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                account_id = save_crm_account(DB_PATH, payload)
                self.send_json({"message": "CRM 거래처를 저장했습니다.", "account_id": account_id})
                return

            if self.path == "/api/crm-task-save":
                if not self.require_permission(user, "crm_manage", "CRM 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                task_id = save_crm_task(DB_PATH, payload, user)
                self.send_json({"message": "CRM 업무를 저장했습니다.", "task_id": task_id})
                return

            if self.path == "/api/crm-task-status":
                if not self.require_permission(user, "crm_view", "CRM 조회"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                task_id = change_crm_task_status(
                    DB_PATH,
                    int(payload.get("id") or 0),
                    str(payload.get("status") or ""),
                    user,
                    comment=str(payload.get("comment") or ""),
                    can_manage=user_has_permission(user, "crm_manage"),
                )
                self.send_json({"message": "CRM 업무 상태를 저장했습니다.", "task_id": task_id})
                return

            if self.path == "/api/crm-task-comment":
                if not self.require_permission(user, "crm_view", "CRM 조회"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                task_id = add_crm_task_comment(
                    DB_PATH,
                    int(payload.get("id") or 0),
                    str(payload.get("body") or ""),
                    user,
                    can_manage=user_has_permission(user, "crm_manage"),
                )
                self.send_json({"message": "CRM 업무 댓글을 저장했습니다.", "task_id": task_id})
                return

            if self.path == "/api/crm-saved-view-save":
                if not self.require_permission(user, "crm_view", "CRM 조회"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                view_id = save_crm_saved_view(DB_PATH, payload, user)
                self.send_json({
                    "message": "CRM 저장뷰를 저장했습니다.",
                    "view_id": view_id,
                    "views": list_crm_saved_views(DB_PATH, int(user.get("id") or 0), scope=str(payload.get("scope") or "tasks")),
                })
                return

            if self.path == "/api/crm-saved-view-delete":
                if not self.require_permission(user, "crm_view", "CRM 조회"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                delete_crm_saved_view(DB_PATH, int(payload.get("id") or 0), user)
                self.send_json({
                    "message": "CRM 저장뷰를 삭제했습니다.",
                    "views": list_crm_saved_views(DB_PATH, int(user.get("id") or 0), scope=str(payload.get("scope") or "tasks")),
                })
                return

            if self.path == "/api/crm-messenger-user-save":
                if not self.require_permission(user, "crm_message_manage", "CRM 메신저 연동"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                mapping_id = save_crm_messenger_user(DB_PATH, payload)
                self.send_json({"message": "메신저 사용자 매핑을 저장했습니다.", "mapping_id": mapping_id})
                return

            if self.path == "/api/crm-webhook-token-rotate":
                if not self.require_permission(user, "crm_message_manage", "CRM 메신저 연동"):
                    return
                token = rotate_crm_webhook_token()
                self.send_json({
                    "message": "웹훅 토큰을 재발급했습니다. 이전 토큰은 즉시 사용할 수 없습니다.",
                    "webhook": {
                        "path": "/api/crm-messenger-webhook",
                        "url": crm_webhook_public_url(self),
                        "token": token,
                    },
                })
                return

            if self.path == "/api/users-save":
                if not self.require_permission(user, "user_admin", "사용자 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                user_id = save_user_account(payload, user)
                self.send_json({"message": "사용자 계정을 저장했습니다.", "user_id": user_id, "users": list_users()})
                return

            if self.path == "/api/backup-create":
                if not self.require_permission(user, "backup_manage", "백업 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
                backup_dir = clean_payload_text(payload, "backup_dir") if isinstance(payload, dict) else ""
                backup = create_workhub_backup("selected" if backup_dir else "manual", backup_dir=backup_dir or None)
                self.send_json({"message": "백업 파일을 생성했습니다.", "backup": backup})
                return

            if self.path == "/api/backup-settings":
                if not self.require_permission(user, "backup_manage", "백업 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
                if not isinstance(payload, dict):
                    raise ValueError("백업 설정 형식이 올바르지 않습니다.")
                settings = save_backup_settings(payload)
                self.send_json({
                    "message": "백업 설정을 저장했습니다.",
                    "settings": settings,
                    "backup_dir": str(backup_dir_path(settings)),
                })
                return

            if self.path == "/api/system-update-check":
                if not self.require_permission(user, "system_update", "시스템 업데이트"):
                    return
                self.send_json(system_update_payload(fetch=True))
                return

            if self.path == "/api/system-update-apply":
                if not self.require_permission(user, "system_update", "시스템 업데이트"):
                    return
                self.send_json(apply_system_update())
                return

            if self.path == "/api/backup-delete":
                if not self.require_permission(user, "backup_manage", "백업 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                path = backup_path_from_name(clean_payload_text(payload, "name"))
                if not path.exists():
                    raise FileNotFoundError("삭제할 백업 파일을 찾지 못했습니다.")
                path.unlink()
                self.send_json({"message": "백업 파일을 삭제했습니다.", "name": path.name})
                return

            if self.path == "/api/backup-restore":
                if not self.require_permission(user, "backup_manage", "백업 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                path = backup_path_from_name(clean_payload_text(payload, "name"))
                if not path.exists():
                    raise FileNotFoundError("복원할 백업 파일을 찾지 못했습니다.")
                self.send_json(restore_workhub_backup(path))
                return

            if self.path == "/api/leave-request":
                if not self.require_permission(user, "leave_view", "연차 조회"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                request_id = create_leave_request(user, payload)
                self.send_json({"message": "연차 신청이 저장되었습니다.", "request_id": request_id})
                return

            if self.path == "/api/leave-decision":
                if not any(user_has_permission(user, permission) for permission in ("leave_approve", "leave_approve_team", "leave_approve_director", "leave_approve_ceo", "leave_director_override", "leave_manage")):
                    self.send_json({"error": "연차 승인 권한이 없습니다."}, status=403)
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                decide_leave_request(int(payload.get("request_id", 0)), user, clean_payload_text(payload, "decision"), clean_payload_text(payload, "comment"))
                self.send_json({"message": "연차 신청을 처리했습니다."})
                return

            if self.path == "/api/leave-cancel":
                if not self.require_permission(user, "leave_view", "\uC5F0\uCC28 \uC870\uD68C"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                cancel_leave_request(int(payload.get("request_id", 0)), user, clean_payload_text(payload, "reason"))
                self.send_json({"message": "\uC5F0\uCC28 \uC2E0\uCCAD\uC744 \uCDE8\uC18C\uD588\uC2B5\uB2C8\uB2E4."})
                return

            if self.path == "/api/leave-accrual":
                if not self.require_permission(user, "leave_manage", "\uC5F0\uCC28 \uAD00\uB9AC"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                year = int(payload.get("year") or date.today().year)
                default_days = float(payload.get("default_days") or 15)
                count = apply_annual_leave_accrual(year, actor=user, default_days=default_days)
                self.send_json({"message": f"{year}\uB144 \uC5F0\uCC28 \uC790\uB3D9 \uBC1C\uC0DD\uC744 {count}\uBA85\uC5D0\uAC8C \uC801\uC6A9\uD588\uC2B5\uB2C8\uB2E4.", "count": count})
                return

            if self.path == "/api/leave-balance":
                if not self.require_permission(user, "leave_manage", "연차 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                set_leave_balance(payload, user)
                self.send_json({"message": "직원 연차 기준을 저장했습니다."})
                return

            if self.path == "/api/leave-historical":
                if not self.require_permission(user, "leave_manage", "연차 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                count = add_historical_leave_usage(payload, user)
                self.send_json({"message": f"기존 사용 연차 {count}건을 등록했습니다.", "count": count})
                return

            if self.path == "/api/mail-settings":
                if not self.require_admin(user, "메일 기본정보 저장"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                naver_email = normalize_naver_email(payload.get("naver_email"))
                if not naver_email:
                    raise ValueError("네이버 메일 아이디를 입력해주세요.")
                naver_password = str(payload.get("naver_password") or "") if payload.get("save_credentials", True) else ""
                save_mail_settings(naver_email, naver_password, bulk_settings=payload)
                settings = load_mail_settings(include_password=False)
                self.send_json({
                    "message": "메일 기본정보를 저장했습니다.",
                    "naver_email": settings.get("naver_email", ""),
                    "has_password": settings.get("has_password", False),
                    "smtp_host": settings.get("smtp_host", NAVER_SMTP_HOST),
                    "smtp_port": settings.get("smtp_port", NAVER_SMTP_PORT),
                    "smtp_security": settings.get("smtp_security", "ssl"),
                    "bulk_batch_size": settings.get("bulk_batch_size", 20),
                    "bulk_send_interval_seconds": settings.get("bulk_send_interval_seconds", 15),
                    "bulk_batch_pause_minutes": settings.get("bulk_batch_pause_minutes", 5),
                    "bulk_test_recipient": settings.get("bulk_test_recipient", ""),
                })
                return

            if self.path == "/api/mail-test":
                if not self.require_admin(user, "메일 테스트 발송"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                send_mail_test(payload)
                self.send_json({"message": "테스트 메일을 발송했습니다."})
                return

            if self.path == "/api/mail-send":
                if not self.require_permission(user, "mail_send", "메일 발송"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                send_general_mail(payload)
                self.send_json({"message": "메일 발송이 완료되었습니다."})
                return

            if self.path == "/api/cs-mail":
                if not self.require_permission(user, "mail_send", "메일 발송"):
                    return
                attachments: list[dict[str, object]] = []
                if self.headers.get("Content-Type", "").lower().startswith("multipart/form-data"):
                    fields = parse_multipart(self.headers, self.rfile)
                    raw_payload = fields.get("payload", "{}")
                    payload = json.loads(str(raw_payload or "{}"))
                    attachments = collect_mail_attachments(fields)
                else:
                    length = int(self.headers.get("Content-Length", "0"))
                    payload = json.loads(self.rfile.read(length).decode("utf-8"))
                send_cs_mail(payload, attachments=attachments)
                case_id = int(payload.get("case_id") or 0)
                if case_id:
                    mark_cs_case_mail_sent(case_id, payload)
                else:
                    case_id = save_cs_case(payload, status="메일발송", mail_sent=True)
                self.send_json({"message": "CS 요청 메일 전송 및 DB 저장이 완료되었습니다.", "case_id": case_id})
                return

            if self.path == "/api/cs-case":
                if not self.require_permission(user, "ledger_edit", "대장 수정"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                case_id = save_cs_case(payload, status=clean_payload_text(payload, "status") or "접수")
                self.send_json({"message": "CS건을 DB에 저장했습니다.", "case_id": case_id})
                return

            if self.path == "/api/cs-case-update":
                if not self.require_permission(user, "ledger_edit", "대장 수정"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                case_id = int(payload.get("id", 0))
                if not case_id:
                    raise ValueError("수정할 CS건 ID가 없습니다.")
                update_cs_case(case_id, payload)
                self.send_json({"message": "CS 처리내용을 저장했습니다.", "case_id": case_id})
                return

            if self.path == "/api/cs-cases-delete":
                if not self.require_permission(user, "ledger_delete", "대장 삭제"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                case_ids = [int(value) for value in payload.get("ids", []) if str(value).isdigit()]
                if not case_ids:
                    raise ValueError("삭제할 CS 처리대장 행이 없습니다.")
                deleted = delete_cs_cases(case_ids)
                self.send_json({"message": f"CS 처리대장 선택 주문 {deleted}건을 삭제했습니다.", "deleted": deleted})
                return

            if self.path == "/api/management-record-update":
                if not self.require_permission(user, "ledger_edit", "대장 수정"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                record_id = int(payload.get("id", 0))
                if not record_id:
                    raise ValueError("수정할 통합관리대장 행 ID가 없습니다.")
                update_management_record(record_id, payload)
                self.send_json({"message": "통합관리대장 행을 저장했습니다.", "record_id": record_id})
                return

            if self.path == "/api/management-records-delete":
                if not self.require_permission(user, "ledger_delete", "대장 삭제"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                record_ids = [int(value) for value in payload.get("ids", []) if str(value).isdigit()]
                if not record_ids:
                    raise ValueError("삭제할 통합관리대장 행이 없습니다.")
                deleted = delete_management_records(record_ids)
                self.send_json({"message": f"통합관리대장 선택 주문 {deleted}건을 삭제했습니다.", "deleted": deleted})
                return

            if self.path == "/api/management-to-cs":
                if not self.require_permission(user, "cs_receive", "CS접수"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                record_id = int(payload.get("id", 0))
                if not record_id:
                    raise ValueError("CS접수할 통합관리대장 행 ID가 없습니다.")
                case_id = create_cs_case_from_management(record_id)
                case = get_cs_case(case_id)
                self.send_json({
                    "message": "CS 처리대장에 접수했습니다.",
                    "case_id": case_id,
                    "case": case,
                    "mail_prompt": vendor_cs_mail_prompt(case),
                })
                return

            if self.path == "/api/import-shipment-save":
                if not self.require_permission(user, "import_shipment_manage", "수입제품 진행 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                shipment_id = save_import_shipment(payload)
                self.send_json({"message": "수입제품 출고 진행 상황을 저장했습니다.", "shipment_id": shipment_id})
                return

            if self.path == "/api/import-shipment-complete":
                if not self.require_permission(user, "import_shipment_manage", "수입제품 진행 관리"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                complete_import_shipment(int(payload.get("id") or 0))
                self.send_json({"message": "수입제품 출고 진행 건을 완료 처리했습니다."})
                return

            if self.path == "/api/management-export":
                if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                rows, filename_stem = management_export_rows_from_payload(payload)
                if not rows:
                    raise ValueError("엑셀로 다운로드할 통합관리대장 데이터가 없습니다.")
                data = management_workbook_bytes_from_template(rows)
                filename = quote(management_export_filename(filename_stem, payload))
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return

            if self.path == "/api/cs-cases-export":
                if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                rows, filename_stem = ledger_export_rows_from_payload(payload)
                if not rows:
                    raise ValueError("엑셀로 다운로드할 CS 처리대장 데이터가 없습니다.")
                data = workbook_bytes_from_rows(rows, LEDGER_EXPORT_COLUMNS, "CS처리대장")
                filename = quote(f"{filename_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return

            if self.path == "/api/vendor-contact":
                if not self.require_permission(user, "mail_send", "메일 발송"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                contacts = save_vendor_contact(
                    str(payload.get("vendor_name", "")),
                    str(payload.get("email", "")),
                    str(payload.get("vendor_type", "purchase")),
                )
                self.send_json({"message": "업체 메일 주소를 저장했습니다.", "contacts": contacts})
                return

            if self.path == "/api/shared-file-delete":
                if not self.require_admin(user, "업무 파일 삭제"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                delete_shared_file(payload.get("id"))
                self.send_json({"message": "업무 파일을 삭제했습니다.", "files": list_shared_files()})
                return

            if self.path == "/api/vehicle-receipt":
                if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                    return
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                output_path = generate_vehicle_receipt(
                    supplier=payload.get("supplier", ""),
                    items=payload.get("items", []),
                    freight_payment=payload.get("freight_payment", "선불"),
                    receipt_type=payload.get("receipt_type", "일반"),
                    request_note=payload.get("request_note", ""),
                    delivery_place=payload.get("delivery_place", ""),
                    manager=payload.get("manager", ""),
                    output_dir=DOWNLOAD_DIR,
                    output_date=parse_receipt_date(payload.get("receipt_date", "")),
                )
                register_order_download(output_path, "차량인수증")
                filename = quote(output_path.name)
                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.send_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{filename}",
                )
                data = output_path.read_bytes()
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return

            if urlsplit(self.path).path == "/api/shared-file-upload":
                if not self.require_admin(user, "업무 파일 업로드"):
                    return
                fields = parse_multipart(self.headers, self.rfile)
                upload_path = save_uploaded_shared_file(fields, "file")
                uploaded_by = str(user.get("display_name") or user.get("username") or "")
                save_shared_file(upload_path, original_uploaded_filename(upload_path.name), uploaded_by)
                self.send_json({"message": "업무 파일을 저장했습니다.", "files": list_shared_files()})
                return

            fields = parse_multipart(self.headers, self.rfile)

            if self.path == "/api/backup-restore-upload":
                if not self.require_permission(user, "backup_manage", "백업 관리"):
                    return
                upload_path = save_uploaded_backup_zip(fields, "file")
                self.send_json(restore_workhub_backup(upload_path))
                return

            if self.path == "/api/vendor-contacts-import":
                if not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                upload_path = save_uploaded_file(fields, "file")
                contacts, saved_count = import_vendor_contacts_from_workbook(upload_path)
                self.send_json({
                    "message": f"업체 메일 주소 {saved_count}건을 저장했습니다.",
                    "saved_count": saved_count,
                    "contacts": contacts,
                })
                return

            if self.path == "/api/sales-report-upload":
                if not self.require_permission(user, "sales_report_manage", "매출표 업로드"):
                    return
                upload_path = save_uploaded_sales_report_file(fields, "file")
                uploaded_by = str(user.get("display_name") or user.get("username") or "")
                saved = save_sales_report_file(upload_path, original_uploaded_filename(upload_path.name), uploaded_by)
                self.send_json({
                    "message": "매출표를 저장했습니다.",
                    "file": saved,
                    "files": list_sales_report_uploads(),
                })
                return

            if self.path == "/api/cs-cases-import-preview":
                if not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                upload_path = save_uploaded_file(fields, "file")
                self.send_json(preview_cs_cases_import(upload_path))
                return

            if self.path == "/api/cs-cases-import":
                mode = fields.get("mode", "daily")
                mode = mode if isinstance(mode, str) and mode == "replace" else "daily"
                if mode == "replace":
                    if not self.require_admin(user, "CS처리대장 전체 데이터 교체 업로드"):
                        return
                elif not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                upload_path = save_uploaded_file(fields, "file")
                corrections = parse_import_corrections(fields)
                inserted, skipped = import_cs_cases_from_workbook(upload_path, mode=mode, corrections=corrections)
                mode_label = "전체 교체" if mode == "replace" else "일일 추가"
                self.send_json({
                    "message": f"CS 처리대장 {mode_label} {inserted}건 완료, 중복 {skipped}건 제외했습니다.",
                    "mode": mode,
                    "inserted": inserted,
                    "skipped": skipped,
                })
                return

            if self.path == "/api/management-import-preview":
                if not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                upload_path = save_uploaded_file(fields, "file")
                self.send_json(preview_management_import(upload_path))
                return

            if self.path == "/api/management-import":
                mode = fields.get("mode", "daily")
                mode = mode if isinstance(mode, str) and mode == "replace" else "daily"
                if mode == "replace":
                    if not self.require_admin(user, "통합관리대장 전체 데이터 교체 업로드"):
                        return
                elif not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                upload_path = save_uploaded_file(fields, "file")
                corrections = parse_import_corrections(fields)
                inserted, skipped = import_management_workbook(upload_path, mode=mode, corrections=corrections)
                mode_label = "전체 교체" if mode == "replace" else "일일 추가"
                self.send_json({
                    "message": f"통합관리대장 {mode_label} {inserted}건 완료, 중복 {skipped}건 제외했습니다.",
                    "mode": mode,
                    "inserted": inserted,
                    "skipped": skipped,
                })
                return

            if self.path == "/api/delivery-summary":
                if not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                upload_path = save_uploaded_file(fields)
                sort_mode = fields.get("sort", "name")
                if sort_mode not in {"name", "count", "first"}:
                    sort_mode = "name"
                summary_payload = build_summary_payload(upload_path, sort_mode=str(sort_mode))
                text = summary_payload["text"]
                lines = [line for line in text.splitlines() if " - " in line]
                self.send_json({
                    "text": text,
                    "approved_text": summary_payload.get("approved_text", text),
                    "safe_number_candidates": summary_payload.get("safe_number_candidates", []),
                    "line_count": len(lines),
                })
                return

            if self.path == "/api/invoice-export":
                if not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                    return
                upload_path = save_uploaded_file(fields)
                preview_rows = extract_invoice_rows(upload_path)
                output_path = export_invoice_numbers(upload_path, DOWNLOAD_DIR)
                register_order_download(output_path, "송장번호 추출")
                filename = quote(output_path.name)
                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.send_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{filename}",
                )
                self.send_header("X-Row-Count", str(len(preview_rows)))
                data = output_path.read_bytes()
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return

            if self.path == "/api/lotte-order-form":
                if not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                    return
                source_path = save_uploaded_file(fields, "file")
                if not LOTTE_TEMPLATE.exists():
                    raise FileNotFoundError(f"롯데택배 발주서 양식을 찾지 못했습니다: {LOTTE_TEMPLATE}")
                output_path = convert_lotte_order_form(source_path, LOTTE_TEMPLATE, DOWNLOAD_DIR)
                register_order_download(output_path, "롯데택배 발주서 변환")
                filename = quote(output_path.name)
                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.send_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{filename}",
                )
                data = output_path.read_bytes()
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return

            if self.path == "/api/sales-vendor-summary":
                if not self.require_permission(user, "excel_upload", "엑셀 업로드"):
                    return
                if not self.require_permission(user, "excel_download", "엑셀 다운로드"):
                    return
                source_path = save_uploaded_file(fields, "file")
                output_path = convert_sales_vendor_workbook(source_path, DOWNLOAD_DIR)
                register_order_download(output_path, "매입/매출별 테이터 정리")
                filename = quote(output_path.name)
                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.send_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{filename}",
                )
                data = output_path.read_bytes()
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return

            self.send_error(404)
        except PermissionError as exc:
            self.send_json({"error": str(exc)}, status=403)
        except Exception as exc:  # noqa: BLE001
            self.send_json({"error": str(exc)}, status=400)

    def send_json(self, payload: dict, status: int = 200) -> None:
        self.send_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
            status,
        )

    def send_bytes(self, data: bytes, content_type: str, status: int = 200) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "same-origin")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format: str, *args) -> None:
        return


def run(host: str = "127.0.0.1", port: int = 8765) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    init_db()
    start_backup_scheduler()
    server = ThreadingHTTPServer((host, port), WorkhubHandler)
    print(f"(주)소일브릿지 발주 업무자동화 앱 실행 중: http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    selected_port = int(os.environ.get("WORKHUB_PORT", sys.argv[1] if len(sys.argv) > 1 else 8765))
    selected_host = os.environ.get("WORKHUB_HOST", "127.0.0.1")
    run(host=selected_host, port=selected_port)
