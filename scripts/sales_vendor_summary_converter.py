from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SUPPLY_HEADERS = [
    "순서",
    "매입거래처",
    "매출거래처",
    "거래구분",
    "장부입력확인",
    "주문일",
    "출고일",
    "주문자",
    "발신자연락처",
    "수령자",
    "수령자연락처",
    "제 품 명",
    "수량",
    "상 세 주 소",
    "택배사**",
    "배송번호",
    "특이(요청)사항",
    "주문상품고유번호",
    "상품코드",
    "주문번호",
    "고객선택옵션",
]

DEFAULT_PURCHASE_VENDOR = "(주)소일브릿지(본사)"
OUTPUT_PREFIX = "주소일브릿지_매출처별_정리"

COLUMN_WIDTHS = [
    6,
    22,
    22,
    14,
    14,
    18,
    14,
    16,
    16,
    18,
    18,
    42,
    8,
    48,
    14,
    22,
    26,
    20,
    16,
    24,
    22,
]


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", _clean_text(value)).lower()


def _normalize_vendor(value: object) -> str:
    return re.sub(r"\s+", " ", _clean_text(value)) or "미지정"


def _display_quantity(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_source_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:19] if "%H" in pattern else text[:10], pattern).date()
        except ValueError:
            continue
    match = re.search(r"(20\d{2})[-./년\s]+(\d{1,2})[-./월\s]+(\d{1,2})", text)
    if not match:
        return None
    return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))


def _format_output_filename(rows: list[dict[str, object]]) -> str:
    output_date = None
    for row in rows:
        output_date = _parse_source_date(row.get("주문일"))
        if output_date:
            break
    if output_date is None:
        output_date = date.today()
    return f"{OUTPUT_PREFIX} {output_date.year}년 {output_date.month}월 {output_date.day}일.xlsx"


def _find_header_row(worksheet) -> tuple[int, dict[str, int]]:
    required = {_normalize_header(header): header for header in SUPPLY_HEADERS}
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        indexes: dict[str, int] = {}
        for column_index, value in enumerate(row, start=1):
            normalized = _normalize_header(value)
            if normalized in required:
                indexes[required[normalized]] = column_index
        missing = [header for header in SUPPLY_HEADERS if header not in indexes]
        if not missing:
            return row_index, indexes
    raise ValueError("주소일브릿지 Supply 형식의 헤더를 찾지 못했습니다.")


def _extract_rows(source_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook["Supply"] if "Supply" in workbook.sheetnames else workbook.active
    header_row, indexes = _find_header_row(worksheet)
    rows: list[dict[str, object]] = []

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        record = {
            header: row[indexes[header] - 1] if indexes[header] - 1 < len(row) else None
            for header in SUPPLY_HEADERS
        }
        if not any(record.values()):
            continue
        if not any(record.get(header) for header in ("매입거래처", "매출거래처", "제 품 명", "수령자", "배송번호")):
            continue
        rows.append(record)

    if not rows:
        raise ValueError("정리할 주소일브릿지 주문 데이터가 없습니다.")
    return rows


def _summary_text(rows: list[dict[str, object]]) -> str:
    purchase_groups: OrderedDict[str, list[dict[str, object]]] = OrderedDict()
    for row in rows:
        purchase_vendor = _normalize_vendor(row.get("매입거래처"))
        purchase_groups.setdefault(purchase_vendor, []).append(row)

    lines: list[str] = []
    for group_index, (purchase_vendor, group_rows) in enumerate(purchase_groups.items()):
        if group_index:
            lines.append("")
        if purchase_vendor != DEFAULT_PURCHASE_VENDOR:
            lines.append(f"★{purchase_vendor} 매입건★")

        product_counts: OrderedDict[tuple[str, str], int] = OrderedDict()
        for row in group_rows:
            product_name = _clean_text(row.get("제 품 명")) or "상품명 없음"
            quantity = _display_quantity(row.get("수량"))
            key = (product_name, quantity)
            product_counts[key] = product_counts.get(key, 0) + 1

        for (product_name, quantity), count in product_counts.items():
            quantity_text = f"{quantity}개" if quantity else "수량 없음"
            lines.append(f"{product_name} - {quantity_text} ({count}건)")

    return "\n".join(lines)


def _safe_sheet_title(title: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", " ", title).strip() or "미지정"
    base = base[:31]
    candidate = base
    suffix = 1
    while candidate in used:
        suffix_text = f"_{suffix}"
        candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used.add(candidate)
    return candidate


def _style_sheet(worksheet, data_start_row: int, max_row: int) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor="70AD47")
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    worksheet["A1"].fill = title_fill
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 26

    worksheet["A3"].font = Font(bold=True, color="1F2937")
    worksheet["A3"].fill = summary_fill
    worksheet["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    for col_index, width in enumerate(COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    for cell in worksheet[data_start_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in worksheet.iter_rows(min_row=data_start_row + 1, max_row=max_row, max_col=len(SUPPLY_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in {12, 14, 17, 21})
            cell.border = border

    worksheet.freeze_panes = f"A{data_start_row + 1}"
    worksheet.auto_filter.ref = f"A{data_start_row}:{get_column_letter(len(SUPPLY_HEADERS))}{max_row}"


def convert_sales_vendor_workbook(source_path: str | Path, output_dir: str | Path) -> Path:
    source_path = Path(source_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    rows = _extract_rows(source_path)
    sales_groups: OrderedDict[str, list[dict[str, object]]] = OrderedDict()
    for row in rows:
        sales_vendor = _normalize_vendor(row.get("매출거래처"))
        sales_groups.setdefault(sales_vendor, []).append(row)

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)
    used_titles: set[str] = set()

    for sales_vendor in sorted(sales_groups):
        group_rows = sales_groups[sales_vendor]
        sheet_title = _safe_sheet_title(sales_vendor, used_titles)
        worksheet = output_workbook.create_sheet(sheet_title)

        worksheet.merge_cells("A1:H1")
        worksheet["A1"] = f"{sales_vendor} 요약"

        summary_text = _summary_text(group_rows)
        summary_line_count = max(1, len(summary_text.splitlines()))
        summary_end_row = 3 + summary_line_count - 1
        worksheet.merge_cells(start_row=3, start_column=1, end_row=summary_end_row, end_column=8)
        worksheet["A3"] = summary_text
        for row_index in range(3, summary_end_row + 1):
            worksheet.row_dimensions[row_index].height = 19

        header_row = summary_end_row + 2
        for column_index, header in enumerate(SUPPLY_HEADERS, start=1):
            worksheet.cell(header_row, column_index, header)
        for row_offset, record in enumerate(group_rows, start=1):
            for column_index, header in enumerate(SUPPLY_HEADERS, start=1):
                worksheet.cell(header_row + row_offset, column_index, record.get(header))

        _style_sheet(worksheet, header_row, worksheet.max_row)

    output_path = output_dir / _format_output_filename(rows)
    output_workbook.save(output_path)
    return output_path
