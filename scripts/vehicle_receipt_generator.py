from __future__ import annotations

import argparse
import json
import os
import re
import sys
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.properties import PageSetupProperties


if getattr(sys, "frozen", False):
    ROOT = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    RUNTIME_ROOT = Path(os.environ.get("LOCALAPPDATA", str(Path.home()))) / "Workhub"
else:
    ROOT = Path(__file__).resolve().parents[1]
    RUNTIME_ROOT = ROOT

RUNTIME_ROOT = Path(os.environ.get("WORKHUB_DATA_DIR", str(RUNTIME_ROOT)))

DEFAULT_TEMPLATE = ROOT / "templates" / "vehicle_receipt_template.xlsx"
DEFAULT_OUTPUT_DIR = RUNTIME_ROOT / "output" / "vehicle_receipts"

PRODUCT_START_ROW = 8
PRODUCT_TEMPLATE_ROWS = 31
MIN_PRODUCT_ROWS = 5
MODNI_MIN_PRODUCT_ROWS = 15
PRODUCT_ROW_HEIGHT = 130
PRODUCT_FONT_SIZE = 30
TOP_ROW_HEIGHT = 80
INFO_ROW_HEIGHT = 130
NO_COL = 2
PRODUCT_COL = 3
QUANTITY_COL = 7
NOTE_COL = 10
FREIGHT_PAYMENT_OPTIONS = {"선불", "후불"}
RECEIPT_TYPE_OPTIONS = {"일반", "모드니 전용"}
REQUEST_FONT_COLOR = "FF003399"
TITLE_FONT_COLOR = "FF003399"
WEEKDAYS_KO = ("월", "화", "수", "목", "금", "토", "일")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_items(raw_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for raw in raw_items:
        product_name = clean_text(raw.get("product_name"))
        quantity = clean_text(raw.get("quantity"))
        pack_quantity = clean_text(raw.get("pack_quantity"))
        if not product_name and not quantity and not pack_quantity:
            continue
        items.append(
            {
                "product_name": product_name,
                "quantity": quantity,
                "pack_quantity": pack_quantity,
            }
        )
    return items


def normalize_freight_payment(value: Any) -> str:
    freight_payment = clean_text(value)
    return freight_payment if freight_payment in FREIGHT_PAYMENT_OPTIONS else "선불"


def normalize_receipt_type(value: Any) -> str:
    receipt_type = clean_text(value)
    return receipt_type if receipt_type in RECEIPT_TYPE_OPTIONS else "일반"


def first_int(value: str) -> int | None:
    match = re.search(r"\d+", value.replace(",", ""))
    return int(match.group()) if match else None


def format_box_note(quantity: str, pack_quantity: str) -> str:
    pack_count = first_int(pack_quantity)
    if not pack_count or pack_count <= 0:
        return ""

    quantity_count = first_int(quantity)
    if not quantity_count or quantity_count <= 0:
        return f"(입수량 : {pack_count}EA)"

    boxes, remainder = divmod(quantity_count, pack_count)
    if remainder:
        return f"{boxes}박스+{remainder}EA\n(입수량 : {pack_count}EA)"
    return f"{boxes}박스\n(입수량 : {pack_count}EA)"


def format_quantity(value: str) -> str:
    quantity = clean_text(value)
    if not quantity:
        return ""
    if re.search(r"ea$", quantity, flags=re.IGNORECASE):
        return quantity
    quantity_count = first_int(quantity)
    if quantity_count:
        return f"{quantity_count}EA"
    return f"{quantity}EA"


def safe_filename_text(value: str, fallback: str = "미입력") -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\r\n]+', "_", clean_text(value))
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned or fallback


def format_date_with_weekday(value: date) -> str:
    weekday = WEEKDAYS_KO[value.weekday()]
    return f"{value.year}년 {value.month}월 {value.day}일({weekday})"


def set_font_size(cell, size: int) -> None:
    font = copy(cell.font)
    font.size = size
    cell.font = font


def remove_stale_merged_cells(worksheet, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, worksheet.max_column + 1):
            if isinstance(worksheet.cell(row, col), MergedCell):
                worksheet._cells.pop((row, col), None)


def copy_row_style(worksheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = worksheet.cell(source_row, col)
        target = worksheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.border = copy(source.border)
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height


def merged_ranges_intersecting_rows(worksheet, start_row: int, end_row: int) -> list[str]:
    intersecting: list[str] = []
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row >= start_row and merged_range.min_row <= end_row:
            intersecting.append(str(merged_range))
    return intersecting


def shrink_product_area(worksheet, wanted_rows: int) -> None:
    if wanted_rows >= PRODUCT_TEMPLATE_ROWS:
        return

    delete_start = PRODUCT_START_ROW + wanted_rows
    delete_count = PRODUCT_TEMPLATE_ROWS - wanted_rows
    delete_end = delete_start + delete_count - 1

    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row >= delete_start or (
            merged_range.max_row >= delete_start and merged_range.min_row <= delete_end
        ):
            worksheet.unmerge_cells(str(merged_range))

    worksheet.delete_rows(delete_start, delete_count)


def expand_product_area(worksheet, wanted_rows: int) -> None:
    if wanted_rows <= PRODUCT_TEMPLATE_ROWS:
        return

    insert_at = PRODUCT_START_ROW + PRODUCT_TEMPLATE_ROWS
    extra_rows = wanted_rows - PRODUCT_TEMPLATE_ROWS
    worksheet.insert_rows(insert_at, extra_rows)

    source_row = PRODUCT_START_ROW
    for row in range(insert_at, insert_at + extra_rows):
        copy_row_style(worksheet, source_row, row, worksheet.max_column)
        worksheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        worksheet.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        worksheet.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)


def clear_product_rows(worksheet, product_rows: int) -> None:
    for row in range(PRODUCT_START_ROW, PRODUCT_START_ROW + product_rows):
        worksheet.cell(row, NO_COL).value = None
        worksheet.cell(row, PRODUCT_COL).value = None
        worksheet.cell(row, QUANTITY_COL).value = None
        worksheet.cell(row, NOTE_COL).value = None


def rebuild_bottom_layout(worksheet, bottom_start: int, request_text: str = "") -> int:
    has_request = bool(clean_text(request_text))
    request_offset = 1 if has_request else 0
    last_row = bottom_start + 3 + request_offset

    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row >= bottom_start:
            worksheet.unmerge_cells(str(merged_range))

    if worksheet.max_row > last_row:
        worksheet.delete_rows(last_row + 1, worksheet.max_row - last_row)

    remove_stale_merged_cells(worksheet, bottom_start, last_row)

    for row in range(bottom_start, last_row + 1):
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row, col).value = None

    request_row = bottom_start if has_request else None
    phone_row = bottom_start + request_offset
    blank_row = bottom_start + 1 + request_offset
    delivery_row = bottom_start + 2 + request_offset
    manager_row = bottom_start + 3 + request_offset

    if request_row:
        worksheet.merge_cells(start_row=request_row, start_column=2, end_row=request_row, end_column=13)
        worksheet.cell(request_row, 2).value = f"요청사항(필독) : {clean_text(request_text)}"
        worksheet.cell(request_row, 2).font = Font(bold=True, size=30, color=REQUEST_FONT_COLOR)
        worksheet.cell(request_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[request_row].height = 80

    worksheet.merge_cells(start_row=phone_row, start_column=2, end_row=phone_row, end_column=13)
    worksheet.merge_cells(start_row=blank_row, start_column=2, end_row=blank_row, end_column=11)
    worksheet.merge_cells(start_row=blank_row, start_column=12, end_row=blank_row, end_column=13)
    worksheet.merge_cells(start_row=delivery_row, start_column=2, end_row=delivery_row, end_column=3)
    worksheet.merge_cells(start_row=delivery_row, start_column=4, end_row=delivery_row, end_column=11)
    worksheet.merge_cells(start_row=delivery_row, start_column=12, end_row=manager_row, end_column=13)
    worksheet.merge_cells(start_row=manager_row, start_column=2, end_row=manager_row, end_column=3)
    worksheet.merge_cells(start_row=manager_row, start_column=4, end_row=manager_row, end_column=11)

    worksheet.cell(phone_row, 2).value = "010-3663-0838 / 사진촬영 후 문자 부탁드립니다"
    worksheet.cell(blank_row, 2).value = "납품 주소 및 인수자 연락처"
    worksheet.cell(blank_row, 2).font = Font(bold=True, size=24)
    worksheet.cell(blank_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.cell(blank_row, 12).value = "인수자 확인"
    worksheet.cell(blank_row, 12).font = Font(bold=True, size=24)
    worksheet.cell(delivery_row, 2).value = "납품장소"
    worksheet.cell(manager_row, 2).value = "담당자명"

    wide_text = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_text = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.cell(delivery_row, 4).alignment = wide_text
    worksheet.cell(manager_row, 4).alignment = wide_text
    worksheet.cell(blank_row, 12).alignment = center_text
    worksheet.row_dimensions[phone_row].height = 80
    worksheet.row_dimensions[blank_row].height = 60
    worksheet.row_dimensions[delivery_row].height = INFO_ROW_HEIGHT
    worksheet.row_dimensions[manager_row].height = INFO_ROW_HEIGHT

    return last_row


def dated_filename(output_date: date, supplier: str) -> str:
    return (
        f"★차량인수증_★ (주)소일브릿지 - "
        f"{format_date_with_weekday(output_date)} - {safe_filename_text(supplier, '공급받는자')}.xlsx"
    )


def set_print_layout(worksheet, last_row: int) -> None:
    worksheet.print_area = f"B1:M{last_row}"
    worksheet.column_dimensions["A"].hidden = True
    worksheet.column_dimensions["A"].width = 0
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.fitToWidth = None
    worksheet.page_setup.fitToHeight = None
    worksheet.page_setup.scale = 100
    if worksheet.sheet_properties.pageSetUpPr is None:
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties()
    worksheet.sheet_properties.pageSetUpPr.fitToPage = False
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.sheet_view.view = "pageBreakPreview"


def generate_vehicle_receipt(
    supplier: str,
    items: list[dict[str, Any]],
    delivery_place: str,
    manager: str,
    freight_payment: str = "선불",
    receipt_type: str = "일반",
    request_note: str = "",
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    template_path: Path = DEFAULT_TEMPLATE,
    output_date: date | None = None,
) -> Path:
    output_date = output_date or date.today()
    parsed_items = parse_items(items)
    freight_payment = normalize_freight_payment(freight_payment)
    receipt_type = normalize_receipt_type(receipt_type)
    min_product_rows = MODNI_MIN_PRODUCT_ROWS if receipt_type == "모드니 전용" else MIN_PRODUCT_ROWS
    product_rows = max(min_product_rows, len(parsed_items))

    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    expand_product_area(worksheet, product_rows)
    shrink_product_area(worksheet, product_rows)
    clear_product_rows(worksheet, product_rows)

    for row in range(1, 6):
        worksheet.row_dimensions[row].height = TOP_ROW_HEIGHT

    worksheet["B1"] = "인수증(공급받는자용)"
    worksheet["B1"].font = Font(bold=True, size=28, color=TITLE_FONT_COLOR)
    worksheet["B1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet["D2"] = format_date_with_weekday(output_date)
    worksheet["D3"] = supplier
    worksheet["K1"] = freight_payment

    center_text = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet["D2"].alignment = center_text
    worksheet["D3"].alignment = center_text
    set_font_size(worksheet["D3"], 48)
    worksheet["K1"].alignment = center_text

    for idx in range(product_rows):
        row = PRODUCT_START_ROW + idx
        worksheet.row_dimensions[row].height = PRODUCT_ROW_HEIGHT
        worksheet.cell(row, NO_COL).value = idx + 1
        worksheet.cell(row, NO_COL).alignment = center_text
        worksheet.cell(row, PRODUCT_COL).alignment = center_text
        worksheet.cell(row, QUANTITY_COL).alignment = center_text
        worksheet.cell(row, NOTE_COL).alignment = center_text
        for col in (NO_COL, PRODUCT_COL, QUANTITY_COL, NOTE_COL):
            set_font_size(worksheet.cell(row, col), PRODUCT_FONT_SIZE)
        if idx < len(parsed_items):
            worksheet.cell(row, PRODUCT_COL).value = parsed_items[idx]["product_name"]
            worksheet.cell(row, QUANTITY_COL).value = format_quantity(parsed_items[idx]["quantity"])
            worksheet.cell(row, NOTE_COL).value = format_box_note(
                parsed_items[idx]["quantity"],
                parsed_items[idx]["pack_quantity"],
            )

    bottom_start = PRODUCT_START_ROW + product_rows
    request_offset = 1 if clean_text(request_note) else 0
    last_row = rebuild_bottom_layout(worksheet, bottom_start, request_note)
    worksheet.cell(bottom_start + 2 + request_offset, 4).value = delivery_place
    worksheet.cell(bottom_start + 3 + request_offset, 4).value = manager
    worksheet.cell(bottom_start + 2 + request_offset, 4).alignment = center_text
    worksheet.cell(bottom_start + 3 + request_offset, 4).alignment = center_text

    set_print_layout(worksheet, last_row)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / dated_filename(output_date, supplier)
    workbook.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="차량인수증 엑셀을 생성합니다.")
    parser.add_argument("--supplier", required=True, help="공급받는자")
    parser.add_argument("--items-json", required=True, help="제품 목록 JSON")
    parser.add_argument("--delivery-place", required=True, help="납품장소")
    parser.add_argument("--manager", required=True, help="담당자명")
    parser.add_argument("--freight-payment", default="선불", choices=sorted(FREIGHT_PAYMENT_OPTIONS), help="운임비용")
    parser.add_argument("--receipt-type", default="일반", choices=sorted(RECEIPT_TYPE_OPTIONS), help="차량인수증 타입")
    parser.add_argument("--request-note", default="", help="요청사항")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="결과 저장 폴더")
    args = parser.parse_args()

    items = json.loads(args.items_json)
    output_path = generate_vehicle_receipt(
        supplier=args.supplier,
        items=items,
        delivery_place=args.delivery_place,
        manager=args.manager,
        freight_payment=args.freight_payment,
        receipt_type=args.receipt_type,
        request_note=args.request_note,
        output_dir=Path(args.output_dir),
    )
    print(f"저장 완료: {output_path}")


if __name__ == "__main__":
    main()
