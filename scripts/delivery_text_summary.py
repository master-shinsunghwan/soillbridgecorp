from __future__ import annotations

import argparse
import re
from collections import OrderedDict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "output" / "delivery_text"

PRODUCT_HEADERS = ("제 품 명", "제품명", "상품명", "품명", "주문상품명")
QUANTITY_HEADERS = ("수 량", "수량", "주문수량", "개수")
ORDER_HEADERS = ("주문번호", "주문 번호", "order_no", "order number")
RECEIVER_NAME_HEADERS = ("받으시는분", "받으시는 분", "수취인", "수령인", "수령자", "받는분")
RECEIVER_PHONE_HEADERS = ("받으시는분 연락처", "받으시는분연락처", "수취인연락처", "수령인연락처", "수령자연락처", "수령자 연락처", "받는분연락처")
DETAIL_ADDRESS_HEADERS = ("상 세 주 소", "상세주소", "상세 주소", "주소", "받으시는분주소")

WEEKDAYS_KO = ("월", "화", "수", "목", "금", "토", "일")

DEFAULT_ALIASES = {
    "[아이제나흐] 2IN1 레트로 글라스 대용량 믹서기 1.5L / EZH-666QEWR": "레트로 글라스 믹서기",
    "[아이제나흐] 스테넬(STENEL-STAINLESS+PANEL) 양면도마": "스테넬 양면도마",
    "☆33cm☆[아이제나흐] 무쇠 그리들팬 33cm": "무쇠 그리들팬 33cm",
    "★39cm★[아이제나흐] 무쇠 그리들팬 39cm": "무쇠 그리들팬 39cm",
    "노르디쿡 IH 무쇠팬 28cm + 뒤집개": "무쇠팬 28cm + 뒤집개",
    "아이제나흐 베틴 IH 스텐냄비 26곰솥냄비": "베틴 IH 스텐냄비 26곰솥냄비",
    "아이제나흐 베틴 IH 스텐냄비 4종(18cm편수+20cm양수+26cm전골+26cm곰솥)": "베틴 IH 스텐냄비 4종(18cm편수+20cm양수+26cm전골+26cm곰솥)",
    "아이제나흐 에센 도마(대)": "에센 도마(대)",
}


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", clean_cell(value)).lower()


def normalize_package_address(value: Any) -> str:
    return re.sub(r"[\s,.\-]+", "", clean_cell(value)).lower()


def find_column(headers: list[Any], candidates: tuple[str, ...], label: str) -> int:
    normalized = [normalize_header(header) for header in headers]
    for candidate in candidates:
        target = normalize_header(candidate)
        if target in normalized:
            return normalized.index(target)
    raise ValueError(f"{label} 컬럼을 찾지 못했습니다. 후보: {', '.join(candidates)}")


def normalize_product_name(raw_name: str) -> str:
    name = clean_cell(raw_name)
    if name in DEFAULT_ALIASES:
        return DEFAULT_ALIASES[name]

    name = re.sub(r"^[★☆]\s*([^★☆]{1,20})\s*[★☆]\s*", "", name).strip()
    name = re.sub(r"\s*/\s*[A-Z0-9_-]{4,}\s*$", "", name).strip()
    name = re.sub(r"\s+", " ", name).strip()
    return DEFAULT_ALIASES.get(name, name)


def parse_quantity(value: Any) -> int:
    if value is None or value == "":
        return 1
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value))
    return int(match.group(0)) if match else 1


def infer_delivery_date(order_numbers: list[str], fallback: date | None = None) -> date:
    for order_number in order_numbers:
        match = re.search(r"(\d{6})", order_number)
        if not match:
            continue
        raw = match.group(1)
        year = 2000 + int(raw[:2])
        month = int(raw[2:4])
        day = int(raw[4:6])
        try:
            return date(year, month, day)
        except ValueError:
            continue
    return fallback or date.today()


def format_title(delivery_date: date) -> str:
    weekday = WEEKDAYS_KO[delivery_date.weekday()]
    return f"★{delivery_date.month}월{delivery_date.day}일({weekday}) 개별 택배건 전달드립니다★"


def parse_delivery_workbook(
    path: Path,
    sheet_name: str | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows))
        product_idx = find_column(headers, PRODUCT_HEADERS, "상품명")
        quantity_idx = find_column(headers, QUANTITY_HEADERS, "수량")
        receiver_name_idx = find_column(headers, RECEIVER_NAME_HEADERS, "받으시는분 성함")
        receiver_phone_idx = find_column(headers, RECEIVER_PHONE_HEADERS, "받으시는분 연락처")
        detail_address_idx = find_column(headers, DETAIL_ADDRESS_HEADERS, "상세주소")

        parsed_rows: list[dict[str, Any]] = []

        for row in rows:
            if not any(row):
                continue
            product_name = normalize_product_name(clean_cell(row[product_idx]))
            if not product_name:
                continue
            quantity = parse_quantity(row[quantity_idx])
            receiver_name = clean_cell(row[receiver_name_idx])
            receiver_phone = clean_cell(row[receiver_phone_idx])
            detail_address = clean_cell(row[detail_address_idx])
            recipient_key = (receiver_name, receiver_phone, detail_address)
            parsed_rows.append(
                {
                    "product_name": product_name,
                    "quantity": quantity,
                    "recipient_key": recipient_key,
                    "safe_recipient_key": (receiver_name, normalize_package_address(detail_address)),
                    "receiver_name": receiver_name,
                    "receiver_phone": receiver_phone,
                    "detail_address": detail_address,
                }
            )

        return parsed_rows, list(workbook.sheetnames)
    finally:
        workbook.close()


def recipient_count_map(parsed_rows: list[dict[str, Any]]) -> OrderedDict[tuple[str, str, str], int]:
    recipient_counts: OrderedDict[tuple[str, str, str], int] = OrderedDict()
    for parsed_row in parsed_rows:
        recipient_key = parsed_row["recipient_key"]
        if all(recipient_key):
            recipient_counts[recipient_key] = recipient_counts.get(recipient_key, 0) + 1
    return recipient_counts


def safe_number_candidates(parsed_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: OrderedDict[tuple[str, str], list[dict[str, Any]]] = OrderedDict()
    for parsed_row in parsed_rows:
        safe_key = parsed_row["safe_recipient_key"]
        if all(safe_key):
            grouped.setdefault(safe_key, []).append(parsed_row)

    candidates: list[dict[str, Any]] = []
    for rows in grouped.values():
        phones = []
        for row in rows:
            phone = row["receiver_phone"]
            if phone and phone not in phones:
                phones.append(phone)
        if len(rows) < 2 or len(phones) < 2:
            continue
        first = rows[0]
        candidates.append(
            {
                "name": first["receiver_name"],
                "address": first["detail_address"],
                "phones": phones,
                "items": [
                    f"{row['product_name']} - {row['quantity']}개"
                    for row in rows
                ],
            }
        )
    return candidates


def render_summary(
    parsed_rows: list[dict[str, Any]],
    sort_mode: str = "name",
    approved_safe_keys: set[tuple[str, str]] | None = None,
) -> str:
    approved_safe_keys = approved_safe_keys or set()
    recipient_counts = recipient_count_map(parsed_rows)
    grouped: OrderedDict[tuple[str, int], int] = OrderedDict()
    combined_recipients: OrderedDict[tuple[Any, ...], list[tuple[str, int]]] = OrderedDict()
    for parsed_row in parsed_rows:
        key = (parsed_row["product_name"], parsed_row["quantity"])
        recipient_key = parsed_row["recipient_key"]
        safe_key = parsed_row["safe_recipient_key"]
        if safe_key in approved_safe_keys:
            combined_recipients.setdefault(("safe", *safe_key), []).append(key)
        elif recipient_counts.get(recipient_key, 0) > 1:
            combined_recipients.setdefault(("strict", *recipient_key), []).append(key)
        else:
            grouped[key] = grouped.get(key, 0) + 1

    delivery_date = date.today()
    lines = [format_title(delivery_date), ""]

    items = list(grouped.items())
    if sort_mode == "name":
        items.sort(key=lambda item: (item[0][0], item[0][1]))
    elif sort_mode == "count":
        items.sort(key=lambda item: (-item[1], item[0][0], item[0][1]))

    for combined_items in combined_recipients.values():
        parts = [
            f"{product_name} - {quantity}개"
            for product_name, quantity in combined_items
        ]
        lines.append(f"★합포장 확인 요청★{' + '.join(parts)}")

    lines.extend(
        f"{product_name} - {quantity}개 ({count}건)"
        for (product_name, quantity), count in items
    )
    return "\n".join(lines)


def build_summary_payload(
    path: Path,
    sheet_name: str | None = None,
    sort_mode: str = "name",
) -> dict[str, Any]:
    parsed_rows, sheet_names = parse_delivery_workbook(path, sheet_name)
    candidates = safe_number_candidates(parsed_rows)
    approved_keys = {
        (
            candidate["name"],
            normalize_package_address(candidate["address"]),
        )
        for candidate in candidates
    }
    text = render_summary(parsed_rows, sort_mode)
    approved_text = render_summary(parsed_rows, sort_mode, approved_keys) if candidates else text
    return {
        "text": text,
        "approved_text": approved_text,
        "safe_number_candidates": candidates,
        "sheet_names": sheet_names,
    }


def summarize_workbook(
    path: Path,
    sheet_name: str | None = None,
    sort_mode: str = "name",
) -> tuple[str, list[str]]:
    payload = build_summary_payload(path, sheet_name, sort_mode)
    return payload["text"], payload["sheet_names"]


def build_output_path(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{input_path.stem}_택배건_요약.txt"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="주소일브릿지 엑셀 파일을 상품별 개별 택배건 전달 텍스트로 변환합니다."
    )
    parser.add_argument("xlsx_path", help="입력 엑셀 파일 경로")
    parser.add_argument("--sheet", help="시트명. 생략하면 첫 번째 시트를 사용합니다.")
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="결과 텍스트 저장 폴더",
    )
    parser.add_argument(
        "--sort",
        choices=("name", "count", "first"),
        default="name",
        help="정렬 방식: name=상품명순, count=건수순, first=엑셀 등장순",
    )
    args = parser.parse_args()

    input_path = Path(args.xlsx_path).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"입력 파일이 없습니다: {input_path}")

    text, _sheet_names = summarize_workbook(input_path, args.sheet, args.sort)
    output_path = build_output_path(input_path, Path(args.output_dir))
    output_path.write_text(text, encoding="utf-8-sig")
    print(text)
    print()
    print(f"저장 완료: {output_path}")


if __name__ == "__main__":
    main()
