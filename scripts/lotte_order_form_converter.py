from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "output" / "lotte_order_forms"

SOURCE_TO_TEMPLATE = {
    "판매사": "주문번호",
    "보내는분": "보내는분",
    "보내는분연락처": "보내는분연락처",
    "받으시는분": "받으시는분",
    "받으시는분 연락처": "받으시는분 연락처",
    "제 품 명": "제 품 명",
    "수 량": "택배수량",
    "상 세 주 소": "상 세 주 소",
    "운임": "운임",
    "배송메모": "특이(요청)사항",
}


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", clean_cell(value)).lower()


def header_map(headers: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, value in enumerate(headers, start=1):
        normalized = normalize_header(value)
        if normalized:
            result[normalized] = idx
    return result


def find_header(headers: list[Any], name: str) -> int:
    normalized = normalize_header(name)
    mapped = header_map(headers)
    if normalized not in mapped:
        raise ValueError(f"필수 컬럼을 찾지 못했습니다: {name}")
    return mapped[normalized]


def make_order_labels(vendor_names: list[str]) -> list[str]:
    counts = Counter(vendor_names)
    sequence: defaultdict[str, int] = defaultdict(int)
    labels: list[str] = []

    for vendor in vendor_names:
        if counts[vendor] <= 1:
            labels.append(vendor)
            continue
        sequence[vendor] += 1
        labels.append(f"{vendor}-{sequence[vendor]}")

    return labels


def copy_row_style(worksheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = worksheet.cell(source_row, col)
        target = worksheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)
        if source.border:
            target.border = copy(source.border)

    if worksheet.row_dimensions[source_row].height:
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height


def clear_template_body(worksheet, start_row: int = 2) -> None:
    for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row):
        for cell in row:
            cell.value = None


def dated_output_name(template_path: Path, output_date: date) -> str:
    month_day = f"{output_date.month}월 {output_date.day}일"
    name = template_path.name

    name = re.sub(r"2026년\s*\d{1,2}월\s*\d{0,2}\s*일", f"2026년 {month_day}", name)
    name = re.sub(r"2026년\s*월\s*일", f"2026년 {month_day}", name)

    if name == template_path.name and "2026년" in name:
        name = name.replace("2026년", f"2026년 {month_day}", 1)

    return name


def convert_lotte_order_form(
    source_path: Path,
    template_path: Path,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    output_date: date | None = None,
) -> Path:
    source_wb = load_workbook(source_path, read_only=True, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]
    template_wb = load_workbook(template_path)
    template_ws = template_wb[template_wb.sheetnames[0]]

    source_headers = [cell.value for cell in next(source_ws.iter_rows(min_row=1, max_row=1))]
    template_headers = [template_ws.cell(1, col).value for col in range(1, template_ws.max_column + 1)]

    source_columns = {
        source_header: find_header(source_headers, source_header)
        for source_header in SOURCE_TO_TEMPLATE
        if normalize_header(source_header) in header_map(source_headers)
    }
    template_columns = {
        template_header: find_header(template_headers, template_header)
        for template_header in SOURCE_TO_TEMPLATE.values()
    }

    required_source_headers = ("판매사", "받으시는분", "받으시는분 연락처", "제 품 명", "수 량", "상 세 주 소")
    for header in required_source_headers:
        find_header(source_headers, header)

    rows: list[dict[str, Any]] = []
    vendor_names: list[str] = []
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        vendor = clean_cell(row[find_header(source_headers, "판매사") - 1])
        vendor_names.append(vendor)
        row_data: dict[str, Any] = {}
        for source_header, template_header in SOURCE_TO_TEMPLATE.items():
            source_col = source_columns.get(source_header)
            if source_col is None:
                row_data[template_header] = ""
            else:
                row_data[template_header] = row[source_col - 1]
        rows.append(row_data)

    order_labels = make_order_labels(vendor_names)
    for idx, label in enumerate(order_labels):
        rows[idx]["주문번호"] = label

    clear_template_body(template_ws)
    style_row = 2
    for row_offset, row_data in enumerate(rows, start=2):
        copy_row_style(template_ws, style_row, row_offset, template_ws.max_column)
        for template_header, value in row_data.items():
            template_ws.cell(row_offset, template_columns[template_header]).value = value

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / dated_output_name(template_path, output_date or date.today())
    template_wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="주소일브릿지 엑셀 내용을 롯데택배 발주서 양식에 맞춰 출력합니다."
    )
    parser.add_argument("source_xlsx", help="주소일브릿지 원본 엑셀 파일")
    parser.add_argument("template_xlsx", help="롯데택배 발주서 양식 엑셀 파일")
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="결과 엑셀 저장 폴더",
    )
    args = parser.parse_args()

    output_path = convert_lotte_order_form(
        Path(args.source_xlsx).expanduser().resolve(),
        Path(args.template_xlsx).expanduser().resolve(),
        Path(args.output_dir),
    )
    print(f"저장 완료: {output_path}")


if __name__ == "__main__":
    main()
