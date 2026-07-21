from __future__ import annotations

import argparse
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "output" / "invoice_numbers"

NAME_HEADERS = ("수하인명", "받으시는분", "받는분", "수취인", "고객명", "이름")
INVOICE_HEADERS = ("운송장번호", "송장번호", "송장 번호", "택배번호", "tracking number")
ORDER_HEADERS = ("주문번호", "주문 번호", "order number", "order_no")
ADDRESS_HEADERS = ("수하인기본주소", "수하인 기본주소", "받으시는분주소", "주소", "기본주소")


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).strip().split())


def normalize_header(value: Any) -> str:
    return clean_cell(value).replace(" ", "").lower()


def find_column(headers: list[Any], candidates: tuple[str, ...], label: str) -> int:
    normalized = [normalize_header(header) for header in headers]
    for candidate in candidates:
        target = normalize_header(candidate)
        if target in normalized:
            return normalized.index(target)
    raise ValueError(f"{label} 컬럼을 찾지 못했습니다. 후보: {', '.join(candidates)}")


def extract_invoice_rows(path: Path, sheet_name: str | None = None) -> list[tuple[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    if (
        getattr(worksheet, "max_row", None) == 1
        and getattr(worksheet, "max_column", None) == 1
        and hasattr(worksheet, "reset_dimensions")
    ):
        worksheet.reset_dimensions()

    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    name_idx = find_column(headers, NAME_HEADERS, "이름")
    invoice_idx = find_column(headers, INVOICE_HEADERS, "송장번호")
    order_idx = find_column(headers, ORDER_HEADERS, "주문번호")
    address_idx = find_column(headers, ADDRESS_HEADERS, "수하인기본주소")

    grouped: OrderedDict[tuple[str, str, str], list[str]] = OrderedDict()

    for row in rows:
        if not any(row):
            continue

        name = clean_cell(row[name_idx])
        invoice_number = clean_cell(row[invoice_idx])
        order_number = clean_cell(row[order_idx])
        address = clean_cell(row[address_idx])
        if not name or not invoice_number:
            continue

        key = (order_number, name, address)
        if key not in grouped:
            grouped[key] = []

        grouped[key].append(invoice_number)

    return [(key[1], " / ".join(invoice_numbers)) for key, invoice_numbers in grouped.items()]


def build_output_path(input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{input_path.stem}_송장번호_추출.xlsx"


def save_invoice_workbook(rows: list[tuple[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "송장번호"

    headers = ("수하인명", "송장번호")
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="1B3A6F")
    border_alignment = Alignment(vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = border_alignment

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 56

    for row_idx in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def export_invoice_numbers(
    input_path: Path,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    sheet_name: str | None = None,
) -> Path:
    rows = extract_invoice_rows(input_path, sheet_name)
    output_path = build_output_path(input_path, output_dir)
    save_invoice_workbook(rows, output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="출고송장 엑셀에서 수하인명별 송장번호를 추출해 새 엑셀로 저장합니다."
    )
    parser.add_argument("xlsx_path", help="입력 엑셀 파일 경로")
    parser.add_argument("--sheet", help="시트명. 생략하면 첫 번째 시트를 사용합니다.")
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="결과 엑셀 저장 폴더",
    )
    args = parser.parse_args()

    input_path = Path(args.xlsx_path).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"입력 파일이 없습니다: {input_path}")

    output_path = export_invoice_numbers(input_path, Path(args.output_dir), args.sheet)
    print(f"저장 완료: {output_path}")


if __name__ == "__main__":
    main()
